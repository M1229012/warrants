# -*- coding: utf-8 -*-
"""
MoneyDJ 現股分點籌碼回測 / 排名系統
=================================

目前追蹤分點：
1. 富邦-新店    9661
2. 永豐金-板新  9A9J

資料來源：MoneyDJ 現股籌碼頁 stock-chip0002-5
功能：
1. 抓上市 + 上櫃股票清單
2. 對每檔股票抓指定區間分點買賣超排行
3. 篩選指定分點代號
4. 建立買超排名、賣超排名、絕對值排名與分點彙總
5. 將結果與快取輸出到 Google Sheet
6. 可部署於 GitHub Actions

注意：
目前確認 stock-chip0002-5 回傳的是每檔股票「區間買超 TOP15 + 區間賣超 TOP15」。
因此本程式能抓到的是：目標分點有進入該股票區間買超 / 賣超 TOP15 的資料。
若目標分點沒有進 TOP15，該股票不會命中。

本地執行：
python stock_branch_backtest.py

指定日期：
python stock_branch_backtest.py --start 2026/05/07 --end 2026/06/07

測試前 100 檔：
python stock_branch_backtest.py --limit 100

依賴：
pip install requests pandas openpyxl gspread google-auth lxml html5lib
"""

import argparse
import json
import os
import re
import time
import threading
from io import StringIO
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════
# 基本設定
# ══════════════════════════════════════════════════════════════════════

DEFAULT_OUTPUT_DIR = "output" if os.getenv("GITHUB_ACTIONS", "").strip().lower() == "true" else r"C:\Users\chen1_ukw0m7r\Downloads"
OUTPUT_DIR = os.getenv("OUTPUT_DIR", DEFAULT_OUTPUT_DIR)
CACHE_DIR = os.getenv("CACHE_DIR", os.path.join(OUTPUT_DIR, "stock_branch_cache"))
CACHE_ENCODING = "utf-8-sig"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

# MoneyDJ API
STOCK_CHIP5_URL = (
    "https://pscnetsecrwd.moneydj.com/b2brwdCommon/jsondata"
    "/8b/80/97/twstockdata.xdjjson"
)
REVISION = os.getenv("MONEYDJ_STOCK_REVISION", "d9b05f1a-9b09-4435-a592-9db7a855bd8b")

# 執行設定
MAX_WORKERS = int(os.getenv("MAX_WORKERS", os.getenv("STOCK_BRANCH_WORKERS", "18")))
REQUEST_TIMEOUT = int(os.getenv("REQUEST_TIMEOUT", "15"))
REQUEST_RETRY = int(os.getenv("REQUEST_RETRY", "2"))
LOOKBACK_DAYS = int(os.getenv("LOOKBACK_DAYS", os.getenv("STOCK_LOOKBACK_DAYS", "31")))
STOCK_LIMIT = int(os.getenv("STOCK_LIMIT", "0"))
INCLUDE_ETF = os.getenv("INCLUDE_ETF", "0").strip().lower() in ("1", "true", "yes")
REFRESH_STOCK_LIST = os.getenv("REFRESH_STOCK_LIST", "1").strip().lower() not in ("0", "false", "no")
USE_CACHE = os.getenv("USE_CACHE", "1").strip().lower() not in ("0", "false", "no")
CAPTURE_ALL_TOP15 = os.getenv("CAPTURE_ALL_TOP15", "0").strip().lower() in ("1", "true", "yes")

# 目前只追蹤兩個現股分點
TARGET_BROKERS = {
    "富邦新店": {
        "分點名稱": "富邦-新店",
        "券商代號": "9661",
    },
    "永豐金板新": {
        "分點名稱": "永豐金-板新",
        "券商代號": "9A9J",
    },
}

# 可用環境變數覆蓋，例如：TARGET_BROKERS_JSON={"富邦新店":{"分點名稱":"富邦-新店","券商代號":"9661"}}
TARGET_BROKERS_JSON = os.getenv("TARGET_BROKERS_JSON", "").strip()
if TARGET_BROKERS_JSON:
    try:
        parsed = json.loads(TARGET_BROKERS_JSON)
        if isinstance(parsed, dict) and parsed:
            TARGET_BROKERS = parsed
    except Exception as e:
        print(f"⚠️ TARGET_BROKERS_JSON 解析失敗，沿用預設分點：{type(e).__name__}: {e}")

TARGET_BROKER_CODES = {
    str(info.get("券商代號", "")).strip().lower(): label
    for label, info in TARGET_BROKERS.items()
    if str(info.get("券商代號", "")).strip()
}

# 快取路徑
STOCK_LIST_CACHE_PATH = os.path.join(CACHE_DIR, "stock_list_cache.csv")
HIT_CACHE_PATH = os.path.join(CACHE_DIR, "stock_branch_hit_cache.csv")
ALL_TOP15_CACHE_PATH = os.path.join(CACHE_DIR, "stock_branch_all_top15_cache.csv")

# Google Sheet 設定
GOOGLE_SHEET_NAME = os.getenv("GOOGLE_SHEET_NAME", "現股分點籌碼")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", "").strip()
GSHEET_CACHE_ENABLED = os.getenv("GSHEET_CACHE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_RESULT_ENABLED = os.getenv("GSHEET_RESULT_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_CHUNK_ROWS = int(os.getenv("GSHEET_CHUNK_ROWS", "3000"))
GSHEET_WRITE_SLEEP_SECONDS = float(os.getenv("GSHEET_WRITE_SLEEP_SECONDS", "1.1"))
GSHEET_MAX_RETRIES = int(os.getenv("GSHEET_MAX_RETRIES", "6"))
GSHEET_RETRY_BASE_SECONDS = float(os.getenv("GSHEET_RETRY_BASE_SECONDS", "10"))

CACHE_SHEET_NAME_MAP = {
    "stock_list_cache.csv": "快取_現股清單",
    "stock_branch_hit_cache.csv": "快取_現股分點命中",
    "stock_branch_all_top15_cache.csv": "快取_現股全排行TOP15",
}

# Excel / Google Sheet 色彩
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BUY_FILL = PatternFill("solid", fgColor="F4CCCC")
SELL_FILL = PatternFill("solid", fgColor="D9EAD3")
INFO_FILL = PatternFill("solid", fgColor="FFF2CC")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")
THIN_SIDE = Side(style="thin", color="B7B7B7")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


# ══════════════════════════════════════════════════════════════════════
# requests session
# ══════════════════════════════════════════════════════════════════════

_THREAD_LOCAL = threading.local()

HEADERS_BASE = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/148.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
}


def get_thread_session():
    session = getattr(_THREAD_LOCAL, "session", None)
    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=1)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _THREAD_LOCAL.session = session
    return session


# ══════════════════════════════════════════════════════════════════════
# 日期 / 數值工具
# ══════════════════════════════════════════════════════════════════════


def parse_date(date_value):
    try:
        if date_value is None:
            return None
        s = str(date_value).strip().replace("-", "/")
        if not s or s == "-":
            return None
        parts = s.split("/")
        if len(parts) != 3:
            return None
        return datetime(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None


def normalize_date_str(date_value):
    dt = parse_date(date_value)
    return dt.strftime("%Y/%m/%d") if dt else str(date_value).strip()


def fmt_date(dt):
    return dt.strftime("%Y/%m/%d")


def default_date_range(days=31):
    end_dt = datetime.today()
    start_dt = end_dt - timedelta(days=days)
    return fmt_date(start_dt), fmt_date(end_dt)


def to_float(value, default=0.0):
    try:
        s = str(value).replace(",", "").strip()
        if s in ("", "-", "None", "nan", "null"):
            return default
        return float(s)
    except Exception:
        return default


def to_int(value, default=0):
    try:
        return int(round(to_float(value, default=default)))
    except Exception:
        return default


def fmt_amount(value):
    try:
        return f"{int(round(float(value))):,}"
    except Exception:
        return str(value)


def safe_filename_text(text):
    text = str(text).strip()
    return re.sub(r"[\\/:*?\"<>|]", "_", text)


def normalize_broker_code(code):
    return str(code).strip().lower()


def fix_mojibake(value):
    """
    修正少數情境下 MoneyDJ 中文被 pandas / CSV 顯示成亂碼的狀況。
    原本已正常的中文會直接回傳。
    """
    s = str(value).strip()
    try:
        fixed = s.encode("latin1").decode("utf-8")
        if re.search(r"[\u4e00-\u9fff]", fixed):
            return fixed
    except Exception:
        pass
    return s


# ══════════════════════════════════════════════════════════════════════
# Google Sheet 工具
# ══════════════════════════════════════════════════════════════════════

_GSHEET_CLIENT = None
_GSHEET_SPREADSHEET = None
_GSHEET_LAST_WRITE_TS = 0.0


def gsheet_enabled():
    return bool(os.getenv("GCP_SERVICE_KEY", "").strip())


def is_gsheet_quota_error(exc):
    msg = str(exc)
    return (
        "429" in msg
        or "Quota exceeded" in msg
        or "RESOURCE_EXHAUSTED" in msg
        or "Write requests per minute" in msg
    )


def gsheet_write_sleep():
    global _GSHEET_LAST_WRITE_TS
    if GSHEET_WRITE_SLEEP_SECONDS <= 0:
        return
    now = time.time()
    elapsed = now - _GSHEET_LAST_WRITE_TS
    if elapsed < GSHEET_WRITE_SLEEP_SECONDS:
        time.sleep(GSHEET_WRITE_SLEEP_SECONDS - elapsed)
    _GSHEET_LAST_WRITE_TS = time.time()


def gsheet_api_call(description, func, *args, **kwargs):
    last_error = None
    for attempt in range(1, GSHEET_MAX_RETRIES + 1):
        try:
            gsheet_write_sleep()
            return func(*args, **kwargs)
        except Exception as e:
            last_error = e
            if not is_gsheet_quota_error(e):
                raise
            wait_seconds = min(90, GSHEET_RETRY_BASE_SECONDS * attempt)
            print(f"  ⚠️ Google Sheet 寫入配額限制：{description}，第 {attempt}/{GSHEET_MAX_RETRIES} 次重試，等待 {wait_seconds:.0f} 秒")
            time.sleep(wait_seconds)
    raise last_error


def get_gsheet_client():
    global _GSHEET_CLIENT
    if _GSHEET_CLIENT is not None:
        return _GSHEET_CLIENT

    service_key = os.getenv("GCP_SERVICE_KEY", "").strip()
    if not service_key:
        return None

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        info = json.loads(service_key)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        _GSHEET_CLIENT = gspread.authorize(creds)
        return _GSHEET_CLIENT
    except Exception as e:
        print(f"  ⚠️ Google Sheet 金鑰初始化失敗：{type(e).__name__}: {e}")
        return None


def get_gsheet_spreadsheet():
    global _GSHEET_SPREADSHEET
    if _GSHEET_SPREADSHEET is not None:
        return _GSHEET_SPREADSHEET

    gc = get_gsheet_client()
    if gc is None:
        return None

    try:
        if GOOGLE_SHEET_ID:
            _GSHEET_SPREADSHEET = gc.open_by_key(GOOGLE_SHEET_ID)
        else:
            _GSHEET_SPREADSHEET = gc.open(GOOGLE_SHEET_NAME)
        return _GSHEET_SPREADSHEET
    except Exception:
        try:
            _GSHEET_SPREADSHEET = gc.create(GOOGLE_SHEET_NAME)
            print(f"  ✅ 已建立 Google Sheet：{GOOGLE_SHEET_NAME}")
            return _GSHEET_SPREADSHEET
        except Exception as e:
            print(f"  ⚠️ Google Sheet 開啟/建立失敗：{type(e).__name__}: {e}")
            return None


def safe_worksheet_title(title):
    title = str(title).strip()
    for ch in [":", "\\", "/", "?", "*", "[", "]"]:
        title = title.replace(ch, "_")
    return title[:100] if title else "工作表"


def cache_sheet_name_from_path(path):
    base = os.path.basename(str(path))
    return CACHE_SHEET_NAME_MAP.get(base, safe_worksheet_title(f"快取_{os.path.splitext(base)[0]}"))


def get_or_create_worksheet(title, rows=100, cols=20):
    sh = get_gsheet_spreadsheet()
    if sh is None:
        return None

    title = safe_worksheet_title(title)
    try:
        return sh.worksheet(title)
    except Exception:
        try:
            return gsheet_api_call(
                f"建立工作表 {title}",
                sh.add_worksheet,
                title=title,
                rows=max(int(rows), 1),
                cols=max(int(cols), 1),
            )
        except Exception as e:
            print(f"  ⚠️ 建立工作表失敗：{title}，原因：{type(e).__name__}: {e}")
            return None


def clean_gsheet_value(value):
    if pd.isna(value):
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value)


def add_gsheet_text_prefix(value):
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    if s.startswith("'") or s.startswith("="):
        return s
    return "'" + s


def strip_gsheet_text_prefix(value):
    if value is None:
        return ""
    s = str(value)
    return s[1:] if s.startswith("'") else s


def dataframe_to_gsheet_values(df):
    if df is None or df.empty:
        return [[""]]

    df2 = df.copy().fillna("")
    headers = list(df2.columns)
    values = [headers]

    text_cols = set()
    for idx, col in enumerate(headers):
        if any(key in str(col) for key in ["代號", "股票代號", "分點代號", "券商代號"]):
            text_cols.add(idx)

    for _, row in df2.iterrows():
        row_values = []
        for idx, value in enumerate(row.tolist()):
            v = clean_gsheet_value(value)
            if idx in text_cols:
                v = add_gsheet_text_prefix(v)
            row_values.append(v)
        values.append(row_values)

    return values


def write_values_to_worksheet(ws, values):
    if ws is None:
        return False
    if not values:
        values = [[""]]

    row_count = max(len(values), 1)
    col_count = max(max((len(row) for row in values), default=1), 1)

    normalized_values = []
    for row in values:
        row = list(row)
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        normalized_values.append(row)

    try:
        gsheet_api_call(f"調整工作表大小 {ws.title}", ws.resize, rows=row_count, cols=col_count)
        gsheet_api_call(f"清除工作表 {ws.title}", ws.clear)

        for start in range(0, len(normalized_values), GSHEET_CHUNK_ROWS):
            chunk = normalized_values[start:start + GSHEET_CHUNK_ROWS]
            start_row = start + 1
            gsheet_api_call(
                f"寫入工作表 {ws.title} A{start_row}",
                ws.update,
                values=chunk,
                range_name=f"A{start_row}",
                value_input_option="USER_ENTERED",
            )
        return True
    except Exception as e:
        print(f"  ⚠️ Google Sheet 寫入失敗：{ws.title}，原因：{type(e).__name__}: {e}")
        return False


def write_dataframe_to_gsheet(df, title, enabled=True):
    if not enabled or not gsheet_enabled():
        return
    if df is None:
        return

    values = dataframe_to_gsheet_values(df)
    max_cols = max(max((len(row) for row in values), default=1), 1)
    ws = get_or_create_worksheet(title, rows=max(len(values), 100), cols=max(max_cols, 20))

    if write_values_to_worksheet(ws, values):
        print(f"  ☁️ 已同步到 Google Sheet：{title}，共 {max(len(values) - 1, 0):,} 筆")


def read_cache_from_gsheet(path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return pd.DataFrame()

    title = cache_sheet_name_from_path(path)
    try:
        sh = get_gsheet_spreadsheet()
        if sh is None:
            return pd.DataFrame()
        ws = sh.worksheet(title)
        values = ws.get_all_values()
        if not values or len(values) < 2:
            return pd.DataFrame()

        headers = [str(h).strip() for h in values[0]]
        rows = values[1:]
        n_cols = len(headers)
        fixed_rows = []
        for row in rows:
            row = list(row)
            if len(row) < n_cols:
                row += [""] * (n_cols - len(row))
            elif len(row) > n_cols:
                row = row[:n_cols]
            fixed_rows.append(row)

        df = pd.DataFrame(fixed_rows, columns=headers).fillna("")
        for col in df.columns:
            if "代號" in str(col):
                df[col] = df[col].map(strip_gsheet_text_prefix)
        print(f"  ☁️ 已從 Google Sheet 讀取快取：{title}，共 {len(df):,} 筆")
        return df
    except Exception:
        return pd.DataFrame()


def write_cache_to_gsheet(df, path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return
    if df is None:
        return
    title = cache_sheet_name_from_path(path)
    write_dataframe_to_gsheet(df, title, enabled=True)


# ══════════════════════════════════════════════════════════════════════
# 快取讀寫
# ══════════════════════════════════════════════════════════════════════


def read_cache_csv(path):
    if not USE_CACHE:
        return pd.DataFrame()

    local_first = os.getenv("GITHUB_ACTIONS", "").strip().lower() != "true"

    if local_first and os.path.exists(path):
        try:
            return pd.read_csv(path, dtype=str, encoding=CACHE_ENCODING).fillna("")
        except Exception as e:
            print(f"  ⚠️ 本機快取讀取失敗：{path}，原因：{type(e).__name__}: {e}")

    df_gsheet = read_cache_from_gsheet(path)
    if df_gsheet is not None and not df_gsheet.empty:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            df_gsheet.to_csv(path, index=False, encoding=CACHE_ENCODING)
        except Exception:
            pass
        return df_gsheet

    if os.path.exists(path):
        try:
            return pd.read_csv(path, dtype=str, encoding=CACHE_ENCODING).fillna("")
        except Exception:
            return pd.DataFrame()

    return pd.DataFrame()


def write_cache_csv(df, path):
    if not USE_CACHE or df is None:
        return

    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        df.to_csv(path, index=False, encoding=CACHE_ENCODING)
        print(f"  💾 已更新快取：{path}，共 {len(df):,} 筆")
    except Exception as e:
        print(f"  ⚠️ 快取寫入失敗：{path}，原因：{type(e).__name__}: {e}")

    write_cache_to_gsheet(df, path)


# ══════════════════════════════════════════════════════════════════════
# 股票清單
# ══════════════════════════════════════════════════════════════════════


def should_skip_security(code, name, row_values):
    if INCLUDE_ETF:
        return False

    code = str(code).strip()
    name = str(name).strip().upper()
    text = " ".join([str(x).strip().upper() for x in row_values])

    if code.startswith("00"):
        return True

    skip_keywords = [
        "ETF", "ETN", "受益證券", "存託憑證", "認購", "認售", "權證",
        "牛證", "熊證", "債券", "指數投資證券",
    ]

    if any(k.upper() in name for k in skip_keywords):
        return True
    if any(k.upper() in text for k in skip_keywords):
        return True

    return False


def fetch_stock_list_live():
    print("【Step 1】抓上市 + 上櫃股票清單...")

    modes = [
        ("上市", "2"),
        ("上櫃", "4"),
    ]

    stocks = []
    seen = set()

    for market_name, mode in modes:
        url = f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}"
        try:
            session = get_thread_session()
            r = session.get(url, headers={"User-Agent": HEADERS_BASE["User-Agent"]}, timeout=30)
            r.raise_for_status()
            r.encoding = "cp950"

            tables = pd.read_html(StringIO(r.text))
            df = tables[0].iloc[2:].reset_index(drop=True)

            count = 0
            for row in df.itertuples(index=False, name=None):
                cell = str(row[0]).strip()

                if "　" in cell:
                    code, name = cell.split("　", 1)
                    code = code.strip()
                    name = name.strip()
                else:
                    m = re.match(r"^(\d{4})\s+(.+)$", cell)
                    if not m:
                        continue
                    code = m.group(1).strip()
                    name = m.group(2).strip()

                if not (len(code) == 4 and code.isdigit()):
                    continue

                if should_skip_security(code, name, row):
                    continue

                if code in seen:
                    continue

                seen.add(code)
                stocks.append({
                    "股票代號": code,
                    "股票名稱": name,
                    "市場": market_name,
                    "更新時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
                })
                count += 1

            print(f"  ✅ {market_name}：{count:,} 檔")

        except Exception as e:
            print(f"  ⚠️ {market_name} 股票清單抓取失敗：{type(e).__name__}: {e}")

    print(f"  ✅ 股票清單完成，共 {len(stocks):,} 檔")
    return pd.DataFrame(stocks)


def load_stock_list():
    cached = read_cache_csv(STOCK_LIST_CACHE_PATH)

    if cached is not None and not cached.empty and not REFRESH_STOCK_LIST:
        print("【Step 1】讀取股票清單快取...")
        print(f"  ✅ 已讀取股票清單快取：{len(cached):,} 檔")
        return cached

    live_df = fetch_stock_list_live()
    if live_df.empty and cached is not None and not cached.empty:
        print("  ⚠️ 即時股票清單失敗，改用既有快取。")
        return cached

    write_cache_csv(live_df, STOCK_LIST_CACHE_PATH)
    return live_df


# ══════════════════════════════════════════════════════════════════════
# MoneyDJ stock-chip0002-5
# ══════════════════════════════════════════════════════════════════════


def build_referer(stock_code):
    return (
        "https://pscnetsecrwd.moneydj.com/b2brwd/page/1000/chip/0002"
        f"?sym=AS{stock_code}&symidxq={stock_code}.TW&symidbsr={stock_code}&bid="
    )


def extract_result_rows(data):
    rows = []
    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            rs = item.get("ResultSet", {})
            if isinstance(rs, dict):
                result = rs.get("Result", [])
                if isinstance(result, list):
                    rows.extend(result)
    elif isinstance(data, dict):
        rs = data.get("ResultSet", {})
        if isinstance(rs, dict):
            result = rs.get("Result", [])
            if isinstance(result, list):
                rows.extend(result)
    return rows


def parse_stock_chip5_row(row, idx, stock, start_date, end_date, run_id, update_time):
    data_date = normalize_date_str(row.get("V1", ""))
    broker_code = str(row.get("V2", "")).strip()
    broker_name = fix_mojibake(row.get("V3", ""))

    buy_shares = to_float(row.get("V4", 0))
    sell_shares = to_float(row.get("V5", 0))
    close_price = to_float(row.get("V6", 0))
    total_volume = to_float(row.get("V7", 0))

    net_shares = buy_shares - sell_shares
    buy_lots = buy_shares / 1000
    sell_lots = sell_shares / 1000
    net_lots = net_shares / 1000

    estimated_buy_amount = buy_shares * close_price
    estimated_sell_amount = sell_shares * close_price
    estimated_net_amount = net_shares * close_price

    turnover_ratio = abs(net_shares) / total_volume * 100 if total_volume > 0 else 0
    side = "買超" if net_shares > 0 else "賣超" if net_shares < 0 else "平盤"
    api_rank_type = "買超榜" if idx < 15 else "賣超榜"
    api_rank = idx + 1 if idx < 15 else idx - 14

    target_label = TARGET_BROKER_CODES.get(normalize_broker_code(broker_code), "")

    return {
        "run_id": run_id,
        "更新時間": update_time,
        "資料日期": data_date,
        "區間起日": start_date,
        "區間迄日": end_date,
        "市場": str(stock.get("市場", "")).strip(),
        "股票代號": str(stock.get("股票代號", "")).strip(),
        "股票名稱": str(stock.get("股票名稱", "")).strip(),
        "分點標籤": target_label,
        "分點代號": broker_code,
        "分點名稱": broker_name,
        "方向": side,
        "API榜別": api_rank_type,
        "API排名": api_rank,
        "買進股數": int(round(buy_shares)),
        "賣出股數": int(round(sell_shares)),
        "買進張數": round(buy_lots, 3),
        "賣出張數": round(sell_lots, 3),
        "買賣超股數": int(round(net_shares)),
        "買賣超張數": round(net_lots, 3),
        "收盤價": close_price,
        "估算買進金額": int(round(estimated_buy_amount)),
        "估算賣出金額": int(round(estimated_sell_amount)),
        "估算買賣超金額": int(round(estimated_net_amount)),
        "成交比重%": round(turnover_ratio, 2),
        "區間總成交股數": int(round(total_volume)),
    }


def fetch_stock_chip5(stock, start_date, end_date, run_id, update_time):
    stock_code = str(stock.get("股票代號", "")).strip()
    if not stock_code:
        return [], []

    params = {
        "a": stock_code,
        "x": "stock-chip0002-5",
        "c": start_date,
        "d": end_date,
        "revision": REVISION,
    }

    headers = dict(HEADERS_BASE)
    headers["Referer"] = build_referer(stock_code)

    last_error = None
    for attempt in range(REQUEST_RETRY + 1):
        try:
            session = get_thread_session()
            r = session.get(STOCK_CHIP5_URL, params=params, headers=headers, timeout=REQUEST_TIMEOUT)

            if r.status_code != 200:
                last_error = f"HTTP {r.status_code}"
                time.sleep(0.5 + attempt)
                continue

            data = r.json()
            rows = extract_result_rows(data)

            if not rows:
                return [], []

            all_top15_rows = []
            hit_rows = []

            for idx, row in enumerate(rows):
                parsed = parse_stock_chip5_row(row, idx, stock, start_date, end_date, run_id, update_time)
                all_top15_rows.append(parsed)

                if normalize_broker_code(parsed.get("分點代號", "")) in TARGET_BROKER_CODES:
                    hit_rows.append(parsed)

            return hit_rows, all_top15_rows

        except Exception as e:
            last_error = f"{type(e).__name__}: {e}"
            time.sleep(0.5 + attempt)

    print(f"  ⚠️ {stock_code} {stock.get('股票名稱', '')} 抓取失敗：{last_error}")
    return [], []


# ══════════════════════════════════════════════════════════════════════
# 資料整理
# ══════════════════════════════════════════════════════════════════════


def merge_hit_cache(old_df, new_df):
    if new_df is None or new_df.empty:
        return old_df if old_df is not None else pd.DataFrame()

    if old_df is None or old_df.empty:
        combined = new_df.copy()
    else:
        combined = pd.concat([old_df, new_df], ignore_index=True)

    key_cols = ["資料日期", "區間起日", "區間迄日", "股票代號", "分點代號", "API榜別"]
    for col in key_cols:
        if col not in combined.columns:
            combined[col] = ""

    combined = combined.drop_duplicates(subset=key_cols, keep="last").reset_index(drop=True)
    return combined


def build_rankings(hit_df):
    if hit_df is None or hit_df.empty:
        empty = pd.DataFrame()
        return empty, empty, empty, empty

    df = hit_df.copy()

    numeric_cols = [
        "買進股數", "賣出股數", "買進張數", "賣出張數", "買賣超股數", "買賣超張數",
        "收盤價", "估算買進金額", "估算賣出金額", "估算買賣超金額", "成交比重%", "區間總成交股數",
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    buy_rank = df[df["買賣超股數"] > 0].copy()
    buy_rank = buy_rank.sort_values(["買賣超股數", "估算買賣超金額"], ascending=[False, False]).reset_index(drop=True)
    if not buy_rank.empty:
        buy_rank.insert(0, "排名", range(1, len(buy_rank) + 1))

    sell_rank = df[df["買賣超股數"] < 0].copy()
    sell_rank = sell_rank.sort_values(["買賣超股數", "估算買賣超金額"], ascending=[True, True]).reset_index(drop=True)
    if not sell_rank.empty:
        sell_rank.insert(0, "排名", range(1, len(sell_rank) + 1))

    abs_rank = df.copy()
    abs_rank["買賣超張數絕對值"] = abs_rank["買賣超張數"].abs()
    abs_rank = abs_rank.sort_values(["買賣超張數絕對值", "估算買賣超金額"], ascending=[False, False]).reset_index(drop=True)
    if not abs_rank.empty:
        abs_rank.insert(0, "排名", range(1, len(abs_rank) + 1))

    summary = df.groupby(["分點標籤", "分點代號", "分點名稱"], dropna=False).agg({
        "股票代號": "count",
        "買進股數": "sum",
        "賣出股數": "sum",
        "買賣超股數": "sum",
        "買進張數": "sum",
        "賣出張數": "sum",
        "買賣超張數": "sum",
        "估算買進金額": "sum",
        "估算賣出金額": "sum",
        "估算買賣超金額": "sum",
    }).reset_index()
    summary = summary.rename(columns={"股票代號": "命中股票數"})
    summary = summary.sort_values("估算買賣超金額", ascending=False).reset_index(drop=True)

    return buy_rank, sell_rank, abs_rank, summary


def build_run_info(start_date, end_date, stock_count, hit_count, all_top15_count, run_id, update_time):
    broker_text = "；".join([
        f"{info.get('券商代號', '')} {info.get('分點名稱', label)}"
        for label, info in TARGET_BROKERS.items()
    ])

    rows = [
        {"項目": "系統名稱", "內容": "MoneyDJ 現股分點籌碼排名"},
        {"項目": "run_id", "內容": run_id},
        {"項目": "更新時間", "內容": update_time},
        {"項目": "區間起日", "內容": start_date},
        {"項目": "區間迄日", "內容": end_date},
        {"項目": "追蹤分點", "內容": broker_text},
        {"項目": "掃描股票數", "內容": stock_count},
        {"項目": "命中筆數", "內容": hit_count},
        {"項目": "全排行TOP15筆數", "內容": all_top15_count},
        {"項目": "資料說明", "內容": "stock-chip0002-5 目前回傳每檔股票區間買超TOP15與賣超TOP15；若分點未進TOP15，該股票不會命中。"},
    ]
    return pd.DataFrame(rows)


def build_color_note():
    return pd.DataFrame([
        {"類別": "買超", "說明": "買賣超股數 > 0，Excel/Google Sheet 排名以買超由大到小。"},
        {"類別": "賣超", "說明": "買賣超股數 < 0，Excel/Google Sheet 排名以賣超由大到小。"},
        {"類別": "估算金額", "說明": "MoneyDJ 此 API 回傳股數與收盤價，本程式使用 股數 × 收盤價 估算金額，不是逐筆成交均價。"},
        {"類別": "資料限制", "說明": "目前 API 回傳 TOP15 + TOP15，不等於所有分點完整資料。"},
    ])


# ══════════════════════════════════════════════════════════════════════
# Excel 輸出與格式
# ══════════════════════════════════════════════════════════════════════


def autosize_worksheet(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 36)


def style_worksheet(ws):
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row in ws.iter_rows(min_row=2):
        direction = None
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "方向" in headers:
            direction_idx = headers.index("方向") + 1
            direction = ws.cell(row[0].row, direction_idx).value

        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if direction == "買超":
                cell.fill = BUY_FILL
            elif direction == "賣超":
                cell.fill = SELL_FILL

    autosize_worksheet(ws)

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 22


def save_excel(output_path, sheets):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            if df is None:
                df = pd.DataFrame()
            safe_name = safe_worksheet_title(sheet_name)[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        style_worksheet(ws)
    wb.save(output_path)
    print(f"  📄 已輸出 Excel：{output_path}")


# ══════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════


def run(start_date=None, end_date=None, stock_limit=0, workers=None):
    if not start_date or not end_date:
        default_start, default_end = default_date_range(LOOKBACK_DAYS)
        start_date = start_date or default_start
        end_date = end_date or default_end

    start_date = normalize_date_str(start_date)
    end_date = normalize_date_str(end_date)
    workers = workers or MAX_WORKERS

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    print("MoneyDJ 現股分點籌碼排名系統")
    print(f"區間：{start_date} ~ {end_date}")
    print("追蹤分點：")
    for label, info in TARGET_BROKERS.items():
        print(f"  - {info.get('券商代號', '')} {info.get('分點名稱', label)}")
    print(f"輸出資料夾：{OUTPUT_DIR}")
    print(f"快取資料夾：{CACHE_DIR}")
    print(f"執行緒：{workers}")

    stock_df = load_stock_list()
    if stock_df.empty:
        raise RuntimeError("股票清單為空，無法執行。")

    stocks = stock_df.to_dict("records")
    if stock_limit and stock_limit > 0:
        stocks = stocks[:stock_limit]
        print(f"  ⚠️ 測試模式：只掃前 {len(stocks):,} 檔")

    print("\n【Step 2】逐檔股票掃描 MoneyDJ stock-chip0002-5...")

    all_hits = []
    all_top15_rows = []
    done = 0

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {
            executor.submit(fetch_stock_chip5, stock, start_date, end_date, run_id, update_time): stock
            for stock in stocks
        }

        for future in as_completed(futures):
            done += 1
            stock = futures[future]

            try:
                hits, top15_rows = future.result()
            except Exception:
                hits, top15_rows = [], []

            if hits:
                all_hits.extend(hits)
                print(f"  ✅ 命中：{stock.get('股票代號', '')} {stock.get('股票名稱', '')}，{len(hits)} 筆")

            if CAPTURE_ALL_TOP15 and top15_rows:
                all_top15_rows.extend(top15_rows)

            if done % 100 == 0:
                print(f"  進度：{done:,}/{len(stocks):,}，目前命中 {len(all_hits):,} 筆")

    print("\n【Step 3】整理排名...")

    hit_df = pd.DataFrame(all_hits)
    all_top15_df = pd.DataFrame(all_top15_rows)

    if hit_df.empty:
        print("⚠️ 沒有找到任何命中資料。")
        print("可能原因：")
        print("1. 9661 富邦-新店、9A9J 永豐金-板新在此區間沒有進任何股票買超 / 賣超 TOP15")
        print("2. stock-chip0002-5 回傳 TOP15，不是全分點完整資料")
        print("3. 日期區間太短，可以改成 2～3 個月測試")
    else:
        print(f"  ✅ 命中資料：{len(hit_df):,} 筆")

    buy_rank, sell_rank, abs_rank, summary_df = build_rankings(hit_df)
    run_info_df = build_run_info(start_date, end_date, len(stocks), len(hit_df), len(all_top15_df), run_id, update_time)
    color_note_df = build_color_note()

    # 合併命中快取
    old_hit_cache = read_cache_csv(HIT_CACHE_PATH)
    hit_cache_df = merge_hit_cache(old_hit_cache, hit_df)
    if not hit_cache_df.empty:
        write_cache_csv(hit_cache_df, HIT_CACHE_PATH)

    if CAPTURE_ALL_TOP15 and not all_top15_df.empty:
        write_cache_csv(all_top15_df, ALL_TOP15_CACHE_PATH)

    # 輸出結果 Excel
    start_safe = start_date.replace("/", "")
    end_safe = end_date.replace("/", "")
    output_xlsx = os.path.join(OUTPUT_DIR, f"stock_branch_{start_safe}_{end_safe}.xlsx")

    result_sheets = {
        "執行資訊": run_info_df,
        "全部命中": hit_df,
        "買超排名": buy_rank,
        "賣超排名": sell_rank,
        "買賣超絕對值排名": abs_rank,
        "分點彙總": summary_df,
        "顏色說明": color_note_df,
    }

    save_excel(output_xlsx, result_sheets)

    # 同步結果到 Google Sheet
    if GSHEET_RESULT_ENABLED and gsheet_enabled():
        print("\n【Step 4】同步結果到 Google Sheet...")
        for sheet_name, df in result_sheets.items():
            write_dataframe_to_gsheet(df, sheet_name, enabled=True)
    else:
        print("\n⚠️ 未設定 GCP_SERVICE_KEY 或 GSHEET_RESULT_ENABLED=0，略過 Google Sheet 結果同步")

    # 終端摘要
    print("\n========== 完成 ==========")
    print(f"掃描股票數：{len(stocks):,}")
    print(f"命中總筆數：{len(hit_df):,}")
    print(f"買超筆數：{len(buy_rank):,}")
    print(f"賣超筆數：{len(sell_rank):,}")
    print(f"Excel：{output_xlsx}")

    if not buy_rank.empty:
        print("\n【買超前 20】")
        cols = ["排名", "股票代號", "股票名稱", "分點代號", "分點名稱", "買進張數", "賣出張數", "買賣超張數", "成交比重%", "收盤價", "估算買賣超金額"]
        cols = [c for c in cols if c in buy_rank.columns]
        print(buy_rank[cols].head(20).to_string(index=False))

    if not sell_rank.empty:
        print("\n【賣超前 20】")
        cols = ["排名", "股票代號", "股票名稱", "分點代號", "分點名稱", "買進張數", "賣出張數", "買賣超張數", "成交比重%", "收盤價", "估算買賣超金額"]
        cols = [c for c in cols if c in sell_rank.columns]
        print(sell_rank[cols].head(20).to_string(index=False))

    return output_xlsx


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", default=os.getenv("START_DATE", ""), help="開始日期，例如 2026/05/07")
    parser.add_argument("--end", default=os.getenv("END_DATE", ""), help="結束日期，例如 2026/06/07")
    parser.add_argument("--limit", type=int, default=int(os.getenv("STOCK_LIMIT", str(STOCK_LIMIT))), help="只掃前 N 檔股票，0 代表全市場")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help="同時抓取執行緒數")
    args = parser.parse_args()

    run(
        start_date=args.start.strip() or None,
        end_date=args.end.strip() or None,
        stock_limit=args.limit,
        workers=args.workers,
    )


if __name__ == "__main__":
    main()
