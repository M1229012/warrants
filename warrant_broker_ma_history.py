"""
FinMind 完整歷史權證分點均線勝率統計
======================================

固定流程：
1. 逐日下載 FinMind 全市場權證分點 Parquet。
2. 驗證、過濾 26 個目標分點，將不含身分判斷的交易事實逐日保存。
3. 由 FinMind Summary 優先、隨附 MOPS 含下市權證官方歷史種子備援，
   日期感知配對身分；GitHub runner 不需連線 MOPS 舊主機。
4. 僅由 compact_daily 重建 ABCDE 事件與跨事件 FIFO。
5. 以 TaiwanStockPriceAdj 建立事件日 MA5/MA10/MA20 條件。
6. 輸出 Parquet 與獨立 Excel，不使用 MoneyDJ 或 Google Sheet。

必要環境變數：
    FINMIND_API_0714=<FinMind token>

必要套件：
    pip install requests pandas openpyxl pyarrow

必要隨附檔案：
    mops_warrant_identity_seed.csv.gz

可選維護參數：
    FORCE_RECALCULATE_STATS=1
    FORCE_REDOWNLOAD_DATES=2023-06-21,2023-06-26
    MA20_FLAT_TOLERANCE_PCT=0.001
"""

from __future__ import annotations

import io
import json
import math
import os
import re
import statistics
import sys
import tempfile
import threading
import time
from collections import Counter, defaultdict, deque
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import requests
    from requests.adapters import HTTPAdapter
except ImportError:  # 允許 --self-test 在未安裝網路套件的環境執行。
    requests = None
    HTTPAdapter = None


PROGRAM_VERSION = "1.5.0"
SCHEMA_VERSION = "ma-history-v2-fact-compact"
IDENTITY_INDEX_SCHEMA_VERSION = "identity-v10-contract-settlement"
MOPS_IDENTITY_SEED_VERSION = "2026-07-30-contract-settlement-v1"
WARRANT_DATASET = "TaiwanStockWarrantTradingDailyReport"
PRICE_DATASET = "TaiwanStockPriceAdj"
WARRANT_PRICE_DATASET = "TaiwanStockPrice"
AMOUNT_THRESH = 1_000_000
FINMIND_DOCUMENTED_START = "2023-06-21"
WARRANT_SUMMARY_HISTORY_START = os.getenv(
    "WARRANT_SUMMARY_HISTORY_START_DATE", "2011-01-03"
).strip()
WARRANT_SUMMARY_CACHE_PREFIX = "TaiwanStockInfoWithWarrantSummary_history"
WARRANT_SUMMARY_TARGET_CACHE_PREFIX = (
    "TaiwanStockInfoWithWarrantSummary_target"
)
MOPS_WARRANT_HISTORY_CACHE_PREFIX = "MOPS_warrant_identity"
MOPS_WARRANT_URL = os.getenv(
    "MOPS_WARRANT_URL",
    "https://mopsov.twse.com.tw/mops/web/ajax_t90sb01",
).strip()
MOPS_WARRANT_LANDING_URL = os.getenv(
    "MOPS_WARRANT_LANDING_URL",
    "https://mopsov.twse.com.tw/mops/web/t90sb01",
).strip()
MOPS_PAGE_SIZE = max(int(os.getenv("MOPS_WARRANT_PAGE_SIZE", "10000")), 100)
MOPS_REQUEST_DELAY_SECONDS = max(
    float(os.getenv("MOPS_REQUEST_DELAY_SECONDS", "0.5")), 0
)
ENABLE_MOPS_NETWORK_REFRESH = os.getenv(
    "ENABLE_MOPS_NETWORK_REFRESH", "0"
).strip().lower() in ("1", "true", "yes")

BASE_DIR = Path(os.getenv("MA_HISTORY_BASE_DIR", ".")).resolve()
MOPS_IDENTITY_SEED_PATH = Path(
    os.getenv(
        "MOPS_IDENTITY_SEED_PATH",
        str(BASE_DIR / "mops_warrant_identity_seed.csv.gz"),
    )
).resolve()
CACHE_ROOT = Path(
    os.getenv("MA_HISTORY_CACHE_DIR", str(BASE_DIR / "warrant_cache" / "ma_history"))
).resolve()
RAW_DAILY_DIR = CACHE_ROOT / "raw_daily"
COMPACT_DAILY_DIR = CACHE_ROOT / "compact_daily"
STOCK_PRICE_DIR = CACHE_ROOT / "stock_price"
WARRANT_PRICE_DIR = CACHE_ROOT / "warrant_price"
RESULTS_DIR = CACHE_ROOT / "results"
METADATA_DIR = CACHE_ROOT / "metadata"
STATE_PATH = CACHE_ROOT / "state.json"
OUTPUT_DIR = Path(os.getenv("OUTPUT_DIR", str(BASE_DIR / "output"))).resolve()

FINMIND_TOKEN = (
    os.getenv("FINMIND_API_0714", "").strip()
    or os.getenv("FINMIND_TOKEN", "").strip()
)
FINMIND_API_BASE = os.getenv(
    "FINMIND_API_BASE", "https://api.finmindtrade.com/api/v4"
).rstrip("/")
FINMIND_DATA_URL = f"{FINMIND_API_BASE}/data"
FINMIND_DATALIST_URL = f"{FINMIND_API_BASE}/datalist"
FINMIND_STORAGE_OBJECTS_URL = f"{FINMIND_API_BASE}/storage_objects"
REQUEST_CONNECT_TIMEOUT = float(os.getenv("FINMIND_CONNECT_TIMEOUT", "10"))
REQUEST_READ_TIMEOUT = float(os.getenv("FINMIND_READ_TIMEOUT", "180"))
MAX_RETRIES = max(int(os.getenv("FINMIND_MAX_RETRIES", "4")), 1)
RETRY_BASE_SECONDS = max(float(os.getenv("FINMIND_RETRY_BASE_SECONDS", "2")), 0.1)
METADATA_CACHE_HOURS = max(float(os.getenv("FINMIND_METADATA_CACHE_HOURS", "12")), 0)
IDENTITY_INDEX_CACHE_HOURS = max(
    float(os.getenv("IDENTITY_INDEX_CACHE_HOURS", "168")), 0
)
PROGRESS_EVERY_DAYS = max(int(os.getenv("PROGRESS_EVERY_DAYS", "20")), 1)
LOCAL_ERROR_CIRCUIT_BREAKER = max(
    int(os.getenv("LOCAL_ERROR_CIRCUIT_BREAKER", "5")), 2
)
MAX_IDENTITY_UNMATCHED_PCT = max(
    float(os.getenv("MAX_IDENTITY_UNMATCHED_PCT", "5")), 0
)
MA20_FLAT_TOLERANCE_PCT = float(
    os.getenv("MA20_FLAT_TOLERANCE_PCT", "0.001")
)
OPEN_POSITION_EVALUATION_MONTHS = max(
    int(os.getenv("OPEN_POSITION_EVALUATION_MONTHS", "2")), 1
)
WARRANT_CASH_SETTLEMENT_TAX_RATE = float(
    os.getenv("WARRANT_CASH_SETTLEMENT_TAX_RATE", "0.001")
)
if not 0 <= WARRANT_CASH_SETTLEMENT_TAX_RATE < 1:
    raise ValueError("WARRANT_CASH_SETTLEMENT_TAX_RATE 必須介於 0（含）與 1（不含）")
FORCE_RECALCULATE_STATS = os.getenv(
    "FORCE_RECALCULATE_STATS", "0"
).strip().lower() in ("1", "true", "yes")
FORCE_REDOWNLOAD_DATES = os.getenv(
    "FORCE_REDOWNLOAD_DATES", ""
).strip()
KEEP_RAW_DAILY = os.getenv(
    "KEEP_RAW_DAILY", "1"
).strip().lower() not in ("0", "false", "no")

AMOUNT_CLASS_SPECS = [
    ("A", "100萬至未滿160萬", 1_000_000, 1_600_000),
    ("B", "160萬至未滿250萬", 1_600_000, 2_500_000),
    ("C", "250萬至未滿500萬", 2_500_000, 5_000_000),
    ("D", "500萬至未滿1000萬", 5_000_000, 10_000_000),
    ("E", "1000萬以上", 10_000_000, None),
]
AMOUNT_CLASS_LABELS = {code: label for code, label, _, _ in AMOUNT_CLASS_SPECS}

TARGET_PATTERNS = {
    "富邦公益": r"富邦.*公益",
    "富邦敦南": r"富邦.*敦南",
    "富邦仁愛": r"富邦.*仁愛",
    "新光": r"^新光$",
    "永豐金內湖": r"永豐.*內湖",
    "永豐金竹北": r"永豐.*竹北",
    "永豐金竹科": r"永豐.*竹科",
    "永豐金市政": r"永豐.*市政",
    "永豐金信義": r"永豐.*信義",
    "華南永昌台中": r"華南.*台中",
    "華南永昌淡水": r"華南.*淡水",
    "福邦證券": r"^福邦證券",
    "群益東大": r"群益.*東大",
    "群益金鼎古亭": r"群益.*古亭",
    "群益金鼎新竹": r"群益.*新竹",
    "元大內湖民權": r"元大.*(內湖.*民權|民權)",
    "元大南屯": r"元大.*南屯",
    "元大汐止": r"元大.*汐止",
    "元大虎尾": r"元大.*虎尾",
    "元大彰化民生": r"元大.*彰化民生",
    "兆豐板橋": r"兆豐.*板橋",
    "凱基士林": r"凱基.*士林",
    "凱基中山": r"凱基.*中山",
    "國票敦北法人": r"國票.*(敦北|法人)",
    "統一三多": r"統一.*三多",
    "第一金中壢": r"第一金.*中壢",
}

FALLBACK_BROKERS = {
    "富邦公益": ("富邦-公益", "961F"),
    "富邦敦南": ("富邦-敦南", "9663"),
    "富邦仁愛": ("富邦-仁愛", "9676"),
    "新光": ("新光", "8560"),
    "永豐金內湖": ("永豐金-內湖", "9A9g"),
    "永豐金竹北": ("永豐金-竹北", "9A9P"),
    "永豐金竹科": ("永豐金-竹科", "9A9X"),
    "永豐金市政": ("永豐金-市政", "9A9W"),
    "永豐金信義": ("永豐金-信義", "9A9R"),
    "華南永昌台中": ("華南永昌-台中", "9302"),
    "華南永昌淡水": ("華南永昌-淡水", "9316"),
    "福邦證券": ("福邦證券", "6480"),
    "群益東大": ("群益金鼎-東大", "9135"),
    "群益金鼎古亭": ("群益金鼎-古亭", "918C"),
    "群益金鼎新竹": ("群益金鼎-新竹", "9186"),
    "元大內湖民權": ("元大-內湖民權", "9867"),
    "元大南屯": ("元大-南屯", "9853"),
    "元大汐止": ("元大-汐止", "989Q"),
    "元大虎尾": ("元大-虎尾", "980l"),
    "元大彰化民生": ("元大-彰化民生", "989J"),
    "兆豐板橋": ("兆豐-板橋", "700B"),
    "凱基士林": ("凱基-士林", "9238"),
    "凱基中山": ("凱基-中山", "9229"),
    "國票敦北法人": ("國票-敦北法人", "779c"),
    "統一三多": ("統一-三多", "585Q"),
    "第一金中壢": ("第一金-中壢", "538Y"),
}

RAW_REQUIRED_COLUMNS = {
    "securities_trader",
    "price",
    "buy",
    "sell",
    "securities_trader_id",
    "stock_id",
    "date",
}
COMPACT_FACT_COLUMNS = [
    "權證代號",
    "分點",
    "分點名稱",
    "券商代號",
    "日期",
    "買進股數",
    "賣出股數",
    "買進金額",
    "賣出金額",
    "買超股數",
    "買超金額",
]
COMPACT_IDENTITY_COLUMNS = [
    "權證名稱",
    "權證類型",
    "標的股",
    "標的名稱",
    "身分配對狀態",
    "身分配對錯誤原因",
]
COMPACT_COLUMNS = COMPACT_FACT_COLUMNS + COMPACT_IDENTITY_COLUMNS

MA_POSITION_NAMES = {
    "000": "低於全部均線",
    "001": "僅高於MA20",
    "010": "僅高於MA10",
    "011": "高於MA10與MA20",
    "100": "僅高於MA5",
    "101": "高於MA5與MA20",
    "110": "高於MA5與MA10",
    "111": "高於全部均線",
}

CONDITION_SPECS = [
    ("高於MA5", lambda d: d["高於MA5"] == True),
    ("低於MA5", lambda d: d["高於MA5"] == False),
    ("高於MA10", lambda d: d["高於MA10"] == True),
    ("低於MA10", lambda d: d["高於MA10"] == False),
    ("高於MA20", lambda d: d["高於MA20"] == True),
    ("低於MA20", lambda d: d["高於MA20"] == False),
    ("剛站上MA5", lambda d: d["剛站上MA5"] == True),
    ("剛站上MA10", lambda d: d["剛站上MA10"] == True),
    ("剛站上MA20", lambda d: d["剛站上MA20"] == True),
    ("剛跌破MA5", lambda d: d["剛跌破MA5"] == True),
    ("剛跌破MA10", lambda d: d["剛跌破MA10"] == True),
    ("剛跌破MA20", lambda d: d["剛跌破MA20"] == True),
    ("高於全部均線", lambda d: d["MA位置代碼"] == "111"),
    ("低於全部均線", lambda d: d["MA位置代碼"] == "000"),
    ("均線多頭排列", lambda d: d["均線多頭排列"] == True),
    ("均線空頭排列", lambda d: d["均線空頭排列"] == True),
    ("MA20上彎", lambda d: d["MA20方向"] == "上彎"),
    ("MA20走平", lambda d: d["MA20方向"] == "走平"),
    ("MA20下彎", lambda d: d["MA20方向"] == "下彎"),
]

_thread_local = threading.local()


class FinMindQuotaExceeded(RuntimeError):
    """FinMind 回傳 402；歷史分片可保留，等待額度恢復後續抓。"""


def ensure_directories() -> None:
    for path in (
        RAW_DAILY_DIR,
        COMPACT_DAILY_DIR,
        STOCK_PRICE_DIR,
        WARRANT_PRICE_DIR,
        RESULTS_DIR,
        METADATA_DIR,
        OUTPUT_DIR,
    ):
        path.mkdir(parents=True, exist_ok=True)


def normalize_code(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text.endswith(".0") and text[:-2].isalnum():
        text = text[:-2]
    return re.sub(r"\s+", "", text)


def normalize_broker_code(value: Any) -> str:
    return normalize_code(value)


def normalize_name(value: Any) -> str:
    return str(value or "").strip().replace(" ", "").replace("　", "")


def parse_date(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    parsed = pd.to_datetime(value, errors="coerce")
    return None if pd.isna(parsed) else parsed.to_pydatetime()


def iso_date(value: Any) -> str:
    parsed = parse_date(value)
    return parsed.strftime("%Y-%m-%d") if parsed else ""


def atomic_write_json(payload: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(
        f"{path.name}.tmp.{os.getpid()}.{threading.get_ident()}"
    )
    try:
        with temp_path.open("w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2, sort_keys=True)
            handle.flush()
            os.fsync(handle.fileno())
        with temp_path.open("r", encoding="utf-8") as handle:
            json.load(handle)
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def atomic_write_parquet(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_name(
        f"{path.name}.tmp.{os.getpid()}.{threading.get_ident()}"
    )
    try:
        frame.to_parquet(temp_path, index=False)
        check = pd.read_parquet(temp_path)
        if list(check.columns) != list(frame.columns):
            raise RuntimeError(f"Parquet 寫入驗證欄位不一致：{path}")
        os.replace(temp_path, path)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def read_parquet(path: Path) -> pd.DataFrame:
    if not path.exists() or path.stat().st_size <= 0:
        raise FileNotFoundError(str(path))
    return pd.read_parquet(path)


def default_state() -> dict[str, Any]:
    return {
        "schema_version": SCHEMA_VERSION,
        "program_version": PROGRAM_VERSION,
        "first_available_date": "",
        "latest_available_date": "",
        "successful_dates": [],
        "failed_dates": [],
        "confirmed_empty_dates": [],
        "last_completed_date": "",
        "last_statistics_time": "",
        "raw_rows_by_date": {},
        "compact_rows_by_date": {},
        "deduplicated_rows_by_date": {},
        "invalid_numeric_rows_by_date": {},
        "failure_reasons": {},
        "availability_discovery_source": "",
        "current_stage": "",
        "stage_updated_at": "",
    }


def load_state() -> dict[str, Any]:
    state = default_state()
    if STATE_PATH.exists():
        try:
            saved = json.loads(STATE_PATH.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                state.update(saved)
        except Exception as exc:
            damaged = STATE_PATH.with_suffix(
                f".damaged-{datetime.now():%Y%m%d%H%M%S}.json"
            )
            os.replace(STATE_PATH, damaged)
            print(f"⚠️ state.json 損壞，已移至 {damaged.name}：{exc}")
    for key in ("successful_dates", "failed_dates", "confirmed_empty_dates"):
        state[key] = sorted({iso_date(x) for x in state.get(key, []) if iso_date(x)})
    for key in (
        "raw_rows_by_date",
        "compact_rows_by_date",
        "deduplicated_rows_by_date",
        "invalid_numeric_rows_by_date",
    ):
        if not isinstance(state.get(key), dict):
            state[key] = {}
    if not isinstance(state.get("failure_reasons"), dict):
        state["failure_reasons"] = {}
    state["schema_version"] = SCHEMA_VERSION
    state["program_version"] = PROGRAM_VERSION
    return state


def save_state(state: dict[str, Any]) -> None:
    for key in ("successful_dates", "failed_dates", "confirmed_empty_dates"):
        state[key] = sorted(set(state.get(key, [])))
    atomic_write_json(state, STATE_PATH)


def set_run_stage(state: dict[str, Any], stage: str) -> None:
    state["current_stage"] = stage
    state["stage_updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_state(state)


def get_session() -> Any:
    if requests is None or HTTPAdapter is None:
        raise RuntimeError("缺少 requests；請執行 pip install requests")
    session = getattr(_thread_local, "session", None)
    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=8, pool_maxsize=8, max_retries=0)
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        _thread_local.session = session
    return session


def finmind_headers() -> dict[str, str]:
    if not FINMIND_TOKEN:
        raise RuntimeError("缺少 FINMIND_API_0714（或 FINMIND_TOKEN）環境變數")
    return {
        "Authorization": f"Bearer {FINMIND_TOKEN}",
        "User-Agent": "warrant-broker-ma-history/1.0",
        "Accept": "application/json, application/octet-stream, */*",
    }


def request_with_retries(
    url: str,
    *,
    params: dict[str, Any],
    description: str,
    attempts: Optional[int] = None,
) -> Any:
    total_attempts = max(int(attempts or MAX_RETRIES), 1)
    last_error: Optional[BaseException] = None
    for attempt in range(1, total_attempts + 1):
        try:
            response = get_session().get(
                url,
                params=params,
                headers=finmind_headers(),
                timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
            )
            if response.status_code == 402:
                raise FinMindQuotaExceeded(
                    "FinMind API 使用次數已達上限（HTTP 402）"
                )
            if response.status_code in (401, 403):
                raise RuntimeError(
                    f"FinMind 權限或額度錯誤（HTTP {response.status_code}）"
                )
            if response.status_code >= 500 or response.status_code in (408, 429):
                response.raise_for_status()
            return response
        except FinMindQuotaExceeded:
            raise
        except Exception as exc:
            last_error = exc
            if attempt < total_attempts:
                delay = min(RETRY_BASE_SECONDS * (2 ** (attempt - 1)), 30)
                print(
                    f"  ⚠️ {description} 第 {attempt}/{total_attempts} 次失敗："
                    f"{type(exc).__name__}: {exc}；{delay:.1f} 秒後重試"
                )
                time.sleep(delay)
    raise RuntimeError(
        f"{description} 重試 {total_attempts} 次仍失敗："
        f"{type(last_error).__name__}: {last_error}"
    )


def metadata_cache_path(cache_key: str) -> Path:
    safe_key = re.sub(r"[^0-9A-Za-z_.-]", "_", str(cache_key or "").strip())
    if not safe_key:
        raise ValueError("metadata cache key 不得為空")
    return METADATA_DIR / f"{safe_key}.parquet"


def fetch_dataset(
    dataset: str,
    *,
    params: Optional[dict[str, Any]] = None,
    cache: bool = False,
    force: bool = False,
    cache_key: str = "",
) -> pd.DataFrame:
    started_at = time.perf_counter()
    parameter_note = ""
    if params:
        interesting = [
            f"{key}={value}"
            for key, value in params.items()
            if key in ("data_id", "start_date", "end_date") and value not in (None, "")
        ]
        if interesting:
            parameter_note = f"（{', '.join(interesting)}）"
    cache_path = metadata_cache_path(cache_key or dataset)
    if cache and cache_path.exists() and not force:
        age_hours = (time.time() - cache_path.stat().st_mtime) / 3600
        if age_hours <= METADATA_CACHE_HOURS:
            try:
                cached = read_parquet(cache_path)
                if not cached.empty:
                    print(
                        f"  📦 {dataset}{parameter_note}：沿用 metadata 快取 "
                        f"{len(cached):,} 筆｜{time.perf_counter() - started_at:.1f} 秒",
                        flush=True,
                    )
                    return cached
            except Exception:
                pass
    print(f"  ⏳ 下載 FinMind {dataset}{parameter_note}...", flush=True)
    query: dict[str, Any] = {"dataset": dataset}
    if params:
        query.update({key: value for key, value in params.items() if value not in (None, "")})
    try:
        response = request_with_retries(
            FINMIND_DATA_URL, params=query, description=f"FinMind {dataset}"
        )
        payload = response.json()
        status = int(payload.get("status", response.status_code) or response.status_code)
        if status != 200:
            raise RuntimeError(payload.get("msg") or f"status={status}")
        frame = pd.DataFrame(payload.get("data") or [])
        if frame.empty:
            raise RuntimeError("回傳空資料")
        if cache:
            atomic_write_parquet(frame, cache_path)
        print(
            f"  ✅ FinMind {dataset}{parameter_note}：{len(frame):,} 筆｜"
            f"{time.perf_counter() - started_at:.1f} 秒",
            flush=True,
        )
        return frame
    except Exception:
        if cache and cache_path.exists():
            cached = read_parquet(cache_path)
            if not cached.empty:
                print(
                    f"  ⚠️ {dataset} 更新失敗，沿用既有中繼資料快取 "
                    f"{len(cached):,} 筆｜{time.perf_counter() - started_at:.1f} 秒",
                    flush=True,
                )
                return cached
        raise


def warrant_summary_chunk_ranges(
    start_date: str,
    end_date: str,
) -> list[tuple[str, str]]:
    start = parse_date(start_date)
    end = parse_date(end_date)
    if start is None or end is None or start.date() > end.date():
        raise ValueError(
            f"權證 Summary 歷史範圍無效：{start_date}～{end_date}"
        )
    cursor = start.date()
    final_day = end.date()
    ranges: list[tuple[str, str]] = []
    while cursor <= final_day:
        next_year = date(cursor.year + 1, 1, 1)
        chunk_end = min(next_year, final_day)
        ranges.append((cursor.isoformat(), chunk_end.isoformat()))
        if chunk_end >= final_day:
            break
        # 相鄰年度在 1 月 1 日重疊一天；合併時去重，可兼容 API 的
        # end_date 為含首不含尾或首尾皆含兩種實作。
        cursor = chunk_end
    return ranges


def fetch_warrant_summary_history(
    start_date: str = WARRANT_SUMMARY_HISTORY_START,
    end_date: str = "",
) -> pd.DataFrame:
    """明確帶日期、逐年下載完整權證代號重用歷史。"""
    normalized_start = iso_date(start_date)
    normalized_end = iso_date(end_date) or datetime.now().date().isoformat()
    if not normalized_start or not normalized_end:
        raise ValueError("權證 Summary 歷史起訖日期無效")
    required = {"stock_id", "target_stock_id", "type", "date", "end_date"}
    frames: list[pd.DataFrame] = []
    ranges = warrant_summary_chunk_ranges(normalized_start, normalized_end)
    print(
        f"  ⏳ 分段下載權證標的完整歷史：{normalized_start}～"
        f"{normalized_end}｜{len(ranges)} 段",
        flush=True,
    )
    for index, (chunk_start, chunk_end) in enumerate(ranges, start=1):
        cache_key = (
            f"{WARRANT_SUMMARY_CACHE_PREFIX}_"
            f"{chunk_start.replace('-', '')}"
        )
        frame = fetch_dataset(
            "TaiwanStockInfoWithWarrantSummary",
            params={
                "start_date": chunk_start,
                "end_date": chunk_end,
            },
            cache=True,
            cache_key=cache_key,
        )
        missing = required - set(frame.columns)
        if missing:
            raise RuntimeError(
                f"權證 Summary {chunk_start}～{chunk_end} 缺少欄位："
                f"{sorted(missing)}"
            )
        work = frame.copy()
        work["date"] = work["date"].map(iso_date)
        work["end_date"] = work["end_date"].map(iso_date)
        if (
            work["date"].eq("").any()
            or work["end_date"].eq("").any()
            or work["date"].isna().any()
            or work["end_date"].isna().any()
        ):
            raise RuntimeError(
                f"權證 Summary {chunk_start}～{chunk_end} 含無效上市／下市日期"
            )
        outside = (work["date"] < chunk_start) | (work["date"] > chunk_end)
        if outside.any():
            raise RuntimeError(
                "FinMind 權證 Summary 未遵守請求的年度範圍："
                f"要求 {chunk_start}～{chunk_end}，"
                f"實際 {work['date'].min()}～{work['date'].max()}。"
                "禁止快取不完整或跨段資料。"
            )
        frames.append(work)
        print(
            f"    Summary 歷史進度：{index}/{len(ranges)}｜"
            f"{chunk_start}～{chunk_end}｜{len(work):,} 筆",
            flush=True,
        )
    if not frames:
        raise RuntimeError("權證 Summary 完整歷史沒有任何分段資料")
    combined = pd.concat(frames, ignore_index=True)
    dedup_keys = [
        "stock_id",
        "target_stock_id",
        "type",
        "date",
        "end_date",
    ]
    combined = (
        combined.drop_duplicates(dedup_keys, keep="last")
        .sort_values(["date", "stock_id", "end_date"])
        .reset_index(drop=True)
    )
    earliest = min(combined["date"], default="")
    if not earliest or earliest > FINMIND_DOCUMENTED_START:
        raise RuntimeError(
            "權證 Summary 歷史仍不完整："
            f"最早上市日 {earliest or '-'} 晚於分點資料起始日 "
            f"{FINMIND_DOCUMENTED_START}。禁止建立身分索引。"
        )
    print(
        f"  ✅ 權證標的完整歷史合併：{len(combined):,} 筆｜"
        f"最早上市 {earliest}｜最新上市 {max(combined['date'], default='-')}",
        flush=True,
    )
    return combined


def warrant_summary_target_cache_paths(
    target_code: str,
) -> tuple[Path, Path]:
    target = normalize_code(target_code)
    if not target:
        raise ValueError("母股代號不得為空")
    stem = f"{WARRANT_SUMMARY_TARGET_CACHE_PREFIX}_{target}"
    return (
        metadata_cache_path(stem),
        METADATA_DIR / f"{stem}.empty.json",
    )


def fetch_warrant_summary_target_history(
    target_code: str,
    start_date: str = WARRANT_SUMMARY_HISTORY_START,
    end_date: str = "",
) -> pd.DataFrame:
    """
    依母股 data_id 反查權證歷史。

    FinMind 官方文件明確指出 data_id 為母股代號，且此查法包含已到期及
    代號重用的歷史權證。不帶 data_id 的全表在實際服務上可能漏掉舊區間，
    因此此函式作為精確補抓；回傳 target_stock_id 不一致時一律拒收。
    """
    target = normalize_code(target_code)
    normalized_start = iso_date(start_date)
    normalized_end = iso_date(end_date) or datetime.now().date().isoformat()
    if not target or not normalized_start or not normalized_end:
        raise ValueError("母股分片查詢參數無效")
    cache_path, empty_marker = warrant_summary_target_cache_paths(target)
    max_age = IDENTITY_INDEX_CACHE_HOURS
    for path, is_empty_marker in (
        (cache_path, False),
        (empty_marker, True),
    ):
        if not path.exists() or path.stat().st_size <= 0:
            continue
        age_hours = max((time.time() - path.stat().st_mtime) / 3600, 0)
        if age_hours > max_age:
            continue
        if is_empty_marker:
            print(
                f"    📦 母股 {target}：沿用空結果標記｜"
                f"快取年齡 {age_hours:.1f} 小時",
                flush=True,
            )
            return pd.DataFrame(
                columns=[
                    "stock_id",
                    "target_stock_id",
                    "type",
                    "date",
                    "end_date",
                ]
            )
        cached = read_parquet(path)
        if not cached.empty:
            print(
                f"    📦 母股 {target}：沿用歷史分片 "
                f"{len(cached):,} 筆｜快取年齡 {age_hours:.1f} 小時",
                flush=True,
            )
            return cached

    started_at = time.perf_counter()
    print(f"    ⏳ 母股 {target}：反查權證完整歷史...", flush=True)
    response = request_with_retries(
        FINMIND_DATA_URL,
        params={
            "dataset": "TaiwanStockInfoWithWarrantSummary",
            "data_id": target,
            "start_date": normalized_start,
            "end_date": normalized_end,
        },
        description=f"FinMind 權證 Summary 母股 {target}",
    )
    payload = response.json()
    status = int(payload.get("status", response.status_code) or response.status_code)
    if status != 200:
        raise RuntimeError(payload.get("msg") or f"status={status}")
    frame = pd.DataFrame(payload.get("data") or [])
    if frame.empty:
        atomic_write_json(
            {
                "target_stock_id": target,
                "start_date": normalized_start,
                "end_date": normalized_end,
                "queried_at": datetime.now().isoformat(timespec="seconds"),
            },
            empty_marker,
        )
        print(
            f"    ℹ️ 母股 {target}：沒有權證歷史｜"
            f"{time.perf_counter() - started_at:.1f} 秒",
            flush=True,
        )
        return pd.DataFrame(
            columns=[
                "stock_id",
                "target_stock_id",
                "type",
                "date",
                "end_date",
            ]
        )

    required = {"stock_id", "target_stock_id", "type", "date", "end_date"}
    missing = required - set(frame.columns)
    if missing:
        raise RuntimeError(
            f"母股 {target} 的權證 Summary 缺少欄位：{sorted(missing)}"
        )
    work = frame.copy()
    work["stock_id"] = work["stock_id"].map(normalize_code)
    work["target_stock_id"] = work["target_stock_id"].map(normalize_code)
    work["date"] = work["date"].map(iso_date)
    work["end_date"] = work["end_date"].map(iso_date)
    if (
        work["date"].eq("").any()
        or work["end_date"].eq("").any()
        or work["date"].isna().any()
        or work["end_date"].isna().any()
    ):
        raise RuntimeError(f"母股 {target} 的權證 Summary 含無效日期")
    wrong_targets = sorted(
        {
            value
            for value in work["target_stock_id"].astype(str)
            if value != target
        }
    )
    if wrong_targets:
        raise RuntimeError(
            f"FinMind data_id={target} 未依母股過濾，"
            f"回傳其他 target_stock_id={wrong_targets[:5]}；拒絕補入索引"
        )
    outside = (work["date"] < normalized_start) | (
        work["date"] > normalized_end
    )
    if outside.any():
        raise RuntimeError(
            f"母股 {target} 的權證 Summary 超出請求上市日期範圍："
            f"{work['date'].min()}～{work['date'].max()}"
        )
    work = (
        work.drop_duplicates(
            ["stock_id", "target_stock_id", "type", "date", "end_date"],
            keep="last",
        )
        .sort_values(["date", "stock_id", "end_date"])
        .reset_index(drop=True)
    )
    atomic_write_parquet(work, cache_path)
    if empty_marker.exists():
        empty_marker.unlink(missing_ok=True)
    print(
        f"    ✅ 母股 {target}：{len(work):,} 筆｜"
        f"{time.perf_counter() - started_at:.1f} 秒",
        flush=True,
    )
    return work


def trading_dates() -> list[str]:
    frame = fetch_dataset("TaiwanStockTradingDate", cache=True)
    if "date" not in frame.columns:
        raise RuntimeError("TaiwanStockTradingDate 缺少 date 欄位")
    dates = sorted({iso_date(value) for value in frame["date"] if iso_date(value)})
    today = datetime.now().date()
    return [value for value in dates if parse_date(value).date() <= today]


def discover_dataset_start_from_metadata() -> tuple[str, str]:
    """
    優先由 FinMind datalist metadata 取得資料起始日。

    datalist 在不同 FinMind 版本的巢狀欄位略有差異，因此以資料集名稱定位
    record，再從 start/min/first/begin 類欄位抽日期。若 metadata 未提供日期，
    才退回官方文件所載的資料集起始日；實際全市場日檔仍逐日由
    storage_objects 回應確認。
    """

    def walk(value: Any) -> Iterable[dict[str, Any]]:
        if isinstance(value, dict):
            yield value
            for child in value.values():
                yield from walk(child)
        elif isinstance(value, list):
            for child in value:
                yield from walk(child)

    started_at = time.perf_counter()
    print("  ⏳ 查詢 FinMind datalist 資料範圍...", flush=True)
    try:
        response = request_with_retries(
            FINMIND_DATALIST_URL,
            params={"dataset": WARRANT_DATASET},
            description="FinMind datalist metadata",
        )
        if response.status_code == 200:
            payload = response.json()
            candidates: list[str] = []
            for record in walk(payload):
                if WARRANT_DATASET not in {
                    str(item) for item in record.values() if isinstance(item, str)
                }:
                    continue
                for key, value in record.items():
                    key_lower = str(key).lower()
                    if any(
                        token in key_lower
                        for token in ("start", "first", "begin", "minimum", "min_date")
                    ):
                        parsed = iso_date(value)
                        if parsed:
                            candidates.append(parsed)
            if candidates:
                print(
                    f"  ✅ datalist 起始日：{min(candidates)}｜"
                    f"{time.perf_counter() - started_at:.1f} 秒",
                    flush=True,
                )
                return min(candidates), "FinMind datalist metadata"
    except Exception as exc:
        print(
            f"  ⚠️ datalist 未提供可用起始日，改用官方資料集範圍："
            f"{type(exc).__name__}: {exc}"
        )
    print(
        f"  ℹ️ datalist 無可用起始日，採官方備援 {FINMIND_DOCUMENTED_START}｜"
        f"{time.perf_counter() - started_at:.1f} 秒",
        flush=True,
    )
    return FINMIND_DOCUMENTED_START, "FinMind 官方資料集範圍備援"


def _response_message(response: Any) -> str:
    try:
        payload = response.json()
        if isinstance(payload, dict):
            return str(payload.get("msg") or payload.get("message") or payload)
    except Exception:
        pass
    return (response.text or "")[:500]


def response_is_parquet(response: Any) -> bool:
    content = response.content or b""
    content_type = str(response.headers.get("Content-Type", "")).lower()
    return (
        "parquet" in content_type
        or (len(content) >= 8 and content[:4] == b"PAR1" and content[-4:] == b"PAR1")
    )


@dataclass
class DownloadResult:
    status: str
    frame: pd.DataFrame
    reason: str = ""


def download_warrant_day(day: str) -> DownloadResult:
    last_reason = ""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = request_with_retries(
                FINMIND_STORAGE_OBJECTS_URL,
                params={"dataset": WARRANT_DATASET, "date": day},
                description=f"全市場權證分點 {day}",
                attempts=1,
            )
            if response_is_parquet(response):
                frame = pd.read_parquet(io.BytesIO(response.content))
                missing = RAW_REQUIRED_COLUMNS - set(frame.columns)
                if missing:
                    raise RuntimeError(f"schema 缺少欄位：{sorted(missing)}")
                if frame.empty:
                    raise RuntimeError("Parquet 為空，無法確認為有效完整日檔")
                actual_dates = {
                    iso_date(value) for value in frame["date"] if iso_date(value)
                }
                if actual_dates != {day}:
                    raise RuntimeError(
                        f"日檔日期不符：預期 {day}，實際 {sorted(actual_dates)}"
                    )
                return DownloadResult("ok", frame)

            message = _response_message(response)
            message_lower = message.lower()
            explicit_no_data = (
                response.status_code in (400, 404)
                and any(
                    token in message_lower
                    for token in (
                        "not found",
                        "no data",
                        "does not exist",
                        "尚未",
                        "無資料",
                        "不存在",
                    )
                )
            )
            if explicit_no_data:
                return DownloadResult("empty", pd.DataFrame(), message)
            raise RuntimeError(
                f"HTTP {response.status_code}，內容不是 Parquet：{message}"
            )
        except Exception as exc:
            last_reason = f"{type(exc).__name__}: {exc}"
            if attempt < MAX_RETRIES:
                delay = min(RETRY_BASE_SECONDS * (2 ** (attempt - 1)), 30)
                print(
                    f"  ⚠️ 全市場權證分點 {day} 第 {attempt}/{MAX_RETRIES} 次失敗："
                    f"{last_reason}；{delay:.1f} 秒後重試"
                )
                time.sleep(delay)
    return DownloadResult("error", pd.DataFrame(), last_reason)


def match_target_broker(name: str) -> str:
    for label, pattern in TARGET_PATTERNS.items():
        if re.search(pattern, str(name or "").strip()):
            return label
    return ""


def build_broker_map() -> dict[str, tuple[str, str]]:
    frame = fetch_dataset("TaiwanSecuritiesTraderInfo", cache=True)
    by_code: dict[str, tuple[str, str]] = {}
    if {"securities_trader_id", "securities_trader"}.issubset(frame.columns):
        for code, name in frame[
            ["securities_trader_id", "securities_trader"]
        ].itertuples(index=False, name=None):
            normalized = normalize_broker_code(code)
            if normalized:
                by_code[normalized] = (str(name or "").strip(), str(code or "").strip())
    result: dict[str, tuple[str, str]] = {}
    for label in TARGET_PATTERNS:
        fallback_name, fallback_code = FALLBACK_BROKERS[label]
        result[label] = by_code.get(
            normalize_broker_code(fallback_code), (fallback_name, fallback_code)
        )
    for _, (name, raw_code) in by_code.items():
        label = match_target_broker(name)
        if label:
            canonical = FALLBACK_BROKERS[label][1]
            result[label] = (name, canonical or raw_code)
    return result


def build_stock_master(frame: pd.DataFrame) -> dict[str, str]:
    if not {"stock_id", "stock_name"}.issubset(frame.columns):
        return {}
    work = frame.copy()
    work["_date"] = pd.to_datetime(
        work["date"] if "date" in work.columns else "", errors="coerce"
    )
    work["_row"] = range(len(work))
    work["stock_id"] = work["stock_id"].map(normalize_code)
    work["stock_name"] = work["stock_name"].fillna("").astype(str).str.strip()
    work = work[(work["stock_id"] != "") & (work["stock_name"] != "")]
    work = work.sort_values(["stock_id", "_date", "_row"]).drop_duplicates(
        "stock_id", keep="last"
    )
    return dict(zip(work["stock_id"], work["stock_name"]))


class StockAliasResolver:
    """以最長前綴字典取代逐筆線性掃描，並快取重複權證名稱。"""

    def __init__(
        self,
        by_prefix: dict[str, tuple[str, str, bool]],
        *,
        memo_limit: int = 200_000,
    ):
        self.by_prefix = by_prefix
        self.prefix_lengths = sorted(
            {len(prefix) for prefix in by_prefix}, reverse=True
        )
        self.memo: dict[str, Optional[tuple[str, str, int, bool]]] = {}
        self.memo_limit = max(int(memo_limit), 1)

    def resolve(self, warrant_name: str) -> Optional[tuple[str, str, int, bool]]:
        name = normalize_name(warrant_name)
        if not name:
            return None
        if name in self.memo:
            return self.memo[name]

        if not name.startswith(
            ("台灣50正", "臺灣50正", "台灣50反", "臺灣50反")
        ) and name.startswith(
            ("台灣50", "臺灣50", "元大台灣50", "元大臺灣50")
        ):
            result: Optional[tuple[str, str, int, bool]] = (
                "0050",
                "元大台灣50",
                4,
                True,
            )
        else:
            result = None
            for length in self.prefix_lengths:
                if length > len(name):
                    continue
                matched = self.by_prefix.get(name[:length])
                if matched is not None:
                    code, stock_name, exact = matched
                    result = code, stock_name, length, exact
                    break

        if len(self.memo) >= self.memo_limit:
            self.memo.clear()
        self.memo[name] = result
        return result


def stock_alias_resolver(stock_master: dict[str, str]) -> StockAliasResolver:
    exact_names = {normalize_name(name) for name in stock_master.values() if normalize_name(name)}
    suffixes = (
        "半導體", "科技", "電子", "光電", "精密", "材料", "生技", "醫療",
        "資訊", "電腦", "通信", "通訊", "電機", "機械", "工業", "實業",
        "企業", "國際", "控股", "投控", "建設", "營造", "食品", "鋼鐵",
    )
    candidates: dict[str, list[tuple[str, str, bool]]] = defaultdict(list)
    for code, name in stock_master.items():
        normalized = normalize_name(name)
        aliases = {normalized}
        stripped = normalized
        for suffix in suffixes:
            if stripped.endswith(suffix) and len(stripped) > len(suffix) + 1:
                aliases.add(stripped[: -len(suffix)])
        for length in range(min(4, len(normalized)), 1, -1):
            prefix = normalized[:length]
            if prefix not in exact_names or prefix == normalized:
                aliases.add(prefix)
        for alias in aliases:
            if len(alias) >= 2 and alias not in {"台灣", "臺灣"}:
                candidates[alias].append((code, name, alias == normalized))
    candidates["台灣50"].append(("0050", "元大台灣50", True))
    candidates["臺灣50"].append(("0050", "元大台灣50", True))

    by_prefix: dict[str, tuple[str, str, bool]] = {}
    for prefix, records in candidates.items():
        # 同前綴優先正式全名，再優先較完整名稱；最後以代號穩定排序。
        by_prefix[prefix] = sorted(
            records,
            key=lambda item: (
                -int(item[2]),
                -len(normalize_name(item[1])),
                item[0],
            ),
        )[0]
    return StockAliasResolver(by_prefix)


def resolve_underlying_from_name(
    warrant_name: str, resolver: StockAliasResolver
) -> Optional[tuple[str, str, int, bool]]:
    return resolver.resolve(warrant_name)


def build_warrant_names_by_code(
    info: pd.DataFrame,
    required_codes: Optional[set[str]] = None,
    *,
    show_progress: bool = False,
) -> dict[str, list[str]]:
    names_by_code: dict[str, list[str]] = defaultdict(list)
    if not {"stock_id", "stock_name"}.issubset(info.columns):
        return names_by_code
    total = len(info)
    if show_progress:
        print(f"  ⏳ 建立權證名稱對照：來源 {total:,} 筆...", flush=True)
    for row_number, (raw_code, raw_name) in enumerate(
        info[["stock_id", "stock_name"]].itertuples(index=False, name=None),
        start=1,
    ):
        code = normalize_code(raw_code)
        if required_codes is not None and code not in required_codes:
            continue
        name = str(raw_name or "").strip()
        if code and name and name not in names_by_code[code]:
            names_by_code[code].append(name)
        if show_progress and row_number % 50_000 == 0:
            print(
                f"    權證名稱對照進度：{row_number:,}/{total:,}",
                flush=True,
            )
    if show_progress:
        print(
            f"  ✅ 權證名稱對照完成：{len(names_by_code):,} 個代號",
            flush=True,
        )
    return names_by_code


def roc_date_to_iso(value: Any) -> str:
    """將 MOPS 的民國日期（例如 112/11/30）轉成 ISO 日期。"""
    text = str(value or "").strip()
    matched = re.fullmatch(r"(\d{2,3})[./-](\d{1,2})[./-](\d{1,2})", text)
    if not matched:
        return iso_date(text)
    year, month, day = (int(part) for part in matched.groups())
    try:
        return date(year + 1911, month, day).isoformat()
    except ValueError:
        return ""


def iso_month_to_roc(value: str) -> str:
    parsed = parse_date(f"{value}-01" if re.fullmatch(r"\d{4}-\d{2}", value) else value)
    if parsed is None:
        raise ValueError(f"無效年月：{value}")
    roc_year = parsed.year - 1911
    if roc_year <= 0:
        raise ValueError(f"MOPS 不支援民國前日期：{value}")
    return f"{roc_year:03d}{parsed.month:02d}"


def month_range(start_day: str, end_day: str) -> list[str]:
    start = parse_date(start_day)
    end = parse_date(end_day)
    if start is None or end is None or start > end:
        raise ValueError(f"無效月份範圍：{start_day}～{end_day}")
    cursor = date(start.year, start.month, 1)
    final = date(end.year, end.month, 1)
    values: list[str] = []
    while cursor <= final:
        values.append(cursor.strftime("%Y-%m"))
        cursor = (
            date(cursor.year + 1, 1, 1)
            if cursor.month == 12
            else date(cursor.year, cursor.month + 1, 1)
        )
    return values


def normalize_mops_warrant_code(value: Any) -> str:
    """
    MOPS 的下市清單會在實際六碼交易代號後附加內部重用識別碼，
    HTML 又可能再附小寫 e（上市）或 c（上櫃）。

    例如 MOPS 的 04472UA、044720F，在 FinMind 分點日檔實際分別是
    04472U、044720；同一六碼的各段上市區間互不重疊。
    """
    text = re.sub(r"\s+", "", str(value or "").strip())
    if re.fullmatch(r"[0-9A-Za-z]+[ec]", text) and text[-1].islower():
        text = text[:-1]
    return normalize_code(text[:6])


class MopsWarrantTableParser(HTMLParser):
    """只解析 MOPS 回應中的 hasBorder 資料表，不依賴 lxml/BeautifulSoup。"""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.in_target_table = False
        self.table_depth = 0
        self.in_row = False
        self.in_cell = False
        self.cell_parts: list[str] = []
        self.row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(
        self, tag: str, attrs: list[tuple[str, Optional[str]]]
    ) -> None:
        lowered = tag.lower()
        attributes = {key.lower(): value or "" for key, value in attrs}
        if lowered == "table":
            if self.in_target_table:
                self.table_depth += 1
            elif "hasBorder" in attributes.get("class", "").split():
                self.in_target_table = True
                self.table_depth = 1
            return
        if not self.in_target_table:
            return
        if lowered == "tr":
            self.in_row = True
            self.row = []
        elif lowered in ("th", "td") and self.in_row:
            self.in_cell = True
            self.cell_parts = []
        elif lowered == "br" and self.in_cell:
            self.cell_parts.append(" ")

    def handle_data(self, data: str) -> None:
        if self.in_target_table and self.in_cell:
            self.cell_parts.append(data)

    def handle_endtag(self, tag: str) -> None:
        lowered = tag.lower()
        if not self.in_target_table:
            return
        if lowered in ("th", "td") and self.in_cell:
            value = re.sub(r"\s+", " ", "".join(self.cell_parts)).strip()
            self.row.append(value)
            self.in_cell = False
            self.cell_parts = []
        elif lowered == "tr" and self.in_row:
            if self.row:
                self.rows.append(self.row)
            self.in_row = False
            self.row = []
        elif lowered == "table":
            self.table_depth -= 1
            if self.table_depth <= 0:
                self.in_target_table = False
                self.table_depth = 0


MOPS_IDENTITY_COLUMNS = [
    "code",
    "name",
    "warrant_type",
    "target_code",
    "target_name",
    "list_date",
    "end_date",
    "status",
    "reason",
]
MOPS_CONTRACT_COLUMNS = [
    "market",
    "fulfillment_end_date",
    "settlement_method",
    "exercise_ratio",
    "fulfillment_price",
    "settlement_price",
    "settlement_record_found",
]
MOPS_SEED_COLUMNS = MOPS_IDENTITY_COLUMNS + MOPS_CONTRACT_COLUMNS


def resolve_mops_identity_seed_path(*, required: bool = True) -> Optional[Path]:
    candidates = [
        MOPS_IDENTITY_SEED_PATH,
        Path(__file__).resolve().with_name("mops_warrant_identity_seed.csv.gz"),
    ]
    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.exists() and resolved.stat().st_size > 0:
            return resolved
    if required:
        raise RuntimeError(
            "缺少隨程式提供的 MOPS 官方歷史身分種子檔："
            "mops_warrant_identity_seed.csv.gz。請將它與 "
            "warrant_broker_ma_history.py 一起放在 repository 根目錄；"
            "程式不會再依賴 GitHub runner 無法連線的 mopsov 主機。"
        )
    return None


def load_bundled_mops_identity_seed() -> pd.DataFrame:
    path = resolve_mops_identity_seed_path(required=True)
    try:
        frame = pd.read_csv(
            path,
            compression="gzip",
            dtype=str,
            keep_default_na=False,
        )
    except Exception as exc:
        raise RuntimeError(
            f"MOPS 官方歷史身分種子檔無法讀取：{path.name}｜"
            f"{type(exc).__name__}: {exc}"
        ) from exc
    missing = set(MOPS_SEED_COLUMNS) - set(frame.columns)
    if missing:
        raise RuntimeError(
            f"MOPS 官方歷史身分種子檔缺少欄位：{sorted(missing)}"
        )
    work = frame[MOPS_SEED_COLUMNS].copy()
    for column in MOPS_SEED_COLUMNS:
        work[column] = work[column].fillna("").astype(str).str.strip()
    work["code"] = work["code"].map(normalize_code)
    work["target_code"] = work["target_code"].map(normalize_code)
    listed_dates = pd.to_datetime(
        work["list_date"], format="%Y-%m-%d", errors="coerce"
    )
    ended_dates = pd.to_datetime(
        work["end_date"], format="%Y-%m-%d", errors="coerce"
    )
    work["list_date"] = listed_dates.dt.strftime("%Y-%m-%d").fillna("")
    work["end_date"] = ended_dates.dt.strftime("%Y-%m-%d").fillna("")
    fulfillment_dates = pd.to_datetime(
        work["fulfillment_end_date"], format="%Y-%m-%d", errors="coerce"
    )
    work["fulfillment_end_date"] = (
        fulfillment_dates.dt.strftime("%Y-%m-%d").fillna("")
    )
    for column in ("exercise_ratio", "fulfillment_price", "settlement_price"):
        work[column] = pd.to_numeric(work[column], errors="coerce")
    settlement_found = work["settlement_record_found"].str.lower().isin(
        {"1", "true", "yes"}
    )
    work["settlement_record_found"] = settlement_found
    invalid = (
        work["code"].eq("")
        | work["code"].str.len().ne(6)
        | work["target_code"].eq("")
        | work["list_date"].eq("")
        | work["end_date"].eq("")
        | (work["list_date"] > work["end_date"])
        | work["status"].ne("Matched")
        | ~work["market"].isin(["1", "2"])
        | work["fulfillment_end_date"].eq("")
    )
    invalid_settlement = settlement_found & (
        work["exercise_ratio"].isna()
        | (work["exercise_ratio"] <= 0)
        | work["fulfillment_price"].isna()
        | (work["fulfillment_price"] < 0)
        | work["settlement_price"].isna()
        | (work["settlement_price"] < 0)
    )
    invalid |= invalid_settlement
    if invalid.any():
        raise RuntimeError(
            "MOPS 官方歷史身分種子檔含無效代號、日期或狀態："
            f"{work.loc[invalid].head(5).to_dict('records')}"
        )
    work = work.drop_duplicates(
        ["code", "list_date", "end_date", "target_code"],
        keep="last",
    ).reset_index(drop=True)
    if len(work) < 200_000:
        raise RuntimeError(
            f"MOPS 官方歷史身分種子檔筆數異常：{len(work):,} < 200,000"
        )
    expected_samples = {
        ("030650", "2023-06-21"): "2330",
        ("030961", "2023-06-21"): "2330",
        ("034676", "2023-06-21"): "8478",
        ("034755", "2023-06-21"): "8478",
        ("035628", "2023-06-21"): "IX0001",
        ("04472U", "2023-06-21"): "4763",
        ("04900U", "2023-06-21"): "2603",
        ("05159U", "2023-06-21"): "2603",
        ("05575U", "2023-06-21"): "2454",
        ("06037U", "2023-06-21"): "3443",
    }
    bad_samples: dict[str, str] = {}
    for (code, trade_day), expected_target in expected_samples.items():
        matches = work[
            (work["code"] == code)
            & (work["list_date"] <= trade_day)
            & (work["end_date"] >= trade_day)
        ]
        targets = sorted(set(matches["target_code"]))
        if targets != [expected_target]:
            bad_samples[f"{code}@{trade_day}"] = ",".join(targets) or "Unmatched"
    if bad_samples:
        raise RuntimeError(
            f"MOPS 官方歷史身分種子檔樣本驗證失敗：{bad_samples}"
        )
    print(
        f"  📦 載入隨附 MOPS 官方歷史身分種子："
        f"{len(work):,} 筆｜官方到期結算 "
        f"{int(work['settlement_record_found'].sum()):,} 筆｜{path.name}",
        flush=True,
    )
    return work


def parse_mops_warrant_html(html: str, market: str) -> pd.DataFrame:
    parser = MopsWarrantTableParser()
    parser.feed(html or "")
    header: Optional[list[str]] = None
    data_rows: list[list[str]] = []
    for row in parser.rows:
        normalized = [normalize_name(value) for value in row]
        if normalized and normalized[0] == "權證代號":
            header = normalized
            continue
        if header is not None and len(row) >= len(header):
            data_rows.append(row[: len(header)])
    if header is None:
        return pd.DataFrame(columns=MOPS_IDENTITY_COLUMNS)

    def column_index(*candidates: str) -> int:
        for candidate in candidates:
            if candidate in header:
                return header.index(candidate)
        raise RuntimeError(
            f"MOPS 權證表缺少欄位 {candidates}；實際欄位={header}"
        )

    code_index = column_index("權證代號")
    name_index = column_index("權證簡稱")
    type_index = column_index("權證類型")
    list_index = column_index("上市日期", "上櫃日期")
    # FinMind Summary 的 end_date 語意是最後交易日，不是履約截止日；
    # MOPS 查詢雖按履約截止月份切片，身分有效區間必須使用最後交易日。
    end_index = column_index("最後交易日", "履約截止日")
    target_code_index = column_index("標的代號")
    target_name_index = column_index("標的名稱")
    records: list[dict[str, str]] = []
    source_name = "MOPS上市官方歷史" if str(market) == "1" else "MOPS上櫃官方歷史"
    for row in data_rows:
        code = normalize_mops_warrant_code(row[code_index])
        target_code = normalize_code(row[target_code_index])
        list_date = roc_date_to_iso(row[list_index])
        end_date = roc_date_to_iso(row[end_index])
        if not code or not target_code or not list_date or not end_date:
            continue
        records.append(
            {
                "code": code,
                "name": str(row[name_index] or "").strip(),
                "warrant_type": str(row[type_index] or "").strip(),
                "target_code": target_code,
                "target_name": str(row[target_name_index] or "").strip(),
                "list_date": list_date,
                "end_date": end_date,
                "status": "Matched",
                "reason": f"{source_name}（含下市權證正式區間）",
            }
        )
    return pd.DataFrame(records, columns=MOPS_IDENTITY_COLUMNS).drop_duplicates(
        ["code", "list_date", "end_date", "target_code"],
        keep="last",
    )


def mops_identity_cache_path(market: str, period: str) -> Path:
    safe_period = re.sub(r"[^0-9A-Za-z_-]", "_", period)
    return METADATA_DIR / (
        f"{MOPS_WARRANT_HISTORY_CACHE_PREFIX}_{market}_{safe_period}.parquet"
    )


def mops_headers() -> dict[str, str]:
    return {
        "User-Agent": f"warrant-broker-ma-history/{PROGRAM_VERSION}",
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.5",
        "Referer": MOPS_WARRANT_LANDING_URL,
    }


def prepare_mops_session() -> Any:
    session = get_session()
    if getattr(_thread_local, "mops_session_ready", False):
        return session
    response = session.get(
        MOPS_WARRANT_LANDING_URL,
        headers=mops_headers(),
        timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
    )
    response.raise_for_status()
    _thread_local.mops_session_ready = True
    return session


def request_mops_warrant_page(
    market: str,
    *,
    roc_month: str = "",
    page: int = 1,
    active: bool = False,
) -> tuple[str, int]:
    payload: dict[str, str] = {
        "step": "1",
        "firstin": "1",
        "off": "1",
        "r": str(market),
        "pageno": str(page),
        "pagesize": str(MOPS_PAGE_SIZE),
    }
    if not active:
        payload.update(
            {
                "rc": "0",
                "start_date": roc_month,
                "end_date": roc_month,
            }
        )
    last_error: Optional[BaseException] = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = prepare_mops_session().post(
                MOPS_WARRANT_URL,
                data=payload,
                headers=mops_headers(),
                timeout=(REQUEST_CONNECT_TIMEOUT, REQUEST_READ_TIMEOUT),
            )
            response.raise_for_status()
            response.encoding = response.apparent_encoding or "utf-8"
            html = response.text
            if "FOR SECURITY REASONS" in html or "錯誤代碼" in html:
                raise RuntimeError("MOPS 拒絕此查詢")
            page_match = re.search(
                r"&nbsp;\s*&nbsp;\s*\d+\s*/\s*(\d+)", html
            )
            total_pages = int(page_match.group(1)) if page_match else 1
            has_table = "class='hasBorder'" in html or 'class="hasBorder"' in html
            explicit_empty = any(
                phrase in html
                for phrase in ("查無資料", "無符合", "沒有符合", "無資料")
            )
            if not has_table and not explicit_empty:
                raise RuntimeError("MOPS 回應未包含資料表")
            return html, max(total_pages, 1)
        except Exception as exc:
            last_error = exc
            _thread_local.mops_session_ready = False
            if attempt < MAX_RETRIES:
                delay = min(RETRY_BASE_SECONDS * (2 ** (attempt - 1)), 30)
                print(
                    f"    ⚠️ MOPS 第 {attempt}/{MAX_RETRIES} 次失敗："
                    f"{type(exc).__name__}: {exc}；{delay:.1f} 秒後重試",
                    flush=True,
                )
                time.sleep(delay)
    raise RuntimeError(
        f"MOPS 權證歷史重試 {MAX_RETRIES} 次仍失敗："
        f"{type(last_error).__name__}: {last_error}"
    )


def fetch_mops_warrant_identity(
    market: str,
    *,
    month: str = "",
    active: bool = False,
) -> pd.DataFrame:
    period = "active" if active else iso_month_to_roc(month)
    path = mops_identity_cache_path(market, period)
    if path.exists() and path.stat().st_size > 0:
        age_hours = max((time.time() - path.stat().st_mtime) / 3600, 0)
        if not active or age_hours <= METADATA_CACHE_HOURS:
            try:
                cached = read_parquet(path)
                missing = set(MOPS_IDENTITY_COLUMNS) - set(cached.columns)
                if not missing and (not active or not cached.empty):
                    return cached[MOPS_IDENTITY_COLUMNS].copy()
            except Exception as exc:
                print(
                    f"    ⚠️ MOPS 快取 {path.name} 損壞，重新下載："
                    f"{type(exc).__name__}: {exc}",
                    flush=True,
                )

    label = "目前仍掛牌" if active else f"到期月 {month}"
    market_name = "上市" if str(market) == "1" else "上櫃"
    print(f"    ⏳ MOPS {market_name}{label}權證...", flush=True)
    first_html, total_pages = request_mops_warrant_page(
        market,
        roc_month="" if active else period,
        page=1,
        active=active,
    )
    frames = [parse_mops_warrant_html(first_html, market)]
    for page in range(2, total_pages + 1):
        if MOPS_REQUEST_DELAY_SECONDS:
            time.sleep(MOPS_REQUEST_DELAY_SECONDS)
        html, _ = request_mops_warrant_page(
            market,
            roc_month="" if active else period,
            page=page,
            active=active,
        )
        frames.append(parse_mops_warrant_html(html, market))
        print(
            f"      MOPS {market_name}{label}分頁：{page}/{total_pages}",
            flush=True,
        )
    frame = pd.concat(frames, ignore_index=True).drop_duplicates(
        ["code", "list_date", "end_date", "target_code"],
        keep="last",
    )
    if frame.empty and active:
        raise RuntimeError(f"MOPS {market_name}{label}回傳空白身分資料")
    atomic_write_parquet(frame[MOPS_IDENTITY_COLUMNS], path)
    print(
        f"    {'↪️' if frame.empty else '✅'} "
        f"MOPS {market_name}{label}：{len(frame):,} 筆",
        flush=True,
    )
    return frame[MOPS_IDENTITY_COLUMNS].copy()


def mops_frames_to_identity_records(
    frames: Iterable[pd.DataFrame],
) -> list["WarrantIdentity"]:
    records: list[WarrantIdentity] = []
    for frame in frames:
        if frame.empty:
            continue
        missing = set(MOPS_IDENTITY_COLUMNS) - set(frame.columns)
        if missing:
            raise RuntimeError(f"MOPS 身分快取缺少欄位：{sorted(missing)}")
        work = frame[MOPS_IDENTITY_COLUMNS].copy()
        work["code"] = work["code"].map(normalize_mops_warrant_code)
        for values in work.itertuples(
            index=False, name=None
        ):
            records.append(
                WarrantIdentity(
                    *[
                        "" if value is None or pd.isna(value) else str(value)
                        for value in values
                    ]
                )
            )
    return records


def load_cached_mops_identity_frames() -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    for path in sorted(
        METADATA_DIR.glob(f"{MOPS_WARRANT_HISTORY_CACHE_PREFIX}_*.parquet")
    ):
        try:
            frame = read_parquet(path)
            if not frame.empty:
                frames.append(frame[MOPS_IDENTITY_COLUMNS].copy())
        except Exception as exc:
            print(
                f"  ⚠️ 略過損壞的 MOPS 身分快取 {path.name}："
                f"{type(exc).__name__}: {exc}",
                flush=True,
            )
    return frames


@dataclass(frozen=True)
class WarrantIdentity:
    code: str
    name: str
    warrant_type: str
    target_code: str
    target_name: str
    list_date: str
    end_date: str
    status: str
    reason: str


class WarrantIdentityIndex:
    def __init__(self, records: Iterable[WarrantIdentity]):
        self.by_code: dict[str, list[WarrantIdentity]] = defaultdict(list)
        for record in records:
            self.by_code[record.code].append(record)
        for code in self.by_code:
            self.by_code[code].sort(key=lambda rec: (rec.list_date, rec.end_date))

    def records(self) -> list[WarrantIdentity]:
        return [
            record
            for code in sorted(self.by_code)
            for record in self.by_code[code]
        ]

    def resolve(self, warrant_code: str, trade_date: str) -> WarrantIdentity:
        code = normalize_code(warrant_code)
        matches = [
            rec
            for rec in self.by_code.get(code, [])
            if rec.list_date <= trade_date <= rec.end_date
        ]
        if not matches:
            return WarrantIdentity(
                code,
                "",
                "",
                "",
                "",
                "",
                "",
                "Unmatched",
                "交易日找不到有效權證上市區間",
            )
        target_codes = {rec.target_code for rec in matches if rec.target_code}
        if len(target_codes) > 1:
            return WarrantIdentity(
                code,
                matches[-1].name,
                matches[-1].warrant_type,
                "",
                "",
                matches[-1].list_date,
                matches[-1].end_date,
                "Ambiguous",
                f"同一交易日命中多個標的代號：{sorted(target_codes)}",
            )
        chosen = max(matches, key=lambda rec: rec.list_date)
        return chosen


@dataclass(frozen=True)
class WarrantContract:
    code: str
    warrant_type: str
    target_code: str
    list_date: str
    end_date: str
    fulfillment_end_date: str
    settlement_method: str
    exercise_ratio: Optional[float]
    fulfillment_price: Optional[float]
    settlement_price: Optional[float]
    settlement_record_found: bool
    market: str


class WarrantContractIndex:
    def __init__(self, records: Iterable[WarrantContract]):
        self.by_code: dict[str, list[WarrantContract]] = defaultdict(list)
        for record in records:
            self.by_code[record.code].append(record)
        for code in self.by_code:
            self.by_code[code].sort(key=lambda rec: (rec.list_date, rec.end_date))

    def resolve(self, warrant_code: str, buy_day: str) -> Optional[WarrantContract]:
        code = normalize_code(warrant_code)
        matches = [
            record
            for record in self.by_code.get(code, [])
            if record.list_date <= buy_day <= record.end_date
        ]
        return max(matches, key=lambda rec: rec.list_date) if matches else None


def optional_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        numeric = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return numeric if math.isfinite(numeric) else None


def build_warrant_contract_index(
    seed: Optional[pd.DataFrame] = None,
) -> WarrantContractIndex:
    work = seed if seed is not None else load_bundled_mops_identity_seed()
    records = [
        WarrantContract(
            code=normalize_code(row.code),
            warrant_type=str(row.warrant_type or "").strip(),
            target_code=normalize_code(row.target_code),
            list_date=iso_date(row.list_date),
            end_date=iso_date(row.end_date),
            fulfillment_end_date=iso_date(row.fulfillment_end_date),
            settlement_method=str(row.settlement_method or "").strip(),
            exercise_ratio=optional_float(row.exercise_ratio),
            fulfillment_price=optional_float(row.fulfillment_price),
            settlement_price=optional_float(row.settlement_price),
            settlement_record_found=(
                row.settlement_record_found is True
                or str(row.settlement_record_found).strip().lower()
                in {"1", "true", "yes"}
            ),
            market=str(row.market or "").strip(),
        )
        for row in work[MOPS_SEED_COLUMNS].itertuples(index=False)
    ]
    return WarrantContractIndex(records)


def identity_index_cache_path() -> Path:
    return METADATA_DIR / "warrant_identity_intervals.parquet"


def build_identity_records_from_summary(
    summary: pd.DataFrame,
    names_by_code: dict[str, list[str]],
    resolver: StockAliasResolver,
    stock_master: dict[str, str],
    *,
    progress_label: str = "",
) -> list[WarrantIdentity]:
    records: list[WarrantIdentity] = []
    if summary.empty:
        return records
    required = {"stock_id", "target_stock_id", "type", "date", "end_date"}
    missing = required - set(summary.columns)
    if missing:
        raise RuntimeError(f"權證標的歷史對照缺少欄位：{sorted(missing)}")
    work = summary.copy().fillna("")
    total_intervals = len(work)
    if progress_label:
        print(
            f"  ⏳ {progress_label}：{total_intervals:,} 筆...",
            flush=True,
        )
    for row_number, (
        raw_code,
        raw_target,
        raw_type,
        raw_listed,
        raw_ended,
    ) in enumerate(
        work[
            ["stock_id", "target_stock_id", "type", "date", "end_date"]
        ].itertuples(index=False, name=None),
        start=1,
    ):
        code = normalize_code(raw_code)
        target = normalize_code(raw_target)
        warrant_type = str(raw_type or "").strip()
        listed = iso_date(raw_listed)
        ended = iso_date(raw_ended)
        if not code or not listed or not ended:
            continue
        names = names_by_code.get(code, [])
        name = names[0] if names else code
        resolved_candidates = [
            (candidate, resolve_underlying_from_name(candidate, resolver))
            for candidate in names
        ]
        resolved_candidates = [item for item in resolved_candidates if item[1]]
        matching_names = [
            candidate
            for candidate, resolved in resolved_candidates
            if resolved and resolved[0] == target
        ]
        if matching_names:
            name = max(matching_names, key=len)
        elif target:
            # TaiwanStockInfoWithWarrant 沒有日期區間；代號重用時其中的名稱
            # 可能屬於較新的商品。若名稱無法印證這筆日期化母股，寧可顯示
            # 代號，也不能拿目前名稱反向覆蓋正式 target_stock_id。
            name = code
        resolved = resolve_underlying_from_name(name, resolver)
        target_name = stock_master.get(target, "")
        status = "Matched"
        reason = "FinMind target_stock_id＋上市區間"
        if resolved:
            resolved_code, resolved_name, prefix_len, exact = resolved
            high_confidence = exact or prefix_len >= 3
            if not target and high_confidence:
                target, target_name = resolved_code, resolved_name
                reason = "權證名稱解析補足標的"
            elif target and resolved_code != target and high_confidence:
                reason = (
                    "FinMind target_stock_id＋上市區間"
                    "（非日期化權證名稱不一致，保留正式母股）"
                )
        if not target:
            status = "Unmatched"
            reason = "缺少 target_stock_id 且權證名稱無法可靠解析"
        records.append(
            WarrantIdentity(
                code,
                name,
                warrant_type,
                target,
                target_name,
                listed,
                ended,
                status,
                reason,
            )
        )
        if progress_label and row_number % 5_000 == 0:
            print(
                f"    身分配對進度：{row_number:,}/{total_intervals:,}｜"
                f"已建立 {len(records):,} 筆",
                flush=True,
            )
    return records


def deduplicate_identity_records(
    records: Iterable[WarrantIdentity],
) -> dict[tuple[str, str, str, str], WarrantIdentity]:
    deduplicated: dict[tuple[str, str, str, str], WarrantIdentity] = {}
    for record in records:
        # target_code 必須留在鍵內；同代號、同區間卻有不同母股時，
        # WarrantIdentityIndex.resolve 會回 Ambiguous，禁止任意挑一筆。
        key = (
            record.code,
            record.list_date,
            record.end_date,
            record.target_code,
        )
        previous = deduplicated.get(key)
        if previous is None or (
            record.status == "Matched",
            bool(record.target_code),
            len(record.name),
        ) >= (
            previous.status == "Matched",
            bool(previous.target_code),
            len(previous.name),
        ):
            deduplicated[key] = record
    return deduplicated


def save_warrant_identity_index(
    records: Iterable[WarrantIdentity],
) -> WarrantIdentityIndex:
    deduplicated = deduplicate_identity_records(records)
    frame = pd.DataFrame([record.__dict__ for record in deduplicated.values()])
    if frame.empty:
        raise RuntimeError("不得保存空白權證身分索引")
    frame["_identity_schema_version"] = IDENTITY_INDEX_SCHEMA_VERSION
    frame["_summary_history_start"] = WARRANT_SUMMARY_HISTORY_START
    frame["_summary_history_end"] = datetime.now().date().isoformat()
    frame["_mops_seed_version"] = MOPS_IDENTITY_SEED_VERSION
    frame["_built_at"] = datetime.now().isoformat(timespec="seconds")
    atomic_write_parquet(frame, identity_index_cache_path())
    return WarrantIdentityIndex(deduplicated.values())


def load_cached_target_summary_frames() -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    for path in sorted(
        METADATA_DIR.glob(
            f"{WARRANT_SUMMARY_TARGET_CACHE_PREFIX}_*.parquet"
        )
    ):
        try:
            frame = read_parquet(path)
            if not frame.empty:
                frames.append(frame)
        except Exception as exc:
            print(
                f"  ⚠️ 略過損壞的母股歷史分片 {path.name}："
                f"{type(exc).__name__}: {exc}",
                flush=True,
            )
    return frames


def load_cached_warrant_identity_index() -> Optional[WarrantIdentityIndex]:
    path = identity_index_cache_path()
    if not path.exists() or path.stat().st_size <= 0:
        return None
    age_hours = max((time.time() - path.stat().st_mtime) / 3600, 0)
    if age_hours > IDENTITY_INDEX_CACHE_HOURS:
        print(
            f"  ℹ️ 權證身分索引快取已逾 {IDENTITY_INDEX_CACHE_HOURS:g} 小時，重新建立。",
            flush=True,
        )
        return None

    summary_history_paths = sorted(
        METADATA_DIR.glob(f"{WARRANT_SUMMARY_CACHE_PREFIX}_*.parquet")
    )
    if not summary_history_paths:
        print(
            "  ℹ️ 尚無明確日期範圍的權證 Summary 歷史快取，重新建立索引。",
            flush=True,
        )
        return None
    source_paths = [
        metadata_cache_path("TaiwanStockInfoWithWarrant"),
        metadata_cache_path("TaiwanStockInfo"),
        *summary_history_paths,
        *sorted(
            METADATA_DIR.glob(
                f"{WARRANT_SUMMARY_TARGET_CACHE_PREFIX}_*.parquet"
            )
        ),
    ]
    if ENABLE_MOPS_NETWORK_REFRESH:
        source_paths.extend(
            sorted(
                METADATA_DIR.glob(
                    f"{MOPS_WARRANT_HISTORY_CACHE_PREFIX}_*.parquet"
                )
            )
        )
    newer_sources = [
        source.name
        for source in source_paths
        if source.exists() and source.stat().st_mtime > path.stat().st_mtime
    ]
    if newer_sources:
        print(
            f"  ℹ️ 權證身分來源快取較新，重新建立索引：{newer_sources}",
            flush=True,
        )
        return None

    required_columns = {
        "code",
        "name",
        "warrant_type",
        "target_code",
        "target_name",
        "list_date",
        "end_date",
        "status",
        "reason",
        "_identity_schema_version",
        "_summary_history_start",
        "_mops_seed_version",
    }
    try:
        frame = read_parquet(path)
        missing = required_columns - set(frame.columns)
        versions = set(
            frame["_identity_schema_version"].dropna().astype(str).unique()
        ) if "_identity_schema_version" in frame.columns else set()
        if frame.empty or missing or versions != {IDENTITY_INDEX_SCHEMA_VERSION}:
            print(
                f"  ℹ️ 權證身分索引快取 schema 不相容，重新建立："
                f"missing={sorted(missing)} versions={sorted(versions)}",
                flush=True,
            )
            return None
        history_starts = set(
            frame["_summary_history_start"].dropna().astype(str).unique()
        )
        if history_starts != {WARRANT_SUMMARY_HISTORY_START}:
            print(
                "  ℹ️ 權證身分索引歷史起日不符，重新建立："
                f"cache={sorted(history_starts)} "
                f"expected={WARRANT_SUMMARY_HISTORY_START}",
                flush=True,
            )
            return None
        seed_versions = set(
            frame["_mops_seed_version"].dropna().astype(str).unique()
        )
        if seed_versions != {MOPS_IDENTITY_SEED_VERSION}:
            print(
                "  ℹ️ 權證身分索引的 MOPS 種子版本不符，重新建立："
                f"cache={sorted(seed_versions)} "
                f"expected={MOPS_IDENTITY_SEED_VERSION}",
                flush=True,
            )
            return None
        listed_dates = pd.to_datetime(
            frame["list_date"], format="%Y-%m-%d", errors="coerce"
        )
        ended_dates = pd.to_datetime(
            frame["end_date"], format="%Y-%m-%d", errors="coerce"
        )
        invalid_identity_rows = (
            frame["code"].fillna("").astype(str).str.strip().eq("")
            | listed_dates.isna()
            | ended_dates.isna()
            | (listed_dates > ended_dates)
        )
        if invalid_identity_rows.any():
            raise ValueError("索引包含無效代號或日期")
        records: list[WarrantIdentity] = []
        for values in frame[
            [
                "code",
                "name",
                "warrant_type",
                "target_code",
                "target_name",
                "list_date",
                "end_date",
                "status",
                "reason",
            ]
        ].itertuples(index=False, name=None):
            record = WarrantIdentity(
                *[
                    "" if value is None or pd.isna(value) else str(value)
                    for value in values
                ]
            )
            records.append(record)
        print(
            f"  📦 沿用權證日期化身分索引：{len(records):,} 筆｜"
            f"快取年齡 {age_hours:.1f} 小時",
            flush=True,
        )
        return WarrantIdentityIndex(records)
    except Exception as exc:
        print(
            f"  ⚠️ 權證身分索引快取讀取失敗，重新建立："
            f"{type(exc).__name__}: {exc}",
            flush=True,
        )
        return None


def build_warrant_identity_index() -> WarrantIdentityIndex:
    cached_index = load_cached_warrant_identity_index()
    if cached_index is not None:
        return cached_index

    started_at = time.perf_counter()
    print("【初始化】建立日期化權證身分索引...", flush=True)
    info = fetch_dataset("TaiwanStockInfoWithWarrant", cache=True)
    summary = fetch_warrant_summary_history()
    stock_info = fetch_dataset("TaiwanStockInfo", cache=True)
    required = {"stock_id", "target_stock_id", "type", "date", "end_date"}
    missing = required - set(summary.columns)
    if missing:
        raise RuntimeError(f"權證標的歷史對照缺少欄位：{sorted(missing)}")

    # 標的 resolver 只能由真正股票／ETF 主檔建立；TaiwanStockInfoWithWarrant
    # 含大量權證本身，混入後會令別名數暴增並造成近似平方級掃描。
    stock_master = build_stock_master(stock_info)
    resolver = stock_alias_resolver(stock_master)
    print(
        f"  ✅ 標的主檔 {len(stock_master):,} 檔｜"
        f"前綴索引 {len(resolver.by_prefix):,} 個",
        flush=True,
    )
    names_by_code = build_warrant_names_by_code(
        info,
        show_progress=True,
    )
    shard_frames = load_cached_target_summary_frames()
    if shard_frames:
        summary = pd.concat([summary, *shard_frames], ignore_index=True)
        summary = summary.drop_duplicates(
            ["stock_id", "target_stock_id", "type", "date", "end_date"],
            keep="last",
        )
        print(
            f"  📦 合併既有母股歷史分片：{len(shard_frames):,} 片｜"
            f"合計 {len(summary):,} 筆",
            flush=True,
        )
    records = build_identity_records_from_summary(
        summary,
        names_by_code,
        resolver,
        stock_master,
        progress_label="配對全部權證歷史區間",
    )
    seed_frame = load_bundled_mops_identity_seed()
    seed_records = mops_frames_to_identity_records([seed_frame])
    records.extend(seed_records)
    if ENABLE_MOPS_NETWORK_REFRESH:
        mops_frames = load_cached_mops_identity_frames()
        if mops_frames:
            mops_records = mops_frames_to_identity_records(mops_frames)
            records.extend(mops_records)
            print(
                f"  📦 合併 MOPS 官方歷史身分快取："
                f"{len(mops_frames):,} 片｜{len(mops_records):,} 筆",
                flush=True,
            )
    identity_index = save_warrant_identity_index(records)
    print(
        f"  ✅ 日期化權證身分索引完成："
        f"{len(identity_index.records()):,} 筆｜"
        f"{time.perf_counter() - started_at:.1f} 秒",
        flush=True,
    )
    return identity_index


def identity_pair_is_matched(
    identity_index: WarrantIdentityIndex,
    pair: tuple[str, str],
) -> bool:
    code, trade_day = pair
    return identity_index.resolve(code, trade_day).status == "Matched"


def backfill_identity_from_target_shards(
    identity_index: WarrantIdentityIndex,
    pairs: Iterable[tuple[str, str]],
) -> WarrantIdentityIndex:
    """
    以權證名稱只找出「應查詢哪些母股」，再向 FinMind 逐母股取回正式區間。

    名稱解析結果不會直接成為身分；只有母股分片實際回傳相同權證代號，
    且交易日在該列 date～end_date 內，resolve 才會變成 Matched。
    """
    normalized_pairs = {
        (normalize_code(code), iso_date(trade_day) or "")
        for code, trade_day in pairs
        if normalize_code(code) and iso_date(trade_day)
    }
    unresolved_pairs = {
        pair
        for pair in normalized_pairs
        if not identity_pair_is_matched(identity_index, pair)
    }
    if not unresolved_pairs:
        return identity_index

    unresolved_codes = {code for code, _ in unresolved_pairs}
    print(
        "  🧩 啟動 FinMind 母股分片歷史補抓："
        f"{len(unresolved_codes):,} 個權證代號｜"
        f"{len(unresolved_pairs):,} 個代號日期組合",
        flush=True,
    )
    info = fetch_dataset("TaiwanStockInfoWithWarrant", cache=True)
    stock_info = fetch_dataset("TaiwanStockInfo", cache=True)
    stock_master = build_stock_master(stock_info)
    resolver = stock_alias_resolver(stock_master)
    required_names = build_warrant_names_by_code(
        info,
        required_codes=unresolved_codes,
    )
    all_names = build_warrant_names_by_code(info)

    candidate_targets: set[str] = set()
    codes_without_candidate: set[str] = set()
    for code in sorted(unresolved_codes):
        code_targets: set[str] = {
            record.target_code
            for record in identity_index.by_code.get(code, [])
            if record.target_code
        }
        for name in required_names.get(code, []):
            resolved = resolve_underlying_from_name(name, resolver)
            if resolved is None:
                continue
            target, _, prefix_len, exact = resolved
            if exact or prefix_len >= 3:
                code_targets.add(target)
        if not code_targets:
            codes_without_candidate.add(code)
        candidate_targets.update(code_targets)

    print(
        f"  🧭 待查母股 {len(candidate_targets):,} 檔｜"
        f"無可靠名稱候選 {len(codes_without_candidate):,} 個權證代號",
        flush=True,
    )
    matched_before = sum(
        identity_pair_is_matched(identity_index, pair)
        for pair in normalized_pairs
    )

    def fetch_target_round(
        targets: set[str],
        round_label: str,
    ) -> list[pd.DataFrame]:
        frames: list[pd.DataFrame] = []
        consecutive_errors = 0
        failed_targets: list[str] = []
        ordered_targets = sorted(
            {normalize_code(target) for target in targets if normalize_code(target)}
        )
        if not ordered_targets:
            return frames
        print(
            f"  📚 {round_label}：共 {len(ordered_targets):,} 檔母股",
            flush=True,
        )
        for index, target in enumerate(ordered_targets, start=1):
            while True:
                try:
                    shard = fetch_warrant_summary_target_history(target)
                    consecutive_errors = 0
                    if not shard.empty:
                        frames.append(shard)
                    break
                except FinMindQuotaExceeded:
                    wait_seconds = max(
                        int(
                            os.getenv(
                                "FINMIND_QUOTA_WAIT_SECONDS",
                                "3660",
                            )
                        ),
                        60,
                    )
                    print(
                        "    ⏸️ FinMind API 額度已用完；"
                        f"已完成分片均已保存，等待 {wait_seconds:,} 秒後"
                        f"從母股 {target} 接續。若先達 GitHub 截止時間，"
                        "下一個 workflow 會從快取續跑。",
                        flush=True,
                    )
                    time.sleep(wait_seconds)
                except Exception as exc:
                    consecutive_errors += 1
                    failed_targets.append(target)
                    print(
                        f"    ⚠️ 母股 {target} 歷史分片失敗："
                        f"{type(exc).__name__}: {exc}",
                        flush=True,
                    )
                    if consecutive_errors >= LOCAL_ERROR_CIRCUIT_BREAKER:
                        raise RuntimeError(
                            "母股歷史分片連續失敗 "
                            f"{consecutive_errors} 次，停止避免消耗 API 額度"
                        ) from exc
                    break
            if index % 25 == 0 or index == len(ordered_targets):
                print(
                    f"  📚 {round_label}進度：{index:,}/"
                    f"{len(ordered_targets):,}｜"
                    f"有資料 {len(frames):,}｜"
                    f"失敗 {len(failed_targets):,}",
                    flush=True,
                )
        return frames

    def merge_shard_frames(
        base_index: WarrantIdentityIndex,
        frames: list[pd.DataFrame],
        progress_label: str,
    ) -> tuple[WarrantIdentityIndex, int]:
        if not frames:
            return base_index, 0
        shard_summary = pd.concat(frames, ignore_index=True).drop_duplicates(
            ["stock_id", "target_stock_id", "type", "date", "end_date"],
            keep="last",
        )
        new_records = build_identity_records_from_summary(
            shard_summary,
            all_names,
            resolver,
            stock_master,
            progress_label=progress_label,
        )
        return (
            save_warrant_identity_index(
                [*base_index.records(), *new_records]
            ),
            len(new_records),
        )

    candidate_frames = fetch_target_round(
        candidate_targets,
        "名稱候選母股補抓",
    )
    repaired_index, candidate_record_count = merge_shard_frames(
        identity_index,
        candidate_frames,
        "配對名稱候選母股區間",
    )
    matched_after_candidates = sum(
        identity_pair_is_matched(repaired_index, pair)
        for pair in normalized_pairs
    )
    remaining_after_candidates = len(normalized_pairs) - matched_after_candidates
    remaining_pct = (
        remaining_after_candidates / len(normalized_pairs) * 100
        if normalized_pairs
        else 0.0
    )
    exhaustive_record_count = 0
    if remaining_pct > MAX_IDENTITY_UNMATCHED_PCT:
        # TaiwanStockInfoWithWarrant 的名稱不是日期化資料。代號重用後，舊名稱
        # 可能完全消失，因此不能只依目前名稱找母股。第二輪遍歷真正股票主檔
        # 以及全表 Summary 已知的所有母股；FinMind 每個 data_id 回傳的正式
        # date～end_date 才是最後身分依據。
        known_summary_targets = {
            record.target_code
            for record in repaired_index.records()
            if record.target_code
        }
        exhaustive_targets = (
            set(stock_master)
            | known_summary_targets
        ) - candidate_targets
        print(
            "  🔄 名稱候選不足："
            f"仍有 {remaining_after_candidates:,}/"
            f"{len(normalized_pairs):,} 組未配對（{remaining_pct:.2f}%）；"
            f"改查其餘 {len(exhaustive_targets):,} 檔已知母股。",
            flush=True,
        )
        exhaustive_frames = fetch_target_round(
            exhaustive_targets,
            "全母股歷史補抓",
        )
        repaired_index, exhaustive_record_count = merge_shard_frames(
            repaired_index,
            exhaustive_frames,
            "配對全母股補回區間",
        )

    matched_after = sum(
        identity_pair_is_matched(repaired_index, pair)
        for pair in normalized_pairs
    )
    print(
        "  ✅ 母股分片補抓完成："
        f"候選區間 {candidate_record_count:,} 筆｜"
        f"全母股區間 {exhaustive_record_count:,} 筆｜"
        f"本批已配對 {matched_before:,} → {matched_after:,}/"
        f"{len(normalized_pairs):,}",
        flush=True,
    )
    return repaired_index


def backfill_identity_from_official_mops(
    identity_index: WarrantIdentityIndex,
    pairs: Iterable[tuple[str, str]],
) -> WarrantIdentityIndex:
    """
    以 MOPS「認購(售)權證基本資料彙總表（含下市權證）」補足正式區間。

    FinMind Summary 實際只回傳目前可見的代號區間；代號重用時，2023 年交易
    可能只查到 2025 年的新權證。MOPS 的下市彙總表則明確包含權證代號、
    上市/上櫃日、履約截止日與標的代號，因此可作日期化身分的官方備援。
    每個到期月份、每個市場都立即寫入獨立 Parquet，runner 中止後可續抓。
    """
    normalized_pairs = {
        (normalize_code(code), iso_date(trade_day) or "")
        for code, trade_day in pairs
        if normalize_code(code) and iso_date(trade_day)
    }
    unresolved_pairs = {
        pair
        for pair in normalized_pairs
        if not identity_pair_is_matched(identity_index, pair)
    }
    if not unresolved_pairs:
        return identity_index

    # 先使用 repository 內隨附的官方歷史種子；這條路徑完全不需要連線
    # mopsov.twse.com.tw，因此 GitHub-hosted runner 不受該站封鎖／逾時影響。
    unresolved_codes = {code for code, _ in unresolved_pairs}
    seed_frame = load_bundled_mops_identity_seed()
    relevant_seed = seed_frame[seed_frame["code"].isin(unresolved_codes)].copy()
    if not relevant_seed.empty:
        identity_index = save_warrant_identity_index(
            [
                *identity_index.records(),
                *mops_frames_to_identity_records([relevant_seed]),
            ]
        )
        unresolved_pairs = {
            pair
            for pair in normalized_pairs
            if not identity_pair_is_matched(identity_index, pair)
        }
        print(
            "  🧭 隨附 MOPS 官方歷史種子補配後："
            f"剩餘 {len(unresolved_pairs):,}/{len(normalized_pairs):,} 組",
            flush=True,
        )
    if not unresolved_pairs:
        return identity_index
    if not ENABLE_MOPS_NETWORK_REFRESH:
        print(
            "  ℹ️ ENABLE_MOPS_NETWORK_REFRESH=0："
            "不從 GitHub runner 連線受限的 mopsov 主機；"
            "未命中的新資料由 FinMind Summary 或後續種子更新處理。",
            flush=True,
        )
        return identity_index

    earliest_trade_day = min(day for _, day in unresolved_pairs)
    latest_month_day = max(
        datetime.now().date().isoformat(),
        max(day for _, day in unresolved_pairs),
    )
    months = month_range(earliest_trade_day, latest_month_day)
    matched_before = len(normalized_pairs) - len(unresolved_pairs)
    print(
        "  🏛️ 啟動 MOPS 官方含下市權證歷史補抓："
        f"{len(unresolved_pairs):,} 個代號日期組合｜"
        f"到期月份 {months[0]}～{months[-1]}（{len(months):,} 月）",
        flush=True,
    )

    frames: list[pd.DataFrame] = []
    for month_index, month in enumerate(months, start=1):
        for market in ("1", "2"):
            frames.append(
                fetch_mops_warrant_identity(
                    market,
                    month=month,
                    active=False,
                )
            )
            if MOPS_REQUEST_DELAY_SECONDS:
                time.sleep(MOPS_REQUEST_DELAY_SECONDS)
        if month_index % 6 == 0 or month_index == len(months):
            print(
                f"  🏛️ MOPS 歷史補抓進度：{month_index:,}/{len(months):,} 月",
                flush=True,
            )

    official_records = mops_frames_to_identity_records(frames)
    repaired_index = save_warrant_identity_index(
        [*identity_index.records(), *official_records]
    )
    remaining_pairs = {
        pair
        for pair in normalized_pairs
        if not identity_pair_is_matched(repaired_index, pair)
    }

    # 未到期權證不會出現在「已下市」月份清單；僅在確實仍有缺口時抓一次
    # MOPS 目前掛牌清單。active 快取依 METADATA_CACHE_HOURS 更新。
    active_record_count = 0
    if remaining_pairs:
        print(
            f"  🏛️ 歷史到期批次後仍有 {len(remaining_pairs):,} 組未配對；"
            "補抓目前掛牌權證。",
            flush=True,
        )
        active_frames = [
            fetch_mops_warrant_identity(market, active=True)
            for market in ("1", "2")
        ]
        active_records = mops_frames_to_identity_records(active_frames)
        active_record_count = len(active_records)
        repaired_index = save_warrant_identity_index(
            [*repaired_index.records(), *active_records]
        )

    matched_after = sum(
        identity_pair_is_matched(repaired_index, pair)
        for pair in normalized_pairs
    )
    print(
        "  ✅ MOPS 官方身分補抓完成："
        f"歷史區間 {len(official_records):,} 筆｜"
        f"目前掛牌 {active_record_count:,} 筆｜"
        f"本批已配對 {matched_before:,} → {matched_after:,}/"
        f"{len(normalized_pairs):,}",
        flush=True,
    )
    return repaired_index


def normalize_warrant_day(
    raw: pd.DataFrame,
    day: str,
    broker_map: dict[str, tuple[str, str]],
) -> tuple[pd.DataFrame, int, int]:
    broker_by_code: dict[str, tuple[str, str, str]] = {}
    for label, (name, code) in broker_map.items():
        broker_by_code[normalize_broker_code(code)] = (label, name, code)
    work = raw.copy()
    work["stock_id"] = work["stock_id"].map(normalize_code)
    work["securities_trader_id"] = work["securities_trader_id"].map(
        normalize_broker_code
    )
    work = work[work["securities_trader_id"].isin(broker_by_code)].copy()
    if work.empty:
        return pd.DataFrame(columns=COMPACT_FACT_COLUMNS), 0, 0

    work["price"] = pd.to_numeric(work["price"], errors="coerce")
    work["buy"] = pd.to_numeric(work["buy"], errors="coerce")
    work["sell"] = pd.to_numeric(work["sell"], errors="coerce")
    invalid_numeric = (
        work[["price", "buy", "sell"]].isna().any(axis=1)
        | (work["price"] <= 0)
        | (work["buy"] < 0)
        | (work["sell"] < 0)
        | work["buy"].mod(1).ne(0)
        | work["sell"].mod(1).ne(0)
    )
    invalid_count = int(invalid_numeric.sum())
    if invalid_count:
        print(
            f"  ⚠️ {day} 略過 {invalid_count:,} 列無效價量資料；"
            "其餘有效列繼續處理。",
            flush=True,
        )
        work = work[~invalid_numeric].copy()
    if work.empty:
        raise RuntimeError("目標分點價量列全部無效，無法建立可信 compact")
    work["buy"] = work["buy"].astype("int64")
    work["sell"] = work["sell"].astype("int64")
    work = work[(work["buy"] != 0) | (work["sell"] != 0)].copy()
    work["_buy_amount"] = (work["price"] * work["buy"]).round().astype("int64")
    work["_sell_amount"] = (work["price"] * work["sell"]).round().astype("int64")
    work["_date"] = work["date"].map(iso_date)
    work["_broker_name"] = (
        work["securities_trader"].fillna("").astype(str).str.strip()
    )
    grouped = (
        work.groupby(
            ["stock_id", "securities_trader_id", "_date"],
            as_index=False,
            sort=False,
        )
        .agg(
            買進股數=("buy", "sum"),
            賣出股數=("sell", "sum"),
            買進金額=("_buy_amount", "sum"),
            賣出金額=("_sell_amount", "sum"),
            FinMind分點名稱=("_broker_name", "last"),
        )
    )
    before_dedup = len(grouped)
    records: list[dict[str, Any]] = []
    # name=None 很重要：Pandas namedtuple 會把前導底線欄位 `_date`
    # 自動改名成 `_2`，再用 row._asdict() 讀 `_date` 會造成 KeyError。
    for (
        warrant_code,
        broker_code,
        trade_day,
        buy_qty,
        sell_qty,
        buy_amount,
        sell_amount,
        finmind_broker_name,
    ) in grouped.itertuples(index=False, name=None):
        trade_day = trade_day or day
        label, configured_name, canonical_code = broker_by_code[broker_code]
        broker_name = finmind_broker_name or configured_name
        buy_qty = int(buy_qty)
        sell_qty = int(sell_qty)
        buy_amount = int(buy_amount)
        sell_amount = int(sell_amount)
        records.append(
            {
                "權證代號": warrant_code,
                "分點": label,
                "分點名稱": broker_name,
                "券商代號": canonical_code,
                "日期": trade_day,
                "買進股數": buy_qty,
                "賣出股數": sell_qty,
                "買進金額": buy_amount,
                "賣出金額": sell_amount,
                "買超股數": buy_qty - sell_qty,
                "買超金額": buy_amount - sell_amount,
            }
        )
    result = pd.DataFrame(records, columns=COMPACT_FACT_COLUMNS)
    duplicate_key = ["權證代號", "券商代號", "日期"]
    duplicates = result.duplicated(duplicate_key, keep=False)
    if duplicates.any():
        sample = result.loc[duplicates, duplicate_key].head(10).to_dict("records")
        raise RuntimeError(f"精簡資料出現重複權證＋券商＋日期：{sample}")
    result = result.sort_values(
        ["日期", "分點", "權證代號"]
    ).reset_index(drop=True)
    return result, before_dedup, invalid_count


def raw_path(day: str) -> Path:
    return RAW_DAILY_DIR / f"{day}.parquet"


def compact_path(day: str) -> Path:
    return COMPACT_DAILY_DIR / f"{day}.parquet"


def load_valid_raw_day(day: str) -> Optional[pd.DataFrame]:
    """讀取精簡失敗後留下的 raw，讓下次執行不必重新呼叫 FinMind。"""
    path = raw_path(day)
    if not path.exists() or path.stat().st_size <= 0:
        return None
    try:
        raw = read_parquet(path)
        missing = RAW_REQUIRED_COLUMNS - set(raw.columns)
        actual_dates = {iso_date(value) for value in raw["date"] if iso_date(value)}
        if raw.empty or missing or actual_dates != {day}:
            print(
                f"  ⚠️ {day} 暫存 raw 驗證失敗，將重新下載："
                f"missing={sorted(missing)} dates={sorted(actual_dates)}"
            )
            return None
        return raw
    except Exception as exc:
        print(
            f"  ⚠️ {day} 暫存 raw 無法讀取，將重新下載："
            f"{type(exc).__name__}: {exc}"
        )
        return None


def validate_saved_day(day: str) -> tuple[bool, str, int, int]:
    try:
        raw_rows = 0
        if KEEP_RAW_DAILY:
            raw = read_parquet(raw_path(day))
            missing_raw = RAW_REQUIRED_COLUMNS - set(raw.columns)
            if missing_raw:
                return False, f"raw 缺少欄位 {sorted(missing_raw)}", 0, 0
            raw_dates = {iso_date(value) for value in raw["date"] if iso_date(value)}
            if raw.empty or raw_dates != {day}:
                return False, f"raw 日期不符或為空：{sorted(raw_dates)}", 0, 0
            raw_rows = len(raw)
        compact = read_parquet(compact_path(day))
        missing_compact = set(COMPACT_FACT_COLUMNS) - set(compact.columns)
        if missing_compact:
            return False, f"compact 缺少欄位 {sorted(missing_compact)}", 0, 0
        compact_dates = {
            iso_date(value) for value in compact["日期"] if iso_date(value)
        }
        if compact_dates and compact_dates != {day}:
            return False, f"compact 日期不符：{sorted(compact_dates)}", 0, 0
        duplicate_key = ["權證代號", "券商代號", "日期"]
        if compact.duplicated(duplicate_key).any():
            return False, "compact 出現重複權證＋券商＋日期", 0, 0
        return True, "", raw_rows, len(compact)
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}", 0, 0


def audit_successful_files(state: dict[str, Any]) -> list[str]:
    valid: list[str] = []
    invalid: list[str] = []
    for day in list(state["successful_dates"]):
        ok, reason, raw_rows, compact_rows = validate_saved_day(day)
        if ok:
            valid.append(day)
            if KEEP_RAW_DAILY:
                state["raw_rows_by_date"][day] = raw_rows
            state["compact_rows_by_date"][day] = compact_rows
            state["deduplicated_rows_by_date"][day] = compact_rows
        else:
            invalid.append(day)
            state["invalid_numeric_rows_by_date"].pop(day, None)
            state["failure_reasons"][day] = f"啟動檢查失敗：{reason}"
            print(f"  ⚠️ {day} 成功檔案損壞，排入重新下載：{reason}")
    state["successful_dates"] = sorted(valid)
    state["failed_dates"] = sorted(set(state["failed_dates"]) | set(invalid))
    save_state(state)
    return invalid


def parse_force_dates(text: str) -> list[str]:
    values = [
        iso_date(part.strip())
        for part in re.split(r"[,;；、\s]+", text or "")
        if part.strip()
    ]
    invalid = [part for part in values if not part]
    if invalid:
        raise ValueError(f"FORCE_REDOWNLOAD_DATES 含無效日期：{invalid}")
    return sorted(set(value for value in values if value))


def discover_expected_dates(state: dict[str, Any]) -> list[str]:
    dates = trading_dates()
    override = iso_date(os.getenv("FINMIND_WARRANT_DATA_START_DATE", ""))
    metadata_start, source = discover_dataset_start_from_metadata()
    start = override or metadata_start
    state["availability_discovery_source"] = (
        "FINMIND_WARRANT_DATA_START_DATE" if override else source
    )
    expected = [day for day in dates if day >= start]
    if not expected:
        raise RuntimeError(f"交易日清單沒有 {start} 以後的日期")
    # 實際最早／最新「可用全市場日檔」在下載後以成功日期更新；這裡只建立候選交易日。
    state["first_available_date"] = min(state["successful_dates"], default="")
    state["latest_available_date"] = max(state["successful_dates"], default="")
    return expected


def mark_day_status(
    state: dict[str, Any],
    day: str,
    status: str,
    reason: str = "",
) -> None:
    for key in ("successful_dates", "failed_dates", "confirmed_empty_dates"):
        if day in state[key]:
            state[key].remove(day)
    if status == "ok":
        state["successful_dates"].append(day)
        state["failure_reasons"].pop(day, None)
        state["last_completed_date"] = day
    elif status == "empty":
        state["confirmed_empty_dates"].append(day)
        state["failure_reasons"].pop(day, None)
    else:
        state["failed_dates"].append(day)
        state["failure_reasons"][day] = reason
    state["first_available_date"] = min(state["successful_dates"], default="")
    state["latest_available_date"] = max(state["successful_dates"], default="")
    save_state(state)


def print_download_progress(
    state: dict[str, Any], new_compact_rows: int, processed: int
) -> None:
    successful = state["successful_dates"]
    print(
        "\n【下載進度摘要】\n"
        f"  本次已處理日數：{processed:,}\n"
        f"  目前成功日數：{len(successful):,}\n"
        f"  失敗日數：{len(state['failed_dates']):,}\n"
        f"  確認無資料日數：{len(state['confirmed_empty_dates']):,}\n"
        f"  最早成功日：{min(successful, default='-')}\n"
        f"  最新成功日：{max(successful, default='-')}\n"
        f"  本次新增列數：{new_compact_rows:,}\n"
        f"  累積精簡列數："
        f"{sum(int(v or 0) for v in state['compact_rows_by_date'].values()):,}\n"
    )


def download_and_compact_history(
    state: dict[str, Any],
    expected_dates: list[str],
    broker_map: dict[str, tuple[str, str]],
) -> None:
    if FORCE_RECALCULATE_STATS:
        print("ℹ️ FORCE_RECALCULATE_STATS=1：跳過全市場日檔網路下載。")
        return

    force_dates = parse_force_dates(FORCE_REDOWNLOAD_DATES)
    # 最新交易日可能在日檔發布前回覆「無資料」。最近 5 個交易日的 empty
    # 每次都重新探測，避免把「尚未發布」永久當成歷史缺檔。
    recent_probe_dates = set(expected_dates[-5:])
    for day in list(state["confirmed_empty_dates"]):
        if day in recent_probe_dates:
            state["confirmed_empty_dates"].remove(day)
    expected_set = set(expected_dates)
    out_of_range = [day for day in force_dates if day not in expected_set]
    if out_of_range:
        print(f"  ⚠️ 強制日期不在候選交易日清單，仍會嘗試：{out_of_range}")
        expected_dates = sorted(set(expected_dates) | set(out_of_range))

    for day in force_dates:
        for key in ("successful_dates", "failed_dates", "confirmed_empty_dates"):
            if day in state[key]:
                state[key].remove(day)
        state["raw_rows_by_date"].pop(day, None)
        state["compact_rows_by_date"].pop(day, None)
        state["deduplicated_rows_by_date"].pop(day, None)
        state["invalid_numeric_rows_by_date"].pop(day, None)
    save_state(state)

    completed = set(state["successful_dates"]) | set(state["confirmed_empty_dates"])
    failed_priority = [day for day in state["failed_dates"] if day in expected_dates]
    forced_priority = [day for day in force_dates if day in expected_dates]
    remaining = [day for day in expected_dates if day not in completed]
    queue = list(dict.fromkeys(forced_priority + failed_priority + remaining))

    if not queue:
        print("✅ 所有候選交易日已有明確狀態；沒有缺少的全市場日檔。")
        return

    print(
        f"【階段一】待處理 {len(queue):,} 個交易日；"
        f"失敗日優先 {len(failed_priority):,} 日。"
    )
    processed = 0
    new_compact_rows = 0
    last_local_error_signature = ""
    consecutive_local_errors = 0
    for day in queue:
        state["invalid_numeric_rows_by_date"].pop(day, None)
        try:
            cached_raw = load_valid_raw_day(day)
            result = (
                DownloadResult("ok", cached_raw, "沿用精簡失敗後保存的 raw")
                if cached_raw is not None
                else download_warrant_day(day)
            )
            if result.status == "empty":
                mark_day_status(state, day, "empty", result.reason)
                print(f"  ↪️ {day}：FinMind 明確回覆無全市場日檔")
            elif result.status != "ok":
                mark_day_status(state, day, "error", result.reason)
                print(f"  ❌ {day}：{result.reason}")
            else:
                raw = result.frame
                # 先保存 raw，再進行精簡。精簡或身分配對失敗時 raw 會留下，
                # 下次直接重用；compact 驗證成功後才依 KEEP_RAW_DAILY 決定刪除。
                if cached_raw is None:
                    atomic_write_parquet(raw, raw_path(day))
                compact, pre_dedup_rows, invalid_numeric_rows = normalize_warrant_day(
                    raw, day, broker_map
                )
                atomic_write_parquet(compact, compact_path(day))
                ok, reason, raw_rows, compact_rows = validate_saved_day(day)
                if not ok:
                    raise RuntimeError(f"落盤後驗證失敗：{reason}")
                state["raw_rows_by_date"][day] = len(raw)
                state["compact_rows_by_date"][day] = compact_rows
                state["deduplicated_rows_by_date"][day] = compact_rows
                state["invalid_numeric_rows_by_date"][day] = invalid_numeric_rows
                mark_day_status(state, day, "ok")
                if not KEEP_RAW_DAILY:
                    try:
                        raw_path(day).unlink(missing_ok=True)
                    except Exception as exc:
                        print(
                            f"  ⚠️ {day} compact 已成功，但暫存 raw 刪除失敗："
                            f"{type(exc).__name__}: {exc}"
                        )
                new_compact_rows += compact_rows
                print(
                    f"  ✅ {day}：原始 {len(raw):,}｜"
                    f"26分點事實列 {compact_rows:,}"
                )
            last_local_error_signature = ""
            consecutive_local_errors = 0
        except KeyboardInterrupt:
            print("\n⏹️ 使用者中止；已完成日期均已保存。")
            raise
        except Exception as exc:
            signature = f"{type(exc).__name__}: {exc}"
            mark_day_status(
                state, day, "error", signature
            )
            print(f"  ❌ {day}：{signature}")
            if signature == last_local_error_signature:
                consecutive_local_errors += 1
            else:
                last_local_error_signature = signature
                consecutive_local_errors = 1
            if consecutive_local_errors >= LOCAL_ERROR_CIRCUIT_BREAKER:
                raise RuntimeError(
                    f"連續 {consecutive_local_errors} 個交易日發生相同本機處理錯誤，"
                    f"已啟動斷路器避免持續浪費 runner 時間：{signature}"
                ) from exc
        processed += 1
        if processed % PROGRESS_EVERY_DAYS == 0:
            print_download_progress(state, new_compact_rows, processed)
    print_download_progress(state, new_compact_rows, processed)


def classify_amount_class(total_amount: Any) -> tuple[str, str]:
    try:
        amount = int(total_amount or 0)
    except Exception:
        amount = 0
    for code, label, lower, upper in AMOUNT_CLASS_SPECS:
        if amount >= lower and (upper is None or amount < upper):
            return code, label
    return "", ""


def event_unique_id(broker_code: str, underlying: str, event_day: str) -> str:
    return f"{normalize_code(broker_code)}|{normalize_code(underlying)}|{event_day}"


def preferred_group_text(values: Iterable[Any], fallback: str = "") -> str:
    """
    從同一穩定事件鍵的顯示名稱中選出可重現的代表值。

    分點名稱、標的名稱可能因資料來源、星號或歷史改名而不同；它們只能用來
    顯示，不能拆分事件。優先取出現次數最多者，同次數時取較短再按字典序。
    """
    texts = [
        str(value).strip()
        for value in values
        if value is not None
        and not pd.isna(value)
        and str(value).strip()
    ]
    if not texts:
        return str(fallback or "").strip()
    counts = Counter(texts)
    return min(
        counts,
        key=lambda text: (-counts[text], len(text), text),
    )


def enrich_compact_identity(
    frame: pd.DataFrame,
    identity_index: WarrantIdentityIndex,
) -> pd.DataFrame:
    """由 compact 事實欄位即時計算身分；忽略舊檔內曾寫入的配對結果。"""
    missing = set(COMPACT_FACT_COLUMNS) - set(frame.columns)
    if missing:
        raise RuntimeError(f"compact 事實資料缺少欄位：{sorted(missing)}")
    work = frame[COMPACT_FACT_COLUMNS].copy()
    if work.empty:
        return pd.DataFrame(columns=COMPACT_COLUMNS)
    work["權證代號"] = work["權證代號"].map(normalize_code)
    work["日期"] = work["日期"].map(iso_date)
    identity_by_key: dict[tuple[str, str], WarrantIdentity] = {}
    for warrant_code, trade_day in work[
        ["權證代號", "日期"]
    ].drop_duplicates().itertuples(index=False, name=None):
        key = (warrant_code, trade_day)
        identity_by_key[key] = identity_index.resolve(warrant_code, trade_day)
    identities = [
        identity_by_key[(warrant_code, trade_day)]
        for warrant_code, trade_day in work[
            ["權證代號", "日期"]
        ].itertuples(index=False, name=None)
    ]
    work["權證名稱"] = [identity.name for identity in identities]
    work["權證類型"] = [identity.warrant_type for identity in identities]
    work["標的股"] = [identity.target_code for identity in identities]
    work["標的名稱"] = [identity.target_name for identity in identities]
    work["身分配對狀態"] = [identity.status for identity in identities]
    work["身分配對錯誤原因"] = [identity.reason for identity in identities]
    return work[COMPACT_COLUMNS]


def print_identity_diagnostic(
    reason_counts: Counter[str],
    examples: list[tuple[str, str]],
    identity_index: WarrantIdentityIndex,
) -> None:
    print("  🔍 身分配對診斷：", flush=True)
    for reason, count in reason_counts.most_common(5):
        print(f"    原因｜{reason}：{count:,}", flush=True)
    for warrant_code, trade_day in examples[:5]:
        intervals = identity_index.by_code.get(normalize_code(warrant_code), [])
        preview = [
            (
                record.list_date,
                record.end_date,
                record.warrant_type,
                record.target_code,
            )
            for record in intervals[:3]
        ]
        print(
            f"    {warrant_code} @ {trade_day}｜索引區間 {len(intervals)} 筆："
            f"{preview}",
            flush=True,
        )


def preflight_identity_quality(
    state: dict[str, Any],
    identity_index: WarrantIdentityIndex,
    sample_days: int = 5,
    *,
    allow_official_backfill: bool = True,
) -> WarrantIdentityIndex:
    """在繼續下載前抽查既有 compact，避免錯誤索引一路跑完整段歷史。"""
    successful = sorted(state["successful_dates"])
    if not successful:
        return identity_index
    sample_count = min(max(sample_days, 1), len(successful))
    if sample_count == 1:
        selected = [successful[0]]
    else:
        positions = {
            round(index * (len(successful) - 1) / (sample_count - 1))
            for index in range(sample_count)
        }
        selected = [successful[position] for position in sorted(positions)]

    total_rows = 0
    unmatched_rows = 0
    reason_counts: Counter[str] = Counter()
    examples: list[tuple[str, str]] = []
    seen_examples: set[tuple[str, str]] = set()
    bad_pairs: set[tuple[str, str]] = set()
    for day in selected:
        enriched = enrich_compact_identity(
            read_parquet(compact_path(day)),
            identity_index,
        )
        total_rows += len(enriched)
        bad = enriched[enriched["身分配對狀態"] != "Matched"]
        unmatched_rows += len(bad)
        reason_counts.update(
            bad["身分配對錯誤原因"]
            .fillna("未提供原因")
            .astype(str)
            .value_counts()
            .to_dict()
        )
        for warrant_code, trade_day in bad[
            ["權證代號", "日期"]
        ].drop_duplicates().itertuples(index=False, name=None):
            key = (normalize_code(warrant_code), iso_date(trade_day) or "")
            if key[0] and key[1]:
                bad_pairs.add(key)
            if key not in seen_examples and len(examples) < 5:
                seen_examples.add(key)
                examples.append(key)

    unmatched_pct = unmatched_rows / total_rows * 100 if total_rows else 0.0
    print(
        f"  🩺 身分配對下載前抽查：{len(selected)} 日｜總列 {total_rows:,}｜"
        f"待確認 {unmatched_rows:,}（{unmatched_pct:.2f}%）",
        flush=True,
    )
    if unmatched_rows:
        print_identity_diagnostic(reason_counts, examples, identity_index)
    if total_rows and unmatched_pct > MAX_IDENTITY_UNMATCHED_PCT:
        if allow_official_backfill and bad_pairs:
            repaired_index = backfill_identity_from_official_mops(
                identity_index,
                bad_pairs,
            )
            return preflight_identity_quality(
                state,
                repaired_index,
                sample_days=sample_days,
                allow_official_backfill=False,
            )
        raise RuntimeError(
            f"下載前抽查的身分未配對率 {unmatched_pct:.2f}% 超過允許上限 "
            f"{MAX_IDENTITY_UNMATCHED_PCT:.2f}%；已停止繼續下載。"
            "既有 compact_daily 的交易事實仍可保留，修正索引後不必重抓。"
        )
    return identity_index


def build_events_for_day(frame: pd.DataFrame, day: str) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    valid = frame[
        (frame["身分配對狀態"] == "Matched")
        & frame["權證類型"].astype(str).str.contains("認購", na=False)
        & (frame["標的股"].astype(str).str.strip() != "")
        & (pd.to_numeric(frame["買進金額"], errors="coerce").fillna(0) > 0)
        & (pd.to_numeric(frame["買進股數"], errors="coerce").fillna(0) > 0)
    ].copy()
    if valid.empty:
        return []
    # 事件唯一性與 FIFO 都以代號事實為準。歷史資料中的分點名稱、標的名稱
    # 可能因星號、簡稱或資料來源不同而改變；若把顯示名稱放入 groupby，
    # 同一「券商＋標的＋日」會被拆成多筆，卻得到相同事件唯一 ID。
    valid["_事件券商代號"] = valid["券商代號"].map(normalize_code)
    valid["_事件標的股"] = valid["標的股"].map(normalize_code)
    valid["_事件日"] = [
        iso_date(value) or day
        for value in valid["日期"]
    ]
    valid = valid[
        (valid["_事件券商代號"] != "")
        & (valid["_事件標的股"] != "")
        & (valid["_事件日"] != "")
    ].copy()
    if valid.empty:
        return []
    events: list[dict[str, Any]] = []
    group_columns = [
        "_事件券商代號",
        "_事件標的股",
        "_事件日",
    ]
    for key, group in valid.groupby(group_columns, dropna=False, sort=False):
        broker_code, underlying, event_day = key
        broker = preferred_group_text(group["分點"], broker_code)
        broker_name = preferred_group_text(group["分點名稱"], broker)
        underlying_name = preferred_group_text(
            group["標的名稱"],
            underlying,
        )
        lots: list[dict[str, Any]] = []
        max_single = 0
        for warrant_code, warrant_group in group.groupby(
            "權證代號",
            dropna=False,
            sort=False,
        ):
            normalized_warrant_code = normalize_code(warrant_code)
            if not normalized_warrant_code:
                continue
            warrant_name = preferred_group_text(
                warrant_group["權證名稱"],
                normalized_warrant_code,
            )
            amount = int(
                pd.to_numeric(
                    warrant_group["買進金額"],
                    errors="coerce",
                ).fillna(0).sum()
            )
            quantity = int(
                pd.to_numeric(
                    warrant_group["買進股數"],
                    errors="coerce",
                ).fillna(0).sum()
            )
            if amount <= 0 or quantity <= 0:
                continue
            max_single = max(max_single, amount)
            lots.append(
                {
                    "買進日": event_day,
                    "權證代號": normalized_warrant_code,
                    "權證名稱": warrant_name,
                    "金額": amount,
                    "股數": quantity,
                }
            )
        if not lots or max_single < AMOUNT_THRESH:
            continue
        total_amount = sum(lot["金額"] for lot in lots)
        net_buy_quantity = int(
            pd.to_numeric(group["買超股數"], errors="coerce").fillna(0).sum()
        )
        code, label = classify_amount_class(total_amount)
        if not code:
            continue
        events.append(
            {
                "事件唯一ID": event_unique_id(broker_code, underlying, event_day),
                "分點": broker,
                "分點名稱": broker_name,
                "券商代號": broker_code,
                "標的股": underlying,
                "標的名稱": underlying_name,
                "事件日": event_day,
                "事件代碼": code,
                "事件類型": f"{code}-{label}",
                "單日累積買進金額": int(total_amount),
                "買超股數": net_buy_quantity,
                "涵蓋權證數": len({lot["權證代號"] for lot in lots}),
                "權證清單": "；".join(
                    f"{lot['權證代號']} {lot['權證名稱']}" for lot in lots
                ),
                "最大單檔買進金額": int(max_single),
                "lots": lots,
            }
        )
    return events


def checkpoint_frame(
    events: list[dict[str, Any]], last_day: str, checkpoint_type: str
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for event in events:
        row = {
            key: value
            for key, value in event.items()
            if key not in ("lots", "_fifo_allocations")
        }
        lots = event.get("lots", [])
        row["FIFO部位JSON"] = json.dumps(lots, ensure_ascii=False)
        row["尚未出清FIFO部位JSON"] = json.dumps(
            [lot for lot in lots if int(lot.get("剩餘股數", lot.get("股數", 0)) or 0) > 0],
            ensure_ascii=False,
        )
        row["已完成出清資料JSON"] = json.dumps(
            event.get("_fifo_allocations", []), ensure_ascii=False
        )
        row["最後處理日期"] = last_day
        row["checkpoint類型"] = checkpoint_type
        row["程式版本"] = PROGRAM_VERSION
        row["schema版本"] = SCHEMA_VERSION
        rows.append(row)
    return pd.DataFrame(rows)


def load_compact_frames(
    state: dict[str, Any],
    identity_index: WarrantIdentityIndex,
) -> tuple[list[pd.DataFrame], list[dict[str, Any]]]:
    frames: list[pd.DataFrame] = []
    events: list[dict[str, Any]] = []
    identity_rows = 0
    unmatched_rows = 0
    call_rows = 0
    put_rows = 0
    reason_counts: Counter[str] = Counter()
    diagnostic_examples: list[tuple[str, str]] = []
    seen_examples: set[tuple[str, str]] = set()
    successful = sorted(state["successful_dates"])
    stored_by_day: list[tuple[str, pd.DataFrame]] = []
    unresolved_pairs: set[tuple[str, str]] = set()
    print(
        "  ⏳ 掃描全部 compact 的權證代號日期，檢查是否需要 MOPS 官方歷史補抓...",
        flush=True,
    )
    for index, day in enumerate(successful, start=1):
        stored = read_parquet(compact_path(day))
        stored_by_day.append((day, stored))
        if {"權證代號", "日期"}.issubset(stored.columns):
            for warrant_code, trade_day in stored[
                ["權證代號", "日期"]
            ].drop_duplicates().itertuples(index=False, name=None):
                pair = (
                    normalize_code(warrant_code),
                    iso_date(trade_day) or "",
                )
                if (
                    pair[0]
                    and pair[1]
                    and not identity_pair_is_matched(identity_index, pair)
                ):
                    unresolved_pairs.add(pair)
        if index % PROGRESS_EVERY_DAYS == 0:
            print(
                f"    compact 身分掃描：{index:,}/{len(successful):,}｜"
                f"待補代號日期 {len(unresolved_pairs):,}",
                flush=True,
            )
    if unresolved_pairs:
        identity_index = backfill_identity_from_official_mops(
            identity_index,
            unresolved_pairs,
        )

    for index, (day, stored) in enumerate(stored_by_day, start=1):
        frame = enrich_compact_identity(stored, identity_index)
        frames.append(frame)
        identity_rows += len(frame)
        call_rows += int(
            frame["權證類型"].astype(str).str.contains("認購", na=False).sum()
        )
        put_rows += int(
            frame["權證類型"].astype(str).str.contains("認售", na=False).sum()
        )
        bad = frame[frame["身分配對狀態"] != "Matched"]
        unmatched_rows += len(bad)
        reason_counts.update(
            bad["身分配對錯誤原因"]
            .fillna("未提供原因")
            .astype(str)
            .value_counts()
            .to_dict()
        )
        for warrant_code, trade_day in bad[
            ["權證代號", "日期"]
        ].drop_duplicates().itertuples(index=False, name=None):
            key = (normalize_code(warrant_code), iso_date(trade_day) or "")
            if key not in seen_examples and len(diagnostic_examples) < 5:
                seen_examples.add(key)
                diagnostic_examples.append(key)
        events.extend(build_events_for_day(frame, day))
        if index % PROGRESS_EVERY_DAYS == 0:
            checkpoint = checkpoint_frame(events, day, "event-build")
            atomic_write_parquet(
                checkpoint, RESULTS_DIR / "event_checkpoint.parquet"
            )
            print(
                f"  💾 事件 checkpoint：{day}｜累積事件 {len(events):,}"
            )
    unmatched_pct = (
        unmatched_rows / identity_rows * 100 if identity_rows else 0.0
    )
    print(
        f"  🧭 階段二身分配對：總列 {identity_rows:,}｜"
        f"認購 {call_rows:,}｜認售 {put_rows:,}｜"
        f"待確認 {unmatched_rows:,}（{unmatched_pct:.2f}%）",
        flush=True,
    )
    if unmatched_rows:
        print_identity_diagnostic(
            reason_counts,
            diagnostic_examples,
            identity_index,
        )
    if identity_rows and unmatched_pct > MAX_IDENTITY_UNMATCHED_PCT:
        raise RuntimeError(
            f"身分未配對率 {unmatched_pct:.2f}% 超過允許上限 "
            f"{MAX_IDENTITY_UNMATCHED_PCT:.2f}%；已停止產生不可信報表。"
            "compact_daily 僅保存交易事實，修正索引後可直接重跑階段二，"
            "不必重新下載日檔。"
        )
    if successful:
        checkpoint = checkpoint_frame(events, successful[-1], "event-build")
        atomic_write_parquet(checkpoint, RESULTS_DIR / "event_checkpoint.parquet")
    ids = [event["事件唯一ID"] for event in events]
    duplicates = [key for key, count in Counter(ids).items() if count > 1]
    if duplicates:
        raise AssertionError(f"事件唯一ID重複：{duplicates[:10]}")
    return frames, events


def build_sale_rows(
    compact_frames: Iterable[pd.DataFrame],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    sales: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for frame in compact_frames:
        if frame.empty:
            continue
        work = frame[
            pd.to_numeric(frame["賣出股數"], errors="coerce").fillna(0) > 0
        ]
        for row in work[
            ["券商代號", "權證代號", "日期", "賣出股數", "賣出金額"]
        ].itertuples(index=False, name=None):
            broker_code, warrant_code, day, quantity, amount = row
            key = (normalize_code(broker_code), normalize_code(warrant_code))
            sales[key].append(
                {
                    "日期": iso_date(day),
                    "賣出股數": int(quantity or 0),
                    "賣出金額": float(amount or 0),
                }
            )
    for key in sales:
        sales[key].sort(key=lambda row: row["日期"])
    return sales


def simulate_group_outcomes_fifo(
    events: list[dict[str, Any]],
    sale_rows: dict[tuple[str, str], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    """沿用主程式的「券商代號＋權證代號、跨事件、舊事件優先」FIFO。"""
    queues: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for event_sequence, event in enumerate(events):
        event_day = iso_date(event["事件日"])
        event.update(
            {
                "目前狀態": "未出清",
                "減碼日": None,
                "出清日": None,
                "出清報酬%": None,
                "持有天數": None,
                "原始股數": 0,
                "剩餘股數": 0,
                "累計賣出股數": 0,
                "已實現賣出金額": 0.0,
                "已實現成本": 0.0,
                "_fifo_allocations": [],
            }
        )
        valid_lots = []
        for lot_sequence, lot in enumerate(event.get("lots", [])):
            quantity = int(lot.get("股數", 0) or 0)
            amount = float(lot.get("金額", 0) or 0)
            warrant_code = normalize_code(lot.get("權證代號"))
            broker_code = normalize_code(event.get("券商代號"))
            if quantity <= 0 or amount <= 0 or not warrant_code or not broker_code:
                continue
            lot.update(
                {
                    "均價": amount / quantity,
                    "剩餘股數": quantity,
                    "累計賣出股數": 0,
                    "已實現賣出金額": 0.0,
                    "已實現成本": 0.0,
                }
            )
            valid_lots.append(lot)
            queues[(broker_code, warrant_code)].append(
                {
                    "event": event,
                    "event_sequence": event_sequence,
                    "lot": lot,
                    "lot_sequence": lot_sequence,
                    "event_day": event_day,
                    "buy_day": iso_date(lot.get("買進日")) or event_day,
                }
            )
        event["lots"] = valid_lots
        event["原始股數"] = sum(int(lot["股數"]) for lot in valid_lots)
        event["剩餘股數"] = event["原始股數"]

    for key, queue in queues.items():
        queue.sort(
            key=lambda item: (
                item["event_day"],
                item["buy_day"],
                item["event_sequence"],
                item["lot_sequence"],
            )
        )
        for sale in sale_rows.get(key, []):
            sell_day = iso_date(sale["日期"])
            sell_quantity = int(sale["賣出股數"] or 0)
            sell_amount = float(sale["賣出金額"] or 0)
            if not sell_day or sell_quantity <= 0:
                continue
            sell_price = sell_amount / sell_quantity
            remaining_sale = sell_quantity
            for reference in queue:
                if remaining_sale <= 0:
                    break
                # 與現行主程式相同：事件日當日買賣不互扣。
                if sell_day <= reference["event_day"]:
                    continue
                lot = reference["lot"]
                remaining_lot = int(lot.get("剩餘股數", 0) or 0)
                if remaining_lot <= 0:
                    continue
                allocated = min(remaining_sale, remaining_lot)
                revenue = allocated * sell_price
                cost = allocated * float(lot["均價"])
                lot["剩餘股數"] -= allocated
                lot["累計賣出股數"] += allocated
                lot["已實現賣出金額"] += revenue
                lot["已實現成本"] += cost
                reference["event"]["_fifo_allocations"].append(
                    {
                        "日期": sell_day,
                        "股數": allocated,
                        "賣出金額": revenue,
                        "成本": cost,
                    }
                )
                remaining_sale -= allocated

    for event in events:
        original = sum(int(lot["股數"]) for lot in event["lots"])
        remaining = sum(int(lot["剩餘股數"]) for lot in event["lots"])
        sold = max(original - remaining, 0)
        by_day: dict[str, dict[str, float]] = defaultdict(
            lambda: {"股數": 0, "賣出金額": 0.0, "成本": 0.0}
        )
        for allocation in event["_fifo_allocations"]:
            values = by_day[allocation["日期"]]
            values["股數"] += int(allocation["股數"])
            values["賣出金額"] += float(allocation["賣出金額"])
            values["成本"] += float(allocation["成本"])
        event["原始股數"] = original
        event["剩餘股數"] = remaining
        event["累計賣出股數"] = sold
        event["已實現賣出金額"] = sum(v["賣出金額"] for v in by_day.values())
        event["已實現成本"] = sum(v["成本"] for v in by_day.values())
        if sold <= 0:
            event["目前狀態"] = "未出清"
            continue
        running = original
        cumulative_revenue = 0.0
        for sell_day in sorted(by_day):
            values = by_day[sell_day]
            running = max(running - int(values["股數"]), 0)
            cumulative_revenue += values["賣出金額"]
            if running > 0 and event["減碼日"] is None:
                event["減碼日"] = sell_day
            if running <= 0:
                total_cost = float(
                    sum(float(lot.get("金額", 0) or 0) for lot in event["lots"])
                )
                event["出清日"] = sell_day
                event["出清報酬%"] = (
                    round((cumulative_revenue - total_cost) / total_cost * 100, 2)
                    if total_cost
                    else None
                )
                event["持有天數"] = (
                    parse_date(sell_day) - parse_date(event["事件日"])
                ).days
                break
        event["目前狀態"] = "已出清" if remaining <= 0 else "減碼"
    return events


def add_calendar_months(day: str, months: int) -> str:
    parsed = pd.Timestamp(day)
    return (parsed + pd.DateOffset(months=months)).strftime("%Y-%m-%d")


def warrant_price_cache_path(contract: WarrantContract) -> Path:
    safe_code = re.sub(r"[^0-9A-Z_-]", "_", contract.code)
    suffix = f"{contract.list_date}_{contract.end_date}".replace("-", "")
    return WARRANT_PRICE_DIR / f"{safe_code}_{suffix}.parquet"


def normalize_warrant_price_frame(
    frame: pd.DataFrame, warrant_code: str
) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame(columns=["date", "stock_id", "close"])
    if "date" not in frame.columns or "close" not in frame.columns:
        raise RuntimeError(
            f"{WARRANT_PRICE_DATASET} schema 缺少 date／close："
            f"{list(frame.columns)}"
        )
    work = pd.DataFrame(
        {
            "date": frame["date"].map(iso_date),
            "stock_id": normalize_code(warrant_code),
            "close": pd.to_numeric(frame["close"], errors="coerce"),
        }
    )
    return (
        work[
            work["date"].ne("")
            & work["close"].notna()
            & work["close"].ge(0)
        ]
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )


def latest_warrant_market_price(
    contract: WarrantContract,
    valuation_day: str,
) -> tuple[float, str]:
    """取得合約本輪上市期間、統計截止日前最後一筆有效收盤價。"""
    query_end = min(valuation_day, contract.end_date)
    path = warrant_price_cache_path(contract)
    status_path = path.with_suffix(".json")
    try:
        cached = normalize_warrant_price_frame(read_parquet(path), contract.code)
    except Exception:
        cached = pd.DataFrame(columns=["date", "stock_id", "close"])
    status: dict[str, Any] = {}
    if status_path.exists():
        try:
            status = json.loads(status_path.read_text(encoding="utf-8"))
        except Exception:
            status = {}
    covered = (
        status.get("code") == contract.code
        and status.get("list_date") == contract.list_date
        and status.get("requested_end", "") >= query_end
    )
    if not covered:
        fetched = fetch_dataset(
            WARRANT_PRICE_DATASET,
            params={
                "data_id": contract.code,
                "start_date": contract.list_date,
                "end_date": query_end,
            },
            cache=False,
        )
        normalized = normalize_warrant_price_frame(fetched, contract.code)
        merged = (
            pd.concat([cached, normalized], ignore_index=True)
            .sort_values("date")
            .drop_duplicates("date", keep="last")
            .reset_index(drop=True)
        )
        atomic_write_parquet(merged, path)
        atomic_write_json(
            {
                "code": contract.code,
                "list_date": contract.list_date,
                "end_date": contract.end_date,
                "requested_end": query_end,
                "updated_at": datetime.now().isoformat(timespec="seconds"),
            },
            status_path,
        )
        cached = merged
    available = cached[
        (cached["date"] >= contract.list_date)
        & (cached["date"] <= query_end)
    ]
    if available.empty:
        raise RuntimeError(
            f"{contract.code} 在 {contract.list_date}～{query_end} "
            "查不到可用的權證收盤價"
        )
    latest = available.iloc[-1]
    price = float(latest["close"])
    if price < 0 or not math.isfinite(price):
        raise RuntimeError(f"{contract.code} 最後收盤價不合法：{price}")
    return price, str(latest["date"])


def warrant_cash_settlement_per_unit(contract: WarrantContract) -> float:
    if not contract.settlement_record_found:
        return 0.0
    values = (
        contract.exercise_ratio,
        contract.fulfillment_price,
        contract.settlement_price,
    )
    if any(value is None or not math.isfinite(float(value)) for value in values):
        raise RuntimeError(f"{contract.code} 官方結算紀錄缺少價格或行使比例")
    ratio = float(contract.exercise_ratio)
    strike = float(contract.fulfillment_price)
    settlement = float(contract.settlement_price)
    is_put = "認售" in contract.warrant_type
    intrinsic = (
        max(strike - settlement, 0.0)
        if is_put
        else max(settlement - strike, 0.0)
    )
    gross = intrinsic * ratio
    return gross * (1 - WARRANT_CASH_SETTLEMENT_TAX_RATE)


def evaluate_open_positions_after_two_months(
    events: list[dict[str, Any]],
    contract_index: WarrantContractIndex,
    valuation_day: str,
    *,
    price_loader: Any = latest_warrant_market_price,
) -> list[dict[str, Any]]:
    """為已出清及滿兩個曆月的未結部位建立可供勝率統計的認列結果。"""
    if not valuation_day:
        raise RuntimeError("缺少統計截止日，無法評價滿兩個月的未結部位")
    failures: list[str] = []
    market_price_cache: dict[tuple[str, str, str], tuple[float, str]] = {}
    aged_count = 0
    for event in events:
        event_day = iso_date(event.get("事件日"))
        threshold_day = add_calendar_months(
            event_day, OPEN_POSITION_EVALUATION_MONTHS
        )
        total_cost = float(
            sum(float(lot.get("金額", 0) or 0) for lot in event.get("lots", []))
        )
        realized = float(event.get("已實現賣出金額", 0) or 0)
        event.update(
            {
                "兩月門檻日": threshold_day,
                "勝敗認列狀態": "待認列",
                "勝敗結果": "待認列",
                "認列日": None,
                "認列報酬%": None,
                "認列持有天數": None,
                "總成本": total_cost,
                "已實現回收金額": realized,
                "剩餘部位評價金額": None,
                "估算總回收金額": None,
                "評價資料狀態": "NotDue",
                "評價錯誤原因": "",
                "評價明細": "",
            }
        )
        if event.get("目前狀態") == "已出清":
            event.update(
                {
                    "勝敗認列狀態": "實際FIFO出清",
                    "勝敗結果": result_tag(event.get("出清報酬%")),
                    "認列日": event.get("出清日"),
                    "認列報酬%": event.get("出清報酬%"),
                    "認列持有天數": event.get("持有天數"),
                    "剩餘部位評價金額": 0.0,
                    "估算總回收金額": realized,
                    "評價資料狀態": "Actual",
                    "評價明細": "實際FIFO完全出清",
                }
            )
            continue
        if valuation_day < threshold_day:
            event["評價錯誤原因"] = (
                f"統計截止日尚未滿 {OPEN_POSITION_EVALUATION_MONTHS} 個曆月"
            )
            continue

        aged_count += 1
        remaining_value = 0.0
        details: list[dict[str, Any]] = []
        statuses: set[str] = set()
        event_failures: list[str] = []
        for lot in event.get("lots", []):
            remaining_quantity = int(lot.get("剩餘股數", 0) or 0)
            if remaining_quantity <= 0:
                continue
            code = normalize_code(lot.get("權證代號"))
            buy_day = iso_date(lot.get("買進日")) or event_day
            contract = contract_index.resolve(code, buy_day)
            if contract is None:
                event_failures.append(f"{code}@{buy_day} 找不到該輪官方權證契約")
                continue
            expired = valuation_day >= contract.fulfillment_end_date
            if expired:
                per_unit = warrant_cash_settlement_per_unit(contract)
                value_source = (
                    "MOPS官方到期結算"
                    if contract.settlement_record_found
                    else "MOPS到期日已過且無自動現金結算紀錄"
                )
                value_day = contract.fulfillment_end_date
                statuses.add(
                    "到期結算" if contract.settlement_record_found else "到期無給付"
                )
            else:
                cache_key = (contract.code, contract.list_date, valuation_day)
                try:
                    if cache_key not in market_price_cache:
                        market_price_cache[cache_key] = price_loader(
                            contract, valuation_day
                        )
                        price_count = len(market_price_cache)
                        if price_count == 1 or price_count % 25 == 0:
                            print(
                                f"    權證市價評價進度：{price_count:,} 組｜"
                                f"目前 {contract.code}",
                                flush=True,
                            )
                    per_unit, value_day = market_price_cache[cache_key]
                except Exception as exc:
                    event_failures.append(
                        f"{code}@{buy_day} 市價評價失敗："
                        f"{type(exc).__name__}: {exc}"
                    )
                    continue
                value_source = f"FinMind {WARRANT_PRICE_DATASET} 收盤價"
                statuses.add("兩月市價評價")
            lot_value = remaining_quantity * float(per_unit)
            remaining_value += lot_value
            details.append(
                {
                    "權證代號": code,
                    "剩餘股數": remaining_quantity,
                    "每單位可回收金額": round(float(per_unit), 8),
                    "評價金額": round(lot_value, 2),
                    "評價日": value_day,
                    "來源": value_source,
                    "最後交易日": contract.end_date,
                    "履約截止日": contract.fulfillment_end_date,
                }
            )
        if event_failures:
            reason = "；".join(event_failures)
            event["評價資料狀態"] = "Error"
            event["評價錯誤原因"] = reason
            failures.append(f"{event.get('事件唯一ID')}｜{reason}")
            continue
        estimated_recovery = realized + remaining_value
        if total_cost <= 0:
            reason = "事件總成本小於等於 0"
            event["評價資料狀態"] = "Error"
            event["評價錯誤原因"] = reason
            failures.append(f"{event.get('事件唯一ID')}｜{reason}")
            continue
        recognized_return = round(
            (estimated_recovery - total_cost) / total_cost * 100, 2
        )
        if statuses == {"到期無給付"}:
            recognition_status = "到期無給付"
        elif "兩月市價評價" in statuses:
            recognition_status = "兩月市價評價"
        else:
            recognition_status = "到期結算"
        recognition_day = (
            valuation_day
            if "兩月市價評價" in statuses
            else max(
                (str(item["評價日"]) for item in details),
                default=valuation_day,
            )
        )
        event.update(
            {
                "勝敗認列狀態": recognition_status,
                "勝敗結果": result_tag(recognized_return),
                "認列日": recognition_day,
                "認列報酬%": recognized_return,
                "認列持有天數": (
                    parse_date(recognition_day) - parse_date(event_day)
                ).days,
                "剩餘部位評價金額": round(remaining_value, 2),
                "估算總回收金額": round(estimated_recovery, 2),
                "評價資料狀態": "OK",
                "評價錯誤原因": "",
                "評價明細": json.dumps(details, ensure_ascii=False),
            }
        )
    if failures:
        examples = "\n".join(f"  - {item}" for item in failures[:10])
        raise RuntimeError(
            f"滿 {OPEN_POSITION_EVALUATION_MONTHS} 個月事件仍有 "
            f"{len(failures):,} 筆無法可靠認列；為避免勝率失真已停止產表：\n"
            f"{examples}"
        )
    print(
        f"  💹 滿 {OPEN_POSITION_EVALUATION_MONTHS} 個月未結部位評價完成："
        f"{aged_count:,} 筆｜權證市價查詢 {len(market_price_cache):,} 組",
        flush=True,
    )
    return events


def price_cache_path(stock_id: str) -> Path:
    safe_code = re.sub(r"[^0-9A-Z_-]", "_", normalize_code(stock_id))
    return STOCK_PRICE_DIR / f"{safe_code}.parquet"


def normalize_price_frame(frame: pd.DataFrame, stock_id: str) -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame(columns=["date", "stock_id", "close"])
    date_column = next(
        (column for column in ("date", "Date", "日期") if column in frame.columns),
        None,
    )
    close_column = next(
        (
            column
            for column in ("close", "Close", "close_price", "收盤價")
            if column in frame.columns
        ),
        None,
    )
    if not date_column or not close_column:
        raise RuntimeError(
            f"{PRICE_DATASET} schema 缺少日期或還原收盤價：{list(frame.columns)}"
        )
    work = pd.DataFrame(
        {
            "date": frame[date_column].map(iso_date),
            "stock_id": normalize_code(stock_id),
            "close": pd.to_numeric(frame[close_column], errors="coerce"),
        }
    )
    work = work[
        (work["date"] != "")
        & work["close"].notna()
        & (work["close"] > 0)
    ].copy()
    return (
        work.sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )


def contiguous_missing_ranges(
    expected_dates: list[str], cached_dates: set[str]
) -> list[tuple[str, str]]:
    missing = [day for day in expected_dates if day not in cached_dates]
    if not missing:
        return []
    position = {day: index for index, day in enumerate(expected_dates)}
    ranges: list[tuple[str, str]] = []
    start = previous = missing[0]
    for day in missing[1:]:
        if position[day] != position[previous] + 1:
            ranges.append((start, previous))
            start = day
        previous = day
    ranges.append((start, previous))
    return ranges


def ensure_stock_price(
    stock_id: str,
    start_day: str,
    end_day: str,
    all_trading_dates: list[str],
) -> pd.DataFrame:
    path = price_cache_path(stock_id)
    try:
        cached = normalize_price_frame(read_parquet(path), stock_id)
    except Exception:
        cached = pd.DataFrame(columns=["date", "stock_id", "close"])
    expected = [
        day for day in all_trading_dates if start_day <= day <= end_day
    ]
    cached_dates = set(cached["date"]) if not cached.empty else set()
    ranges = contiguous_missing_ranges(expected, cached_dates)
    pieces = [cached]
    for range_start, range_end in ranges:
        try:
            fetched = fetch_dataset(
                PRICE_DATASET,
                params={
                    "data_id": stock_id,
                    "start_date": range_start,
                    "end_date": range_end,
                },
                cache=False,
            )
            normalized = normalize_price_frame(fetched, stock_id)
            pieces.append(normalized)
        except Exception as exc:
            print(
                f"  ⚠️ {stock_id} 還原股價缺口 {range_start}~{range_end}："
                f"{type(exc).__name__}: {exc}"
            )
    merged = (
        pd.concat(pieces, ignore_index=True)
        .sort_values("date")
        .drop_duplicates("date", keep="last")
        .reset_index(drop=True)
    )
    if not merged.empty:
        atomic_write_parquet(merged, path)
    return merged


def find_lookback_start(
    event_day: str, all_trading_dates: list[str], trading_days: int = 80
) -> str:
    prior = [day for day in all_trading_dates if day < event_day]
    return prior[max(len(prior) - trading_days, 0)] if prior else event_day


def build_trading_calendar(
    all_trading_dates: list[str],
) -> tuple[list[str], dict[str, int]]:
    """交易日曆只正規化與排序一次，供所有事件 O(1) 查位置。"""
    calendar = sorted(
        {
            normalized
            for value in all_trading_dates
            if (normalized := iso_date(value))
        }
    )
    return calendar, {day: index for index, day in enumerate(calendar)}


def prepare_ma_table(price_frame: pd.DataFrame) -> pd.DataFrame:
    work = normalize_price_frame(
        price_frame,
        str(price_frame["stock_id"].iloc[0]) if not price_frame.empty else "",
    )
    if work.empty:
        return work
    work = work.sort_values("date").reset_index(drop=True)
    for window in (5, 10, 20):
        work[f"ma{window}"] = work["close"].rolling(
            window=window, min_periods=window
        ).mean()
    work["prev_close"] = work["close"].shift(1)
    work["prev_ma5"] = work["ma5"].shift(1)
    work["prev_ma10"] = work["ma10"].shift(1)
    work["prev_ma20"] = work["ma20"].shift(1)
    work["ma20_5d_ago"] = work["ma20"].shift(5)
    return work


def missing_ma_values(reason: str) -> dict[str, Any]:
    values: dict[str, Any] = {
        "事件日收盤價": None,
        "MA5": None,
        "MA10": None,
        "MA20": None,
        "昨日收盤價": None,
        "昨日MA5": None,
        "昨日MA10": None,
        "昨日MA20": None,
        "5個交易日前MA20": None,
        "高於MA5": None,
        "高於MA10": None,
        "高於MA20": None,
        "剛站上MA5": None,
        "剛站上MA10": None,
        "剛站上MA20": None,
        "剛跌破MA5": None,
        "剛跌破MA10": None,
        "剛跌破MA20": None,
        "MA位置代碼": "",
        "MA位置名稱": "",
        "MA20方向": "",
        "均線多頭排列": None,
        "均線空頭排列": None,
        "均線資料狀態": "Missing",
        "均線錯誤原因": reason,
    }
    return values


def calculate_ma_values(
    ma_table: pd.DataFrame,
    event_day: str,
    tolerance: float = MA20_FLAT_TOLERANCE_PCT,
    expected_trading_dates: Optional[list[str]] = None,
    trading_calendar: Optional[tuple[list[str], dict[str, int]]] = None,
    available_dates: Optional[set[str]] = None,
    ma_rows_by_date: Optional[dict[str, dict[str, Any]]] = None,
) -> dict[str, Any]:
    if ma_table.empty:
        return missing_ma_values("無還原股價資料")
    if trading_calendar is None and expected_trading_dates is not None:
        trading_calendar = build_trading_calendar(expected_trading_dates)
    if trading_calendar is not None:
        calendar, calendar_index = trading_calendar
        event_index = calendar_index.get(event_day)
        if event_index is None:
            return missing_ma_values("事件日不在市場交易日曆中")
        if event_index < 24:
            return missing_ma_values("事件日前不足25個市場交易日，無法計算完整均線")
        required_dates = calendar[event_index - 24 : event_index + 1]
        if available_dates is None:
            available_dates = set(ma_table["date"].dropna().astype(str))
        missing_dates = [
            value for value in required_dates if value not in available_dates
        ]
        if missing_dates:
            preview = "、".join(missing_dates[:5])
            suffix = "…" if len(missing_dates) > 5 else ""
            return missing_ma_values(
                f"事件日前25個市場交易日股價有缺口：{preview}{suffix}"
            )
    if ma_rows_by_date is None:
        rows = ma_table[ma_table["date"] == event_day]
        if rows.empty:
            return missing_ma_values("事件日沒有還原收盤價；禁止使用事件日後價格補值")
        row: Any = rows.iloc[-1]
    else:
        row = ma_rows_by_date.get(event_day)
        if row is None:
            return missing_ma_values("事件日沒有還原收盤價；禁止使用事件日後價格補值")
    required = [
        "close",
        "ma5",
        "ma10",
        "ma20",
        "prev_close",
        "prev_ma5",
        "prev_ma10",
        "prev_ma20",
        "ma20_5d_ago",
    ]
    missing = [column for column in required if pd.isna(row[column])]
    if missing:
        return missing_ma_values(
            f"事件日前歷史不足，無法計算：{','.join(missing)}"
        )
    close = float(row["close"])
    ma5 = float(row["ma5"])
    ma10 = float(row["ma10"])
    ma20 = float(row["ma20"])
    prev_close = float(row["prev_close"])
    prev_ma5 = float(row["prev_ma5"])
    prev_ma10 = float(row["prev_ma10"])
    prev_ma20 = float(row["prev_ma20"])
    old_ma20 = float(row["ma20_5d_ago"])
    above5 = close > ma5
    above10 = close > ma10
    above20 = close > ma20
    position = f"{int(above5)}{int(above10)}{int(above20)}"
    base = abs(old_ma20)
    relative_difference = (
        (ma20 - old_ma20) / base if base > 0 else ma20 - old_ma20
    )
    if abs(relative_difference) <= tolerance:
        direction = "走平"
    elif relative_difference > 0:
        direction = "上彎"
    else:
        direction = "下彎"
    return {
        "事件日收盤價": close,
        "MA5": ma5,
        "MA10": ma10,
        "MA20": ma20,
        "昨日收盤價": prev_close,
        "昨日MA5": prev_ma5,
        "昨日MA10": prev_ma10,
        "昨日MA20": prev_ma20,
        "5個交易日前MA20": old_ma20,
        "高於MA5": above5,
        "高於MA10": above10,
        "高於MA20": above20,
        "剛站上MA5": close > ma5 and prev_close <= prev_ma5,
        "剛站上MA10": close > ma10 and prev_close <= prev_ma10,
        "剛站上MA20": close > ma20 and prev_close <= prev_ma20,
        "剛跌破MA5": close < ma5 and prev_close >= prev_ma5,
        "剛跌破MA10": close < ma10 and prev_close >= prev_ma10,
        "剛跌破MA20": close < ma20 and prev_close >= prev_ma20,
        "MA位置代碼": position,
        "MA位置名稱": MA_POSITION_NAMES[position],
        "MA20方向": direction,
        "均線多頭排列": ma5 > ma10 > ma20,
        "均線空頭排列": ma5 < ma10 < ma20,
        "均線資料狀態": "OK",
        "均線錯誤原因": "",
    }


def attach_ma_to_events(
    events: list[dict[str, Any]],
    all_trading_dates: list[str],
    statistics_day: str = "",
) -> list[dict[str, Any]]:
    trading_calendar = build_trading_calendar(all_trading_dates)
    by_stock: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        by_stock[normalize_code(event["標的股"])].append(event)
    latest_event_day = max(
        (event["事件日"] for event in events),
        default="",
    )
    latest_day = max(
        value
        for value in (
            iso_date(statistics_day),
            latest_event_day,
            max(all_trading_dates, default=iso_date(datetime.now())),
        )
        if value
    )
    for index, (stock_id, stock_events) in enumerate(sorted(by_stock.items()), start=1):
        earliest_event = min(event["事件日"] for event in stock_events)
        start_day = find_lookback_start(earliest_event, all_trading_dates, 80)
        price_frame = ensure_stock_price(
            stock_id, start_day, latest_day, all_trading_dates
        )
        ma_table = prepare_ma_table(price_frame)
        ma_rows_by_date = (
            ma_table.set_index("date", drop=False).to_dict(orient="index")
            if not ma_table.empty
            else {}
        )
        available_dates = set(ma_rows_by_date)
        for event in stock_events:
            event.update(
                calculate_ma_values(
                    ma_table,
                    event["事件日"],
                    trading_calendar=trading_calendar,
                    available_dates=available_dates,
                    ma_rows_by_date=ma_rows_by_date,
                )
            )
        if index % 20 == 0 or index == len(by_stock):
            print(f"  📈 還原股價／均線進度：{index:,}/{len(by_stock):,} 個標的")
    return events


EVENT_OUTPUT_COLUMNS = [
    "事件唯一ID",
    "分點",
    "分點名稱",
    "券商代號",
    "標的股",
    "標的名稱",
    "事件日",
    "事件代碼",
    "事件類型",
    "單日累積買進金額",
    "買超股數",
    "涵蓋權證數",
    "權證清單",
    "目前狀態",
    "出清日",
    "出清報酬%",
    "持有天數",
    "兩月門檻日",
    "勝敗認列狀態",
    "勝敗結果",
    "認列日",
    "認列報酬%",
    "認列持有天數",
    "總成本",
    "已實現回收金額",
    "剩餘部位評價金額",
    "估算總回收金額",
    "評價資料狀態",
    "評價錯誤原因",
    "評價明細",
    "事件日收盤價",
    "MA5",
    "MA10",
    "MA20",
    "昨日收盤價",
    "昨日MA5",
    "昨日MA10",
    "昨日MA20",
    "5個交易日前MA20",
    "高於MA5",
    "高於MA10",
    "高於MA20",
    "剛站上MA5",
    "剛站上MA10",
    "剛站上MA20",
    "剛跌破MA5",
    "剛跌破MA10",
    "剛跌破MA20",
    "MA位置代碼",
    "MA位置名稱",
    "MA20方向",
    "均線多頭排列",
    "均線空頭排列",
    "均線資料狀態",
    "均線錯誤原因",
]


def validate_event_outcomes(frame: pd.DataFrame) -> None:
    """驗證狀態與出清欄位彼此一致，避免用恆真式掩蓋資料錯誤。"""
    if frame.empty:
        return
    allowed_statuses = {"已出清", "未出清", "減碼"}
    statuses = set(frame["目前狀態"].dropna().astype(str))
    if frame["目前狀態"].isna().any() or not statuses.issubset(allowed_statuses):
        invalid = statuses - allowed_statuses
        if frame["目前狀態"].isna().any():
            invalid.add("<空值>")
        raise AssertionError(f"目前狀態不合法：{sorted(invalid)}")

    closed = frame[frame["目前狀態"] == "已出清"]
    if (
        closed["出清日"].fillna("").astype(str).str.strip().eq("").any()
        or pd.to_numeric(closed["出清報酬%"], errors="coerce").isna().any()
        or pd.to_numeric(closed["持有天數"], errors="coerce").isna().any()
    ):
        raise AssertionError("已出清事件缺少出清日、出清報酬或持有天數")

    open_events = frame[frame["目前狀態"].isin(["未出清", "減碼"])]
    if (
        open_events["出清日"]
        .fillna("")
        .astype(str)
        .str.strip()
        .ne("")
        .any()
    ):
        raise AssertionError("未出清或減碼事件不應填入出清日")

    allowed_recognition = {
        "實際FIFO出清",
        "兩月市價評價",
        "到期結算",
        "到期無給付",
        "待認列",
    }
    recognition_statuses = set(frame["勝敗認列狀態"].dropna().astype(str))
    if (
        frame["勝敗認列狀態"].isna().any()
        or not recognition_statuses.issubset(allowed_recognition)
    ):
        raise AssertionError(
            f"勝敗認列狀態不合法：{sorted(recognition_statuses - allowed_recognition)}"
        )
    resolved = frame[frame["勝敗認列狀態"] != "待認列"]
    if (
        resolved["認列日"].fillna("").astype(str).str.strip().eq("").any()
        or pd.to_numeric(resolved["認列報酬%"], errors="coerce").isna().any()
        or pd.to_numeric(resolved["認列持有天數"], errors="coerce").isna().any()
        or (~resolved["勝敗結果"].isin(["勝", "敗", "平手"])).any()
    ):
        raise AssertionError("已認列事件缺少認列日、報酬、持有天數或勝敗結果")
    expected_tags = pd.to_numeric(
        resolved["認列報酬%"], errors="coerce"
    ).map(result_tag)
    if not expected_tags.equals(resolved["勝敗結果"]):
        raise AssertionError("勝敗結果與認列報酬%不一致")
    pending = frame[frame["勝敗認列狀態"] == "待認列"]
    if (
        pd.to_numeric(pending["認列報酬%"], errors="coerce").notna().any()
        or pending["勝敗結果"].ne("待認列").any()
    ):
        raise AssertionError("待認列事件不得預填認列報酬或勝敗結果")


def events_to_frame(events: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for event in events:
        rows.append({column: event.get(column) for column in EVENT_OUTPUT_COLUMNS})
    frame = pd.DataFrame(rows, columns=EVENT_OUTPUT_COLUMNS)
    if frame.empty:
        return frame
    if frame["事件唯一ID"].duplicated().any():
        raise AssertionError("同一券商、標的、事件日產生重複事件")
    invalid_codes = set(frame["事件代碼"]) - set("ABCDE")
    if invalid_codes:
        raise AssertionError(f"事件代碼不在 A~E：{invalid_codes}")
    validate_event_outcomes(frame)
    valid_ma = frame["均線資料狀態"] == "OK"
    calculated_codes = (
        frame.loc[valid_ma, ["高於MA5", "高於MA10", "高於MA20"]]
        .astype(int)
        .astype(str)
        .agg("".join, axis=1)
    )
    if not calculated_codes.equals(frame.loc[valid_ma, "MA位置代碼"]):
        raise AssertionError("MA位置代碼與高於MA5／10／20不一致")
    return frame


def sample_grade(closed_count: int) -> str:
    if closed_count < 15:
        return "樣本不足"
    if closed_count < 30:
        return "初步觀察"
    if closed_count < 60:
        return "具參考性"
    return "相對穩定"


def result_tag(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    numeric = float(value)
    if numeric > 0:
        return "勝"
    if numeric < 0:
        return "敗"
    return "平手"


def summarize_event_group(
    group: pd.DataFrame,
    *,
    broker: str,
    condition: str = "",
    position_code: str = "",
    event_code: str = "ALL",
) -> dict[str, Any]:
    closed = group[group["目前狀態"] == "已出清"].copy()
    resolved = group[group["勝敗認列狀態"] != "待認列"].copy()
    returns = pd.to_numeric(resolved["認列報酬%"], errors="coerce").dropna()
    holding = pd.to_numeric(resolved["認列持有天數"], errors="coerce").dropna()
    resolved = resolved.loc[returns.index] if len(returns) else resolved.iloc[0:0]
    result_tags = returns.map(result_tag)
    total_amount = int(
        pd.to_numeric(group["單日累積買進金額"], errors="coerce")
        .fillna(0)
        .sum()
    )
    resolved_amount_series = pd.to_numeric(
        resolved["總成本"], errors="coerce"
    ).fillna(0)
    resolved_amount = int(resolved_amount_series.sum())
    pnl = (
        float((resolved_amount_series * returns / 100).sum())
        if len(returns)
        else 0.0
    )
    event_type = (
        "全部-A+B+C+D+E合併"
        if event_code == "ALL"
        else f"{event_code}-{AMOUNT_CLASS_LABELS[event_code]}"
    )
    closed_count = len(closed)
    resolved_count = len(resolved)
    pending_count = len(group) - resolved_count
    row = {
        "分點": broker,
        "均線條件": condition,
        "MA位置代碼": position_code,
        "事件代碼": event_code,
        "事件類型": event_type,
        "事件總數": len(group),
        "已出清筆數": closed_count,
        "未出清筆數": len(group) - closed_count,
        "勝敗樣本數": resolved_count,
        "實際出清筆數": int(
            (resolved["勝敗認列狀態"] == "實際FIFO出清").sum()
        ),
        "兩月以上評價筆數": int(
            (resolved["勝敗認列狀態"] == "兩月市價評價").sum()
        ),
        "到期結算筆數": int(
            resolved["勝敗認列狀態"].isin(["到期結算", "到期無給付"]).sum()
        ),
        "尚未認列筆數": pending_count,
        "勝筆數": int((result_tags == "勝").sum()),
        "敗筆數": int((result_tags == "敗").sum()),
        "平手筆數": int((result_tags == "平手").sum()),
        "勝率": round(float((result_tags == "勝").sum()) / resolved_count * 100, 2)
        if resolved_count
        else None,
        "平均持有天數": round(float(holding.mean()), 2) if len(holding) else None,
        "中位數持有天數": round(float(holding.median()), 2) if len(holding) else None,
        "平均報酬%": round(float(returns.mean()), 2) if len(returns) else None,
        "中位數報酬%": round(float(returns.median()), 2) if len(returns) else None,
        "加權報酬%": round(pnl / resolved_amount * 100, 2)
        if resolved_amount
        else None,
        "總買進金額": total_amount,
        "已認列買進金額": resolved_amount,
        "已出清買進金額": int(
            pd.to_numeric(
                closed["單日累積買進金額"], errors="coerce"
            ).fillna(0).sum()
        ),
        "估算損益金額": int(round(pnl)),
        "最高報酬%": round(float(returns.max()), 2) if len(returns) else None,
        "最低報酬%": round(float(returns.min()), 2) if len(returns) else None,
        "樣本等級": sample_grade(resolved_count),
        "資料起始日": group["事件日"].min() if len(group) else None,
        "資料結束日": group["事件日"].max() if len(group) else None,
    }
    if row["勝筆數"] + row["敗筆數"] + row["平手筆數"] != resolved_count:
        raise AssertionError("勝＋敗＋平手不等於勝敗樣本數")
    if resolved_count + pending_count != len(group):
        raise AssertionError("勝敗樣本數＋尚未認列筆數不等於事件總數")
    return row


def build_condition_statistics(events: pd.DataFrame) -> pd.DataFrame:
    valid = events[events["均線資料狀態"] == "OK"].copy()
    brokers = sorted(set(events["分點"].dropna())) if not events.empty else []
    rows: list[dict[str, Any]] = []
    for broker in brokers:
        broker_group = valid[valid["分點"] == broker]
        for condition_name, predicate in CONDITION_SPECS:
            condition_group = broker_group[predicate(broker_group)]
            for code in ["ALL", "A", "B", "C", "D", "E"]:
                group = (
                    condition_group
                    if code == "ALL"
                    else condition_group[condition_group["事件代碼"] == code]
                )
                rows.append(
                    summarize_event_group(
                        group,
                        broker=broker,
                        condition=condition_name,
                        event_code=code,
                    )
                )
    return pd.DataFrame(rows)


def build_position_statistics(events: pd.DataFrame) -> pd.DataFrame:
    valid = events[events["均線資料狀態"] == "OK"].copy()
    brokers = sorted(set(events["分點"].dropna())) if not events.empty else []
    rows: list[dict[str, Any]] = []
    for broker in brokers:
        broker_group = valid[valid["分點"] == broker]
        for position_code in MA_POSITION_NAMES:
            position_group = broker_group[
                broker_group["MA位置代碼"] == position_code
            ]
            for code in ["ALL", "A", "B", "C", "D", "E"]:
                group = (
                    position_group
                    if code == "ALL"
                    else position_group[position_group["事件代碼"] == code]
                )
                rows.append(
                    summarize_event_group(
                        group,
                        broker=broker,
                        position_code=position_code,
                        event_code=code,
                    )
                )
    return pd.DataFrame(rows)


def best_qualified(
    stats: pd.DataFrame, label_column: str
) -> str:
    if stats.empty:
        return "樣本不足，暫不判定"
    eligible = stats[
        (stats["事件代碼"] == "ALL")
        & (stats["勝敗樣本數"] >= 15)
        & stats["勝率"].notna()
    ]
    if eligible.empty:
        return "樣本不足，暫不判定"
    best = eligible.sort_values(
        ["勝率", "勝敗樣本數", "事件總數"],
        ascending=[False, False, False],
    ).iloc[0]
    return (
        f"{best[label_column]}（勝率{best['勝率']:.2f}%，"
        f"勝敗樣本{int(best['勝敗樣本數'])}筆）"
    )


def build_broker_profiles(
    events: pd.DataFrame,
    condition_stats: pd.DataFrame,
    position_stats: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for broker, group in events.groupby("分點", sort=True):
        closed = group[group["目前狀態"] == "已出清"].copy()
        resolved = group[group["勝敗認列狀態"] != "待認列"].copy()
        returns = pd.to_numeric(resolved["認列報酬%"], errors="coerce")
        amounts = pd.to_numeric(
            resolved["總成本"], errors="coerce"
        ).fillna(0)
        holding = pd.to_numeric(
            resolved["認列持有天數"], errors="coerce"
        ).dropna()
        valid_positions = group.loc[
            group["均線資料狀態"] == "OK", "MA位置代碼"
        ]
        most_position = (
            valid_positions.value_counts().index[0]
            if not valid_positions.empty
            else "Missing"
        )
        pnl = float((amounts * returns.fillna(0) / 100).sum())
        total_closed_amount = float(amounts[returns.notna()].sum())
        rows.append(
            {
                "分點": broker,
                "資料起始日": group["事件日"].min(),
                "資料結束日": group["事件日"].max(),
                "事件總數": len(group),
                "已出清事件數": len(closed),
                "未出清事件數": len(group) - len(closed),
                "勝敗樣本數": len(resolved),
                "兩月以上評價事件數": int(
                    (resolved["勝敗認列狀態"] == "兩月市價評價").sum()
                ),
                "到期結算事件數": int(
                    resolved["勝敗認列狀態"]
                    .isin(["到期結算", "到期無給付"])
                    .sum()
                ),
                "尚未認列事件數": len(group) - len(resolved),
                "最常出現MA位置": most_position,
                "最高勝率均線條件": best_qualified(
                    condition_stats[condition_stats["分點"] == broker], "均線條件"
                ),
                "最高勝率MA位置": best_qualified(
                    position_stats[position_stats["分點"] == broker], "MA位置代碼"
                ),
                "平均持有天數": round(float(holding.mean()), 2)
                if len(holding)
                else None,
                "中位數持有天數": round(float(holding.median()), 2)
                if len(holding)
                else None,
                "平均報酬率": round(float(returns.mean()), 2)
                if returns.notna().any()
                else None,
                "中位數報酬率": round(float(returns.median()), 2)
                if returns.notna().any()
                else None,
                "加權報酬率": round(pnl / total_closed_amount * 100, 2)
                if total_closed_amount
                else None,
            }
        )
    return pd.DataFrame(rows)


def build_integrity_report(
    state: dict[str, Any],
    expected_dates: list[str],
    events: pd.DataFrame,
) -> pd.DataFrame:
    successful = set(state["successful_dates"])
    failed = set(state["failed_dates"])
    empty = set(state["confirmed_empty_dates"])
    unprocessed = sorted(set(expected_dates) - successful - failed - empty)
    missing_ma = (
        int((events["均線資料狀態"] == "Missing").sum())
        if not events.empty
        else 0
    )
    complete_ma = (
        int((events["均線資料狀態"] == "OK").sum())
        if not events.empty
        else 0
    )
    closed = (
        int((events["目前狀態"] == "已出清").sum())
        if not events.empty
        else 0
    )
    resolved = (
        int((events["勝敗認列狀態"] != "待認列").sum())
        if not events.empty
        else 0
    )
    missing_reason_summary = "-"
    if missing_ma:
        reasons = (
            events.loc[
                events["均線資料狀態"] == "Missing",
                "均線錯誤原因",
            ]
            .fillna("未提供原因")
            .astype(str)
            .str.strip()
            .replace("", "未提供原因")
            .str.split("：", n=1)
            .str[0]
            .value_counts()
        )
        top_reasons = reasons.head(10)
        parts = [f"{reason}：{int(count):,}" for reason, count in top_reasons.items()]
        remaining_count = int(reasons.iloc[10:].sum())
        if remaining_count:
            parts.append(f"其他原因：{remaining_count:,}")
        missing_reason_summary = "；".join(parts)
    actual_first = min(successful, default="")
    actual_latest = max(successful, default="")
    complete = not failed and not unprocessed
    mops_identity_cache_count = len(
        list(
            METADATA_DIR.glob(
                f"{MOPS_WARRANT_HISTORY_CACHE_PREFIX}_*.parquet"
            )
        )
    )
    rows = [
        ("歷史資料狀態", "完整" if complete else "尚未完整"),
        ("FinMind最早可用日期", actual_first or "尚未確認"),
        ("FinMind最新可用日期", actual_latest or "尚未確認"),
        (
            "可用範圍探索來源",
            state.get("availability_discovery_source", ""),
        ),
        (
            "權證身分來源",
            "FinMind Summary 優先；隨附 MOPS 含下市權證官方歷史、"
            "契約與到期結算種子備援",
        ),
        (
            "MOPS官方身分種子檔",
            (
                resolve_mops_identity_seed_path(required=False).name
                if resolve_mops_identity_seed_path(required=False)
                else "缺少"
            ),
        ),
        ("MOPS官方身分快取分片數", mops_identity_cache_count),
        ("候選範圍起始日", min(expected_dates, default="")),
        ("候選範圍結束日", max(expected_dates, default="")),
        ("預期交易日數", len(expected_dates)),
        ("成功日數", len(successful)),
        ("失敗日數", len(failed)),
        ("確認無資料日數", len(empty)),
        ("尚未處理日數", len(unprocessed)),
        ("失敗日期清單", "、".join(sorted(failed)) or "-"),
        ("尚未處理日期清單", "、".join(unprocessed) or "-"),
        (
            "原始資料總列數",
            sum(int(v or 0) for v in state["raw_rows_by_date"].values()),
        ),
        (
            "過濾26分點後列數",
            sum(int(v or 0) for v in state["compact_rows_by_date"].values()),
        ),
        (
            "去重複後列數",
            sum(int(v or 0) for v in state["deduplicated_rows_by_date"].values()),
        ),
        (
            "無效價量略過列數（已記錄成功日）",
            sum(
                int(v or 0)
                for v in state["invalid_numeric_rows_by_date"].values()
            ),
        ),
        (
            "無效價量列數已記錄日數",
            len(state["invalid_numeric_rows_by_date"]),
        ),
        ("事件總數", len(events)),
        ("有完整均線事件數", complete_ma),
        ("均線Missing事件數", missing_ma),
        ("均線Missing原因分布（前10類）", missing_reason_summary),
        ("已出清事件數", closed),
        ("未出清事件數", len(events) - closed),
        ("勝敗樣本數", resolved),
        (
            f"滿{OPEN_POSITION_EVALUATION_MONTHS}月市價評價事件數",
            int((events["勝敗認列狀態"] == "兩月市價評價").sum())
            if not events.empty
            else 0,
        ),
        (
            "到期結算／無給付事件數",
            int(
                events["勝敗認列狀態"]
                .isin(["到期結算", "到期無給付"])
                .sum()
            )
            if not events.empty
            else 0,
        ),
        ("尚未認列事件數", len(events) - resolved),
        ("勝敗統計截止日", actual_latest),
        ("最早事件日", events["事件日"].min() if not events.empty else ""),
        ("最新事件日", events["事件日"].max() if not events.empty else ""),
        ("程式版本", PROGRAM_VERSION),
        ("schema版本", SCHEMA_VERSION),
        ("產生時間", datetime.now().isoformat(timespec="seconds")),
    ]
    if not events.empty:
        for broker, count in events["分點"].value_counts().sort_index().items():
            rows.append((f"各分點事件數｜{broker}", int(count)))
    return pd.DataFrame(rows, columns=["項目", "內容"])


def dataframe_for_excel(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    for column in output.columns:
        if output[column].dtype == "object":
            output[column] = output[column].map(
                lambda value: ""
                if value is None
                else value
                if isinstance(value, (str, int, float, bool, datetime, date))
                else json.dumps(value, ensure_ascii=False)
            )
    return output


def style_excel(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sample_rows = min(worksheet.max_row, 300)
        for column_index in range(1, worksheet.max_column + 1):
            values = [
                str(worksheet.cell(row_index, column_index).value or "")
                for row_index in range(1, sample_rows + 1)
            ]
            width = min(max(max((len(value) for value in values), default=0) + 2, 10), 42)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=False)
    workbook.save(path)


def write_outputs(
    events: pd.DataFrame,
    condition_stats: pd.DataFrame,
    position_stats: pd.DataFrame,
    profiles: pd.DataFrame,
    integrity: pd.DataFrame,
) -> Path:
    atomic_write_parquet(events, RESULTS_DIR / "events.parquet")
    atomic_write_parquet(
        condition_stats, RESULTS_DIR / "ma_condition_stats.parquet"
    )
    atomic_write_parquet(
        position_stats, RESULTS_DIR / "ma_position_stats.parquet"
    )
    atomic_write_parquet(profiles, RESULTS_DIR / "broker_profiles.parquet")
    output_path = OUTPUT_DIR / f"warrant_broker_ma_history_{datetime.now():%Y%m%d}.xlsx"
    temporary = output_path.with_name(
        f"{output_path.name}.tmp.{os.getpid()}.xlsx"
    )
    try:
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            dataframe_for_excel(events).to_excel(
                writer, sheet_name="歷史均線事件明細", index=False
            )
            dataframe_for_excel(condition_stats).to_excel(
                writer, sheet_name="分點均線條件統計", index=False
            )
            dataframe_for_excel(position_stats).to_excel(
                writer, sheet_name="分點MA位置統計", index=False
            )
            dataframe_for_excel(profiles).to_excel(
                writer, sheet_name="分點操作指紋", index=False
            )
            dataframe_for_excel(integrity).to_excel(
                writer, sheet_name="歷史資料完整性", index=False
            )
        style_excel(temporary)
        os.replace(temporary, output_path)
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)
    return output_path


def validate_final_results(
    events: pd.DataFrame,
    condition_stats: pd.DataFrame,
    position_stats: pd.DataFrame,
) -> None:
    if not events.empty:
        if events["事件唯一ID"].duplicated().any():
            raise AssertionError("事件唯一ID不得重複")
        if not set(events["事件代碼"]).issubset(set("ABCDE")):
            raise AssertionError("每筆事件只能屬於 A~E 其中一類")
        validate_event_outcomes(events)
    for frame, label in (
        (condition_stats, "均線條件"),
        (position_stats, "MA位置"),
    ):
        if frame.empty:
            continue
        bad = frame[
            frame["勝筆數"] + frame["敗筆數"] + frame["平手筆數"]
            != frame["勝敗樣本數"]
        ]
        if not bad.empty:
            raise AssertionError(f"{label}統計勝敗平手驗證失敗")
        if (frame["事件總數"] != frame["已出清筆數"] + frame["未出清筆數"]).any():
            raise AssertionError(f"{label}統計事件數驗證失敗")
        if (
            frame["事件總數"]
            != frame["勝敗樣本數"] + frame["尚未認列筆數"]
        ).any():
            raise AssertionError(f"{label}統計勝敗認列事件數驗證失敗")

    valid = events[events["均線資料狀態"] == "OK"].copy()
    brokers = sorted(set(events["分點"].dropna())) if not events.empty else []
    event_codes = ["ALL", "A", "B", "C", "D", "E"]
    condition_predicates = dict(CONDITION_SPECS)
    valid_by_broker = {
        broker: valid[valid["分點"] == broker]
        for broker in brokers
    }
    condition_groups = {
        (broker, condition): broker_group[predicate(broker_group)]
        for broker, broker_group in valid_by_broker.items()
        for condition, predicate in CONDITION_SPECS
    }
    position_groups = {
        (broker, position): broker_group[
            broker_group["MA位置代碼"] == position
        ]
        for broker, broker_group in valid_by_broker.items()
        for position in MA_POSITION_NAMES
    }

    expected_condition_keys = {
        (broker, condition, code)
        for broker in brokers
        for condition in condition_predicates
        for code in event_codes
    }
    actual_condition_keys = [
        (row["分點"], row["均線條件"], row["事件代碼"])
        for _, row in condition_stats.iterrows()
    ]
    if (
        len(actual_condition_keys) != len(set(actual_condition_keys))
        or set(actual_condition_keys) != expected_condition_keys
    ):
        raise AssertionError("均線條件統計的分點／條件／事件代碼索引不完整或重複")

    expected_position_keys = {
        (broker, position, code)
        for broker in brokers
        for position in MA_POSITION_NAMES
        for code in event_codes
    }
    actual_position_keys = [
        (row["分點"], row["MA位置代碼"], row["事件代碼"])
        for _, row in position_stats.iterrows()
    ]
    if (
        len(actual_position_keys) != len(set(actual_position_keys))
        or set(actual_position_keys) != expected_position_keys
    ):
        raise AssertionError("MA位置統計的分點／位置／事件代碼索引不完整或重複")

    def values_match(actual: Any, expected: Any) -> bool:
        if pd.isna(actual) and pd.isna(expected):
            return True
        if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
            return math.isclose(float(actual), float(expected), rel_tol=1e-12, abs_tol=1e-9)
        return actual == expected

    def verify_row(
        actual: pd.Series,
        source_group: pd.DataFrame,
        *,
        label: str,
        condition: str = "",
        position_code: str = "",
    ) -> None:
        expected = summarize_event_group(
            source_group,
            broker=actual["分點"],
            condition=condition,
            position_code=position_code,
            event_code=actual["事件代碼"],
        )
        for column, expected_value in expected.items():
            if column not in actual.index or not values_match(
                actual[column], expected_value
            ):
                raise AssertionError(
                    f"{label}統計無法追溯事件明細："
                    f"{actual['分點']}／{condition or position_code}／"
                    f"{actual['事件代碼']}／{column}"
                )

    for _, row in condition_stats.iterrows():
        source = condition_groups[(row["分點"], row["均線條件"])]
        if row["事件代碼"] != "ALL":
            source = source[source["事件代碼"] == row["事件代碼"]]
        verify_row(
            row,
            source,
            label="均線條件",
            condition=row["均線條件"],
        )

    for _, row in position_stats.iterrows():
        source = position_groups[(row["分點"], row["MA位置代碼"])]
        if row["事件代碼"] != "ALL":
            source = source[source["事件代碼"] == row["事件代碼"]]
        verify_row(
            row,
            source,
            label="MA位置",
            position_code=row["MA位置代碼"],
        )


def make_test_ma_row(**overrides: Any) -> pd.DataFrame:
    row = {
        "date": "2026-01-30",
        "stock_id": "2330",
        "close": 120.0,
        "ma5": 115.0,
        "ma10": 110.0,
        "ma20": 105.0,
        "prev_close": 100.0,
        "prev_ma5": 104.0,
        "prev_ma10": 103.0,
        "prev_ma20": 101.0,
        "ma20_5d_ago": 100.0,
    }
    row.update(overrides)
    return pd.DataFrame([row])


def run_self_tests() -> dict[str, str]:
    """不連網的小型測試；失敗時直接拋出 AssertionError。"""
    bundled_seed_test = load_bundled_mops_identity_seed()
    bundled_seed_index = WarrantIdentityIndex(
        mops_frames_to_identity_records(
            [
                bundled_seed_test[
                    bundled_seed_test["code"].isin(
                        ["030650", "034676", "04472U"]
                    )
                ]
            ]
        )
    )
    assert bundled_seed_index.resolve(
        "030650", "2023-06-21"
    ).target_code == "2330"
    assert bundled_seed_index.resolve(
        "030650", "2025-08-06"
    ).target_code == "4958"
    assert bundled_seed_index.resolve(
        "034676", "2023-06-21"
    ).target_code == "8478"
    assert bundled_seed_index.resolve(
        "04472U", "2023-06-21"
    ).target_code == "4763"
    assert bundled_seed_index.resolve(
        "04472U", "2025-08-06"
    ).target_code == "3037"
    original_seed_save_identity = globals()["save_warrant_identity_index"]
    original_seed_network_fetch = globals()["fetch_mops_warrant_identity"]
    globals()["save_warrant_identity_index"] = (
        lambda records: WarrantIdentityIndex(records)
    )
    globals()["fetch_mops_warrant_identity"] = (
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("停用網路更新時不得呼叫 MOPS")
        )
    )
    try:
        seed_only_repaired = backfill_identity_from_official_mops(
            WarrantIdentityIndex(
                [
                    WarrantIdentity(
                        "030650",
                        "臻鼎元大52購03",
                        "認購",
                        "4958",
                        "臻鼎-KY",
                        "2025-08-05",
                        "2026-02-04",
                        "Matched",
                        "較新的重用代號",
                    )
                ]
            ),
            {("030650", "2023-06-21")},
        )
    finally:
        globals()["save_warrant_identity_index"] = original_seed_save_identity
        globals()["fetch_mops_warrant_identity"] = original_seed_network_fetch
    assert seed_only_repaired.resolve(
        "030650", "2023-06-21"
    ).target_code == "2330"

    assert roc_date_to_iso("112/11/30") == "2023-11-30"
    assert iso_month_to_roc("2023-06") == "11206"
    assert normalize_mops_warrant_code("04472UAe") == "04472U"
    assert normalize_mops_warrant_code("044720F") == "044720"
    assert month_range("2023-11-15", "2024-02-01") == [
        "2023-11",
        "2023-12",
        "2024-01",
        "2024-02",
    ]
    mops_html_test = """
    <html><body><table class='hasBorder'>
      <tr>
        <th>權證代號</th><th>權證簡稱</th><th>權證類型</th>
        <th>上市日期</th><th>最後交易日</th><th>履約截止日</th>
        <th>標的<br>代號</th><th>標的<br>名稱</th>
      </tr>
      <tr class='cColor'>
        <td>030650e</td><td>台積電凱基2A購11</td><td>認購</td>
        <td>111/11/30</td><td>112/09/27</td><td>112/10/02</td>
        <td>2330</td><td>台積電</td>
      </tr>
      <tr class='lColor'>
        <td>07925P</td><td>仁寶元富83售01</td><td>認售</td>
        <td>107/06/11</td><td>108/03/07</td><td>108/03/11</td>
        <td>2324</td><td>仁寶</td>
      </tr>
      <tr class='cColor'>
        <td>034676e</td><td>東哥凱基28購03</td><td>認購</td>
        <td>112/02/01</td><td>112/08/29</td><td>112/08/31</td>
        <td>8478</td><td>東哥遊艇</td>
      </tr>
      <tr class='lColor'>
        <td>030961e</td><td>台積電凱基33購01</td><td>認購</td>
        <td>111/12/05</td><td>113/02/29</td><td>113/03/04</td>
        <td>2330</td><td>台積電</td>
      </tr>
      <tr class='lColor'>
        <td>035628e</td><td>臺股指凱基2A購20</td><td>認購</td>
        <td>112/02/09</td><td>112/10/05</td><td>112/10/11</td>
        <td>IX0001</td><td>臺股指數</td>
      </tr>
    </table></body></html>
    """
    parsed_mops_test = parse_mops_warrant_html(mops_html_test, "1")
    assert list(parsed_mops_test["code"]) == [
        "030650",
        "07925P",
        "034676",
        "030961",
        "035628",
    ]
    assert parsed_mops_test.iloc[0]["list_date"] == "2022-11-30"
    assert parsed_mops_test.iloc[0]["end_date"] == "2023-09-27"
    mops_identity_test = WarrantIdentityIndex(
        mops_frames_to_identity_records([parsed_mops_test])
    ).resolve("030650", "2023-06-21")
    assert mops_identity_test.status == "Matched"
    assert mops_identity_test.target_code == "2330"
    assert "MOPS上市官方歷史" in mops_identity_test.reason
    assert WarrantIdentityIndex(
        mops_frames_to_identity_records([parsed_mops_test])
    ).resolve("034676", "2023-06-21").target_code == "8478"
    assert WarrantIdentityIndex(
        mops_frames_to_identity_records([parsed_mops_test])
    ).resolve("030961", "2023-06-21").target_code == "2330"
    assert WarrantIdentityIndex(
        mops_frames_to_identity_records([parsed_mops_test])
    ).resolve("035628", "2023-06-21").target_code == "IX0001"
    same_target_name_variants = WarrantIdentityIndex(
        [
            WarrantIdentity(
                "TEST01",
                "智通",
                "認購",
                "8932",
                "智通",
                "2023-01-01",
                "2023-12-31",
                "Matched",
                "測試",
            ),
            WarrantIdentity(
                "TEST01",
                "智通*",
                "認購",
                "8932",
                "智通*",
                "2023-01-01",
                "2023-12-31",
                "Matched",
                "測試",
            ),
        ]
    ).resolve("TEST01", "2023-06-21")
    assert same_target_name_variants.status == "Matched"
    assert same_target_name_variants.target_code == "8932"

    test_ranges = warrant_summary_chunk_ranges(
        "2022-06-01",
        "2024-02-01",
    )
    assert test_ranges == [
        ("2022-06-01", "2023-01-01"),
        ("2023-01-01", "2024-01-01"),
        ("2024-01-01", "2024-02-01"),
    ]
    summary_calls: list[tuple[dict[str, Any], str]] = []
    original_fetch_dataset = globals()["fetch_dataset"]

    def fake_summary_fetch(
        dataset: str,
        *,
        params: Optional[dict[str, Any]] = None,
        cache: bool = False,
        force: bool = False,
        cache_key: str = "",
    ) -> pd.DataFrame:
        assert dataset == "TaiwanStockInfoWithWarrantSummary"
        assert params and params.get("start_date") and params.get("end_date")
        assert cache and cache_key.startswith(WARRANT_SUMMARY_CACHE_PREFIX)
        summary_calls.append((dict(params), cache_key))
        chunk_start = str(params["start_date"])
        chunk_end = str(params["end_date"])
        rows = [
            {
                "stock_id": f"W{chunk_start[:4]}",
                "target_stock_id": "2330",
                "type": "認購",
                "date": chunk_start,
                "end_date": chunk_end,
            }
        ]
        if chunk_end == "2023-01-01" or chunk_start == "2023-01-01":
            rows.append(
                {
                    "stock_id": "WDUP",
                    "target_stock_id": "2330",
                    "type": "認購",
                    "date": "2023-01-01",
                    "end_date": "2023-12-31",
                }
            )
        return pd.DataFrame(rows)

    globals()["fetch_dataset"] = fake_summary_fetch
    try:
        summary_test = fetch_warrant_summary_history(
            "2022-06-01",
            "2024-02-01",
        )
    finally:
        globals()["fetch_dataset"] = original_fetch_dataset
    assert len(summary_calls) == 3
    assert len({cache_key for _, cache_key in summary_calls}) == 3
    assert len(summary_test[summary_test["stock_id"] == "WDUP"]) == 1

    def fake_ignored_range(
        dataset: str,
        *,
        params: Optional[dict[str, Any]] = None,
        cache: bool = False,
        force: bool = False,
        cache_key: str = "",
    ) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {
                    "stock_id": "WBAD",
                    "target_stock_id": "2330",
                    "type": "認購",
                    "date": "2025-08-01",
                    "end_date": "2026-02-01",
                }
            ]
        )

    globals()["fetch_dataset"] = fake_ignored_range
    try:
        try:
            fetch_warrant_summary_history("2022-01-01", "2022-12-31")
        except RuntimeError as exc:
            assert "未遵守請求的年度範圍" in str(exc)
        else:
            raise AssertionError("Summary 忽略日期範圍時必須停止")
    finally:
        globals()["fetch_dataset"] = original_fetch_dataset

    original_request_with_retries = globals()["request_with_retries"]
    original_target_cache_paths = globals()["warrant_summary_target_cache_paths"]
    original_atomic_write_parquet = globals()["atomic_write_parquet"]
    target_query_calls: list[dict[str, Any]] = []

    class FakeSummaryResponse:
        status_code = 200

        def __init__(self, rows: list[dict[str, Any]]):
            self.rows = rows

        def json(self) -> dict[str, Any]:
            return {"status": 200, "data": self.rows}

    def fake_target_cache_paths(target_code: str) -> tuple[Path, Path]:
        return (
            Path(f"__self_test_{target_code}.parquet"),
            Path(f"__self_test_{target_code}.empty.json"),
        )

    def fake_target_request(
        url: str,
        *,
        params: dict[str, Any],
        description: str,
        attempts: Optional[int] = None,
    ) -> FakeSummaryResponse:
        target_query_calls.append(dict(params))
        return FakeSummaryResponse(
            [
                {
                    "stock_id": "WREUSE",
                    "target_stock_id": "2330",
                    "type": "認購",
                    "date": "2023-01-01",
                    "end_date": "2023-12-31",
                }
            ]
        )

    globals()["request_with_retries"] = fake_target_request
    globals()["warrant_summary_target_cache_paths"] = fake_target_cache_paths
    globals()["atomic_write_parquet"] = lambda frame, path: None
    try:
        target_shard_test = fetch_warrant_summary_target_history(
            "2330",
            "2023-01-01",
            "2023-12-31",
        )
    finally:
        globals()["request_with_retries"] = original_request_with_retries
        globals()["warrant_summary_target_cache_paths"] = (
            original_target_cache_paths
        )
        globals()["atomic_write_parquet"] = original_atomic_write_parquet
    assert len(target_shard_test) == 1
    assert target_query_calls[-1]["data_id"] == "2330"

    original_save_identity = globals()["save_warrant_identity_index"]
    original_target_fetch = globals()["fetch_warrant_summary_target_history"]
    original_sleep = time.sleep
    quota_pause_calls: list[float] = []
    quota_raised = False
    base_reused_index = WarrantIdentityIndex(
        [
            WarrantIdentity(
                "WREUSE",
                "臻鼎測試購01",
                "認購",
                "4958",
                "臻鼎-KY",
                "2025-08-01",
                "2026-02-01",
                "Matched",
                "目前區間",
            )
        ]
    )

    def fake_support_fetch(
        dataset: str,
        *,
        params: Optional[dict[str, Any]] = None,
        cache: bool = False,
        force: bool = False,
        cache_key: str = "",
    ) -> pd.DataFrame:
        if dataset == "TaiwanStockInfoWithWarrant":
            return pd.DataFrame(
                [
                    {
                        "stock_id": "WREUSE",
                        # 只保留 2025 新一輪商品名稱，模擬 FinMind 總覽
                        # 已遺失 2023 舊名稱的真實狀況。
                        "stock_name": "臻鼎測試購01",
                    }
                ]
            )
        if dataset == "TaiwanStockInfo":
            return pd.DataFrame(
                [
                    {
                        "stock_id": "2330",
                        "stock_name": "台積電",
                        "date": "2021-01-01",
                    },
                    {
                        "stock_id": "4958",
                        "stock_name": "臻鼎-KY",
                        "date": "2021-01-01",
                    },
                ]
            )
        raise AssertionError(dataset)

    def fake_target_history(
        target_code: str,
        start_date: str = WARRANT_SUMMARY_HISTORY_START,
        end_date: str = "",
    ) -> pd.DataFrame:
        nonlocal quota_raised
        if target_code == "4958" and not quota_raised:
            quota_raised = True
            raise FinMindQuotaExceeded("測試額度用完")
        if target_code == "2330":
            return pd.DataFrame(
                [
                    {
                        "stock_id": "WREUSE",
                        "target_stock_id": "2330",
                        "type": "認購",
                        "date": "2023-01-01",
                        "end_date": "2023-12-31",
                    }
                ]
            )
        return pd.DataFrame(
            columns=[
                "stock_id",
                "target_stock_id",
                "type",
                "date",
                "end_date",
            ]
        )

    globals()["fetch_dataset"] = fake_support_fetch
    globals()["fetch_warrant_summary_target_history"] = fake_target_history
    globals()["save_warrant_identity_index"] = (
        lambda records: WarrantIdentityIndex(records)
    )
    time.sleep = lambda seconds: quota_pause_calls.append(seconds)
    try:
        repaired_reused_index = backfill_identity_from_target_shards(
            base_reused_index,
            {("WREUSE", "2023-06-21")},
        )
    finally:
        globals()["fetch_dataset"] = original_fetch_dataset
        globals()["fetch_warrant_summary_target_history"] = original_target_fetch
        globals()["save_warrant_identity_index"] = original_save_identity
        time.sleep = original_sleep
    repaired_identity = repaired_reused_index.resolve(
        "WREUSE",
        "2023-06-21",
    )
    assert repaired_identity.status == "Matched"
    assert repaired_identity.target_code == "2330"
    assert quota_pause_calls and quota_pause_calls[0] >= 60
    explicit_target_records = build_identity_records_from_summary(
        pd.DataFrame(
            [
                {
                    "stock_id": "WREUSE",
                    "target_stock_id": "2330",
                    "type": "認購",
                    "date": "2023-01-01",
                    "end_date": "2023-12-31",
                }
            ]
        ),
        {"WREUSE": ["臻鼎測試購01"]},
        stock_alias_resolver(
            {
                "2330": "台積電",
                "4958": "臻鼎-KY",
            }
        ),
        {
            "2330": "台積電",
            "4958": "臻鼎-KY",
        },
    )
    assert explicit_target_records[0].target_code == "2330"
    assert explicit_target_records[0].name == "WREUSE"

    resolver_test = stock_alias_resolver(
        {
            "1111": "台積",
            "2330": "台積電",
            "0050": "元大台灣50",
        }
    )
    resolved_test = resolve_underlying_from_name("台積電測試購01", resolver_test)
    assert resolved_test is not None and resolved_test[0] == "2330"
    for sequence in range(5_000):
        resolved = resolve_underlying_from_name(
            f"台積電效能測試購{sequence}", resolver_test
        )
        assert resolved is not None and resolved[0] == "2330"
    memo_size_after_first_pass = len(resolver_test.memo)
    resolve_underlying_from_name("台積電測試購01", resolver_test)
    assert len(resolver_test.memo) == memo_size_after_first_pass

    raw_test = pd.DataFrame(
        [
            {
                "securities_trader": "測試分點",
                "price": 10.0,
                "buy": 100,
                "sell": 20,
                "securities_trader_id": "T001",
                "stock_id": "W1",
                "date": "2026-01-02",
            },
            {
                "securities_trader": "測試分點",
                "price": 11.0,
                "buy": 50,
                "sell": 0,
                "securities_trader_id": "T001",
                "stock_id": "W1",
                "date": "2026-01-02",
            },
            {
                "securities_trader": "測試分點",
                "price": "壞值",
                "buy": 999,
                "sell": 0,
                "securities_trader_id": "T001",
                "stock_id": "W1",
                "date": "2026-01-02",
            },
        ]
    )
    identity_test = WarrantIdentityIndex(
        [
            WarrantIdentity(
                "W1",
                "測試購01",
                "認購",
                "2330",
                "台積電",
                "2025-01-01",
                "2026-12-31",
                "Matched",
                "測試身分",
            ),
            WarrantIdentity(
                "WPUT",
                "測試售01",
                "認售",
                "2330",
                "台積電",
                "2025-01-01",
                "2026-12-31",
                "Matched",
                "測試認售身分",
            ),
        ]
    )
    normalized_test, _, invalid_numeric_rows = normalize_warrant_day(
        raw_test,
        "2026-01-02",
        {"測試分點": ("測試分點", "T001")},
    )
    assert len(normalized_test) == 1
    assert normalized_test.iloc[0]["日期"] == "2026-01-02"
    assert int(normalized_test.iloc[0]["買進股數"]) == 150
    assert invalid_numeric_rows == 1
    assert list(normalized_test.columns) == COMPACT_FACT_COLUMNS
    legacy_compact = normalized_test.copy()
    legacy_compact["權證名稱"] = "舊錯誤名稱"
    legacy_compact["標的股"] = "9999"
    legacy_compact["標的名稱"] = "舊錯誤標的"
    legacy_compact["身分配對狀態"] = "Unmatched"
    legacy_compact["身分配對錯誤原因"] = "舊錯誤結果"
    enriched_test = enrich_compact_identity(legacy_compact, identity_test)
    assert enriched_test.iloc[0]["標的股"] == "2330"
    assert enriched_test.iloc[0]["權證類型"] == "認購"
    put_fact = normalized_test.copy()
    put_fact["權證代號"] = "WPUT"
    put_fact["買進金額"] = 2_000_000
    enriched_put = enrich_compact_identity(put_fact, identity_test)
    assert enriched_put.iloc[0]["身分配對狀態"] == "Matched"
    assert enriched_put.iloc[0]["權證類型"] == "認售"
    assert build_events_for_day(enriched_put, "2026-01-02") == []

    high = calculate_ma_values(make_test_ma_row(), "2026-01-30")
    assert high["MA位置代碼"] == "111"
    assert high["均線多頭排列"] is True
    assert high["剛站上MA20"] is True

    low = calculate_ma_values(
        make_test_ma_row(
            close=80,
            ma5=90,
            ma10=100,
            ma20=110,
            prev_close=120,
            prev_ma5=115,
            prev_ma10=110,
            prev_ma20=105,
            ma20_5d_ago=109,
        ),
        "2026-01-30",
    )
    assert low["MA位置代碼"] == "000"
    assert low["均線空頭排列"] is True
    assert low["剛跌破MA10"] is True

    missing = calculate_ma_values(pd.DataFrame(), "2026-01-30")
    assert missing["均線資料狀態"] == "Missing"

    short_dates = [
        value.strftime("%Y-%m-%d")
        for value in pd.bdate_range("2026-01-01", periods=19)
    ]
    short_prices = pd.DataFrame(
        {
            "date": short_dates,
            "stock_id": "2330",
            "close": [100 + value for value in range(len(short_dates))],
        }
    )
    short_ma = prepare_ma_table(short_prices)
    short_result = calculate_ma_values(
        short_ma,
        short_dates[-1],
        expected_trading_dates=short_dates,
    )
    assert short_result["均線資料狀態"] == "Missing"
    assert "不足25個市場交易日" in short_result["均線錯誤原因"]

    gap_dates = [
        value.strftime("%Y-%m-%d")
        for value in pd.bdate_range("2026-01-01", periods=30)
    ]
    missing_price_day = gap_dates[-10]
    gap_prices = pd.DataFrame(
        {
            "date": [
                value for value in gap_dates if value != missing_price_day
            ],
            "stock_id": "2330",
            "close": [
                100 + sequence
                for sequence, value in enumerate(gap_dates)
                if value != missing_price_day
            ],
        }
    )
    gap_ma = prepare_ma_table(gap_prices)
    gap_result = calculate_ma_values(
        gap_ma,
        gap_dates[-1],
        expected_trading_dates=gap_dates,
    )
    assert gap_result["均線資料狀態"] == "Missing"
    assert missing_price_day in gap_result["均線錯誤原因"]
    gap_calendar = build_trading_calendar(gap_dates)
    gap_rows_by_date = gap_ma.set_index("date", drop=False).to_dict(orient="index")
    optimized_gap_result = calculate_ma_values(
        gap_ma,
        gap_dates[-1],
        trading_calendar=gap_calendar,
        available_dates=set(gap_rows_by_date),
        ma_rows_by_date=gap_rows_by_date,
    )
    assert optimized_gap_result == gap_result

    compact = pd.DataFrame(
        [
            {
                "權證代號": "W1",
                "權證名稱": "測試購01",
                "權證類型": "認購",
                "標的股": "2330",
                "標的名稱": "台積電",
                "分點": "測試分點",
                "分點名稱": "測試分點",
                "券商代號": "T001",
                "日期": "2026-01-02",
                "買進股數": 100,
                "賣出股數": 0,
                "買進金額": 1_100_000,
                "賣出金額": 0,
                "買超股數": 100,
                "買超金額": 1_100_000,
                "身分配對狀態": "Matched",
                "身分配對錯誤原因": "",
            },
            {
                "權證代號": "W2",
                "權證名稱": "測試購02",
                "權證類型": "認購",
                "標的股": "2330",
                "標的名稱": "台積電",
                "分點": "測試分點",
                "分點名稱": "測試分點",
                "券商代號": "T001",
                "日期": "2026-01-02",
                "買進股數": 50,
                "賣出股數": 20,
                "買進金額": 500_000,
                "賣出金額": 200_000,
                "買超股數": 30,
                "買超金額": 300_000,
                "身分配對狀態": "Matched",
                "身分配對錯誤原因": "",
            },
        ],
        columns=COMPACT_COLUMNS,
    )
    built = build_events_for_day(compact, "2026-01-02")
    assert len(built) == 1, "同券商＋標的＋日必須去重成一事件"
    assert built[0]["事件代碼"] == "B"
    assert built[0]["買超股數"] == 130
    built_again = build_events_for_day(compact, "2026-01-02")
    assert [item["事件唯一ID"] for item in built] == [
        item["事件唯一ID"] for item in built_again
    ], "相同精簡資料重算不得增加事件"

    display_name_variants = compact.copy()
    display_name_variants.loc[1, "分點名稱"] = "測試分點（舊名）"
    display_name_variants.loc[1, "券商代號"] = "t001"
    display_name_variants.loc[1, "標的名稱"] = "台積電*"
    display_name_variants.loc[1, "買進金額"] = 1_000_000
    display_name_variants.loc[1, "買超金額"] = 800_000
    merged_variant_events = build_events_for_day(
        display_name_variants,
        "2026-01-02",
    )
    assert len(merged_variant_events) == 1, (
        "顯示名稱或代號大小寫不同不得拆成重複事件"
    )
    assert merged_variant_events[0]["事件唯一ID"] == "T001|2330|2026-01-02"
    assert merged_variant_events[0]["單日累積買進金額"] == 2_100_000
    assert merged_variant_events[0]["事件代碼"] == "B"
    assert merged_variant_events[0]["涵蓋權證數"] == 2
    assert len(merged_variant_events[0]["lots"]) == 2
    assert merged_variant_events[0]["分點名稱"] == "測試分點"
    assert merged_variant_events[0]["標的名稱"] == "台積電"

    sales = {
        ("T001", "W1"): [
            {"日期": "2026-01-03", "賣出股數": 100, "賣出金額": 1_210_000}
        ],
        ("T001", "W2"): [
            {"日期": "2026-01-03", "賣出股數": 50, "賣出金額": 550_000}
        ],
    }
    closed_event = simulate_group_outcomes_fifo(built, sales)[0]
    assert closed_event["目前狀態"] == "已出清"
    assert closed_event["出清報酬%"] == 10.0
    closed_event = evaluate_open_positions_after_two_months(
        [closed_event],
        WarrantContractIndex([]),
        "2026-04-30",
        price_loader=lambda contract, day: (0.0, day),
    )[0]
    assert closed_event["勝敗認列狀態"] == "實際FIFO出清"
    assert closed_event["勝敗結果"] == "勝"

    open_event = {
        **built[0],
        "事件唯一ID": "T001|2317|2026-01-04",
        "標的股": "2317",
        "事件日": "2026-01-04",
        "lots": [
            {
                "買進日": "2026-01-04",
                "權證代號": "W3",
                "權證名稱": "測試購03",
                "金額": 1_000_000,
                "股數": 100,
            }
        ],
    }
    open_result = simulate_group_outcomes_fifo([open_event], {})[0]
    assert open_result["目前狀態"] == "未出清"
    recent_result = evaluate_open_positions_after_two_months(
        [open_result],
        WarrantContractIndex([]),
        "2026-02-28",
        price_loader=lambda contract, day: (0.0, day),
    )[0]
    assert recent_result["勝敗認列狀態"] == "待認列"

    active_contract = WarrantContract(
        "W3",
        "認購",
        "2317",
        "2025-12-01",
        "2026-12-01",
        "2026-12-05",
        "4",
        0.1,
        100.0,
        None,
        False,
        "1",
    )
    aged_result = evaluate_open_positions_after_two_months(
        [open_result],
        WarrantContractIndex([active_contract]),
        "2026-04-30",
        price_loader=lambda contract, day: (12_000.0, "2026-04-29"),
    )[0]
    assert aged_result["勝敗認列狀態"] == "兩月市價評價"
    assert aged_result["認列報酬%"] == 20.0
    assert aged_result["勝敗結果"] == "勝"

    expired_contract = WarrantContract(
        "W3",
        "認購",
        "2317",
        "2025-12-01",
        "2026-02-27",
        "2026-03-05",
        "4",
        0.1,
        100.0,
        120.0,
        True,
        "1",
    )
    expired_result = evaluate_open_positions_after_two_months(
        [open_result],
        WarrantContractIndex([expired_contract]),
        "2026-04-30",
        price_loader=lambda contract, day: (_ for _ in ()).throw(
            AssertionError("到期權證不得查市價")
        ),
    )[0]
    expected_settlement = (120.0 - 100.0) * 0.1 * (1 - 0.001)
    assert expired_result["勝敗認列狀態"] == "到期結算"
    assert expired_result["剩餘部位評價金額"] == round(
        expected_settlement * 100, 2
    )

    no_payment_contract = WarrantContract(
        "W3",
        "認購",
        "2317",
        "2025-12-01",
        "2026-02-27",
        "2026-03-05",
        "5",
        0.1,
        100.0,
        None,
        False,
        "1",
    )
    no_payment_result = evaluate_open_positions_after_two_months(
        [open_result],
        WarrantContractIndex([no_payment_contract]),
        "2026-04-30",
        price_loader=lambda contract, day: (999.0, day),
    )[0]
    assert no_payment_result["勝敗認列狀態"] == "到期無給付"
    assert no_payment_result["認列報酬%"] == -100.0
    try:
        evaluate_open_positions_after_two_months(
            [open_result],
            WarrantContractIndex([]),
            "2026-04-30",
            price_loader=lambda contract, day: (1.0, day),
        )
    except RuntimeError as exc:
        assert "無法可靠認列" in str(exc)
    else:
        raise AssertionError("滿兩月事件缺少契約時不得悄悄排除於勝率")

    closed_event.update(high)
    event_frame = events_to_frame([closed_event])
    condition_frame = build_condition_statistics(event_frame)
    position_frame = build_position_statistics(event_frame)
    profile_frame = build_broker_profiles(
        event_frame, condition_frame, position_frame
    )
    validate_final_results(event_frame, condition_frame, position_frame)
    broken_condition_frame = condition_frame.copy()
    broken_condition_frame.loc[0, "事件總數"] += 1
    broken_condition_frame.loc[0, "未出清筆數"] += 1
    try:
        validate_final_results(
            event_frame,
            broken_condition_frame,
            position_frame,
        )
    except AssertionError:
        pass
    else:
        raise AssertionError("統計交叉檢核未抓到無法追溯事件明細的數值")
    assert len(condition_frame) == len(CONDITION_SPECS) * 6
    assert len(position_frame) == len(MA_POSITION_NAMES) * 6
    assert len(profile_frame) == 1

    statistic_events = pd.DataFrame(
        [
            {
                "分點": "測試分點",
                "目前狀態": "已出清",
                "出清報酬%": 10,
                "持有天數": 1,
                "勝敗認列狀態": "實際FIFO出清",
                "認列報酬%": 10,
                "認列持有天數": 1,
                "總成本": 100,
                "單日累積買進金額": 100,
                "事件日": "2026-01-01",
            },
            {
                "分點": "測試分點",
                "目前狀態": "已出清",
                "出清報酬%": -5,
                "持有天數": 3,
                "勝敗認列狀態": "實際FIFO出清",
                "認列報酬%": -5,
                "認列持有天數": 3,
                "總成本": 100,
                "單日累積買進金額": 100,
                "事件日": "2026-01-02",
            },
            {
                "分點": "測試分點",
                "目前狀態": "已出清",
                "出清報酬%": 0,
                "持有天數": 100,
                "勝敗認列狀態": "實際FIFO出清",
                "認列報酬%": 0,
                "認列持有天數": 100,
                "總成本": 100,
                "單日累積買進金額": 100,
                "事件日": "2026-01-03",
            },
        ]
    )
    summary = summarize_event_group(
        statistic_events, broker="測試分點", event_code="ALL"
    )
    assert summary["勝率"] == round(1 / 3 * 100, 2)
    assert summary["平均持有天數"] == round(104 / 3, 2)
    assert summary["中位數持有天數"] == 3.0
    assert summary["勝筆數"] == 1 and summary["敗筆數"] == 1
    assert summary["平手筆數"] == 1
    return {
        "隨附MOPS官方種子完整性": "PASS",
        "隨附種子處理權證代號重用": "PASS",
        "MOPS第七碼重用標記正規化": "PASS",
        "同母股名稱差異不誤判歧義": "PASS",
        "GitHub不連MOPS仍可補配": "PASS",
        "MOPS民國年月與月份切片": "PASS",
        "MOPS官方HTML欄位解析": "PASS",
        "MOPS下市代號尾碼正規化": "PASS",
        "MOPS日期化身分命中2023交易日": "PASS",
        "Summary明確日期年度切片": "PASS",
        "Summary年度合併去重": "PASS",
        "Summary忽略日期時拒絕快取": "PASS",
        "Summary母股data_id分片查詢": "PASS",
        "目前名稱失效時全母股日期補配": "PASS",
        "母股補抓額度耗盡等待續跑": "PASS",
        "日期化母股不被目前名稱覆蓋": "PASS",
        "標的前綴字典與名稱memo": "PASS",
        "FinMind日檔精簡與_date欄位": "PASS",
        "均線位置與排列": "PASS",
        "剛站上MA20": "PASS",
        "剛跌破MA10": "PASS",
        "不足歷史Missing": "PASS",
        "真實19交易日均線Missing": "PASS",
        "股價缺口均線Missing": "PASS",
        "壞數值列隔離": "PASS",
        "compact事實與階段二身分重算": "PASS",
        "認售有身分但不建立事件": "PASS",
        "事件去重複": "PASS",
        "事件顯示名稱變體穩定聚合": "PASS",
        "ABCDE分類": "PASS",
        "事件買超股數語意": "PASS",
        "FIFO已出清": "PASS",
        "FIFO未出清": "PASS",
        "兩個曆月前不提前認列": "PASS",
        "滿兩月未結部位市價認列": "PASS",
        "到期官方結算金額認列": "PASS",
        "到期無自動給付按零元認列": "PASS",
        "滿兩月無法評價時禁止產表": "PASS",
        "勝率": "PASS",
        "平均與中位持有天數": "PASS",
        "統計可追溯交叉檢核": "PASS",
        "完整統計管線": "PASS",
    }


def rebuild_statistics(
    state: dict[str, Any], expected_dates: list[str]
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    print("【階段二】只讀 compact_daily，重新配對身分並重建事件、FIFO、均線與統計...")
    set_run_stage(state, "identity_history")
    print("  ⏳ 載入或建立日期化權證身分索引...", flush=True)
    identity_index = build_warrant_identity_index()
    compact_frames, events = load_compact_frames(state, identity_index)
    set_run_stage(state, "statistics")
    sales = build_sale_rows(compact_frames)
    events = simulate_group_outcomes_fifo(events, sales)
    last_day = max(state["successful_dates"], default="")
    print(
        f"  ⏳ 評價截至 {last_day} 已滿 "
        f"{OPEN_POSITION_EVALUATION_MONTHS} 個曆月的未結部位...",
        flush=True,
    )
    contract_index = build_warrant_contract_index()
    events = evaluate_open_positions_after_two_months(
        events,
        contract_index,
        last_day,
    )
    atomic_write_parquet(
        checkpoint_frame(events, last_day, "fifo-and-two-month-valuation-complete"),
        RESULTS_DIR / "fifo_checkpoint.parquet",
    )
    dates = trading_dates()
    events = attach_ma_to_events(
        events, dates, statistics_day=max(state["successful_dates"], default="")
    )
    events_frame = events_to_frame(events)
    condition_stats = build_condition_statistics(events_frame)
    position_stats = build_position_statistics(events_frame)
    profiles = build_broker_profiles(
        events_frame, condition_stats, position_stats
    )
    validate_final_results(events_frame, condition_stats, position_stats)
    integrity = build_integrity_report(state, expected_dates, events_frame)
    return events_frame, condition_stats, position_stats, profiles, integrity


def main() -> int:
    ensure_directories()
    if "--self-test" in sys.argv:
        results = run_self_tests()
        for name, result in results.items():
            print(f"{name}: {result}")
        return 0
    print(
        "FinMind 完整歷史權證分點均線勝率統計\n"
        f"程式版本：{PROGRAM_VERSION}｜schema：{SCHEMA_VERSION}\n"
        f"快取目錄：{CACHE_ROOT}",
        flush=True,
    )
    state = load_state()
    set_run_stage(state, "initialization")
    initialization_started = time.perf_counter()
    print("【初始化 1/3】檢查 state 與成功日期檔案...", flush=True)
    audit_successful_files(state)
    print(
        f"  ✅ state 檢查完成｜成功 {len(state['successful_dates']):,} 日｜"
        f"失敗 {len(state['failed_dates']):,} 日｜"
        f"{time.perf_counter() - initialization_started:.1f} 秒",
        flush=True,
    )
    print("【初始化 2/3】探索 FinMind 預期交易日期...", flush=True)
    expected_dates = discover_expected_dates(state)
    print(
        f"  ✅ 預期日期 {len(expected_dates):,} 日｜"
        f"{min(expected_dates, default='-')}～{max(expected_dates, default='-')}",
        flush=True,
    )
    if not FORCE_RECALCULATE_STATS:
        print("【初始化 3/3】建立 26 分點券商代號對照...", flush=True)
        broker_map = build_broker_map()
        if len(broker_map) != len(TARGET_PATTERNS):
            raise RuntimeError(
                f"目標分點對照不完整：{len(broker_map)}/{len(TARGET_PATTERNS)}"
            )
        print(f"  ✅ 目標分點對照完成：{len(broker_map):,} 個", flush=True)
        if state["successful_dates"]:
            print("【下載前檢查】抽查既有 compact 的身分配對品質...", flush=True)
            set_run_stage(state, "identity_history")
            preflight_identity_quality(
                state,
                build_warrant_identity_index(),
            )
        print(
            f"  ✅ 初始化全部完成｜{time.perf_counter() - initialization_started:.1f} 秒",
            flush=True,
        )
        set_run_stage(state, "downloading")
        download_and_compact_history(state, expected_dates, broker_map)
    elif not state["successful_dates"]:
        raise RuntimeError(
            "FORCE_RECALCULATE_STATS=1，但本機沒有任何成功 compact_daily"
        )
    (
        events,
        condition_stats,
        position_stats,
        profiles,
        integrity,
    ) = rebuild_statistics(state, expected_dates)
    output_path = write_outputs(
        events, condition_stats, position_stats, profiles, integrity
    )
    state["last_statistics_time"] = datetime.now().isoformat(timespec="seconds")
    set_run_stage(state, "complete")
    print(
        "\n✅ 完成\n"
        f"  事件：{len(events):,}\n"
        f"  均線條件統計：{len(condition_stats):,}\n"
        f"  MA位置統計：{len(position_stats):,}\n"
        f"  Excel：{output_path}"
    )
    if state["failed_dates"]:
        print(
            "  ⚠️ 歷史資料尚未完整；失敗日期："
            + "、".join(state["failed_dates"])
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
