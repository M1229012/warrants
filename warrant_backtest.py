"""
認購權證特定分點買超回測系統
=================================
新制金額強度分類：同一分點 + 同一標的 + 同一天為 1 筆事件。
進入條件：該事件內至少 1 檔權證單日買進金額 >= 100萬。
分類依據：同一分點 + 同一標的 + 同一天的所有權證買進金額合計。
A：單日累積買進金額 100–159萬
B：單日累積買進金額 160–249萬
C：單日累積買進金額 250–499萬
D：單日累積買進金額 500–999萬
E：單日累積買進金額 >= 1000萬

事件規則：同一分點 + 同一標的 + 同一天只會產生 A / B / C / D / E 其中一類；不同分點交易同一標的互不排除。每次符合條件的事件各自成一筆，後續賣出依剩餘股數 FIFO 扣減；金額僅用於成本與損益。

輸出 Excel：
1. A_基礎買超
2. B_明顯買超
3. C_強勢買超
4. D_大額布局
5. E_超大額布局
6. 勝率統計
7. 近兩月買賣金額排行
8. 近兩月分點數排行
9. 券商查詢
10. 快取_TOP15共識淨買超
11. 快取_TOP15部位明細
12. 快取_近7日權證分點共識TOP15（近7／14／21日精選13分點排名；僅 RUN_MODE=2 更新）
13. 快取_近10日分點買賣明細（僅 RUN_MODE=2 全市場分點模式更新）
14. 顏色說明

執行：python warrant_backtest.py
依賴：pip install requests pandas openpyxl
"""

import json, re, time, os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from io import StringIO

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ══════════════════════════════════════════════════════════════════════
# 設定
# ══════════════════════════════════════════════════════════════════════

DEFAULT_OUTPUT_DIR = "output" if os.getenv("GITHUB_ACTIONS", "").strip().lower() == "true" else r"C:\Users\chen1_ukw0m7r\Downloads"
OUTPUT_DIR = os.getenv("OUTPUT_DIR", DEFAULT_OUTPUT_DIR)
AMOUNT_THRESH = 1_000_000

# 新制金額強度分類：
# 事件單位為「同一分點 + 同一標的 + 同一天」。
# 進入條件為事件內至少一檔權證單日買進金額 >= AMOUNT_THRESH；
# 分類依據為同標的單日累積買進金額。
AMOUNT_CLASS_SPECS = [
    ("A", "基礎買超", 1_000_000, 1_600_000),
    ("B", "明顯買超", 1_600_000, 2_500_000),
    ("C", "強勢買超", 2_500_000, 5_000_000),
    ("D", "大額布局", 5_000_000, 10_000_000),
    ("E", "超大額布局", 10_000_000, None),
]
AMOUNT_CLASS_CODES = [spec[0] for spec in AMOUNT_CLASS_SPECS]
AMOUNT_CLASS_LABELS = {code: label for code, label, _, _ in AMOUNT_CLASS_SPECS}
AMOUNT_CLASS_SHEET_NAMES = {
    f"{code}_{label}"
    for code, label, _, _ in AMOUNT_CLASS_SPECS
}
MAX_WORKERS   = 50
RECENT_RANKING_DAYS = 62
SELL_DETAIL_DAYS = int(os.getenv("SELL_DETAIL_DAYS", "3"))
D_WINDOW_DAYS = 10

# 快取設定：
# 第一次執行沒有快取時會完整爬取並建立快取；
# 第二次之後會優先讀取快取，只針對最近有出現目標分點的候選組合補抓新資料。
USE_CACHE = os.getenv("USE_CACHE", "1").strip().lower() not in ("0", "false", "no")
FORCE_FULL_CACHE_REFRESH = os.getenv("FORCE_FULL_CACHE_REFRESH", "0").strip().lower() in ("1", "true", "yes")
CACHE_RECENT_SCAN_DAYS = int(os.getenv("CACHE_RECENT_SCAN_DAYS", "50"))
PRICE_WORKERS = int(os.getenv("PRICE_WORKERS", "80"))
PRESCAN_WORKERS = int(os.getenv("PRESCAN_WORKERS", "60"))
FIND_BROKER_WORKERS = int(os.getenv("FIND_BROKER_WORKERS", "40"))

# 加速模式：
# 1. 有候選組合快取時，仍會每天補掃全市場最近 CACHE_RECENT_SCAN_DAYS 天，
#    用來發現新權證 / 新候選組合；舊候選資料則優先使用快取，避免重抓完整歷史。
#    FAST_SKIP_RECENT_PRESCAN 僅保留為相容舊設定，不再作為每日主流程的跳過依據。
# 2. B / C / D 工作表的 D+ 欄位只使用標的股價格，預設不再額外抓群組事件中每一檔權證價格。
#    若未來需要群組事件權證明細價格，可設定 FETCH_GROUP_WARRANT_PRICES=1。
FAST_SKIP_RECENT_PRESCAN = os.getenv("FAST_SKIP_RECENT_PRESCAN", "0").strip().lower() not in ("0", "false", "no")
FETCH_GROUP_WARRANT_PRICES = os.getenv("FETCH_GROUP_WARRANT_PRICES", "0").strip().lower() in ("1", "true", "yes")

CACHE_DIR = os.getenv("CACHE_DIR", os.path.join(OUTPUT_DIR, "warrant_cache"))
CACHE_ENCODING = "utf-8-sig"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

WARRANTS_CACHE_PATH   = os.path.join(CACHE_DIR, "warrants_cache.csv")
BROKER_MAP_CACHE_PATH = os.path.join(CACHE_DIR, "broker_map_cache.csv")
CANDIDATES_CACHE_PATH = os.path.join(CACHE_DIR, "candidates_cache.csv")
CANDIDATES_CACHE_ALL_PATH = CANDIDATES_CACHE_PATH
CANDIDATES_CACHE_SELECTED5_PATH = os.path.join(CACHE_DIR, "candidates_cache_selected5.csv")
HISTORY_CACHE_PATH    = os.path.join(CACHE_DIR, "broker_warrant_history_cache.csv")
PRICE_CACHE_PATH      = os.path.join(CACHE_DIR, "price_cache.csv")
PRICE_PREFETCH_STATE_PATH = os.path.join(CACHE_DIR, "price_prefetch_state.csv")
PRESCAN_STATUS_PATH = os.path.join(CACHE_DIR, "prescan_status.csv")
PRESCAN_REFRESH_KEYS_PATH = os.path.join(CACHE_DIR, "prescan_refresh_keys.csv")

# 自動價格預抓：
# 若當日分點資料尚未出現在 API4 預掃描結果中，正式報表流程會先停止，
# 改成只用既有快取資料把所有可能需要的價格先補進 price_cache。
# 同一天已經預抓過且分點資料仍未出來時，下一次執行會快速結束，不再重複抓價格。
AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING = os.getenv("AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING", "1").strip().lower() not in ("0", "false", "no")
PRICE_PREFETCH_STATE_ENABLED = os.getenv("PRICE_PREFETCH_STATE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
PRICE_PREFETCH_SKIP_IF_DONE_TODAY = os.getenv("PRICE_PREFETCH_SKIP_IF_DONE_TODAY", "1").strip().lower() not in ("0", "false", "no")
PRICE_PREFETCH_FORCE = os.getenv("PRICE_PREFETCH_FORCE", "0").strip().lower() in ("1", "true", "yes")
PRICE_PREFETCH_LOOKBACK_DAYS = int(os.getenv("PRICE_PREFETCH_LOOKBACK_DAYS", "30"))
PRICE_PREFETCH_TARGET_DATE = os.getenv("PRICE_PREFETCH_TARGET_DATE", "").strip()
# 價格預抓防呆：
# 盤後價格可能比分點 API4 更早更新；若今天分點資料尚未出現，但今天價格也尚未寫進快取，
# 即使 price_prefetch_state 已記錄 done，也會再嘗試預抓，避免早盤/盤中先跑過後，盤後價格不再更新。
PRICE_PREFETCH_RETRY_UNTIL_TARGET_PRICE = os.getenv("PRICE_PREFETCH_RETRY_UNTIL_TARGET_PRICE", "1").strip().lower() not in ("0", "false", "no")
PRICE_PREFETCH_MIN_TARGET_PRICE_CODES = int(os.getenv("PRICE_PREFETCH_MIN_TARGET_PRICE_CODES", "1"))


# 長期留單補價 / 修正勝率設定：
# longterm 模式會直接從「快取_分點歷史」用 FIFO 重建未出清部位，
# 篩選持有超過 120 / 150 / 180 天的留單，補最新可用價格，並輸出修正勝率。
LONGTERM_OPEN_DAYS = [
    int(x.strip())
    for x in re.split(r"[,;；、\n\r\t]+", os.getenv("LONGTERM_OPEN_DAYS", "120,150,180"))
    if x.strip().isdigit()
]
if not LONGTERM_OPEN_DAYS:
    LONGTERM_OPEN_DAYS = [120, 150, 180]
LONGTERM_OPEN_DAYS = sorted(set(max(int(x), 1) for x in LONGTERM_OPEN_DAYS))
LONGTERM_PRICE_LOOKBACK_DAYS = int(os.getenv("LONGTERM_PRICE_LOOKBACK_DAYS", "420"))
LONGTERM_PRICE_STALE_DAYS = int(os.getenv("LONGTERM_PRICE_STALE_DAYS", "10"))
LONGTERM_ZERO_PRICE_THRESHOLD = float(os.getenv("LONGTERM_ZERO_PRICE_THRESHOLD", "0.05"))
LONGTERM_TARGET_DATE = os.getenv("LONGTERM_TARGET_DATE", "").strip()
LONGTERM_MAX_DETAIL_ROWS = int(os.getenv("LONGTERM_MAX_DETAIL_ROWS", "0"))
# longterm 的完整明細 / 分點彙總只輸出 Excel，避免占用 Google Sheet 格數。
# Google Sheet 預設只更新既有「勝率統計」工作表底部的長期留單修正勝率區塊。
LONGTERM_UPDATE_WINRATE_SHEET = os.getenv("LONGTERM_UPDATE_WINRATE_SHEET", os.getenv("LONGTERM_UPLOAD_TO_GSHEET", "1")).strip().lower() not in ("0", "false", "no")
LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET = os.getenv("LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET", "0").strip().lower() in ("1", "true", "yes")
LONGTERM_WINRATE_SHEET_DAYS = int(os.getenv("LONGTERM_WINRATE_SHEET_DAYS", str(LONGTERM_OPEN_DAYS[0] if LONGTERM_OPEN_DAYS else 120)) or "120")
LONGTERM_WINRATE_SHEET_HEADER = os.getenv("LONGTERM_WINRATE_SHEET_HEADER", "修正勝率").strip() or "修正勝率"

# 工作流模式：
# daily：每日自動流程。先輕量探測，分點資料沒出來就只補價格並停止；不啟用 MoneyDJ Repair。
# longterm：長期留單補價入口（保留模式，不影響每日流程）。
# repair：完整修補模式。忽略 prescan 狀態、強制重掃並啟用 MoneyDJ Search Repair。
WORKFLOW_MODE = os.getenv("WORKFLOW_MODE", "daily").strip().lower() or "daily"
if WORKFLOW_MODE not in ("daily", "longterm", "repair"):
    print(f"  ⚠️ WORKFLOW_MODE={WORKFLOW_MODE} 不支援，改用 daily。")
    WORKFLOW_MODE = "daily"

# daily 輕量探測：先用近期有活動的權證少量查 API4，避免分點資料尚未更新時白跑全市場 prescan。
DAILY_LIGHT_PROBE_ENABLED = os.getenv("DAILY_LIGHT_PROBE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
DAILY_LIGHT_PROBE_SAMPLE_SIZE = int(os.getenv("DAILY_LIGHT_PROBE_SAMPLE_SIZE", "300"))
DAILY_LIGHT_PROBE_HISTORY_TRADING_DAYS = int(os.getenv("DAILY_LIGHT_PROBE_HISTORY_TRADING_DAYS", "5"))
DAILY_LIGHT_PROBE_SCAN_DAYS = int(os.getenv("DAILY_LIGHT_PROBE_SCAN_DAYS", "2"))
DAILY_LIGHT_PROBE_WORKERS = int(os.getenv("DAILY_LIGHT_PROBE_WORKERS", str(min(PRESCAN_WORKERS, 80))))

# 價格快取同步加速：
# 價格快取筆數很大時，若每次都整張重寫 Google Sheet 會拖慢整體執行。
# 本機 CSV 仍會完整保存；Google Sheet 則在快取很大且本次只有部分代號更新時，改用增量 append。
PRICE_CACHE_GSHEET_INCREMENTAL_APPEND = os.getenv("PRICE_CACHE_GSHEET_INCREMENTAL_APPEND", "1").strip().lower() not in ("0", "false", "no")
PRICE_CACHE_FULL_SYNC_THRESHOLD_ROWS = int(os.getenv("PRICE_CACHE_FULL_SYNC_THRESHOLD_ROWS", "80000"))

# TOP15 圖片用固定資料集：
# 主程式在同一次 RUN 內，會把 TOP15 所需的所有部位明細、賣出扣減、價格快照與報酬率
# 一次整理成「快取_TOP15部位明細」與「快取_TOP15共識淨買超」。
# 圖片程式之後應只讀這兩張固定資料集，不要再從 A/B/C/D/E 或快取歷史即時計算。
TOP15_CACHE_ENABLED = os.getenv("TOP15_CACHE_ENABLED", os.getenv("TOP15_RETURN_CACHE_ENABLED", "1")).strip().lower() not in ("0", "false", "no")
TOP15_POSITION_DETAIL_SHEET = os.getenv("TOP15_POSITION_DETAIL_SHEET", "快取_TOP15部位明細")
TOP15_CONSENSUS_SHEET = os.getenv("TOP15_CONSENSUS_SHEET", "快取_TOP15共識淨買超")
TOP15_LOOKBACK_TRADING_DAYS = int(os.getenv("TOP15_LOOKBACK_TRADING_DAYS", os.getenv("TOP15_RETURN_LOOKBACK_TRADING_DAYS", os.getenv("LOOKBACK_TRADING_DAYS", "22"))))
TOP15_PRICE_LOOKBACK_DAYS = int(os.getenv("TOP15_PRICE_LOOKBACK_DAYS", os.getenv("TOP15_RETURN_PRICE_LOOKBACK_DAYS", "75")))
TOP15_PRICE_STALE_DAYS = int(os.getenv("TOP15_PRICE_STALE_DAYS", os.getenv("TOP15_RETURN_PRICE_STALE_DAYS", "10")))
TOP15_FAIL_ON_MISSING_PRICE = os.getenv("TOP15_FAIL_ON_MISSING_PRICE", "1").strip().lower() not in ("0", "false", "no")
# 若 TOP15 剩餘部位因多日未造市 / 無成交而缺少有效權證價格，
# 預設不讓 RUN 失敗，而是保留淨買超成本，但該筆部位不納入報酬率估算。
TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN = os.getenv("TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN", "1").strip().lower() not in ("0", "false", "no")
TOP15_TARGET_DATE = os.getenv("TOP15_TARGET_DATE", "").strip()

# 近 7／14／21 日「指定精選分點」權證共識買賣超 TOP15：
# 這張工作表只會在 RUN_MODE=2 完整分點清單模式建立 / 更新，
# 但實際排名只統計 WARRANT_CONSENSUS_SELECTED_BROKERS_DEFAULT 指定的 13 個精選分點。
# RUN_MODE=1 精選分點模式不會建立這張 sheet，因此同步到 Google Sheet 時也不會動到既有工作表。
WARRANT_CONSENSUS_7D_ENABLED = os.getenv("WARRANT_CONSENSUS_7D_ENABLED", "1").strip().lower() not in ("0", "false", "no")
WARRANT_CONSENSUS_7D_SHEET = os.getenv("WARRANT_CONSENSUS_7D_SHEET", "快取_近7日權證分點共識TOP15")
WARRANT_CONSENSUS_7D_DAYS = int(os.getenv("WARRANT_CONSENSUS_7D_DAYS", "7"))
WARRANT_CONSENSUS_14D_DAYS = int(os.getenv("WARRANT_CONSENSUS_14D_DAYS", "14"))
WARRANT_CONSENSUS_21D_DAYS = int(os.getenv("WARRANT_CONSENSUS_21D_DAYS", "21"))
WARRANT_CONSENSUS_7D_TOP_N = int(os.getenv("WARRANT_CONSENSUS_7D_TOP_N", "15"))
WARRANT_CONSENSUS_SELECTED_BROKERS_DEFAULT = [
    "元大南屯",
    "華南永昌台中",
    "新光",
    "統一三多",
    "永豐金竹科",
    "福邦證券",
    "群益金鼎新竹",
    "凱基士林",
    "元大內湖民權",
    "群益金鼎古亭",
    "兆豐板橋",
    "富邦敦南",
    "永豐金內湖",
]
WARRANT_CONSENSUS_SELECTED_BROKERS_ENV = os.getenv("WARRANT_CONSENSUS_SELECTED_BROKERS", "").strip()

# 近 10 日「單一分點 + 標的股」買賣明細快取：
# 這張工作表只會在 RUN_MODE=2 完整分點清單模式建立 / 更新。
# RUN_MODE=1 精選分點模式不會建立這張 sheet，因此同步到 Google Sheet 時也不會動到既有工作表。
# 不分類 A/B/C/D/E，只要 API5 / 快取_分點歷史有抓到資料，就依分點與標的股合併統計。
BROKER_10D_DETAIL_ENABLED = os.getenv("BROKER_10D_DETAIL_ENABLED", "1").strip().lower() not in ("0", "false", "no")
BROKER_10D_DETAIL_SHEET = os.getenv("BROKER_10D_DETAIL_SHEET", "快取_近10日分點買賣明細")
# 近 10 日「全分點」勝率排行：
# 這張工作表只會在 RUN_MODE=2 完整分點清單模式建立 / 更新。
# RUN_MODE=1 精選分點模式不會建立這張 sheet，因此同步到 Google Sheet 時也不會動到既有工作表。
BROKER_10D_WINRATE_RANK_ENABLED = os.getenv("BROKER_10D_WINRATE_RANK_ENABLED", "1").strip().lower() not in ("0", "false", "no")
BROKER_10D_WINRATE_RANK_SHEET = os.getenv("BROKER_10D_WINRATE_RANK_SHEET", "快取_近10日分點勝率排行")
BROKER_10D_DETAIL_DAYS = int(os.getenv("BROKER_10D_DETAIL_DAYS", "10"))
BROKER_10D_PRICE_LOOKBACK_DAYS = int(os.getenv("BROKER_10D_PRICE_LOOKBACK_DAYS", "90"))
# 近10日明細價格補抓加速：先抓較短區間；完全沒有價格時才補抓完整區間。
BROKER_10D_PRICE_FAST_LOOKBACK_DAYS = int(os.getenv("BROKER_10D_PRICE_FAST_LOOKBACK_DAYS", "30"))
BROKER_10D_PRICE_STALE_DAYS = int(os.getenv("BROKER_10D_PRICE_STALE_DAYS", "10"))
# 近10日分點圖卡新增「現股10日」欄位後，需同步預抓標的股最新收盤價。
# 這裡抓較短區間即可涵蓋 10 日漲跌幅起訖價，避免只更新權證價卻讓現股10日停在前一交易日。
BROKER_10D_UNDERLYING_PRICE_LOOKBACK_DAYS = int(os.getenv("BROKER_10D_UNDERLYING_PRICE_LOOKBACK_DAYS", "35"))
# 價格預抓完整模式：
# 當分點資料尚未更新到今天時，預抓模式不只補事件 / TOP15 / 近10日圖卡目前會用到的價格，
# 也會把既有分點歷史快取中所有權證與標的股的最新價格先補進 price_cache。
# 這樣盤後重跑時，所有可能被後續報表 / 圖卡用到的價格都會先準備好。
PRICE_PREFETCH_ALL_ITEM_PRICES = os.getenv("PRICE_PREFETCH_ALL_ITEM_PRICES", "1").strip().lower() not in ("0", "false", "no")
PRICE_PREFETCH_ALL_WARRANT_PRICE_LOOKBACK_DAYS = int(os.getenv("PRICE_PREFETCH_ALL_WARRANT_PRICE_LOOKBACK_DAYS", str(BROKER_10D_PRICE_LOOKBACK_DAYS)))
PRICE_PREFETCH_ALL_UNDERLYING_PRICE_LOOKBACK_DAYS = int(os.getenv("PRICE_PREFETCH_ALL_UNDERLYING_PRICE_LOOKBACK_DAYS", str(BROKER_10D_UNDERLYING_PRICE_LOOKBACK_DAYS)))
PRICE_PREFETCH_ALL_REQUIRE_TARGET_DATE = os.getenv("PRICE_PREFETCH_ALL_REQUIRE_TARGET_DATE", "1").strip().lower() not in ("0", "false", "no")
# 預設不再為所有純賣超權證補抓最新價；賣超報酬優先使用 API5 歷史 FIFO 成本。
# 若賣超完全找不到歷史買進成本，仍會補抓最新價作為備援成本，避免報酬率空白。
BROKER_10D_FETCH_ALL_TRADED_WARRANT_PRICES = os.getenv("BROKER_10D_FETCH_ALL_TRADED_WARRANT_PRICES", "0").strip().lower() in ("1", "true", "yes")
BROKER_10D_FETCH_SELL_FALLBACK_PRICES = os.getenv("BROKER_10D_FETCH_SELL_FALLBACK_PRICES", "1").strip().lower() not in ("0", "false", "no")

# 執行模式：
# RUN_MODE=1：精選 5 分點模式。只追蹤 SELECTED_TARGET_LABELS，但會對這 5 間分點做全市場最近資料補掃，
#             讓今日買賣超明細盡量完整，例如元大南屯今日賣南亞科所有相關認購權證。
# RUN_MODE=2：完整清單模式。使用目前 TARGET_PATTERNS 內所有分點，維持原本完整分點清單邏輯。
RUN_MODE = int(os.getenv("RUN_MODE", os.getenv("BROKER_RUN_MODE", "1")) or "1")
SELECTED_TARGET_LABELS_DEFAULT = [
    "華南永昌台中",
    "元大南屯",
    "富邦敦南",
    "永豐金內湖",
    "新光",
]
SELECTED_TARGET_LABELS_ENV = os.getenv("SELECTED_TARGET_LABELS", "").strip()
SELECTED_FULL_SCAN_DAYS = int(os.getenv("SELECTED_FULL_SCAN_DAYS", str(CACHE_RECENT_SCAN_DAYS)))
# RUN_MODE=1 / RUN_MODE=2 加速追蹤設定：
# 舊版 RUN_MODE=1 會建立「所有認購權證 × 精選分點」，再用 SELECTED_REFRESH_ALL_WARRANTS
# 每次全部重打 API5，速度會非常慢。新版改成：
# 1. 先用 API4 掃最近有目標分點活動的標的股。
# 2. 再展開成「該標的所有認購權證 × 追蹤分點」。
# 3. 若歷史快取已有該候選資料，就走增量判斷，不再每天重抓 250 日 API5。
# 下列兩個舊環境變數保留相容，但預設流程已改成「活動標的擴展 + 快取增量」。
SELECTED_FORCE_ALL_WARRANTS = os.getenv("SELECTED_FORCE_ALL_WARRANTS", "1").strip().lower() not in ("0", "false", "no")
SELECTED_REFRESH_ALL_WARRANTS = os.getenv("SELECTED_REFRESH_ALL_WARRANTS", "1").strip().lower() not in ("0", "false", "no")
EXPAND_ACTIVE_UNDERLYING_WARRANTS = os.getenv("EXPAND_ACTIVE_UNDERLYING_WARRANTS", "1").strip().lower() not in ("0", "false", "no")

# 全面增量更新設定：
# 只要快取已有該「權證代號 + 券商代號」，就不再無差別重抓。
# API4 近期直接掃到有活動的候選，只有在快取最後日期落後時才補抓。
CACHE_INCREMENTAL_UPDATE_ENABLED = os.getenv("CACHE_INCREMENTAL_UPDATE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
CACHE_INCREMENTAL_REFRESH_LAG_DAYS = int(os.getenv("CACHE_INCREMENTAL_REFRESH_LAG_DAYS", "0"))
HISTORY_GSHEET_INCREMENTAL_APPEND = os.getenv("HISTORY_GSHEET_INCREMENTAL_APPEND", "1").strip().lower() not in ("0", "false", "no")
HISTORY_CACHE_FULL_SYNC_THRESHOLD_ROWS = int(os.getenv("HISTORY_CACHE_FULL_SYNC_THRESHOLD_ROWS", "200000"))

# 快取自動清理：
# - 分點歷史與價格只保留最近指定的「實際交易日」；預設皆為 200 個交易日。
# - 自動移除不在 FULL_TARGET_PATTERNS / FULL_FALLBACK 的分點資料。
# - 候選組合只保留目前仍上市的權證與目前設定中的分點。
# - 清理結果會同步到本機 CSV、Supabase；若大型快取仍使用 Google Sheet，也會整表覆蓋刪除舊資料。
CACHE_AUTO_PRUNE_ENABLED = os.getenv("CACHE_AUTO_PRUNE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
HISTORY_RETENTION_TRADING_DAYS = max(int(os.getenv("HISTORY_RETENTION_TRADING_DAYS", "200")), 1)
PRICE_RETENTION_TRADING_DAYS = max(int(os.getenv("PRICE_RETENTION_TRADING_DAYS", "200")), 1)
SUPABASE_AUTO_PRUNE_ENABLED = os.getenv("SUPABASE_AUTO_PRUNE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC = os.getenv("CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC", "1").strip().lower() not in ("0", "false", "no")
API5_HISTORY_LIMIT = max(int(os.getenv("API5_HISTORY_LIMIT", str(HISTORY_RETENTION_TRADING_DAYS))), 1)

# prescan_all() 會更新這個集合，主流程用它判斷哪些候選組合需要重新 api5_get。
# 注意：新版只把「API4 直接掃到近期有活動」的 key 放進來。
# 活動標的擴展出的 key 若已有快取，不再強制刷新，避免 RUN_MODE=2 候選爆量。
PRESCAN_REFRESH_KEYS = set()
# 這個集合是「本次 API4 直接候選 + 活動標的擴展候選」。
# 若某候選沒有歷史快取，只有在這個集合內才補抓；避免舊候選快取裡的全市場空候選全部打 API5。
PRESCAN_MISSING_FETCH_KEYS = set()
# API4 預掃描觀察到的目標分點最新活動日期。
# 這用來判斷「今天分點資料是否已經出來」。若尚未出來，主流程會自動改成價格預抓模式。
PRESCAN_LATEST_ACTIVITY_DATE = None
PRESCAN_TODAY_ACTIVITY_FOUND = False
PRESCAN_STATUS_LAST_RECORD = {}

# MoneyDJ Search 自動補漏模式：
# 當 OpenAPI / API4 最新交易日落後本次報表目標日時，不再只做價格預抓後結束，
# 而是自動啟用 MoneyDJ Search 補漏模式，強制用 MoneyDJ API5 補抓可能漏掉的候選組合。
# 這個模式同時適用 RUN_MODE=1 精選五分點與 RUN_MODE=2 全分點。
MONEYDJ_SEARCH_REPAIR_ENABLED = os.getenv("MONEYDJ_SEARCH_REPAIR_ENABLED", "1").strip().lower() not in ("0", "false", "no")
MONEYDJ_SEARCH_REPAIR_RECENT_HISTORY_DAYS = int(os.getenv("MONEYDJ_SEARCH_REPAIR_RECENT_HISTORY_DAYS", "45"))
MONEYDJ_SEARCH_REPAIR_INCLUDE_OPEN_POSITION = os.getenv("MONEYDJ_SEARCH_REPAIR_INCLUDE_OPEN_POSITION", "1").strip().lower() not in ("0", "false", "no")
MONEYDJ_SEARCH_REPAIR_SELECTED_FULL_POOL = os.getenv("MONEYDJ_SEARCH_REPAIR_SELECTED_FULL_POOL", "1").strip().lower() not in ("0", "false", "no")
MONEYDJ_SEARCH_REPAIR_MAX_FETCH_KEYS = int(os.getenv("MONEYDJ_SEARCH_REPAIR_MAX_FETCH_KEYS", "0"))
# API4 / OpenAPI 尚未更新到報表目標日時，額外用「近期有動作 / 尚有庫存的標的股」
# 展開成同標的所有認購權證 × 同分點，並用 API5 強制檢查今天是否已有新買賣。
# 這不是全市場無差別掃描，而是先從分點歷史快取找出高機率有動作的標的股，再補抓該標的權證池。
MONEYDJ_SEARCH_REPAIR_HISTORY_UNDERLYING_EXPAND_ENABLED = os.getenv("MONEYDJ_SEARCH_REPAIR_HISTORY_UNDERLYING_EXPAND_ENABLED", "1").strip().lower() not in ("0", "false", "no")
MONEYDJ_SEARCH_REPAIR_FORCE_HISTORY_UNDERLYING_FETCH = os.getenv("MONEYDJ_SEARCH_REPAIR_FORCE_HISTORY_UNDERLYING_FETCH", "1").strip().lower() not in ("0", "false", "no")
# 若真的要 100% 不漏 API4 未更新時的全新陌生標的，可手動開啟下列設定；預設關閉，避免全市場掃描太慢。
MONEYDJ_SEARCH_REPAIR_FULL_POOL_FORCE_FETCH = os.getenv("MONEYDJ_SEARCH_REPAIR_FULL_POOL_FORCE_FETCH", "0").strip().lower() in ("1", "true", "yes")
MONEYDJ_SEARCH_REPAIR_FULL_POOL_MAX_FETCH = int(os.getenv("MONEYDJ_SEARCH_REPAIR_FULL_POOL_MAX_FETCH", "0"))
MONEYDJ_SEARCH_REPAIR_ACTIVE = False
MONEYDJ_SEARCH_REPAIR_TARGET_DATE = ""
MONEYDJ_SEARCH_REPAIR_OPENAPI_LATEST_DATE = ""
MONEYDJ_SEARCH_REPAIR_FETCH_KEYS = set()
MONEYDJ_SEARCH_REPAIR_DISCOVERY_FETCH_KEYS = set()
MONEYDJ_SEARCH_REPAIR_REASON = ""

TARGET_PATTERNS = {
    "富邦公益":       r"富邦.*公益",
    "富邦敦南":       r"富邦.*敦南",
    "富邦仁愛":       r"富邦.*仁愛",
    "新光":           r"^新光$",
    "永豐金內湖":     r"永豐.*內湖",
    "永豐金竹北":     r"永豐.*竹北",
    "永豐金竹科":     r"永豐.*竹科",
    "永豐金市政":     r"永豐.*市政",
    "永豐金信義":     r"永豐.*信義",
    "華南永昌台中":   r"華南.*台中",
    "華南永昌淡水":   r"華南.*淡水",
    "福邦證券":       r"^福邦證券",
    "群益東大":       r"群益.*東大",
    "群益金鼎古亭":   r"群益.*古亭",
    "群益金鼎新竹":   r"群益.*新竹",
    "元大內湖民權":   r"元大.*(內湖.*民權|民權)",
    "元大南屯":       r"元大.*南屯",
    "元大汐止":       r"元大.*汐止",
    "元大虎尾":       r"元大.*虎尾",
    "元大彰化民生":   r"元大.*彰化民生",
    "兆豐板橋":       r"兆豐.*板橋",
    "凱基士林":       r"凱基.*士林",
    "凱基中山":       r"凱基.*中山",
    "國票敦北法人":   r"國票.*(敦北|法人)",
    "統一三多":       r"統一.*三多",
    "第一金中壢":     r"第一金.*中壢",
}

FALLBACK = {
    "富邦公益":       ("富邦-公益",       "961F"),
    "富邦敦南":       ("富邦-敦南",       "9663"),
    "富邦仁愛":       ("富邦-仁愛",       "9676"),
    "新光":           ("新光",           "8560"),
    "永豐金內湖":     ("永豐金-內湖",     "9A9g"),
    "永豐金竹北":     ("永豐金-竹北",     "9A9P"),
    "永豐金竹科":     ("永豐金-竹科",     "9A9X"),
    "永豐金市政":     ("永豐金-市政",     "9A9W"),
    "永豐金信義":     ("永豐金-信義",     "9A9R"),
    "華南永昌台中":   ("華南永昌-台中",   "9302"),
    "華南永昌淡水":   ("華南永昌-淡水",   "9316"),
    "福邦證券":       ("福邦證券",       "6480"),
    "群益東大":       ("群益金鼎-東大",   "9135"),
    "群益金鼎古亭":   ("群益金鼎-古亭",   "918C"),
    "群益金鼎新竹":   ("群益金鼎-新竹",   "9186"),
    "元大內湖民權":   ("元大-內湖民權",   "9867"),
    "元大南屯":       ("元大-南屯",       "9853"),
    "元大汐止":       ("元大-汐止",       "989Q"),
    "元大虎尾":       ("元大-虎尾",       "980l"),
    "元大彰化民生":   ("元大-彰化民生",   "989J"),
    "兆豐板橋":       ("兆豐-板橋",       "700B"),
    "凱基士林":       ("凱基-士林",       "9238"),
    "凱基中山":       ("凱基-中山",       "9229"),
    "國票敦北法人":   ("國票-敦北法人",   "779c"),
    "統一三多":       ("統一-三多",       "585Q"),
    "第一金中壢":     ("第一金-中壢",      "538Y"),
}

FULL_TARGET_PATTERNS = dict(TARGET_PATTERNS)
FULL_FALLBACK = dict(FALLBACK)
CURRENT_LIVE_WARRANT_CODES = set()
LIVE_WARRANT_SNAPSHOT_READY = False
LIVE_WARRANT_MARKET_SUCCESS_COUNT = 0
LIVE_WARRANT_MARKET_EXPECTED_COUNT = 2


def parse_selected_target_labels():
    if SELECTED_TARGET_LABELS_ENV:
        labels = [
            x.strip()
            for x in re.split(r"[,;；、\n\r\t]+", SELECTED_TARGET_LABELS_ENV)
            if x.strip()
        ]
    else:
        labels = list(SELECTED_TARGET_LABELS_DEFAULT)

    out = []
    for label in labels:
        if label not in out:
            out.append(label)

    return out


def configure_run_mode():
    """
    依 RUN_MODE 切換分點範圍與候選快取。

    RUN_MODE=1：
    - 只保留 SELECTED_TARGET_LABELS 指定的精選分點。
    - 候選快取使用 candidates_cache_selected5.csv，避免與完整清單模式混在一起。
    - 歷史分點快取仍共用 broker_warrant_history_cache.csv，因為 key 是權證代號 + 券商代號 + 日期，
      可讓精選模式抓到的新資料補強整體歷史資料。

    RUN_MODE=2：
    - 使用完整 TARGET_PATTERNS 分點清單。
    - 候選快取使用原本 candidates_cache.csv。
    """
    global RUN_MODE, TARGET_PATTERNS, FALLBACK, CANDIDATES_CACHE_PATH

    try:
        RUN_MODE = int(os.getenv("RUN_MODE", os.getenv("BROKER_RUN_MODE", str(RUN_MODE))) or "1")
    except Exception:
        RUN_MODE = 1

    if RUN_MODE not in (1, 2):
        print(f"  ⚠️ RUN_MODE={RUN_MODE} 不支援，改用 RUN_MODE=1 精選分點模式。")
        RUN_MODE = 1

    if RUN_MODE == 1:
        selected_labels = parse_selected_target_labels()
        missing = [label for label in selected_labels if label not in FULL_TARGET_PATTERNS]

        if missing:
            print(f"  ⚠️ SELECTED_TARGET_LABELS 中有不存在於 TARGET_PATTERNS 的分點，已略過：{missing}")

        active_labels = [label for label in selected_labels if label in FULL_TARGET_PATTERNS]

        if not active_labels:
            print("  ⚠️ 精選分點清單為空，改用預設 5 間分點。")
            active_labels = [
                label for label in SELECTED_TARGET_LABELS_DEFAULT
                if label in FULL_TARGET_PATTERNS
            ]

        TARGET_PATTERNS = {
            label: FULL_TARGET_PATTERNS[label]
            for label in active_labels
        }
        FALLBACK = {
            label: FULL_FALLBACK[label]
            for label in active_labels
            if label in FULL_FALLBACK
        }
        CANDIDATES_CACHE_PATH = CANDIDATES_CACHE_SELECTED5_PATH
        print("  ✅ RUN_MODE=1：精選分點全市場追蹤模式")
        print(f"  ✅ 精選分點：{', '.join(TARGET_PATTERNS.keys())}")
        print(f"  ✅ 候選快取：{CANDIDATES_CACHE_PATH}")
    else:
        TARGET_PATTERNS = dict(FULL_TARGET_PATTERNS)
        FALLBACK = dict(FULL_FALLBACK)
        CANDIDATES_CACHE_PATH = CANDIDATES_CACHE_ALL_PATH
        print("  ✅ RUN_MODE=2：完整分點清單模式")
        print(f"  ✅ 分點數：{len(TARGET_PATTERNS)}")
        print(f"  ✅ 候選快取：{CANDIDATES_CACHE_PATH}")


def filter_broker_map_for_active_targets(broker_map):
    if not broker_map:
        return {}

    active_labels = set(TARGET_PATTERNS.keys())
    return {
        label: value
        for label, value in broker_map.items()
        if label in active_labels
    }


def filter_candidates_by_broker_map(candidates, broker_map):
    if not candidates:
        return []

    if not broker_map:
        return []

    allowed_labels = set(broker_map.keys())
    allowed_codes = {str(code).strip() for _, code in broker_map.values()}

    out = []
    for c in candidates:
        try:
            label = str(c[4]).strip()
            broker_code = str(c[6]).strip()
        except Exception:
            continue

        if label in allowed_labels or broker_code in allowed_codes:
            out.append(c)

    return out


HDR = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "*/*",
    "Referer": "https://pscnetsecrwd.moneydj.com/",
}

_THREAD_LOCAL = threading.local()


def get_thread_session():
    """
    每個執行緒各自建立並重用 requests.Session。

    目的：
    1. 避免 api4_get / api5_get 每次呼叫都重新建立連線。
    2. 在 ThreadPoolExecutor 多執行緒抓資料時，每個 thread 使用自己的 Session，
       避免多執行緒共用同一個 Session 造成不穩定。
    3. 不改變任何抓資料邏輯，只改善大量 API 請求時的連線重用效率。
    """
    session = getattr(_THREAD_LOCAL, "session", None)

    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=1)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _THREAD_LOCAL.session = session

    return session


API4 = ("https://pscnetsecrwd.moneydj.com/b2brwdCommon/jsondata"
        "/9b/6e/0a/TwWarrantData.xdjjson"
        "?a={code}&x=warrant-chip0002-4&c={start}&d={end}&revision=2018_07_31_1")

API5 = ("https://pscnetsecrwd.moneydj.com/b2brwdCommon/jsondata"
        "/d8/f5/27/twWarrantData.xdjjson"
        "?x=warrant-chip0002-5&c={limit}&a={warrant}&b={broker}&revision=2018_07_31_1")

# Excel 顏色（柔和舒適版）
RED    = PatternFill("solid", fgColor="F4CCCC")
GREEN  = PatternFill("solid", fgColor="D9EAD3")
BLUE   = PatternFill("solid", fgColor="D9EAF7")
ORANGE = PatternFill("solid", fgColor="FCE5CD")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GRAY   = PatternFill("solid", fgColor="E7E6E6")
WHITE  = PatternFill("solid", fgColor="FFFFFF")

# 出清且獲利時使用的外框（不改底色，只用外框凸顯）
PROFIT_EXIT_SIDE = Side(style="thick", color="C00000")
LOSS_EXIT_SIDE   = Side(style="thick", color="38761D")


# ══════════════════════════════════════════════════════════════════════
# 工具函式
# ══════════════════════════════════════════════════════════════════════

def parse_date(date_str):
    try:
        if date_str is None:
            return None
        s = str(date_str).strip()
        if not s or s == "-":
            return None
        s = s.replace("-", "/")
        parts = s.split("/")
        if len(parts) != 3:
            return None
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        return datetime(y, m, d)
    except:
        return None


def normalize_date_str(date_str):
    dt = parse_date(date_str)
    return dt.strftime("%Y/%m/%d") if dt else str(date_str).strip()


def add_months(dt, months):
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    return datetime(year, month, 1)


def iter_month_starts(start_dt, end_dt):
    start_dt = datetime(start_dt.year, start_dt.month, 1)
    end_dt   = datetime(end_dt.year, end_dt.month, 1)

    cur = start_dt
    while cur <= end_dt:
        yield cur.strftime("%Y%m01")
        cur = add_months(cur, 1)


def fmt_pct(v):
    return "-" if v is None else f"{v:+.2f}%"


def fmt_num(v):
    return "-" if v is None else v


def fmt_amount(v):
    if v is None:
        return "-"
    try:
        return f"{int(round(float(v))):,}"
    except:
        return str(v)


def calc_result_tag(return_pct):
    if return_pct is None:
        return "未出清"
    if return_pct > 0:
        return "勝"
    if return_pct < 0:
        return "敗"
    return "平手"


def match_target(name):
    for label, pat in TARGET_PATTERNS.items():
        if re.search(pat, name):
            return label
    return ""


def api4_get(code, start, end):
    try:
        session = get_thread_session()
        r = session.get(API4.format(code=code, start=start, end=end), headers=HDR, timeout=(5, 12))
        data = json.loads(r.content.decode("utf-8"))
        rows = []
        for item in (data if isinstance(data, list) else [data]):
            rows.extend(item.get("ResultSet", {}).get("Result", []))
        return rows
    except:
        return []


def api5_get(warrant, broker):
    try:
        session = get_thread_session()
        r = session.get(API5.format(warrant=warrant, broker=broker, limit=API5_HISTORY_LIMIT), headers=HDR, timeout=(5, 12))
        data = json.loads(r.content.decode("utf-8"))
        rs = data[0].get("ResultSet", {}) if isinstance(data, list) else data.get("ResultSet", {})
        return rs.get("Result", [])
    except:
        return []


def safe_price_float(x):
    try:
        s = str(x).replace(",", "").replace("--", "").replace("X", "").strip()

        if s in ["", "-", "---", "除權息", "None", "nan", "null"]:
            return None

        v = float(s)

        # 權證 / 股價不應該用 0 當有效收盤價。
        # 測試時發現部分權證會回傳 0.0，不能拿來計算 D+。
        if v <= 0:
            return None

        return v
    except:
        return None


def merge_price_dicts(*dicts):
    merged = {}

    for prices in dicts:
        if not prices:
            continue

        for d, p in prices.items():
            if p is not None and p > 0:
                merged[d] = p

    return merged


def normalize_price_code(code):
    s = str(code).strip()

    if s.endswith(".0"):
        s = s[:-2]

    s = "".join(ch for ch in s if ch.isdigit())

    if not s:
        return ""

    # 權證通常為 6 碼；股票通常為 4 碼。
    # 若是 5 碼權證，很可能是 Excel / pandas 吃掉前導 0，補回 6 碼。
    if len(s) == 5:
        return s.zfill(6)

    return s




def configured_broker_labels_for_scope(scope="all"):
    scope = str(scope or "all").strip().lower()
    if scope == "selected5":
        return {
            label for label in parse_selected_target_labels()
            if label in FULL_TARGET_PATTERNS
        }
    return set(FULL_TARGET_PATTERNS.keys())


def configured_broker_codes_for_scope(scope="all"):
    labels = configured_broker_labels_for_scope(scope)
    return {
        str(FULL_FALLBACK[label][1]).strip()
        for label in labels
        if label in FULL_FALLBACK and str(FULL_FALLBACK[label][1]).strip()
    }


def candidate_scope_from_path(path):
    base = os.path.basename(str(path))
    return "selected5" if base == "candidates_cache_selected5.csv" else "all"


def recent_trading_date_cutoff_from_series(series, keep_days):
    keep_days = max(int(keep_days or 1), 1)
    parsed = pd.to_datetime(series.astype(str).str.replace("/", "-", regex=False), errors="coerce")
    valid_dates = sorted({d.date() for d in parsed.dropna().tolist()}, reverse=True)
    if not valid_dates:
        return None
    return valid_dates[min(keep_days - 1, len(valid_dates) - 1)]


def prune_price_cache_dataframe(df):
    stats = {"before": 0, "after": 0, "removed": 0, "cutoff": ""}
    if df is None or df.empty:
        return pd.DataFrame(columns=["代號", "日期", "收盤價"]), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    required = {"代號", "日期", "收盤價"}
    if not required.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    out["_parsed_date"] = pd.to_datetime(
        out["日期"].astype(str).str.replace("/", "-", regex=False),
        errors="coerce",
    )
    out = out[out["_parsed_date"].notna()].copy()
    cutoff = recent_trading_date_cutoff_from_series(out["日期"], PRICE_RETENTION_TRADING_DAYS)
    if cutoff is not None:
        out = out[out["_parsed_date"].dt.date >= cutoff].copy()
        stats["cutoff"] = cutoff.strftime("%Y/%m/%d")

    out["代號"] = out["代號"].map(normalize_price_code)
    out = out[out["代號"].astype(str).str.strip() != ""].copy()
    out["日期"] = out["_parsed_date"].dt.strftime("%Y/%m/%d")
    out["收盤價"] = pd.to_numeric(out["收盤價"], errors="coerce")
    out = out[out["收盤價"].fillna(0) > 0].copy()
    out = out.drop(columns=["_parsed_date"], errors="ignore")
    out = out.drop_duplicates(subset=["代號", "日期"], keep="last")
    out = out.sort_values(["代號", "日期"]).reset_index(drop=True)
    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out[["代號", "日期", "收盤價"]], stats


def prune_history_cache_dataframe(df):
    stats = {
        "before": 0,
        "after": 0,
        "removed": 0,
        "removed_broker": 0,
        "removed_date": 0,
        "cutoff": "",
    }
    if df is None or df.empty:
        return pd.DataFrame(), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    required = {"分點", "券商代號", "日期"}
    if not required.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    allowed_labels = configured_broker_labels_for_scope("all")
    allowed_codes = configured_broker_codes_for_scope("all")
    label_ok = out["分點"].astype(str).str.strip().isin(allowed_labels)
    code_ok = out["券商代號"].astype(str).str.strip().isin(allowed_codes)
    broker_mask = label_ok | code_ok
    stats["removed_broker"] = int((~broker_mask).sum())
    out = out[broker_mask].copy()

    out["_parsed_date"] = pd.to_datetime(
        out["日期"].astype(str).str.replace("/", "-", regex=False),
        errors="coerce",
    )
    invalid_date_count = int(out["_parsed_date"].isna().sum())
    out = out[out["_parsed_date"].notna()].copy()

    cutoff = recent_trading_date_cutoff_from_series(out["日期"], HISTORY_RETENTION_TRADING_DAYS)
    before_date_filter = len(out)
    if cutoff is not None:
        out = out[out["_parsed_date"].dt.date >= cutoff].copy()
        stats["cutoff"] = cutoff.strftime("%Y/%m/%d")
    stats["removed_date"] = invalid_date_count + max(before_date_filter - len(out), 0)

    out["日期"] = out["_parsed_date"].dt.strftime("%Y/%m/%d")
    out = out.drop(columns=["_parsed_date"], errors="ignore")
    if {"權證代號", "券商代號", "日期"}.issubset(set(out.columns)):
        out = out.drop_duplicates(
            subset=["權證代號", "券商代號", "日期"],
            keep="last",
        )
        out = out.sort_values(["權證代號", "券商代號", "日期"]).reset_index(drop=True)

    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out, stats


def prune_candidates_dataframe(df, path=None, valid_warrant_codes=None):
    stats = {"before": 0, "after": 0, "removed": 0, "removed_broker": 0, "removed_warrant": 0}
    if df is None or df.empty:
        return pd.DataFrame(), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    required = {"權證代號", "分點", "券商代號"}
    if not required.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    scope = candidate_scope_from_path(path or CANDIDATES_CACHE_PATH)
    allowed_labels = configured_broker_labels_for_scope(scope)
    allowed_codes = configured_broker_codes_for_scope(scope)
    label_ok = out["分點"].astype(str).str.strip().isin(allowed_labels)
    code_ok = out["券商代號"].astype(str).str.strip().isin(allowed_codes)
    broker_mask = label_ok | code_ok
    stats["removed_broker"] = int((~broker_mask).sum())
    out = out[broker_mask].copy()

    codes = {
        str(code).strip() for code in (valid_warrant_codes or set())
        if str(code).strip()
    }
    if codes:
        warrant_mask = out["權證代號"].astype(str).str.strip().isin(codes)
        stats["removed_warrant"] = int((~warrant_mask).sum())
        out = out[warrant_mask].copy()

    out = out.drop_duplicates(subset=["權證代號", "券商代號"], keep="last")
    out = out.sort_values(["權證代號", "券商代號"]).reset_index(drop=True)
    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out, stats


def prune_broker_map_dataframe(df, scope=None):
    stats = {"before": 0, "after": 0, "removed": 0}
    if df is None or df.empty:
        return pd.DataFrame(), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    if not {"分點", "券商代號"}.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    scope = scope or ("selected5" if RUN_MODE == 1 else "all")
    labels = configured_broker_labels_for_scope(scope)
    codes = configured_broker_codes_for_scope(scope)
    mask = out["分點"].astype(str).str.strip().isin(labels) | out["券商代號"].astype(str).str.strip().isin(codes)
    out = out[mask].copy()
    out = out.drop_duplicates(subset=["分點"], keep="last").reset_index(drop=True)
    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out, stats

def price_code_variants(code):
    code = normalize_price_code(code)

    if not code:
        return []

    variants = [code]

    no_zero = code.lstrip("0")
    if no_zero and no_zero != code:
        variants.append(no_zero)

    out = []
    for v in variants:
        if v not in out:
            out.append(v)

    return out


def yahoo_symbol_variants(code):
    symbols = []

    for c in price_code_variants(code):
        symbols.append(f"{c}.TW")
        symbols.append(f"{c}.TWO")

    out = []
    for s in symbols:
        if s not in out:
            out.append(s)

    return out


def fetch_twse_stock_day_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_dt in iter_month_starts(start_dt, end_dt):
        try:
            session = get_thread_session()
            rp = session.get(
                f"https://www.twse.com.tw/exchangeReport/STOCK_DAY"
                f"?response=json&date={month_dt}&stockNo={code}",
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=8
            )
            data = rp.json()

            for row in data.get("data", []):
                try:
                    parts = str(row[0]).split("/")
                    dk = f"{int(parts[0]) + 1911}/{int(parts[1]):02d}/{int(parts[2]):02d}"
                    close_price = safe_price_float(row[6])

                    if close_price is not None:
                        prices[dk] = close_price
                except:
                    pass
        except:
            pass

    return prices


def fetch_tpex_new_trading_stock_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_start in iter_month_starts(start_dt, end_dt):
        try:
            month_dt = datetime.strptime(month_start, "%Y%m%d")
            date_str = month_dt.strftime("%Y/%m/01")

            urls = [
                (
                    "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock"
                    f"?code={code}&date={date_str}&response=json"
                ),
                (
                    "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock"
                    f"?code={code}&date={date_str}&type=EW&response=json"
                ),
            ]

            for url in urls:
                try:
                    session = get_thread_session()
                    rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
                    data = rp.json()

                    rows = []

                    if isinstance(data, dict):
                        if isinstance(data.get("tables"), list):
                            for table in data.get("tables", []):
                                if isinstance(table, dict) and isinstance(table.get("data"), list):
                                    rows.extend(table.get("data", []))

                        if isinstance(data.get("data"), list):
                            rows.extend(data.get("data", []))

                        if isinstance(data.get("aaData"), list):
                            rows.extend(data.get("aaData", []))

                    for row in rows:
                        try:
                            if not isinstance(row, (list, tuple)) or not row:
                                continue

                            raw_date = str(row[0]).strip().replace("-", "/")
                            parts = raw_date.split("/")

                            if len(parts) != 3:
                                continue

                            y = int(parts[0])

                            if y < 1911:
                                y += 1911

                            dk = f"{y}/{int(parts[1]):02d}/{int(parts[2]):02d}"

                            close_price = None

                            # 測試結果顯示 TPEx 新版 tradingStock 是 70xxxx 權證與上櫃股最穩來源。
                            # 常見收盤價欄位落在 6 / 5 / 4 / 7 / 3，逐一嘗試。
                            for idx in [6, 5, 4, 7, 3]:
                                if idx < len(row):
                                    v = safe_price_float(row[idx])

                                    if v is not None:
                                        close_price = v
                                        break

                            if close_price is not None:
                                prices[dk] = close_price
                        except:
                            pass

                except:
                    pass

        except:
            pass

    return prices


def fetch_tpex_old_st43_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_start in iter_month_starts(start_dt, end_dt):
        try:
            month_dt = datetime.strptime(month_start, "%Y%m%d")
            roc_month = f"{month_dt.year - 1911}/{month_dt.month:02d}"

            url = (
                "https://www.tpex.org.tw/web/stock/aftertrading/daily_trading_info/"
                f"st43_result.php?l=zh-tw&d={roc_month}&stkno={code}"
            )

            session = get_thread_session()
            rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            data = rp.json()

            rows = data.get("aaData", []) if isinstance(data, dict) else []

            for row in rows:
                try:
                    if not isinstance(row, (list, tuple)) or not row:
                        continue

                    raw_date = str(row[0]).strip().replace("-", "/")
                    parts = raw_date.split("/")

                    if len(parts) != 3:
                        continue

                    y = int(parts[0])

                    if y < 1911:
                        y += 1911

                    dk = f"{y}/{int(parts[1]):02d}/{int(parts[2]):02d}"

                    close_price = None

                    for idx in [6, 5, 4, 7, 3]:
                        if idx < len(row):
                            v = safe_price_float(row[idx])

                            if v is not None:
                                close_price = v
                                break

                    if close_price is not None:
                        prices[dk] = close_price
                except:
                    pass
        except:
            pass

    return prices


def fetch_yahoo_chart_prices(symbol, start_dt=None, end_dt=None, host="query1"):
    today = datetime.today()
    prices = {}

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    period1 = int(start_dt.timestamp())
    period2 = int((end_dt + timedelta(days=1)).timestamp())

    url = (
        f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}"
        f"?period1={period1}&period2={period2}&interval=1d&events=history"
    )

    try:
        session = get_thread_session()
        rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
        data = rp.json()

        result = data.get("chart", {}).get("result", [])

        if not result:
            return prices

        result = result[0]
        timestamps = result.get("timestamp", [])
        quote = result.get("indicators", {}).get("quote", [{}])[0]
        closes = quote.get("close", [])

        for ts, close_price in zip(timestamps, closes):
            v = safe_price_float(close_price)

            if v is None:
                continue

            dt = datetime.fromtimestamp(int(ts))
            prices[dt.strftime("%Y/%m/%d")] = v
    except:
        pass

    return prices


def fetch_yahoo_range_prices(symbol):
    prices = {}

    # 只在官方來源不足時才會進入 Yahoo，這裡用 5y + max 做快速備援。
    for range_value in ["5y", "max"]:
        url = (
            f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            f"?range={range_value}&interval=1d&events=history"
        )

        try:
            session = get_thread_session()
            rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            data = rp.json()

            result = data.get("chart", {}).get("result", [])

            if not result:
                continue

            result = result[0]
            timestamps = result.get("timestamp", [])
            quote = result.get("indicators", {}).get("quote", [{}])[0]
            closes = quote.get("close", [])

            for ts, close_price in zip(timestamps, closes):
                v = safe_price_float(close_price)

                if v is None:
                    continue

                dt = datetime.fromtimestamp(int(ts))
                prices[dt.strftime("%Y/%m/%d")] = v

            if prices:
                break
        except:
            pass

    return prices


def fetch_yahoo_download_prices(symbol, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    period1 = int(start_dt.timestamp())
    period2 = int((end_dt + timedelta(days=1)).timestamp())

    url = (
        f"https://query1.finance.yahoo.com/v7/finance/download/{symbol}"
        f"?period1={period1}&period2={period2}&interval=1d&events=history&includeAdjustedClose=true"
    )

    try:
        session = get_thread_session()
        rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)

        if rp.status_code != 200 or "Date" not in rp.text:
            return prices

        df = pd.read_csv(StringIO(rp.text))

        if "Date" not in df.columns or "Close" not in df.columns:
            return prices

        for _, row in df.iterrows():
            dt = parse_date(row["Date"])
            close_price = safe_price_float(row["Close"])

            if dt and close_price is not None:
                prices[dt.strftime("%Y/%m/%d")] = close_price
    except:
        pass

    return prices


def fetch_yahoo_prices(code, start_dt=None, end_dt=None):
    """
    Yahoo 備援：
    1. 同時測補零版與去零版，例如 064390.TW / 64390.TW。
    2. 同時測 .TW / .TWO。
    3. 優先 query1 chart，其次 query2、range、download。
    """
    prices = {}

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_chart_prices(symbol, start_dt, end_dt, host="query1")

        if p:
            return p

        p = fetch_yahoo_chart_prices(symbol, start_dt, end_dt, host="query2")

        if p:
            return p

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_range_prices(symbol)

        if p:
            return p

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_download_prices(symbol, start_dt, end_dt)

        if p:
            return p

    return prices


def prices_need_yahoo_fallback(prices, start_dt=None, end_dt=None):
    """
    判斷官方來源價格是否不足，需要 Yahoo 備援。

    只用 len(prices) < 2 不夠，因為有些來源雖然有幾筆，
    但覆蓋不到 D+1 ~ D+20 的日期，Excel 還是會出現大量 權:- / 標:-。
    """
    if not prices:
        return True

    valid_dates = sorted([d for d, p in prices.items() if p is not None and p > 0])

    if not valid_dates:
        return True

    # 如果測試區間超過 30 天，但有效價格少於 10 筆，代表覆蓋率明顯不足。
    if start_dt and end_dt:
        try:
            span_days = (end_dt - start_dt).days
            if span_days >= 30 and len(valid_dates) < 10:
                return True
        except:
            pass

    # 如果最後一筆價格離需要的結束日太遠，也要補 Yahoo。
    if end_dt:
        try:
            latest_dt = parse_date(valid_dates[-1])
            target_end = min(end_dt, datetime.today())

            if latest_dt and (target_end - latest_dt).days > 10:
                return True
        except:
            pass

    return False


def fetch_twse_prices(code, start_dt=None, end_dt=None):
    """
    統一價格抓取函式（保留原函式名稱，避免改動其他流程）：

    速度與準確率策略：
    1. 先用官方資料源：
       - 上市股票 / 上市權證：TWSE STOCK_DAY
       - 上櫃股票 / 上櫃權證：TPEx tradingStock
    2. 若官方來源覆蓋率不足，再補 Yahoo。
    3. 價格 <= 0 一律不採用。
    4. 權證 5 碼會自動補回 6 碼，避免前導 0 被吃掉。
    """
    code = normalize_price_code(code)

    if not code:
        return {}

    prices = {}

    # 先根據代號型態決定優先順序，減少不必要請求。
    # 70xxxx 權證大多走 TPEx；0xxxxx 權證大多走 TWSE。
    if len(code) == 6 and code.startswith("7"):
        prices = merge_price_dicts(
            prices,
            fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_old_st43_prices(code, start_dt, end_dt)
            )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_twse_stock_day_prices(code, start_dt, end_dt)
            )

    elif len(code) == 6 and code.startswith("0"):
        prices = merge_price_dicts(
            prices,
            fetch_twse_stock_day_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
            )

    else:
        prices = merge_price_dicts(
            prices,
            fetch_twse_stock_day_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
            )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_old_st43_prices(code, start_dt, end_dt)
            )

    # 官方來源覆蓋率不足時才用 Yahoo，兼顧速度與完整度。
    if prices_need_yahoo_fallback(prices, start_dt, end_dt):
        prices = merge_price_dicts(
            prices,
            fetch_yahoo_prices(code, start_dt, end_dt)
        )

    return prices


def get_price_nearest(prices, date):
    date = normalize_date_str(date)

    if date in prices:
        return prices[date]

    before = [d for d in sorted(prices) if d <= date]
    return prices[before[-1]] if before else None







# ══════════════════════════════════════════════════════════════════════
# Google Sheet 快取 / 結果同步工具（GitHub Actions 部署用）
# ══════════════════════════════════════════════════════════════════════

GOOGLE_SHEET_NAME = os.getenv("GOOGLE_SHEET_NAME", "權證分點資料_NEW")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", os.getenv("GSHEET_ID", "")).strip()
GSHEET_CACHE_ENABLED = os.getenv("GSHEET_CACHE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_RESULT_ENABLED = os.getenv("GSHEET_RESULT_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_CHUNK_ROWS = int(os.getenv("GSHEET_CHUNK_ROWS", "3000"))
GSHEET_SYNC_CACHE_ON_READ = os.getenv("GSHEET_SYNC_CACHE_ON_READ", "0").strip().lower() in ("1", "true", "yes")

# Google Sheet 結果表自動封存 / 保留設定：
# 1. 每日賣出明細預設只在主試算表保留最近 60 個實際交易日，可改成 30 / 60 / 90 或其他正整數。
# 2. TOP15 兩張快取表預設各資料範圍保留最近 5 個統計日期。
# 3. 超出保留範圍的舊資料會先寫入獨立的年度封存試算表，封存成功後才從主表移除。
# 4. 封存試算表依年份自動建立；單一年度封存檔接近儲存格上限時，會自動切換到 _02、_03...。
# 5. 同步前會把所有工作表縮到實際有內容的列數 / 欄數；後續每張寫入表也會直接 resize 成實際大小。
GSHEET_RESULT_ARCHIVE_ENABLED = os.getenv("GSHEET_RESULT_ARCHIVE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_ARCHIVE_NAME_PREFIX = os.getenv(
    "GSHEET_ARCHIVE_NAME_PREFIX",
    f"{GOOGLE_SHEET_NAME}_歷史封存",
).strip() or f"{GOOGLE_SHEET_NAME}_歷史封存"
GSHEET_ARCHIVE_SHARE_EMAILS = os.getenv("GSHEET_ARCHIVE_SHARE_EMAILS", "").strip()
# 若服務帳號需要在指定的 Shared Drive / 資料夾內建立封存檔，可填 Google Drive 資料夾 ID。
GSHEET_ARCHIVE_FOLDER_ID = os.getenv("GSHEET_ARCHIVE_FOLDER_ID", "").strip()
# 服務帳號通常沒有可用的個人 Drive 儲存空間；若 gc.create() 回傳
#「The user's Drive storage quota has been exceeded」，請先用自己的 Google 帳號建立一份
# 封存試算表並分享給服務帳號，再把試算表 ID 填入下列環境變數。
# 可填單一 ID，也可用逗號 / 分號分隔多個 ID；第一份接近格數上限時會自動換到下一份。
GSHEET_ARCHIVE_SPREADSHEET_IDS_RAW = os.getenv(
    "GSHEET_ARCHIVE_SPREADSHEET_IDS",
    os.getenv("GSHEET_ARCHIVE_SPREADSHEET_ID", ""),
).strip()
GSHEET_ARCHIVE_SPREADSHEET_IDS = [
    item.strip()
    for item in re.split(r"[,;；、\n\r\t]+", GSHEET_ARCHIVE_SPREADSHEET_IDS_RAW)
    if item.strip()
]
# 未提供既有封存試算表 ID 時，是否允許服務帳號自行建立新檔。
# 若 Drive quota 不足，程式會在第一次 403 後自動停用本次執行的建立動作，不再連續嘗試 _02～_99。
GSHEET_ARCHIVE_ALLOW_CREATE = os.getenv("GSHEET_ARCHIVE_ALLOW_CREATE", "1").strip().lower() not in ("0", "false", "no")
GSHEET_ARCHIVE_MAX_CELLS = max(int(os.getenv("GSHEET_ARCHIVE_MAX_CELLS", "8500000")), 100000)
GSHEET_ARCHIVE_YEARLY_SPLIT = os.getenv("GSHEET_ARCHIVE_YEARLY_SPLIT", "1").strip().lower() not in ("0", "false", "no")
GSHEET_TOP15_KEEP_STAT_DATES = max(int(os.getenv("GSHEET_TOP15_KEEP_STAT_DATES", "5")), 1)
GSHEET_DAILY_SELL_KEEP_TRADING_DAYS = max(int(os.getenv("GSHEET_DAILY_SELL_KEEP_TRADING_DAYS", "30")), 1)
GSHEET_COMPACT_BLANK_GRID_ENABLED = os.getenv("GSHEET_COMPACT_BLANK_GRID_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_COMPACT_MIN_ROWS = max(int(os.getenv("GSHEET_COMPACT_MIN_ROWS", "1")), 1)
GSHEET_COMPACT_MIN_COLS = max(int(os.getenv("GSHEET_COMPACT_MIN_COLS", "1")), 1)

# Google Sheets API 有「每分鐘寫入請求」限制。
# 結果工作表很多、又要同步格式時，如果沒有節流與 429 重試，
# 會出現後面工作表建立 / 寫入失敗，甚至因先刪後建導致工作表消失。
GSHEET_WRITE_SLEEP_SECONDS = float(os.getenv("GSHEET_WRITE_SLEEP_SECONDS", "1.25"))
GSHEET_MAX_RETRIES = int(os.getenv("GSHEET_MAX_RETRIES", "6"))
GSHEET_RETRY_BASE_SECONDS = float(os.getenv("GSHEET_RETRY_BASE_SECONDS", "12"))

_GSHEET_CLIENT = None
_GSHEET_SPREADSHEET = None
_GSHEET_LAST_WRITE_TS = 0.0
_GSHEET_ARCHIVE_SPREADSHEETS = {}
_GSHEET_ARCHIVE_PERMISSION_SYNCED = set()
_GSHEET_ARCHIVE_CREATE_BLOCKED = False
_GSHEET_ARCHIVE_CREATE_BLOCK_REASON = ""
_GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED = False

CACHE_SHEET_NAME_MAP = {
    "warrants_cache.csv": "快取_權證清單",
    "broker_map_cache.csv": "快取_分點代號",
    "candidates_cache.csv": "快取_候選組合",
    "candidates_cache_selected5.csv": "快取_候選組合_精選5",
    "broker_warrant_history_cache.csv": "快取_分點歷史",
    "price_cache.csv": "快取_價格",
    "price_prefetch_state.csv": "快取_價格預抓狀態",
    "prescan_status.csv": "快取_候選掃描狀態",
    "prescan_refresh_keys.csv": "快取_候選掃描Keys",
}


# ══════════════════════════════════════════════════════════════════════
# Supabase 快取同步工具（大型快取分流用）
# ══════════════════════════════════════════════════════════════════════

SUPABASE_CACHE_ENABLED = os.getenv("SUPABASE_CACHE_ENABLED", "0").strip().lower() in ("1", "true", "yes")
SUPABASE_DB_URL = os.getenv("SUPABASE_DB_URL", "").strip()
SUPABASE_SCHEMA = os.getenv("SUPABASE_SCHEMA", "warrant_cache").strip() or "warrant_cache"
SUPABASE_BATCH_SIZE = int(os.getenv("SUPABASE_BATCH_SIZE", "5000"))
# 啟用 Supabase 後，大型快取預設不再同步到 Google Sheet，避免 1,000 萬格上限。
# Excel / Google Sheet 結果表仍照原本流程同步，不受影響。
SUPABASE_CACHE_SKIP_GSHEET_SYNC = os.getenv("SUPABASE_CACHE_SKIP_GSHEET_SYNC", "1").strip().lower() not in ("0", "false", "no")

_SUPABASE_AVAILABLE = None
_SUPABASE_EMPTY_CACHE_KEYS = set()


def supabase_enabled():
    return bool(SUPABASE_CACHE_ENABLED and SUPABASE_DB_URL)


def supabase_import_psycopg():
    global _SUPABASE_AVAILABLE

    if not supabase_enabled():
        return None

    if _SUPABASE_AVAILABLE is False:
        return None

    try:
        import psycopg
        _SUPABASE_AVAILABLE = True
        return psycopg
    except Exception as e:
        _SUPABASE_AVAILABLE = False
        print(f"  ⚠️ Supabase 快取停用：無法載入 psycopg，原因：{type(e).__name__}: {e}")
        return None


def get_supabase_conn():
    psycopg = supabase_import_psycopg()
    if psycopg is None:
        return None

    try:
        return psycopg.connect(
            SUPABASE_DB_URL,
            connect_timeout=20,
            autocommit=True,
            prepare_threshold=None,
        )
    except Exception as e:
        print(f"  ⚠️ Supabase 連線失敗：{type(e).__name__}: {e}")
        return None


def supabase_cache_kind_and_scope(path):
    base = os.path.basename(str(path))

    if base == "price_cache.csv":
        return "price", ""
    if base == "broker_warrant_history_cache.csv":
        return "history", ""
    if base == "candidates_cache.csv":
        return "candidates", "all"
    if base == "candidates_cache_selected5.csv":
        return "candidates", "selected5"
    if base == "warrants_cache.csv":
        return "warrants", ""
    if base == "broker_map_cache.csv":
        return "broker_map", "selected5" if RUN_MODE == 1 else "all"

    return "", ""


def supabase_cache_supported(path):
    kind, _ = supabase_cache_kind_and_scope(path)
    return bool(kind)


def supabase_cache_identifier(path):
    kind, scope = supabase_cache_kind_and_scope(path)
    return f"{kind}:{scope}" if scope else kind


def supabase_should_skip_gsheet_cache(path):
    return bool(supabase_enabled() and SUPABASE_CACHE_SKIP_GSHEET_SYNC and supabase_cache_supported(path))


def _supabase_date_for_db(value):
    dt = parse_date(value)
    return dt.strftime("%Y-%m-%d") if dt else None


def _supabase_date_for_cache(value):
    if value is None:
        return ""
    try:
        if hasattr(value, "strftime"):
            return value.strftime("%Y/%m/%d")
    except Exception:
        pass
    return normalize_date_str(value)


def _supabase_int(value):
    try:
        if value is None:
            return 0
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return 0
        return int(float(s))
    except Exception:
        return 0


def _supabase_float(value):
    try:
        if value is None:
            return None
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return None
        v = float(s)
        return v if v > 0 else None
    except Exception:
        return None


def _supabase_text(value):
    if value is None:
        return ""
    return strip_gsheet_text_prefix(str(value)).strip()


def _supabase_table(name):
    # schema 與 table 名稱由程式固定產生，不吃使用者輸入值。
    return f"{SUPABASE_SCHEMA}.{name}"




def _supabase_sync_rows(sql, rows, label="", pre_statements=None):
    conn = get_supabase_conn()
    if conn is None:
        return False

    pre_statements = pre_statements or []
    rows = list(rows or [])

    try:
        batch_size = max(int(SUPABASE_BATCH_SIZE or 5000), 1)
        with conn.transaction():
            with conn.cursor() as cur:
                for pre_sql, pre_params in pre_statements:
                    cur.execute(pre_sql, pre_params or ())

                total = len(rows)
                for start in range(0, total, batch_size):
                    batch = rows[start:start + batch_size]
                    cur.executemany(sql, batch)
                    if total >= batch_size * 2:
                        print(f"  🗄️ Supabase 寫入中：{label} {min(start + len(batch), total):,}/{total:,}")
        return True
    except Exception as e:
        print(f"  ⚠️ Supabase 快取同步失敗：{label}，原因：{type(e).__name__}: {e}")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass


def prune_supabase_cache_tables():
    """依目前程式設定，直接清除 Supabase 中超期日期、已移除分點與失效候選。"""
    if not supabase_enabled() or not SUPABASE_AUTO_PRUNE_ENABLED or not CACHE_AUTO_PRUNE_ENABLED:
        return False

    conn = get_supabase_conn()
    if conn is None:
        return False

    all_labels = sorted(configured_broker_labels_for_scope("all"))
    all_codes = sorted(configured_broker_codes_for_scope("all"))
    selected_labels = sorted(configured_broker_labels_for_scope("selected5"))
    selected_codes = sorted(configured_broker_codes_for_scope("selected5"))

    try:
        with conn.transaction():
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    delete from {_supabase_table('broker_warrant_history')}
                    where trade_date < (
                        select min(trade_date)
                        from (
                            select distinct trade_date
                            from {_supabase_table('broker_warrant_history')}
                            order by trade_date desc
                            limit %s
                        ) recent_dates
                    )
                    """,
                    (HISTORY_RETENTION_TRADING_DAYS,),
                )
                history_date_deleted = max(cur.rowcount or 0, 0)

                cur.execute(
                    f"""
                    delete from {_supabase_table('broker_warrant_history')}
                    where not (
                        broker_label = any(%s::text[])
                        or broker_code = any(%s::text[])
                    )
                    """,
                    (all_labels, all_codes),
                )
                history_broker_deleted = max(cur.rowcount or 0, 0)

                cur.execute(
                    f"""
                    delete from {_supabase_table('price_cache')}
                    where trade_date < (
                        select min(trade_date)
                        from (
                            select distinct trade_date
                            from {_supabase_table('price_cache')}
                            order by trade_date desc
                            limit %s
                        ) recent_dates
                    )
                    """,
                    (PRICE_RETENTION_TRADING_DAYS,),
                )
                price_deleted = max(cur.rowcount or 0, 0)

                cur.execute(
                    f"""
                    delete from {_supabase_table('candidates')}
                    where scope = 'all'
                      and not (
                          broker_label = any(%s::text[])
                          or broker_code = any(%s::text[])
                      )
                    """,
                    (all_labels, all_codes),
                )
                candidates_all_deleted = max(cur.rowcount or 0, 0)

                cur.execute(
                    f"""
                    delete from {_supabase_table('candidates')}
                    where scope = 'selected5'
                      and not (
                          broker_label = any(%s::text[])
                          or broker_code = any(%s::text[])
                      )
                    """,
                    (selected_labels, selected_codes),
                )
                candidates_selected_deleted = max(cur.rowcount or 0, 0)

                if LIVE_WARRANT_SNAPSHOT_READY:
                    cur.execute(
                        f"""
                        delete from {_supabase_table('candidates')} c
                        where not exists (
                            select 1
                            from {_supabase_table('warrants')} w
                            where w.warrant_code = c.warrant_code
                        )
                        """
                    )
                    candidates_warrant_deleted = max(cur.rowcount or 0, 0)
                else:
                    candidates_warrant_deleted = 0

                cur.execute(
                    f"""
                    delete from {_supabase_table('broker_map')}
                    where scope = 'all'
                      and not (
                          broker_label = any(%s::text[])
                          or broker_code = any(%s::text[])
                      )
                    """,
                    (all_labels, all_codes),
                )
                broker_map_all_deleted = max(cur.rowcount or 0, 0)

                cur.execute(
                    f"""
                    delete from {_supabase_table('broker_map')}
                    where scope = 'selected5'
                      and not (
                          broker_label = any(%s::text[])
                          or broker_code = any(%s::text[])
                      )
                    """,
                    (selected_labels, selected_codes),
                )
                broker_map_selected_deleted = max(cur.rowcount or 0, 0)

        total_deleted = sum([
            history_date_deleted,
            history_broker_deleted,
            price_deleted,
            candidates_all_deleted,
            candidates_selected_deleted,
            candidates_warrant_deleted,
            broker_map_all_deleted,
            broker_map_selected_deleted,
        ])
        print(
            "  🧹 Supabase 自動清理完成："
            f"歷史超期 {history_date_deleted:,}｜歷史移除分點 {history_broker_deleted:,}｜"
            f"價格超期 {price_deleted:,}｜候選移除分點 {candidates_all_deleted + candidates_selected_deleted:,}｜"
            f"候選失效權證 {candidates_warrant_deleted:,}｜分點代號 {broker_map_all_deleted + broker_map_selected_deleted:,}｜"
            f"合計 {total_deleted:,} 筆"
        )
        return True
    except Exception as e:
        print(f"  ⚠️ Supabase 自動清理失敗：{type(e).__name__}: {e}")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass

def _supabase_fetch_dataframe(sql, params=None):
    conn = get_supabase_conn()
    if conn is None:
        return pd.DataFrame()

    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(sql, params or ())
                rows = cur.fetchall()
                headers = [getattr(desc, "name", desc[0]) for desc in cur.description] if cur.description else []
        if not rows or not headers:
            return pd.DataFrame()
        return pd.DataFrame(rows, columns=headers).fillna("")
    except Exception as e:
        print(f"  ⚠️ Supabase 快取讀取失敗：{type(e).__name__}: {e}")
        return pd.DataFrame()
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _supabase_executemany(sql, rows, label=""):
    if not rows:
        return False

    conn = get_supabase_conn()
    if conn is None:
        return False

    try:
        total = len(rows)
        batch_size = max(int(SUPABASE_BATCH_SIZE or 5000), 1)
        with conn:
            with conn.cursor() as cur:
                for start in range(0, total, batch_size):
                    batch = rows[start:start + batch_size]
                    cur.executemany(sql, batch)
                    if total >= batch_size * 2:
                        print(f"  🗄️ Supabase 寫入中：{label} {min(start + len(batch), total):,}/{total:,}")
        return True
    except Exception as e:
        print(f"  ⚠️ Supabase 快取寫入失敗：{label}，原因：{type(e).__name__}: {e}")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass


def read_cache_from_supabase(path):
    if not supabase_enabled() or not supabase_cache_supported(path):
        return pd.DataFrame()

    kind, scope = supabase_cache_kind_and_scope(path)
    cache_id = supabase_cache_identifier(path)

    if kind == "price":
        df = _supabase_fetch_dataframe(
            f"""
            with recent_dates as (
                select distinct trade_date
                from {_supabase_table('price_cache')}
                order by trade_date desc
                limit %s
            ), cutoff as (
                select min(trade_date) as min_date from recent_dates
            )
            select code as "代號",
                   to_char(trade_date, 'YYYY/MM/DD') as "日期",
                   close_price as "收盤價"
            from {_supabase_table('price_cache')}
            where trade_date >= (select min_date from cutoff)
            order by code, trade_date
            """,
            (PRICE_RETENTION_TRADING_DAYS,),
        )
    elif kind == "history":
        allowed_labels = sorted(configured_broker_labels_for_scope("all"))
        allowed_codes = sorted(configured_broker_codes_for_scope("all"))
        df = _supabase_fetch_dataframe(
            f"""
            with recent_dates as (
                select distinct trade_date
                from {_supabase_table('broker_warrant_history')}
                order by trade_date desc
                limit %s
            ), cutoff as (
                select min(trade_date) as min_date from recent_dates
            )
            select warrant_code as "權證代號",
                   warrant_name as "權證名稱",
                   underlying_code as "標的股",
                   underlying_name as "標的名稱",
                   broker_label as "分點",
                   broker_name as "分點名稱",
                   broker_code as "券商代號",
                   to_char(trade_date, 'YYYY/MM/DD') as "日期",
                   buy_shares as "買進股數",
                   sell_shares as "賣出股數",
                   buy_amount as "買進金額",
                   sell_amount as "賣出金額",
                   net_buy_shares as "買超股數",
                   net_buy_amount as "買超金額"
            from {_supabase_table('broker_warrant_history')}
            where trade_date >= (select min_date from cutoff)
              and (
                  broker_label = any(%s::text[])
                  or broker_code = any(%s::text[])
              )
            order by warrant_code, broker_code, trade_date
            """,
            (HISTORY_RETENTION_TRADING_DAYS, allowed_labels, allowed_codes),
        )
    elif kind == "candidates":
        allowed_labels = sorted(configured_broker_labels_for_scope(scope))
        allowed_codes = sorted(configured_broker_codes_for_scope(scope))
        valid_warrant_sql = ""
        if LIVE_WARRANT_SNAPSHOT_READY:
            valid_warrant_sql = f"""
              and exists (
                  select 1
                  from {_supabase_table('warrants')} w
                  where w.warrant_code = c.warrant_code
              )
            """
        df = _supabase_fetch_dataframe(
            f"""
            select c.warrant_code as "權證代號",
                   c.warrant_name as "權證名稱",
                   c.underlying_code as "標的股",
                   c.underlying_name as "標的名稱",
                   c.broker_label as "分點",
                   c.broker_name as "分點名稱",
                   c.broker_code as "券商代號"
            from {_supabase_table('candidates')} c
            where c.scope = %s
              and (
                  c.broker_label = any(%s::text[])
                  or c.broker_code = any(%s::text[])
              )
              {valid_warrant_sql}
            order by c.warrant_code, c.broker_code
            """,
            (scope, allowed_labels, allowed_codes),
        )
    elif kind == "warrants":
        df = _supabase_fetch_dataframe(
            f"""
            select warrant_code as "代號",
                   warrant_name as "名稱",
                   underlying_code as "標的股",
                   underlying_name as "標的名稱"
            from {_supabase_table('warrants')}
            order by warrant_code
            """
        )
    elif kind == "broker_map":
        allowed_labels = sorted(configured_broker_labels_for_scope(scope))
        allowed_codes = sorted(configured_broker_codes_for_scope(scope))
        df = _supabase_fetch_dataframe(
            f"""
            select broker_label as "分點",
                   broker_name as "分點名稱",
                   broker_code as "券商代號"
            from {_supabase_table('broker_map')}
            where scope = %s
              and (
                  broker_label = any(%s::text[])
                  or broker_code = any(%s::text[])
              )
            order by broker_label
            """,
            (scope, allowed_labels, allowed_codes),
        )
    else:
        return pd.DataFrame()

    if df is None or df.empty:
        _SUPABASE_EMPTY_CACHE_KEYS.add(cache_id)
        return pd.DataFrame()

    print(f"  🗄️ 已從 Supabase 讀取快取：{cache_sheet_name_from_path(path)}，共 {len(df):,} 筆")
    return df.fillna("")

def write_cache_to_supabase(df, path, full_df=None):
    if not supabase_enabled() or not supabase_cache_supported(path):
        return False

    kind, scope = supabase_cache_kind_and_scope(path)
    cache_id = supabase_cache_identifier(path)
    df2 = (df.copy() if df is not None else pd.DataFrame()).fillna("")
    full_df2 = (full_df.copy() if full_df is not None else df2.copy()).fillna("")
    pre_statements = []

    if kind == "history":
        full_df2 = fix_known_underlying_info_dataframe(full_df2, "權證名稱", "標的股", "標的名稱")
        full_df2, _ = prune_history_cache_dataframe(full_df2)
        df2 = fix_known_underlying_info_dataframe(df2, "權證名稱", "標的股", "標的名稱")
        df2, _ = prune_history_cache_dataframe(df2)
    elif kind == "candidates":
        full_df2 = fix_known_underlying_info_dataframe(full_df2, "權證名稱", "標的股", "標的名稱")
        valid_codes = CURRENT_LIVE_WARRANT_CODES if LIVE_WARRANT_SNAPSHOT_READY else None
        full_df2, _ = prune_candidates_dataframe(full_df2, path=path, valid_warrant_codes=valid_codes)
        df2 = full_df2.copy()
    elif kind == "warrants":
        full_df2 = fix_known_underlying_info_dataframe(full_df2, "名稱", "標的股", "標的名稱")
        df2 = full_df2.copy()
    elif kind == "price":
        full_df2, _ = prune_price_cache_dataframe(full_df2)
        df2, _ = prune_price_cache_dataframe(df2)
    elif kind == "broker_map":
        full_df2, _ = prune_broker_map_dataframe(full_df2, scope=scope)
        df2 = full_df2.copy()

    if kind == "price":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            code = normalize_price_code(row.get("代號", ""))
            trade_date = _supabase_date_for_db(row.get("日期", ""))
            close_price = _supabase_float(row.get("收盤價", ""))
            if code and trade_date and close_price is not None:
                rows.append((code, trade_date, close_price))

        if not full_df2.empty and "日期" in full_df2.columns:
            cutoff = recent_trading_date_cutoff_from_series(full_df2["日期"], PRICE_RETENTION_TRADING_DAYS)
            if cutoff:
                pre_statements.append((
                    f"delete from {_supabase_table('price_cache')} where trade_date < %s",
                    (cutoff,),
                ))

        sql = f"""
            insert into {_supabase_table('price_cache')} as pc (code, trade_date, close_price, updated_at)
            values (%s, %s, %s, now())
            on conflict (code, trade_date) do update set
                close_price = excluded.close_price,
                updated_at = now()
            where pc.close_price is distinct from excluded.close_price
        """
        label = "快取_價格"

    elif kind == "history":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("權證代號", ""))
            broker_code = _supabase_text(row.get("券商代號", ""))
            trade_date = _supabase_date_for_db(row.get("日期", ""))
            if not warrant_code or not broker_code or not trade_date:
                continue
            rows.append((
                warrant_code,
                _supabase_text(row.get("權證名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
                _supabase_text(row.get("分點", "")),
                _supabase_text(row.get("分點名稱", "")),
                broker_code,
                trade_date,
                _supabase_int(row.get("買進股數", 0)),
                _supabase_int(row.get("賣出股數", 0)),
                _supabase_int(row.get("買進金額", 0)),
                _supabase_int(row.get("賣出金額", 0)),
                _supabase_int(row.get("買超股數", 0)),
                _supabase_int(row.get("買超金額", 0)),
            ))

        if not full_df2.empty and "日期" in full_df2.columns:
            cutoff = recent_trading_date_cutoff_from_series(full_df2["日期"], HISTORY_RETENTION_TRADING_DAYS)
            if cutoff:
                pre_statements.append((
                    f"delete from {_supabase_table('broker_warrant_history')} where trade_date < %s",
                    (cutoff,),
                ))

        allowed_labels = sorted(configured_broker_labels_for_scope("all"))
        allowed_codes = sorted(configured_broker_codes_for_scope("all"))
        pre_statements.append((
            f"""
            delete from {_supabase_table('broker_warrant_history')}
            where not (
                broker_label = any(%s::text[])
                or broker_code = any(%s::text[])
            )
            """,
            (allowed_labels, allowed_codes),
        ))

        sql = f"""
            insert into {_supabase_table('broker_warrant_history')} as h
            (warrant_code, warrant_name, underlying_code, underlying_name,
             broker_label, broker_name, broker_code, trade_date,
             buy_shares, sell_shares, buy_amount, sell_amount,
             net_buy_shares, net_buy_amount, updated_at)
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now())
            on conflict (warrant_code, broker_code, trade_date) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                broker_label = excluded.broker_label,
                broker_name = excluded.broker_name,
                buy_shares = excluded.buy_shares,
                sell_shares = excluded.sell_shares,
                buy_amount = excluded.buy_amount,
                sell_amount = excluded.sell_amount,
                net_buy_shares = excluded.net_buy_shares,
                net_buy_amount = excluded.net_buy_amount,
                updated_at = now()
            where (
                h.warrant_name,
                h.underlying_code,
                h.underlying_name,
                h.broker_label,
                h.broker_name,
                h.buy_shares,
                h.sell_shares,
                h.buy_amount,
                h.sell_amount,
                h.net_buy_shares,
                h.net_buy_amount
            ) is distinct from (
                excluded.warrant_name,
                excluded.underlying_code,
                excluded.underlying_name,
                excluded.broker_label,
                excluded.broker_name,
                excluded.buy_shares,
                excluded.sell_shares,
                excluded.buy_amount,
                excluded.sell_amount,
                excluded.net_buy_shares,
                excluded.net_buy_amount
            )
        """
        label = "快取_分點歷史"

    elif kind == "candidates":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("權證代號", ""))
            broker_code = _supabase_text(row.get("券商代號", ""))
            if not warrant_code or not broker_code:
                continue
            rows.append((
                scope,
                warrant_code,
                _supabase_text(row.get("權證名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
                _supabase_text(row.get("分點", "")),
                _supabase_text(row.get("分點名稱", "")),
                broker_code,
            ))
        sql = f"""
            insert into {_supabase_table('candidates')} as c
            (scope, warrant_code, warrant_name, underlying_code, underlying_name,
             broker_label, broker_name, broker_code, updated_at)
            values (%s, %s, %s, %s, %s, %s, %s, %s, now())
            on conflict (scope, warrant_code, broker_code) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                broker_label = excluded.broker_label,
                broker_name = excluded.broker_name,
                updated_at = now()
            where (
                c.warrant_name,
                c.underlying_code,
                c.underlying_name,
                c.broker_label,
                c.broker_name
            ) is distinct from (
                excluded.warrant_name,
                excluded.underlying_code,
                excluded.underlying_name,
                excluded.broker_label,
                excluded.broker_name
            )
        """
        label = f"快取_候選組合({scope})"

    elif kind == "warrants":
        rows = []
        seen_codes = set()
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("代號", ""))
            if not warrant_code or warrant_code in seen_codes:
                continue
            seen_codes.add(warrant_code)
            rows.append((
                warrant_code,
                _supabase_text(row.get("名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
            ))
        pre_statements.append((f"delete from {_supabase_table('warrants')}", ()))
        sql = f"""
            insert into {_supabase_table('warrants')}
            (warrant_code, warrant_name, underlying_code, underlying_name, updated_at)
            values (%s, %s, %s, %s, now())
            on conflict (warrant_code) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                updated_at = now()
        """
        label = "快取_權證清單"

    elif kind == "broker_map":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            broker_label = _supabase_text(row.get("分點", ""))
            if not broker_label:
                continue
            rows.append((
                scope,
                broker_label,
                _supabase_text(row.get("分點名稱", "")),
                _supabase_text(row.get("券商代號", "")),
            ))
        pre_statements.append((
            f"delete from {_supabase_table('broker_map')} where scope = %s",
            (scope,),
        ))
        sql = f"""
            insert into {_supabase_table('broker_map')}
            (scope, broker_label, broker_name, broker_code, updated_at)
            values (%s, %s, %s, %s, now())
            on conflict (scope, broker_label) do update set
                broker_name = excluded.broker_name,
                broker_code = excluded.broker_code,
                updated_at = now()
        """
        label = f"快取_分點代號({scope})"
    else:
        return False

    ok = _supabase_sync_rows(sql, rows, label=label, pre_statements=pre_statements)
    if ok:
        _SUPABASE_EMPTY_CACHE_KEYS.discard(cache_id)
        print(f"  🗄️ 已同步快取到 Supabase：{label}，本次寫入 {len(rows):,} 筆")
    return ok

def get_gcp_service_key():
    return os.getenv("GCP_SERVICE_KEY", os.getenv("GCP_SERVICE_KEY_NEW", "")).strip()


def gsheet_enabled():
    return bool(get_gcp_service_key())


def is_gsheet_quota_error(exc):
    msg = str(exc)
    return (
        "429" in msg
        or "Quota exceeded" in msg
        or "RESOURCE_EXHAUSTED" in msg
        or "Write requests per minute" in msg
    )


def gsheet_write_sleep():
    """
    Google Sheets 寫入節流。

    這不是改資料邏輯，而是避免短時間連續建立 / 清除 / 寫入 / 套格式
    造成 429 quota exceeded。
    """
    global _GSHEET_LAST_WRITE_TS

    if GSHEET_WRITE_SLEEP_SECONDS <= 0:
        return

    now = time.time()
    elapsed = now - _GSHEET_LAST_WRITE_TS

    if elapsed < GSHEET_WRITE_SLEEP_SECONDS:
        time.sleep(GSHEET_WRITE_SLEEP_SECONDS - elapsed)

    _GSHEET_LAST_WRITE_TS = time.time()


def gsheet_api_call(description, func, *args, **kwargs):
    """
    所有 Google Sheet 寫入動作統一走這裡：
    1. 先節流
    2. 遇到 429 自動等待重試
    3. 不讓暫時 quota 造成後續工作表消失
    """
    last_error = None

    for attempt in range(1, GSHEET_MAX_RETRIES + 1):
        try:
            gsheet_write_sleep()
            return func(*args, **kwargs)
        except Exception as e:
            last_error = e

            if not is_gsheet_quota_error(e):
                raise

            wait_seconds = min(90, GSHEET_RETRY_BASE_SECONDS * attempt)
            print(
                f"  ⚠️ Google Sheet 觸發寫入配額限制，{description} 第 {attempt}/{GSHEET_MAX_RETRIES} 次重試，"
                f"等待 {wait_seconds:.0f} 秒..."
            )
            time.sleep(wait_seconds)

    raise last_error


def get_gsheet_client():
    global _GSHEET_CLIENT

    if _GSHEET_CLIENT is not None:
        return _GSHEET_CLIENT

    service_key = get_gcp_service_key()

    if not service_key:
        return None

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        info = json.loads(service_key)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        _GSHEET_CLIENT = gspread.authorize(creds)
        return _GSHEET_CLIENT
    except Exception as e:
        print(f"  ⚠️ Google Sheet 金鑰初始化失敗：{type(e).__name__}: {e}")
        return None


def get_gsheet_spreadsheet():
    global _GSHEET_SPREADSHEET

    if _GSHEET_SPREADSHEET is not None:
        return _GSHEET_SPREADSHEET

    gc = get_gsheet_client()

    if gc is None:
        return None

    try:
        if GOOGLE_SHEET_ID:
            _GSHEET_SPREADSHEET = gc.open_by_key(GOOGLE_SHEET_ID)
        else:
            _GSHEET_SPREADSHEET = gc.open(GOOGLE_SHEET_NAME)
        return _GSHEET_SPREADSHEET
    except Exception as e:
        try:
            _GSHEET_SPREADSHEET = gc.create(GOOGLE_SHEET_NAME)
            print(f"  ✅ 已建立 Google Sheet：{GOOGLE_SHEET_NAME}")
            return _GSHEET_SPREADSHEET
        except Exception as e2:
            print(f"  ⚠️ Google Sheet 開啟/建立失敗：{type(e).__name__}: {e} / {type(e2).__name__}: {e2}")
            return None



def spreadsheet_grid_cell_count(spreadsheet):
    """計算整份 Google 試算表目前配置的格數（包含空白格）。"""
    if spreadsheet is None:
        return 0

    total = 0
    try:
        for ws in spreadsheet.worksheets():
            total += max(int(getattr(ws, "row_count", 1) or 1), 1) * max(
                int(getattr(ws, "col_count", 1) or 1),
                1,
            )
    except Exception:
        return total

    return total


def _worksheet_values_with_formulas(ws):
    """
    讀取工作表實際內容，優先保留公式字串。

    使用 FORMULA 可避免公式結果剛好是空字串時，被誤判成多餘空白列而刪除。
    """
    if ws is None:
        return []

    try:
        return ws.get_all_values(value_render_option="FORMULA") or []
    except TypeError:
        try:
            return ws.get_all_values() or []
        except Exception:
            return []
    except Exception:
        try:
            return ws.get_all_values() or []
        except Exception:
            return []


def worksheet_used_grid_size(ws):
    """回傳工作表最後一個有值或有公式的實際列數 / 欄數。"""
    values = _worksheet_values_with_formulas(ws)
    last_row = 0
    last_col = 0

    for row_idx, row in enumerate(values, start=1):
        row_last_col = 0
        for col_idx, value in enumerate(list(row), start=1):
            if str(value).strip() != "":
                row_last_col = col_idx

        if row_last_col > 0:
            last_row = row_idx
            last_col = max(last_col, row_last_col)

    return (
        max(last_row, GSHEET_COMPACT_MIN_ROWS),
        max(last_col, GSHEET_COMPACT_MIN_COLS),
    )


def compact_spreadsheet_blank_grid(spreadsheet=None, label="主試算表"):
    """
    把每張工作表縮到實際有內容 / 公式的列數與欄數。

    Google Sheets 的 1,000 萬格上限會把空白預留列與欄也算進去，因此不能只 clear，
    必須真的 resize 才能釋放格數。
    """
    if not GSHEET_COMPACT_BLANK_GRID_ENABLED:
        return False

    spreadsheet = spreadsheet or get_gsheet_spreadsheet()
    if spreadsheet is None:
        return False

    before_cells = spreadsheet_grid_cell_count(spreadsheet)
    changed_count = 0

    try:
        worksheets = spreadsheet.worksheets()
    except Exception as e:
        print(f"  ⚠️ Google Sheet 空白格清理失敗：{label}，原因：{type(e).__name__}: {e}")
        return False

    for ws in worksheets:
        try:
            used_rows, used_cols = worksheet_used_grid_size(ws)
            current_rows = max(int(getattr(ws, "row_count", 1) or 1), 1)
            current_cols = max(int(getattr(ws, "col_count", 1) or 1), 1)

            if current_rows == used_rows and current_cols == used_cols:
                continue

            gsheet_api_call(
                f"縮減空白列欄 {label}/{ws.title}",
                ws.resize,
                rows=used_rows,
                cols=used_cols,
            )
            changed_count += 1
            print(
                f"  🧹 已縮減工作表空白格：{ws.title}｜"
                f"{current_rows:,}×{current_cols:,} → {used_rows:,}×{used_cols:,}"
            )
        except Exception as e:
            print(f"  ⚠️ 工作表空白格縮減失敗：{getattr(ws, 'title', '-')}，原因：{type(e).__name__}: {e}")

    after_cells = spreadsheet_grid_cell_count(spreadsheet)
    if changed_count > 0:
        print(
            f"  🧹 Google Sheet 空白列欄清理完成：{label}｜"
            f"{before_cells:,} 格 → {after_cells:,} 格｜釋放 {max(before_cells - after_cells, 0):,} 格"
        )

    return changed_count > 0

def safe_worksheet_title(title):
    title = str(title).strip()
    bad_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in bad_chars:
        title = title.replace(ch, "_")
    return title[:100] if title else "工作表"


def cache_sheet_name_from_path(path):
    base = os.path.basename(str(path))
    return CACHE_SHEET_NAME_MAP.get(base, safe_worksheet_title(f"快取_{os.path.splitext(base)[0]}"))


def get_or_create_worksheet(title, rows=100, cols=20):
    sh = get_gsheet_spreadsheet()

    if sh is None:
        return None

    title = safe_worksheet_title(title)

    try:
        return sh.worksheet(title)
    except Exception:
        try:
            return gsheet_api_call(
                f"建立工作表 {title}",
                sh.add_worksheet,
                title=title,
                rows=max(int(rows), 1),
                cols=max(int(cols), 1),
            )
        except Exception as e:
            print(f"  ⚠️ 建立工作表失敗：{title}，原因：{type(e).__name__}: {e}")
            return None


def get_or_recreate_result_worksheet(title, rows=100, cols=20):
    """
    取得結果工作表。

    重要修正：
    以前為了清掉舊格式，會先刪除再重建結果工作表。
    但 Google Sheets API 有每分鐘寫入限制，一旦刪除後重建遇到 429，
    會導致「券商查詢」、「ABCDE組合勝率」等工作表直接消失。

    這版改成不刪除工作表：
    1. 已存在就沿用
    2. 不存在才建立
    3. 寫入前會清除舊格式與資料，再重新寫入與套樣式
    """
    sh = get_gsheet_spreadsheet()

    if sh is None:
        return None

    title = safe_worksheet_title(title)
    rows = max(int(rows), 1)
    cols = max(int(cols), 1)

    try:
        return sh.worksheet(title)
    except Exception:
        try:
            return gsheet_api_call(
                f"建立結果工作表 {title}",
                sh.add_worksheet,
                title=title,
                rows=rows,
                cols=cols,
            )
        except Exception as e:
            print(f"  ⚠️ 建立結果工作表失敗：{title}，原因：{type(e).__name__}: {e}")
            return None


GSHEET_TEXT_HEADER_KEYWORDS = (
    "權證代號",
    "權證代碼",
    "權證清單",
    "券商代號",
    "券商代碼",
    "股票代號",
    "股票代碼",
    "標的股",
    "標的代號",
    "標的代碼",
    "證券代號",
    "證券代碼",
    "商品代號",
    "商品代碼",
    "代號",
)


def clean_gsheet_value(value):
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        return value

    return str(value)


def is_gsheet_text_header(header):
    header = str(header).strip()
    return any(keyword in header for keyword in GSHEET_TEXT_HEADER_KEYWORDS)


def _strip_trailing_excel_decimal_for_code(s):
    s = str(s).strip()

    if s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            return head

    return s


def normalize_gsheet_code_text_value(header, value):
    """
    Google Sheet 寫入 / 讀回時的代號欄位修正。

    目的：
    1. 權證代號若被 Google Sheet 轉成 5 碼，例如 30004，補回 030004。
    2. 權證清單內的每一個 5 碼權證代號也補回 6 碼。
    3. 股票代號 / 標的股 / 券商代號只維持文字，不任意補零，避免破壞原代號。
    4. 公式欄位不可處理，避免 =IFERROR(...) 被改成純文字。
    """
    if value is None:
        return ""

    s = str(value).strip()

    if s == "":
        return ""

    if s.startswith("="):
        return s

    had_prefix = s.startswith("'")
    if had_prefix:
        s = s[1:]

    header = str(header).strip()
    s = _strip_trailing_excel_decimal_for_code(s)

    if "權證清單" in header:
        # 權證清單常見格式：30004 海華國票...；60234 文華統一...
        # 若 Google Sheet 已經把 030004 轉成 30004，這裡將每段開頭的 5 碼權證補回 6 碼。
        def repl(match):
            token = match.group(0)
            return token.zfill(6) if token.isdigit() and len(token) == 5 else token

        s = re.sub(r"(?<!\d)\d{5}(?!\d)", repl, s)
        return ("'" + s) if had_prefix else s

    if "權證代號" in header or "權證代碼" in header:
        if s.isdigit() and len(s) == 5:
            s = s.zfill(6)
        return ("'" + s) if had_prefix else s

    # 快取_權證清單 的欄位名稱是「代號」，價格快取也可能是「代號」。
    # 只有 5 碼純數字才視為權證前導 0 被吃掉，補回 6 碼；4 碼股票不補。
    if header in ("代號", "證券代號", "證券代碼", "商品代號", "商品代碼"):
        if s.isdigit() and len(s) == 5:
            s = s.zfill(6)
        return ("'" + s) if had_prefix else s

    return ("'" + s) if had_prefix else s


def add_gsheet_text_prefix(value):
    """
    讓 Google Sheet 用文字格式寫入代號欄位。

    若使用 USER_ENTERED 寫入 064390，Google Sheet 可能會自動轉成數字 64390；
    對權證代號 / 券商代號欄位加上前導單引號，可保留開頭 0。
    Google Sheet 顯示時不會顯示這個單引號，只會把儲存格視為文字。

    注意：如果儲存格內容是公式，不能加單引號，否則 Google Sheet 會把公式當成純文字顯示。
    例如「券商查詢」工作表的前 15 名查詢欄位就是公式欄位，必須保持 =IFERROR(...) 可計算。
    """
    if value is None:
        return ""

    s = str(value).strip()

    if s == "":
        return ""

    if s.startswith("="):
        return s

    if s.startswith("'"):
        return s

    return "'" + s


def strip_gsheet_text_prefix(value):
    if value is None:
        return ""

    s = str(value)

    if s.startswith("'"):
        return s[1:]

    return s


def normalize_gsheet_values_for_text_columns(values):
    """
    依據表頭自動判斷權證代號 / 券商代號 / 代號欄位，
    寫入 Google Sheet 前強制改成文字，避免前導 0 被吃掉。

    支援：
    1. 快取工作表：表頭通常在第 1 列。
    2. 一般結果工作表：表頭多數也在第 1 列。
    3. 例如 ABCDE組合勝率 這種表頭在第 4 列的工作表，會掃前 10 列找表頭。
    """
    if not values:
        return values

    out = [list(row) for row in values]
    header_rows = []

    scan_limit = min(len(out), 10)
    for row_idx in range(scan_limit):
        row = out[row_idx]
        text_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_text_header(header):
                text_cols.append(col_idx)

        if text_cols:
            header_rows.append((row_idx, text_cols))

    if not header_rows:
        return out

    for header_row_idx, text_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(out)

        for row_idx in range(header_row_idx + 1, end_row):
            for col_idx in text_cols:
                if col_idx >= len(out[row_idx]):
                    continue

                cell_value = out[row_idx][col_idx]
                if isinstance(cell_value, str) and cell_value.strip().startswith("="):
                    continue

                header = out[header_row_idx][col_idx] if col_idx < len(out[header_row_idx]) else ""
                cell_value = normalize_gsheet_code_text_value(header, cell_value)
                out[row_idx][col_idx] = add_gsheet_text_prefix(cell_value)

    return out


def collect_gsheet_text_column_ranges(values):
    """
    根據工作表前 10 列的表頭，找出需要強制使用「純文字」格式的欄位範圍。

    這裡除了「權證代號 / 券商代號 / 代號」之外，也包含「權證清單」、
    「股票代號」、「標的股」等欄位，避免 Google Sheet 把 0 開頭的代號自動轉成數字。
    """
    ranges = []

    if not values:
        return ranges

    scan_limit = min(len(values), 10)
    header_rows = []

    for row_idx in range(scan_limit):
        row = list(values[row_idx])
        text_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_text_header(header):
                text_cols.append(col_idx)

        if text_cols:
            header_rows.append((row_idx, text_cols))

    for header_row_idx, text_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(values)

        if end_row <= header_row_idx + 1:
            continue

        for col_idx in text_cols:
            ranges.append({
                "start_row": header_row_idx + 1,
                "end_row": end_row,
                "start_col": col_idx,
                "end_col": col_idx + 1,
            })

    return ranges


def apply_text_format_to_gsheet(gws, values):
    """
    將 Google Sheet 中的代號相關欄位套用純文字格式。

    write_values_to_worksheet() 已經會在代號欄位前加單引號，這裡再補上
    Google Sheets 的 TEXT numberFormat，雙重避免權證代號、股票代號、券商代號
    或權證清單中的 0 開頭代號被吃掉。

    注意：「券商查詢」工作表的資料列是公式查詢結果，不可把公式欄位套成純文字，
    否則 Google Sheet 會顯示 =IFERROR(...) 文字而不是計算結果。
    權證代號與權證清單的文字格式會保留在「券商查詢資料」隱藏工作表中。
    """
    if gws is None or not values:
        return

    if str(getattr(gws, "title", "")).strip() == "券商查詢":
        return

    requests = []

    for r in collect_gsheet_text_column_ranges(values):
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": gws.id,
                    "startRowIndex": r["start_row"],
                    "endRowIndex": r["end_row"],
                    "startColumnIndex": r["start_col"],
                    "endColumnIndex": r["end_col"],
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "TEXT"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat",
            }
        })

    _gsheet_batch_update(requests)



GSHEET_COMMA_NUMBER_HEADER_KEYWORDS = (
    "排名",
    "金額",
    "股數",
    "張數",
    "筆數",
    "事件數",
    "權證檔數",
    "買進分點數",
    "涵蓋權證數",
    "持有天數",
    "價格筆數",
    "成本",
    "市值",
    "損益",
)

GSHEET_COMMA_NUMBER_EXCLUDE_KEYWORDS = (
    "代號",
    "日期",
    "名稱",
    "清單",
    "類型",
    "狀態",
    "均價",
    "勝率",
    "占比",
    "比例",
    "報酬%",
    "獲利%",
)

GSHEET_DECIMAL_NUMBER_HEADER_KEYWORDS = (
    "均價",
    "收盤價",
    "最新權證價格",
    "權證價格",
    "股價",
    "報酬率",
    "報酬%",
    "獲利%",
    "勝率",
    "占比",
    "比例",
    "平均",
)

GSHEET_DECIMAL_NUMBER_EXCLUDE_KEYWORDS = (
    "日期",
    "代號",
    "名稱",
    "清單",
    "類型",
    "狀態",
    "文字",
    "說明",
    "來源",
)


def is_gsheet_percent_header(header):
    header = str(header).strip()

    if not header:
        return False

    if "文字" in header or "日期" in header or "代號" in header:
        return False

    return ("%" in header) or ("勝率" in header) or ("占比" in header) or ("比例" in header)


def is_gsheet_decimal_number_header(header):
    """
    判斷 Google Sheet 中需要數字格式但不一定要整數千分位的欄位。

    主要修正：
    1. 買進均價 / 減碼均價 / 出清均價不可被誤套日期格式。
    2. 減碼獲利% / 出清獲利% 不可被誤套日期格式。
    3. 報酬率、勝率、占比等欄位要保持數字或百分比顯示。
    """
    header = str(header).strip()

    if not header:
        return False

    if header.startswith("D+"):
        return False

    if is_gsheet_text_header(header):
        return False

    for keyword in GSHEET_DECIMAL_NUMBER_EXCLUDE_KEYWORDS:
        if keyword in header:
            return False

    return any(keyword in header for keyword in GSHEET_DECIMAL_NUMBER_HEADER_KEYWORDS)


def is_gsheet_numeric_format_header(header):
    return is_gsheet_comma_number_header(header) or is_gsheet_decimal_number_header(header)


def is_gsheet_comma_number_header(header):
    """
    判斷 Google Sheet 結果工作表中，哪些欄位要用千分位逗號顯示。

    注意：
    1. 權證代號 / 券商代號 / 代號欄位不能套數字格式，否則開頭 0 會被吃掉。
    2. 日期、名稱、清單、百分比、均價等欄位不套用千分位。
    3. 這個函式只用在「結果工作表」同步，不改快取工作表邏輯。
    """
    header = str(header).strip()

    if not header:
        return False

    if header.startswith("D+"):
        return False

    if is_gsheet_text_header(header):
        return False

    for keyword in GSHEET_COMMA_NUMBER_EXCLUDE_KEYWORDS:
        if keyword in header:
            return False

    return any(keyword in header for keyword in GSHEET_COMMA_NUMBER_HEADER_KEYWORDS)


def _parse_comma_number_for_gsheet(value):
    """
    將 1,234 / 12,345.67 這類字串轉成數字，讓 Google Sheet 可以搭配 numberFormat 顯示逗號。
    公式、空值、百分比、文字說明都保持原樣。
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return value

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return value

    s = str(value).strip()

    if s == "" or s == "-":
        return value

    if s.startswith("="):
        return value

    if "%" in s or "\n" in s or "；" in s:
        return value

    raw = s.replace(",", "").strip()

    if raw.startswith("+"):
        raw = raw[1:]

    if raw.startswith("-"):
        sign = "-"
        num_part = raw[1:]
    else:
        sign = ""
        num_part = raw

    if not num_part:
        return value

    if num_part.replace(".", "", 1).isdigit():
        try:
            if "." in num_part:
                return float(sign + num_part)
            return int(sign + num_part)
        except Exception:
            return value

    return value


def normalize_result_values_for_comma_numbers(values):
    """
    Google Sheet 結果同步前，將需要千分位的欄位轉成數字。

    原因：若直接把 "1,234" 用 USER_ENTERED 寫入 Google Sheet，可能被解析成數字但顯示為 1234，
    或在部分語系下變成文字。這裡先依表頭把金額 / 股數 / 張數 / 筆數等欄位轉成數字，
    後續再用 Google Sheets numberFormat 套 #,##0，確保畫面會顯示逗號。
    """
    if not values:
        return values

    out = [list(row) for row in values]
    header_rows = []

    scan_limit = min(len(out), 10)
    for row_idx in range(scan_limit):
        row = out[row_idx]
        number_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_numeric_format_header(header):
                number_cols.append(col_idx)

        if number_cols:
            header_rows.append((row_idx, number_cols))

    if not header_rows:
        return out

    for header_row_idx, number_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(out)

        for row_idx in range(header_row_idx + 1, end_row):
            for col_idx in number_cols:
                if col_idx >= len(out[row_idx]):
                    continue

                out[row_idx][col_idx] = _parse_comma_number_for_gsheet(out[row_idx][col_idx])

    return out


def _gsheet_number_format_for_header(header):
    header = str(header).strip()

    # 重要：Excel / Google Sheet 用 USER_ENTERED 寫入「+7.00%」這種字串時，
    # 儲存格底層值會變成 0.07。這類欄位必須套用真正的 PERCENT 格式，
    # 才會顯示成 +7.00%，不能用 NUMBER + 文字百分號，否則會被顯示成 +0.07%。
    #
    # 但「報酬率」這種沒有 % 符號的 TOP15 快取欄位，程式內多為 70.30 這種百分點數值，
    # 所以不歸類為 PERCENT，仍以一般數字顯示，避免變成 7,030%。
    if is_gsheet_percent_header(header):
        # 勝率 / 占比 / 比例不是報酬率，不應該顯示 + 號。
        if "勝率" in header or "占比" in header or "比例" in header:
            return {
                "type": "PERCENT",
                "pattern": '0.00%',
            }

        return {
            "type": "PERCENT",
            "pattern": '+0.00%;-0.00%;0.00%',
        }

    if is_gsheet_decimal_number_header(header):
        return {
            "type": "NUMBER",
            "pattern": "#,##0.00",
        }

    return {
        "type": "NUMBER",
        "pattern": "#,##0",
    }


def _gsheet_number_pattern_for_header(header):
    # 保留舊函式名稱供既有呼叫相容；實際格式型別請用 _gsheet_number_format_for_header()。
    return _gsheet_number_format_for_header(header).get("pattern", "#,##0")


def _format_source_rows_from_values_or_xlsx(ws_xlsx, values=None):
    """
    Google Sheet 格式套用時的表頭來源。

    upsert 模式會在 Google Sheet 寫入前新增「資料範圍」欄位，
    因此不能再只看原本 Excel 的欄位位置，否則日期格式會往左/往右套錯，
    造成買進均價、買超張數、獲利% 被顯示成 1900 年代日期。
    """
    if values:
        rows = [list(row) for row in values]
        max_row = max(len(rows), 1)
        max_col = max(max((len(row) for row in rows), default=1), 1)
        return rows, max_row, max_col

    rows = []
    if ws_xlsx is not None:
        scan_limit = min(ws_xlsx.max_row, 10)
        for row_idx in range(1, scan_limit + 1):
            rows.append([ws_xlsx.cell(row_idx, col_idx).value for col_idx in range(1, ws_xlsx.max_column + 1)])
        return rows, ws_xlsx.max_row, ws_xlsx.max_column

    return [], 1, 1


def apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=None):
    """
    將 Google Sheet 結果工作表的金額 / 股數 / 張數 / 成本 / 均價 / 獲利% 等欄位套用正確數字格式。

    重點修正：
    - upsert 後的實際 Google Sheet 表頭可能比原 Excel 多「資料範圍」欄。
    - 這裡改用實際寫入的 values 找欄位位置，避免格式位移。
    - 所有金額 / 成本 / 市值 / 損益使用千分位逗號顯示。
    - 均價、報酬率、獲利%、勝率等欄位也強制套數字/百分比格式，避免被日期格式污染。
    """
    if gws is None:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    source_rows, max_row, max_col = _format_source_rows_from_values_or_xlsx(ws_xlsx, values)
    header_rows = []
    scan_limit = min(len(source_rows), 10)

    for row_idx in range(1, scan_limit + 1):
        row = source_rows[row_idx - 1]
        number_cols = []

        for col_idx in range(1, max_col + 1):
            header = row[col_idx - 1] if col_idx - 1 < len(row) else ""

            if is_gsheet_numeric_format_header(header):
                number_cols.append((col_idx, _gsheet_number_pattern_for_header(header)))

        if number_cols:
            header_rows.append((row_idx, number_cols))

    if not header_rows:
        return

    requests = []

    for idx, (header_row_idx, number_cols) in enumerate(header_rows):
        if idx + 1 < len(header_rows):
            end_row = header_rows[idx + 1][0] - 1
        else:
            end_row = max_row

        start_data_row = header_row_idx + 1

        if start_data_row > end_row:
            continue

        for col_idx, pattern in number_cols:
            header = ""
            try:
                row = source_rows[header_row_idx - 1]
                header = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            except Exception:
                header = ""

            number_format = _gsheet_number_format_for_header(header)

            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_data_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": col_idx - 1,
                        "endColumnIndex": col_idx,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": number_format
                        }
                    },
                    "fields": "userEnteredFormat.numberFormat",
                }
            })

    _gsheet_batch_update(requests)


GSHEET_DATE_HEADER_KEYWORDS = (
    "日期",
    "買進日",
    "賣出日",
    "事件日",
    "起始日",
    "結束日",
    "減碼日",
    "出清日",
    "最近買進日",
    "第一筆日期",
    "最後筆日期",
)


def is_gsheet_date_header(header):
    """
    判斷 Google Sheet 中哪些欄位應該以日期格式顯示。

    主要修正「券商查詢」工作表的「最近買進日」：
    Google Sheet 公式 INDEX/MATCH 從資料表抓到日期時，常會以日期序號顯示，
    例如 46160。這裡統一把日期欄位套成 yyyy/mm/dd，讓畫面顯示正常日期。
    """
    header = str(header).strip()

    if not header:
        return False

    if "天數" in header:
        return False

    return any(keyword in header for keyword in GSHEET_DATE_HEADER_KEYWORDS)


def apply_date_format_to_gsheet(ws_xlsx, gws, values=None):
    """
    將日期相關欄位套用 Google Sheets 日期格式 yyyy/mm/dd。

    upsert 模式會新增「資料範圍」欄位，因此日期欄位置必須以實際寫入 values 的表頭判斷，
    不可只看原本 Excel 欄位位置。這可避免買進均價 / 張數 / 獲利% 被誤套成日期格式。
    """
    if gws is None:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    source_rows, max_row, max_col = _format_source_rows_from_values_or_xlsx(ws_xlsx, values)
    header_rows = []
    scan_limit = min(len(source_rows), 10)

    for row_idx in range(1, scan_limit + 1):
        row = source_rows[row_idx - 1]
        date_cols = []

        for col_idx in range(1, max_col + 1):
            header = row[col_idx - 1] if col_idx - 1 < len(row) else ""

            if is_gsheet_date_header(header):
                date_cols.append(col_idx)

        if date_cols:
            header_rows.append((row_idx, date_cols))

    if not header_rows:
        return

    requests = []

    for idx, (header_row_idx, date_cols) in enumerate(header_rows):
        if idx + 1 < len(header_rows):
            end_row = header_rows[idx + 1][0] - 1
        else:
            end_row = max_row

        start_data_row = header_row_idx + 1

        if start_data_row > end_row:
            continue

        for col_idx in date_cols:
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_data_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": col_idx - 1,
                        "endColumnIndex": col_idx,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": {
                                "type": "DATE",
                                "pattern": "yyyy/mm/dd",
                            }
                        }
                    },
                    "fields": "userEnteredFormat.numberFormat",
                }
            })

    _gsheet_batch_update(requests)


def _gsheet_pixel_width_for_header(header):
    header = str(header).strip()

    if not header:
        return None

    if header == "資料範圍":
        return 105

    # 有權證資訊的欄位加寬，避免新增「資料範圍」後欄寬位移造成內容被擠在一起。
    if "權證清單" in header or "權證集合" in header:
        return 520
    if "分點明細_JSON" in header:
        return 700
    if "權證名稱" in header:
        return 190
    if "權證代號" in header or "權證代碼" in header:
        return 105
    if "權證檔數" in header:
        return 95

    if "標的名稱" in header or header == "股票名稱":
        return 140
    if "標的股" in header or "標的代號" in header:
        return 95
    if "分點名稱" in header or header == "分點":
        return 140
    if "券商代號" in header:
        return 95

    if "日期" in header or header in ("買進日", "事件日", "起始日", "結束日", "減碼日", "出清日"):
        return 105

    if any(k in header for k in ("金額", "成本", "市值", "損益")):
        return 125

    if any(k in header for k in ("均價", "報酬率", "獲利%", "勝率", "占比", "比例")):
        return 105

    if any(k in header for k in ("股數", "張數", "筆數", "天數", "排名")):
        return 85

    return None


def apply_header_widths_to_gsheet(gws, values=None):
    """
    依照實際 Google Sheet 表頭重新調整欄寬。

    因為 upsert 會新增「資料範圍」欄位，不能只沿用原 Excel 欄寬；
    否則有權證名稱 / 權證清單的欄位會被擠到太窄。
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    header_row_idx = _find_simple_header_row(values)
    if header_row_idx is None or header_row_idx >= len(values):
        return

    headers = list(values[header_row_idx])
    requests = []

    for col_idx, header in enumerate(headers):
        pixel_size = _gsheet_pixel_width_for_header(header)
        if not pixel_size:
            continue

        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex": col_idx + 1,
                },
                "properties": {
                    "pixelSize": pixel_size,
                },
                "fields": "pixelSize",
            }
        })

    _gsheet_batch_update(requests)


def reset_worksheet_before_value_write(ws, row_count, col_count):
    """
    寫入值之前先清掉舊格式、舊資料驗證與舊合併範圍。

    這可以解決「券商查詢」公式被舊的純文字格式吃掉，
    同時避免因刪除 / 重建工作表造成 429 後工作表消失。
    """
    if ws is None:
        return

    try:
        sheet_id = int(ws.id)
    except Exception:
        return

    requests = [
        {
            "unmergeCells": {
                "range": {
                    "sheetId": sheet_id,
                }
            }
        },
        {
            "updateCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": max(int(row_count), 1),
                    "startColumnIndex": 0,
                    "endColumnIndex": max(int(col_count), 1),
                },
                "fields": "userEnteredFormat,dataValidation",
            }
        },
    ]

    _gsheet_batch_update(requests)


def write_values_to_worksheet(ws, values):
    if ws is None:
        return False

    if not values:
        values = [[""]]

    row_count = max(len(values), 1)
    col_count = max(max((len(row) for row in values), default=1), 1)

    normalized_values = []
    for row in values:
        row = list(row)
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        normalized_values.append([clean_gsheet_value(v) for v in row])

    normalized_values = normalize_gsheet_values_for_text_columns(normalized_values)

    try:
        # 不再使用 delete/recreate。先 resize，再清格式，最後寫入值。
        # 這樣遇到 429 時不會把既有工作表刪掉。
        gsheet_api_call(
            f"調整工作表大小 {ws.title}",
            ws.resize,
            rows=max(row_count, 1),
            cols=max(col_count, 1),
        )

        reset_worksheet_before_value_write(ws, row_count, col_count)

        for start in range(0, len(normalized_values), GSHEET_CHUNK_ROWS):
            chunk = normalized_values[start:start + GSHEET_CHUNK_ROWS]
            start_row = start + 1
            cell_range = f"A{start_row}"
            gsheet_api_call(
                f"寫入工作表資料 {ws.title} A{start_row}",
                ws.update,
                values=chunk,
                range_name=cell_range,
                value_input_option="USER_ENTERED",
            )

        apply_text_format_to_gsheet(ws, normalized_values)

        return True
    except Exception as e:
        print(f"  ⚠️ Google Sheet 寫入失敗：{ws.title}，原因：{type(e).__name__}: {e}")
        return False


def read_cache_from_gsheet(path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return pd.DataFrame()

    title = cache_sheet_name_from_path(path)

    try:
        sh = get_gsheet_spreadsheet()
        if sh is None:
            return pd.DataFrame()

        ws = sh.worksheet(title)
        values = ws.get_all_values()

        if not values or len(values) < 2:
            return pd.DataFrame()

        headers = [str(h).strip() for h in values[0]]
        rows = values[1:]

        if not headers or all(h == "" for h in headers):
            return pd.DataFrame()

        fixed_rows = []
        n_cols = len(headers)
        for row in rows:
            row = list(row)
            if len(row) < n_cols:
                row = row + [""] * (n_cols - len(row))
            elif len(row) > n_cols:
                row = row[:n_cols]
            fixed_rows.append(row)

        df = pd.DataFrame(fixed_rows, columns=headers).fillna("")

        for col in df.columns:
            if is_gsheet_text_header(col):
                df[col] = df[col].map(strip_gsheet_text_prefix)
                df[col] = df[col].map(lambda v, _col=col: normalize_gsheet_code_text_value(_col, v))

        print(f"  ☁️ 已從 Google Sheet 讀取快取：{title}，共 {len(df):,} 筆")
        return df
    except Exception:
        return pd.DataFrame()


def write_cache_to_gsheet(df, path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return

    if df is None:
        return

    try:
        title = cache_sheet_name_from_path(path)
        df2 = df.copy().fillna("")
        values = [list(df2.columns)] + df2.astype(str).values.tolist()
        ws = get_or_create_worksheet(title, rows=max(len(values), 100), cols=max(len(df2.columns), 20))

        if write_values_to_worksheet(ws, values):
            print(f"  ☁️ 已同步快取到 Google Sheet：{title}，共 {len(df2):,} 筆")
    except Exception as e:
        print(f"  ⚠️ 快取同步到 Google Sheet 失敗：{path}，原因：{type(e).__name__}: {e}")



def _hex_to_gsheet_color(hex_value):
    if not hex_value:
        return None

    s = str(hex_value).strip()

    if not s or s in ("00000000", "FFFFFFFF"):
        # 00000000 在 openpyxl 常代表無填色，不是真的黑色背景。
        return None

    if len(s) == 8:
        s = s[-6:]

    if len(s) != 6:
        return None

    try:
        return {
            "red": int(s[0:2], 16) / 255,
            "green": int(s[2:4], 16) / 255,
            "blue": int(s[4:6], 16) / 255,
        }
    except Exception:
        return None


def _openpyxl_color_to_gsheet(color_obj):
    if color_obj is None:
        return None

    try:
        if color_obj.type == "rgb" and color_obj.rgb:
            return _hex_to_gsheet_color(color_obj.rgb)
    except Exception:
        pass

    return None


def _openpyxl_fill_to_gsheet(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return None

        color = _openpyxl_color_to_gsheet(fill.fgColor)
        return color
    except Exception:
        return None


def _openpyxl_font_to_gsheet(cell):
    text_format = {}

    try:
        font = cell.font

        if font is None:
            return text_format

        if font.bold is not None:
            text_format["bold"] = bool(font.bold)

        if font.italic is not None:
            text_format["italic"] = bool(font.italic)

        if font.sz:
            text_format["fontSize"] = int(float(font.sz))

        color = _openpyxl_color_to_gsheet(font.color)
        if color:
            text_format["foregroundColor"] = color
    except Exception:
        pass

    return text_format


def _openpyxl_alignment_to_gsheet(cell):
    fmt = {}

    try:
        alignment = cell.alignment

        horizontal_map = {
            "center": "CENTER",
            "left": "LEFT",
            "right": "RIGHT",
        }
        vertical_map = {
            "center": "MIDDLE",
            "top": "TOP",
            "bottom": "BOTTOM",
        }

        if alignment.horizontal in horizontal_map:
            fmt["horizontalAlignment"] = horizontal_map[alignment.horizontal]

        if alignment.vertical in vertical_map:
            fmt["verticalAlignment"] = vertical_map[alignment.vertical]

        if alignment.wrap_text:
            fmt["wrapStrategy"] = "WRAP"
    except Exception:
        pass

    return fmt


def _openpyxl_number_format_to_gsheet(cell):
    try:
        nf = str(cell.number_format or "").strip()

        if not nf or nf == "General":
            return None

        if "%" in nf:
            return {"type": "PERCENT", "pattern": nf}

        if "#" in nf or "0" in nf:
            return {"type": "NUMBER", "pattern": nf}
    except Exception:
        pass

    return None


def _cell_gsheet_format(cell):
    fmt = {}

    bg = _openpyxl_fill_to_gsheet(cell)
    if bg:
        fmt["backgroundColor"] = bg

    text_format = _openpyxl_font_to_gsheet(cell)
    if text_format:
        fmt["textFormat"] = text_format

    align_format = _openpyxl_alignment_to_gsheet(cell)
    fmt.update(align_format)

    number_format = _openpyxl_number_format_to_gsheet(cell)
    if number_format:
        fmt["numberFormat"] = number_format

    return fmt


def _format_fields(fmt):
    fields = []

    if "backgroundColor" in fmt:
        fields.append("userEnteredFormat.backgroundColor")
    if "textFormat" in fmt:
        for key in fmt["textFormat"].keys():
            fields.append(f"userEnteredFormat.textFormat.{key}")
    if "horizontalAlignment" in fmt:
        fields.append("userEnteredFormat.horizontalAlignment")
    if "verticalAlignment" in fmt:
        fields.append("userEnteredFormat.verticalAlignment")
    if "wrapStrategy" in fmt:
        fields.append("userEnteredFormat.wrapStrategy")
    if "numberFormat" in fmt:
        fields.append("userEnteredFormat.numberFormat")

    return ",".join(fields)


def _gsheet_batch_update(requests, chunk_size=1000):
    if not requests:
        return

    sh = get_gsheet_spreadsheet()

    if sh is None:
        return

    for start in range(0, len(requests), chunk_size):
        chunk = requests[start:start + chunk_size]
        try:
            gsheet_api_call(
                f"套用 Google Sheet 格式 batchUpdate {start + 1}-{start + len(chunk)}",
                sh.batch_update,
                {"requests": chunk},
            )
        except Exception as e:
            print(f"  ⚠️ Google Sheet 格式套用失敗：{type(e).__name__}: {e}")
            return


def _openpyxl_freeze_to_grid_properties(ws_xlsx):
    frozen_rows = 0
    frozen_cols = 0

    try:
        pane = ws_xlsx.freeze_panes

        if pane:
            from openpyxl.utils.cell import coordinate_to_tuple
            row, col = coordinate_to_tuple(str(pane))
            frozen_rows = max(row - 1, 0)
            frozen_cols = max(col - 1, 0)
    except Exception:
        pass

    return frozen_rows, frozen_cols


def _excel_width_to_pixels(width):
    try:
        return max(20, int(float(width) * 7 + 5))
    except Exception:
        return None


def _excel_height_to_pixels(height):
    try:
        return max(18, int(float(height) * 1.333))
    except Exception:
        return None


def normalize_formula_for_gsheet(value):
    if not isinstance(value, str):
        return value

    if not value.startswith("="):
        return value

    # Google Sheets 對中文工作表名稱建議加單引號，避免公式解析失敗。
    value = value.replace("INDEX(券商查詢資料!", "INDEX('券商查詢資料'!")
    value = value.replace(",券商查詢資料!", ",'券商查詢資料'!")
    value = value.replace("MATCH($B$2&\"|\"&$A", "MATCH($B$2&\"|\"&$A")
    return value


def apply_excel_style_to_gsheet(ws_xlsx, gws):
    """
    將 openpyxl 產生的 Excel 樣式轉成 Google Sheets 格式。

    會同步：
    1. 背景色、字體粗細 / 字色 / 字級
    2. 文字置中、換行、數字格式
    3. 欄寬、列高、凍結列 / 欄
    4. 合併儲存格
    5. 隱藏工作表
    6. 券商查詢 B2 下拉選單

    Google Sheets API 與 Excel 格式模型不同，因此外框只保留主要視覺效果，
    但 A/B/C/D/E 的紅綠藍橘狀態色、標頭色與查詢頁互動功能會完整保留。
    """
    if gws is None:
        return

    sheet_id = int(gws.id)
    requests = []

    frozen_rows, frozen_cols = _openpyxl_freeze_to_grid_properties(ws_xlsx)

    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "hidden": bool(ws_xlsx.sheet_state == "hidden"),
                "gridProperties": {
                    "frozenRowCount": frozen_rows,
                    "frozenColumnCount": frozen_cols,
                },
            },
            "fields": "hidden,gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
        }
    })

    # 先清除舊合併範圍，避免重複執行時 mergeCells 失敗。
    requests.append({
        "unmergeCells": {
            "range": {
                "sheetId": sheet_id,
            }
        }
    })

    # 合併儲存格。
    for merged_range in ws_xlsx.merged_cells.ranges:
        try:
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": merged_range.min_row - 1,
                        "endRowIndex": merged_range.max_row,
                        "startColumnIndex": merged_range.min_col - 1,
                        "endColumnIndex": merged_range.max_col,
                    },
                    "mergeType": "MERGE_ALL",
                }
            })
        except Exception:
            pass

    # 欄寬。
    for col_idx in range(1, ws_xlsx.max_column + 1):
        letter = get_column_letter(col_idx)
        width = ws_xlsx.column_dimensions[letter].width
        pixel_size = _excel_width_to_pixels(width) if width else None

        if pixel_size:
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": col_idx - 1,
                        "endIndex": col_idx,
                    },
                    "properties": {
                        "pixelSize": pixel_size,
                    },
                    "fields": "pixelSize",
                }
            })

    # 列高。
    for row_idx in range(1, ws_xlsx.max_row + 1):
        height = ws_xlsx.row_dimensions[row_idx].height
        pixel_size = _excel_height_to_pixels(height) if height else None

        if pixel_size:
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "ROWS",
                        "startIndex": row_idx - 1,
                        "endIndex": row_idx,
                    },
                    "properties": {
                        "pixelSize": pixel_size,
                    },
                    "fields": "pixelSize",
                }
            })

    # 逐列壓縮相同格式的連續儲存格，減少 batchUpdate 請求數量。
    import json as _json

    for row_idx in range(1, ws_xlsx.max_row + 1):
        run_start = None
        run_fmt = None
        run_key = None

        for col_idx in range(1, ws_xlsx.max_column + 2):
            if col_idx <= ws_xlsx.max_column:
                cell = ws_xlsx.cell(row_idx, col_idx)
                fmt = _cell_gsheet_format(cell)
                key = _json.dumps(fmt, sort_keys=True, ensure_ascii=False) if fmt else None
            else:
                fmt = None
                key = None

            if key and run_key is None:
                run_start = col_idx
                run_fmt = fmt
                run_key = key
            elif key and key == run_key:
                continue
            else:
                if run_key and run_start is not None and run_fmt:
                    fields = _format_fields(run_fmt)
                    if fields:
                        requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": row_idx - 1,
                                    "endRowIndex": row_idx,
                                    "startColumnIndex": run_start - 1,
                                    "endColumnIndex": col_idx - 1,
                                },
                                "cell": {
                                    "userEnteredFormat": run_fmt,
                                },
                                "fields": fields,
                            }
                        })

                if key:
                    run_start = col_idx
                    run_fmt = fmt
                    run_key = key
                else:
                    run_start = None
                    run_fmt = None
                    run_key = None

    # Google Sheets 版券商查詢：B2 下拉選單。
    if ws_xlsx.title == "券商查詢":
        try:
            # 使用 ONE_OF_LIST 直接寫入券商清單，比 ONE_OF_RANGE 更穩定，
            # 可避免部分環境下跨工作表範圍驗證未顯示下拉箭頭。
            sh = get_gsheet_spreadsheet()
            data_ws = sh.worksheet("券商查詢資料") if sh else None
            broker_values = []

            if data_ws:
                for value in data_ws.col_values(16)[1:]:
                    value = str(value).strip()
                    if value and value not in broker_values:
                        broker_values.append(value)

            if broker_values:
                requests.append({
                    "setDataValidation": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "endRowIndex": 2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 2,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": broker}
                                    for broker in broker_values
                                ],
                            },
                            "showCustomUi": True,
                            "strict": True,
                        },
                    }
                })
        except Exception as e:
            print(f"  ⚠️ 券商查詢下拉選單建立失敗：{type(e).__name__}: {e}")

    _gsheet_batch_update(requests)



# ══════════════════════════════════════════════════════════════════════
# Google Sheet 結果同步：補充 / 更新模式
# ══════════════════════════════════════════════════════════════════════

GSHEET_RESULT_UPSERT_ENABLED = os.getenv("GSHEET_RESULT_UPSERT_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_LEGACY_SCOPE_LABEL = os.getenv("GSHEET_LEGACY_SCOPE_LABEL", "未標記舊資料").strip() or "未標記舊資料"

GSHEET_RESULT_UPSERT_TITLES = {
    *AMOUNT_CLASS_SHEET_NAMES,
    "每日賣出明細",
    "近兩月買賣金額排行",
    "近兩月分點數排行",
    TOP15_POSITION_DETAIL_SHEET,
    TOP15_CONSENSUS_SHEET,
    BROKER_10D_DETAIL_SHEET,
    BROKER_10D_WINRATE_RANK_SHEET,
}

GSHEET_RESULT_OVERWRITE_TITLES = {
    WARRANT_CONSENSUS_7D_SHEET,
    "券商查詢",
    "券商查詢資料",
    "價格抓取狀態",
    "顏色說明",
}

# RUN_MODE=1 精選五分點模式只用來產出每日圖卡與精選資料，
# 不應覆蓋全分點模式才需要維護的查詢頁與勝率統計頁。
# 這些工作表只在 RUN_MODE=2 全分點模式同步到 Google Sheet。
GSHEET_RUN_MODE1_SKIP_RESULT_TITLES = {
    "券商查詢",
    "券商查詢資料",
    "勝率統計",
    "ABCDE組合勝率",
}


def should_skip_result_sheet_in_run_mode(title):
    title = safe_worksheet_title(title)
    return RUN_MODE == 1 and title in GSHEET_RUN_MODE1_SKIP_RESULT_TITLES


def get_result_data_scope():
    """
    Google Sheet 結果同步用的資料範圍標籤。

    目的：
    - RUN_MODE=1 跑精選五分點時，只更新「資料範圍=精選五分點」的資料。
    - RUN_MODE=2 跑完整分點時，只更新「資料範圍=全分點」的資料。
    - 同一張 Google Sheet 內不同資料範圍可以並存，避免五分點覆蓋全分點資料。
    """
    return "精選五分點" if RUN_MODE == 1 else "全分點"


def should_upsert_result_sheet(title):
    title = safe_worksheet_title(title)

    if not GSHEET_RESULT_UPSERT_ENABLED:
        return False

    if title in GSHEET_RESULT_OVERWRITE_TITLES:
        return False

    if title in GSHEET_RESULT_UPSERT_TITLES:
        return True

    return False


def read_existing_worksheet_values(title):
    try:
        sh = get_gsheet_spreadsheet()
        if sh is None:
            return []
        ws = sh.worksheet(safe_worksheet_title(title))
        values = ws.get_all_values()
        return values or []
    except Exception:
        return []


def _find_simple_header_row(values):
    if not values:
        return None

    scan_limit = min(len(values), 10)
    key_headers = {
        "統計日期", "事件類型", "日期", "排名", "權證代號", "標的股", "分點",
        "買進金額", "賣出金額", "淨買超成本", "排名類型",
    }

    for idx in range(scan_limit):
        row = [str(x).strip() for x in values[idx]]
        non_empty = [x for x in row if x]
        if not non_empty:
            continue
        if len(set(row) & key_headers) >= 2:
            return idx

    return 0 if values and any(str(x).strip() for x in values[0]) else None


def _values_to_records(values, header_row_idx=0, default_scope=None):
    if not values or header_row_idx is None or header_row_idx >= len(values):
        return [], []

    headers = [str(h).strip() for h in values[header_row_idx]]
    if not headers or all(h == "" for h in headers):
        return [], []

    # 去掉尾端完全空白表頭，避免 Google Sheet 舊資料多出空欄造成 key 錯亂。
    while headers and headers[-1] == "":
        headers.pop()

    records = []
    for raw_row in values[header_row_idx + 1:]:
        row = list(raw_row)
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[:len(headers)]

        if not any(str(v).strip() for v in row):
            continue

        rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(headers)}
        if default_scope is not None:
            rec["資料範圍"] = str(rec.get("資料範圍", "") or default_scope).strip()
        records.append(rec)

    return headers, records


def _records_to_values(headers, records):
    out = [list(headers)]
    for rec in records:
        out.append([rec.get(h, "") for h in headers])
    return out


def _ensure_scope_header(headers):
    headers = [str(h).strip() for h in headers]
    if "資料範圍" not in headers:
        return ["資料範圍"] + headers
    return headers


def _sheet_upsert_key_columns(title, headers):
    title = safe_worksheet_title(title)
    hset = set(headers)

    def keep(cols):
        return [c for c in cols if c in hset]

    if title in AMOUNT_CLASS_SHEET_NAMES:
        return keep(["資料範圍", "事件類型", "分點", "券商代號", "標的股", "事件日", "權證清單"])

    if title == TOP15_CONSENSUS_SHEET:
        return keep(["資料範圍", "統計日期", "標的股"])

    if title == TOP15_POSITION_DETAIL_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號", "標的股", "權證代號", "事件", "事件日", "買進日"])

    if title == WARRANT_CONSENSUS_7D_SHEET:
        # 近7／14／21日共識皆為「標的層級」排名；統計天數必須納入 key，避免三個期間互相覆蓋。
        return keep(["資料範圍", "統計日期", "統計天數", "排名類型", "標的股"])

    if title == BROKER_10D_DETAIL_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號", "標的股"])

    if title == BROKER_10D_WINRATE_RANK_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號"])


    if title == "每日賣出明細":
        return keep(["資料範圍", "日期", "分點", "券商代號", "標的股", "權證代號", "事件", "狀態", "事件日"])

    if title == "近兩月買賣金額排行":
        return keep(["資料範圍", "權證代號", "買進分點"])

    if title == "近兩月分點數排行":
        return keep(["資料範圍", "標的股", "買進分點清單"])

    # 通用備援：至少要有資料範圍，加上一些穩定欄位。
    cols = keep([
        "資料範圍", "統計日期", "日期", "事件類型", "分點", "券商代號",
        "標的股", "權證代號", "買進日", "事件日", "起始日", "結束日",
        "排名類型",
    ])
    return cols if len(cols) >= 2 else []



def normalize_underlying_code_for_group(value, fallback_text=""):
    """
    近7日共識與 Google Sheet upsert 用的標的股代號正規化。

    修正目的：
    同一標的可能因 Google Sheet / pandas 讀寫變成 2408、'2408、2408.0，
    若直接拿原字串當 group key，就會造成本週 TOP15 同標的重複出現。
    這裡統一轉成穩定代號後再合併與排序。
    """
    candidates = []
    for raw in (value, fallback_text):
        if raw is None:
            continue
        s = strip_gsheet_text_prefix(str(raw)).strip()
        if not s:
            continue
        candidates.append(s)

    for s in candidates:
        s = s.replace("，", ",").replace(",", "").strip()
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]

        m = re.match(r"^(\d{4,6})(?:\s|$)", s)
        if m:
            return m.group(1)

        if s.isdigit() and 4 <= len(s) <= 6:
            return s

        digits = "".join(ch for ch in s if ch.isdigit())
        if 4 <= len(digits) <= 6:
            return digits

    return ""

def _record_key(rec, key_cols):
    parts = []
    for col in key_cols:
        value = rec.get(col, "")
        if col in ("日期", "統計日期", "買進日", "事件日", "起始日", "結束日", "第一筆日期", "最後筆日期"):
            value = normalize_date_str(strip_gsheet_text_prefix(value))
        elif col in ("標的股", "標的代號", "標的代碼"):
            value = normalize_underlying_code_for_group(value)
        else:
            value = strip_gsheet_text_prefix(value)
        parts.append(str(value).strip())
    return tuple(parts)



def _result_retention_policy(title):
    """回傳需要裁切的日期欄與保留的不同日期數。"""
    title = safe_worksheet_title(title)

    if title == "每日賣出明細":
        return {
            "date_col": "日期",
            "keep_count": GSHEET_DAILY_SELL_KEEP_TRADING_DAYS,
            "label": "交易日",
        }

    if title in (TOP15_POSITION_DETAIL_SHEET, TOP15_CONSENSUS_SHEET):
        return {
            "date_col": "統計日期",
            "keep_count": GSHEET_TOP15_KEEP_STAT_DATES,
            "label": "統計日期",
        }

    return None


def _parse_result_record_date(value):
    """解析結果表日期；同時支援 yyyy/mm/dd 與 Google Sheet 日期序號。"""
    raw = strip_gsheet_text_prefix(value)
    dt = parse_date(raw)
    if dt is not None:
        return dt

    try:
        serial = float(str(raw).replace(",", "").strip())
        if 20000 <= serial <= 100000:
            return datetime(1899, 12, 30) + timedelta(days=serial)
    except Exception:
        pass

    return None


def _split_records_by_recent_dates(records, date_col, keep_count):
    """
    依「資料範圍」分開保留最近 N 個不同日期。

    精選五分點與全分點可能不是同一天更新，因此不能用全表共同日期裁切，
    否則更新較頻繁的一邊會把另一邊的日期提前擠出主表。
    """
    keep_count = max(int(keep_count or 1), 1)
    dates_by_scope = {}

    for rec in records:
        scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip() or GSHEET_LEGACY_SCOPE_LABEL
        dt = _parse_result_record_date(rec.get(date_col, ""))
        if dt:
            dates_by_scope.setdefault(scope, set()).add(dt.date())

    keep_dates_by_scope = {
        scope: set(sorted(date_set, reverse=True)[:keep_count])
        for scope, date_set in dates_by_scope.items()
    }

    retained = []
    archive = []

    for rec in records:
        scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip() or GSHEET_LEGACY_SCOPE_LABEL
        dt = _parse_result_record_date(rec.get(date_col, ""))

        # 日期無法辨識時留在主表，避免誤刪舊版或人工資料。
        if dt is None:
            retained.append(rec)
            continue

        if dt.date() in keep_dates_by_scope.get(scope, set()):
            retained.append(rec)
        else:
            archive.append(rec)

    return retained, archive


def _archive_year_label(rec, date_col):
    if not GSHEET_ARCHIVE_YEARLY_SPLIT:
        return "全部"

    dt = _parse_result_record_date(rec.get(date_col, ""))
    return str(dt.year) if dt else "未分類"


def _archive_spreadsheet_name(year_label, part_no=1):
    base = f"{GSHEET_ARCHIVE_NAME_PREFIX}_{year_label}"
    return base if int(part_no) <= 1 else f"{base}_{int(part_no):02d}"


def _explicit_archive_share_emails():
    if not GSHEET_ARCHIVE_SHARE_EMAILS:
        return []

    out = []
    for email in re.split(r"[,;；、\n\r\t]+", GSHEET_ARCHIVE_SHARE_EMAILS):
        email = str(email).strip()
        if email and "@" in email and email not in out:
            out.append(email)
    return out


def _sync_archive_permissions(archive_sh):
    """
    新建立的封存檔由服務帳號擁有時，盡量複製主試算表既有使用者 / 群組權限；
    也可透過 GSHEET_ARCHIVE_SHARE_EMAILS 明確指定要分享的信箱。
    """
    if archive_sh is None:
        return

    archive_id = str(getattr(archive_sh, "id", "") or "")
    if archive_id and archive_id in _GSHEET_ARCHIVE_PERMISSION_SYNCED:
        return

    share_targets = {}

    try:
        primary_sh = get_gsheet_spreadsheet()
        if primary_sh is not None and hasattr(primary_sh, "list_permissions"):
            for permission in primary_sh.list_permissions() or []:
                ptype = str(permission.get("type", "")).strip()
                role = str(permission.get("role", "")).strip()
                email = str(permission.get("emailAddress", "")).strip()

                if ptype not in ("user", "group") or not email:
                    continue
                if role not in ("owner", "reader", "writer", "commenter"):
                    continue

                # 主檔 owner 在封存檔改授予 writer；commenter 則降成 reader。
                if role == "owner":
                    mapped_role = "writer"
                elif role == "commenter":
                    mapped_role = "reader"
                else:
                    mapped_role = role
                share_targets[(ptype, email)] = mapped_role
    except Exception:
        pass

    for email in _explicit_archive_share_emails():
        share_targets[("user", email)] = "writer"

    for (ptype, email), role in share_targets.items():
        try:
            archive_sh.share(
                email,
                perm_type=ptype,
                role=role,
                notify=False,
            )
        except Exception:
            pass

    if archive_id:
        _GSHEET_ARCHIVE_PERMISSION_SYNCED.add(archive_id)


def is_gsheet_drive_storage_quota_error(exc):
    """判斷是否為 Google Drive 永久性儲存空間不足，而不是可藉由換檔名重試的錯誤。"""
    msg = str(exc or "").strip().lower()
    return (
        "drive storage quota has been exceeded" in msg
        or "storage quota has been exceeded" in msg
        or "storagequota" in msg
        or ("403" in msg and "drive" in msg and "quota" in msg)
    )


def _configured_archive_spreadsheet_id(part_no):
    try:
        idx = int(part_no) - 1
    except Exception:
        return ""

    if 0 <= idx < len(GSHEET_ARCHIVE_SPREADSHEET_IDS):
        return str(GSHEET_ARCHIVE_SPREADSHEET_IDS[idx]).strip()
    return ""


def _archive_worksheet_title(title, year_label):
    """
    使用固定封存試算表 ID 時，同一份檔案可能保存多個年度，
    因此把年度放進工作表名稱，避免 2026 與 2027 資料混在同一張分頁。
    """
    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        return safe_worksheet_title(f"{title}_{year_label}")
    return safe_worksheet_title(title)


def _mark_archive_create_blocked(exc):
    global _GSHEET_ARCHIVE_CREATE_BLOCKED
    global _GSHEET_ARCHIVE_CREATE_BLOCK_REASON
    global _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED

    _GSHEET_ARCHIVE_CREATE_BLOCKED = True
    _GSHEET_ARCHIVE_CREATE_BLOCK_REASON = str(exc or "Google Drive storage quota exceeded").strip()

    if not _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED:
        print(
            "  ⚠️ Google Drive 儲存空間不足，本次執行已停止自動建立新的封存試算表；"
            "不會再嘗試 _02、_03……。請先用自己的 Google 帳號建立封存試算表、"
            "分享給服務帳號，並設定 GSHEET_ARCHIVE_SPREADSHEET_ID。"
        )
        _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED = True


def _open_archive_spreadsheet(year_label, part_no=1, create_if_missing=False):
    key = (str(year_label), int(part_no))
    if key in _GSHEET_ARCHIVE_SPREADSHEETS:
        return _GSHEET_ARCHIVE_SPREADSHEETS[key]

    gc = get_gsheet_client()
    if gc is None:
        return None

    # 優先使用使用者自行建立、由使用者帳號持有的封存試算表。
    # 這不會消耗服務帳號的個人 Drive 配額。
    configured_id = _configured_archive_spreadsheet_id(part_no)
    if configured_id:
        try:
            sh = gc.open_by_key(configured_id)
            _GSHEET_ARCHIVE_SPREADSHEETS[key] = sh
            _sync_archive_permissions(sh)
            return sh
        except Exception as e:
            print(
                f"  ⚠️ 開啟指定 Google Sheet 封存檔失敗：分卷={int(part_no):02d}｜"
                f"ID={configured_id}｜原因：{type(e).__name__}: {e}"
            )
            return None

    # 已設定固定封存檔清單時，只使用清單內的檔案；不再由服務帳號另外建立新檔。
    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        return None

    name = _archive_spreadsheet_name(year_label, part_no)
    sh = None

    try:
        sh = gc.open(name)
    except Exception:
        if not create_if_missing or not GSHEET_ARCHIVE_ALLOW_CREATE:
            return None

        # Drive quota 屬於永久性錯誤，本次執行第一次遇到後便停止建立，避免連續刷出 _02～_99。
        if _GSHEET_ARCHIVE_CREATE_BLOCKED:
            return None

        try:
            if GSHEET_ARCHIVE_FOLDER_ID:
                try:
                    sh = gc.create(name, folder_id=GSHEET_ARCHIVE_FOLDER_ID)
                except TypeError:
                    # 相容較舊版 gspread；舊版 create() 沒有 folder_id 參數。
                    sh = gc.create(name)
            else:
                sh = gc.create(name)
            print(f"  📦 已建立 Google Sheet 年度封存檔：{name}")
        except Exception as e:
            if is_gsheet_drive_storage_quota_error(e):
                _mark_archive_create_blocked(e)
            else:
                print(f"  ⚠️ 建立 Google Sheet 封存檔失敗：{name}，原因：{type(e).__name__}: {e}")
            return None

    _GSHEET_ARCHIVE_SPREADSHEETS[key] = sh
    _sync_archive_permissions(sh)
    return sh


def _is_blank_archive_worksheet(ws):
    try:
        values = _worksheet_values_with_formulas(ws)
        return not any(str(v).strip() for row in values for v in row)
    except Exception:
        return False


def _prepare_archive_target_worksheet(archive_sh, title, target_rows, target_cols):
    """
    檢查指定封存檔是否還容得下這張工作表；若可容納，回傳可寫入的 worksheet。
    """
    if archive_sh is None:
        return None

    title = safe_worksheet_title(title)
    target_rows = max(int(target_rows), 1)
    target_cols = max(int(target_cols), 1)

    existing_ws = None
    try:
        existing_ws = archive_sh.worksheet(title)
    except Exception:
        existing_ws = None

    total_cells = spreadsheet_grid_cell_count(archive_sh)
    current_target_cells = 0
    blank_default_ws = None

    if existing_ws is not None:
        current_target_cells = max(int(existing_ws.row_count), 1) * max(int(existing_ws.col_count), 1)
    else:
        try:
            sheets = archive_sh.worksheets()
            if len(sheets) == 1 and _is_blank_archive_worksheet(sheets[0]):
                blank_default_ws = sheets[0]
                current_target_cells = max(int(blank_default_ws.row_count), 1) * max(int(blank_default_ws.col_count), 1)
        except Exception:
            blank_default_ws = None

    projected_cells = total_cells - current_target_cells + target_rows * target_cols
    if projected_cells > GSHEET_ARCHIVE_MAX_CELLS:
        return None

    if existing_ws is not None:
        return existing_ws

    if blank_default_ws is not None:
        try:
            gsheet_api_call(
                f"重新命名封存工作表 {title}",
                blank_default_ws.update_title,
                title,
            )
            return blank_default_ws
        except Exception:
            pass

    try:
        return gsheet_api_call(
            f"建立封存工作表 {title}",
            archive_sh.add_worksheet,
            title=title,
            rows=target_rows,
            cols=target_cols,
        )
    except Exception as e:
        print(f"  ⚠️ 建立封存工作表失敗：{title}，原因：{type(e).__name__}: {e}")
        return None


def _read_archive_worksheet_values(archive_sh, title):
    if archive_sh is None:
        return []

    try:
        ws = archive_sh.worksheet(safe_worksheet_title(title))
        return ws.get_all_values() or []
    except Exception:
        return []


def _merge_records_for_archive(title, incoming_headers, incoming_records, old_values):
    old_headers, old_records = _values_to_records(
        old_values,
        header_row_idx=0,
        default_scope=GSHEET_LEGACY_SCOPE_LABEL,
    )
    old_headers = _ensure_scope_header(old_headers) if old_headers else []

    headers = []
    for h in list(incoming_headers) + list(old_headers):
        h = str(h).strip()
        if h and h not in headers:
            headers.append(h)

    key_cols = _sheet_upsert_key_columns(title, headers)
    if not key_cols or "資料範圍" not in key_cols:
        # 理論上目前三張封存表一定有穩定 key；若未來欄位改名，至少保留原順序追加。
        return headers, old_records + list(incoming_records)

    merged_map = {}
    order = []

    for rec in list(old_records) + list(incoming_records):
        if not rec.get("資料範圍"):
            rec["資料範圍"] = GSHEET_LEGACY_SCOPE_LABEL

        key = _record_key(rec, key_cols)
        if not any(key):
            continue

        if key not in merged_map:
            order.append(key)
        merged_map[key] = rec

    return headers, [merged_map[key] for key in order]


def _write_values_to_archive_worksheet(ws, values):
    """封存檔使用的精簡寫入：完整覆蓋、保留代號前導 0，並把格數縮到實際大小。"""
    if ws is None:
        return False

    if not values:
        values = [[""]]

    row_count = max(len(values), 1)
    col_count = max(max((len(row) for row in values), default=1), 1)
    normalized_values = []

    for row in values:
        row = list(row)
        if len(row) < col_count:
            row += [""] * (col_count - len(row))
        normalized_values.append([clean_gsheet_value(v) for v in row])

    normalized_values = normalize_gsheet_values_for_text_columns(normalized_values)

    try:
        gsheet_api_call(
            f"調整封存工作表大小 {ws.title}",
            ws.resize,
            rows=row_count,
            cols=col_count,
        )
        gsheet_api_call(f"清除封存工作表 {ws.title}", ws.clear)

        for start in range(0, len(normalized_values), GSHEET_CHUNK_ROWS):
            chunk = normalized_values[start:start + GSHEET_CHUNK_ROWS]
            gsheet_api_call(
                f"寫入封存工作表 {ws.title} A{start + 1}",
                ws.update,
                values=chunk,
                range_name=f"A{start + 1}",
                value_input_option="USER_ENTERED",
            )

        # 再 resize 一次，確保 clear / update 後沒有殘留多餘空白列欄。
        gsheet_api_call(
            f"最終縮減封存工作表 {ws.title}",
            ws.resize,
            rows=row_count,
            cols=col_count,
        )
        return True
    except Exception as e:
        print(f"  ⚠️ Google Sheet 封存寫入失敗：{ws.title}，原因：{type(e).__name__}: {e}")
        return False


def _archive_records_to_year_part(title, headers, records, year_label):
    """
    把同一年度資料寫入可容納的封存分卷。

    - 有設定 GSHEET_ARCHIVE_SPREADSHEET_IDS：依序使用既有試算表 ID。
    - 未設定 ID：沿用原本依名稱開啟 / 建立年度封存檔。
    - 一旦確認 Drive quota 不足，立即停止，不再嘗試 _02～_99。
    """
    if not records:
        return True

    max_parts = len(GSHEET_ARCHIVE_SPREADSHEET_IDS) if GSHEET_ARCHIVE_SPREADSHEET_IDS else 99
    archive_ws_title = _archive_worksheet_title(title, year_label)

    for part_no in range(1, max_parts + 1):
        archive_sh = _open_archive_spreadsheet(
            year_label,
            part_no=part_no,
            create_if_missing=True,
        )

        if archive_sh is None:
            if _GSHEET_ARCHIVE_CREATE_BLOCKED:
                break
            continue

        old_values = _read_archive_worksheet_values(archive_sh, archive_ws_title)
        merged_headers, merged_records = _merge_records_for_archive(
            title,
            headers,
            records,
            old_values,
        )
        merged_values = _records_to_values(merged_headers, merged_records)
        target_rows = max(len(merged_values), 1)
        target_cols = max(max((len(row) for row in merged_values), default=1), 1)

        ws = _prepare_archive_target_worksheet(
            archive_sh,
            archive_ws_title,
            target_rows=target_rows,
            target_cols=target_cols,
        )
        if ws is None:
            # 這一卷預估超過安全格數，改試下一個既有 ID 或下一個可建立分卷。
            continue

        if _write_values_to_archive_worksheet(ws, merged_values):
            archive_url = str(getattr(archive_sh, "url", "") or "").strip()
            print(
                f"  📦 Google Sheet 舊資料封存完成：{title}｜年度={year_label}｜"
                f"分卷={part_no:02d}｜工作表={archive_ws_title}｜本次 {len(records):,} 筆｜"
                f"封存表合計 {len(merged_records):,} 筆"
                + (f"｜{archive_url}" if archive_url else "")
            )
            return True

    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        print(
            f"  ⚠️ Google Sheet 舊資料封存失敗：{title}｜年度={year_label}，"
            "已設定的封存試算表 ID 均無法使用或已達安全格數。"
        )
    elif not _GSHEET_ARCHIVE_CREATE_BLOCKED:
        print(f"  ⚠️ Google Sheet 舊資料封存失敗：{title}｜年度={year_label}，找不到可用封存分卷。")
    return False


def archive_result_records(title, headers, records, date_col):
    """
    將主表裁切出的舊資料依年度寫入獨立 Google Sheet。

    只有全部年度都封存成功時才回傳 True；任一年度失敗時，主表會保留全部資料避免遺失。
    """
    if not records:
        return True

    if not GSHEET_RESULT_ARCHIVE_ENABLED:
        print(f"  ⚠️ 已停用 Google Sheet 自動封存，{title} 舊資料暫不從主表移除。")
        return False

    grouped = {}
    for rec in records:
        year_label = _archive_year_label(rec, date_col)
        grouped.setdefault(year_label, []).append(rec)

    all_ok = True
    for year_label in sorted(grouped.keys()):
        ok = _archive_records_to_year_part(
            title,
            headers,
            grouped[year_label],
            year_label,
        )
        all_ok = all_ok and ok

    return all_ok


def apply_result_retention_and_archive(title, headers, merged_records):
    policy = _result_retention_policy(title)
    if not policy:
        return merged_records, 0

    retained_records, archive_records = _split_records_by_recent_dates(
        merged_records,
        date_col=policy["date_col"],
        keep_count=policy["keep_count"],
    )

    if not archive_records:
        return retained_records, 0

    archive_ok = archive_result_records(
        title,
        headers,
        archive_records,
        date_col=policy["date_col"],
    )

    if not archive_ok:
        print(
            f"  ⚠️ {title} 有 {len(archive_records):,} 筆舊資料尚未完成封存，"
            "本次仍保留在主試算表，避免資料遺失。"
        )
        return merged_records, 0

    print(
        f"  🧹 Google Sheet 主表保留策略：{title}｜"
        f"每個資料範圍保留最近 {policy['keep_count']} 個{policy['label']}｜"
        f"主表 {len(retained_records):,} 筆｜已移至封存 {len(archive_records):,} 筆"
    )
    return retained_records, len(archive_records)


def merge_result_values_for_gsheet(title, new_values):
    """
    將本次 Excel 結果與 Google Sheet 既有結果做 upsert 合併。

    規則：
    - 本次資料會加上「資料範圍」：精選五分點 / 全分點。
    - 舊資料若沒有「資料範圍」，會保留為「未標記舊資料」，不會被本次資料誤覆蓋。
    - key 重複時，以本次新資料為準。
    - key 不重複時，舊資料保留。
    - 每日賣出明細與 TOP15 快取超過主表保留日期的資料，會先封存到獨立 Google Sheet；
      封存成功後才從主表移除。
    """
    if not should_upsert_result_sheet(title):
        return new_values

    header_row_idx = _find_simple_header_row(new_values)
    if header_row_idx is None:
        return new_values

    if header_row_idx != 0:
        # 目前只對第一列就是表頭的資料表啟用 upsert。
        # 多段式報表例如勝率統計、ABCDE組合勝率，仍維持原本覆蓋模式，避免破壞版面與公式。
        return new_values

    current_scope = get_result_data_scope()
    new_headers, new_records = _values_to_records(new_values, header_row_idx=0, default_scope=current_scope)
    if not new_headers:
        return new_values

    new_headers = _ensure_scope_header(new_headers)
    for rec in new_records:
        rec["資料範圍"] = current_scope

    old_values = read_existing_worksheet_values(title)
    old_headers, old_records = _values_to_records(
        old_values,
        header_row_idx=0,
        default_scope=GSHEET_LEGACY_SCOPE_LABEL,
    )
    old_headers = _ensure_scope_header(old_headers) if old_headers else []

    headers = []
    for h in new_headers + old_headers:
        h = str(h).strip()
        if h and h not in headers:
            headers.append(h)

    if not headers:
        return new_values

    key_cols = _sheet_upsert_key_columns(title, headers)
    if not key_cols or "資料範圍" not in key_cols:
        return new_values

    old_map = {}
    old_order = []
    for rec in old_records:
        if not rec.get("資料範圍"):
            rec["資料範圍"] = GSHEET_LEGACY_SCOPE_LABEL
        key = _record_key(rec, key_cols)
        if not any(key):
            continue
        if key not in old_map:
            old_order.append(key)
        old_map[key] = rec

    new_map = {}
    new_order = []
    for rec in new_records:
        rec["資料範圍"] = current_scope
        key = _record_key(rec, key_cols)
        if not any(key):
            continue
        if key not in new_map:
            new_order.append(key)
        new_map[key] = rec

    merged_records = []
    used_keys = set()

    # 本次資料排在最前面，畫面上可以優先看到最新結果。
    for key in new_order:
        merged_records.append(new_map[key])
        used_keys.add(key)

    # 舊資料中沒有被本次 key 覆蓋的保留下來，避免五分點覆蓋全分點。
    for key in old_order:
        if key in used_keys:
            continue
        merged_records.append(old_map[key])
        used_keys.add(key)

    retained_records, archived_count = apply_result_retention_and_archive(
        title,
        headers,
        merged_records,
    )
    merged_values = _records_to_values(headers, retained_records)

    print(
        f"  ☁️ Google Sheet upsert：{safe_worksheet_title(title)}｜"
        f"資料範圍={current_scope}｜本次 {len(new_records):,} 筆｜舊資料 {len(old_records):,} 筆｜"
        f"合併後 {len(merged_records):,} 筆｜主表保留 {len(retained_records):,} 筆｜封存 {archived_count:,} 筆"
    )

    return merged_values




def apply_safe_result_table_style_to_gsheet(gws, values=None):
    """
    針對 upsert 結果表套用安全版基本樣式。

    為什麼不用 apply_excel_style_to_gsheet：
    upsert 會在最左邊新增「資料範圍」欄位，而且會把舊資料與本次資料合併。
    如果繼續照原 Excel 欄位位置套樣式，日期格式會套到錯欄，
    例如「買進均價」會被顯示成 1899/12/30。

    這個函式只根據實際寫入 Google Sheet 的 values 表頭套安全樣式：
    - 凍結表頭列
    - 表頭粗體、淡黃色底、置中
    - 資料列垂直置中、換行
    - 不套任何日期或數字格式；日期 / 數字格式由後面的專用函式處理
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    header_row_idx = _find_simple_header_row(values)
    if header_row_idx is None:
        return

    max_cols = max(max((len(row) for row in values), default=1), 1)
    max_rows = max(len(values), 1)

    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": header_row_idx + 1,
                    },
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row_idx,
                    "endRowIndex": header_row_idx + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 1.0, "green": 0.949, "blue": 0.8},
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "WRAP",
                    }
                },
                "fields": "userEnteredFormat.backgroundColor,userEnteredFormat.textFormat.bold,userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy",
            }
        },
    ]

    if max_rows > header_row_idx + 1:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row_idx + 1,
                    "endRowIndex": max_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "CLIP",
                    }
                },
                "fields": "userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy",
            }
        })
        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": header_row_idx + 1,
                    "endIndex": max_rows,
                },
                "properties": {
                    "pixelSize": 30,
                },
                "fields": "pixelSize",
            }
        })

    _gsheet_batch_update(requests)


def clear_all_number_formats_for_written_range(gws, values=None):
    """
    寫入結果後，先把實際資料範圍內的 numberFormat 全部清掉。

    這是為了處理舊版已經把「買進均價 / 減碼均價 / 出清獲利%」誤套成日期格式的工作表。
    單純寫入值不一定會移除舊格式，因此必須先清 numberFormat，再由日期 / 數字專用函式重套。
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    max_rows = max(len(values), 1)
    max_cols = max(max((len(row) for row in values), default=1), 1)

    requests = [{
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": max_rows,
                "startColumnIndex": 0,
                "endColumnIndex": max_cols,
            },
            "cell": {
                "userEnteredFormat": {}
            },
            "fields": "userEnteredFormat.numberFormat",
        }
    }]

    # 這裡只清 numberFormat，不動其他樣式。
    # 後續 TEXT / DATE / NUMBER 專用函式會依實際表頭重新套正確格式。
    _gsheet_batch_update(requests)



def _openpyxl_border_to_gsheet(cell):
    """
    只把 Excel 裡真正有設定的外框轉到 Google Sheet。

    用途：保留「減碼獲利% / 出清獲利%」的粗紅 / 粗綠外框。
    一般沒有外框的儲存格不處理，避免新增資料範圍欄位後破壞既有版面。
    """
    try:
        border = cell.border
    except Exception:
        return None

    if border is None:
        return None

    def side_to_gsheet(side):
        try:
            style = str(side.style or "").strip()
        except Exception:
            style = ""

        if not style:
            return None

        style_map = {
            "hair": "DOTTED",
            "dotted": "DOTTED",
            "dashDot": "DASHED",
            "dashDotDot": "DASHED",
            "dashed": "DASHED",
            "mediumDashDot": "DASHED",
            "mediumDashDotDot": "DASHED",
            "mediumDashed": "DASHED",
            "thin": "SOLID",
            "medium": "SOLID_MEDIUM",
            "thick": "SOLID_THICK",
            "double": "DOUBLE",
        }

        out = {"style": style_map.get(style, "SOLID")}
        color = _openpyxl_color_to_gsheet(getattr(side, "color", None))
        if color:
            out["color"] = color
        return out

    out = {}
    for attr, gname in [
        ("top", "top"),
        ("bottom", "bottom"),
        ("left", "left"),
        ("right", "right"),
    ]:
        side_fmt = side_to_gsheet(getattr(border, attr, None))
        if side_fmt:
            out[gname] = side_fmt

    return out or None


def _cell_gsheet_visual_format(cell):
    """
    只同步視覺樣式，不同步 numberFormat。

    目的：
    1. 保留 A/B/C/D/E 原本 D+1～D+20 的紅綠藍橘配色。
    2. 保留獲利欄位的粗紅 / 粗綠外框。
    3. 不把 Excel 舊欄位位置的日期格式帶進 Google Sheet，避免「買進均價」再次變成 1899/12/30。
    """
    fmt = {}

    bg = _openpyxl_fill_to_gsheet(cell)
    if bg:
        fmt["backgroundColor"] = bg

    text_format = _openpyxl_font_to_gsheet(cell)
    if text_format:
        fmt["textFormat"] = text_format

    align_format = _openpyxl_alignment_to_gsheet(cell)
    fmt.update(align_format)

    borders = _openpyxl_border_to_gsheet(cell)
    if borders:
        fmt["borders"] = borders

    return fmt


def _format_fields_visual(fmt):
    fields = []

    if "backgroundColor" in fmt:
        fields.append("userEnteredFormat.backgroundColor")
    if "textFormat" in fmt:
        for key in fmt["textFormat"].keys():
            fields.append(f"userEnteredFormat.textFormat.{key}")
    if "horizontalAlignment" in fmt:
        fields.append("userEnteredFormat.horizontalAlignment")
    if "verticalAlignment" in fmt:
        fields.append("userEnteredFormat.verticalAlignment")
    if "wrapStrategy" in fmt:
        fields.append("userEnteredFormat.wrapStrategy")
    if "borders" in fmt:
        fields.append("userEnteredFormat.borders")

    return ",".join(fields)


def _source_record_key_to_excel_row_map(title, source_values, final_headers):
    """
    建立 upsert 後資料列 key -> 原 Excel row number 對照。

    upsert 會把「資料範圍」插到最左邊，但原本 Excel 沒有這欄。
    這個 mapping 用 key 找回原 Excel 列，讓 Google Sheet 可以照原 Excel 的狀態色套回去，
    而不是用錯位的欄位位置硬套樣式。
    """
    header_row_idx = _find_simple_header_row(source_values)
    if header_row_idx is None:
        return {}, []

    source_headers = [str(h).strip() for h in source_values[header_row_idx]]
    while source_headers and source_headers[-1] == "":
        source_headers.pop()

    key_cols = _sheet_upsert_key_columns(title, final_headers)
    if not key_cols:
        return {}, source_headers

    current_scope = get_result_data_scope()
    out = {}

    for offset, raw_row in enumerate(source_values[header_row_idx + 1:], start=header_row_idx + 2):
        row = list(raw_row)
        if len(row) < len(source_headers):
            row = row + [""] * (len(source_headers) - len(row))
        elif len(row) > len(source_headers):
            row = row[:len(source_headers)]

        if not any(str(v).strip() for v in row):
            continue

        rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(source_headers)}
        rec["資料範圍"] = str(rec.get("資料範圍", "") or current_scope).strip()
        key = _record_key(rec, key_cols)
        if any(key):
            out[key] = offset

    return out, source_headers


def _dplus_status_format_from_text(value):
    """
    依 D+ 欄位文字補回基本紅綠色。

    這只是保護機制：
    - 本次新資料會優先從原 Excel 樣式精準套回紅 / 綠 / 藍 / 橘。
    - 舊資料沒有原 Excel 樣式可參照時，至少依文字中的正負報酬補回紅綠底色。
    """
    s = str(value or "").strip()
    if not s:
        return None

    nums = []
    for m in re.finditer(r"([+-]?\d+(?:\.\d+)?)%", s):
        try:
            nums.append(float(m.group(1)))
        except Exception:
            pass

    if not nums:
        return None

    # 只要有正報酬，就以台股習慣用紅色；否則用綠色。
    fill = RED if any(v > 0 for v in nums) else GREEN
    color = _openpyxl_color_to_gsheet(fill.fgColor)
    if not color:
        return None

    return {
        "backgroundColor": color,
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
        "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}},
    }


def apply_upsert_original_excel_visual_style_to_gsheet(ws_xlsx, gws, source_values=None, final_values=None, title=""):
    """
    upsert 工作表專用：在不套錯日期 / 數字格式的前提下，把原本 Excel 的配色套回 Google Sheet。

    修正重點：
    - 不再因為新增「資料範圍」欄位而放棄原本配色。
    - 透過表頭名稱與 upsert key 對齊原 Excel 列，避免欄位位移。
    - 只套背景色、字型、對齊與外框；numberFormat 仍交給後面的日期 / 數字專用函式處理。
    """
    if gws is None or ws_xlsx is None or not source_values or not final_values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    title = safe_worksheet_title(title)
    final_header_row_idx = _find_simple_header_row(final_values)
    if final_header_row_idx is None:
        return

    final_headers = [str(h).strip() for h in final_values[final_header_row_idx]]
    while final_headers and final_headers[-1] == "":
        final_headers.pop()

    source_key_to_row, source_headers = _source_record_key_to_excel_row_map(title, source_values, final_headers)
    if not final_headers:
        return

    source_col_by_header = {h: idx + 1 for idx, h in enumerate(source_headers) if h}
    key_cols = _sheet_upsert_key_columns(title, final_headers)
    if not key_cols:
        return

    current_scope = get_result_data_scope()
    requests = []

    # 先處理表頭：如果原 Excel 表頭有黃底等樣式，依欄名套回；資料範圍欄則維持安全樣式。
    for final_col_idx, header in enumerate(final_headers, start=1):
        source_col_idx = source_col_by_header.get(header)
        if not source_col_idx:
            continue

        src_cell = ws_xlsx.cell(1, source_col_idx)
        fmt = _cell_gsheet_visual_format(src_cell)
        fields = _format_fields_visual(fmt)
        if fmt and fields:
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": final_header_row_idx,
                        "endRowIndex": final_header_row_idx + 1,
                        "startColumnIndex": final_col_idx - 1,
                        "endColumnIndex": final_col_idx,
                    },
                    "cell": {"userEnteredFormat": fmt},
                    "fields": fields,
                }
            })

    final_records = []
    for raw_row in final_values[final_header_row_idx + 1:]:
        row = list(raw_row)
        if len(row) < len(final_headers):
            row = row + [""] * (len(final_headers) - len(row))
        elif len(row) > len(final_headers):
            row = row[:len(final_headers)]

        rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(final_headers)}
        final_records.append((row, rec))

    import json as _json

    for row_offset, (row_values, rec) in enumerate(final_records, start=final_header_row_idx + 2):
        rec_scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip()
        key = _record_key(rec, key_cols)
        src_excel_row = source_key_to_row.get(key)

        run_start = None
        run_fmt = None
        run_key = None

        for final_col_idx in range(1, len(final_headers) + 2):
            fmt = None
            if final_col_idx <= len(final_headers):
                header = final_headers[final_col_idx - 1]
                source_col_idx = source_col_by_header.get(header)

                if src_excel_row and source_col_idx and rec_scope == current_scope:
                    src_cell = ws_xlsx.cell(src_excel_row, source_col_idx)
                    fmt = _cell_gsheet_visual_format(src_cell)
                elif str(header).startswith("D+"):
                    # 舊資料或沒有對到原 Excel key 的資料，至少補回 D+ 欄位紅綠底色。
                    cell_value = row_values[final_col_idx - 1] if final_col_idx - 1 < len(row_values) else ""
                    fmt = _dplus_status_format_from_text(cell_value)

            key_json = _json.dumps(fmt, sort_keys=True, ensure_ascii=False) if fmt else None

            if key_json and run_key is None:
                run_start = final_col_idx
                run_fmt = fmt
                run_key = key_json
            elif key_json and key_json == run_key:
                continue
            else:
                if run_key and run_start is not None and run_fmt:
                    fields = _format_fields_visual(run_fmt)
                    if fields:
                        requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": row_offset - 1,
                                    "endRowIndex": row_offset,
                                    "startColumnIndex": run_start - 1,
                                    "endColumnIndex": final_col_idx - 1,
                                },
                                "cell": {"userEnteredFormat": run_fmt},
                                "fields": fields,
                            }
                        })

                if key_json:
                    run_start = final_col_idx
                    run_fmt = fmt
                    run_key = key_json
                else:
                    run_start = None
                    run_fmt = None
                    run_key = None

    _gsheet_batch_update(requests)

def upload_excel_to_google_sheet(xlsx_path):
    if not GSHEET_RESULT_ENABLED or not gsheet_enabled():
        print("  ⚠️ 未設定 GCP_SERVICE_KEY，略過 Google Sheet 結果同步")
        return

    try:
        from openpyxl import load_workbook

        wb = load_workbook(xlsx_path, data_only=False)

        primary_sh = get_gsheet_spreadsheet()
        if primary_sh is not None:
            print(
                f"  ☁️ Google Sheet 目標試算表：{GOOGLE_SHEET_NAME}" + (f"｜ID={GOOGLE_SHEET_ID}" if GOOGLE_SHEET_ID else "")
            )
            print(
                "  ⚙️ Google Sheet 主表保留設定："
                f"每日賣出明細最近 {GSHEET_DAILY_SELL_KEEP_TRADING_DAYS} 個交易日｜"
                f"TOP15 最近 {GSHEET_TOP15_KEEP_STAT_DATES} 個統計日期｜"
                f"自動封存={'開啟' if GSHEET_RESULT_ARCHIVE_ENABLED else '關閉'}"
            )
            compact_spreadsheet_blank_grid(primary_sh, label="主試算表（同步前）")

        for ws_xlsx in wb.worksheets:
            title = safe_worksheet_title(ws_xlsx.title)

            if should_skip_result_sheet_in_run_mode(title):
                print(
                    f"  ✅ RUN_MODE=1 精選五分點模式：略過同步結果到 Google Sheet：{title}，"
                    "保留既有全分點資料。"
                )
                continue

            values = []

            for row in ws_xlsx.iter_rows(values_only=False):
                row_values = []
                for cell in row:
                    value = cell.value
                    value = normalize_formula_for_gsheet(value)
                    row_values.append(clean_gsheet_value(value))
                values.append(row_values)

            if not values:
                values = [[""]]

            source_values = [list(row) for row in values]
            values = merge_result_values_for_gsheet(title, values)
            values = normalize_result_values_for_comma_numbers(values)

            max_cols = max(max((len(row) for row in values), default=1), 1)
            # 只配置實際需要的格數，避免建立工作表時先多占 100 列 / 20 欄。
            gws = get_or_recreate_result_worksheet(
                title,
                rows=max(len(values), 1),
                cols=max(max_cols, 1),
            )

            if write_values_to_worksheet(gws, values):
                if should_upsert_result_sheet(title):
                    # upsert 表會新增「資料範圍」欄，不能再照原 Excel 欄位位置硬套 numberFormat。
                    # 但原本 A/B/C/D/E 的 D+ 紅綠藍橘配色必須保留，所以改成：
                    # 1. 先清掉舊錯誤 numberFormat
                    # 2. 套基本表格樣式
                    # 3. 依 upsert key 與表頭名稱，把原 Excel 視覺配色精準套回來
                    # 4. 最後再依實際 Google Sheet 表頭重套文字 / 數字 / 日期格式
                    clear_all_number_formats_for_written_range(gws, values=values)
                    apply_safe_result_table_style_to_gsheet(gws, values=values)
                    apply_upsert_original_excel_visual_style_to_gsheet(
                        ws_xlsx,
                        gws,
                        source_values=source_values,
                        final_values=values,
                        title=title,
                    )
                    apply_text_format_to_gsheet(gws, values)
                    apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=values)
                    apply_date_format_to_gsheet(ws_xlsx, gws, values=values)
                    apply_header_widths_to_gsheet(gws, values=values)
                else:
                    apply_excel_style_to_gsheet(ws_xlsx, gws)
                    apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=values)
                    apply_date_format_to_gsheet(ws_xlsx, gws, values=values)
                    apply_header_widths_to_gsheet(gws, values=values)
                print(f"  ☁️ 已同步結果到 Google Sheet：{title}")

        # 每張已寫入的工作表都已由 write_values_to_worksheet() resize 成實際大小；
        # 這裡只回報同步後配置格數，不再重讀全部工作表，避免大型試算表多做一次完整掃描。
        if primary_sh is not None:
            print(f"  🧮 Google Sheet 主試算表同步後配置格數：{spreadsheet_grid_cell_count(primary_sh):,}")

        try:
            sh = get_gsheet_spreadsheet()
            if sh is not None:
                tmp_ws = sh.worksheet("__tmp_delete_guard__")
                if len(sh.worksheets()) > 1:
                    gsheet_api_call("刪除暫時工作表 __tmp_delete_guard__", sh.del_worksheet, tmp_ws)
        except Exception:
            pass

    except Exception as e:
        print(f"  ⚠️ Excel 同步 Google Sheet 失敗：{type(e).__name__}: {e}")



# ══════════════════════════════════════════════════════════════════════
# 快取工具：避免每次重爬舊資料
# ══════════════════════════════════════════════════════════════════════

def cache_enabled():
    return USE_CACHE and not FORCE_FULL_CACHE_REFRESH


def read_cache_csv(path):
    if not cache_enabled():
        return pd.DataFrame()

    # Supabase 啟用時，大型快取優先從 Supabase 讀取。
    # 若 Supabase 尚未有資料，才退回原本 Google Sheet / 本機 CSV，方便第一次遷移。
    df_from_supabase = read_cache_from_supabase(path)

    if df_from_supabase is not None and not df_from_supabase.empty:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            df_from_supabase.to_csv(path, index=False, encoding=CACHE_ENCODING)
        except Exception:
            pass
        return df_from_supabase

    # 本機執行時：優先讀本機快取，並順手上傳到 Google Sheet，方便先把本機快取種到雲端。
    # GitHub Actions 執行時：優先讀 Google Sheet 快取，因為 runner 通常是乾淨環境。
    local_first = os.getenv("GITHUB_ACTIONS", "").strip().lower() != "true"

    if local_first and os.path.exists(path):
        try:
            df = pd.read_csv(path, dtype=str, encoding=CACHE_ENCODING).fillna("")
            if GSHEET_SYNC_CACHE_ON_READ and not supabase_should_skip_gsheet_cache(path):
                write_cache_to_gsheet(df, path)
            return df
        except Exception as e:
            print(f"  ⚠️ 本機快取讀取失敗：{path}，原因：{e}")

    df_from_gsheet = read_cache_from_gsheet(path)

    if df_from_gsheet is not None and not df_from_gsheet.empty:
        try:
            os.makedirs(os.path.dirname(path), exist_ok=True)
            df_from_gsheet.to_csv(path, index=False, encoding=CACHE_ENCODING)
        except Exception:
            pass
        return df_from_gsheet

    if not os.path.exists(path):
        return pd.DataFrame()

    try:
        df = pd.read_csv(path, dtype=str, encoding=CACHE_ENCODING).fillna("")
        if GSHEET_SYNC_CACHE_ON_READ and not supabase_should_skip_gsheet_cache(path):
            write_cache_to_gsheet(df, path)
        return df
    except Exception as e:
        print(f"  ⚠️ 快取讀取失敗：{path}，原因：{e}")
        return pd.DataFrame()


def write_cache_csv(
    df,
    path,
    sync_gsheet=True,
    supabase_df=None,
    sync_supabase=True,
    force_gsheet_sync=False,
    force_supabase_sync=False,
):
    if not USE_CACHE:
        return

    cache_changed = True

    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        new_csv_text = df.to_csv(index=False)
        new_csv_bytes = new_csv_text.encode(CACHE_ENCODING)

        if os.path.exists(path):
            try:
                with open(path, "rb") as f:
                    old_csv_bytes = f.read()

                if old_csv_bytes == new_csv_bytes:
                    cache_changed = False
            except Exception:
                cache_changed = True

        if cache_changed:
            with open(path, "wb") as f:
                f.write(new_csv_bytes)
        else:
            print(f"  ✅ 快取內容未變動：{path}")
    except Exception as e:
        print(f"  ⚠️ 快取寫入失敗：{path}，原因：{e}")
        return

    cache_id = supabase_cache_identifier(path)
    need_supabase_seed = cache_id in _SUPABASE_EMPTY_CACHE_KEYS

    if sync_supabase and (cache_changed or need_supabase_seed or force_supabase_sync):
        upload_df = df if need_supabase_seed or supabase_df is None else supabase_df
        write_cache_to_supabase(upload_df, path, full_df=df)

    if sync_gsheet:
        if supabase_should_skip_gsheet_cache(path):
            print(f"  🗄️ 已啟用 Supabase 快取，略過大型快取同步到 Google Sheet：{cache_sheet_name_from_path(path)}")
        else:
            if cache_changed or force_gsheet_sync:
                write_cache_to_gsheet(df, path)
            else:
                print(f"  ✅ 快取內容未變動，略過 Google Sheet 同步：{path}")

def load_price_cache():
    """讀取價格持久化快取，並只保留最近 PRICE_RETENTION_TRADING_DAYS 個交易日。"""
    df = read_cache_csv(PRICE_CACHE_PATH)

    if df.empty:
        return {}

    required_cols = ["代號", "日期", "收盤價"]
    for col in required_cols:
        if col not in df.columns:
            print(f"  ⚠️ 價格快取欄位不完整，缺少：{col}")
            return {}

    df, prune_stats = prune_price_cache_dataframe(df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 價格快取記憶體清理：移除 {prune_stats['removed']:,} 筆｜"
            f"保留最近 {PRICE_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    price_cache = {}

    for row in df.itertuples(index=False):
        row = row._asdict()
        code = normalize_price_code(row.get("代號", ""))

        if not code:
            continue

        date_str = normalize_date_str(row.get("日期", ""))
        dt = parse_date(date_str)

        if not dt:
            continue

        price = safe_price_float(row.get("收盤價", ""))

        if price is None:
            continue

        price_cache.setdefault(code, {})[dt.strftime("%Y/%m/%d")] = price

    return price_cache

def append_price_cache_rows_to_gsheet(canonical_price_cache, changed_codes):
    """
    將本次新增 / 更新的價格資料增量 append 到 Google Sheet「快取_價格」。

    本機 price_cache.csv 仍會完整保存，因此不會遺失舊價格資料；
    Google Sheet 端改用 append 可避免每次把十幾萬筆價格快取整張重寫。
    讀回時 load_price_cache() 仍會依「代號 + 日期」覆蓋同日價格，因此少量重複列不會影響計算正確性。
    """
    if not PRICE_CACHE_GSHEET_INCREMENTAL_APPEND:
        return False

    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return False

    if not canonical_price_cache or not changed_codes:
        return False

    changed_codes = {
        normalize_price_code(code)
        for code in changed_codes
        if normalize_price_code(code)
    }

    if not changed_codes:
        return False

    rows = []
    for code in sorted(changed_codes):
        prices = canonical_price_cache.get(code, {})
        if not prices:
            continue

        for date_str in sorted(prices.keys()):
            dt = parse_date(date_str)
            price = safe_price_float(prices.get(date_str))

            if not dt or price is None:
                continue

            rows.append({
                "代號": code,
                "日期": dt.strftime("%Y/%m/%d"),
                "收盤價": price,
            })

    if not rows:
        return False

    try:
        title = cache_sheet_name_from_path(PRICE_CACHE_PATH)
        headers = ["代號", "日期", "收盤價"]
        ws = get_or_create_worksheet(title, rows=max(len(rows) + 1, 100), cols=len(headers))
        if ws is None:
            return False

        try:
            existing_headers = ws.row_values(1)
        except Exception:
            existing_headers = []

        if not existing_headers or all(str(x).strip() == "" for x in existing_headers):
            header_values = normalize_gsheet_values_for_text_columns([headers])
            gsheet_api_call(
                f"寫入價格快取表頭 {title}",
                ws.update,
                values=header_values,
                range_name="A1",
                value_input_option="USER_ENTERED",
            )

        values = [[row.get(h, "") for h in headers] for row in rows]
        normalized = normalize_gsheet_values_for_text_columns([headers] + values)[1:]

        for start in range(0, len(normalized), GSHEET_CHUNK_ROWS):
            chunk = normalized[start:start + GSHEET_CHUNK_ROWS]
            gsheet_api_call(
                f"增量追加價格快取 {title} {start + 1}-{start + len(chunk)}",
                ws.append_rows,
                chunk,
                value_input_option="USER_ENTERED",
            )

        print(f"  ☁️ 已增量追加價格快取到 Google Sheet：{title}，本次 {len(rows):,} 筆")
        return True
    except Exception as e:
        print(f"  ⚠️ 價格快取增量追加到 Google Sheet 失敗：{type(e).__name__}: {e}")
        return False


def save_price_cache(price_cache, changed_codes=None):
    """寫入價格快取，並自動裁切為最近 PRICE_RETENTION_TRADING_DAYS 個交易日。"""
    if not USE_CACHE or not price_cache:
        return

    canonical = {}

    for code, prices in price_cache.items():
        norm_code = normalize_price_code(code)

        if not norm_code or not prices:
            continue

        for date_str, price in prices.items():
            dt = parse_date(date_str)
            price = safe_price_float(price)

            if not dt or price is None:
                continue

            canonical.setdefault(norm_code, {})[dt.strftime("%Y/%m/%d")] = price

    rows = []

    for code in sorted(canonical.keys()):
        for date_str in sorted(canonical[code].keys()):
            rows.append({
                "代號": code,
                "日期": date_str,
                "收盤價": canonical[code][date_str],
            })

    if not rows:
        return

    df = pd.DataFrame(rows, columns=["代號", "日期", "收盤價"])
    df, prune_stats = prune_price_cache_dataframe(df)

    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 價格快取自動清理：移除 {prune_stats['removed']:,} 筆｜"
            f"保留最近 {PRICE_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    canonical = {}
    for row in df.itertuples(index=False):
        row = row._asdict()
        code = normalize_price_code(row.get("代號", ""))
        date_str = normalize_date_str(row.get("日期", ""))
        price = safe_price_float(row.get("收盤價", ""))
        if code and parse_date(date_str) and price is not None:
            canonical.setdefault(code, {})[date_str] = price

    normalized_changed_codes = {
        normalize_price_code(code)
        for code in (changed_codes or [])
        if normalize_price_code(code)
    }

    if normalized_changed_codes:
        supabase_delta_df = df[df["代號"].astype(str).isin(normalized_changed_codes)].copy()
    else:
        supabase_delta_df = df.copy()

    use_incremental_gsheet = (
        PRICE_CACHE_GSHEET_INCREMENTAL_APPEND
        and bool(normalized_changed_codes)
        and len(df) >= PRICE_CACHE_FULL_SYNC_THRESHOLD_ROWS
        and prune_stats.get("removed", 0) == 0
    )

    if use_incremental_gsheet:
        write_cache_csv(
            df,
            PRICE_CACHE_PATH,
            sync_gsheet=False,
            supabase_df=supabase_delta_df,
        )
        if supabase_should_skip_gsheet_cache(PRICE_CACHE_PATH):
            print("  🗄️ 已啟用 Supabase 快取，略過價格快取增量同步到 Google Sheet。")
        elif not append_price_cache_rows_to_gsheet(canonical, normalized_changed_codes):
            write_cache_to_gsheet(df, PRICE_CACHE_PATH)
    else:
        write_cache_csv(
            df,
            PRICE_CACHE_PATH,
            supabase_df=supabase_delta_df,
            force_gsheet_sync=bool(prune_stats.get("removed", 0) and CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC),
            force_supabase_sync=bool(prune_stats.get("removed", 0)),
        )

    print(f"  💾 已更新價格快取：{PRICE_CACHE_PATH}，共 {len(df):,} 筆")

def get_cached_prices_for_code(price_cache, code):
    """
    從價格快取中取出指定代號的價格。

    同時支援補零版與去零版查找，最後回傳單一合併後 dict。
    """
    out = {}
    norm_code = normalize_price_code(code)

    if not norm_code:
        return out

    lookup_codes = []
    for c in price_code_variants(norm_code):
        if c and c not in lookup_codes:
            lookup_codes.append(c)

        no_zero = c.lstrip("0")
        if no_zero and no_zero not in lookup_codes:
            lookup_codes.append(no_zero)

        norm_c = normalize_price_code(c)
        if norm_c and norm_c not in lookup_codes:
            lookup_codes.append(norm_c)

    for c in lookup_codes:
        cached = price_cache.get(c)

        if not cached:
            continue

        out = merge_price_dicts(out, cached)

    return out


def add_price_aliases(price_cache, code, prices):
    """
    在記憶體 price_cache 中建立補零 / 去零別名，
    避免 Excel 或 pandas 吃掉前導 0 時查不到。
    """
    if not prices:
        prices = {}

    norm_code = normalize_price_code(code)

    if not norm_code:
        return

    price_cache[norm_code] = prices

    no_zero = norm_code.lstrip("0")

    if no_zero:
        price_cache[no_zero] = prices

    raw_code = str(code).strip()

    if raw_code:
        price_cache[raw_code] = prices



def candidate_key_from_tuple(c):
    return (str(c[0]).strip(), str(c[6]).strip())


def candidate_key_from_values(warrant_code, broker_code):
    return (str(warrant_code).strip(), str(broker_code).strip())


def save_warrants_cache(warrants):
    if not USE_CACHE or not warrants:
        return

    df = pd.DataFrame(warrants)

    wanted_cols = ["代號", "名稱", "標的股", "標的名稱"]
    for col in wanted_cols:
        if col not in df.columns:
            df[col] = ""

    df = fix_known_underlying_info_dataframe(df, "名稱", "標的股", "標的名稱")
    df["代號"] = df["代號"].astype(str).str.strip()
    df = df[df["代號"] != ""].drop_duplicates(subset=["代號"], keep="last")
    df = df.sort_values("代號").reset_index(drop=True)

    write_cache_csv(df[wanted_cols], WARRANTS_CACHE_PATH, force_supabase_sync=FORCE_FULL_CACHE_REFRESH)
    print(f"  💾 已更新權證清單快取：{WARRANTS_CACHE_PATH}，目前有效 {len(df):,} 支")

def load_warrants_cache():
    df = read_cache_csv(WARRANTS_CACHE_PATH)

    if df.empty:
        return []

    required_cols = ["代號", "名稱", "標的股", "標的名稱"]
    for col in required_cols:
        if col not in df.columns:
            return []

    warrants = []
    for row in df.itertuples(index=False):
        row = row._asdict()
        code = str(row["代號"]).strip()
        name = str(row["名稱"]).strip()

        if not code or not name:
            continue

        underlying_code, underlying_name = correct_underlying_info_by_warrant_name(
            name,
            row.get("標的股", ""),
            row.get("標的名稱", ""),
        )

        warrants.append({
            "代號": code,
            "名稱": name,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
        })

    return warrants


def save_broker_map_cache(broker_map):
    if not USE_CACHE or not broker_map:
        return

    rows = []
    for label, (name, code) in broker_map.items():
        rows.append({
            "分點": label,
            "分點名稱": name,
            "券商代號": code,
        })

    df = pd.DataFrame(rows)
    write_cache_csv(df, BROKER_MAP_CACHE_PATH)
    print(f"  💾 已更新分點代號快取：{BROKER_MAP_CACHE_PATH}")


def load_broker_map_cache():
    df = read_cache_csv(BROKER_MAP_CACHE_PATH)

    if df.empty:
        return {}

    required_cols = ["分點", "分點名稱", "券商代號"]
    for col in required_cols:
        if col not in df.columns:
            return {}

    broker_map = {}

    for row in df.itertuples(index=False):
        row = row._asdict()
        label = str(row["分點"]).strip()
        name = str(row["分點名稱"]).strip()
        code = str(row["券商代號"]).strip()

        if label and name and code and label in TARGET_PATTERNS:
            broker_map[label] = (name, code)

    missing = [k for k in TARGET_PATTERNS if k not in broker_map]
    if missing:
        print(f"  ⚠️ 分點代號快取不完整，缺少：{missing}")
        return {}

    return broker_map


def save_candidates_cache(candidates):
    if not USE_CACHE or not candidates:
        return False

    rows = []

    for c in candidates:
        underlying_code, underlying_name = correct_underlying_info_by_warrant_name(c[1], c[2], c[3])

        rows.append({
            "權證代號": c[0],
            "權證名稱": c[1],
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "分點": c[4],
            "分點名稱": c[5],
            "券商代號": c[6],
        })

    df = pd.DataFrame(rows)
    valid_codes = CURRENT_LIVE_WARRANT_CODES if LIVE_WARRANT_SNAPSHOT_READY else None
    df, prune_stats = prune_candidates_dataframe(
        df,
        path=CANDIDATES_CACHE_PATH,
        valid_warrant_codes=valid_codes,
    )
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 候選組合自動清理：移除 {prune_stats['removed']:,} 組｜"
            f"移除分點 {prune_stats.get('removed_broker', 0):,}｜"
            f"失效權證 {prune_stats.get('removed_warrant', 0):,}"
        )

    write_cache_csv(
        df,
        CANDIDATES_CACHE_PATH,
        force_gsheet_sync=bool(prune_stats.get("removed", 0) and CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC),
        force_supabase_sync=FORCE_FULL_CACHE_REFRESH,
    )
    print(f"  💾 已更新候選組合快取：{CANDIDATES_CACHE_PATH}，共 {len(df):,} 組")
    return True


def load_candidates_cache():
    df = read_cache_csv(CANDIDATES_CACHE_PATH)

    if df.empty:
        return []

    required_cols = ["權證代號", "權證名稱", "標的股", "標的名稱", "分點", "分點名稱", "券商代號"]
    for col in required_cols:
        if col not in df.columns:
            return []

    valid_codes = CURRENT_LIVE_WARRANT_CODES if LIVE_WARRANT_SNAPSHOT_READY else None
    df, prune_stats = prune_candidates_dataframe(
        df,
        path=CANDIDATES_CACHE_PATH,
        valid_warrant_codes=valid_codes,
    )
    if prune_stats.get("removed", 0) > 0:
        print(f"  🧹 讀取候選快取時略過已移除分點／失效權證：{prune_stats['removed']:,} 組")

    candidates = []

    for row in df.itertuples(index=False):
        row = row._asdict()
        warrant_code = str(row["權證代號"]).strip()
        broker_code = str(row["券商代號"]).strip()

        if not warrant_code or not broker_code:
            continue

        warrant_name = str(row["權證名稱"]).strip()
        underlying_code, underlying_name = correct_underlying_info_by_warrant_name(
            warrant_name,
            row.get("標的股", ""),
            row.get("標的名稱", ""),
        )

        candidates.append((
            warrant_code,
            warrant_name,
            underlying_code,
            underlying_name,
            str(row["分點"]).strip(),
            str(row["分點名稱"]).strip(),
            broker_code,
        ))


    return candidates


# ══════════════════════════════════════════════════════════════════════
# daily / repair 工作流：prescan 狀態與 refresh keys 快取
# ══════════════════════════════════════════════════════════════════════

def workflow_is_daily():
    return WORKFLOW_MODE == "daily"


def workflow_is_repair():
    return WORKFLOW_MODE == "repair"


def workflow_is_longterm():
    return WORKFLOW_MODE == "longterm"


def current_prescan_scope():
    return "selected5" if RUN_MODE == 1 else "all"


def current_prescan_target_date(target_date=None):
    return normalize_price_prefetch_target_date(target_date)


def load_prescan_status_df():
    df = read_cache_csv(PRESCAN_STATUS_PATH)
    if df is None or df.empty:
        return pd.DataFrame()
    return df.fillna("")


def write_prescan_status(status, target_date=None, reason="", **extra):
    """記錄今日 50 天候選掃描狀態；只有 status=success 才允許日內跳過 prescan。"""
    if not USE_CACHE:
        return

    target_date = current_prescan_target_date(target_date)
    scope = current_prescan_scope()
    latest_label = ""
    if PRESCAN_LATEST_ACTIVITY_DATE:
        latest_label = PRESCAN_LATEST_ACTIVITY_DATE.strftime("%Y/%m/%d")

    headers = [
        "日期", "資料範圍", "RUN_MODE", "狀態", "最新分點活動日", "今日分點資料是否出現",
        "候選組合數", "API4直接候選數", "允許缺快取補抓數", "候選快取是否寫入成功",
        "更新時間", "錯誤原因",
    ]

    old_df = read_cache_csv(PRESCAN_STATUS_PATH)
    rows = []
    if old_df is not None and not old_df.empty:
        for _, old_row in old_df.fillna("").iterrows():
            rows.append({h: old_row.get(h, "") for h in headers})

    rec = {h: "" for h in headers}
    rec.update({
        "日期": target_date,
        "資料範圍": scope,
        "RUN_MODE": str(RUN_MODE),
        "狀態": str(status or "").strip(),
        "最新分點活動日": latest_label,
        "今日分點資料是否出現": "1" if has_today_broker_data_from_prescan(target_date) else "0",
        "候選組合數": extra.get("candidate_count", ""),
        "API4直接候選數": extra.get("direct_candidate_count", len(PRESCAN_REFRESH_KEYS or [])),
        "允許缺快取補抓數": extra.get("missing_fetch_key_count", len(PRESCAN_MISSING_FETCH_KEYS or [])),
        "候選快取是否寫入成功": "1" if extra.get("candidate_cache_saved", False) else "0",
        "更新時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S"),
        "錯誤原因": reason,
    })
    rows.append(rec)

    df = pd.DataFrame(rows, columns=headers).fillna("")
    df = df.drop_duplicates(subset=["日期", "資料範圍", "RUN_MODE"], keep="last").reset_index(drop=True)
    write_cache_csv(df, PRESCAN_STATUS_PATH)


def find_valid_prescan_success_record(target_date=None):
    """只接受同日、同 scope、同 RUN_MODE 且 status=success 的 prescan 狀態。"""
    if workflow_is_repair():
        return None

    target_date = current_prescan_target_date(target_date)
    scope = current_prescan_scope()
    df = load_prescan_status_df()
    if df.empty:
        return None

    for row in df.itertuples(index=False):
        rd = row._asdict()
        if normalize_date_str(rd.get("日期", "")) != target_date:
            continue
        if str(rd.get("資料範圍", "")).strip() != scope:
            continue
        if str(rd.get("RUN_MODE", "")).strip() != str(RUN_MODE):
            continue
        if str(rd.get("狀態", "")).strip().lower() != "success":
            continue
        return rd

    return None


def _split_candidate_key(key):
    if isinstance(key, (tuple, list)) and len(key) >= 2:
        return str(key[0]).strip(), str(key[1]).strip()
    s = str(key or "").strip()
    for sep in ("|", "::", ","):
        if sep in s:
            a, b = s.split(sep, 1)
            return a.strip(), b.strip()
    return "", ""


def write_prescan_refresh_keys(refresh_keys=None, missing_fetch_keys=None, target_date=None):
    """將 PRESCAN_REFRESH_KEYS / PRESCAN_MISSING_FETCH_KEYS 獨立保存，避免狀態表單格塞爆。"""
    if not USE_CACHE:
        return False

    target_date = current_prescan_target_date(target_date)
    scope = current_prescan_scope()
    now_s = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    rows = []

    def add_rows(keys, key_type):
        for key in sorted(keys or []):
            warrant_code, broker_code = _split_candidate_key(key)
            if not warrant_code or not broker_code:
                continue
            rows.append({
                "日期": target_date,
                "資料範圍": scope,
                "RUN_MODE": str(RUN_MODE),
                "key類型": key_type,
                "權證代號": warrant_code,
                "券商代號": broker_code,
                "更新時間": now_s,
            })

    add_rows(refresh_keys or set(), "refresh")
    add_rows(missing_fetch_keys or set(), "missing_fetch")

    headers = ["日期", "資料範圍", "RUN_MODE", "key類型", "權證代號", "券商代號", "更新時間"]
    old_df = read_cache_csv(PRESCAN_REFRESH_KEYS_PATH)
    keep_rows = []
    if old_df is not None and not old_df.empty:
        for _, old_row in old_df.fillna("").iterrows():
            rd = {h: old_row.get(h, "") for h in headers}
            same_scope = (
                normalize_date_str(rd.get("日期", "")) == target_date
                and str(rd.get("資料範圍", "")).strip() == scope
                and str(rd.get("RUN_MODE", "")).strip() == str(RUN_MODE)
            )
            if not same_scope:
                keep_rows.append(rd)

    df = pd.DataFrame(keep_rows + rows, columns=headers).fillna("")
    write_cache_csv(df, PRESCAN_REFRESH_KEYS_PATH)
    print(f"  💾 已保存 prescan keys：refresh {len(refresh_keys or []):,} 組｜missing_fetch {len(missing_fetch_keys or []):,} 組")
    return True


def load_prescan_refresh_keys(target_date=None):
    """讀回同日、同 scope、同 RUN_MODE 的 refresh keys；不符即視為無效。"""
    target_date = current_prescan_target_date(target_date)
    scope = current_prescan_scope()
    df = read_cache_csv(PRESCAN_REFRESH_KEYS_PATH)
    if df is None or df.empty:
        return set(), set(), False

    required_cols = {"日期", "資料範圍", "RUN_MODE", "key類型", "權證代號", "券商代號"}
    if not required_cols.issubset(set(df.columns)):
        return set(), set(), False

    refresh_keys = set()
    missing_keys = set()
    for _, row in df.fillna("").iterrows():
        if normalize_date_str(row.get("日期", "")) != target_date:
            continue
        if str(row.get("資料範圍", "")).strip() != scope:
            continue
        if str(row.get("RUN_MODE", "")).strip() != str(RUN_MODE):
            continue
        key = candidate_key_from_values(row.get("權證代號", ""), row.get("券商代號", ""))
        if not key:
            continue
        key_type = str(row.get("key類型", "")).strip().lower()
        if key_type == "refresh":
            refresh_keys.add(key)
        elif key_type == "missing_fetch":
            missing_keys.add(key)

    return refresh_keys, missing_keys, bool(refresh_keys or missing_keys)


def try_load_prescan_success_cache(broker_map, target_date=None):
    """daily 第二次執行時，若 prescan 狀態成功，直接讀候選快取與 keys。"""
    global PRESCAN_REFRESH_KEYS, PRESCAN_MISSING_FETCH_KEYS
    global PRESCAN_LATEST_ACTIVITY_DATE, PRESCAN_TODAY_ACTIVITY_FOUND, PRESCAN_STATUS_LAST_RECORD

    record = find_valid_prescan_success_record(target_date)
    if not record:
        return None

    refresh_keys, missing_keys, keys_ok = load_prescan_refresh_keys(target_date)
    if not keys_ok:
        print("  ⚠️ 找到 prescan success，但找不到同日 / 同 scope 的 prescan keys；本次重新掃描，避免 API5 漏補。")
        return None

    cached_candidates = filter_candidates_by_broker_map(load_candidates_cache(), broker_map)
    if not cached_candidates:
        print("  ⚠️ 找到 prescan success，但候選快取為空；本次重新掃描。")
        return None

    PRESCAN_REFRESH_KEYS = refresh_keys
    PRESCAN_MISSING_FETCH_KEYS = missing_keys
    latest_dt = parse_date(record.get("最新分點活動日", ""))
    PRESCAN_LATEST_ACTIVITY_DATE = latest_dt
    PRESCAN_TODAY_ACTIVITY_FOUND = str(record.get("今日分點資料是否出現", "")).strip() in ("1", "true", "True", "是", "已出現")
    PRESCAN_STATUS_LAST_RECORD = dict(record)

    print(
        f"  ✅ 今日 prescan 已成功完成，跳過 50 天掃描："
        f"候選 {len(cached_candidates):,} 組｜refresh keys {len(PRESCAN_REFRESH_KEYS):,}｜missing keys {len(PRESCAN_MISSING_FETCH_KEYS):,}"
    )
    return cached_candidates


def _select_light_probe_warrants_from_history(warrants, broker_map):
    """從最近 3~5 個交易日有活動的權證挑探測樣本；沒有樣本時回傳空列表。"""
    history_df = load_history_cache()
    if history_df is None or history_df.empty:
        return []

    required_cols = {"權證代號", "券商代號", "日期"}
    if not required_cols.issubset(set(history_df.columns)):
        return []

    broker_codes = {str(code).strip() for _, code in broker_map.values() if str(code).strip()}
    df = history_df[["權證代號", "券商代號", "日期"]].copy().fillna("")
    df["日期_dt"] = df["日期"].apply(parse_date)
    df = df[df["日期_dt"].notna()]
    if broker_codes:
        df = df[df["券商代號"].astype(str).str.strip().isin(broker_codes)]
    if df.empty:
        return []

    unique_dates = sorted({d.date() for d in df["日期_dt"] if d})
    recent_dates = set(unique_dates[-max(DAILY_LIGHT_PROBE_HISTORY_TRADING_DAYS, 1):])
    df = df[df["日期_dt"].apply(lambda d: d.date() in recent_dates if d else False)]
    if df.empty:
        return []

    # 依最近日期與出現次數排序，優先挑近期確定有活動的權證。
    grouped = df.groupby("權證代號", dropna=False).agg(
        latest_dt=("日期_dt", "max"),
        count=("日期", "count"),
    ).reset_index()
    grouped["權證代號"] = grouped["權證代號"].astype(str).str.strip()
    grouped = grouped[grouped["權證代號"] != ""]
    grouped = grouped.sort_values(["latest_dt", "count"], ascending=[False, False])

    warrant_map = {str(w.get("代號", "")).strip(): w for w in warrants or []}
    sample = []
    for code in grouped["權證代號"].tolist():
        w = warrant_map.get(code)
        if w:
            sample.append(w)
        if len(sample) >= max(DAILY_LIGHT_PROBE_SAMPLE_SIZE, 1):
            break

    return sample


def light_probe_today_broker_data(warrants, broker_map, target_date=None):
    """
    daily 正式 prescan 前的輕量探測。
    回傳：True=今日資料已出現；False=尚未出現；None=無樣本，應直接正式 prescan。
    """
    global PRESCAN_LATEST_ACTIVITY_DATE, PRESCAN_TODAY_ACTIVITY_FOUND

    if not workflow_is_daily() or not DAILY_LIGHT_PROBE_ENABLED:
        return None

    target_date = current_prescan_target_date(target_date)
    target_dt = parse_date(target_date)
    if not target_dt:
        return None

    sample_warrants = _select_light_probe_warrants_from_history(warrants, broker_map)
    if not sample_warrants:
        print("  ⚠️ 輕量探測：快取_分點歷史沒有可用樣本，跳過探測並進入正式 prescan。")
        return None

    broker_codes_set = {code for _, code in broker_map.values()}
    start_s = (target_dt - timedelta(days=max(DAILY_LIGHT_PROBE_SCAN_DAYS, 1))).strftime("%Y/%m/%d")
    end_s = target_dt.strftime("%Y/%m/%d")
    target_s = target_dt.strftime("%Y/%m/%d")
    found_today = False
    latest_dt = None
    done = 0
    timeout_or_error = 0
    empty_rows = 0

    print(f"【Step 3a-0】輕量探測今日分點資料：樣本 {len(sample_warrants):,} 支，區間 {start_s}~{end_s}")

    def probe_one(w):
        local_latest = None
        local_today = False
        row_count = 0
        try:
            rows = api4_get(w["代號"], start_s, end_s)
        except Exception:
            return None, False, 0, True

        for row in rows or []:
            row_count += 1
            bcode = row.get("V2", "")
            if bcode not in broker_codes_set:
                continue
            row_date = normalize_date_str(row.get("V1", ""))
            row_dt = parse_date(row_date)
            if row_dt and (local_latest is None or row_dt > local_latest):
                local_latest = row_dt
            if row_date == target_s:
                local_today = True
        return local_latest, local_today, row_count, False

    stop_probe_event = threading.Event()

    def probe_one_with_stop(w):
        if stop_probe_event.is_set():
            return None, False, 0, False
        return probe_one(w)

    ex = ThreadPoolExecutor(max_workers=max(1, DAILY_LIGHT_PROBE_WORKERS))
    futures = {ex.submit(probe_one_with_stop, w): w for w in sample_warrants}
    try:
        for future in as_completed(futures):
            done += 1
            try:
                row_latest, row_today, row_count, had_error = future.result()
            except Exception:
                row_latest, row_today, row_count, had_error = None, False, 0, True

            if had_error:
                timeout_or_error += 1
            if not row_count:
                empty_rows += 1
            if row_latest and (latest_dt is None or row_latest > latest_dt):
                latest_dt = row_latest
            if row_today:
                found_today = True
                stop_probe_event.set()
                # 已確認今日資料出現，取消尚未開始的探測請求，避免探測階段打好打滿。
                try:
                    ex.shutdown(wait=False, cancel_futures=True)
                except TypeError:
                    ex.shutdown(wait=False)
                break
    finally:
        if not found_today:
            ex.shutdown(wait=True)

    if latest_dt and (PRESCAN_LATEST_ACTIVITY_DATE is None or latest_dt > PRESCAN_LATEST_ACTIVITY_DATE):
        PRESCAN_LATEST_ACTIVITY_DATE = latest_dt
    if found_today:
        PRESCAN_TODAY_ACTIVITY_FOUND = True

    latest_label = latest_dt.strftime("%Y/%m/%d") if latest_dt else "-"
    print(
        f"  ✅ 輕量探測完成：今日資料 {'已出現' if found_today else '尚未出現'}｜"
        f"最新活動日 {latest_label}｜錯誤/逾時 {timeout_or_error}｜空回應 {empty_rows}"
    )
    return bool(found_today)


def merge_candidates(old_candidates, new_candidates):
    merged = {}

    for c in old_candidates:
        merged[candidate_key_from_tuple(c)] = c

    for c in new_candidates:
        merged[candidate_key_from_tuple(c)] = c

    return list(merged.values())


def load_history_cache():
    df = read_cache_csv(HISTORY_CACHE_PATH)

    if df.empty:
        return pd.DataFrame()

    required_cols = [
        "權證代號", "權證名稱", "標的股", "標的名稱",
        "分點", "分點名稱", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
        "買超股數", "買超金額",
    ]

    for col in required_cols:
        if col not in df.columns:
            print(f"  ⚠️ 原始分點資料快取欄位不完整，缺少：{col}")
            return pd.DataFrame()

    out_df = df[required_cols].copy()
    out_df = fix_known_underlying_info_dataframe(out_df, "權證名稱", "標的股", "標的名稱")
    out_df["日期"] = out_df["日期"].map(normalize_date_str)
    out_df = out_df.drop_duplicates(
        subset=["權證代號", "券商代號", "日期"],
        keep="last"
    ).reset_index(drop=True)

    out_df, prune_stats = prune_history_cache_dataframe(out_df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 分點歷史記憶體清理：移除 {prune_stats['removed']:,} 筆｜"
            f"超期/異常日期 {prune_stats.get('removed_date', 0):,}｜"
            f"已移除分點 {prune_stats.get('removed_broker', 0):,}｜"
            f"保留最近 {HISTORY_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    return out_df

def history_cache_keys(history_df):
    keys = set()

    if history_df is None or history_df.empty:
        return keys

    for row in history_df[["權證代號", "券商代號"]].drop_duplicates().itertuples(index=False):
        row_dict = row._asdict()
        keys.add(candidate_key_from_values(row_dict["權證代號"], row_dict["券商代號"]))

    return keys


def history_cache_latest_dates(history_df):
    """
    回傳每組「權證代號 + 券商代號」在歷史快取中的最後日期。
    有快取且最後日期夠新時，就不用重打 API5。
    """
    latest = {}

    if history_df is None or history_df.empty:
        return latest

    needed = {"權證代號", "券商代號", "日期"}
    if not needed.issubset(set(history_df.columns)):
        return latest

    df = history_df[["權證代號", "券商代號", "日期"]].copy().fillna("")
    df["日期"] = df["日期"].map(normalize_date_str)

    for (warrant_code, broker_code), g in df.groupby(["權證代號", "券商代號"], dropna=False):
        dates = []
        for d in g["日期"].tolist():
            dt = parse_date(d)
            if dt:
                dates.append(dt)
        if dates:
            latest[candidate_key_from_values(warrant_code, broker_code)] = max(dates)

    return latest


def get_incremental_refresh_target_dt():
    lag_days = max(int(CACHE_INCREMENTAL_REFRESH_LAG_DAYS or 0), 0)
    return datetime.today() - timedelta(days=lag_days)


def should_fetch_candidate_incremental(key, history_keys, history_latest_map, direct_refresh_keys, missing_fetch_keys=None):
    """
    增量抓取判斷：
    1. 沒快取：只有本次近期活動候選 / 活動標的擴展候選才抓。
       這可以清掉舊候選快取裡大量從未成交的空候選。
    2. 關閉增量：維持舊邏輯，direct refresh 命中就抓。
    3. 有快取但不是 API4 直接近期活動候選：不抓，直接用快取。
    4. 有快取且 API4 直接近期活動候選：只有快取最後日期落後目標日期才抓。
    """
    missing_fetch_keys = missing_fetch_keys or set()

    if key not in history_keys:
        if not CACHE_INCREMENTAL_UPDATE_ENABLED:
            return True
        return key in missing_fetch_keys

    if not CACHE_INCREMENTAL_UPDATE_ENABLED:
        return key in direct_refresh_keys

    if key not in direct_refresh_keys:
        return False

    latest_dt = history_latest_map.get(key)
    if latest_dt is None:
        return True

    target_dt = get_incremental_refresh_target_dt()
    return latest_dt.date() < target_dt.date()


def item_to_history_rows(item):
    rows = []
    df = item["df"]

    underlying_code, underlying_name = correct_underlying_info_by_warrant_name(
        item.get("warrant_name", ""),
        item.get("underlying_code", ""),
        item.get("underlying_name", ""),
    )

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        rows.append({
            "權證代號": item["warrant_code"],
            "權證名稱": item["warrant_name"],
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "分點": item["broker_label"],
            "分點名稱": item["broker_name"],
            "券商代號": item["broker_code"],
            "日期": normalize_date_str(row_dict["日期"]),
            "買進股數": int(row_dict["買進股數"]),
            "賣出股數": int(row_dict["賣出股數"]),
            "買進金額": int(row_dict["買進金額"]),
            "賣出金額": int(row_dict["賣出金額"]),
            "買超股數": int(row_dict["買超股數"]),
            "買超金額": int(row_dict["買超金額"]),
        })

    return rows


def merge_items_into_history_cache(history_df, new_items):
    if not new_items:
        return history_df if history_df is not None else pd.DataFrame()

    new_rows = []
    new_keys = set()

    for item in new_items:
        new_keys.add(candidate_key_from_values(item["warrant_code"], item["broker_code"]))
        new_rows.extend(item_to_history_rows(item))

    new_df = pd.DataFrame(new_rows)

    if new_df.empty:
        return history_df if history_df is not None else pd.DataFrame()

    if history_df is None or history_df.empty:
        combined = new_df
    else:
        history_df = history_df.copy()
        remove_mask = pd.Series(
            [candidate_key_from_values(w, b) in new_keys for w, b in zip(history_df["權證代號"], history_df["券商代號"])],
            index=history_df.index
        )
        old_keep_df = history_df[~remove_mask].copy()
        combined = pd.concat([old_keep_df, new_df], ignore_index=True)

    numeric_cols = ["買進股數", "賣出股數", "買進金額", "賣出金額", "買超股數", "買超金額"]
    for col in numeric_cols:
        combined[col] = pd.to_numeric(combined[col], errors="coerce").fillna(0).astype(int)

    combined["日期"] = combined["日期"].map(normalize_date_str)

    combined = combined.drop_duplicates(
        subset=["權證代號", "券商代號", "日期"],
        keep="last"
    )

    combined = combined.sort_values(
        ["權證代號", "券商代號", "日期"]
    ).reset_index(drop=True)

    return combined


def append_history_rows_to_gsheet(new_rows):
    """
    將本次 API5 新增 / 更新的原始分點資料增量 append 到 Google Sheet。
    下次讀回時 load_history_cache() 會用 權證代號 + 券商代號 + 日期 去重，保留最後一筆。
    """
    if not HISTORY_GSHEET_INCREMENTAL_APPEND:
        return False

    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return False

    if not new_rows:
        return False

    try:
        title = cache_sheet_name_from_path(HISTORY_CACHE_PATH)
        headers = [
            "權證代號", "權證名稱", "標的股", "標的名稱",
            "分點", "分點名稱", "券商代號", "日期",
            "買進股數", "賣出股數", "買進金額", "賣出金額",
            "買超股數", "買超金額",
        ]

        ws = get_or_create_worksheet(title, rows=max(len(new_rows) + 1, 100), cols=len(headers))
        if ws is None:
            return False

        try:
            existing_headers = ws.row_values(1)
        except Exception:
            existing_headers = []

        if not existing_headers or all(str(x).strip() == "" for x in existing_headers):
            header_values = normalize_gsheet_values_for_text_columns([headers])
            gsheet_api_call(
                f"寫入原始分點歷史快取表頭 {title}",
                ws.update,
                values=header_values,
                range_name="A1",
                value_input_option="USER_ENTERED",
            )

        values = []
        for row in new_rows:
            values.append([row.get(h, "") for h in headers])

        normalized = normalize_gsheet_values_for_text_columns([headers] + values)[1:]

        for start in range(0, len(normalized), GSHEET_CHUNK_ROWS):
            chunk = normalized[start:start + GSHEET_CHUNK_ROWS]
            gsheet_api_call(
                f"增量追加原始分點歷史 {title} {start + 1}-{start + len(chunk)}",
                ws.append_rows,
                chunk,
                value_input_option="USER_ENTERED",
            )

        print(f"  ☁️ 已增量追加快取到 Google Sheet：{title}，本次 {len(new_rows):,} 筆")
        return True
    except Exception as e:
        print(f"  ⚠️ 原始分點資料快取增量追加到 Google Sheet 失敗：{type(e).__name__}: {e}")
        return False


def save_history_cache(history_df, fetched_items=None, previous_history_empty=False):
    if not USE_CACHE or history_df is None or history_df.empty:
        return history_df

    history_df, prune_stats = prune_history_cache_dataframe(history_df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 分點歷史自動清理：移除 {prune_stats['removed']:,} 筆｜"
            f"超期/異常日期 {prune_stats.get('removed_date', 0):,}｜"
            f"已移除分點 {prune_stats.get('removed_broker', 0):,}｜"
            f"保留最近 {HISTORY_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    delta_rows = []
    for item in (fetched_items or []):
        delta_rows.extend(item_to_history_rows(item))
    delta_df = pd.DataFrame(delta_rows)
    if not delta_df.empty:
        delta_df, _ = prune_history_cache_dataframe(delta_df)
    else:
        delta_df = history_df.copy() if previous_history_empty else pd.DataFrame(columns=history_df.columns)

    do_full_gsheet_sync = True

    if (
        HISTORY_GSHEET_INCREMENTAL_APPEND
        and not previous_history_empty
        and len(history_df) > HISTORY_CACHE_FULL_SYNC_THRESHOLD_ROWS
        and prune_stats.get("removed", 0) == 0
    ):
        do_full_gsheet_sync = False

    write_cache_csv(
        history_df,
        HISTORY_CACHE_PATH,
        sync_gsheet=do_full_gsheet_sync,
        supabase_df=delta_df,
        force_gsheet_sync=bool(prune_stats.get("removed", 0) and CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC),
        force_supabase_sync=bool(prune_stats.get("removed", 0)),
    )

    if not do_full_gsheet_sync:
        if supabase_should_skip_gsheet_cache(HISTORY_CACHE_PATH):
            print("  🗄️ 已啟用 Supabase 快取，略過原始分點歷史增量同步到 Google Sheet。")
        else:
            append_history_rows_to_gsheet(delta_rows)

    print(f"  💾 已更新原始分點資料快取：{HISTORY_CACHE_PATH}，共 {len(history_df):,} 筆")
    return history_df

def items_from_history_cache(history_df, candidate_filter=None):
    items = []

    if history_df is None or history_df.empty:
        return items

    df = history_df.copy().fillna("")

    if candidate_filter:
        mask = pd.Series(
            [candidate_key_from_values(w, b) in candidate_filter for w, b in zip(df["權證代號"], df["券商代號"])],
            index=df.index
        )
        df = df[mask].copy()

    if df.empty:
        return items

    numeric_cols = ["買進股數", "賣出股數", "買進金額", "賣出金額", "買超股數", "買超金額"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    group_cols = ["權證代號", "權證名稱", "標的股", "標的名稱", "分點", "分點名稱", "券商代號"]

    for key, g in df.groupby(group_cols, dropna=False):
        warrant_code, warrant_name, underlying_code, underlying_name, broker_label, broker_name, broker_code = key

        item_df = g[[
            "日期", "買進股數", "賣出股數",
            "買進金額", "賣出金額", "買超股數", "買超金額"
        ]].copy()

        item_df["日期"] = item_df["日期"].map(normalize_date_str)
        item_df = item_df.sort_values("日期").reset_index(drop=True)

        item = {
            "warrant_code": str(warrant_code).strip(),
            "warrant_name": str(warrant_name).strip(),
            "underlying_code": str(underlying_code).strip(),
            "underlying_name": str(underlying_name).strip(),
            "broker_label": str(broker_label).strip(),
            "broker_name": str(broker_name).strip(),
            "broker_code": str(broker_code).strip(),
            "df": item_df,
        }

        items.append(item)

    return items



# ══════════════════════════════════════════════════════════════════════
# Step 1：取所有認購權證 + 標的股代號
# ══════════════════════════════════════════════════════════════════════

def build_stock_map(df):
    stock_map = {}

    for row in df.itertuples(index=False, name=None):
        cell = str(row[0]).strip()

        if "　" in cell:
            parts = cell.split("　", 1)
            code, name = parts[0].strip(), parts[1].strip()
        else:
            m = re.match(r"^(\d{4})\s+(.+)$", cell)
            if m:
                code, name = m.group(1), m.group(2)
            else:
                continue

        if len(code) == 4 and code.isdigit():
            stock_map[name] = code

    return stock_map


def normalize_stock_name_text(s):
    return str(s).strip().replace(" ", "").replace("　", "")


# 特殊標的優先對照：
# 有些 ETF 權證名稱會使用市場慣用簡稱，例如「台灣50台新5B購17」。
# 若直接丟進一般股票名稱 alias，比對時可能被「台灣大哥大」的短 alias「台灣」吃到，
# 導致 0050 台灣50 被誤判成 3045 台灣大哥大。
# 因此這裡先處理已知 ETF / 特殊標的，且一定要放在一般 resolver 比對前面。
SPECIAL_UNDERLYING_PREFIX_RULES = [
    {
        "prefixes": [
            "台灣50",
            "臺灣50",
            "元大台灣50",
            "元大臺灣50",
        ],
        "exclude_prefixes": [
            "台灣50正",
            "臺灣50正",
            "元大台灣50正",
            "元大臺灣50正",
            "台灣50反",
            "臺灣50反",
            "元大台灣50反",
            "元大臺灣50反",
        ],
        "code": "0050",
        "name": "元大台灣50",
    },
]


def get_special_underlying_info_from_warrant_name(warrant_name):
    wname = normalize_stock_name_text(warrant_name)

    if not wname:
        return "", ""

    for rule in SPECIAL_UNDERLYING_PREFIX_RULES:
        prefixes = [normalize_stock_name_text(x) for x in rule.get("prefixes", []) if str(x).strip()]
        exclude_prefixes = [normalize_stock_name_text(x) for x in rule.get("exclude_prefixes", []) if str(x).strip()]

        if exclude_prefixes and any(wname.startswith(prefix) for prefix in exclude_prefixes):
            continue

        if prefixes and any(wname.startswith(prefix) for prefix in prefixes):
            return str(rule.get("code", "")).strip(), str(rule.get("name", "")).strip()

    return "", ""


def correct_underlying_info_by_warrant_name(warrant_name, underlying_code="", underlying_name=""):
    special_code, special_name = get_special_underlying_info_from_warrant_name(warrant_name)

    if special_code:
        return special_code, special_name

    return str(underlying_code or "").strip(), str(underlying_name or "").strip()


def fix_known_underlying_info_dataframe(df, warrant_name_col, underlying_code_col, underlying_name_col):
    """
    針對已知 ETF / 特殊標的修正既有快取中的標的股代號。

    這裡只處理明確命中的特殊規則，不改一般股票 / 權證辨識邏輯。
    目的：避免舊快取或 Supabase 讀回的「台灣50 → 3045」繼續污染後續候選與歷史資料。
    """
    if df is None or df.empty:
        return df

    needed = {warrant_name_col, underlying_code_col, underlying_name_col}
    if not needed.issubset(set(df.columns)):
        return df

    out = df.copy()
    names = out[warrant_name_col].astype(str).map(normalize_stock_name_text)

    for rule in SPECIAL_UNDERLYING_PREFIX_RULES:
        prefixes = [normalize_stock_name_text(x) for x in rule.get("prefixes", []) if str(x).strip()]
        exclude_prefixes = [normalize_stock_name_text(x) for x in rule.get("exclude_prefixes", []) if str(x).strip()]

        if not prefixes:
            continue

        mask = False
        for prefix in prefixes:
            mask = mask | names.str.startswith(prefix, na=False)

        for prefix in exclude_prefixes:
            mask = mask & ~names.str.startswith(prefix, na=False)

        if mask.any():
            out.loc[mask, underlying_code_col] = str(rule.get("code", "")).strip()
            out.loc[mask, underlying_name_col] = str(rule.get("name", "")).strip()

    return out


def make_stock_aliases(stock_name, exact_stock_names=None):
    """
    建立股票名稱候選別名，但避免把某一檔股票的簡稱撞到另一檔真實股票名稱。

    修正重點：
    1. 8028 昇陽半導體可產生「昇陽半」，讓「昇陽半XXX購」正確對到 8028。
    2. 8028 昇陽半導體不可產生「昇陽」，因為「昇陽」本身是 3266。
    3. 後續比對會用最長前綴優先，因此「昇陽半」會優先於「昇陽」。
    """
    name = normalize_stock_name_text(stock_name)
    aliases = set()

    if not name:
        return aliases

    exact_stock_names = exact_stock_names or set()

    aliases.add(name)

    suffixes = [
        "半導體", "科技", "電子", "光電", "精密", "材料", "生技", "醫療",
        "資訊", "電腦", "通信", "通訊", "電機", "機械", "工業", "實業",
        "企業", "國際", "控股", "投控", "控", "建設", "營造", "食品",
        "鋼鐵", "化學", "化工", "紡織", "玻璃", "塑膠", "水泥",
    ]

    stripped = name
    changed = True

    while changed:
        changed = False
        for suffix in suffixes:
            if stripped.endswith(suffix) and len(stripped) > len(suffix) + 1:
                candidate = stripped[:-len(suffix)]
                stripped = candidate

                # 如果切出來的簡稱剛好是另一檔股票的完整名稱，就不要加入。
                # 例如「昇陽半導體」切成「昇陽」，但「昇陽」本身是 3266。
                if candidate not in exact_stock_names or candidate == name:
                    aliases.add(candidate)

                changed = True
                break

    for n in range(min(4, len(name)), 1, -1):
        candidate = name[:n]

        # 前綴簡稱若撞到另一檔真實股票名稱，也不要加入。
        if candidate in exact_stock_names and candidate != name:
            continue

        aliases.add(candidate)

    # 「台灣 / 臺灣」這類過短且常見的前綴不適合作為標的辨識 alias。
    # 例如台灣大哥大若產生「台灣」alias，會誤吃「台灣50」ETF 權證。
    dangerous_aliases = {"台灣", "臺灣"}

    return {a for a in aliases if len(a) >= 2 and a not in dangerous_aliases}


def build_underlying_resolver(stock_map):
    """
    預先建立完整股名與安全 alias 對照表。

    重要：
    原本 find_underlying_info() 會先用完整股名比對，導致「昇陽半XXX購」
    先被短股名「昇陽」吃到，誤判成 3266。
    這版改成完整股名與 alias 全部放在同一個候選表，統一採「最長前綴優先」。
    因此「昇陽半」會優先於「昇陽」。
    """
    exact_stock_names = set()

    for stock_name in stock_map.keys():
        sname = normalize_stock_name_text(stock_name)
        if sname:
            exact_stock_names.add(sname)

    candidates = []
    seen = set()

    def add_candidate(prefix, stock_code, stock_name, is_exact_alias):
        prefix_norm = normalize_stock_name_text(prefix)
        stock_name_norm = normalize_stock_name_text(stock_name)

        if not prefix_norm:
            return

        key = (prefix_norm, str(stock_code), stock_name_norm)

        if key in seen:
            return

        seen.add(key)

        candidates.append({
            "prefix": prefix_norm,
            "prefix_len": len(prefix_norm),
            "is_exact_alias": 1 if is_exact_alias else 0,
            "stock_name_len": len(stock_name_norm),
            "stock_code": stock_code,
            "stock_name": stock_name,
        })

    for stock_name, stock_code in stock_map.items():
        stock_name_norm = normalize_stock_name_text(stock_name)

        # 完整股名也是候選，但不再獨立提前回傳，避免短完整股名壓過較長 alias。
        add_candidate(stock_name_norm, stock_code, stock_name, True)

        for alias in make_stock_aliases(stock_name, exact_stock_names):
            alias_norm = normalize_stock_name_text(alias)
            add_candidate(alias_norm, stock_code, stock_name, alias_norm == stock_name_norm)

    candidates = sorted(
        candidates,
        key=lambda x: (
            x["prefix_len"],
            x["is_exact_alias"],
            -x["stock_name_len"],
        ),
        reverse=True
    )

    return candidates


def find_underlying_info(warrant_name, stock_map, resolver=None):
    wname = normalize_stock_name_text(warrant_name)

    if not wname:
        return "", ""

    # 特殊標的優先處理，避免「台灣50」被一般股票短 alias「台灣」誤判成 3045 台灣大哥大。
    special_code, special_name = get_special_underlying_info_from_warrant_name(warrant_name)
    if special_code:
        return special_code, special_name

    if resolver is None:
        resolver = build_underlying_resolver(stock_map)

    for rec in resolver:
        prefix = rec["prefix"]

        if prefix and wname.startswith(prefix):
            return rec["stock_code"], rec["stock_name"]

    return "", ""




def get_all_call_warrants_live():
    global LIVE_WARRANT_MARKET_SUCCESS_COUNT

    print("【Step 1】取所有認購權證清單...")
    warrants = []
    LIVE_WARRANT_MARKET_SUCCESS_COUNT = 0

    # strMode=2：上市有價證券
    # strMode=4：上櫃有價證券
    # 兩邊都要抓，否則上櫃標的的權證，例如 70xxxx 權證會漏掉。
    isin_modes = [
        ("上市", "2"),
        ("上櫃", "4"),
    ]

    all_dfs = []
    stock_map = {}

    for market_name, mode in isin_modes:
        try:
            resp = requests.get(
                f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}",
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=30
            )
            resp.raise_for_status()
            resp.encoding = "cp950"

            tables = pd.read_html(StringIO(resp.text))
            df = tables[0].iloc[2:].reset_index(drop=True)

            all_dfs.append((market_name, df))
            stock_map.update(build_stock_map(df))
            LIVE_WARRANT_MARKET_SUCCESS_COUNT += 1

            print(f"  ✅ 已取得{market_name} ISIN 清單：{len(df)} 筆")

        except Exception as e:
            print(f"  ⚠️ {market_name} ISIN 清單取得失敗：{e}")

    # 只建立一次 resolver，避免每檔權證都重新掃全部股票與 alias，保持執行速度。
    underlying_resolver = build_underlying_resolver(stock_map)

    seen_warrants = set()

    for market_name, df in all_dfs:
        for row in df.itertuples(index=False, name=None):
            cell = str(row[0]).strip()

            if "\u3000" in cell:
                parts = cell.split("\u3000", 1)
                code, name = parts[0].strip(), parts[1].strip()
            else:
                m = re.match(r"^(\d{6})\s+(.+)$", cell)
                if m:
                    code, name = m.group(1), m.group(2)
                else:
                    continue

            if len(code) == 6 and code.isdigit() and "購" in name:
                if code in seen_warrants:
                    continue

                seen_warrants.add(code)
                underlying, underlying_name = find_underlying_info(name, stock_map, underlying_resolver)

                warrants.append({
                    "代號": code,
                    "名稱": name,
                    "標的股": underlying,
                    "標的名稱": underlying_name
                })

    print(f"  ✅ 共 {len(warrants)} 支認購權證")
    return warrants




def get_all_call_warrants():
    global CURRENT_LIVE_WARRANT_CODES, LIVE_WARRANT_SNAPSHOT_READY

    cached_warrants = load_warrants_cache()

    if cached_warrants:
        print("【Step 1】讀取認購權證清單快取...")
        print(f"  ✅ 已讀取權證清單快取：{len(cached_warrants)} 支")
        print("  🔄 即時更新今日認購權證清單；成功時以今日清單完整覆蓋，移除已到期權證...")
        live_warrants = get_all_call_warrants_live()

        if not live_warrants or LIVE_WARRANT_MARKET_SUCCESS_COUNT < LIVE_WARRANT_MARKET_EXPECTED_COUNT:
            LIVE_WARRANT_SNAPSHOT_READY = False
            CURRENT_LIVE_WARRANT_CODES = {
                str(w.get("代號", "")).strip()
                for w in cached_warrants
                if str(w.get("代號", "")).strip()
            }
            if live_warrants:
                print(
                    f"  ⚠️ 即時權證清單只有 {LIVE_WARRANT_MARKET_SUCCESS_COUNT}/{LIVE_WARRANT_MARKET_EXPECTED_COUNT} 個市場成功，"
                    "為避免把另一市場的有效權證誤刪，暫時沿用既有權證清單快取。"
                )
            else:
                print("  ⚠️ 即時權證清單取得失敗，為避免誤刪資料，暫時沿用既有權證清單快取。")
            return cached_warrants

        warrants = live_warrants
        LIVE_WARRANT_SNAPSHOT_READY = True
        CURRENT_LIVE_WARRANT_CODES = {
            str(w.get("代號", "")).strip()
            for w in warrants
            if str(w.get("代號", "")).strip()
        }
        save_warrants_cache(warrants)

        old_codes = {
            str(w.get("代號", "")).strip()
            for w in cached_warrants
            if str(w.get("代號", "")).strip()
        }
        removed_count = len(old_codes - CURRENT_LIVE_WARRANT_CODES)
        new_count = len(CURRENT_LIVE_WARRANT_CODES - old_codes)
        print(
            f"  ✅ 權證清單更新完成：原快取 {len(old_codes):,} 支｜"
            f"今日有效 {len(warrants):,} 支｜新增 {new_count:,} 支｜移除已失效 {removed_count:,} 支"
        )
        return warrants

    warrants = get_all_call_warrants_live()
    if warrants and LIVE_WARRANT_MARKET_SUCCESS_COUNT >= LIVE_WARRANT_MARKET_EXPECTED_COUNT:
        LIVE_WARRANT_SNAPSHOT_READY = True
        CURRENT_LIVE_WARRANT_CODES = {
            str(w.get("代號", "")).strip()
            for w in warrants
            if str(w.get("代號", "")).strip()
        }
        save_warrants_cache(warrants)
    else:
        LIVE_WARRANT_SNAPSHOT_READY = False
        CURRENT_LIVE_WARRANT_CODES = {
            str(w.get("代號", "")).strip()
            for w in (warrants or [])
            if str(w.get("代號", "")).strip()
        }
        if warrants:
            print(
                f"  ⚠️ 即時權證清單只有 {LIVE_WARRANT_MARKET_SUCCESS_COUNT}/{LIVE_WARRANT_MARKET_EXPECTED_COUNT} 個市場成功，"
                "本次可繼續使用已取得資料，但不會執行失效權證刪除。"
            )
    return warrants


# ══════════════════════════════════════════════════════════════════════
# Step 2：找目標分點券商代號
# ══════════════════════════════════════════════════════════════════════

def find_broker_codes_live(warrants):
    print("【Step 2】掃描目標分點券商代號...")

    today = datetime.today()
    end_s   = today.strftime("%Y/%m/%d")
    start_s = (today - timedelta(days=3)).strftime("%Y/%m/%d")
    found = {}

    def scan_one(code):
        hits = {}
        for row in api4_get(code, start_s, end_s):
            name = row.get("V3", "")
            label = match_target(name)
            if label and label not in found:
                hits[label] = (name, row.get("V2", ""))
        return hits

    with ThreadPoolExecutor(max_workers=FIND_BROKER_WORKERS) as ex:
        futures = {ex.submit(scan_one, w["代號"]): w for w in warrants[:300]}

        for future in as_completed(futures):
            try:
                result = future.result()
            except:
                result = {}

            for label, (name, code) in result.items():
                if label not in found:
                    found[label] = (name, code)

            if len(found) == len(TARGET_PATTERNS):
                for pending_future in futures:
                    if not pending_future.done():
                        pending_future.cancel()
                break

    for label, (name, code) in found.items():
        print(f"    ✅ {label:14s} → {name} ({code})")

    missing = [k for k in TARGET_PATTERNS if k not in found]

    if missing:
        print(f"    備援補入：{missing}")
        for k in missing:
            if k in FALLBACK:
                found[k] = FALLBACK[k]
                print(f"    ✅ {k:14s} → {FALLBACK[k][0]} ({FALLBACK[k][1]}) [備援]")

    return found




def find_broker_codes(warrants):
    broker_map = load_broker_map_cache()

    if broker_map:
        print("【Step 2】讀取目標分點券商代號快取...")
        for label, (name, code) in broker_map.items():
            print(f"    ✅ {label:14s} → {name} ({code}) [快取]")
        return broker_map

    broker_map = find_broker_codes_live(warrants)
    save_broker_map_cache(broker_map)
    return broker_map


# ══════════════════════════════════════════════════════════════════════
# Step 3a：預篩有目標分點出現的 (權證, 分點) 組合
# ══════════════════════════════════════════════════════════════════════

def prescan_all_live(warrants, broker_map, scan_days=40):
    global PRESCAN_LATEST_ACTIVITY_DATE, PRESCAN_TODAY_ACTIVITY_FOUND

    print("【Step 3a】預篩：找有目標分點的權證...")

    today = datetime.today()
    today_s = today.strftime("%Y/%m/%d")
    end_s   = today.strftime("%Y/%m/%d")
    start_s = (today - timedelta(days=scan_days)).strftime("%Y/%m/%d")

    PRESCAN_LATEST_ACTIVITY_DATE = None
    PRESCAN_TODAY_ACTIVITY_FOUND = False

    broker_codes_set = {code for _, code in broker_map.values()}
    code_to_label    = {code: label for label, (_, code) in broker_map.items()}

    candidates = []
    done = 0

    def prescan_one(w):
        hits = []
        latest_dt = None
        today_found = False

        for row in api4_get(w["代號"], start_s, end_s):
            bcode = row.get("V2", "")

            if bcode in broker_codes_set:
                label = code_to_label.get(bcode, "")

                if label:
                    row_date = normalize_date_str(row.get("V1", ""))
                    row_dt = parse_date(row_date)
                    if row_dt and (latest_dt is None or row_dt > latest_dt):
                        latest_dt = row_dt
                    if row_date == today_s:
                        today_found = True

                    bname = next((n for l, (n, c) in broker_map.items() if c == bcode), bcode)
                    hits.append((w["代號"], w["名稱"], w["標的股"], w.get("標的名稱", ""), label, bname, bcode))

        return hits, latest_dt, today_found

    with ThreadPoolExecutor(max_workers=PRESCAN_WORKERS) as ex:
        futures = {ex.submit(prescan_one, w): w for w in warrants}

        for future in as_completed(futures):
            done += 1

            try:
                result, latest_dt, today_found = future.result()
            except:
                result, latest_dt, today_found = [], None, False

            if latest_dt and (PRESCAN_LATEST_ACTIVITY_DATE is None or latest_dt > PRESCAN_LATEST_ACTIVITY_DATE):
                PRESCAN_LATEST_ACTIVITY_DATE = latest_dt
            if today_found:
                PRESCAN_TODAY_ACTIVITY_FOUND = True

            for hit in result:
                candidates.append(hit)

            if done % 1000 == 0:
                print(f"  [{done:,}/{len(warrants):,}] 預篩中，候選 {len(candidates)} 組...")

    unique_candidates = []
    seen = set()

    for c in candidates:
        key = (c[0], c[6])
        if key not in seen:
            seen.add(key)
            unique_candidates.append(c)

    latest_label = PRESCAN_LATEST_ACTIVITY_DATE.strftime("%Y/%m/%d") if PRESCAN_LATEST_ACTIVITY_DATE else "-"
    print(f"  ✅ 預篩完成：{len(warrants)} 支 → {len(candidates)} 組候選，去重後 {len(unique_candidates)} 組")
    print(f"  ✅ API4 目標分點最新活動日：{latest_label}｜今日資料：{'已出現' if PRESCAN_TODAY_ACTIVITY_FOUND else '尚未出現'}")
    return unique_candidates




def collect_underlying_codes_from_candidates(candidates):
    """從 API4 prescan 掃到的候選中取得近期有活動的標的股。"""
    underlying_codes = set()

    if not candidates:
        return underlying_codes

    for c in candidates:
        try:
            underlying_code = str(c[2]).strip()
        except Exception:
            underlying_code = ""

        if not underlying_code:
            continue

        norm_code = normalize_underlying_code_for_group(underlying_code)
        underlying_codes.add(norm_code or underlying_code)

    return {code for code in underlying_codes if str(code).strip()}


def collect_underlying_broker_pairs_from_candidates(candidates):
    """
    從 API4 prescan 候選整理出「標的股 + 券商代號」活動配對。

    舊版只要某標的近期有活動，就展開成「該標的所有認購權證 × 所有追蹤分點」，
    RUN_MODE=2 會膨脹到百萬組候選。

    新版只展開同標的 + 同一個近期真的有活動的分點。
    例如 API4 掃到「元大南屯 × 南亞科」，才展開：
    南亞科所有認購權證 × 元大南屯。
    """
    pairs = set()

    if not candidates:
        return pairs

    for c in candidates:
        try:
            underlying_code = normalize_underlying_code_for_group(c[2]) or str(c[2]).strip()
            broker_code = str(c[6]).strip()
        except Exception:
            underlying_code = ""
            broker_code = ""

        if underlying_code and broker_code:
            pairs.add((underlying_code, broker_code))

    return pairs


def build_underlying_expanded_candidates(warrants, broker_map, underlying_codes, source_label="", active_pairs=None):
    """
    依近期有活動的標的股建立候選池。

    核心邏輯：
    1. API4 先找出最近有活動的「標的股 + 分點」。
    2. 只對同一個活動配對展開：同標的所有認購權證 × 同分點。
    3. 若 active_pairs 沒有傳入，才退回舊的「標的 × 追蹤分點」相容模式。

    這樣仍可抓到「同一分點在同一標的多檔權證分散式買賣超」，
    但不會在 RUN_MODE=2 膨脹成全市場權證 × 全部分點。
    """
    if not warrants or not broker_map or not underlying_codes:
        return []

    normalized_underlyings = set()
    for code in underlying_codes:
        norm_code = normalize_underlying_code_for_group(code)
        if norm_code:
            normalized_underlyings.add(norm_code)
        elif str(code).strip():
            normalized_underlyings.add(str(code).strip())

    if not normalized_underlyings:
        return []

    active_pairs = active_pairs or set()
    active_pairs = {
        (normalize_underlying_code_for_group(u) or str(u).strip(), str(b).strip())
        for u, b in active_pairs
        if str(u).strip() and str(b).strip()
    }

    broker_code_to_info = {}
    for label, (broker_name, broker_code) in broker_map.items():
        broker_code = str(broker_code).strip()
        if broker_code:
            broker_code_to_info[broker_code] = (label, str(broker_name).strip(), broker_code)

    title = source_label.strip() or "活動標的"

    if active_pairs:
        print(
            f"【Step 3a】{title} 候選池擴展：近期活動標的 {len(normalized_underlyings):,} 檔，"
            f"活動標的×分點配對 {len(active_pairs):,} 組 "
            f"→ 同標的所有認購權證 × 同活動分點..."
        )
    else:
        print(
            f"【Step 3a】{title} 候選池擴展：近期活動標的 {len(normalized_underlyings):,} 檔 "
            f"→ 同標的所有認購權證 × 追蹤分點 [相容模式]..."
        )

    candidates = []
    seen = set()
    matched_warrants = 0
    matched_pairs = set()

    for w in warrants:
        warrant_code = str(w.get("代號", "")).strip()
        warrant_name = str(w.get("名稱", "")).strip()
        underlying_code = str(w.get("標的股", "")).strip()
        underlying_name = str(w.get("標的名稱", "")).strip()

        if not warrant_code or not warrant_name or not underlying_code:
            continue

        norm_underlying = normalize_underlying_code_for_group(underlying_code) or underlying_code

        if norm_underlying not in normalized_underlyings:
            continue

        if active_pairs:
            active_broker_codes = [
                broker_code
                for u, broker_code in active_pairs
                if u == norm_underlying and broker_code in broker_code_to_info
            ]
            broker_infos = [broker_code_to_info[broker_code] for broker_code in active_broker_codes]
        else:
            broker_infos = [
                (label, str(broker_name).strip(), str(broker_code).strip())
                for label, (broker_name, broker_code) in broker_map.items()
                if str(broker_code).strip()
            ]

        if not broker_infos:
            continue

        matched_warrants += 1

        for label, broker_name, broker_code in broker_infos:
            c = (
                warrant_code,
                warrant_name,
                norm_underlying,
                underlying_name,
                label,
                broker_name,
                broker_code,
            )
            key = candidate_key_from_tuple(c)

            if key in seen:
                continue

            seen.add(key)
            matched_pairs.add((norm_underlying, broker_code))
            candidates.append(c)

    print(
        f"  ✅ {title} 候選池擴展完成：命中權證 {matched_warrants:,} 支，"
        f"命中活動配對 {len(matched_pairs):,} 組 → {len(candidates):,} 組候選"
    )
    return candidates

def build_selected_full_market_candidates(warrants, broker_map):
    """
    RUN_MODE=1 精選分點完整追蹤候選池。

    目的：
    原本候選池依賴 api4 prescan，若某檔權證的分點資料沒有在 api4 回傳清單中出現，
    即使該分點今天實際有大額賣出，後續也不會進入 API5 歷史抓取與每日賣出明細。

    這裡改成針對精選分點建立「所有認購權證 × 精選分點」候選組合，
    再由 API5 抓該分點該權證最近設定天數的歷史，確保像「元大南屯賣南亞科」這類
    分散在多檔權證的大額賣超不會因候選池漏抓而少算。
    """
    print("【Step 3a】RUN_MODE=1 精選分點完整候選池：所有認購權證 × 精選分點...")

    if not warrants or not broker_map:
        return []

    candidates = []
    seen = set()

    for w in warrants:
        warrant_code = str(w.get("代號", "")).strip()
        warrant_name = str(w.get("名稱", "")).strip()
        underlying_code = str(w.get("標的股", "")).strip()
        underlying_name = str(w.get("標的名稱", "")).strip()

        if not warrant_code or not warrant_name:
            continue

        for label, (broker_name, broker_code) in broker_map.items():
            broker_code = str(broker_code).strip()
            if not broker_code:
                continue

            c = (
                warrant_code,
                warrant_name,
                underlying_code,
                underlying_name,
                label,
                str(broker_name).strip(),
                broker_code,
            )
            key = candidate_key_from_tuple(c)

            if key in seen:
                continue

            seen.add(key)
            candidates.append(c)

    print(
        f"  ✅ 精選分點完整候選池建立完成：權證 {len(warrants):,} 支 × 分點 {len(broker_map):,} 間 "
        f"→ {len(candidates):,} 組候選"
    )
    return candidates



def build_candidates_from_history_cache(warrants, broker_map, history_df):
    """
    從共用的 API5 分點歷史快取，自動補回目前執行模式應該擁有的候選組合。

    修正目的：
    - candidates_cache.csv 與 candidates_cache_selected5.csv 雖然分開保存，
      但 broker_warrant_history_cache.csv 是兩種模式共用。
    - 若某組「權證代號 + 券商代號」已經存在共用歷史快取，且該券商屬於目前
      RUN_MODE 的追蹤分點，就不應再因為目前模式的候選快取缺少該組合而被
      items_from_history_cache(..., candidate_filter=...) 濾掉。
    - 候選池只補目前仍上市的認購權證與目前追蹤中的分點，不會把其他分點或
      已失效權證帶入本次結果。

    這個補回動作只使用既有共用歷史資料，不會額外呼叫 API5。
    """
    if not warrants or not broker_map or history_df is None or history_df.empty:
        return []

    required_cols = {
        "權證代號", "權證名稱", "標的股", "標的名稱",
        "分點", "分點名稱", "券商代號",
    }
    if not required_cols.issubset(set(history_df.columns)):
        return []

    broker_code_to_info = {}
    for label, (broker_name, broker_code) in broker_map.items():
        broker_code = str(broker_code).strip()
        if not broker_code:
            continue
        broker_code_to_info[broker_code] = (
            str(label).strip(),
            str(broker_name).strip(),
            broker_code,
        )

    if not broker_code_to_info:
        return []

    warrant_by_code = {}
    for warrant in warrants:
        warrant_code = str(warrant.get("代號", "")).strip()
        if not warrant_code:
            continue
        warrant_by_code[warrant_code] = warrant

    if not warrant_by_code:
        return []

    df = history_df[[
        "權證代號", "權證名稱", "標的股", "標的名稱",
        "分點", "分點名稱", "券商代號",
    ]].copy().fillna("")

    df["權證代號"] = df["權證代號"].astype(str).str.strip()
    df["券商代號"] = df["券商代號"].astype(str).str.strip()
    df = df[
        df["券商代號"].isin(set(broker_code_to_info.keys()))
        & df["權證代號"].isin(set(warrant_by_code.keys()))
    ].copy()

    if df.empty:
        return []

    df = df.drop_duplicates(subset=["權證代號", "券商代號"], keep="last")

    candidates = []
    seen = set()

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        warrant_code = str(row_dict.get("權證代號", "")).strip()
        broker_code = str(row_dict.get("券商代號", "")).strip()

        warrant = warrant_by_code.get(warrant_code)
        broker_info = broker_code_to_info.get(broker_code)
        if not warrant or not broker_info:
            continue

        broker_label, broker_name, broker_code = broker_info
        warrant_name = str(warrant.get("名稱", "") or row_dict.get("權證名稱", "")).strip()
        underlying_code = str(warrant.get("標的股", "") or row_dict.get("標的股", "")).strip()
        underlying_name = str(warrant.get("標的名稱", "") or row_dict.get("標的名稱", "")).strip()

        if not warrant_name:
            continue

        underlying_code, underlying_name = correct_underlying_info_by_warrant_name(
            warrant_name,
            underlying_code,
            underlying_name,
        )

        candidate = (
            warrant_code,
            warrant_name,
            underlying_code,
            underlying_name,
            broker_label,
            broker_name,
            broker_code,
        )
        key = candidate_key_from_tuple(candidate)

        if key in seen:
            continue

        seen.add(key)
        candidates.append(candidate)

    return candidates

def prescan_all(warrants, broker_map):
    global PRESCAN_REFRESH_KEYS, PRESCAN_MISSING_FETCH_KEYS

    broker_map = filter_broker_map_for_active_targets(broker_map)

    if workflow_is_daily():
        cached_from_success = try_load_prescan_success_cache(broker_map)
        if cached_from_success is not None:
            return cached_from_success
    elif workflow_is_repair():
        print("  🔧 WORKFLOW_MODE=repair：忽略 prescan_status，強制重新掃描候選。")

    cached_candidates = filter_candidates_by_broker_map(load_candidates_cache(), broker_map)

    if cached_candidates:
        if RUN_MODE == 1:
            print("【Step 3a】讀取精選分點候選組合快取...")
            print(f"  ✅ 已讀取精選候選組合快取：{len(cached_candidates):,} 組")
        else:
            print("【Step 3a】讀取候選組合快取...")
            print(f"  ✅ 已讀取候選組合快取：{len(cached_candidates):,} 組")

    if FAST_SKIP_RECENT_PRESCAN:
        print("  ⚠️ 偵測到 FAST_SKIP_RECENT_PRESCAN=1，但目前仍會補掃最近資料，避免漏掉新權證 / 新候選組合。")

    # 新版 RUN_MODE=1 / RUN_MODE=2 共用增量流程：
    # 1. 用 API4 補掃最近資料，找出追蹤分點近期有活動的權證與標的。
    # 2. 將活動標的展開成「同標的所有認購權證 × 追蹤分點」。
    # 3. 只有 API4 直接掃到的 key 會放入 PRESCAN_REFRESH_KEYS；
    #    擴展出的 key 若快取已有資料就不強制 API5，只在缺快取時補抓。
    if cached_candidates:
        scan_days = SELECTED_FULL_SCAN_DAYS if RUN_MODE == 1 else CACHE_RECENT_SCAN_DAYS
    else:
        scan_days = max(40, SELECTED_FULL_SCAN_DAYS if RUN_MODE == 1 else CACHE_RECENT_SCAN_DAYS)

    if RUN_MODE == 1:
        print(
            f"  🔄 RUN_MODE=1：補掃全市場最近 {scan_days} 天，"
            "先鎖定精選分點有活動的標的，再展開同標的所有權證。"
        )
    else:
        print(
            f"  🔄 RUN_MODE=2：補掃全市場最近 {scan_days} 天，"
            "先鎖定完整分點清單有活動的標的，再展開同標的所有權證。"
        )

    recent_candidates = prescan_all_live(warrants, broker_map, scan_days=scan_days)
    recent_candidates = filter_candidates_by_broker_map(recent_candidates, broker_map)

    active_underlyings = collect_underlying_codes_from_candidates(recent_candidates)
    active_underlying_broker_pairs = collect_underlying_broker_pairs_from_candidates(recent_candidates)

    if EXPAND_ACTIVE_UNDERLYING_WARRANTS and active_underlyings:
        mode_label = "精選分點活動標的" if RUN_MODE == 1 else "全分點活動標的"
        expanded_candidates = build_underlying_expanded_candidates(
            warrants,
            broker_map,
            active_underlyings,
            source_label=mode_label,
            active_pairs=active_underlying_broker_pairs,
        )
        expanded_candidates = filter_candidates_by_broker_map(expanded_candidates, broker_map)
    else:
        expanded_candidates = []
        if not active_underlyings:
            print("  ⚠️ 最近 API4 未掃到可展開的活動標的，本次只使用既有候選快取或 API4 直接候選。")
        else:
            print("  ⚠️ EXPAND_ACTIVE_UNDERLYING_WARRANTS=0，略過活動標的候選池擴展。")

    refresh_candidates = merge_candidates(recent_candidates, expanded_candidates)
    refresh_candidates = filter_candidates_by_broker_map(refresh_candidates, broker_map)

    # 只有 API4 直接掃到的候選才需要檢查是否更新；擴展候選若快取已有資料不強制刷新。
    PRESCAN_REFRESH_KEYS = {candidate_key_from_tuple(c) for c in recent_candidates}
    # 若候選沒有歷史快取，只有本次近期活動標的相關候選才需要補抓。
    # 這可避免舊版留下的「全市場權證 × 分點」空候選在有候選快取時仍全部打 API5。
    PRESCAN_MISSING_FETCH_KEYS = {candidate_key_from_tuple(c) for c in refresh_candidates}

    merged_candidates = merge_candidates(cached_candidates, refresh_candidates)
    merged_candidates = filter_candidates_by_broker_map(merged_candidates, broker_map)

    candidate_cache_saved = False
    try:
        candidate_cache_saved = bool(save_candidates_cache(merged_candidates))
    except Exception as e:
        candidate_cache_saved = False
        print(f"  ⚠️ 候選組合快取寫入失敗：{type(e).__name__}: {e}")

    target_date = current_prescan_target_date()
    prescan_has_today = has_today_broker_data_from_prescan(target_date)

    if candidate_cache_saved and prescan_has_today and merged_candidates:
        write_prescan_refresh_keys(PRESCAN_REFRESH_KEYS, PRESCAN_MISSING_FETCH_KEYS, target_date=target_date)
        write_prescan_status(
            "success",
            target_date=target_date,
            candidate_count=len(merged_candidates),
            direct_candidate_count=len(PRESCAN_REFRESH_KEYS),
            missing_fetch_key_count=len(PRESCAN_MISSING_FETCH_KEYS),
            candidate_cache_saved=True,
        )
    else:
        status = "broker_data_not_ready" if not prescan_has_today else "failed"
        reason = "API4 尚未看到今日目標分點資料" if not prescan_has_today else "候選快取未成功寫入或候選為空"
        write_prescan_status(
            status,
            target_date=target_date,
            reason=reason,
            candidate_count=len(merged_candidates),
            direct_candidate_count=len(PRESCAN_REFRESH_KEYS),
            missing_fetch_key_count=len(PRESCAN_MISSING_FETCH_KEYS),
            candidate_cache_saved=candidate_cache_saved,
        )

    print(
        f"  ✅ 候選組合完成：快取 {len(cached_candidates):,} 組，"
        f"API4直接候選 {len(recent_candidates):,} 組，"
        f"活動標的×分點配對 {len(active_underlying_broker_pairs):,} 組，"
        f"活動標的擴展 {len(expanded_candidates):,} 組，"
        f"合併後 {len(merged_candidates):,} 組"
    )
    print(f"  ✅ 本次需用 API5 檢查更新的直接候選組合：{len(PRESCAN_REFRESH_KEYS):,} 組")

    return merged_candidates


# ══════════════════════════════════════════════════════════════════════
# 新制說明：不再建立舊版「A_單檔大買」逐筆事件。
# 事件統一由 build_amount_class_events() 依「同分點 × 同標的 × 同一天」產生 A/B/C/D/E。
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# Step 3b：抓候選組合歷史資料
# ══════════════════════════════════════════════════════════════════════

def process_candidate(warrant_code, warrant_name, underlying_code, underlying_name, broker_label, broker_name, broker_code):
    rows = api5_get(warrant_code, broker_code)

    if not rows:
        return None

    records = []

    for row in rows:
        buy_s  = int(float(row.get("V2", 0) or 0))
        sell_s = int(float(row.get("V3", 0) or 0))
        buy_a  = int(float(row.get("V4", 0) or 0) * 1000)
        sell_a = int(float(row.get("V5", 0) or 0) * 1000)

        records.append({
            "日期": normalize_date_str(row.get("V1", "")),
            "買進股數": buy_s,
            "賣出股數": sell_s,
            "買進金額": buy_a,
            "賣出金額": sell_a,
            "買超股數": buy_s - sell_s,
            "買超金額": buy_a - sell_a,
        })

    df = pd.DataFrame(records).sort_values("日期").reset_index(drop=True)

    item = {
        "warrant_code":    warrant_code,
        "warrant_name":    warrant_name,
        "underlying_code": underlying_code,
        "underlying_name": underlying_name,
        "broker_label":    broker_label,
        "broker_name":     broker_name,
        "broker_code":     broker_code,
        "df":              df,
    }

    return item


# ══════════════════════════════════════════════════════════════════════
# 建立每日資料，用於新制 A/B/C/D/E 金額強度分類
# ══════════════════════════════════════════════════════════════════════

def build_daily_records(items):
    daily_records = []

    for item in items:
        if not item["underlying_code"]:
            continue

        for row in item["df"].itertuples(index=False):
            row_dict = row._asdict()
            daily_records.append({
                "日期": row_dict["日期"],
                "分點": item["broker_label"],
                "分點名稱": item["broker_name"],
                "券商代號": item["broker_code"],
                "權證代號": item["warrant_code"],
                "權證名稱": item["warrant_name"],
                "標的股": item["underlying_code"],
                "標的名稱": item.get("underlying_name", ""),
                "買進股數": int(row_dict["買進股數"]),
                "賣出股數": int(row_dict["賣出股數"]),
                "買進金額": int(row_dict["買進金額"]),
                "賣出金額": int(row_dict["賣出金額"]),
                "買超股數": int(row_dict["買超股數"]),
                "買超金額": int(row_dict["買超金額"]),
            })

    return daily_records


def normalize_warrant_code_for_unique(warrant_code):
    """
    用於新制金額強度分類與 FIFO 扣減的權證代號正規化。

    """
    return str(warrant_code or "").strip()


def make_broker_warrant_key(broker_code, warrant_code):
    """
    建立新制金額強度分類與 FIFO 扣減使用的唯一鍵。

    正確單位為「券商代號 + 權證代號」：
    - 新制事件以同分點、同標的、同一天合併。
    - FIFO 扣減時仍需用券商代號與權證代號分開追蹤。
    """
    broker_code = str(broker_code or "").strip()
    warrant_code = normalize_warrant_code_for_unique(warrant_code)

    if not broker_code or not warrant_code:
        return None

    return broker_code, warrant_code


def classify_amount_class(total_amount):
    """
    依同標的單日累積買進金額回傳 A / B / C / D / E 分類。
    """
    try:
        total_amount = int(total_amount or 0)
    except Exception:
        total_amount = 0

    for code, label, lower, upper in AMOUNT_CLASS_SPECS:
        if total_amount >= lower and (upper is None or total_amount < upper):
            return code, label

    return "", ""


def split_amount_class_events(events):
    out = {code: [] for code in AMOUNT_CLASS_CODES}

    for ev in events or []:
        code = str(ev.get("事件代碼", "")).strip()
        if code in out:
            out[code].append(ev)

    return tuple(out[code] for code in AMOUNT_CLASS_CODES)


def iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events=None):
    event_groups = [
        ("A", a_events or []),
        ("B", b_events or []),
        ("C", c_events or []),
        ("D", d_events or []),
    ]

    if e_events is not None:
        event_groups.append(("E", e_events or []))

    return event_groups


def flatten_amount_class_events(a_events, b_events, c_events, d_events, e_events=None):
    events = []
    for _, group_events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        events.extend(group_events or [])
    return events


def build_amount_class_events(daily_records, item_map):
    """
    新制 A/B/C/D/E 金額強度分類。

    事件單位：同一分點 + 同一標的 + 同一天。
    進入條件：該事件內至少一檔權證單日買進金額 >= AMOUNT_THRESH。
    分類依據：同一分點 + 同一標的 + 同一天的所有權證買進金額合計。

    注意：
    1. 這裡使用「買進金額」而不是「買超金額」，避免同日賣出把實際買入力道扣掉。
    2. 單筆 100 萬負責過濾小單；單日累積金額負責判斷資金強度。
    3. 同一天同分點同標的若買多檔權證，只會合併成一筆事件，不會拆成多筆分類。
    """
    events = []

    if not daily_records:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    df = pd.DataFrame(daily_records)

    if df.empty:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    df = df[(df["買進金額"] > 0) & (df["買進股數"] > 0)].copy()

    if df.empty:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    group_cols = ["分點", "分點名稱", "券商代號", "標的股", "標的名稱", "日期"]

    for key, g in df.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_name, broker_code, underlying_code, underlying_name, date = key
        date = normalize_date_str(date)

        warrant_rows = g.groupby(["權證代號", "權證名稱"], as_index=False).agg({
            "買進金額": "sum",
            "買進股數": "sum",
        })

        lots = []
        max_single_amount = 0
        max_single_warrant = ""

        for wr in warrant_rows.itertuples(index=False):
            wr = wr._asdict()
            warrant_code = normalize_warrant_code_for_unique(wr["權證代號"])
            warrant_key = make_broker_warrant_key(broker_code, warrant_code)

            if not warrant_key:
                continue

            buy_amount = int(wr["買進金額"] or 0)
            buy_shares = int(wr["買進股數"] or 0)

            if buy_amount <= 0 or buy_shares <= 0:
                continue

            if buy_amount > max_single_amount:
                max_single_amount = buy_amount
                max_single_warrant = f'{wr["權證代號"]} {wr["權證名稱"]}'

            lots.append({
                "買進日": date,
                "權證代號": wr["權證代號"],
                "權證名稱": wr["權證名稱"],
                "金額": buy_amount,
                "股數": buy_shares,
            })

        if not lots:
            continue

        # 最低進入條件：同日同分點同標的內，至少有一檔權證單筆買進金額 >= 100 萬。
        if max_single_amount < AMOUNT_THRESH:
            continue

        total_amount = int(sum(lot["金額"] for lot in lots))
        total_shares = int(sum(lot["股數"] for lot in lots))
        warrant_count = len(set(normalize_warrant_code_for_unique(lot["權證代號"]) for lot in lots))

        if total_amount < AMOUNT_THRESH or total_shares <= 0:
            continue

        event_code, event_label = classify_amount_class(total_amount)

        if not event_code:
            continue

        event = {
            "事件類型": f"{event_code}-{event_label}",
            "事件代碼": event_code,
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "起始日": date,
            "結束日": date,
            "事件日": date,
            "涵蓋權證數": warrant_count,
            "權證清單": "；".join([f'{lot["權證代號"]} {lot["權證名稱"]}' for lot in lots]),
            "最大單筆金額": max_single_amount,
            "最大單筆權證": max_single_warrant,
            "單日累積買進金額": total_amount,
            "買超金額": total_amount,
            "買超股數": total_shares,
            "買超張數": total_shares // 1000,
            "lots": lots,
        }

        events.append(event)

    events = simulate_group_outcomes_fifo(events, item_map)
    return {code: group for code, group in zip(AMOUNT_CLASS_CODES, split_amount_class_events(events))}








# ══════════════════════════════════════════════════════════════════════
# B / C 群組事件出清推估
# ══════════════════════════════════════════════════════════════════════


_GROUP_OUTCOME_SALE_ROWS_CACHE = {}


def get_group_sale_rows_for_warrant(item_map, broker_code, warrant_code):
    """
    新制金額強度事件 FIFO 扣減使用的賣出資料快取。

    同一個「券商代號 + 權證代號」只整理一次每日賣出資料；
    後續由 simulate_group_outcomes_fifo() 將每一筆賣出依事件時間 FIFO 分配，
    避免同一筆賣出被不同事件重複使用。
    """
    key = (str(broker_code).strip(), str(warrant_code).strip())

    if key in _GROUP_OUTCOME_SALE_ROWS_CACHE:
        return _GROUP_OUTCOME_SALE_ROWS_CACHE[key]

    item = item_map.get(key)
    if not item:
        item = item_map.get((key[0], normalize_price_code(key[1])))

    rows = []

    if item:
        df = item.get("df", pd.DataFrame())

        if df is not None and not df.empty:
            needed_cols = ["日期", "賣出股數", "賣出金額"]

            if all(col in df.columns for col in needed_cols):
                for date, sell_s, sell_a in df[needed_cols].itertuples(index=False, name=None):
                    try:
                        sell_s = int(sell_s)
                        sell_a = int(sell_a)
                    except Exception:
                        continue

                    if sell_s > 0:
                        rows.append({
                            "日期": normalize_date_str(date),
                            "權證代號": key[1],
                            "賣出股數": sell_s,
                            "賣出金額": sell_a,
                        })

    rows = sorted(rows, key=lambda x: (x["日期"], x["權證代號"]))
    _GROUP_OUTCOME_SALE_ROWS_CACHE[key] = rows
    return rows


def simulate_group_outcomes_fifo(events, item_map):
    """
    對同一類別的新制金額強度事件進行「跨事件、逐筆 FIFO」扣減。

    規則：
    1. 每個符合條件的事件各自保留成一筆資料，不合併。
    2. 同一分點 + 同一權證的賣出，只能使用一次。
    3. 賣出依事件成立時間由早到晚扣；舊事件未扣完前，不得先扣新事件。
    4. 同日買賣不互扣：賣出日必須晚於該事件結束日。
    5. 減碼 / 出清只看事件 lots 的剩餘股數；金額只計算已實現報酬。
    """
    if not events:
        return events

    queues = {}

    for event_seq, event in enumerate(events):
        event_start_date = normalize_date_str(
            event.get("起始日") or event.get("事件日") or event.get("結束日") or ""
        )
        event_end_date = normalize_date_str(
            event.get("結束日") or event.get("事件日") or event.get("起始日") or ""
        )
        broker_code = str(event.get("券商代號", "")).strip()

        event["減碼日"] = None
        event["減碼賣出金額"] = None
        event["減碼獲利%"] = None
        event["出清日"] = None
        event["出清賣出金額"] = None
        event["出清獲利%"] = None
        event["持有天數"] = None
        event["狀態"] = "持有"
        event["累計賣出股數"] = 0
        event["已實現賣出金額"] = 0.0
        event["已實現成本"] = 0.0
        event["_fifo_allocations"] = []

        original_total = 0
        valid_lots = []

        for lot_seq, lot in enumerate(event.get("lots", [])):
            try:
                qty = int(lot.get("股數", 0) or 0)
                amount = float(lot.get("金額", 0) or 0)
            except Exception:
                continue

            warrant_code = normalize_warrant_code_for_unique(lot.get("權證代號", ""))
            if qty <= 0 or amount <= 0 or not broker_code or not warrant_code:
                continue

            buy_date = normalize_date_str(lot.get("買進日") or event_end_date)
            avg_cost = amount / qty if qty else 0

            lot["買進日"] = buy_date
            lot["股數"] = qty
            lot["金額"] = amount
            lot["均價"] = avg_cost
            lot["剩餘股數"] = qty
            lot["累計賣出股數"] = 0
            lot["已實現賣出金額"] = 0.0
            lot["已實現成本"] = 0.0

            valid_lots.append(lot)
            original_total += qty

            key = (broker_code, warrant_code)
            queues.setdefault(key, []).append({
                "event": event,
                "event_seq": event_seq,
                "lot": lot,
                "lot_seq": lot_seq,
                "event_start_date": event_start_date,
                "event_end_date": event_end_date,
                "buy_date": buy_date,
            })

        event["lots"] = valid_lots
        event["原始股數"] = original_total
        event["剩餘股數"] = original_total

    # 同一分點 + 權證依舊事件優先，再依 lot 買進日排序。
    for key, queue in queues.items():
        queue.sort(key=lambda ref: (
            ref["event_end_date"],
            ref["buy_date"],
            ref["event_seq"],
            ref["lot_seq"],
        ))

        broker_code, warrant_code = key
        sales = get_group_sale_rows_for_warrant(item_map, broker_code, warrant_code)

        for sale in sales:
            sell_date = normalize_date_str(sale.get("日期", ""))
            sell_qty = int(sale.get("賣出股數", 0) or 0)
            sell_amount = float(sale.get("賣出金額", 0) or 0)

            if not sell_date or sell_qty <= 0:
                continue

            sell_price = sell_amount / sell_qty if sell_qty > 0 else 0
            sell_left = sell_qty

            for ref in queue:
                if sell_left <= 0:
                    break

                # 事件視窗內的買賣已反映在事件成立金額 / 股數中；只扣事件結束日之後的賣出。
                if sell_date <= ref["event_end_date"]:
                    continue

                lot = ref["lot"]
                remaining = int(lot.get("剩餘股數", 0) or 0)
                if remaining <= 0:
                    continue

                alloc = min(sell_left, remaining)
                if alloc <= 0:
                    continue

                alloc_revenue = alloc * sell_price
                alloc_cost = alloc * float(lot.get("均價", 0) or 0)

                lot["剩餘股數"] = remaining - alloc
                lot["累計賣出股數"] = int(lot.get("累計賣出股數", 0) or 0) + alloc
                lot["已實現賣出金額"] = float(lot.get("已實現賣出金額", 0) or 0) + alloc_revenue
                lot["已實現成本"] = float(lot.get("已實現成本", 0) or 0) + alloc_cost

                event = ref["event"]
                event["_fifo_allocations"].append({
                    "日期": sell_date,
                    "股數": alloc,
                    "賣出金額": alloc_revenue,
                    "成本": alloc_cost,
                })

                sell_left -= alloc

    # 依每筆事件自己的剩餘股數判斷持有 / 減碼 / 出清。
    for event in events:
        original_total = int(sum(int(lot.get("股數", 0) or 0) for lot in event.get("lots", [])))
        remaining_total = int(sum(int(lot.get("剩餘股數", 0) or 0) for lot in event.get("lots", [])))
        sold_total = max(original_total - remaining_total, 0)

        allocations_by_date = {}
        for alloc in event.pop("_fifo_allocations", []):
            d = normalize_date_str(alloc.get("日期", ""))
            if not d:
                continue
            rec = allocations_by_date.setdefault(d, {
                "股數": 0,
                "賣出金額": 0.0,
                "成本": 0.0,
            })
            rec["股數"] += int(alloc.get("股數", 0) or 0)
            rec["賣出金額"] += float(alloc.get("賣出金額", 0) or 0)
            rec["成本"] += float(alloc.get("成本", 0) or 0)

        realized_revenue = sum(rec["賣出金額"] for rec in allocations_by_date.values())
        realized_cost = sum(rec["成本"] for rec in allocations_by_date.values())

        event["原始股數"] = original_total
        event["剩餘股數"] = remaining_total
        event["累計賣出股數"] = sold_total
        event["已實現賣出金額"] = realized_revenue
        event["已實現成本"] = realized_cost

        if sold_total <= 0:
            event["狀態"] = "持有"
            continue

        running_remaining = original_total
        cumulative_revenue = 0.0
        cumulative_cost = 0.0

        for sell_date in sorted(allocations_by_date.keys()):
            rec = allocations_by_date[sell_date]
            running_remaining = max(running_remaining - int(rec["股數"]), 0)
            cumulative_revenue += rec["賣出金額"]
            cumulative_cost += rec["成本"]

            if running_remaining > 0 and event.get("減碼日") is None:
                event["減碼日"] = sell_date
                event["減碼賣出金額"] = round(rec["賣出金額"], 0)
                event["減碼獲利%"] = round(
                    (rec["賣出金額"] - rec["成本"]) / rec["成本"] * 100,
                    2,
                ) if rec["成本"] else None

            if running_remaining <= 0:
                event["出清日"] = sell_date
                event["出清賣出金額"] = round(cumulative_revenue, 0)

                total_cost = float(sum(float(lot.get("金額", 0) or 0) for lot in event.get("lots", [])))
                event["出清獲利%"] = round(
                    (cumulative_revenue - total_cost) / total_cost * 100,
                    2,
                ) if total_cost else None

                start_dt = parse_date(event.get("起始日") or event.get("事件日"))
                exit_dt = parse_date(sell_date)
                if start_dt and exit_dt:
                    event["持有天數"] = (exit_dt - start_dt).days
                break

        # 最終狀態仍只依剩餘股數判斷。
        if remaining_total <= 0:
            event["剩餘股數"] = 0
            event["狀態"] = "出清"
        else:
            event["狀態"] = "減碼"

    return events


def simulate_group_outcome(event, item_map):
    """保留舊函式名稱供相容；單筆呼叫時仍使用同一套剩餘股數 FIFO 規則。"""
    result = simulate_group_outcomes_fifo([event], item_map)
    return result[0] if result else event


# ══════════════════════════════════════════════════════════════════════
# 舊版 B/C/D 事件建立邏輯已移除。
# 目前只使用 build_amount_class_events() 的新制 A/B/C/D/E 單日金額強度分類。
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# Step 4：抓收盤價
# ══════════════════════════════════════════════════════════════════════

def fetch_all_prices(a_events, b_events, c_events, d_events, e_events=None):
    print("【Step 4】抓收盤價...")

    code_ranges = {}

    def update_code_range(code, start_dt, end_dt):
        if not code or start_dt is None or end_dt is None:
            return

        code = normalize_price_code(code)

        if not code:
            return

        if end_dt < start_dt:
            end_dt = start_dt

        if code not in code_ranges:
            code_ranges[code] = [start_dt, end_dt]
        else:
            code_ranges[code][0] = min(code_ranges[code][0], start_dt)
            code_ranges[code][1] = max(code_ranges[code][1], end_dt)

    for _, group_events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in group_events:
            dt = parse_date(ev.get("事件日") or ev.get("結束日") or ev.get("起始日"))
            if dt:
                start_dt = dt - timedelta(days=60)
                end_dt = dt + timedelta(days=160)
                update_code_range(ev.get("標的股"), start_dt, end_dt)

                # 新制 A/B/C/D/E 的 D+ 欄位目前只看標的股；若未來需要群組事件權證明細價格，
                # 可設定 FETCH_GROUP_WARRANT_PRICES=1。
                if FETCH_GROUP_WARRANT_PRICES:
                    for lot in ev.get("lots", []):
                        update_code_range(lot.get("權證代號"), start_dt, end_dt)

    all_codes = list(code_ranges.keys())
    price_cache = {}
    total = len(all_codes)

    if total == 0:
        print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
        return price_cache

    persistent_price_cache = load_price_cache()
    print(f"  價格快取讀取：{len(persistent_price_cache):,} 個代號")

    today = datetime.today()
    fetch_plan = {}

    for code in all_codes:
        start_dt, end_dt = code_ranges[code]

        if end_dt > today:
            end_dt = today

        if start_dt > end_dt:
            start_dt = end_dt

        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)

        if cached_prices:
            add_price_aliases(price_cache, code, cached_prices)

            valid_dates = sorted([d for d, p in cached_prices.items() if p is not None and p > 0])
            latest_dt = parse_date(valid_dates[-1]) if valid_dates else None

            # 快取不只看最後日期，也要檢查覆蓋率品質。
            # 若快取筆數太少、有效日期不足，或覆蓋不到需求區間，仍重新補抓完整區間，避免 D+ 欄位大量出現 -。
            if prices_need_yahoo_fallback(cached_prices, start_dt, end_dt):
                fetch_plan[code] = [start_dt, end_dt]
            elif latest_dt and latest_dt < end_dt:
                fetch_start_dt = latest_dt + timedelta(days=1)

                if fetch_start_dt <= end_dt:
                    fetch_plan[code] = [fetch_start_dt, end_dt]
            elif not latest_dt:
                fetch_plan[code] = [start_dt, end_dt]
        else:
            add_price_aliases(price_cache, code, {})
            fetch_plan[code] = [start_dt, end_dt]

    print(f"  價格代號去重後：{total:,} 檔")
    print(f"  價格快取命中：{total - len(fetch_plan):,} 檔")
    print(f"  本次需補抓價格：{len(fetch_plan):,} 檔")

    if not fetch_plan:
        print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
        return price_cache

    def fetch_one(code):
        start_dt, end_dt = fetch_plan[code]
        return code, fetch_twse_prices(code, start_dt, end_dt)

    price_workers = PRICE_WORKERS
    print(f"  價格抓取執行緒：{price_workers}")

    done = 0
    changed_price_codes = set()

    with ThreadPoolExecutor(max_workers=price_workers) as ex:
        futures = {ex.submit(fetch_one, code): code for code in fetch_plan}

        for future in as_completed(futures):
            done += 1

            try:
                code, fetched_prices = future.result()
                old_prices = get_cached_prices_for_code(persistent_price_cache, code)
                merged_prices = merge_price_dicts(old_prices, fetched_prices)

                norm_code = normalize_price_code(code)

                if norm_code:
                    persistent_price_cache[norm_code] = merged_prices
                    if fetched_prices:
                        changed_price_codes.add(norm_code)

                add_price_aliases(price_cache, code, merged_prices)

            except:
                code = futures[future]
                old_prices = get_cached_prices_for_code(persistent_price_cache, code)
                add_price_aliases(price_cache, code, old_prices)

            if done % 20 == 0:
                print(f"  [{done}/{len(fetch_plan)}] 收盤價補抓中...")

    save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)

    print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
    return price_cache

# ══════════════════════════════════════════════════════════════════════
# D+1 ~ D+20
# ══════════════════════════════════════════════════════════════════════

def get_price_series_from_cache(price_cache, code):
    if not code:
        return {}

    code1 = str(code).strip()
    code2 = normalize_price_code(code1)
    code3 = code2.lstrip("0") if code2 else ""

    for c in [code1, code2, code3]:
        if c and c in price_cache:
            return price_cache.get(c, {})

    return {}


def price_dates_after(prices, base_date):
    base_date = normalize_date_str(base_date)
    return [d for d in sorted(prices.keys()) if d > base_date]


def build_dplus_dates(base_date, primary_prices, secondary_prices=None, limit=20):
    """
    建立 D+1 ~ D+20 的交易日序列。

    原本用「權證價格日期 + 標的價格日期」聯集，容易造成 D+ 日期亂跳；
    現在改成：
    1. 優先使用標的股價格日期，因為標的股交易日最完整。
    2. 標的股不足時，用權證價格日期補。
    3. 最後才用兩者聯集。
    """
    base_date = normalize_date_str(base_date)

    primary_dates = price_dates_after(primary_prices or {}, base_date)
    secondary_dates = price_dates_after(secondary_prices or {}, base_date)

    if len(primary_dates) >= limit:
        return primary_dates[:limit]

    merged = sorted(set(primary_dates + secondary_dates))
    return merged[:limit]


def get_price_on_or_before(prices, date):
    return get_price_nearest(prices, date)


def calc_pct_by_base(current_price, base_price):
    if current_price is None or base_price is None:
        return None

    try:
        current_price = float(current_price)
        base_price = float(base_price)

        if base_price <= 0 or current_price <= 0:
            return None

        return round((current_price - base_price) / base_price * 100, 2)
    except:
        return None


def get_buy_avg_as_base(ev):
    try:
        v = float(ev.get("買進均價"))
        return v if v > 0 else None
    except:
        return None


def make_a_day_cells(ev, item_map, price_cache):
    day_values = []
    day_status = []

    item = item_map.get((ev["券商代號"], ev["權證代號"]))
    if not item:
        return [""] * 20, ["none"] * 20

    w_prices = get_price_series_from_cache(price_cache, ev["權證代號"])
    u_prices = get_price_series_from_cache(price_cache, ev["標的股"]) if ev["標的股"] else {}

    buy_date = normalize_date_str(ev["買進日"])

    # 權證基準價：優先使用買進日之前最近收盤價；若沒有，使用實際買進均價補救。
    buy_w = get_price_on_or_before(w_prices, buy_date) if w_prices else None
    if buy_w is None:
        buy_w = get_buy_avg_as_base(ev)

    # 標的基準價：使用買進日之前最近收盤價。
    buy_u = get_price_on_or_before(u_prices, buy_date) if u_prices else None

    # D+ 日期以標的股交易日為主，權證日期為輔。
    future_dates = build_dplus_dates(
        buy_date,
        primary_prices=u_prices,
        secondary_prices=w_prices,
        limit=20
    )

    sell_by_date = item["df"].groupby("日期")["賣出股數"].sum().to_dict()

    stopped = False

    for i, check_date in enumerate(future_dates[:20]):
        check_date = normalize_date_str(check_date)

        # 權證與標的獨立抓價；一邊沒有資料，不影響另一邊。
        w_p = get_price_on_or_before(w_prices, check_date) if w_prices else None
        u_p = get_price_on_or_before(u_prices, check_date) if u_prices else None

        w_chg = calc_pct_by_base(w_p, buy_w)
        u_chg = calc_pct_by_base(u_p, buy_u)

        w_str = fmt_pct(w_chg)
        u_str = fmt_pct(u_chg)

        cell_text = f"權:{w_str}\n標:{u_str}"

        day_sell = int(sell_by_date.get(check_date, 0))

        if ev["出清日"] == check_date:
            status = "exit"
        elif ev["減碼日"] == check_date or day_sell > 0:
            status = "reduce"
        elif w_chg is not None and w_chg > 0:
            status = "win"
        else:
            status = "lose"

        day_values.append(cell_text)
        day_status.append(status)

        if ev["出清日"] and check_date >= ev["出清日"]:
            for _ in range(20 - i - 1):
                day_values.append("")
                day_status.append("none")
            stopped = True
            break

    if not stopped:
        while len(day_values) < 20:
            day_values.append("")
            day_status.append("none")

    return day_values, day_status


def make_group_day_cells(ev, price_cache):
    day_values = []
    day_status = []

    u_prices = get_price_series_from_cache(price_cache, ev["標的股"])
    base_date = normalize_date_str(ev["結束日"])

    buy_u = get_price_on_or_before(u_prices, base_date) if u_prices else None

    future_dates = build_dplus_dates(
        base_date,
        primary_prices=u_prices,
        secondary_prices=None,
        limit=20
    )

    stopped = False

    for i, check_date in enumerate(future_dates[:20]):
        check_date = normalize_date_str(check_date)

        u_p = get_price_on_or_before(u_prices, check_date) if u_prices else None
        u_chg = calc_pct_by_base(u_p, buy_u)

        cell_text = f"標:{fmt_pct(u_chg)}"

        if ev["出清日"] == check_date:
            status = "exit"
        elif ev["減碼日"] == check_date:
            status = "reduce"
        elif u_chg is not None and u_chg > 0:
            status = "win"
        else:
            status = "lose"

        day_values.append(cell_text)
        day_status.append(status)

        if ev["出清日"] and check_date >= ev["出清日"]:
            for _ in range(20 - i - 1):
                day_values.append("")
                day_status.append("none")
            stopped = True
            break

    if not stopped:
        while len(day_values) < 20:
            day_values.append("")
            day_status.append("none")

    return day_values, day_status


def fmt_price_value(v):
    if v is None:
        return "-"

    try:
        v = float(v)
        if v <= 0:
            return "-"
        if v.is_integer():
            return int(v)
        return round(v, 2)
    except:
        return "-"


def get_ma_status_cells(price_cache, underlying_code, base_date):
    """
    依照事件選出日計算標的股技術狀態：
    1. 標的股股價：選出日當天或之前最近一個交易日收盤價
    2. 5MA：若標的股股價 > 當天 5 日均線，顯示 ✓，否則空白
    3. 20MA：若當天 20 日均線 > 前一交易日 20 日均線，顯示 ✓，否則空白
    """
    prices = get_price_series_from_cache(price_cache, underlying_code)

    if not prices:
        return "-", "", ""

    base_date = normalize_date_str(base_date)
    valid_rows = []

    for d, p in prices.items():
        dt = parse_date(d)
        price = safe_price_float(p)

        if not dt or price is None:
            continue

        if normalize_date_str(d) <= base_date:
            valid_rows.append((normalize_date_str(d), price))

    valid_rows = sorted(valid_rows, key=lambda x: x[0])

    if not valid_rows:
        return "-", "", ""

    current_price = valid_rows[-1][1]
    close_values = [p for _, p in valid_rows]

    ma5_mark = ""
    ma20_mark = ""

    if len(close_values) >= 5:
        ma5 = sum(close_values[-5:]) / 5
        if current_price > ma5:
            ma5_mark = "✓"

    if len(close_values) >= 21:
        ma20_today = sum(close_values[-20:]) / 20
        ma20_prev = sum(close_values[-21:-1]) / 20
        if ma20_today > ma20_prev:
            ma20_mark = "✓"

    return fmt_price_value(current_price), ma5_mark, ma20_mark



# ══════════════════════════════════════════════════════════════════════
# TOP15 圖片用：固定部位明細與共識淨買超資料集
# ══════════════════════════════════════════════════════════════════════

def get_latest_price_info_on_or_before(price_cache, code, target_date):
    """
    取得指定代號在 target_date 當天或之前最近一筆有效收盤價。

    回傳：
    - (價格, 日期字串)
    - 找不到則回傳 (None, "")
    """
    prices = get_price_series_from_cache(price_cache, code)

    if not prices:
        return None, ""

    target_str = normalize_date_str(target_date)
    valid = []

    for d, p in prices.items():
        dt = parse_date(d)
        price = safe_price_float(p)

        if not dt or price is None:
            continue

        d_norm = normalize_date_str(d)
        if d_norm <= target_str:
            valid.append((d_norm, price))

    if not valid:
        return None, ""

    valid.sort(key=lambda x: x[0])
    return valid[-1][1], valid[-1][0]


def collect_top15_return_recent_dates(a_events, b_events, c_events, d_events, e_events=None, lookback_days=None):
    """
    從 A/B/C/D/E 事件抓近 N 個有效事件交易日。

    這個範圍要和 TOP15 圖的「近一個月交易日」概念一致，
    讓報酬率快取可以直接對應圖片中的 TOP15 參與分點。
    """
    if lookback_days is None:
        lookback_days = TOP15_LOOKBACK_TRADING_DAYS

    dates = set()

    for _, group_events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in group_events:
            d = parse_date(ev.get("事件日") or ev.get("結束日") or ev.get("起始日") or ev.get("買進日"))
            if d:
                dates.add(normalize_date_str(d.strftime("%Y/%m/%d")))

    recent_dates = sorted(dates, reverse=True)[:max(int(lookback_days), 1)]
    return recent_dates


def _top15_return_event_date(ev, is_a=False):
    if is_a:
        return normalize_date_str(ev.get("買進日") or ev.get("事件日") or "")
    return normalize_date_str(ev.get("事件日") or ev.get("結束日") or ev.get("起始日") or "")


def collect_top15_return_position_lots(a_events, b_events, c_events, d_events, e_events=None, recent_dates=None, item_map=None):
    """
    將近 N 個有效事件交易日內的 A/B/C/D/E 買超事件轉成可計算未實現報酬率的 lot。

    定義：
    - 這裡只看「近 N 個有效事件交易日內」被 A/B/C/D/E 納入的權證部位。
    - 報酬率要表達的是：這批近 N 日買進後，持有到目前的帳面報酬。
    - A 直接使用原本單檔買進毛額。
    - B/C/D 優先回到 item_map 的原始逐日資料抓「毛買進金額 / 毛買進股數」建立 lot；
      後續賣出一律交給 apply_sales_to_top15_return_lots() 統一扣減。
      這樣報酬率就是「當時實際買進均價 vs 目前權證價格」，不會用淨額先扣一次又再扣賣出。
    """
    date_set = set(recent_dates or [])
    lots = []
    item_map = item_map or {}

    def _find_item(broker_code, warrant_code):
        broker_code = str(broker_code or "").strip()
        warrant_code = str(warrant_code or "").strip()

        if not broker_code or not warrant_code:
            return None

        candidates = []
        for wc in [
            warrant_code,
            normalize_warrant_code_for_unique(warrant_code),
            normalize_price_code(warrant_code),
        ]:
            wc = str(wc or "").strip()
            if wc and wc not in candidates:
                candidates.append(wc)

        for wc in candidates:
            item = item_map.get((broker_code, wc))
            if item:
                return item

        return None

    def add_lot(
        event_code, event_type, broker_label, broker_name, broker_code,
        underlying_code, underlying_name, event_date, buy_date,
        warrant_code, warrant_name, buy_amount, buy_qty, source_text
    ):
        event_date = normalize_date_str(event_date)
        buy_date = normalize_date_str(buy_date or event_date)

        if not event_date or event_date not in date_set:
            return False

        warrant_code = normalize_warrant_code_for_unique(warrant_code)
        buy_amount = float(buy_amount or 0)
        buy_qty = float(buy_qty or 0)

        if not warrant_code or buy_amount <= 0 or buy_qty <= 0:
            return False

        lots.append({
            "事件": event_code,
            "事件類型": event_type,
            "事件日": event_date,
            "買進日": buy_date,
            "分點": str(broker_label or "").strip(),
            "分點名稱": str(broker_name or "").strip(),
            "券商代號": str(broker_code or "").strip(),
            "標的股": str(underlying_code or "").strip(),
            "標的名稱": str(underlying_name or "").strip(),
            "權證代號": warrant_code,
            "權證名稱": str(warrant_name or "").strip(),
            "原始股數": buy_qty,
            "原始成本": buy_amount,
            "剩餘股數": buy_qty,
            "剩餘成本": buy_amount,
            "來源": source_text,
        })
        return True

    def add_group_lots_from_history(event_code, event_type, ev, lot, event_date, start_date, end_date):
        """
        B/C/D TOP15 報酬率用毛買進資料建立 lot。
        回傳 True 代表已成功用原始流水建立；False 則外層可退回舊欄位資料。
        """
        event_date = normalize_date_str(event_date)
        if not event_date or event_date not in date_set:
            return False

        warrant_code = normalize_warrant_code_for_unique(lot.get("權證代號", ""))
        if not warrant_code:
            return False

        start_date = normalize_date_str(start_date or event_date)
        end_date = normalize_date_str(end_date or start_date)

        item = _find_item(ev.get("券商代號", ""), warrant_code)
        if not item:
            return False

        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            return False

        added = False
        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2.sort_values("日期").reset_index(drop=True)
        sell_return_map = _daily_sell_fifo_return_map_for_item(item)

        for row in df2.itertuples(index=False):
            row_dict = row._asdict()
            buy_date = normalize_date_str(row_dict.get("日期", ""))

            if not buy_date or buy_date < start_date or buy_date > end_date:
                continue

            buy_qty = float(row_dict.get("買進股數", 0) or 0)
            buy_amount = float(row_dict.get("買進金額", 0) or 0)

            if buy_qty <= 0 or buy_amount <= 0:
                continue

            added = add_lot(
                event_code,
                event_type,
                ev.get("分點", ""),
                ev.get("分點名稱", ""),
                ev.get("券商代號", ""),
                ev.get("標的股", ""),
                ev.get("標的名稱", ""),
                event_date,
                buy_date,
                warrant_code,
                lot.get("權證名稱", ""),
                buy_amount,
                buy_qty,
                f'{event_code} | {buy_date} | {warrant_code} {lot.get("權證名稱", "")}',
            ) or added

        return added

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            event_date = _top15_return_event_date(ev, is_a=False)
            event_type = ev.get("事件類型", f"{event_code}-事件")

            for lot in ev.get("lots", []):
                if event_code == "D":
                    # D 是近 N 日累積事件，報酬率要用該 D 視窗內的實際毛買進流水。
                    buy_start_date = ev.get("起始日") or lot.get("買進日") or event_date
                    buy_end_date = ev.get("結束日") or event_date
                else:
                    # B/C 的 lot 本身已有實際買進日，直接抓該日毛買進流水。
                    buy_start_date = lot.get("買進日") or ev.get("事件日") or ev.get("結束日") or event_date
                    buy_end_date = buy_start_date

                used_history = add_group_lots_from_history(
                    event_code,
                    event_type,
                    ev,
                    lot,
                    event_date,
                    buy_start_date,
                    buy_end_date,
                )

                if used_history:
                    continue

                # 備援：若 item_map 找不到原始流水，才沿用事件內既有 lot 金額。
                # 這個分支正常情況很少用到，保留是避免舊快取缺資料時整批報酬率消失。
                add_lot(
                    event_code,
                    event_type,
                    ev.get("分點", ""),
                    ev.get("分點名稱", ""),
                    ev.get("券商代號", ""),
                    ev.get("標的股", ""),
                    ev.get("標的名稱", ""),
                    event_date,
                    lot.get("買進日") or event_date,
                    lot.get("權證代號", ""),
                    lot.get("權證名稱", ""),
                    lot.get("金額", 0),
                    lot.get("股數", 0),
                    f'{event_code} | {lot.get("權證代號", "")} {lot.get("權證名稱", "")}',
                )

    return lots



def apply_sales_to_top15_return_lots(position_lots, item_map, target_date):
    """
    依照原始分點歷史資料，把近 N 日事件 lot 買進日之後的賣出股數扣掉，得到目前剩餘部位。

    扣減邏輯：
    - 同一分點 + 同一權證代號的 lot 依「買進日」FIFO 扣。
    - 只扣「賣出日 > 買進日」的賣出，避免權證不可當沖時，同日賣出誤扣當日新買。
    - 扣掉的是賣出股數對應的原始成本，不是賣出成交金額。
    - 這裡不回溯計算 22 日以前舊庫存，因為 TOP15 報酬率定義為近 N 日事件部位的帳面報酬。
    """
    if not position_lots:
        return position_lots

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()

    lots_by_key = {}
    for lot in position_lots:
        key = (str(lot.get("券商代號", "")).strip(), str(lot.get("權證代號", "")).strip())
        lots_by_key.setdefault(key, []).append(lot)

    for key, lots in lots_by_key.items():
        broker_code, warrant_code = key
        item = item_map.get((broker_code, warrant_code))

        if not item:
            item = item_map.get((broker_code, normalize_price_code(warrant_code)))

        if not item:
            continue

        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            continue

        lots.sort(key=lambda x: (x.get("買進日", "") or x.get("事件日", ""), x.get("事件日", ""), x.get("權證代號", "")))
        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2.sort_values("日期").reset_index(drop=True)

        for row in df2.itertuples(index=False):
            row_dict = row._asdict()
            sell_date = normalize_date_str(row_dict.get("日期", ""))
            sell_dt = parse_date(sell_date)

            if not sell_dt or sell_dt > target_dt:
                continue

            sell_qty_left = float(row_dict.get("賣出股數", 0) or 0)
            if sell_qty_left <= 0:
                continue

            for lot in lots:
                if sell_qty_left <= 0:
                    break

                buy_dt = parse_date(lot.get("買進日") or lot.get("事件日", ""))
                if not buy_dt or sell_dt <= buy_dt:
                    continue

                remaining_qty = float(lot.get("剩餘股數", 0) or 0)
                remaining_cost = float(lot.get("剩餘成本", 0) or 0)
                original_qty = float(lot.get("原始股數", 0) or 0)
                original_cost = float(lot.get("原始成本", 0) or 0)

                if remaining_qty <= 0 or remaining_cost <= 0 or original_qty <= 0 or original_cost <= 0:
                    continue

                avg_cost = original_cost / original_qty
                alloc_qty = min(sell_qty_left, remaining_qty)
                alloc_cost = min(remaining_cost, alloc_qty * avg_cost)

                lot["剩餘股數"] = max(remaining_qty - alloc_qty, 0)
                lot["剩餘成本"] = max(remaining_cost - alloc_cost, 0)
                sell_qty_left -= alloc_qty

    return position_lots



def ensure_top15_return_warrant_prices(price_cache, position_lots, target_date):
    """
    TOP15 固定資料集需要權證目前價格。

    原本 fetch_all_prices() 為了加速，B/C/D 預設只抓標的股價格，
    因此這裡會針對目前仍有剩餘部位的權證補抓最新價格，並同步回 price_cache.csv / Google Sheet。
    """
    if not position_lots:
        return price_cache

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()

    target_dt = min(target_dt, datetime.today())
    start_dt = target_dt - timedelta(days=max(TOP15_PRICE_LOOKBACK_DAYS, 10))

    needed_codes = sorted({
        normalize_price_code(lot.get("權證代號", ""))
        for lot in position_lots
        if float(lot.get("剩餘成本", 0) or 0) > 0
        and normalize_price_code(lot.get("權證代號", ""))
    })

    if not needed_codes:
        return price_cache

    persistent_price_cache = load_price_cache()
    fetch_plan = []

    for code in needed_codes:
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        current_prices = get_price_series_from_cache(price_cache, code)
        merged_prices = merge_price_dicts(cached_prices, current_prices)

        if merged_prices:
            add_price_aliases(price_cache, code, merged_prices)
            persistent_price_cache[normalize_price_code(code)] = merged_prices

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_dt.strftime("%Y/%m/%d"))
        latest_dt = parse_date(latest_date) if latest_date else None

        need_fetch = latest_price is None
        if latest_dt and (target_dt - latest_dt).days > TOP15_PRICE_STALE_DAYS:
            need_fetch = True

        if need_fetch:
            fetch_plan.append(code)

    print(f"  TOP15固定資料集需檢查權證價格：{len(needed_codes):,} 檔")
    print(f"  TOP15固定資料集需補抓權證價格：{len(fetch_plan):,} 檔")

    if not fetch_plan:
        return price_cache

    def fetch_one(code):
        return code, fetch_twse_prices(code, start_dt, target_dt)

    done = 0
    changed_price_codes = set()
    with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as ex:
        futures = {ex.submit(fetch_one, code): code for code in fetch_plan}

        for future in as_completed(futures):
            done += 1
            code = futures[future]

            try:
                code, fetched_prices = future.result()
            except Exception:
                fetched_prices = {}

            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            merged_prices = merge_price_dicts(old_prices, fetched_prices)

            if merged_prices:
                norm_code = normalize_price_code(code)
                if norm_code:
                    persistent_price_cache[norm_code] = merged_prices
                    if fetched_prices:
                        changed_price_codes.add(norm_code)
                add_price_aliases(price_cache, code, merged_prices)

            if done % 20 == 0:
                print(f"  [{done}/{len(fetch_plan)}] TOP15固定資料集權證價格補抓中...")

    save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)
    return price_cache



def normalize_top15_target_date(target_date=None):
    raw = str(target_date or TOP15_TARGET_DATE or "").strip()

    if raw:
        dt = parse_date(raw)
        if not dt:
            raise RuntimeError(f"TOP15_TARGET_DATE 格式錯誤，請使用 YYYY/MM/DD 或 YYYY-MM-DD：{raw}")
        return dt.strftime("%Y/%m/%d")

    return datetime.today().strftime("%Y/%m/%d")


def top15_safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        s = str(value).replace(",", "").strip()
        if s in ("", "-", "None", "nan", "null"):
            return default
        return float(s)
    except Exception:
        return default


def top15_fmt_amount_wan(value):
    try:
        return f"{float(value) / 10000:.1f}萬"
    except Exception:
        return "0.0萬"


def build_top15_position_detail_and_consensus_rows(a_events, b_events, c_events, d_events, e_events=None, item_map=None, price_cache=None, target_date=None):
    """
    建立 TOP15 圖片用固定資料集。

    產出兩份資料：
    1. 快取_TOP15部位明細：一列代表一筆實際買進 lot，包含原始成本、賣出扣減、剩餘成本、價格與報酬率。
    2. 快取_TOP15共識淨買超：由部位明細加總而成，一列代表一檔標的股，圖片程式可直接讀取排名。

    嚴格規則：
    - 價格日期若距離估值日超過 TOP15_PRICE_STALE_DAYS，視為缺價格。
    - 若剩餘部位缺少有效權證價格，預設保留淨買超成本，但該筆部位不納入報酬率估算。
    - TOP15 總表完全由部位明細加總，不再另外重算。
    """
    item_map = item_map or {}
    price_cache = price_cache or {}

    if not TOP15_CACHE_ENABLED:
        return [], []

    print("【Step 4b】建立 TOP15 圖片用固定資料集...")

    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    recent_dates = collect_top15_return_recent_dates(
        a_events, b_events, c_events, d_events, e_events, TOP15_LOOKBACK_TRADING_DAYS
    )

    if not recent_dates:
        print("  ⚠️ TOP15固定資料集：沒有可用事件日期")
        return [], []

    period_text = f"{min(recent_dates)} ～ {max(recent_dates)}"

    position_lots = collect_top15_return_position_lots(
        a_events, b_events, c_events, d_events, e_events, recent_dates, item_map=item_map
    )

    if not position_lots:
        print("  ⚠️ TOP15固定資料集：沒有可用買超 lot")
        return [], []

    position_lots = apply_sales_to_top15_return_lots(position_lots, item_map, target_date)
    position_lots = [
        lot for lot in position_lots
        if top15_safe_float(lot.get("剩餘成本", 0)) > 0 and top15_safe_float(lot.get("剩餘股數", 0)) > 0
    ]

    if not position_lots:
        print("  ⚠️ TOP15固定資料集：扣除賣出後沒有剩餘部位")
        return [], []

    ensure_top15_return_warrant_prices(price_cache, position_lots, target_date)

    detail_rows = []
    validation_errors = []
    missing_price_rows = []

    for lot in position_lots:
        original_qty = top15_safe_float(lot.get("原始股數", 0))
        original_cost = top15_safe_float(lot.get("原始成本", 0))
        remaining_qty = top15_safe_float(lot.get("剩餘股數", 0))
        remaining_cost = top15_safe_float(lot.get("剩餘成本", 0))

        if original_qty <= 0 or original_cost <= 0 or remaining_qty <= 0 or remaining_cost <= 0:
            continue

        sold_qty = max(original_qty - remaining_qty, 0)
        sold_cost = max(original_cost - remaining_cost, 0)
        original_avg = original_cost / original_qty if original_qty else None
        remaining_avg = remaining_cost / remaining_qty if remaining_qty else None

        if remaining_qty - original_qty > 0.0001:
            validation_errors.append(
                f"剩餘股數大於原始股數：{lot.get('分點')} {lot.get('權證代號')} 原始={original_qty} 剩餘={remaining_qty}"
            )

        if remaining_cost - original_cost > 1:
            validation_errors.append(
                f"剩餘成本大於原始成本：{lot.get('分點')} {lot.get('權證代號')} 原始={original_cost} 剩餘={remaining_cost}"
            )

        latest_price, latest_price_date = get_latest_price_info_on_or_before(
            price_cache,
            lot.get("權證代號", ""),
            target_date,
        )

        price_status = "OK"
        latest_dt = parse_date(latest_price_date) if latest_price_date else None

        if latest_price is None:
            price_status = "缺價格"
        elif latest_dt and target_dt and (target_dt - latest_dt).days > TOP15_PRICE_STALE_DAYS:
            price_status = "價格過舊"
            latest_price = None
            latest_price_date = ""

        if latest_price is None:
            market_value = ""
            unrealized_pnl = ""
            return_pct = ""
            return_text = "-"
            if price_status == "缺價格":
                price_status = "未造市不計報酬率"
            elif price_status == "價格過舊":
                price_status = "價格過舊不計報酬率"

            missing_price_rows.append(
                f"TOP15剩餘部位不計報酬率：{lot.get('分點')} {lot.get('標的股')} {lot.get('權證代號')} {lot.get('權證名稱')}，剩餘成本={round(remaining_cost, 0)}，原因={price_status}"
            )

            if not TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN:
                validation_errors.append(
                    f"TOP15剩餘部位缺價格：{lot.get('分點')} {lot.get('標的股')} {lot.get('權證代號')} {lot.get('權證名稱')}，剩餘成本={round(remaining_cost, 0)}"
                )
        else:
            market_value_float = remaining_qty * latest_price
            unrealized_pnl_float = market_value_float - remaining_cost
            return_pct_float = unrealized_pnl_float / remaining_cost * 100 if remaining_cost > 0 else None
            market_value = round(market_value_float, 0)
            unrealized_pnl = round(unrealized_pnl_float, 0)
            return_pct = "" if return_pct_float is None else round(return_pct_float, 2)
            return_text = "-" if return_pct_float is None else f"{return_pct_float:+.2f}%"

        detail_rows.append({
            "資料範圍": get_result_data_scope(),
            "統計日期": target_date,
            "統計期間": period_text,
            "有效交易日數": len(recent_dates),
            "分點": str(lot.get("分點", "")).strip(),
            "分點名稱": str(lot.get("分點名稱", "")).strip(),
            "券商代號": str(lot.get("券商代號", "")).strip(),
            "標的股": str(lot.get("標的股", "")).strip(),
            "標的名稱": str(lot.get("標的名稱", "")).strip(),
            "事件": str(lot.get("事件", "")).strip(),
            "事件類型": str(lot.get("事件類型", "")).strip(),
            "事件日": normalize_date_str(lot.get("事件日", "")),
            "買進日": normalize_date_str(lot.get("買進日", "")),
            "權證代號": str(lot.get("權證代號", "")).strip(),
            "權證名稱": str(lot.get("權證名稱", "")).strip(),
            "原始股數": round(original_qty, 0),
            "原始成本": round(original_cost, 0),
            "原始均價": "" if original_avg is None else round(original_avg, 4),
            "已扣賣出股數": round(sold_qty, 0),
            "已扣賣出成本": round(sold_cost, 0),
            "剩餘股數": round(remaining_qty, 0),
            "剩餘成本": round(remaining_cost, 0),
            "剩餘均價": "" if remaining_avg is None else round(remaining_avg, 4),
            "最新權證價格": "" if latest_price is None else latest_price,
            "最新價格日期": latest_price_date,
            "目前市值": market_value,
            "未實現損益": unrealized_pnl,
            "報酬率": return_pct,
            "報酬率文字": return_text,
            "價格狀態": price_status,
            "完成狀態": "DONE",
            "來源": str(lot.get("來源", "")).strip(),
            "run_id": run_id,
            "更新時間": update_time,
        })

    if missing_price_rows:
        preview = "\n".join(missing_price_rows[:20])
        extra = "" if len(missing_price_rows) <= 20 else f"\n... 其餘 {len(missing_price_rows) - 20} 筆略"
        print(
            "  ⚠️ TOP15固定資料集：部分權證因多日未造市 / 無有效價格，不納入報酬率估算，但仍保留淨買超成本：\n"
            + preview
            + extra
        )

    if validation_errors and TOP15_FAIL_ON_MISSING_PRICE:
        preview = "\n".join(validation_errors[:20])
        extra = "" if len(validation_errors) <= 20 else f"\n... 其餘 {len(validation_errors) - 20} 筆略"
        raise RuntimeError(
            "TOP15固定資料集驗證失敗，為避免淨買超成本錯誤，本次 RUN 已中止：\n"
            + preview
            + extra
        )

    consensus_rows = build_top15_consensus_rows_from_detail(detail_rows, run_id, update_time)

    print(
        f"  ✅ TOP15固定資料集完成：部位明細 {len(detail_rows):,} 筆，"
        f"共識淨買超 {len(consensus_rows):,} 檔標的"
    )
    return detail_rows, consensus_rows


def build_top15_consensus_rows_from_detail(detail_rows, run_id, update_time):
    """
    由「快取_TOP15部位明細」加總出「快取_TOP15共識淨買超」。
    這裡不再讀 A/B/C/D/E 或分點歷史，確保圖片用資料只來自同一份固定明細。
    """
    if not detail_rows:
        return []

    agg = {}

    for row in detail_rows:
        underlying = str(row.get("標的股", "")).strip()
        if not underlying:
            continue

        key = underlying
        rec = agg.setdefault(key, {
            "資料範圍": row.get("資料範圍", get_result_data_scope()),
            "統計日期": row.get("統計日期", ""),
            "統計期間": row.get("統計期間", ""),
            "有效交易日數": row.get("有效交易日數", ""),
            "標的股": underlying,
            "標的名稱": str(row.get("標的名稱", "")).strip(),
            "淨買超成本": 0.0,
            "可估成本": 0.0,
            "缺價格成本": 0.0,
            "目前市值": 0.0,
            "未實現損益": 0.0,
            "分點": {},
            "事件集合": set(),
            "權證集合": set(),
            "權證清單": [],
            "最新價格日期集合": set(),
            "資料狀態": "OK",
        })

        remaining_cost = top15_safe_float(row.get("剩餘成本", 0))
        market_value = top15_safe_float(row.get("目前市值", 0), 0.0) if row.get("目前市值", "") != "" else 0.0
        pnl = top15_safe_float(row.get("未實現損益", 0), 0.0) if row.get("未實現損益", "") != "" else 0.0
        price_status = str(row.get("價格狀態", "")).strip()
        broker = str(row.get("分點", "")).strip()
        broker_name = str(row.get("分點名稱", "")).strip()
        broker_code = str(row.get("券商代號", "")).strip()
        event_code = str(row.get("事件", "")).strip()
        warrant_code = str(row.get("權證代號", "")).strip()
        warrant_name = str(row.get("權證名稱", "")).strip()
        latest_price_date = str(row.get("最新價格日期", "")).strip()

        rec["淨買超成本"] += remaining_cost

        if price_status == "OK":
            rec["可估成本"] += remaining_cost
            rec["目前市值"] += market_value
            rec["未實現損益"] += pnl
        else:
            rec["缺價格成本"] += remaining_cost
            rec["資料狀態"] = "部分報酬率未估"

        if event_code:
            rec["事件集合"].add(event_code)
        if warrant_code:
            rec["權證集合"].add(warrant_code)
            warrant_label = f"{warrant_code} {warrant_name}".strip()
            if warrant_label and warrant_label not in rec["權證清單"]:
                rec["權證清單"].append(warrant_label)
        if latest_price_date:
            rec["最新價格日期集合"].add(latest_price_date)

        broker_key = (broker, broker_name, broker_code)
        broker_rec = rec["分點"].setdefault(broker_key, {
            "分點": broker,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "淨買超成本": 0.0,
            "可估成本": 0.0,
            "缺價格成本": 0.0,
            "目前市值": 0.0,
            "未實現損益": 0.0,
            "事件集合": set(),
            "權證集合": set(),
        })
        broker_rec["淨買超成本"] += remaining_cost
        if price_status == "OK":
            broker_rec["可估成本"] += remaining_cost
            broker_rec["目前市值"] += market_value
            broker_rec["未實現損益"] += pnl
        else:
            broker_rec["缺價格成本"] += remaining_cost
        if event_code:
            broker_rec["事件集合"].add(event_code)
        if warrant_code:
            broker_rec["權證集合"].add(warrant_code)

    rows = []

    sorted_records = sorted(
        agg.values(),
        key=lambda r: (float(r.get("淨買超成本", 0) or 0), float(r.get("可估成本", 0) or 0)),
        reverse=True,
    )

    for rank, rec in enumerate(sorted_records, 1):
        total_cost = float(rec.get("淨買超成本", 0) or 0)
        estimated_cost = float(rec.get("可估成本", 0) or 0)
        missing_cost = float(rec.get("缺價格成本", 0) or 0)
        market_value = float(rec.get("目前市值", 0) or 0)
        pnl = float(rec.get("未實現損益", 0) or 0)
        return_pct = round(pnl / estimated_cost * 100, 2) if estimated_cost > 0 else None
        coverage_pct = round(estimated_cost / total_cost * 100, 2) if total_cost > 0 else None

        broker_rows = []
        broker_json = []

        for broker_rec in sorted(rec["分點"].values(), key=lambda x: x["淨買超成本"], reverse=True):
            b_cost = float(broker_rec.get("淨買超成本", 0) or 0)
            b_estimated_cost = float(broker_rec.get("可估成本", 0) or 0)
            b_pnl = float(broker_rec.get("未實現損益", 0) or 0)
            b_return_pct = round(b_pnl / b_estimated_cost * 100, 2) if b_estimated_cost > 0 else None
            b_events = "/".join(sorted(x for x in broker_rec["事件集合"] if x))
            b_warrant_count = len(broker_rec["權證集合"])
            b_return_text = "-" if b_return_pct is None else f"{b_return_pct:+.2f}%"

            b_missing_cost = float(broker_rec.get("缺價格成本", 0) or 0)
            b_coverage_text = ""
            if b_missing_cost > 0 and b_cost > 0:
                b_coverage_text = f"｜估值{b_estimated_cost / b_cost * 100:.0f}%"

            broker_rows.append(
                f"{broker_rec['分點']} {top15_fmt_amount_wan(b_cost)}（{b_return_text}{b_coverage_text}｜{b_events}｜{b_warrant_count}檔）"
            )
            broker_json.append({
                "分點": broker_rec["分點"],
                "分點名稱": broker_rec["分點名稱"],
                "券商代號": broker_rec["券商代號"],
                "淨買超成本": round(b_cost, 0),
                "可估成本": round(b_estimated_cost, 0),
                "缺價格成本": round(float(broker_rec.get("缺價格成本", 0) or 0), 0),
                "目前市值": round(float(broker_rec.get("目前市值", 0) or 0), 0),
                "未實現損益": round(b_pnl, 0),
                "報酬率": "" if b_return_pct is None else b_return_pct,
                "事件": b_events,
                "權證檔數": b_warrant_count,
            })

        rows.append({
            "資料範圍": rec.get("資料範圍", get_result_data_scope()),
            "統計日期": rec.get("統計日期", ""),
            "統計期間": rec.get("統計期間", ""),
            "有效交易日數": rec.get("有效交易日數", ""),
            "排名": rank,
            "標的股": rec.get("標的股", ""),
            "標的名稱": rec.get("標的名稱", ""),
            "淨買超成本": round(total_cost, 0),
            "可估成本": round(estimated_cost, 0),
            "缺價格成本": round(missing_cost, 0),
            "目前市值": round(market_value, 0),
            "未實現損益": round(pnl, 0),
            "報酬率": "" if return_pct is None else return_pct,
            "報酬率文字": "-" if return_pct is None else (f"{return_pct:+.2f}%" if missing_cost <= 0 else f"{return_pct:+.2f}%（部分估）"),
            "價格覆蓋率": "" if coverage_pct is None else coverage_pct,
            "價格覆蓋率文字": "-" if coverage_pct is None else f"{coverage_pct:.2f}%",
            "參與分點數": len(rec["分點"]),
            "參與分點明細": "\n".join(broker_rows),
            "事件": "/".join(sorted(x for x in rec["事件集合"] if x)),
            "權證檔數": len(rec["權證集合"]),
            "權證清單": "；".join(rec["權證清單"]),
            "最新價格日期": max(rec["最新價格日期集合"]) if rec["最新價格日期集合"] else "",
            "資料狀態": rec.get("資料狀態", "OK"),
            "完成狀態": "DONE",
            "更新時間": update_time,
            "run_id": run_id,
            "分點明細_JSON": json.dumps(broker_json, ensure_ascii=False),
        })

    return rows


def write_top15_position_detail_sheet(wb, detail_rows):
    """寫入 TOP15 部位明細固定資料集。"""
    ws = wb.create_sheet(TOP15_POSITION_DETAIL_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "有效交易日數",
        "分點", "分點名稱", "券商代號",
        "標的股", "標的名稱",
        "事件", "事件類型", "事件日", "買進日",
        "權證代號", "權證名稱",
        "原始股數", "原始成本", "原始均價",
        "已扣賣出股數", "已扣賣出成本",
        "剩餘股數", "剩餘成本", "剩餘均價",
        "最新權證價格", "最新價格日期",
        "目前市值", "未實現損益", "報酬率", "報酬率文字",
        "價格狀態", "完成狀態", "來源", "run_id", "更新時間",
    ]

    ws.append(headers)

    for row in detail_rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [12, 12, 24, 12, 14, 18, 12, 10, 12, 8, 22, 12, 12, 12, 24, 12, 14, 10, 14, 14, 12, 14, 10, 12, 12, 14, 14, 10, 12, 10, 10, 44, 16, 20]
    _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="價格狀態")


def write_top15_consensus_cache_sheet(wb, consensus_rows):
    """寫入 TOP15 共識淨買超固定資料集，圖片程式應直接讀這張表。"""
    ws = wb.create_sheet(TOP15_CONSENSUS_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "有效交易日數", "排名",
        "標的股", "標的名稱",
        "淨買超成本", "可估成本", "缺價格成本",
        "目前市值", "未實現損益", "報酬率", "報酬率文字",
        "價格覆蓋率", "價格覆蓋率文字",
        "參與分點數", "參與分點明細",
        "事件", "權證檔數", "權證清單",
        "最新價格日期", "資料狀態", "完成狀態",
        "更新時間", "run_id", "分點明細_JSON",
    ]

    ws.append(headers)

    for row in consensus_rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [12, 12, 24, 12, 8, 10, 12, 14, 14, 14, 14, 14, 10, 12, 12, 14, 12, 48, 10, 10, 60, 12, 10, 10, 20, 16, 80]
    _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="資料狀態")


def _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="資料狀態"):
    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    return_col_idx = header_map.get(return_col_name)
    status_col_idx = header_map.get(status_col_name)

    for row in ws.iter_rows(min_row=2):
        pct = None
        status_text = ""

        if return_col_idx:
            try:
                value = row[return_col_idx - 1].value
                if value not in (None, "", "-"):
                    pct = float(value)
            except Exception:
                pct = None

        if status_col_idx:
            status_text = str(row[status_col_idx - 1].value or "").strip()

        if status_text and status_text != "OK":
            row_fill = ORANGE
        elif pct is not None and pct > 0:
            row_fill = RED
        elif pct is not None and pct < 0:
            row_fill = GREEN
        else:
            row_fill = WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 30

    ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════
# Excel 樣式
# ══════════════════════════════════════════════════════════════════════

def style_sheet(ws, col_widths, status_rows=None, header_row=1):
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 24

    fill_map = {
        "win": RED,
        "lose": GREEN,
        "reduce": BLUE,
        "exit": ORANGE,
        "none": PatternFill(),
    }

    if status_rows:
        for r_idx, status_row in enumerate(status_rows, header_row + 1):
            for c_idx, status in enumerate(status_row, 1):
                cell = ws.cell(r_idx, c_idx)
                cell.fill = fill_map.get(status, PatternFill())
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[r_idx].height = 36
    else:
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"


def apply_exit_profit_result_outline(ws, row_idx, col_idx, return_pct):
    """
    針對實際「減碼獲利% / 出清獲利%」欄位加外框：
    獲利% > 0：粗紅色外框
    獲利% < 0：粗綠色外框

    不再框 D+N 橘色出清格，避免把「收盤價漲跌」誤解成實際出清損益。
    """
    if return_pct is None:
        return

    try:
        pct = float(return_pct)
    except:
        return

    if pct > 0:
        side = PROFIT_EXIT_SIDE
    elif pct < 0:
        side = LOSS_EXIT_SIDE
    else:
        return

    cell = ws.cell(row_idx, col_idx)
    cell.border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )


# ══════════════════════════════════════════════════════════════════════
# 建立 Excel：新制 A / B / C / D / E / 勝率統計
# ══════════════════════════════════════════════════════════════════════

def write_group_sheet(wb, sheet_name, events, price_cache, is_c=False):
    ws = wb.create_sheet(sheet_name)

    day_cols = [f"D+{i}" for i in range(1, 21)]

    if is_c:
        headers = [
            "事件類型", "分點", "標的股",
            "起始日", "結束日", "標的股股價", "5MA", "20MA",
            "涵蓋權證數", "權證清單",
            "最大單筆金額", "最大單筆權證",
            "單日累積買進金額", "買超張數",
            "減碼日", "減碼賣出金額", "減碼獲利%",
            "出清日", "出清賣出金額", "出清獲利%",
            "持有天數",
        ] + day_cols
        fixed_len = 21
    else:
        headers = [
            "事件類型", "分點", "標的股",
            "事件日", "標的股股價", "5MA", "20MA",
            "涵蓋權證數", "權證清單",
            "最大單筆金額", "最大單筆權證",
            "單日累積買進金額", "買超張數",
            "減碼日", "減碼賣出金額", "減碼獲利%",
            "出清日", "出清賣出金額", "出清獲利%",
            "持有天數",
        ] + day_cols
        fixed_len = 20

    ws.append(headers)
    status_rows = []
    exit_profit_result_cells = []

    sort_date_field = "結束日" if is_c else "事件日"
    sorted_events = sorted(
        events,
        key=lambda x: (
            -((parse_date(x.get(sort_date_field) or x.get("事件日") or x.get("起始日")) or datetime.min).toordinal()),
            str(x.get("分點", "")),
            str(x.get("標的股", "")),
        )
    )

    for ev in sorted_events:
        day_values, day_status = make_group_day_cells(ev, price_cache)
        ma_base_date = ev.get("結束日") if is_c else ev.get("事件日")
        underlying_price, ma5_mark, ma20_mark = get_ma_status_cells(
            price_cache,
            ev.get("標的股"),
            ma_base_date,
        )

        if is_c:
            row = [
                ev["事件類型"],
                ev["分點"],
                ev["標的股"],
                ev["起始日"],
                ev["結束日"],
                underlying_price,
                ma5_mark,
                ma20_mark,
                ev["涵蓋權證數"],
                ev["權證清單"],
                fmt_amount(ev.get("最大單筆金額")),
                ev.get("最大單筆權證", "-"),
                fmt_amount(ev.get("單日累積買進金額", ev.get("買超金額"))),
                ev["買超張數"],
                ev["減碼日"] or "-",
                fmt_amount(ev["減碼賣出金額"]),
                fmt_pct(ev["減碼獲利%"]),
                ev["出清日"] or "-",
                fmt_amount(ev["出清賣出金額"]),
                fmt_pct(ev["出清獲利%"]),
                fmt_num(ev["持有天數"]),
            ] + day_values
        else:
            row = [
                ev["事件類型"],
                ev["分點"],
                ev["標的股"],
                ev["事件日"],
                underlying_price,
                ma5_mark,
                ma20_mark,
                ev["涵蓋權證數"],
                ev["權證清單"],
                fmt_amount(ev.get("最大單筆金額")),
                ev.get("最大單筆權證", "-"),
                fmt_amount(ev.get("單日累積買進金額", ev.get("買超金額"))),
                ev["買超張數"],
                ev["減碼日"] or "-",
                fmt_amount(ev["減碼賣出金額"]),
                fmt_pct(ev["減碼獲利%"]),
                ev["出清日"] or "-",
                fmt_amount(ev["出清賣出金額"]),
                fmt_pct(ev["出清獲利%"]),
                fmt_num(ev["持有天數"]),
            ] + day_values

        ws.append(row)
        current_row_idx = ws.max_row
        status_rows.append(["none"] * fixed_len + day_status)

        if ev["減碼獲利%"] is not None:
            # 新制 A/B/C/D/E 工作表第 16 欄為「減碼獲利%」；跨日起訖格式第 17 欄為「減碼獲利%」
            reduce_profit_col_idx = 17 if is_c else 16
            exit_profit_result_cells.append((current_row_idx, reduce_profit_col_idx, ev["減碼獲利%"]))

        if ev["出清獲利%"] is not None:
            # 新制 A/B/C/D/E 工作表第 19 欄為「出清獲利%」；跨日起訖格式第 20 欄為「出清獲利%」
            exit_profit_col_idx = 20 if is_c else 19
            exit_profit_result_cells.append((current_row_idx, exit_profit_col_idx, ev["出清獲利%"]))

    if is_c:
        col_widths = [20, 14, 8, 12, 12, 12, 8, 8, 12, 45, 14, 24, 16, 10, 12, 14, 12, 12, 14, 12, 10] + [14] * 20
    else:
        col_widths = [20, 14, 8, 12, 12, 8, 8, 12, 45, 14, 24, 16, 10, 12, 14, 12, 12, 14, 12, 10] + [14] * 20

    style_sheet(ws, col_widths, status_rows)

    for row_idx, col_idx, return_pct in exit_profit_result_cells:
        apply_exit_profit_result_outline(ws, row_idx, col_idx, return_pct)


def _safe_stat_amount(value):
    try:
        if value is None:
            return 0
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return 0
        return int(round(float(s)))
    except Exception:
        return 0


def collect_stat_records(a_events, b_events, c_events, d_events, e_events=None):
    records = []

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            return_pct = ev["出清獲利%"]
            closed = return_pct is not None

            records.append({
                "分點": ev["分點"],
                "事件類型": ev.get("事件類型", f"{event_code}-{AMOUNT_CLASS_LABELS.get(event_code, '事件')}"),
                "事件代碼": event_code,
                "是否出清": closed,
                "結果": calc_result_tag(return_pct),
                "持有天數": ev["持有天數"],
                "報酬%": return_pct,
                "買進金額": _safe_stat_amount(ev.get("買超金額", ev.get("單日累積買進金額", 0))),
            })

    return records

def calc_summary_for_group(g, broker, event_type):
    total_events = len(g)
    closed_g = g[g["是否出清"] == True]
    open_g = g[g["是否出清"] == False]

    closed_count = len(closed_g)
    open_count = len(open_g)

    win_count = int((closed_g["結果"] == "勝").sum()) if closed_count > 0 else 0
    loss_count = int((closed_g["結果"] == "敗").sum()) if closed_count > 0 else 0
    flat_count = int((closed_g["結果"] == "平手").sum()) if closed_count > 0 else 0

    win_rate = round(win_count / closed_count * 100, 2) if closed_count > 0 else None

    avg_holding_days = None
    if closed_count > 0:
        holding_series = pd.to_numeric(closed_g["持有天數"], errors="coerce").dropna()
        if len(holding_series) > 0:
            avg_holding_days = round(float(holding_series.mean()), 2)

    avg_return = None
    weighted_return = None
    max_return = None
    min_return = None
    total_entry_amount = 0
    closed_entry_amount = 0
    estimated_pnl_amount = None
    avg_entry_amount = None

    if "買進金額" in g.columns:
        total_entry_amount = int(pd.to_numeric(g["買進金額"], errors="coerce").fillna(0).sum())

    if closed_count > 0:
        return_series = pd.to_numeric(closed_g["報酬%"], errors="coerce").dropna()
        if len(return_series) > 0:
            avg_return = round(float(return_series.mean()), 2)
            max_return = round(float(return_series.max()), 2)
            min_return = round(float(return_series.min()), 2)

        if "買進金額" in closed_g.columns:
            weighted_df = closed_g.copy()
            weighted_df["報酬%"] = pd.to_numeric(weighted_df["報酬%"], errors="coerce")
            weighted_df["買進金額"] = pd.to_numeric(weighted_df["買進金額"], errors="coerce").fillna(0)
            weighted_df = weighted_df.dropna(subset=["報酬%"] )
            weighted_df = weighted_df[weighted_df["買進金額"] > 0]

            if not weighted_df.empty:
                closed_entry_amount = int(round(float(weighted_df["買進金額"].sum())))
                weighted_pnl = float((weighted_df["買進金額"] * weighted_df["報酬%"] / 100.0).sum())
                estimated_pnl_amount = int(round(weighted_pnl))
                weighted_return = round(weighted_pnl / closed_entry_amount * 100.0, 2) if closed_entry_amount > 0 else None
                avg_entry_amount = int(round(closed_entry_amount / len(weighted_df))) if len(weighted_df) > 0 else None

    return {
        "分點": broker,
        "事件類型": event_type,
        "事件數": total_events,
        "已出清筆數": closed_count,
        "未出清筆數": open_count,
        "勝筆數": win_count,
        "敗筆數": loss_count,
        "平手筆數": flat_count,
        "勝率": win_rate,
        "平均持有天數": avg_holding_days,
        "平均報酬%": avg_return,
        "加權報酬%": weighted_return,
        "總買進金額": total_entry_amount,
        "已出清買進金額": closed_entry_amount,
        "估算損益金額": estimated_pnl_amount,
        "平均單筆買進金額": avg_entry_amount,
        "最高報酬%": max_return,
        "最低報酬%": min_return,
    }


def calc_empty_summary(broker, event_type):
    return {
        "分點": broker,
        "事件類型": event_type,
        "事件數": 0,
        "已出清筆數": 0,
        "未出清筆數": 0,
        "勝筆數": 0,
        "敗筆數": 0,
        "平手筆數": 0,
        "勝率": None,
        "平均持有天數": None,
        "平均報酬%": None,
        "加權報酬%": None,
        "總買進金額": 0,
        "已出清買進金額": 0,
        "估算損益金額": None,
        "平均單筆買進金額": None,
        "最高報酬%": None,
        "最低報酬%": None,
    }


WINRATE_STATS_MIN_TOTAL_EVENTS = int(os.getenv("WINRATE_STATS_MIN_TOTAL_EVENTS", "10"))


def _winrate_sort_value(value):
    try:
        if value is None:
            return -1.0
        return float(value)
    except Exception:
        return -1.0


def sort_brokers_by_winrate_summary(summary_map, broker_order):
    """
    勝率統計工作表的分點排序。

    排序邏輯：
    1. 先看「全部-A+B+C+D+E合併」的事件數是否達到門檻。
       - ABCDE 總事件數 >= WINRATE_STATS_MIN_TOTAL_EVENTS 的分點排前面。
       - ABCDE 總事件數 < WINRATE_STATS_MIN_TOTAL_EVENTS 的分點全部排最後。
    2. 同一組內依「全部-A+B+C+D+E合併」勝率由高到低排序。
    3. 若總勝率一樣，依序比較 A / B / C / D / E 類勝率。
    4. 若勝率都一樣，再以 ABCDE 總事件數多到少排序。
    5. 最後保留原本分點順序作為穩定排序依據。
    """
    original_index = {broker: idx for idx, broker in enumerate(broker_order)}

    def key_func(broker):
        broker_summary = summary_map.get(broker, {})
        all_summary = broker_summary.get("ALL", {})
        total_events = int(all_summary.get("事件數") or 0)
        qualified_rank = 0 if total_events >= WINRATE_STATS_MIN_TOTAL_EVENTS else 1

        return (
            qualified_rank,
            -_winrate_sort_value(all_summary.get("勝率")),
            -_winrate_sort_value(broker_summary.get("A", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("B", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("C", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("D", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("E", {}).get("勝率")),
            -total_events,
            original_index.get(broker, 999999),
        )

    return sorted(broker_order, key=key_func)


def make_summary_map(stat_records):
    if not stat_records:
        stat_df = pd.DataFrame(columns=["分點", "事件代碼", "事件類型", "是否出清", "結果", "持有天數", "報酬%", "買進金額"])
    else:
        stat_df = pd.DataFrame(stat_records)

    broker_order = list(TARGET_PATTERNS.keys())

    if not stat_df.empty:
        for broker in sorted(stat_df["分點"].dropna().unique()):
            if broker not in broker_order:
                broker_order.append(broker)

    summary_map = {}

    event_types = {
        code: f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}"
        for code in AMOUNT_CLASS_CODES
    }
    event_types["ALL"] = "全部-A+B+C+D+E+E合併"

    for broker in broker_order:
        summary_map[broker] = {}

        broker_g = stat_df[stat_df["分點"] == broker] if not stat_df.empty else pd.DataFrame()

        for code in AMOUNT_CLASS_CODES:
            if not broker_g.empty:
                g = broker_g[broker_g["事件代碼"] == code]
            else:
                g = pd.DataFrame()

            if len(g) > 0:
                summary_map[broker][code] = calc_summary_for_group(g, broker, event_types[code])
            else:
                summary_map[broker][code] = calc_empty_summary(broker, event_types[code])

        if len(broker_g) > 0:
            summary_map[broker]["ALL"] = calc_summary_for_group(broker_g, broker, event_types["ALL"])
        else:
            summary_map[broker]["ALL"] = calc_empty_summary(broker, event_types["ALL"])

    broker_order = sort_brokers_by_winrate_summary(summary_map, broker_order)

    return summary_map, broker_order


def write_stats_sheet(wb, a_events, b_events, c_events, d_events, e_events=None):
    ws = wb.create_sheet("勝率統計")

    headers = [
        "分點",
        "事件類型",
        "事件數",
        "已出清筆數",
        "未出清筆數",
        "勝筆數",
        "敗筆數",
        "平手筆數",
        "勝率",
        "平均持有天數",
        "平均報酬%",
        "加權報酬%",
        "總買進金額",
        "已出清買進金額",
        "估算損益金額",
        "平均單筆買進金額",
        "最高報酬%",
        "最低報酬%",
    ]

    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)
    summary_map, broker_order = make_summary_map(stat_records)

    thin_gray = Side(style="thin", color="B7B7B7")
    medium_gray = Side(style="medium", color="999999")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    broker_border = Border(left=medium_gray, right=medium_gray, top=medium_gray, bottom=medium_gray)

    current_row = 1

    for broker in broker_order:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(current_row, 1)
        title_cell.value = f"分點：{broker}"
        title_cell.font = Font(bold=True, color="000000", size=12)
        title_cell.fill = GRAY
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = broker_border
        ws.row_dimensions[current_row].height = 22

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(current_row, col_idx)
            cell.fill = GRAY
            cell.border = broker_border

        current_row += 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[current_row].height = 24

        current_row += 1

        for code in AMOUNT_CLASS_CODES + ["ALL"]:
            row = summary_map[broker][code]

            values = [
                row["分點"],
                row["事件類型"],
                row["事件數"],
                row["已出清筆數"],
                row["未出清筆數"],
                row["勝筆數"],
                row["敗筆數"],
                row["平手筆數"],
                "-" if row["勝率"] is None else f'{row["勝率"]:.2f}%',
                "-" if row["平均持有天數"] is None else row["平均持有天數"],
                "-" if row["平均報酬%"] is None else f'{row["平均報酬%"]:+.2f}%',
                "-" if row["加權報酬%"] is None else f'{row["加權報酬%"]:+.2f}%',
                fmt_amount(row.get("總買進金額")),
                fmt_amount(row.get("已出清買進金額")),
                "-" if row.get("估算損益金額") is None else fmt_amount(row.get("估算損益金額")),
                "-" if row.get("平均單筆買進金額") is None else fmt_amount(row.get("平均單筆買進金額")),
                "-" if row["最高報酬%"] is None else f'{row["最高報酬%"]:+.2f}%',
                "-" if row["最低報酬%"] is None else f'{row["最低報酬%"]:+.2f}%',
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(current_row, col_idx)
                cell.value = value
                cell.font = Font(color="000000", bold=True if code == "ALL" else False)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = normal_border

                if code == "ALL":
                    cell.fill = PatternFill("solid", fgColor="EAF2F8")
                else:
                    cell.fill = WHITE

            ws.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1

    col_widths = [16, 24, 10, 12, 12, 10, 10, 10, 10, 14, 12, 12, 14, 16, 14, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A1"


def write_combo_winrate_sheet(wb, a_events, b_events, c_events, d_events, e_events=None):
    """
    新增「ABCDE組合勝率」工作表。

    統計邏輯：
    1. 以分點為單位。
    2. 組合包含 AB / AC / AD / BC / BD / CD / ABC / ABD / ACD / BCD / ABCDE。
    3. 只有該分點同時具備該組合內所有事件類型，才列入該組合勝率。
    4. 勝率只用「已出清」事件計算，未出清不列入勝敗。

    排版邏輯：
    參考「勝率統計」工作表，每個分點獨立區塊顯示，方便閱讀與截圖。
    """
    ws = wb.create_sheet("ABCDE組合勝率")

    headers = [
        "分點",
        "組合",
        "包含事件",
        "是否同時出現",
        "A事件數",
        "B事件數",
        "C事件數",
        "D事件數",
        "E事件數",
        "組合事件數",
        "已出清筆數",
        "未出清筆數",
        "勝筆數",
        "敗筆數",
        "平手筆數",
        "勝率",
        "平均持有天數",
        "平均報酬%",
        "最高報酬%",
        "最低報酬%",
    ]

    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)

    if not stat_records:
        stat_df = pd.DataFrame(columns=[
            "分點", "事件代碼", "事件類型", "是否出清",
            "結果", "持有天數", "報酬%", "買進金額"
        ])
    else:
        stat_df = pd.DataFrame(stat_records)

    broker_order = list(TARGET_PATTERNS.keys())

    if not stat_df.empty:
        for broker in sorted(stat_df["分點"].dropna().unique()):
            if broker not in broker_order:
                broker_order.append(broker)

    combo_defs = []
    for mask in range(1, 1 << len(AMOUNT_CLASS_CODES)):
        codes = [AMOUNT_CLASS_CODES[idx] for idx in range(len(AMOUNT_CLASS_CODES)) if mask & (1 << idx)]
        if len(codes) >= 2:
            combo_defs.append(("".join(codes), codes))

    thin_gray = Side(style="thin", color="B7B7B7")
    medium_gray = Side(style="medium", color="999999")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    broker_border = Border(left=medium_gray, right=medium_gray, top=medium_gray, bottom=medium_gray)

    current_row = 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    title_cell = ws.cell(current_row, 1)
    title_cell.value = "ABCDE 所有組合勝率統計（依分點）"
    title_cell.font = Font(bold=True, color="000000", size=14)
    title_cell.fill = YELLOW
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = normal_border
    ws.row_dimensions[current_row].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(current_row, col_idx)
        cell.fill = YELLOW
        cell.border = normal_border

    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    note_cell = ws.cell(current_row, 1)
    note_cell.value = "統計邏輯：只有該分點同時具備該組合內所有事件類型，才列入該組合勝率；勝率只用「已出清」事件計算，未出清不列入勝敗。"
    note_cell.font = Font(color="666666")
    note_cell.fill = WHITE
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.border = normal_border
    ws.row_dimensions[current_row].height = 24

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(current_row, col_idx)
        cell.fill = WHITE
        cell.border = normal_border

    current_row += 2

    for broker in broker_order:
        if stat_df.empty:
            broker_g = pd.DataFrame(columns=stat_df.columns)
        else:
            broker_g = stat_df[stat_df["分點"] == broker].copy()

        event_counts = {}
        for code in AMOUNT_CLASS_CODES:
            if broker_g.empty:
                event_counts[code] = 0
            else:
                event_counts[code] = int((broker_g["事件代碼"] == code).sum())

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(current_row, 1)
        title_cell.value = f"分點：{broker}"
        title_cell.font = Font(bold=True, color="000000", size=12)
        title_cell.fill = GRAY
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = broker_border
        ws.row_dimensions[current_row].height = 22

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(current_row, col_idx)
            cell.fill = GRAY
            cell.border = broker_border

        current_row += 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[current_row].height = 24

        current_row += 1

        for combo_name, combo_codes in combo_defs:
            has_all = all(event_counts.get(code, 0) > 0 for code in combo_codes)
            include_text = " + ".join(combo_codes)

            if has_all and not broker_g.empty:
                combo_g = broker_g[broker_g["事件代碼"].isin(combo_codes)].copy()

                combo_event_count = len(combo_g)
                closed_g = combo_g[combo_g["是否出清"] == True]
                open_g = combo_g[combo_g["是否出清"] == False]

                closed_count = len(closed_g)
                open_count = len(open_g)

                win_count = int((closed_g["結果"] == "勝").sum()) if closed_count > 0 else 0
                loss_count = int((closed_g["結果"] == "敗").sum()) if closed_count > 0 else 0
                flat_count = int((closed_g["結果"] == "平手").sum()) if closed_count > 0 else 0

                win_rate = round(win_count / closed_count * 100, 2) if closed_count > 0 else None

                avg_holding_days = None
                if closed_count > 0:
                    holding_series = pd.to_numeric(closed_g["持有天數"], errors="coerce").dropna()
                    if len(holding_series) > 0:
                        avg_holding_days = round(float(holding_series.mean()), 2)

                avg_return = None
                max_return = None
                min_return = None
                if closed_count > 0:
                    return_series = pd.to_numeric(closed_g["報酬%"], errors="coerce").dropna()
                    if len(return_series) > 0:
                        avg_return = round(float(return_series.mean()), 2)
                        max_return = round(float(return_series.max()), 2)
                        min_return = round(float(return_series.min()), 2)

                row_values = [
                    broker,
                    combo_name,
                    include_text,
                    "是",
                    event_counts["A"],
                    event_counts["B"],
                    event_counts["C"],
                    event_counts["D"],
                    event_counts["E"],
                    combo_event_count,
                    closed_count,
                    open_count,
                    win_count,
                    loss_count,
                    flat_count,
                    "-" if win_rate is None else f"{win_rate:.2f}%",
                    "-" if avg_holding_days is None else avg_holding_days,
                    "-" if avg_return is None else f"{avg_return:+.2f}%",
                    "-" if max_return is None else f"{max_return:+.2f}%",
                    "-" if min_return is None else f"{min_return:+.2f}%",
                ]
            else:
                row_values = [
                    broker,
                    combo_name,
                    include_text,
                    "否",
                    event_counts["A"],
                    event_counts["B"],
                    event_counts["C"],
                    event_counts["D"],
                    event_counts["E"],
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                ]

            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(current_row, col_idx)
                cell.value = value
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = normal_border

                if col_idx == 4:
                    if value == "是":
                        cell.fill = PatternFill("solid", fgColor="EAF2F8")
                    else:
                        cell.fill = GRAY
                else:
                    cell.fill = WHITE

            ws.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1

    col_widths = [16, 8, 14, 14, 10, 10, 10, 10, 10, 12, 12, 12, 10, 10, 10, 10, 14, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

def fmt_ratio_value(numerator, denominator):
    try:
        numerator = float(numerator)
        denominator = float(denominator)
        if denominator == 0:
            return "-"
        return f"{(numerator / denominator * 100):.2f}%"
    except:
        return "-"


def collect_recent_trade_date_sets(items, cutoff_dt):
    trade_dates = set()

    for item in items:
        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_dates.add(normalize_date_str(row_dict["日期"]))

    sorted_dates = sorted(trade_dates)
    last_5_dates = set(sorted_dates[-5:])
    last_20_dates = set(sorted_dates[-20:])

    return last_5_dates, last_20_dates


def write_recent_warrant_amount_ranking_sheet(wb, items):
    ws = wb.create_sheet("近兩月買賣金額排行")

    headers = [
        "排名",
        "權證代號",
        "權證名稱",
        "標的股",
        "標的名稱",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "近20日淨買進金額",
        "近5日淨買進金額",
        "近20日占比",
        "近5日占比",
        "買進分點",
        "最近買進日",
    ]

    ws.append(headers)

    cutoff_dt = datetime.today() - timedelta(days=RECENT_RANKING_DAYS)
    last_5_dates, last_20_dates = collect_recent_trade_date_sets(items, cutoff_dt)
    ranking_map = {}

    for item in items:
        warrant_code = item["warrant_code"]
        warrant_name = item["warrant_name"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")
        broker_label = item["broker_label"]

        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_date_str = normalize_date_str(row_dict["日期"])
            buy_amount = int(row_dict["買進金額"])
            sell_amount = int(row_dict["賣出金額"])

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            key = (warrant_code, warrant_name, underlying_code, underlying_name)

            if key not in ranking_map:
                ranking_map[key] = {
                    "權證代號": warrant_code,
                    "權證名稱": warrant_name,
                    "標的股": underlying_code,
                    "標的名稱": underlying_name,
                    "買進金額": 0,
                    "賣出金額": 0,
                    "淨買進金額": 0,
                    "近20日淨買進金額": 0,
                    "近5日淨買進金額": 0,
                    "分點買進金額": {},
                    "分點賣出金額": {},
                    "最近買進日": "",
                }

            rec = ranking_map[key]
            net_amount = buy_amount - sell_amount

            rec["買進金額"] += buy_amount
            rec["賣出金額"] += sell_amount
            rec["淨買進金額"] += net_amount

            if trade_date_str in last_20_dates:
                rec["近20日淨買進金額"] += net_amount

            if trade_date_str in last_5_dates:
                rec["近5日淨買進金額"] += net_amount

            if buy_amount > 0:
                rec["分點買進金額"][broker_label] = rec["分點買進金額"].get(broker_label, 0) + buy_amount
                if not rec["最近買進日"] or trade_date_str > rec["最近買進日"]:
                    rec["最近買進日"] = trade_date_str

            if sell_amount > 0:
                rec["分點賣出金額"][broker_label] = rec["分點賣出金額"].get(broker_label, 0) + sell_amount

    ranking_rows = [
        rec for rec in ranking_map.values()
        if rec["淨買進金額"] > 0
    ]

    ranking_rows = sorted(
        ranking_rows,
        key=lambda x: x["淨買進金額"],
        reverse=True
    )[:20]

    for rank, rec in enumerate(ranking_rows, 1):
        broker_net_rows = []

        for broker, buy_amount in rec["分點買進金額"].items():
            sell_amount = rec["分點賣出金額"].get(broker, 0)
            net_amount = buy_amount - sell_amount

            if net_amount > 0:
                broker_net_rows.append((broker, net_amount))

        broker_net_rows = sorted(
            broker_net_rows,
            key=lambda x: x[1],
            reverse=True
        )

        if broker_net_rows:
            broker_text = "；".join([f"{broker}({fmt_amount(amount)})" for broker, amount in broker_net_rows])
        else:
            broker_text = "-"

        ws.append([
            rank,
            rec["權證代號"],
            rec["權證名稱"],
            rec["標的股"],
            rec["標的名稱"] or "-",
            fmt_amount(rec["買進金額"]),
            fmt_amount(rec["賣出金額"]),
            fmt_amount(rec["淨買進金額"]),
            fmt_amount(rec["近20日淨買進金額"]),
            fmt_amount(rec["近5日淨買進金額"]),
            fmt_ratio_value(rec["近20日淨買進金額"], rec["淨買進金額"]),
            fmt_ratio_value(rec["近5日淨買進金額"], rec["淨買進金額"]),
            broker_text,
            rec["最近買進日"] or "-",
        ])

    col_widths = [8, 12, 24, 10, 12, 16, 16, 18, 18, 18, 12, 12, 70, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"


def write_underlying_broker_count_ranking_sheet(wb, items):
    ws = wb.create_sheet("近兩月分點數排行")

    headers = [
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "買進分點數",
        "買進分點清單",
        "近兩月買進金額",
        "近兩月賣出金額",
        "近兩月淨買進金額",
        "近20日淨買進金額",
        "近5日淨買進金額",
        "近20日占比",
        "近5日占比",
        "最近買進日",
    ]

    ws.append(headers)

    cutoff_dt = datetime.today() - timedelta(days=RECENT_RANKING_DAYS)
    last_5_dates, last_20_dates = collect_recent_trade_date_sets(items, cutoff_dt)
    underlying_map = {}

    for item in items:
        warrant_code = item["warrant_code"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")
        broker_label = item["broker_label"]

        if not underlying_code:
            continue

        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_date_str = normalize_date_str(row_dict["日期"])
            buy_amount = int(row_dict["買進金額"])
            sell_amount = int(row_dict["賣出金額"])
            buy_shares = int(row_dict["買進股數"])
            sell_shares = int(row_dict["賣出股數"])

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            if underlying_code not in underlying_map:
                underlying_map[underlying_code] = {
                    "標的股": underlying_code,
                    "標的名稱": underlying_name,
                    "權證集合": set(),
                    "分點資料": {},
                }

            rec = underlying_map[underlying_code]
            if not rec["標的名稱"] and underlying_name:
                rec["標的名稱"] = underlying_name

            broker_rec = rec["分點資料"].setdefault(broker_label, {
                "買進金額": 0,
                "賣出金額": 0,
                "淨買進金額": 0,
                "近20日淨買進金額": 0,
                "近5日淨買進金額": 0,
                "買進股數": 0,
                "賣出股數": 0,
                "淨買進股數": 0,
                "最近買進日": "",
                "權證集合": set(),
            })

            net_amount = buy_amount - sell_amount
            net_shares = buy_shares - sell_shares

            broker_rec["權證集合"].add(warrant_code)
            broker_rec["買進金額"] += buy_amount
            broker_rec["賣出金額"] += sell_amount
            broker_rec["淨買進金額"] += net_amount
            broker_rec["買進股數"] += buy_shares
            broker_rec["賣出股數"] += sell_shares
            broker_rec["淨買進股數"] += net_shares

            if trade_date_str in last_20_dates:
                broker_rec["近20日淨買進金額"] += net_amount

            if trade_date_str in last_5_dates:
                broker_rec["近5日淨買進金額"] += net_amount

            if buy_amount > 0:
                if not broker_rec["最近買進日"] or trade_date_str > broker_rec["最近買進日"]:
                    broker_rec["最近買進日"] = trade_date_str

    ranking_rows = []

    for rec in underlying_map.values():
        active_broker_rows = []
        active_warrants = set()
        active_buy_amount = 0
        active_sell_amount = 0
        active_net_amount = 0
        active_20_net_amount = 0
        active_5_net_amount = 0
        latest_buy_date = ""

        for broker, broker_rec in rec["分點資料"].items():
            # 分點若已出清，淨買進股數 <= 0，不列入分點數排行與金額統計。
            if broker_rec["淨買進股數"] <= 0:
                continue

            # 若金額面也沒有正向淨買進，避免已賣出獲利但零庫存的分點被列入。
            if broker_rec["淨買進金額"] <= 0:
                continue

            active_broker_rows.append((broker, broker_rec["淨買進金額"]))
            active_warrants.update(broker_rec["權證集合"])
            active_buy_amount += broker_rec["買進金額"]
            active_sell_amount += broker_rec["賣出金額"]
            active_net_amount += broker_rec["淨買進金額"]
            active_20_net_amount += broker_rec["近20日淨買進金額"]
            active_5_net_amount += broker_rec["近5日淨買進金額"]

            if broker_rec["最近買進日"] and (not latest_buy_date or broker_rec["最近買進日"] > latest_buy_date):
                latest_buy_date = broker_rec["最近買進日"]

        if active_net_amount <= 0:
            continue

        if not active_broker_rows:
            continue

        active_broker_rows = sorted(
            active_broker_rows,
            key=lambda x: x[1],
            reverse=True
        )

        rec["權證檔數"] = len(active_warrants)
        rec["買進分點數"] = len(active_broker_rows)
        rec["買進分點清單"] = "；".join([f"{broker}({fmt_amount(amount)})" for broker, amount in active_broker_rows])
        rec["近兩月買進金額"] = active_buy_amount
        rec["近兩月賣出金額"] = active_sell_amount
        rec["近兩月淨買進金額"] = active_net_amount
        rec["近20日淨買進金額"] = active_20_net_amount
        rec["近5日淨買進金額"] = active_5_net_amount
        rec["最近買進日"] = latest_buy_date

        ranking_rows.append(rec)

    ranking_rows = sorted(
        ranking_rows,
        key=lambda x: (x["買進分點數"], x["近兩月淨買進金額"]),
        reverse=True
    )[:20]

    for rank, rec in enumerate(ranking_rows, 1):
        ws.append([
            rank,
            rec["標的股"],
            rec["標的名稱"] or "-",
            rec["權證檔數"],
            rec["買進分點數"],
            rec["買進分點清單"],
            fmt_amount(rec["近兩月買進金額"]),
            fmt_amount(rec["近兩月賣出金額"]),
            fmt_amount(rec["近兩月淨買進金額"]),
            fmt_amount(rec["近20日淨買進金額"]),
            fmt_amount(rec["近5日淨買進金額"]),
            fmt_ratio_value(rec["近20日淨買進金額"], rec["近兩月淨買進金額"]),
            fmt_ratio_value(rec["近5日淨買進金額"], rec["近兩月淨買進金額"]),
            rec["最近買進日"] or "-",
        ])

    col_widths = [8, 10, 12, 10, 12, 70, 16, 16, 18, 18, 18, 12, 12, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"



def write_broker_query_sheet(wb, items):
    """
    新增「券商查詢」工作表：
    可用下拉選單選擇券商，並顯示該券商目前「標的股」相關權證總淨買進金額前 15 名。
    不是單一權證排名，而是同一券商買進同一標的股底下所有相關認購權證的合計排名。

    排名邏輯改回：淨買進金額由高到低排序。
    淨買進金額 = 買進金額 - 賣出金額，代表扣掉賣出後仍留在裡面的推估金額。
    這版不使用 FILTER / SORTBY，改為先在隱藏工作表預先整理每個券商前 15 名，
    再用 INDEX / MATCH 查詢，提高 Excel 相容性。
    """
    data_ws = wb.create_sheet("券商查詢資料")
    data_ws.sheet_state = "hidden"

    data_headers = [
        "分點",
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "權證清單",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "買進張數",
        "賣出張數",
        "淨買進張數",
        "最近買進日",
        "查詢鍵",
    ]
    data_ws.append(data_headers)

    broker_underlying_map = {}

    for item in items:
        broker_label = item["broker_label"]
        warrant_code = item["warrant_code"]
        warrant_name = item["warrant_name"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")

        if not underlying_code:
            continue

        df = item["df"]

        buy_amount = int(df["買進金額"].sum()) if not df.empty else 0
        sell_amount = int(df["賣出金額"].sum()) if not df.empty else 0
        buy_shares = int(df["買進股數"].sum()) if not df.empty else 0
        sell_shares = int(df["賣出股數"].sum()) if not df.empty else 0

        if buy_amount <= 0 and sell_amount <= 0:
            continue

        net_amount = buy_amount - sell_amount
        net_shares = buy_shares - sell_shares

        latest_buy_date = ""
        if not df.empty:
            buy_dates = df[df["買進金額"] > 0]["日期"].dropna().tolist()
            if buy_dates:
                latest_buy_date = max([normalize_date_str(d) for d in buy_dates])

        key = (broker_label, underlying_code)

        if key not in broker_underlying_map:
            broker_underlying_map[key] = {
                "分點": broker_label,
                "標的股": underlying_code,
                "標的名稱": underlying_name,
                "權證集合": set(),
                "權證清單": [],
                "買進金額": 0,
                "賣出金額": 0,
                "淨買進金額": 0,
                "買進股數": 0,
                "賣出股數": 0,
                "淨買進股數": 0,
                "最近買進日": "",
            }

        rec = broker_underlying_map[key]

        if not rec["標的名稱"] and underlying_name:
            rec["標的名稱"] = underlying_name

        rec["權證集合"].add(warrant_code)

        warrant_label = f"{warrant_code} {warrant_name}"
        if warrant_label not in rec["權證清單"]:
            rec["權證清單"].append(warrant_label)

        rec["買進金額"] += buy_amount
        rec["賣出金額"] += sell_amount
        rec["淨買進金額"] += net_amount
        rec["買進股數"] += buy_shares
        rec["賣出股數"] += sell_shares
        rec["淨買進股數"] += net_shares

        if latest_buy_date and (not rec["最近買進日"] or latest_buy_date > rec["最近買進日"]):
            rec["最近買進日"] = latest_buy_date

    broker_map = {}

    for rec in broker_underlying_map.values():
        # 券商查詢主排序改回「淨買進金額」由高到低，
        # 但列入條件必須同時符合：
        # 1. 買進金額 > 0：代表這個券商確實有買這個標的底下的權證
        # 2. 淨買進金額 > 0：避免買很多但賣更多、實際已轉為淨賣出的標的進榜
        # 3. 淨買進股數 > 0：避免金額面為正但張數面已經接近或完全出清
        #
        # 這樣比較符合「主力還留著沒賣的總金額」這個籌碼意圖，
        # 同時也不會出現淨買進金額為負的標的排在前面。
        if rec["買進金額"] <= 0:
            continue

        if rec["淨買進金額"] <= 0:
            continue

        if rec["淨買進股數"] <= 0:
            continue

        broker_label = rec["分點"]
        broker_map.setdefault(broker_label, []).append(rec)

    broker_order = list(TARGET_PATTERNS.keys())
    active_brokers = sorted(broker_map.keys())

    broker_list = []
    for broker in broker_order:
        if broker in active_brokers and broker not in broker_list:
            broker_list.append(broker)

    for broker in active_brokers:
        if broker not in broker_list:
            broker_list.append(broker)

    for broker in broker_list:
        rows = sorted(
            broker_map.get(broker, []),
            key=lambda x: (x["淨買進金額"], x["最近買進日"], x["買進金額"], x["標的股"]),
            reverse=True
        )[:15]

        for rank, rec in enumerate(rows, 1):
            warrant_list = sorted(rec["權證清單"])

            data_ws.append([
                broker,
                rank,
                rec["標的股"],
                rec["標的名稱"] or "-",
                len(rec["權證集合"]),
                "；".join(warrant_list),
                rec["買進金額"],
                rec["賣出金額"],
                rec["淨買進金額"],
                rec["買進股數"] // 1000,
                rec["賣出股數"] // 1000,
                rec["淨買進股數"] // 1000,
                rec["最近買進日"],
                f"{broker}|{rank}",
            ])

    broker_start_col = 16  # P 欄
    data_ws.cell(1, broker_start_col).value = "券商清單"

    for idx, broker in enumerate(broker_list, 2):
        data_ws.cell(idx, broker_start_col).value = broker

    for col_idx, width in enumerate([16, 8, 10, 12, 10, 70, 16, 16, 18, 12, 12, 12, 14, 24], 1):
        data_ws.column_dimensions[get_column_letter(col_idx)].width = width

    data_ws.column_dimensions[get_column_letter(broker_start_col)].width = 18

    ws = wb.create_sheet("券商查詢")

    ws["A1"] = "券商標的股淨買進金額前 15 名查詢"
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(bold=True, size=14, color="000000")
    ws["A1"].fill = YELLOW
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "選擇券商"
    ws["A2"].font = Font(bold=True, color="000000")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = GRAY

    ws["B2"] = broker_list[0] if broker_list else ""
    ws["B2"].font = Font(bold=True, color="000000")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].fill = PatternFill("solid", fgColor="EAF2F8")

    ws["D2"] = "排序邏輯：同一券商買進同一標的股底下所有相關權證合計；先排除淨買進金額<=0或淨買進張數<=0，再依淨買進金額由高到低排序。"
    ws.merge_cells("D2:M2")
    ws["D2"].font = Font(color="666666")
    ws["D2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if broker_list:
        broker_formula = f"'券商查詢資料'!$P$2:$P${len(broker_list) + 1}"
        dv = DataValidation(type="list", formula1=broker_formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws["B2"])

    headers = [
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "權證清單",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "買進張數",
        "賣出張數",
        "淨買進張數",
        "最近買進日",
    ]

    header_row = 5

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    max_data_row = max(data_ws.max_row, 2)

    for i in range(15):
        row_idx = header_row + 1 + i
        rank = i + 1

        ws.cell(row_idx, 1).value = rank

        for col_idx in range(2, 13):
            data_col_idx = col_idx + 1
            data_col_letter = get_column_letter(data_col_idx)

            formula = (
                f'=IFERROR(INDEX(券商查詢資料!${data_col_letter}$2:${data_col_letter}${max_data_row},'
                f'MATCH($B$2&"|"&$A{row_idx},券商查詢資料!$N$2:$N${max_data_row},0)),"")'
            )
            ws.cell(row_idx, col_idx).value = formula

    col_widths = [8, 10, 12, 10, 70, 16, 16, 18, 12, 12, 12, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=header_row + 15, min_col=1, max_col=12):
        for cell in row:
            cell.border = normal_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row > header_row:
                cell.font = Font(color="000000")

    for row_idx in range(header_row + 1, header_row + 16):
        ws.row_dimensions[row_idx].height = 32

    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A6"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except:
        pass


def write_price_status_sheet(wb, price_cache):
    ws = wb.create_sheet("價格抓取狀態")

    headers = [
        "代號",
        "價格筆數",
        "第一筆日期",
        "最後筆日期",
        "狀態",
    ]

    ws.append(headers)

    rows = []

    seen = set()

    for code, prices in price_cache.items():
        code = str(code).strip()

        if not code or code in seen:
            continue

        seen.add(code)

        valid_dates = sorted([d for d, p in prices.items() if p is not None and p > 0])

        if valid_dates:
            rows.append([
                code,
                len(valid_dates),
                valid_dates[0],
                valid_dates[-1],
                "OK",
            ])
        else:
            rows.append([
                code,
                0,
                "-",
                "-",
                "NO DATA",
            ])

    rows = sorted(rows, key=lambda x: (x[4], x[0]))

    for row in rows:
        ws.append(row)

    col_widths = [12, 12, 14, 14, 12]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border

            if cell.column == 5 and cell.value == "NO DATA":
                cell.fill = GREEN

        ws.row_dimensions[row[0].row].height = 20

    ws.freeze_panes = "A2"


def apply_global_amount_comma_format(wb):
    """
    全工作表套用金額千分位格式。

    只要欄位名稱包含「金額」，例如：
    買進金額、賣出金額、淨買進金額、近20日淨買進金額、買超金額，
    就統一顯示為 1,234,567。

    不處理「均價」、「比例」、「獲利%」等欄位，避免小數或百分比被誤改。
    """
    for ws in wb.worksheets:
        # 掃前 10 列，因為大部分工作表標題在第 1 列，
        # 「券商查詢」的欄位標題在第 5 列。
        for header_row in range(1, min(ws.max_row, 10) + 1):
            amount_cols = []

            for col_idx in range(1, ws.max_column + 1):
                header_value = ws.cell(header_row, col_idx).value
                header = str(header_value).strip() if header_value is not None else ""

                if not header:
                    continue

                if "金額" in header:
                    # 排除非金額欄位，避免誤格式化
                    if "%" in header or "比例" in header or "均價" in header:
                        continue

                    amount_cols.append(col_idx)

            if not amount_cols:
                continue

            for col_idx in amount_cols:
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(row_idx, col_idx)

                    if cell.value is None or cell.value == "":
                        continue

                    # 公式欄位不動，只套顯示格式
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.number_format = '#,##0'
                        continue

                    if str(cell.value).strip() == "-":
                        continue

                    try:
                        raw = str(cell.value).replace(",", "").strip()

                        if raw.startswith("-"):
                            numeric_part = raw[1:]
                        else:
                            numeric_part = raw

                        if numeric_part.replace(".", "", 1).isdigit():
                            num = float(raw)

                            if num.is_integer():
                                cell.value = int(num)
                                cell.number_format = '#,##0'
                            else:
                                cell.value = num
                                cell.number_format = '#,##0.00'

                    except:
                        pass


def write_color_legend_sheet(wb):
    ws = wb.create_sheet("顏色說明")

    headers = ["色塊", "顏色用途", "出現位置", "說明"]
    ws.append(headers)

    legend_rows = [
        {
            "fill": YELLOW,
            "用途": "標頭 / 欄位名稱",
            "位置": "所有工作表",
            "說明": "用於表格欄位名稱與重要標題列。",
        },
        {
            "fill": RED,
            "用途": "勝 / 上漲",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表該追蹤日為正報酬或上漲狀態；依台股習慣使用紅色表示上漲。",
        },
        {
            "fill": GREEN,
            "用途": "敗 / 下跌或未上漲",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表該追蹤日為負報酬、下跌或未形成正報酬；依台股習慣使用綠色表示下跌。",
        },
        {
            "fill": BLUE,
            "用途": "減碼",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表程式偵測到該分點後續有賣出，但尚未完全出清。",
        },
        {
            "fill": ORANGE,
            "用途": "出清",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表依 FIFO 推估，該筆事件已經被完全賣出。",
        },
        {
            "fill": WHITE,
            "border": "profit",
            "用途": "粗紅色外框",
            "位置": "A、B、C、D 工作表的「減碼獲利% / 出清獲利%」欄位",
            "說明": "代表該筆事件已減碼或出清，且實際獲利% > 0。",
        },
        {
            "fill": WHITE,
            "border": "loss",
            "用途": "粗綠色外框",
            "位置": "A、B、C、D 工作表的「減碼獲利% / 出清獲利%」欄位",
            "說明": "代表該筆事件已減碼或出清，且實際獲利% < 0。",
        },
        {
            "fill": GRAY,
            "用途": "分點區隔列",
            "位置": "勝率統計工作表",
            "說明": "用於區隔不同分點，讓每個分點的統計區塊更清楚。",
        },
        {
            "fill": PatternFill("solid", fgColor="EAF2F8"),
            "用途": "全部-A+B+C+D+E合併",
            "位置": "勝率統計工作表",
            "說明": "代表該分點 A、B、C、D 四類事件合併後的統計列。",
        },
        {
            "fill": WHITE,
            "用途": "一般資料列",
            "位置": "勝率統計工作表",
            "說明": "一般事件類型統計列，沒有特殊狀態標記。",
        },
    ]

    for row_info in legend_rows:
        ws.append(["", row_info["用途"], row_info["位置"], row_info["說明"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 1).fill = row_info["fill"]

        if row_info.get("border"):
            side = PROFIT_EXIT_SIDE if row_info.get("border") == "profit" else LOSS_EXIT_SIDE
            ws.cell(row_idx, 1).border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side,
            )

    col_widths = [12, 24, 34, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        legend_name = str(row[1].value).strip()
        is_profit_outline_legend = legend_name == "粗紅色外框"
        is_loss_outline_legend = legend_name == "粗綠色外框"

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if is_profit_outline_legend and cell.column == 1:
                cell.border = Border(
                    left=PROFIT_EXIT_SIDE,
                    right=PROFIT_EXIT_SIDE,
                    top=PROFIT_EXIT_SIDE,
                    bottom=PROFIT_EXIT_SIDE,
                )
            elif is_loss_outline_legend and cell.column == 1:
                cell.border = Border(
                    left=LOSS_EXIT_SIDE,
                    right=LOSS_EXIT_SIDE,
                    top=LOSS_EXIT_SIDE,
                    bottom=LOSS_EXIT_SIDE,
                )
            else:
                cell.border = normal_border

        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"




def build_event_warrant_source_map(a_events, b_events, c_events, d_events, e_events=None):
    """
    建立「分點 + 權證代號 -> A/B/C/D/E 事件來源」對照表。

    用途：
    1. 每日賣出明細直接來自原始分點歷史資料 items，不再用 A/B/C/D/E 工作表的減碼日推估。
    2. 若賣出的權證屬於任一新制 A/B/C/D/E 金額強度事件，也能標示它原本歸屬哪一類事件。
    3. 同一權證在不同分點互不混用；同一分點與權證有多次事件時，以事件日較新的事件為準。
    4. 一併保存出清日，讓每日賣出明細可排除已在該筆賣出日前出清的舊事件。
    """
    source_map = {}

    def put_source(broker, warrant_code, event_code, event_type, event_date, exit_date, event_source):
        broker = str(broker).strip()
        warrant_code = normalize_warrant_code_for_unique(warrant_code)

        if not warrant_code:
            return

        key = (broker, warrant_code)
        new_source = {
            "事件": event_code,
            "事件類型": event_type,
            "事件日": normalize_date_str(event_date),
            "出清日": normalize_date_str(ev_exit) if (ev_exit := exit_date) else "",
            "事件來源": event_source,
        }

        if key in source_map:
            old_event_dt = parse_date(source_map[key].get("事件日", ""))
            new_event_dt = parse_date(new_source.get("事件日", ""))

            if old_event_dt and (not new_event_dt or new_event_dt <= old_event_dt):
                return

        source_map[key] = new_source

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            event_type = ev.get("事件類型", "")
            event_date = ev.get("事件日") or ev.get("結束日") or ev.get("起始日")

            for lot in ev.get("lots", []):
                put_source(
                    ev.get("分點", ""),
                    lot.get("權證代號", ""),
                    event_code,
                    event_type,
                    event_date,
                    ev.get("出清日"),
                    f'{event_code} | {lot.get("權證代號", "")} {lot.get("權證名稱", "")}',
                )

    return source_map




def _fmt_daily_sell_return_pct(value):
    if value is None:
        return "-"
    try:
        return f"{float(value):+.2f}%"
    except Exception:
        return "-"


def _daily_sell_fifo_return_map_for_item(item):
    """
    用同一分點 + 同一權證的 API5 歷史資料，替每日賣出明細計算賣出報酬率。

    重點：
    1. 不再只依賴 A/B/C/D/E 事件，未歸類權證也能從快取_分點歷史回頭找買進成本。
    2. 權證不可當沖，同一天先扣舊庫存賣出，再加入當日買進。
    3. 成本優先使用 FIFO；若歷史快取缺少足夠庫存，才用賣出前所有歷史買進均價估算不足部分。
    4. 若完全沒有歷史買進成本，報酬率維持 '-'，並標示「缺歷史買進成本」。
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty or "日期" not in df.columns:
        return {}

    df2 = df.copy()
    df2["日期"] = df2["日期"].map(normalize_date_str)
    df2["dt_parsed"] = df2["日期"].map(parse_date)
    df2 = df2.dropna(subset=["dt_parsed"]).sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []
    out = {}
    historical_buy_qty = 0.0
    historical_buy_amount = 0.0

    for row in df2.itertuples(index=False):
        row_dict = row._asdict()
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0), 0.0)
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0), 0.0)
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0), 0.0)
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0), 0.0)

        # 權證不能當沖：同一天賣出先扣舊庫存，不吃同日買進。
        if sell_amount > 0:
            hist_avg_cost = historical_buy_amount / historical_buy_qty if historical_buy_qty > 0 and historical_buy_amount > 0 else None
            sell_price = sell_amount / sell_qty if sell_qty > 0 else None

            matched_qty = 0.0
            matched_cost = 0.0
            estimated_qty = 0.0
            estimated_cost = 0.0
            unmatched_qty = 0.0
            unmatched_amount = 0.0

            if sell_qty > 0 and sell_price is not None:
                sell_left = sell_qty

                for lot in lots:
                    if sell_left <= 0:
                        break

                    remain_qty = top15_safe_float(lot.get("剩餘股數", 0), 0.0)
                    if remain_qty <= 0:
                        continue

                    alloc_qty = min(sell_left, remain_qty)
                    lot_avg = top15_safe_float(lot.get("均價", 0), 0.0)
                    if alloc_qty <= 0 or lot_avg <= 0:
                        continue

                    lot["剩餘股數"] = remain_qty - alloc_qty
                    sell_left -= alloc_qty
                    matched_qty += alloc_qty
                    matched_cost += alloc_qty * lot_avg

                if sell_left > 0:
                    if hist_avg_cost is not None and hist_avg_cost > 0:
                        estimated_qty = sell_left
                        estimated_cost = sell_left * hist_avg_cost
                    else:
                        unmatched_qty = sell_left
                        unmatched_amount = sell_left * sell_price

                total_cost = matched_cost + estimated_cost
                pnl = sell_amount - total_cost if total_cost > 0 else None
                return_pct = pnl / total_cost * 100 if total_cost > 0 else None

                if total_cost > 0:
                    if estimated_qty > 0 and matched_qty > 0:
                        cost_status = "部分FIFO＋歷史均價估"
                    elif estimated_qty > 0:
                        cost_status = "歷史均價估"
                    elif unmatched_qty > 0:
                        cost_status = "部分成本不足"
                    else:
                        cost_status = "FIFO"
                else:
                    cost_status = "缺歷史買進成本"
            else:
                total_cost = None
                pnl = None
                return_pct = None
                cost_status = "賣出股數異常"

            out[date_str] = {
                "歷史買進股數": round(historical_buy_qty, 0),
                "歷史買進金額": round(historical_buy_amount, 0),
                "歷史平均成本": round(hist_avg_cost, 4) if hist_avg_cost is not None else "",
                "FIFO對應股數": round(matched_qty, 0),
                "估算對應股數": round(estimated_qty, 0),
                "成本不足股數": round(unmatched_qty, 0),
                "成本不足金額": round(unmatched_amount, 0),
                "賣出成本": round(total_cost, 0) if total_cost is not None else "",
                "賣出損益": round(pnl, 0) if pnl is not None else "",
                "報酬率": return_pct,
                "報酬率文字": _fmt_daily_sell_return_pct(return_pct),
                "成本狀態": cost_status,
            }

        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
            })
            historical_buy_qty += buy_qty
            historical_buy_amount += buy_amount

    return out

def write_daily_sell_detail_sheet(wb, items, a_events, b_events, c_events, d_events, e_events=None):
    """
    新增「每日賣出明細」工作表。

    速度修正：
    只輸出最近 SELL_DETAIL_DAYS 天的賣出明細，避免每次把 250 天全歷史賣出資料整張寫入 Google Sheet。
    原始歷史快取與 A/B/C/D/E、近兩月排行、券商查詢仍照舊使用完整 items，不受影響。

    重要修正：
    1. 今日賣超圖片不應再用 A/B/C/D/E 工作表的「減碼日 / 出清日」去反推賣出金額。
       A 類如果只賣 1 張，不能用「減碼均價 × 原始買進張數」估算，否則會把金額放大。
    2. 本表直接使用原始分點歷史資料 item["df"] 的「賣出股數 / 賣出金額」，
       因此會與官方分點資料一致。
    3. B / C / D 事件中的其中一檔權證若發生賣出，也會被列出，不會被群組事件的
       第一個減碼日 / 出清日限制而漏掉。
    4. 賣出報酬率不再只依賴 A/B/C/D/E 事件；未歸類權證會回到同分點 + 同權證歷史買進資料，
       用 FIFO 成本估算賣出損益與報酬率。
    """
    ws = wb.create_sheet("每日賣出明細")

    headers = [
        "日期",
        "分點",
        "分點名稱",
        "券商代號",
        "事件",
        "狀態",
        "標的股",
        "標的名稱",
        "權證代號",
        "權證名稱",
        "賣出張數",
        "賣出股數",
        "賣出金額",
        "賣出均價",
        "報酬率",
        "賣出成本",
        "賣出損益",
        "歷史買進張數",
        "歷史買進股數",
        "歷史買進金額",
        "歷史平均成本",
        "成本狀態",
        "事件日",
        "事件來源",
    ]
    ws.append(headers)

    source_map = build_event_warrant_source_map(a_events, b_events, c_events, d_events, e_events)
    rows = []
    cutoff_dt = datetime.today() - timedelta(days=max(SELL_DETAIL_DAYS - 1, 0))
    cutoff_dt = datetime(cutoff_dt.year, cutoff_dt.month, cutoff_dt.day)

    for item in items:
        df = item.get("df", pd.DataFrame())

        if df is None or df.empty:
            continue

        position = 0
        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2.sort_values("日期").reset_index(drop=True)
        sell_return_map = _daily_sell_fifo_return_map_for_item(item)

        for row in df2.itertuples(index=False):
            row_dict = row._asdict()
            date = normalize_date_str(row_dict.get("日期", ""))
            trade_dt = parse_date(date)
            buy_s = int(row_dict.get("買進股數", 0) or 0)
            sell_s = int(row_dict.get("賣出股數", 0) or 0)
            sell_a = int(row_dict.get("賣出金額", 0) or 0)

            before_position = position
            within_sell_detail_range = trade_dt is not None and trade_dt >= cutoff_dt

            if within_sell_detail_range and sell_s > 0 and sell_a > 0:
                if before_position > 0 and sell_s >= before_position:
                    status = "出清"
                elif before_position > 0:
                    status = "減碼"
                else:
                    status = "賣超"

                warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
                source = source_map.get((str(item.get("broker_label", "")).strip(), warrant_code), {})

                # 事件已在本筆賣出日之前出清，代表本筆屬於重新買進後的部位，不掛舊事件。
                # 使用 < 保留出清日當天的賣出仍可對應原事件。
                src_exit = parse_date(source.get("出清日", ""))
                if src_exit and trade_dt and src_exit < trade_dt:
                    source = {}

                sell_avg = round(sell_a / sell_s, 4) if sell_s > 0 else ""
                sell_return = sell_return_map.get(date, {})
                hist_buy_qty = top15_safe_float(sell_return.get("歷史買進股數", 0), 0.0)

                rows.append({
                    "日期": date,
                    "分點": item.get("broker_label", ""),
                    "分點名稱": item.get("broker_name", ""),
                    "券商代號": item.get("broker_code", ""),
                    "事件": source.get("事件", "未歸類"),
                    "狀態": status,
                    "標的股": item.get("underlying_code", ""),
                    "標的名稱": item.get("underlying_name", ""),
                    "權證代號": item.get("warrant_code", ""),
                    "權證名稱": item.get("warrant_name", ""),
                    "賣出張數": sell_s // 1000,
                    "賣出股數": sell_s,
                    "賣出金額": sell_a,
                    "賣出均價": sell_avg,
                    "報酬率": sell_return.get("報酬率文字", "-"),
                    "賣出成本": sell_return.get("賣出成本", ""),
                    "賣出損益": sell_return.get("賣出損益", ""),
                    "歷史買進張數": int(hist_buy_qty // 1000) if hist_buy_qty > 0 else 0,
                    "歷史買進股數": sell_return.get("歷史買進股數", 0),
                    "歷史買進金額": sell_return.get("歷史買進金額", 0),
                    "歷史平均成本": sell_return.get("歷史平均成本", ""),
                    "成本狀態": sell_return.get("成本狀態", "缺歷史買進成本"),
                    "事件日": source.get("事件日", ""),
                    "事件來源": source.get("事件來源", ""),
                })

            # 權證不可當沖：同日賣出先扣舊庫存，再把當日買進加入庫存。
            if sell_s > 0:
                position = max(position - sell_s, 0)

            if buy_s > 0:
                position += buy_s

    rows = sorted(
        rows,
        key=lambda r: (
            -((parse_date(r.get("日期")) or datetime.min).toordinal()),
            str(r.get("分點", "")),
            str(r.get("標的股", "")),
            -int(r.get("賣出金額", 0) or 0),
        )
    )

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    col_widths = [12, 14, 18, 12, 8, 8, 10, 12, 12, 24, 10, 12, 16, 10, 10, 14, 14, 12, 14, 16, 12, 20, 12, 40]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    fill_map = {
        "減碼": BLUE,
        "出清": ORANGE,
        "賣超": GREEN,
    }

    for row in ws.iter_rows(min_row=2):
        status = str(row[5].value or "").strip()
        row_fill = fill_map.get(status, WHITE)

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 28

    ws.freeze_panes = "A2"



# ══════════════════════════════════════════════════════════════════════
# 近 10 日分點買賣明細快取：單一分點 + 標的股層級
# ══════════════════════════════════════════════════════════════════════

def _fmt_pct_text(value, signed=True):
    if value is None:
        return "-"
    try:
        v = float(value)
    except Exception:
        return "-"
    return f"{v:+.2f}%" if signed else f"{v:.2f}%"


def _near10_window_dates(target_date=None, window_days=None):
    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date)

    if not target_dt:
        target_dt = datetime.today()
        target_date = target_dt.strftime("%Y/%m/%d")

    if window_days is None:
        window_days = BROKER_10D_DETAIL_DAYS

    window_days = max(int(window_days), 1)
    start_dt = target_dt - timedelta(days=window_days - 1)
    start_date = start_dt.strftime("%Y/%m/%d")
    period_text = f"{start_date} ～ {target_date}"
    return target_date, target_dt, start_date, start_dt, window_days, period_text


def _collect_recent_underlying_codes_for_10d(items, target_date=None):
    """
    收集近10日分點買賣明細會用到的標的股代號。

    圖卡新增「現股10日」後，不能只補權證最新價；
    若分點資料仍停在前一交易日，但盤後現股價格已更新，這裡會先把近10日有買賣的標的股最新收盤價補進快取。
    """
    target_date, target_dt, start_date, start_dt, window_days, period_text = _near10_window_dates(target_date)
    codes = set()

    for item in items or []:
        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            continue

        underlying_code = normalize_underlying_code_for_group(item.get("underlying_code", "")) or normalize_price_code(item.get("underlying_code", ""))
        if not underlying_code:
            continue

        has_recent_activity = False
        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            date_str = normalize_date_str(row_dict.get("日期", ""))
            dt = parse_date(date_str)

            if not dt or dt < start_dt or dt > target_dt:
                continue

            buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
            sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))
            buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
            sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))

            if buy_amount > 0 or sell_amount > 0 or buy_qty > 0 or sell_qty > 0:
                has_recent_activity = True
                break

        if has_recent_activity:
            codes.add(underlying_code)

    return codes


def ensure_broker_10d_underlying_prices(price_cache, items, target_date=None):
    """
    近10日分點明細的「現股10日」需要標的股起始價與最新收盤價。

    當 API4 分點資料尚未更新到今天，但今天已盤後時，仍應先把標的股最新價格補進 price_cache，
    讓之後分點資料一出來，圖卡可以直接使用今天的現股收盤價，不會停在前一交易日。
    """
    if not BROKER_10D_DETAIL_ENABLED:
        return price_cache

    codes = _collect_recent_underlying_codes_for_10d(items, target_date)
    if not codes:
        print("  ✅ 近10日分點明細沒有需要預抓的標的股價格。")
        return price_cache

    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date) or datetime.today()
    end_dt = min(target_dt, datetime.today())
    lookback_days = max(int(BROKER_10D_UNDERLYING_PRICE_LOOKBACK_DAYS), 1)
    start_dt = target_dt - timedelta(days=lookback_days)

    persistent_price_cache = load_price_cache()
    fetch_plan = {}

    for code in sorted(codes):
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        in_memory_prices = get_price_series_from_cache(price_cache, code)
        merged_cached = merge_price_dicts(cached_prices, in_memory_prices)

        if merged_cached:
            add_price_aliases(price_cache, code, merged_cached)

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_date)
        latest_dt = parse_date(latest_date) if latest_date else None

        if latest_price is None or latest_dt is None or latest_dt.date() < target_dt.date():
            fetch_plan[code] = (start_dt, end_dt)

    print(f"【Step 4c】近10日分點明細需檢查標的股價格：{len(codes):,} 檔")
    print(f"  近10日分點明細需補抓標的股價格：{len(fetch_plan):,} 檔")

    if not fetch_plan:
        return price_cache

    def fetch_one(code):
        sdt, edt = fetch_plan[code]
        return code, fetch_twse_prices(code, sdt, edt)

    done = 0
    changed_price_codes = set()

    with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as ex:
        futures = {ex.submit(fetch_one, code): code for code in fetch_plan}

        for future in as_completed(futures):
            done += 1
            code = futures[future]

            try:
                _, fetched_prices = future.result()
            except Exception:
                fetched_prices = {}

            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            merged_prices = merge_price_dicts(old_prices, fetched_prices)

            norm_code = normalize_price_code(code)
            if norm_code:
                persistent_price_cache[norm_code] = merged_prices
                if fetched_prices:
                    changed_price_codes.add(norm_code)

            add_price_aliases(price_cache, code, merged_prices)

            if done % 20 == 0:
                print(f"  [{done}/{len(fetch_plan)}] 近10日標的股價格補抓中...")

    save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)
    return price_cache


def _sell_needs_latest_price_fallback_for_item(item, start_dt, target_dt):
    """
    判斷近10日賣超是否真的需要最新權證價格作為成本備援。

    若 API5 歷史中已有賣出前的買進成本，賣超報酬可用 FIFO / 歷史均價估算，
    不需要再補抓最新權證價格；只有在賣出時完全沒有歷史買進成本可參考時，
    才補抓最新價作為備援，避免報酬率空白。
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty or "日期" not in df.columns:
        return False

    df = df.copy()
    df["dt_parsed"] = df["日期"].map(parse_date)
    df = df.dropna(subset=["dt_parsed"]).sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    historical_buy_qty = 0.0
    historical_buy_amount = 0.0

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        dt = row_dict.get("dt_parsed")
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

        in_window = bool(dt and start_dt <= dt <= target_dt)

        # 權證不可當沖：同一天先賣舊庫存，因此這裡要在加入當日買進前先判斷賣出。
        if in_window and (sell_qty > 0 or sell_amount > 0):
            if historical_buy_qty <= 0 or historical_buy_amount <= 0:
                return True

        if buy_qty > 0 and buy_amount > 0:
            historical_buy_qty += buy_qty
            historical_buy_amount += buy_amount

    return False


def _collect_recent_warrant_codes_for_10d(items, target_date=None):
    target_date, target_dt, start_date, start_dt, window_days, period_text = _near10_window_dates(target_date)
    codes = set()
    skipped_no_remaining_position = 0
    skipped_sell_has_cost = 0

    for item in items or []:
        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            continue

        warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
        if not warrant_code:
            continue

        recent_buy_amount = 0.0
        recent_sell_amount = 0.0
        recent_buy_qty = 0.0
        recent_sell_qty = 0.0

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            date_str = normalize_date_str(row_dict.get("日期", ""))
            dt = parse_date(date_str)

            if not dt or dt < start_dt or dt > target_dt:
                continue

            buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
            sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))
            buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
            sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))

            recent_buy_amount += buy_amount
            recent_sell_amount += sell_amount
            recent_buy_qty += buy_qty
            recent_sell_qty += sell_qty

        if recent_buy_amount <= 0 and recent_sell_amount <= 0 and recent_buy_qty <= 0 and recent_sell_qty <= 0:
            continue

        if BROKER_10D_FETCH_ALL_TRADED_WARRANT_PRICES:
            codes.add(warrant_code)
            continue

        # 買超報酬需要「近10日買進後仍未賣掉的部位」目前市值，這類權證一定要補最新價。
        if recent_buy_amount > 0 or recent_buy_qty > 0:
            position_summary = _recent_buy_position_summary_for_item(
                item,
                start_dt,
                target_dt,
                latest_price=None,
            )
            remaining_cost = top15_safe_float(position_summary.get("remaining_cost", 0.0))
            if remaining_cost > 0:
                codes.add(warrant_code)
            else:
                skipped_no_remaining_position += 1
            continue

        # 純賣超通常可用 API5 歷史 FIFO 成本估報酬，不需要市場最新價。
        # 只有完全找不到賣出前買進成本時，才補抓最新價作為備援。
        if recent_sell_amount > 0 or recent_sell_qty > 0:
            if BROKER_10D_FETCH_SELL_FALLBACK_PRICES and _sell_needs_latest_price_fallback_for_item(item, start_dt, target_dt):
                codes.add(warrant_code)
            else:
                skipped_sell_has_cost += 1

    print(
        f"  近10日分點明細價格篩選：需最新價 {len(codes):,} 檔｜"
        f"已排除無剩餘買超部位 {skipped_no_remaining_position:,} 檔｜"
        f"已排除可用歷史成本計算的純賣超 {skipped_sell_has_cost:,} 檔"
    )

    return codes

def ensure_broker_10d_warrant_prices(price_cache, items, target_date=None):
    """
    近10日分點買賣明細需要用「最新權證價格」估算買超部位放到現在的報酬。

    加速修正：
    1. 不再對近10日所有有買賣的權證一律抓價。
       - 近10日買進後仍有剩餘部位：一定補最新價。
       - 純賣超且 API5 歷史已有成本：不補最新價，直接用 FIFO / 歷史成本算報酬。
       - 純賣超但完全沒有歷史成本：才補最新價作為備援。
    2. 價格補抓採兩段式：先抓 BROKER_10D_PRICE_FAST_LOOKBACK_DAYS，完全沒價格才補抓完整 BROKER_10D_PRICE_LOOKBACK_DAYS。
    3. 本機價格快取完整保存；Google Sheet 價格快取可增量 append，避免整張重寫拖慢。
    """
    if not BROKER_10D_DETAIL_ENABLED:
        return price_cache

    codes = _collect_recent_warrant_codes_for_10d(items, target_date)
    if not codes:
        return price_cache

    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date) or datetime.today()
    end_dt = min(target_dt, datetime.today())

    full_lookback_days = max(int(BROKER_10D_PRICE_LOOKBACK_DAYS), 1)
    fast_lookback_days = max(int(BROKER_10D_PRICE_FAST_LOOKBACK_DAYS), 1)
    fast_lookback_days = min(fast_lookback_days, full_lookback_days)
    stale_days = max(int(BROKER_10D_PRICE_STALE_DAYS), 0)

    fast_start_dt = target_dt - timedelta(days=fast_lookback_days)
    full_start_dt = target_dt - timedelta(days=full_lookback_days)

    persistent_price_cache = load_price_cache()
    fetch_plan = {}

    for code in sorted(codes):
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        in_memory_prices = get_price_series_from_cache(price_cache, code)
        merged_cached = merge_price_dicts(cached_prices, in_memory_prices)

        if merged_cached:
            add_price_aliases(price_cache, code, merged_cached)

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_date)
        latest_dt = parse_date(latest_date) if latest_date else None

        if latest_price is None or latest_dt is None or (target_dt - latest_dt).days > stale_days:
            fetch_plan[code] = (fast_start_dt, end_dt, full_start_dt)

    print(f"【Step 4d】近10日分點明細需檢查權證價格：{len(codes):,} 檔")
    print(f"  近10日分點明細需補抓權證價格：{len(fetch_plan):,} 檔")
    print(f"  近10日分點明細價格補抓策略：先抓近 {fast_lookback_days} 天，完全無價格才補抓近 {full_lookback_days} 天")

    if not fetch_plan:
        return price_cache

    def fetch_one(code):
        fast_sdt, edt, full_sdt = fetch_plan[code]
        fetched_prices = fetch_twse_prices(code, fast_sdt, edt)

        # 若短區間完全沒有價格，而且本地快取也沒有任何可用價格，再補抓完整區間。
        old_prices = get_cached_prices_for_code(persistent_price_cache, code)
        merged_after_fast = merge_price_dicts(old_prices, fetched_prices)
        has_any_price = any(
            p is not None and p > 0
            for p in merged_after_fast.values()
        )

        if not has_any_price and full_sdt < fast_sdt:
            full_prices = fetch_twse_prices(code, full_sdt, edt)
            fetched_prices = merge_price_dicts(fetched_prices, full_prices)

        return code, fetched_prices

    done = 0
    changed_price_codes = set()
    with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as ex:
        futures = {ex.submit(fetch_one, code): code for code in fetch_plan}

        for future in as_completed(futures):
            done += 1
            code = futures[future]

            try:
                _, fetched_prices = future.result()
            except Exception:
                fetched_prices = {}

            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            merged_prices = merge_price_dicts(old_prices, fetched_prices)

            if merged_prices:
                norm_code = normalize_price_code(code)
                if norm_code:
                    persistent_price_cache[norm_code] = merged_prices
                    if fetched_prices:
                        changed_price_codes.add(norm_code)
                add_price_aliases(price_cache, code, merged_prices)

            if done % 20 == 0:
                print(f"  [{done}/{len(fetch_plan)}] 近10日分點明細權證價格補抓中...")

    save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)
    return price_cache

def _sell_return_summary_for_item(item, start_dt, target_dt, fallback_price=None):
    """
    用同一分點 + 同一權證的 API5 歷史資料估算近10日賣出實現報酬。

    規則：
    - 權證不可當沖：同一天先賣舊庫存，再把當日買進加入庫存。
    - 賣出成本優先用 FIFO 持倉成本估算。
    - 若近10日賣出找不到足夠舊庫存成本，改用可取得的成本備援估算，避免賣超報酬率空白：
      1. 優先用該權證歷史已出現買進均價
      2. 其次用最新權證價格
      3. 最後用當筆賣出均價，讓報酬率保守落在 0%
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty:
        return {
            "revenue": 0.0,
            "cost": 0.0,
            "unmatched_amount": 0.0,
            "unmatched_qty": 0.0,
        }

    df = df.copy()
    if "日期" not in df.columns:
        return {
            "revenue": 0.0,
            "cost": 0.0,
            "unmatched_amount": 0.0,
            "unmatched_qty": 0.0,
        }

    df["dt_parsed"] = df["日期"].map(parse_date)
    df = df.dropna(subset=["dt_parsed"]).sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []
    revenue = 0.0
    cost = 0.0
    unmatched_amount = 0.0
    unmatched_qty = 0.0

    historical_buy_amount = 0.0
    historical_buy_qty = 0.0

    try:
        fallback_price = float(fallback_price) if fallback_price is not None else None
    except Exception:
        fallback_price = None

    if fallback_price is not None and fallback_price <= 0:
        fallback_price = None

    def fallback_unit_cost(sell_price):
        if historical_buy_qty > 0 and historical_buy_amount > 0:
            return historical_buy_amount / historical_buy_qty
        if fallback_price is not None and fallback_price > 0:
            return fallback_price
        if sell_price and sell_price > 0:
            return sell_price
        return None

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        dt = row_dict.get("dt_parsed")
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

        in_window = bool(dt and start_dt <= dt <= target_dt)

        # 權證不能當沖：同一天先處理賣出，只能扣舊庫存。
        if sell_amount > 0:
            if sell_qty <= 0:
                # API 異常時仍保留數據，不讓近10日賣超報酬率變空白。
                if in_window:
                    revenue += sell_amount
                    cost += sell_amount
                sell_qty = 0
            else:
                sell_price = sell_amount / sell_qty
                sell_left = sell_qty
                allocated_revenue = 0.0
                allocated_cost = 0.0

                for lot in lots:
                    if sell_left <= 0:
                        break
                    if lot.get("剩餘股數", 0) <= 0:
                        continue

                    alloc = min(sell_left, lot["剩餘股數"])
                    if alloc <= 0:
                        continue

                    lot["剩餘股數"] -= alloc
                    sell_left -= alloc
                    allocated_revenue += alloc * sell_price
                    allocated_cost += alloc * lot["均價"]

                if sell_left > 0:
                    fallback_cost_price = fallback_unit_cost(sell_price)
                    if fallback_cost_price is not None and fallback_cost_price > 0:
                        allocated_revenue += sell_left * sell_price
                        allocated_cost += sell_left * fallback_cost_price
                    else:
                        unmatched_qty += sell_left
                        unmatched_amount += sell_left * sell_price

                if in_window:
                    revenue += allocated_revenue
                    cost += allocated_cost

        # 同日買進放到賣出後面，避免當沖錯扣。
        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
            })
            historical_buy_qty += buy_qty
            historical_buy_amount += buy_amount

    return {
        "revenue": revenue,
        "cost": cost,
        "unmatched_amount": unmatched_amount,
        "unmatched_qty": unmatched_qty,
    }


def _recent_buy_position_summary_for_item(item, start_dt, target_dt, latest_price=None):
    """
    計算近10日買進 lot 在統計日仍然留下的真實 FIFO 持倉。

    這個函式用完整歷史跑到統計日：
    - 每天先賣出扣 FIFO 舊庫存，再加入當天買進。
    - 只把「買進日落在近10日視窗內」且統計日尚未被賣掉的剩餘 lot 納入買超持倉報酬。
    - 買超報酬 = (有最新價格的剩餘股數 × 最新權證價格 - 有最新價格的剩餘成本) / 有最新價格的剩餘成本。
    - 缺最新權證價格的剩餘部位不納入報酬率，改在明細備註統計省略檔數。
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty:
        return {
            "remaining_qty": 0.0,
            "remaining_cost": 0.0,
            "market_value": 0.0,
            "missing_price": False,
        }

    df = df.copy()
    if "日期" not in df.columns:
        return {
            "remaining_qty": 0.0,
            "remaining_cost": 0.0,
            "market_value": 0.0,
            "missing_price": False,
        }

    df["dt_parsed"] = df["日期"].map(parse_date)
    df = df.dropna(subset=["dt_parsed"])
    df = df[df["dt_parsed"] <= target_dt].sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        dt = row_dict.get("dt_parsed")
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

        # 權證不能當沖：同一天先賣出，只扣舊庫存。
        if sell_qty > 0 and sell_amount > 0:
            sell_left = sell_qty
            for lot in lots:
                if sell_left <= 0:
                    break
                if lot.get("剩餘股數", 0) <= 0:
                    continue

                alloc = min(sell_left, lot["剩餘股數"])
                if alloc <= 0:
                    continue

                lot["剩餘股數"] -= alloc
                sell_left -= alloc

        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
                "近10日買進": bool(dt and start_dt <= dt <= target_dt),
            })

    remaining_qty = 0.0
    remaining_cost = 0.0

    for lot in lots:
        if not lot.get("近10日買進"):
            continue
        qty = top15_safe_float(lot.get("剩餘股數", 0))
        avg = top15_safe_float(lot.get("均價", 0))
        if qty <= 0 or avg <= 0:
            continue
        remaining_qty += qty
        remaining_cost += qty * avg

    try:
        latest_price = float(latest_price) if latest_price is not None else None
    except Exception:
        latest_price = None

    if latest_price is not None and latest_price > 0 and remaining_qty > 0:
        market_value = remaining_qty * latest_price
        missing_price = False
    else:
        market_value = 0.0
        missing_price = remaining_cost > 0

    return {
        "remaining_qty": remaining_qty,
        "remaining_cost": remaining_cost,
        "market_value": market_value,
        "missing_price": missing_price,
    }


def build_10d_broker_underlying_detail_rows(items, price_cache, target_date=None):
    """
    建立「快取_近10日分點買賣明細」。

    統計單位：資料範圍 + 統計日期 + 單一分點 + 標的股。
    同一分點同一標的底下的所有權證會先完整合併，再計算買賣金額、買超 / 賣超報酬與勝率。
    """
    if not BROKER_10D_DETAIL_ENABLED:
        return None

    print("【Step 4e】建立近10日分點買賣明細快取...")

    if not items:
        print("  ⚠️ 近10日分點買賣明細：沒有 items 資料")
        return []

    target_date, target_dt, start_date, start_dt, window_days, period_text = _near10_window_dates(target_date)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = get_result_data_scope()

    agg = {}

    for item in items:
        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            continue

        warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
        if not warrant_code:
            continue

        warrant_name = str(item.get("warrant_name", "")).strip()
        underlying_code = str(item.get("underlying_code", "")).strip()
        underlying_name = str(item.get("underlying_name", "")).strip()
        broker_label = str(item.get("broker_label", "")).strip()
        broker_name = str(item.get("broker_name", "")).strip()
        broker_code = str(item.get("broker_code", "")).strip()

        if not underlying_code or not broker_label:
            continue

        key = (broker_label, broker_name, broker_code, underlying_code)
        rec = agg.setdefault(key, {
            "資料範圍": scope,
            "統計日期": add_gsheet_text_prefix(target_date),
            "統計期間": period_text,
            "統計天數": window_days,
            "第一筆日期": "",
            "最後筆日期": "",
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "近10日買進股數": 0.0,
            "近10日買進金額": 0.0,
            "近10日賣出股數": 0.0,
            "近10日賣出金額": 0.0,
            "日期集合": set(),
            "權證": {},
            "買超報酬加權分子": 0.0,
            "買超報酬權重": 0.0,
            "買超剩餘股數": 0.0,
            "買超剩餘成本": 0.0,
            "買超目前市值": 0.0,
            "買超報酬有效權證數": 0,
            "買超報酬缺價權證數": 0,
            "買超缺價省略權證清單": [],
            "賣超實現賣出金額": 0.0,
            "賣超實現成本": 0.0,
            "賣超成本不足金額": 0.0,
            "賣超成本不足股數": 0.0,
            "更新時間": update_time,
            "run_id": run_id,
        })

        if underlying_name and not rec.get("標的名稱"):
            rec["標的名稱"] = underlying_name

        item_buy_qty = 0.0
        item_buy_amount = 0.0
        item_sell_qty = 0.0
        item_sell_amount = 0.0
        item_dates = set()

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            date_str = normalize_date_str(row_dict.get("日期", ""))
            dt = parse_date(date_str)

            if not dt or dt < start_dt or dt > target_dt:
                continue

            buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
            sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
            buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
            sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

            if buy_qty <= 0 and sell_qty <= 0 and buy_amount <= 0 and sell_amount <= 0:
                continue

            rec["近10日買進股數"] += buy_qty
            rec["近10日買進金額"] += buy_amount
            rec["近10日賣出股數"] += sell_qty
            rec["近10日賣出金額"] += sell_amount
            rec["日期集合"].add(date_str)

            item_buy_qty += buy_qty
            item_buy_amount += buy_amount
            item_sell_qty += sell_qty
            item_sell_amount += sell_amount
            item_dates.add(date_str)

        if item_buy_amount <= 0 and item_sell_amount <= 0:
            continue

        warrant_rec = rec["權證"].setdefault(warrant_code, {
            "權證代號": warrant_code,
            "權證名稱": warrant_name,
            "買進金額": 0.0,
            "賣出金額": 0.0,
            "買進股數": 0.0,
            "賣出股數": 0.0,
            "日期集合": set(),
            "最新權證價格": None,
            "最新權證價格日": "",
        })
        warrant_rec["買進金額"] += item_buy_amount
        warrant_rec["賣出金額"] += item_sell_amount
        warrant_rec["買進股數"] += item_buy_qty
        warrant_rec["賣出股數"] += item_sell_qty
        warrant_rec["日期集合"].update(item_dates)

        latest_price, latest_price_date = get_latest_price_info_on_or_before(price_cache, warrant_code, target_date)
        if latest_price is not None:
            warrant_rec["最新權證價格"] = latest_price
            warrant_rec["最新權證價格日"] = latest_price_date

        if item_buy_qty > 0 and item_buy_amount > 0:
            position_summary = _recent_buy_position_summary_for_item(
                item,
                start_dt,
                target_dt,
                latest_price=latest_price,
            )
            remaining_cost = top15_safe_float(position_summary.get("remaining_cost", 0.0))
            remaining_qty = top15_safe_float(position_summary.get("remaining_qty", 0.0))
            market_value = top15_safe_float(position_summary.get("market_value", 0.0))

            if remaining_cost > 0:
                if position_summary.get("missing_price"):
                    # 缺最新權證價格的部位直接從買超平均報酬分子 / 分母排除，
                    # 避免用 0 市值把整體報酬率不合理壓低。
                    rec["買超報酬缺價權證數"] += 1
                    rec.setdefault("買超缺價省略權證清單", []).append(f"{warrant_code} {warrant_name}".strip())
                    warrant_rec["買超報酬省略原因"] = "缺最新權證價格，未納入買超平均報酬"
                else:
                    rec["買超剩餘成本"] += remaining_cost
                    rec["買超剩餘股數"] += remaining_qty
                    rec["買超目前市值"] += market_value
                    rec["買超報酬權重"] += remaining_cost
                    rec["買超報酬加權分子"] += market_value - remaining_cost
                    rec["買超報酬有效權證數"] += 1

            warrant_rec["近10日買進剩餘股數"] = remaining_qty
            warrant_rec["近10日買進剩餘成本"] = remaining_cost
            warrant_rec["近10日買進目前市值"] = market_value

        if item_sell_qty > 0 and item_sell_amount > 0:
            sell_summary = _sell_return_summary_for_item(
                item,
                start_dt,
                target_dt,
                fallback_price=latest_price,
            )
            rec["賣超實現賣出金額"] += sell_summary.get("revenue", 0.0)
            rec["賣超實現成本"] += sell_summary.get("cost", 0.0)
            rec["賣超成本不足金額"] += sell_summary.get("unmatched_amount", 0.0)
            rec["賣超成本不足股數"] += sell_summary.get("unmatched_qty", 0.0)

    rows = []

    for rec in agg.values():
        dates = sorted(rec.get("日期集合", set()))
        if not dates:
            continue

        buy_amount = float(rec.get("近10日買進金額", 0) or 0)
        sell_amount = float(rec.get("近10日賣出金額", 0) or 0)
        buy_qty = float(rec.get("近10日買進股數", 0) or 0)
        sell_qty = float(rec.get("近10日賣出股數", 0) or 0)
        net_buy_amount = buy_amount - sell_amount
        net_sell_amount = sell_amount - buy_amount
        net_buy_qty = buy_qty - sell_qty
        net_sell_qty = sell_qty - buy_qty

        buy_return_pct = None
        buy_return_weight = float(rec.get("買超報酬權重", 0) or 0)
        buy_return_numerator = float(rec.get("買超報酬加權分子", 0) or 0)

        if buy_return_weight > 0:
            # 買超平均報酬只使用「有最新價格」的剩餘部位；缺價部位已在上方省略。
            buy_return_pct = buy_return_numerator / buy_return_weight * 100
        elif buy_amount > 0:
            # 近10日有買進但沒有任何可估剩餘部位時，仍給出 0.00%，避免勝率報酬空白。
            buy_return_pct = 0.0

        sell_return_pct = None
        sell_realized_cost = float(rec.get("賣超實現成本", 0) or 0)
        sell_realized_revenue = float(rec.get("賣超實現賣出金額", 0) or 0)

        if sell_realized_cost > 0:
            sell_return_pct = (sell_realized_revenue - sell_realized_cost) / sell_realized_cost * 100
        elif sell_amount > 0:
            # API 賣出資料異常或成本仍不足時，保守帶入 0.00%，確保賣超報酬率一定有值。
            sell_return_pct = 0.0

        if net_buy_amount > 0:
            direction = "買超"
            win_return_pct = buy_return_pct
        elif net_sell_amount > 0:
            direction = "賣超"
            win_return_pct = sell_return_pct
        else:
            direction = "買賣平衡"
            win_return_pct = buy_return_pct if buy_return_pct is not None else sell_return_pct

        if win_return_pct is None:
            win_return_pct = 0.0

        # 分點層級加權報酬使用「主要方向淨額」作為權重：
        # - 買超標的：使用近10日淨買超金額
        # - 賣超標的：使用近10日淨賣超金額
        # - 買賣平衡：退回使用買進 / 賣出金額較大者
        # 這樣可以避免單純平均讓小金額標的過度影響整體分點表現。
        primary_return_weight = net_buy_amount if direction == "買超" else net_sell_amount if direction == "賣超" else max(buy_amount, sell_amount)
        if primary_return_weight <= 0:
            primary_return_weight = max(buy_amount, sell_amount, 0.0)

        broker_buy_return_weight = net_buy_amount if net_buy_amount > 0 else 0.0
        broker_sell_return_weight = net_sell_amount if net_sell_amount > 0 else 0.0

        if win_return_pct > 0:
            result = "勝"
        else:
            # 使用者指定 0% 要算賠錢，所以 <= 0 都是敗。
            result = "敗"

        notes = []
        missing_price_count = int(rec.get("買超報酬缺價權證數", 0) or 0)

        if missing_price_count > 0:
            notes.append(f"買超報酬已省略缺價權證 {missing_price_count} 檔")

        if buy_amount > 0 and buy_return_weight <= 0:
            if missing_price_count > 0:
                notes.append("買超剩餘部位全數缺價或無可估價格，買超報酬以 0.00% 保守帶入")
            else:
                notes.append("近10日買進目前無可估剩餘部位，買超報酬以 0.00% 帶入")

        if sell_amount > 0 and sell_realized_cost <= 0:
            notes.append("賣超成本不足，賣超報酬以 0.00% 帶入")

        note_text = "；".join(notes)

        underlying_10d_return_pct = None
        underlying_start_price = None
        underlying_end_price = None
        underlying_start_price_date = ""
        underlying_end_price_date = ""
        underlying_code_for_price = str(rec.get("標的股", "") or "").strip()

        if underlying_code_for_price:
            underlying_start_price, underlying_start_price_date = get_latest_price_info_on_or_before(
                price_cache,
                underlying_code_for_price,
                start_date,
            )
            underlying_end_price, underlying_end_price_date = get_latest_price_info_on_or_before(
                price_cache,
                underlying_code_for_price,
                target_date,
            )

            try:
                if (
                    underlying_start_price is not None
                    and underlying_end_price is not None
                    and float(underlying_start_price) > 0
                    and float(underlying_end_price) > 0
                ):
                    underlying_10d_return_pct = (float(underlying_end_price) - float(underlying_start_price)) / float(underlying_start_price) * 100
            except Exception:
                underlying_10d_return_pct = None

        warrant_rows = []
        warrant_json = []
        for w in sorted(
            rec.get("權證", {}).values(),
            key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
            reverse=True,
        ):
            warrant_rows.append(
                f"{w.get('權證代號', '')} {w.get('權證名稱', '')}"
                f"｜買{_fmt_wan_text(w.get('買進金額', 0))}／賣{_fmt_wan_text(w.get('賣出金額', 0))}"
            )
            warrant_json.append({
                "權證代號": w.get("權證代號", ""),
                "權證名稱": w.get("權證名稱", ""),
                "買進金額": round(float(w.get("買進金額", 0) or 0), 0),
                "賣出金額": round(float(w.get("賣出金額", 0) or 0), 0),
                "買進股數": round(float(w.get("買進股數", 0) or 0), 0),
                "賣出股數": round(float(w.get("賣出股數", 0) or 0), 0),
                "最新權證價格": w.get("最新權證價格"),
                "最新權證價格日": w.get("最新權證價格日", ""),
                "近10日買進剩餘股數": round(float(w.get("近10日買進剩餘股數", 0) or 0), 0),
                "近10日買進剩餘成本": round(float(w.get("近10日買進剩餘成本", 0) or 0), 0),
                "近10日買進目前市值": round(float(w.get("近10日買進目前市值", 0) or 0), 0),
                "買超報酬省略原因": w.get("買超報酬省略原因", ""),
                "日期數": len(w.get("日期集合", set())),
            })

        rows.append({
            "資料範圍": rec.get("資料範圍", scope),
            "統計日期": add_gsheet_text_prefix(target_date),
            "統計期間": period_text,
            "統計天數": window_days,
            "有效日期數": len(dates),
            "第一筆日期": add_gsheet_text_prefix(dates[0]),
            "最後筆日期": add_gsheet_text_prefix(dates[-1]),
            "分點": rec.get("分點", ""),
            "分點名稱": rec.get("分點名稱", ""),
            "券商代號": rec.get("券商代號", ""),
            "標的股": rec.get("標的股", ""),
            "標的名稱": rec.get("標的名稱", ""),
            "標的10日漲跌幅%": _fmt_pct_text(underlying_10d_return_pct, signed=True) if underlying_10d_return_pct is not None else "-",
            "現股10日報酬率%": _fmt_pct_text(underlying_10d_return_pct, signed=True) if underlying_10d_return_pct is not None else "-",
            "標的10日起始價": round(float(underlying_start_price), 4) if underlying_start_price is not None else "",
            "標的10日收盤價": round(float(underlying_end_price), 4) if underlying_end_price is not None else "",
            "標的10日起始價格日": add_gsheet_text_prefix(underlying_start_price_date) if underlying_start_price_date else "",
            "標的10日收盤價格日": add_gsheet_text_prefix(underlying_end_price_date) if underlying_end_price_date else "",
            "買賣方向": direction,
            "近10日買進股數": round(buy_qty, 0),
            "近10日買進金額": round(buy_amount, 0),
            "近10日賣出股數": round(sell_qty, 0),
            "近10日賣出金額": round(sell_amount, 0),
            "近10日淨買超股數": round(net_buy_qty, 0),
            "近10日淨買超金額": round(net_buy_amount, 0),
            "近10日淨賣超股數": round(net_sell_qty, 0),
            "近10日淨賣超金額": round(net_sell_amount, 0),
            "涉及權證檔數": len(rec.get("權證", {})),
            "權證清單": "；".join(warrant_rows),
            "買超平均報酬%": _fmt_pct_text(buy_return_pct, signed=True),
            "買超剩餘股數": round(float(rec.get("買超剩餘股數", 0) or 0), 0),
            "買超剩餘成本": round(float(rec.get("買超剩餘成本", 0) or 0), 0),
            "買超目前市值": round(float(rec.get("買超目前市值", 0) or 0), 0),
            "買超報酬有效權證數": int(rec.get("買超報酬有效權證數", 0) or 0),
            "買超報酬缺價權證數": int(rec.get("買超報酬缺價權證數", 0) or 0),
            "賣超平均報酬%": _fmt_pct_text(sell_return_pct, signed=True),
            "賣超實現賣出金額": round(float(rec.get("賣超實現賣出金額", 0) or 0), 0),
            "賣超實現成本": round(float(rec.get("賣超實現成本", 0) or 0), 0),
            "賣超成本不足金額": round(float(rec.get("賣超成本不足金額", 0) or 0), 0),
            "用於勝率報酬%": _fmt_pct_text(win_return_pct, signed=True),
            "判定": result,
            "備註": note_text,
            "分點近10日勝率": "-",
            "分點近10日勝筆數": 0,
            "分點近10日敗筆數": 0,
            "分點近10日加權平均報酬%": "-",
            "分點近10日買超加權平均報酬%": "-",
            "分點近10日賣超加權平均報酬%": "-",
            "分點近10日加權平均勝報酬%": "-",
            "分點近10日加權平均敗報酬%": "-",
            "分點近10日盈虧比": "-",
            "分點近10日加權期望值%": "-",
            "分點近10日加權報酬權重金額": 0,
            "更新時間": rec.get("更新時間", update_time),
            "run_id": rec.get("run_id", run_id),
            "權證明細_JSON": json.dumps(warrant_json, ensure_ascii=False),
            "_分點統計報酬率數值": win_return_pct,
            "_分點統計權重": primary_return_weight,
            "_分點統計買超報酬率數值": buy_return_pct,
            "_分點統計買超權重": broker_buy_return_weight,
            "_分點統計賣超報酬率數值": sell_return_pct,
            "_分點統計賣超權重": broker_sell_return_weight,
        })

    broker_stats = {}
    for row in rows:
        broker_key = (row.get("分點", ""), row.get("券商代號", ""))
        stat = broker_stats.setdefault(broker_key, {
            "win": 0,
            "loss": 0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
        })

        if row.get("判定") == "勝":
            stat["win"] += 1
        elif row.get("判定") == "敗":
            stat["loss"] += 1

        ret = row.get("_分點統計報酬率數值")
        weight = top15_safe_float(row.get("_分點統計權重", 0), 0.0)

        if ret is not None and weight > 0:
            ret = top15_safe_float(ret, None)
            if ret is not None:
                stat["return_weighted_numerator"] += ret * weight
                stat["return_weight"] += weight

                if ret > 0:
                    stat["win_return_weighted_numerator"] += ret * weight
                    stat["win_return_weight"] += weight
                else:
                    stat["loss_return_weighted_numerator"] += ret * weight
                    stat["loss_return_weight"] += weight

        buy_ret = row.get("_分點統計買超報酬率數值")
        buy_weight = top15_safe_float(row.get("_分點統計買超權重", 0), 0.0)
        if buy_ret is not None and buy_weight > 0:
            buy_ret = top15_safe_float(buy_ret, None)
            if buy_ret is not None:
                stat["buy_return_weighted_numerator"] += buy_ret * buy_weight
                stat["buy_return_weight"] += buy_weight

        sell_ret = row.get("_分點統計賣超報酬率數值")
        sell_weight = top15_safe_float(row.get("_分點統計賣超權重", 0), 0.0)
        if sell_ret is not None and sell_weight > 0:
            sell_ret = top15_safe_float(sell_ret, None)
            if sell_ret is not None:
                stat["sell_return_weighted_numerator"] += sell_ret * sell_weight
                stat["sell_return_weight"] += sell_weight

    for row in rows:
        broker_key = (row.get("分點", ""), row.get("券商代號", ""))
        stat = broker_stats.get(broker_key, {
            "win": 0,
            "loss": 0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
        })

        total = stat["win"] + stat["loss"]
        win_rate_value = stat["win"] / total * 100 if total else None

        avg_return = (
            stat["return_weighted_numerator"] / stat["return_weight"]
            if stat["return_weight"] > 0 else None
        )
        avg_buy_return = (
            stat["buy_return_weighted_numerator"] / stat["buy_return_weight"]
            if stat["buy_return_weight"] > 0 else None
        )
        avg_sell_return = (
            stat["sell_return_weighted_numerator"] / stat["sell_return_weight"]
            if stat["sell_return_weight"] > 0 else None
        )
        avg_win_return = (
            stat["win_return_weighted_numerator"] / stat["win_return_weight"]
            if stat["win_return_weight"] > 0 else None
        )
        avg_loss_return = (
            stat["loss_return_weighted_numerator"] / stat["loss_return_weight"]
            if stat["loss_return_weight"] > 0 else None
        )

        profit_loss_ratio = None
        if avg_win_return is not None and avg_win_return > 0 and avg_loss_return is not None and avg_loss_return < 0:
            profit_loss_ratio = avg_win_return / abs(avg_loss_return)

        expectancy = None
        if total > 0 and (avg_win_return is not None or avg_loss_return is not None):
            win_rate_decimal = stat["win"] / total
            loss_rate_decimal = stat["loss"] / total
            expectancy = win_rate_decimal * (avg_win_return or 0.0) + loss_rate_decimal * (avg_loss_return or 0.0)
        elif avg_return is not None:
            expectancy = avg_return

        row["分點近10日勝筆數"] = stat["win"]
        row["分點近10日敗筆數"] = stat["loss"]
        row["分點近10日勝率"] = _fmt_pct_text(win_rate_value, signed=False) if win_rate_value is not None else "-"
        row["分點近10日加權平均報酬%"] = _fmt_pct_text(avg_return, signed=True)
        row["分點近10日買超加權平均報酬%"] = _fmt_pct_text(avg_buy_return, signed=True)
        row["分點近10日賣超加權平均報酬%"] = _fmt_pct_text(avg_sell_return, signed=True)
        row["分點近10日加權平均勝報酬%"] = _fmt_pct_text(avg_win_return, signed=True)
        row["分點近10日加權平均敗報酬%"] = _fmt_pct_text(avg_loss_return, signed=True)
        row["分點近10日盈虧比"] = f"{profit_loss_ratio:.2f}" if profit_loss_ratio is not None else "-"
        row["分點近10日加權期望值%"] = _fmt_pct_text(expectancy, signed=True)
        row["分點近10日加權報酬權重金額"] = round(float(stat.get("return_weight", 0.0) or 0.0), 0)

    rows = sorted(
        rows,
        key=lambda r: (
            str(r.get("分點", "")),
            -abs(float(r.get("近10日淨買超金額", 0) or 0)),
            str(r.get("標的股", "")),
        )
    )

    print(f"  ✅ 近10日分點買賣明細完成：{len(rows):,} 筆，統計期間 {period_text}")
    return rows


def write_10d_broker_underlying_detail_sheet(wb, rows):
    """寫入近10日分點買賣明細。rows=None 代表不建立工作表。"""
    if rows is None:
        return

    ws = wb.create_sheet(BROKER_10D_DETAIL_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期",
        "分點", "分點名稱", "券商代號", "標的股", "標的名稱",
        "標的10日漲跌幅%", "現股10日報酬率%", "標的10日起始價", "標的10日收盤價", "標的10日起始價格日", "標的10日收盤價格日",
        "買賣方向",
        "近10日買進股數", "近10日買進金額", "近10日賣出股數", "近10日賣出金額",
        "近10日淨買超股數", "近10日淨買超金額", "近10日淨賣超股數", "近10日淨賣超金額",
        "涉及權證檔數", "權證清單",
        "買超平均報酬%", "買超剩餘股數", "買超剩餘成本", "買超目前市值", "買超報酬有效權證數", "買超報酬缺價權證數",
        "賣超平均報酬%", "賣超實現賣出金額", "賣超實現成本", "賣超成本不足金額",
        "用於勝率報酬%", "判定", "備註", "分點近10日勝率", "分點近10日勝筆數", "分點近10日敗筆數",
        "分點近10日加權平均報酬%", "分點近10日買超加權平均報酬%", "分點近10日賣超加權平均報酬%",
        "分點近10日加權平均勝報酬%", "分點近10日加權平均敗報酬%", "分點近10日盈虧比",
        "分點近10日加權期望值%", "分點近10日加權報酬權重金額",
        "更新時間", "run_id", "權證明細_JSON",
    ]

    ws.append(headers)

    for row in rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [
        12, 12, 24, 10, 12, 12, 12,
        14, 18, 12, 10, 14,
        16, 16, 14, 14, 14, 14,
        10,
        14, 16, 14, 16,
        16, 18, 16, 18,
        12, 72,
        14, 14, 16, 16, 14, 14,
        14, 16, 16, 16,
        14, 10, 36, 14, 14, 14,
        16, 20, 20, 18, 18, 12, 16, 18,
        20, 18, 90,
    ]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 26

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    direction_col = header_map.get("買賣方向")

    for row in ws.iter_rows(min_row=2):
        direction = str(row[direction_col - 1].value or "").strip() if direction_col else ""
        row_fill = RED if direction == "買超" else GREEN if direction == "賣超" else WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 32

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════
# 近 10 日分點勝率排行（僅 RUN_MODE=2 更新）
# ══════════════════════════════════════════════════════════════════════

def _parse_pct_text_to_float(value):
    try:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s or s == "-":
            return None
        s = s.replace("%", "").replace("+", "").replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def build_10d_broker_winrate_rank_rows(broker_10d_detail_rows):
    """
    建立「快取_近10日分點勝率排行」。

    重要規則：
    1. 只在 RUN_MODE=2 完整分點清單模式執行。
    2. RUN_MODE=1 精選分點模式直接回傳 None，build_excel 不會建立該工作表，
       因此 upload_excel_to_google_sheet() 也不會更新 / 清空 Google Sheet 上原本的同名工作表。
    3. 統計來源沿用「快取_近10日分點買賣明細」已完成的分點 + 標的層級結果，
       不重新改動 A/B/C/D/E、TOP15 或每日賣出明細邏輯。
    4. 勝率排序：勝率高到低，其次統計筆數多到少、加權報酬高到低、總交易金額高到低。
    """
    if not BROKER_10D_WINRATE_RANK_ENABLED:
        return None

    if RUN_MODE != 2:
        print("  ✅ RUN_MODE=1 精選分點模式：略過近10日分點勝率排行工作表，避免動到 Google Sheet 既有資料。")
        return None

    if broker_10d_detail_rows is None:
        return None

    print("【Step 4f】建立近10日分點勝率排行（RUN_MODE=2 專用）...")

    if not broker_10d_detail_rows:
        print("  ⚠️ 近10日分點勝率排行：沒有近10日分點明細資料")
        return []

    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = get_result_data_scope()

    agg = {}

    for row in broker_10d_detail_rows:
        broker_label = str(row.get("分點", "")).strip()
        broker_name = str(row.get("分點名稱", "")).strip()
        broker_code = str(row.get("券商代號", "")).strip()

        if not broker_label:
            continue

        key = (broker_label, broker_code)
        rec = agg.setdefault(key, {
            "資料範圍": row.get("資料範圍", scope),
            "統計日期": row.get("統計日期", ""),
            "統計期間": row.get("統計期間", ""),
            "統計天數": row.get("統計天數", ""),
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "win": 0,
            "loss": 0,
            "buy_amount": 0.0,
            "sell_amount": 0.0,
            "net_buy_amount": 0.0,
            "net_sell_amount": 0.0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
            "underlying_codes": set(),
            "warrant_count": 0,
            "detail_rows": [],
            "date_set": set(),
            "first_date": "",
            "last_date": "",
        })

        if broker_name and not rec.get("分點名稱"):
            rec["分點名稱"] = broker_name
        if broker_code and not rec.get("券商代號"):
            rec["券商代號"] = broker_code

        result = str(row.get("判定", "")).strip()
        if result == "勝":
            rec["win"] += 1
        elif result == "敗":
            rec["loss"] += 1

        buy_amount = top15_safe_float(row.get("近10日買進金額", 0), 0.0)
        sell_amount = top15_safe_float(row.get("近10日賣出金額", 0), 0.0)
        net_buy_amount = max(top15_safe_float(row.get("近10日淨買超金額", 0), 0.0), 0.0)
        net_sell_amount = max(top15_safe_float(row.get("近10日淨賣超金額", 0), 0.0), 0.0)
        warrant_count = int(top15_safe_float(row.get("涉及權證檔數", 0), 0.0) or 0)

        rec["buy_amount"] += buy_amount
        rec["sell_amount"] += sell_amount
        rec["net_buy_amount"] += net_buy_amount
        rec["net_sell_amount"] += net_sell_amount
        rec["warrant_count"] += warrant_count

        underlying_code = str(row.get("標的股", "")).strip()
        if underlying_code:
            rec["underlying_codes"].add(underlying_code)

        for d in [row.get("第一筆日期", ""), row.get("最後筆日期", "")]:
            d = normalize_date_str(strip_gsheet_text_prefix(d))
            if parse_date(d):
                rec["date_set"].add(d)

        ret = row.get("_分點統計報酬率數值")
        if ret is None:
            ret = _parse_pct_text_to_float(row.get("用於勝率報酬%", ""))
        weight = top15_safe_float(row.get("_分點統計權重", 0), 0.0)
        if weight <= 0:
            direction = str(row.get("買賣方向", "")).strip()
            if direction == "買超":
                weight = net_buy_amount
            elif direction == "賣超":
                weight = net_sell_amount
            else:
                weight = max(buy_amount, sell_amount, 0.0)

        if ret is not None and weight > 0:
            ret = top15_safe_float(ret, None)
            if ret is not None:
                rec["return_weighted_numerator"] += ret * weight
                rec["return_weight"] += weight
                if ret > 0:
                    rec["win_return_weighted_numerator"] += ret * weight
                    rec["win_return_weight"] += weight
                else:
                    rec["loss_return_weighted_numerator"] += ret * weight
                    rec["loss_return_weight"] += weight

        buy_ret = row.get("_分點統計買超報酬率數值")
        if buy_ret is None:
            buy_ret = _parse_pct_text_to_float(row.get("買超平均報酬%", ""))
        buy_weight = top15_safe_float(row.get("_分點統計買超權重", 0), 0.0)
        if buy_weight <= 0:
            buy_weight = net_buy_amount
        if buy_ret is not None and buy_weight > 0:
            buy_ret = top15_safe_float(buy_ret, None)
            if buy_ret is not None:
                rec["buy_return_weighted_numerator"] += buy_ret * buy_weight
                rec["buy_return_weight"] += buy_weight

        sell_ret = row.get("_分點統計賣超報酬率數值")
        if sell_ret is None:
            sell_ret = _parse_pct_text_to_float(row.get("賣超平均報酬%", ""))
        sell_weight = top15_safe_float(row.get("_分點統計賣超權重", 0), 0.0)
        if sell_weight <= 0:
            sell_weight = net_sell_amount
        if sell_ret is not None and sell_weight > 0:
            sell_ret = top15_safe_float(sell_ret, None)
            if sell_ret is not None:
                rec["sell_return_weighted_numerator"] += sell_ret * sell_weight
                rec["sell_return_weight"] += sell_weight

        detail_amount = max(buy_amount + sell_amount, abs(net_buy_amount), abs(net_sell_amount), 0.0)
        rec["detail_rows"].append({
            "標的股": row.get("標的股", ""),
            "標的名稱": row.get("標的名稱", ""),
            "買賣方向": row.get("買賣方向", ""),
            "判定": result,
            "用於勝率報酬%": row.get("用於勝率報酬%", ""),
            "近10日買進金額": round(buy_amount, 0),
            "近10日賣出金額": round(sell_amount, 0),
            "近10日淨買超金額": round(net_buy_amount, 0),
            "近10日淨賣超金額": round(net_sell_amount, 0),
            "涉及權證檔數": warrant_count,
            "權證清單": row.get("權證清單", ""),
            "排序金額": detail_amount,
        })

    rows = []

    for rec in agg.values():
        total = int(rec.get("win", 0) or 0) + int(rec.get("loss", 0) or 0)
        if total <= 0:
            continue

        win_rate_value = rec["win"] / total * 100 if total else None
        avg_return = rec["return_weighted_numerator"] / rec["return_weight"] if rec["return_weight"] > 0 else None
        avg_buy_return = rec["buy_return_weighted_numerator"] / rec["buy_return_weight"] if rec["buy_return_weight"] > 0 else None
        avg_sell_return = rec["sell_return_weighted_numerator"] / rec["sell_return_weight"] if rec["sell_return_weight"] > 0 else None
        avg_win_return = rec["win_return_weighted_numerator"] / rec["win_return_weight"] if rec["win_return_weight"] > 0 else None
        avg_loss_return = rec["loss_return_weighted_numerator"] / rec["loss_return_weight"] if rec["loss_return_weight"] > 0 else None

        profit_loss_ratio = None
        if avg_win_return is not None and avg_win_return > 0 and avg_loss_return is not None and avg_loss_return < 0:
            profit_loss_ratio = avg_win_return / abs(avg_loss_return)

        expectancy = None
        if total > 0 and (avg_win_return is not None or avg_loss_return is not None):
            win_rate_decimal = rec["win"] / total
            loss_rate_decimal = rec["loss"] / total
            expectancy = win_rate_decimal * (avg_win_return or 0.0) + loss_rate_decimal * (avg_loss_return or 0.0)
        elif avg_return is not None:
            expectancy = avg_return

        date_values = sorted(rec.get("date_set", set()))
        first_date = date_values[0] if date_values else ""
        last_date = date_values[-1] if date_values else ""

        details_sorted = sorted(
            rec.get("detail_rows", []),
            key=lambda x: float(x.get("排序金額", 0) or 0),
            reverse=True,
        )
        main_underlying_rows = []
        detail_json = []
        for d in details_sorted:
            amount = max(
                float(d.get("近10日買進金額", 0) or 0),
                float(d.get("近10日賣出金額", 0) or 0),
                float(d.get("近10日淨買超金額", 0) or 0),
                float(d.get("近10日淨賣超金額", 0) or 0),
            )
            main_underlying_rows.append(
                f"{d.get('標的股', '')} {d.get('標的名稱', '')}".strip()
                + f"｜{d.get('買賣方向', '')}｜{d.get('判定', '')}｜{d.get('用於勝率報酬%', '')}｜{_fmt_wan_text(amount)}"
            )
            detail_json.append({k: v for k, v in d.items() if k != "排序金額"})

        total_trade_amount = float(rec.get("buy_amount", 0.0) or 0.0) + float(rec.get("sell_amount", 0.0) or 0.0)

        rows.append({
            "資料範圍": rec.get("資料範圍", scope),
            "統計日期": rec.get("統計日期", ""),
            "統計期間": rec.get("統計期間", ""),
            "統計天數": rec.get("統計天數", ""),
            "有效日期數": len(date_values),
            "第一筆日期": add_gsheet_text_prefix(first_date) if first_date else "",
            "最後筆日期": add_gsheet_text_prefix(last_date) if last_date else "",
            "排名": 0,
            "分點": rec.get("分點", ""),
            "分點名稱": rec.get("分點名稱", ""),
            "券商代號": rec.get("券商代號", ""),
            "近10日勝率": _fmt_pct_text(win_rate_value, signed=False),
            "近10日勝筆數": int(rec.get("win", 0) or 0),
            "近10日敗筆數": int(rec.get("loss", 0) or 0),
            "近10日統計筆數": total,
            "近10日加權平均報酬%": _fmt_pct_text(avg_return, signed=True),
            "近10日買超加權平均報酬%": _fmt_pct_text(avg_buy_return, signed=True),
            "近10日賣超加權平均報酬%": _fmt_pct_text(avg_sell_return, signed=True),
            "近10日加權平均勝報酬%": _fmt_pct_text(avg_win_return, signed=True),
            "近10日加權平均敗報酬%": _fmt_pct_text(avg_loss_return, signed=True),
            "近10日盈虧比": f"{profit_loss_ratio:.2f}" if profit_loss_ratio is not None else "-",
            "近10日加權期望值%": _fmt_pct_text(expectancy, signed=True),
            "近10日加權報酬權重金額": round(float(rec.get("return_weight", 0.0) or 0.0), 0),
            "近10日買進金額": round(float(rec.get("buy_amount", 0.0) or 0.0), 0),
            "近10日賣出金額": round(float(rec.get("sell_amount", 0.0) or 0.0), 0),
            "近10日總交易金額": round(total_trade_amount, 0),
            "近10日淨買超金額": round(float(rec.get("net_buy_amount", 0.0) or 0.0), 0),
            "近10日淨賣超金額": round(float(rec.get("net_sell_amount", 0.0) or 0.0), 0),
            "涉及標的數": len(rec.get("underlying_codes", set())),
            "涉及權證檔數": int(rec.get("warrant_count", 0) or 0),
            "主要交易標的": "；".join(main_underlying_rows[:10]),
            "排序說明": "勝率高到低，其次統計筆數、加權報酬、總交易金額",
            "更新時間": update_time,
            "run_id": run_id,
            "標的明細_JSON": json.dumps(detail_json, ensure_ascii=False),
            "_勝率數值": win_rate_value if win_rate_value is not None else -999999,
            "_加權報酬數值": avg_return if avg_return is not None else -999999,
            "_總交易金額": total_trade_amount,
        })

    rows = sorted(
        rows,
        key=lambda r: (
            -float(r.get("_勝率數值", -999999) or -999999),
            -int(r.get("近10日統計筆數", 0) or 0),
            -float(r.get("_加權報酬數值", -999999) or -999999),
            -float(r.get("_總交易金額", 0) or 0),
            str(r.get("分點", "")),
        )
    )

    for idx, row in enumerate(rows, start=1):
        row["排名"] = idx

    print(f"  ✅ 近10日分點勝率排行完成：{len(rows):,} 個分點")
    return rows


def write_10d_broker_winrate_rank_sheet(wb, rows):
    """寫入近10日分點勝率排行。rows=None 代表不建立工作表。"""
    if rows is None:
        return

    ws = wb.create_sheet(BROKER_10D_WINRATE_RANK_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期", "排名",
        "分點", "分點名稱", "券商代號",
        "近10日勝率", "近10日勝筆數", "近10日敗筆數", "近10日統計筆數",
        "近10日加權平均報酬%", "近10日買超加權平均報酬%", "近10日賣超加權平均報酬%",
        "近10日加權平均勝報酬%", "近10日加權平均敗報酬%", "近10日盈虧比", "近10日加權期望值%",
        "近10日加權報酬權重金額", "近10日買進金額", "近10日賣出金額", "近10日總交易金額",
        "近10日淨買超金額", "近10日淨賣超金額", "涉及標的數", "涉及權證檔數",
        "主要交易標的", "排序說明", "更新時間", "run_id", "標的明細_JSON",
    ]

    ws.append(headers)

    for row in rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [
        12, 12, 24, 10, 12, 12, 12, 8,
        14, 18, 12,
        12, 12, 12, 12,
        18, 22, 22,
        20, 20, 12, 18,
        18, 16, 16, 16,
        18, 18, 12, 12,
        80, 44, 20, 18, 90,
    ]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 26

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    win_rate_col = header_map.get("近10日勝率")

    for row in ws.iter_rows(min_row=2):
        win_rate_value = _parse_pct_text_to_float(row[win_rate_col - 1].value) if win_rate_col else None
        if win_rate_value is not None and win_rate_value >= 70:
            row_fill = RED
        elif win_rate_value is not None and win_rate_value < 50:
            row_fill = GREEN
        else:
            row_fill = WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 32

    ws.freeze_panes = "A2"


# ══════════════════════════════════════════════════════════════════════
# 近 7／14／21 日精選分點權證共識買賣超 TOP15（僅 RUN_MODE=2 更新）
# ══════════════════════════════════════════════════════════════════════

def _fmt_wan_text(value):
    try:
        return f"{float(value) / 10000:.1f}萬"
    except Exception:
        return "0.0萬"


def parse_warrant_consensus_selected_brokers():
    """取得近 7／14／21 日共識排名專用的精選分點清單。"""
    if WARRANT_CONSENSUS_SELECTED_BROKERS_ENV:
        labels = [
            x.strip()
            for x in re.split(r"[,;；、\n\r\t]+", WARRANT_CONSENSUS_SELECTED_BROKERS_ENV)
            if x.strip()
        ]
    else:
        labels = list(WARRANT_CONSENSUS_SELECTED_BROKERS_DEFAULT)

    out = []
    for label in labels:
        if label in FULL_TARGET_PATTERNS and label not in out:
            out.append(label)

    return out


def build_7d_warrant_consensus_top15_rows(items, target_date=None):
    """
    建立「快取_近7日權證分點共識TOP15」。

    重要規則：
    1. 只在 RUN_MODE=2 完整分點清單模式執行，確保 13 個指定分點的歷史資料都有被更新。
    2. 實際排名只統計 parse_warrant_consensus_selected_brokers() 回傳的精選分點。
    3. 同一張工作表依序放入近 7 日、近 14 日、近 21 日三組排名資料。
    4. 各期間都以「標的股」層級完整合併同標的全部權證，再各自排序取 TOP15。
    5. 共識買超 TOP15：買進金額 - 賣出金額 > 0，依淨買超金額排序。
    6. 共識賣超 TOP15：賣出金額 - 買進金額 > 0，依淨賣超金額排序。
    7. 近 7／14／21 日沿用原本邏輯，皆以日曆日區間計算。
    """
    if not WARRANT_CONSENSUS_7D_ENABLED:
        return None

    if RUN_MODE != 2:
        print("  ✅ RUN_MODE=1 精選分點模式：略過近7／14／21日精選分點共識TOP15工作表，避免動到 Google Sheet 既有資料。")
        return None

    selected_brokers = parse_warrant_consensus_selected_brokers()
    selected_broker_set = set(selected_brokers)
    selected_broker_codes = {
        str(FULL_FALLBACK[label][1]).strip().lower()
        for label in selected_brokers
        if label in FULL_FALLBACK
    }

    print(
        "【Step 4c】建立近7／14／21日精選分點共識買賣超 TOP15"
        f"（標的層級，共 {len(selected_brokers)} 個分點）..."
    )
    print(f"  ✅ 統計分點：{', '.join(selected_brokers)}")

    if not items:
        print("  ⚠️ 近7／14／21日精選分點共識TOP15：沒有 items 資料")
        return []

    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date)

    if not target_dt:
        target_dt = datetime.today()
        target_date = target_dt.strftime("%Y/%m/%d")

    window_days_list = []
    for days in [
        WARRANT_CONSENSUS_7D_DAYS,
        WARRANT_CONSENSUS_14D_DAYS,
        WARRANT_CONSENSUS_21D_DAYS,
    ]:
        days = max(int(days), 1)
        if days not in window_days_list:
            window_days_list.append(days)

    top_n = max(int(WARRANT_CONSENSUS_7D_TOP_N), 1)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = f"精選{len(selected_brokers)}分點"

    def build_rows_for_window(window_days):
        start_dt = target_dt - timedelta(days=window_days - 1)
        start_date = start_dt.strftime("%Y/%m/%d")
        period_text = f"{start_date} ～ {target_date}"

        # group_key = 標的股。
        # 同標的底下所有權證、所有指定精選分點先完整加總，再做 TOP15 排名。
        agg = {}

        for item in items:
            broker_label = str(item.get("broker_label", "")).strip()
            broker_code = str(item.get("broker_code", "")).strip()
            broker_code_key = broker_code.lower()

            if broker_label not in selected_broker_set and broker_code_key not in selected_broker_codes:
                continue

            df = item.get("df", pd.DataFrame())

            if df is None or df.empty:
                continue

            warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
            if not warrant_code:
                continue

            warrant_name = str(item.get("warrant_name", "")).strip()
            raw_underlying_code = str(item.get("underlying_code", "")).strip()
            underlying_name = str(item.get("underlying_name", "")).strip()
            underlying_code = normalize_underlying_code_for_group(raw_underlying_code, underlying_name or warrant_name)
            broker_name = str(item.get("broker_name", "")).strip()

            if not underlying_code:
                continue

            rec = agg.setdefault(underlying_code, {
                "標的股": underlying_code,
                "標的名稱": underlying_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "分點": {},
                "權證": {},
                "日期集合": set(),
            })

            if underlying_name and not rec.get("標的名稱"):
                rec["標的名稱"] = underlying_name

            warrant_rec = rec["權證"].setdefault(warrant_code, {
                "權證代號": warrant_code,
                "權證名稱": warrant_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "日期集合": set(),
            })
            if warrant_name and not warrant_rec.get("權證名稱"):
                warrant_rec["權證名稱"] = warrant_name

            broker_key = (broker_label, broker_name, broker_code)
            broker_rec = rec["分點"].setdefault(broker_key, {
                "分點": broker_label,
                "分點名稱": broker_name,
                "券商代號": broker_code,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "權證": {},
                "日期集合": set(),
            })
            broker_warrant_rec = broker_rec["權證"].setdefault(warrant_code, {
                "權證代號": warrant_code,
                "權證名稱": warrant_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "日期集合": set(),
            })

            for row in df.itertuples(index=False):
                row_dict = row._asdict()
                date_str = normalize_date_str(row_dict.get("日期", ""))
                dt = parse_date(date_str)

                if not dt or dt < start_dt or dt > target_dt:
                    continue

                buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
                sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))
                buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
                sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))

                if buy_amount <= 0 and sell_amount <= 0 and buy_qty <= 0 and sell_qty <= 0:
                    continue

                rec["日期集合"].add(date_str)
                rec["買進金額"] += buy_amount
                rec["賣出金額"] += sell_amount
                rec["買進股數"] += buy_qty
                rec["賣出股數"] += sell_qty

                warrant_rec["日期集合"].add(date_str)
                warrant_rec["買進金額"] += buy_amount
                warrant_rec["賣出金額"] += sell_amount
                warrant_rec["買進股數"] += buy_qty
                warrant_rec["賣出股數"] += sell_qty

                broker_rec["日期集合"].add(date_str)
                broker_rec["買進金額"] += buy_amount
                broker_rec["賣出金額"] += sell_amount
                broker_rec["買進股數"] += buy_qty
                broker_rec["賣出股數"] += sell_qty

                broker_warrant_rec["日期集合"].add(date_str)
                broker_warrant_rec["買進金額"] += buy_amount
                broker_warrant_rec["賣出金額"] += sell_amount
                broker_warrant_rec["買進股數"] += buy_qty
                broker_warrant_rec["賣出股數"] += sell_qty

        records = []

        for rec in agg.values():
            buy_amount = float(rec.get("買進金額", 0) or 0)
            sell_amount = float(rec.get("賣出金額", 0) or 0)
            buy_qty = float(rec.get("買進股數", 0) or 0)
            sell_qty = float(rec.get("賣出股數", 0) or 0)
            net_buy = buy_amount - sell_amount
            net_sell = sell_amount - buy_amount

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            records.append({
                **rec,
                "買進金額": buy_amount,
                "賣出金額": sell_amount,
                "買進股數": buy_qty,
                "賣出股數": sell_qty,
                "淨買超金額": net_buy,
                "淨賣超金額": net_sell,
            })

        # 排名前再做一次標的層級保護性合併。
        merged_records = {}
        for rec in records:
            key = normalize_underlying_code_for_group(rec.get("標的股", ""), rec.get("標的名稱", ""))
            if not key:
                continue

            if key not in merged_records:
                rec["標的股"] = key
                merged_records[key] = rec
                continue

            dst = merged_records[key]
            for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數", "淨買超金額", "淨賣超金額"]:
                dst[numeric_col] = float(dst.get(numeric_col, 0) or 0) + float(rec.get(numeric_col, 0) or 0)

            if not dst.get("標的名稱") and rec.get("標的名稱"):
                dst["標的名稱"] = rec.get("標的名稱")

            dst.setdefault("日期集合", set()).update(rec.get("日期集合", set()))

            for warrant_code, warrant_rec in rec.get("權證", {}).items():
                if warrant_code in dst.get("權證", {}):
                    dw = dst["權證"][warrant_code]
                    for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                        dw[numeric_col] = float(dw.get(numeric_col, 0) or 0) + float(warrant_rec.get(numeric_col, 0) or 0)
                    dw.setdefault("日期集合", set()).update(warrant_rec.get("日期集合", set()))
                else:
                    dst.setdefault("權證", {})[warrant_code] = warrant_rec

            for broker_key, broker_rec in rec.get("分點", {}).items():
                if broker_key not in dst.get("分點", {}):
                    dst.setdefault("分點", {})[broker_key] = broker_rec
                    continue

                db = dst["分點"][broker_key]
                for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                    db[numeric_col] = float(db.get(numeric_col, 0) or 0) + float(broker_rec.get(numeric_col, 0) or 0)
                db.setdefault("日期集合", set()).update(broker_rec.get("日期集合", set()))

                for warrant_code, bw in broker_rec.get("權證", {}).items():
                    if warrant_code in db.get("權證", {}):
                        dbw = db["權證"][warrant_code]
                        for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                            dbw[numeric_col] = float(dbw.get(numeric_col, 0) or 0) + float(bw.get(numeric_col, 0) or 0)
                        dbw.setdefault("日期集合", set()).update(bw.get("日期集合", set()))
                    else:
                        db.setdefault("權證", {})[warrant_code] = bw

        records = list(merged_records.values())

        buy_top = [r for r in records if float(r.get("淨買超金額", 0) or 0) > 0]
        sell_top = [r for r in records if float(r.get("淨賣超金額", 0) or 0) > 0]

        buy_top = sorted(
            buy_top,
            key=lambda r: (float(r.get("淨買超金額", 0) or 0), float(r.get("買進金額", 0) or 0)),
            reverse=True,
        )[:top_n]
        sell_top = sorted(
            sell_top,
            key=lambda r: (float(r.get("淨賣超金額", 0) or 0), float(r.get("賣出金額", 0) or 0)),
            reverse=True,
        )[:top_n]

        def make_rank_rows(rank_type, records_for_rank):
            rows = []
            is_buy = rank_type == "共識買超"

            for rank, rec in enumerate(records_for_rank, 1):
                broker_rows = []
                broker_json = []
                same_direction_count = 0
                opposite_direction_count = 0

                for broker_rec in sorted(
                    rec["分點"].values(),
                    key=lambda x: (
                        (float(x.get("買進金額", 0) or 0) - float(x.get("賣出金額", 0) or 0))
                        if is_buy else
                        (float(x.get("賣出金額", 0) or 0) - float(x.get("買進金額", 0) or 0))
                    ),
                    reverse=True,
                ):
                    b_buy = float(broker_rec.get("買進金額", 0) or 0)
                    b_sell = float(broker_rec.get("賣出金額", 0) or 0)
                    b_net_buy = b_buy - b_sell
                    b_net_sell = b_sell - b_buy
                    direction_amount = b_net_buy if is_buy else b_net_sell

                    if direction_amount > 0:
                        same_direction_count += 1
                        broker_rows.append(f"{broker_rec['分點']} {_fmt_wan_text(direction_amount)}")
                    elif direction_amount < 0:
                        opposite_direction_count += 1

                    broker_warrants = []
                    for bw in sorted(
                        broker_rec.get("權證", {}).values(),
                        key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
                        reverse=True,
                    ):
                        broker_warrants.append({
                            "權證代號": bw.get("權證代號", ""),
                            "權證名稱": bw.get("權證名稱", ""),
                            "買進金額": round(float(bw.get("買進金額", 0) or 0), 0),
                            "賣出金額": round(float(bw.get("賣出金額", 0) or 0), 0),
                            "買進股數": round(float(bw.get("買進股數", 0) or 0), 0),
                            "賣出股數": round(float(bw.get("賣出股數", 0) or 0), 0),
                            "日期數": len(bw.get("日期集合", set())),
                        })

                    broker_json.append({
                        "分點": broker_rec["分點"],
                        "分點名稱": broker_rec["分點名稱"],
                        "券商代號": broker_rec["券商代號"],
                        "買進金額": round(b_buy, 0),
                        "賣出金額": round(b_sell, 0),
                        "淨買超金額": round(b_net_buy, 0),
                        "淨賣超金額": round(b_net_sell, 0),
                        "權證檔數": len(broker_rec.get("權證", {})),
                        "日期數": len(broker_rec.get("日期集合", set())),
                        "權證明細": broker_warrants,
                    })

                rank_amount = float(rec.get("淨買超金額", 0) or 0) if is_buy else float(rec.get("淨賣超金額", 0) or 0)
                dates = sorted(rec.get("日期集合", set()))
                warrant_values = sorted(
                    rec.get("權證", {}).values(),
                    key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
                    reverse=True,
                )
                warrant_list = "；".join([
                    f"{w.get('權證代號', '')} {w.get('權證名稱', '')}"
                    f"｜買{_fmt_wan_text(w.get('買進金額', 0))}／賣{_fmt_wan_text(w.get('賣出金額', 0))}"
                    for w in warrant_values
                ])
                top_warrant = warrant_values[0] if warrant_values else {}

                rows.append({
                    "資料範圍": scope,
                    "統計日期": target_date,
                    "統計期間": period_text,
                    "統計天數": window_days,
                    "有效日期數": len(dates),
                    "第一筆日期": dates[0] if dates else "",
                    "最後筆日期": dates[-1] if dates else "",
                    "排名類型": rank_type,
                    "排名": rank,
                    # 保留舊欄位，避免圖片端或舊公式依欄名讀取時壞掉；但實際排名已是標的層級。
                    "權證代號": top_warrant.get("權證代號", ""),
                    "權證名稱": top_warrant.get("權證名稱", "同標的合計"),
                    "標的股": rec.get("標的股", ""),
                    "標的名稱": rec.get("標的名稱", ""),
                    "權證檔數": len(rec.get("權證", {})),
                    "權證清單": warrant_list,
                    "排名金額": round(rank_amount, 0),
                    "買進金額": round(float(rec.get("買進金額", 0) or 0), 0),
                    "賣出金額": round(float(rec.get("賣出金額", 0) or 0), 0),
                    "淨買超金額": round(float(rec.get("淨買超金額", 0) or 0), 0),
                    "淨賣超金額": round(float(rec.get("淨賣超金額", 0) or 0), 0),
                    "買進股數": round(float(rec.get("買進股數", 0) or 0), 0),
                    "賣出股數": round(float(rec.get("賣出股數", 0) or 0), 0),
                    "參與分點數": len(rec.get("分點", {})),
                    "同向分點數": same_direction_count,
                    "反向分點數": opposite_direction_count,
                    "主要同向分點": "；".join(broker_rows[:8]),
                    "完成狀態": "DONE",
                    "更新時間": update_time,
                    "run_id": run_id,
                    "分點明細_JSON": json.dumps(broker_json, ensure_ascii=False),
                })

            return rows

        rows = []
        rows.extend(make_rank_rows("共識買超", buy_top))
        rows.extend(make_rank_rows("共識賣超", sell_top))

        print(
            f"  ✅ 近{window_days}日精選分點共識TOP15完成："
            f"共識買超 {len(buy_top):,} 檔標的，共識賣超 {len(sell_top):,} 檔標的，"
            f"統計期間 {period_text}"
        )
        return rows

    all_rows = []
    for window_days in window_days_list:
        all_rows.extend(build_rows_for_window(window_days))

    return all_rows


def write_7d_warrant_consensus_top15_sheet(wb, rows):
    """
    寫入近 7／14／21 日精選分點共識買賣超 TOP15。

    rows=None 代表不建立工作表；同一張工作表內由上往下建立三張獨立排名表：
    近 7 日、近 14 日、近 21 日。每張排名表都有自己的標題、統計期間與欄位表頭。
    """
    if rows is None:
        return

    ws = wb.create_sheet(WARRANT_CONSENSUS_7D_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期",
        "排名類型", "排名",
        "權證代號", "權證名稱", "標的股", "標的名稱", "權證檔數", "權證清單",
        "排名金額", "買進金額", "賣出金額", "淨買超金額", "淨賣超金額",
        "買進股數", "賣出股數",
        "參與分點數", "同向分點數", "反向分點數", "主要同向分點",
        "完成狀態", "更新時間", "run_id", "分點明細_JSON",
    ]

    window_order = []
    for days in [
        WARRANT_CONSENSUS_7D_DAYS,
        WARRANT_CONSENSUS_14D_DAYS,
        WARRANT_CONSENSUS_21D_DAYS,
    ]:
        days = max(int(days), 1)
        if days not in window_order:
            window_order.append(days)

    rows_by_window = {days: [] for days in window_order}
    for row in rows or []:
        try:
            days = int(float(row.get("統計天數", 0) or 0))
        except Exception:
            days = 0
        rows_by_window.setdefault(days, []).append(row)

    col_widths = [12, 12, 24, 10, 12, 12, 12, 12, 8, 12, 24, 10, 12, 10, 72, 14, 14, 14, 14, 14, 14, 14, 12, 12, 12, 56, 10, 20, 16, 90]
    thin_gray = Side(style="thin", color="B7B7B7")
    medium_brown = Side(style="medium", color="7F6000")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    section_top_border = Border(left=thin_gray, right=thin_gray, top=medium_brown, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    selected_brokers = parse_warrant_consensus_selected_brokers()
    selected_brokers_text = "、".join(selected_brokers)
    max_col = len(headers)
    first_header_row = None

    for section_index, days in enumerate(window_order):
        section_rows = rows_by_window.get(days, [])

        if section_index > 0:
            ws.append([""] * max_col)
            ws.row_dimensions[ws.max_row].height = 9

        title_row = 1 if section_index == 0 else ws.max_row + 1
        ws.cell(title_row, 1, f"近{days}日精選{len(selected_brokers)}分點權證共識買賣超 TOP15")
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)
        title_cell = ws.cell(title_row, 1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.fill = YELLOW
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = Border(top=medium_brown, bottom=medium_brown)
        ws.row_dimensions[title_row].height = 28

        period_text = str(section_rows[0].get("統計期間", "")).strip() if section_rows else ""
        subtitle_parts = []
        if period_text:
            subtitle_parts.append(f"統計期間：{period_text}")
        subtitle_parts.append(f"統計分點：{selected_brokers_text}")

        subtitle_row = ws.max_row + 1
        ws.cell(subtitle_row, 1, "｜".join(subtitle_parts))
        ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=max_col)
        subtitle_cell = ws.cell(subtitle_row, 1)
        subtitle_cell.font = Font(color="000000")
        subtitle_cell.fill = BLUE
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        subtitle_cell.border = normal_border
        ws.row_dimensions[subtitle_row].height = 32

        header_row = ws.max_row + 1
        ws.append(headers)
        if first_header_row is None:
            first_header_row = header_row

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[header_row].height = 24

        if not section_rows:
            empty_row = ws.max_row + 1
            ws.cell(empty_row, 1, f"近{days}日無符合排名資料")
            ws.merge_cells(start_row=empty_row, start_column=1, end_row=empty_row, end_column=max_col)
            empty_cell = ws.cell(empty_row, 1)
            empty_cell.font = Font(color="666666")
            empty_cell.fill = WHITE
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            empty_cell.border = normal_border
            ws.row_dimensions[empty_row].height = 24
            continue

        previous_rank_type = None
        for record in section_rows:
            ws.append([record.get(h, "") for h in headers])
            current_row = ws.max_row
            rank_type = str(record.get("排名類型", "")).strip()
            row_fill = RED if rank_type == "共識買超" else GREEN if rank_type == "共識賣超" else WHITE
            use_section_top = rank_type != previous_rank_type

            for cell in ws[current_row]:
                cell.font = Font(
                    bold=use_section_top,
                    color="000000",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = section_top_border if use_section_top else normal_border
                cell.fill = row_fill

            # 金額與股數欄使用千分位，維持原本報表的閱讀格式。
            for col_name in [
                "排名金額", "買進金額", "賣出金額", "淨買超金額", "淨賣超金額",
                "買進股數", "賣出股數",
            ]:
                col_idx = headers.index(col_name) + 1
                ws.cell(current_row, col_idx).number_format = '#,##0'

            ws.row_dimensions[current_row].height = 30
            previous_rank_type = rank_type

    ws.freeze_panes = f"A{(first_header_row or 1) + 1}"

def build_excel(a_events, b_events, c_events, d_events, e_events, item_map, price_cache, items, output_path, top15_detail_rows=None, top15_consensus_rows=None, warrant_consensus_7d_rows=None, broker_10d_detail_rows=None, broker_10d_winrate_rank_rows=None):
    print("【Step 5】建立 Excel...")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    write_group_sheet(wb, "A_基礎買超", a_events, price_cache, is_c=False)
    write_group_sheet(wb, "B_明顯買超", b_events, price_cache, is_c=False)
    write_group_sheet(wb, "C_強勢買超", c_events, price_cache, is_c=False)
    write_group_sheet(wb, "D_大額布局", d_events, price_cache, is_c=False)
    write_group_sheet(wb, "E_超大額布局", e_events, price_cache, is_c=False)
    write_daily_sell_detail_sheet(wb, items, a_events, b_events, c_events, d_events, e_events)
    write_top15_consensus_cache_sheet(wb, top15_consensus_rows or [])
    write_top15_position_detail_sheet(wb, top15_detail_rows or [])
    write_7d_warrant_consensus_top15_sheet(wb, warrant_consensus_7d_rows)
    write_10d_broker_underlying_detail_sheet(wb, broker_10d_detail_rows)
    write_10d_broker_winrate_rank_sheet(wb, broker_10d_winrate_rank_rows)
    write_stats_sheet(wb, a_events, b_events, c_events, d_events, e_events)
    write_recent_warrant_amount_ranking_sheet(wb, items)
    write_underlying_broker_count_ranking_sheet(wb, items)
    write_broker_query_sheet(wb, items)
    write_price_status_sheet(wb, price_cache)
    write_color_legend_sheet(wb)
    write_combo_winrate_sheet(wb, a_events, b_events, c_events, d_events, e_events)

    apply_global_amount_comma_format(wb)

    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)

    print(
        f"  ✅ 已存：{output_path} "
        f"（A:{len(a_events)} 筆，B:{len(b_events)} 筆，C:{len(c_events)} 筆，D:{len(d_events)} 筆，E:{len(e_events)} 筆）"
    )



def _collect_all_item_price_codes_for_prefetch(items):
    """
    收集價格預抓完整模式要補的所有代號。

    來源是既有「快取_分點歷史」還原出的 items：
    - 權證代號：所有有分點歷史資料的權證都納入
    - 標的股代號：所有有分點歷史資料的標的股都納入

    注意：這裡只負責收集代號，不改 A/B/C/D/E、TOP15、近10日明細的判斷邏輯。
    """
    warrant_codes = set()
    underlying_codes = set()

    for item in items or []:
        warrant_code = normalize_price_code(item.get("warrant_code", ""))
        if warrant_code:
            warrant_codes.add(warrant_code)

        underlying_code = normalize_underlying_code_for_group(item.get("underlying_code", "")) or normalize_price_code(item.get("underlying_code", ""))
        if underlying_code:
            underlying_codes.add(underlying_code)

    return warrant_codes, underlying_codes


def ensure_price_prefetch_all_item_prices(price_cache, items, target_date=None):
    """
    價格預抓完整模式：補抓所有既有分點歷史項目會牽涉到的價格。

    目的：
    分點資料可能最新只到前一交易日，但盤後價格已經更新到今天。
    這時候預抓模式應該先把所有可能會被後續報表 / 圖卡使用的價格都補進 price_cache，
    不只補目前事件清單、TOP15 或近10日圖卡剛好需要的少數代號。

    判斷：
    - 同一個代號若 price_cache 已有 target_date 當天價格，略過。
    - 若沒有 target_date 當天價格，或完全沒有價格，才補抓。
    - 權證與標的股分開使用不同 lookback，避免權證太久沒成交時完全抓不到價格。
    """
    if not PRICE_PREFETCH_ALL_ITEM_PRICES:
        return price_cache

    warrant_codes, underlying_codes = _collect_all_item_price_codes_for_prefetch(items)

    if not warrant_codes and not underlying_codes:
        print("  ✅ 價格預抓完整模式：沒有可預抓的權證 / 標的股代號。")
        return price_cache

    target_date = normalize_price_prefetch_target_date(target_date)
    target_dt = parse_date(target_date) or datetime.today()
    end_dt = min(target_dt, datetime.today())

    warrant_lookback_days = max(int(PRICE_PREFETCH_ALL_WARRANT_PRICE_LOOKBACK_DAYS), 1)
    underlying_lookback_days = max(int(PRICE_PREFETCH_ALL_UNDERLYING_PRICE_LOOKBACK_DAYS), 1)

    persistent_price_cache = load_price_cache()
    fetch_plan = {}

    def add_to_fetch_plan(code, code_type):
        norm_code = normalize_price_code(code)
        if not norm_code:
            return

        lookback_days = warrant_lookback_days if code_type == "warrant" else underlying_lookback_days
        start_dt = target_dt - timedelta(days=lookback_days)

        cached_prices = get_cached_prices_for_code(persistent_price_cache, norm_code)
        in_memory_prices = get_price_series_from_cache(price_cache, norm_code)
        merged_cached = merge_price_dicts(cached_prices, in_memory_prices)

        if merged_cached:
            add_price_aliases(price_cache, norm_code, merged_cached)
            persistent_price_cache[norm_code] = merged_cached

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, norm_code, target_date)
        latest_dt = parse_date(latest_date) if latest_date else None

        need_fetch = latest_price is None or latest_dt is None
        if not need_fetch and PRICE_PREFETCH_ALL_REQUIRE_TARGET_DATE:
            need_fetch = latest_dt.date() < target_dt.date()

        if need_fetch:
            old_plan = fetch_plan.get(norm_code)
            if old_plan:
                old_start_dt, old_end_dt, old_type = old_plan
                fetch_plan[norm_code] = (min(old_start_dt, start_dt), max(old_end_dt, end_dt), old_type if old_type == "warrant" else code_type)
            else:
                fetch_plan[norm_code] = (start_dt, end_dt, code_type)

    for code in sorted(underlying_codes):
        add_to_fetch_plan(code, "underlying")

    for code in sorted(warrant_codes):
        add_to_fetch_plan(code, "warrant")

    print(
        f"【價格預抓】完整價格預抓：標的股 {len(underlying_codes):,} 檔｜權證 {len(warrant_codes):,} 檔｜"
        f"需補抓 {len(fetch_plan):,} 檔"
    )
    print(
        f"  完整價格預抓 lookback：標的股近 {underlying_lookback_days} 天｜"
        f"權證近 {warrant_lookback_days} 天｜目標日：{target_date}"
    )

    if not fetch_plan:
        return price_cache

    def fetch_one(code):
        start_dt, end_dt, code_type = fetch_plan[code]
        return code, fetch_twse_prices(code, start_dt, end_dt)

    done = 0
    changed_price_codes = set()

    with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as ex:
        futures = {ex.submit(fetch_one, code): code for code in fetch_plan}

        for future in as_completed(futures):
            done += 1
            code = futures[future]

            try:
                _, fetched_prices = future.result()
            except Exception:
                fetched_prices = {}

            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            merged_prices = merge_price_dicts(old_prices, fetched_prices)

            norm_code = normalize_price_code(code)
            if norm_code:
                persistent_price_cache[norm_code] = merged_prices
                if fetched_prices:
                    changed_price_codes.add(norm_code)

            add_price_aliases(price_cache, code, merged_prices)

            if done % 50 == 0:
                print(f"  [{done}/{len(fetch_plan)}] 完整價格預抓中...")

    save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)
    return price_cache


# ══════════════════════════════════════════════════════════════════════
# 自動價格預抓：今日分點資料未出現時，只更新價格快取並快速結束
# ══════════════════════════════════════════════════════════════════════

def normalize_price_prefetch_target_date(target_date=None):
    raw = str(target_date or PRICE_PREFETCH_TARGET_DATE or TOP15_TARGET_DATE or "").strip()

    if raw:
        dt = parse_date(raw)
        if dt:
            return dt.strftime("%Y/%m/%d")

    return datetime.today().strftime("%Y/%m/%d")


def load_price_prefetch_state():
    if not PRICE_PREFETCH_STATE_ENABLED:
        return pd.DataFrame()

    df = read_cache_csv(PRICE_PREFETCH_STATE_PATH)

    if df is None or df.empty:
        return pd.DataFrame()

    required_cols = ["日期", "資料範圍", "模式", "狀態"]
    for col in required_cols:
        if col not in df.columns:
            return pd.DataFrame()

    return df.fillna("")


def is_price_prefetch_done_for_today(target_date=None):
    if not PRICE_PREFETCH_STATE_ENABLED:
        return False

    if PRICE_PREFETCH_FORCE:
        return False

    target_date = normalize_price_prefetch_target_date(target_date)
    scope = get_result_data_scope()
    df = load_price_prefetch_state()

    if df.empty:
        return False

    for row in df.itertuples(index=False):
        rd = row._asdict()
        if normalize_date_str(rd.get("日期", "")) != target_date:
            continue
        if str(rd.get("資料範圍", "")).strip() != scope:
            continue
        if str(rd.get("模式", "")).strip() != "AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING":
            continue
        if str(rd.get("狀態", "")).strip().lower() in ("done", "success", "no_items", "skipped"):
            return True

    return False


def write_price_prefetch_state(record):
    if not PRICE_PREFETCH_STATE_ENABLED or not USE_CACHE:
        return

    headers = [
        "日期", "資料範圍", "模式", "狀態", "原因",
        "候選組合數", "快取歷史筆數", "還原項目數", "A事件數", "B事件數", "C事件數", "D事件數",
        "價格快取代號數", "價格最新日期", "目標日價格代號數", "執行時間", "更新時間",
    ]

    old_df = read_cache_csv(PRICE_PREFETCH_STATE_PATH)
    rows = []

    if old_df is not None and not old_df.empty:
        for _, old_row in old_df.fillna("").iterrows():
            rows.append({h: old_row.get(h, "") for h in headers})

    rec = {h: record.get(h, "") for h in headers}
    rec["更新時間"] = rec.get("更新時間") or datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    rows.append(rec)

    df = pd.DataFrame(rows, columns=headers).fillna("")
    df = df.drop_duplicates(
        subset=["日期", "資料範圍", "模式"],
        keep="last"
    ).reset_index(drop=True)

    write_cache_csv(df, PRICE_PREFETCH_STATE_PATH)


def price_cache_target_date_summary(price_cache, target_date=None):
    """
    統計價格快取是否已經有 target_date 的收盤價。

    目的：
    盤後價格有時會比 API4 分點資料更早更新。若早盤/盤中已執行過價格預抓，
    price_prefetch_state 可能已是 done，但當時 price_cache 還沒有今天收盤價。
    這個函式用來判斷是否應該在盤後再預抓一次最新價格。
    """
    target_date = normalize_price_prefetch_target_date(target_date)
    target_key = normalize_date_str(target_date)

    latest_dt = None
    latest_date = ""
    target_date_code_count = 0

    if not price_cache:
        return {
            "target_date": target_key,
            "target_date_code_count": 0,
            "latest_date": "",
        }

    counted_codes = set()

    for code, prices in price_cache.items():
        norm_code = normalize_price_code(code)
        if not norm_code or not isinstance(prices, dict):
            continue

        has_target_price = False

        for d, p in prices.items():
            dt = parse_date(d)
            price = safe_price_float(p)
            if not dt or price is None:
                continue

            d_norm = dt.strftime("%Y/%m/%d")
            if latest_dt is None or dt > latest_dt:
                latest_dt = dt
                latest_date = d_norm

            if d_norm == target_key:
                has_target_price = True

        if has_target_price and norm_code not in counted_codes:
            counted_codes.add(norm_code)
            target_date_code_count += 1

    return {
        "target_date": target_key,
        "target_date_code_count": target_date_code_count,
        "latest_date": latest_date,
    }


def price_cache_has_target_date_prices(target_date=None, min_codes=None):
    if min_codes is None:
        min_codes = PRICE_PREFETCH_MIN_TARGET_PRICE_CODES

    price_cache = load_price_cache()
    summary = price_cache_target_date_summary(price_cache, target_date)
    return int(summary.get("target_date_code_count", 0) or 0) >= max(int(min_codes or 1), 1), summary


def price_cache_has_exact_target_date_price_for_code(price_cache, code, target_date):
    target_key = normalize_date_str(target_date)
    prices = get_cached_prices_for_code(price_cache, code)

    if not prices:
        return False

    for d, p in prices.items():
        dt = parse_date(d)
        price = safe_price_float(p)

        if not dt or price is None:
            continue

        if dt.strftime("%Y/%m/%d") == target_key:
            return True

    return False


def price_cache_has_required_10d_underlying_target_prices(history_cache_df, candidate_keys=None, target_date=None):
    """
    檢查近10日圖卡「現股10日」會用到的標的股，是否都已經有目標日收盤價。

    修正重點：
    舊版只要 price_cache 裡任一代號有今天價格，就會因 price_prefetch_state=done 而快速結束，
    導致近10日現股標的仍停在前一交易日。這裡改成必須檢查近10日實際需要的標的股。
    """
    target_date = normalize_price_prefetch_target_date(target_date)

    summary = {
        "target_date": target_date,
        "required_underlying_count": 0,
        "target_date_underlying_count": 0,
        "missing_underlying_count": 0,
        "latest_date": "",
        "missing_sample": "",
    }

    if history_cache_df is None or history_cache_df.empty:
        return False, summary

    try:
        if candidate_keys:
            items = items_from_history_cache(history_cache_df, candidate_filter=candidate_keys)
        else:
            items = items_from_history_cache(history_cache_df)
    except Exception:
        items = []

    required_codes = _collect_recent_underlying_codes_for_10d(items, target_date)
    required_codes = {normalize_underlying_code_for_group(c) or normalize_price_code(c) for c in required_codes if str(c).strip()}
    required_codes = {c for c in required_codes if c}

    summary["required_underlying_count"] = len(required_codes)

    if not required_codes:
        generic_has_target, generic_summary = price_cache_has_target_date_prices(target_date)
        summary["latest_date"] = generic_summary.get("latest_date", "")
        return generic_has_target, summary

    price_cache = load_price_cache()
    target_ok_codes = set()
    missing_codes = []
    latest_dt = None
    latest_date = ""

    for code in sorted(required_codes):
        prices = get_cached_prices_for_code(price_cache, code)
        has_target = False

        for d, p in prices.items():
            dt = parse_date(d)
            price = safe_price_float(p)

            if not dt or price is None:
                continue

            d_norm = dt.strftime("%Y/%m/%d")
            if latest_dt is None or dt > latest_dt:
                latest_dt = dt
                latest_date = d_norm

            if d_norm == target_date:
                has_target = True

        if has_target:
            target_ok_codes.add(code)
        else:
            missing_codes.append(code)

    summary["target_date_underlying_count"] = len(target_ok_codes)
    summary["missing_underlying_count"] = len(missing_codes)
    summary["latest_date"] = latest_date
    summary["missing_sample"] = ", ".join(missing_codes[:10])

    return len(missing_codes) == 0, summary


def has_today_broker_data_from_prescan(target_date=None):
    target_date = normalize_price_prefetch_target_date(target_date)

    if PRESCAN_TODAY_ACTIVITY_FOUND:
        return True

    if PRESCAN_LATEST_ACTIVITY_DATE:
        return PRESCAN_LATEST_ACTIVITY_DATE.strftime("%Y/%m/%d") >= target_date

    return False


def normalize_moneydj_search_repair_target_date(target_date=None):
    """
    MoneyDJ Search 補漏模式使用的報表目標日。

    優先順序沿用既有價格預抓 / TOP15 目標日設定，避免新增一套日期邏輯：
    1. 函式傳入 target_date
    2. PRICE_PREFETCH_TARGET_DATE
    3. TOP15_TARGET_DATE
    4. 今天
    """
    return normalize_price_prefetch_target_date(target_date)


def get_openapi_latest_activity_date_str():
    if PRESCAN_LATEST_ACTIVITY_DATE:
        return PRESCAN_LATEST_ACTIVITY_DATE.strftime("%Y/%m/%d")
    return ""


def should_activate_moneydj_search_repair(target_date=None):
    """
    判斷是否啟用 MoneyDJ Search 自動補漏。

    條件：
    - MONEYDJ_SEARCH_REPAIR_ENABLED=1
    - OpenAPI / API4 預掃描最新交易日 < 本次報表目標日

    只要落後，就不讓流程停在價格預抓，而是繼續進入正式報表流程，
    並強制用 MoneyDJ API5 補抓可能漏掉的候選組合。
    """
    if not MONEYDJ_SEARCH_REPAIR_ENABLED and not workflow_is_repair():
        return False

    target_date = normalize_moneydj_search_repair_target_date(target_date)
    target_dt = parse_date(target_date)

    if not target_dt:
        return False

    if PRESCAN_LATEST_ACTIVITY_DATE is None:
        return True

    return PRESCAN_LATEST_ACTIVITY_DATE.date() < target_dt.date()


def activate_moneydj_search_repair_if_needed(target_date=None):
    global MONEYDJ_SEARCH_REPAIR_ACTIVE
    global MONEYDJ_SEARCH_REPAIR_TARGET_DATE
    global MONEYDJ_SEARCH_REPAIR_OPENAPI_LATEST_DATE
    global MONEYDJ_SEARCH_REPAIR_REASON

    if not workflow_is_repair():
        MONEYDJ_SEARCH_REPAIR_ACTIVE = False
        MONEYDJ_SEARCH_REPAIR_TARGET_DATE = normalize_moneydj_search_repair_target_date(target_date)
        MONEYDJ_SEARCH_REPAIR_OPENAPI_LATEST_DATE = get_openapi_latest_activity_date_str() or "-"
        MONEYDJ_SEARCH_REPAIR_REASON = ""
        if WORKFLOW_MODE == "daily":
            print("  ✅ WORKFLOW_MODE=daily：不啟用 MoneyDJ Search Repair；資料未出現時改走價格預抓並停止。")
        return False

    target_date = normalize_moneydj_search_repair_target_date(target_date)
    latest_date = get_openapi_latest_activity_date_str() or "-"

    if not should_activate_moneydj_search_repair(target_date):
        MONEYDJ_SEARCH_REPAIR_ACTIVE = False
        MONEYDJ_SEARCH_REPAIR_TARGET_DATE = target_date
        MONEYDJ_SEARCH_REPAIR_OPENAPI_LATEST_DATE = latest_date
        MONEYDJ_SEARCH_REPAIR_REASON = ""
        return False

    MONEYDJ_SEARCH_REPAIR_ACTIVE = True
    MONEYDJ_SEARCH_REPAIR_TARGET_DATE = target_date
    MONEYDJ_SEARCH_REPAIR_OPENAPI_LATEST_DATE = latest_date
    MONEYDJ_SEARCH_REPAIR_REASON = f"OpenAPI最新交易日 {latest_date} < 報表目標日 {target_date}"

    print(
        "  🔄 自動啟用 MoneyDJ Search 補漏模式："
        f"{MONEYDJ_SEARCH_REPAIR_REASON}。"
    )
    print(
        "  🔄 本次不會停在價格預抓，會繼續用 MoneyDJ API5 補抓可能漏掉的分點資料。"
    )
    return True


def moneydj_search_repair_is_active():
    return bool(MONEYDJ_SEARCH_REPAIR_ACTIVE and (MONEYDJ_SEARCH_REPAIR_ENABLED or workflow_is_repair()))


def _history_latest_and_net_by_candidate(history_cache_df, candidate_keys=None):
    """
    從既有分點歷史快取整理每一個「權證代號 + 券商代號」的：
    1. 最後資料日期
    2. 淨庫存股數
    3. 最後日期買賣金額

    MoneyDJ Search 補漏模式用這個結果挑出最需要強制 API5 補抓的候選：
    - 目前仍有淨庫存者
    - 最近一段時間有活動者
    - API4 直接掃到 / 活動標的擴展出的候選
    """
    out = {}

    if history_cache_df is None or history_cache_df.empty:
        return out

    required_cols = {"權證代號", "券商代號", "日期", "買進股數", "賣出股數", "買進金額", "賣出金額"}
    if not required_cols.issubset(set(history_cache_df.columns)):
        return out

    candidate_keys = candidate_keys or None

    df = history_cache_df[["權證代號", "券商代號", "日期", "買進股數", "賣出股數", "買進金額", "賣出金額"]].copy().fillna("")

    for col in ["買進股數", "賣出股數", "買進金額", "賣出金額"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    for (warrant_code, broker_code), g in df.groupby(["權證代號", "券商代號"], dropna=False):
        key = candidate_key_from_values(warrant_code, broker_code)

        if candidate_keys is not None and key not in candidate_keys:
            continue

        latest_dt = None
        latest_buy_amount = 0
        latest_sell_amount = 0
        net_shares = 0

        for row in g.itertuples(index=False):
            rd = row._asdict()
            dt = parse_date(rd.get("日期", ""))
            buy_shares = _supabase_int(rd.get("買進股數", 0)) if "_supabase_int" in globals() else int(rd.get("買進股數", 0) or 0)
            sell_shares = _supabase_int(rd.get("賣出股數", 0)) if "_supabase_int" in globals() else int(rd.get("賣出股數", 0) or 0)
            buy_amount = _supabase_int(rd.get("買進金額", 0)) if "_supabase_int" in globals() else int(rd.get("買進金額", 0) or 0)
            sell_amount = _supabase_int(rd.get("賣出金額", 0)) if "_supabase_int" in globals() else int(rd.get("賣出金額", 0) or 0)

            net_shares += buy_shares - sell_shares

            if dt and (latest_dt is None or dt > latest_dt):
                latest_dt = dt
                latest_buy_amount = buy_amount
                latest_sell_amount = sell_amount

        if latest_dt:
            out[key] = {
                "latest_dt": latest_dt,
                "latest_date": latest_dt.strftime("%Y/%m/%d"),
                "net_shares": net_shares,
                "has_open_position": net_shares > 0,
                "latest_buy_amount": latest_buy_amount,
                "latest_sell_amount": latest_sell_amount,
                "latest_total_amount": abs(latest_buy_amount) + abs(latest_sell_amount),
            }

    return out


def _safe_int_for_repair(value):
    try:
        if value is None:
            return 0
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return 0
        return int(float(s))
    except Exception:
        return 0


def collect_moneydj_search_repair_underlying_broker_pairs_from_history(history_cache_df, target_date=None):
    """
    API4 / OpenAPI 落後時，用既有分點歷史快取整理「標的股 + 分點」補漏池。

    目的：
    1. 先找出這個分點最近對哪些標的股的權證有動作。
    2. 再把這些標的股展開成「同標的所有認購權證 × 同分點」。
    3. 後續用 API5 檢查今天是否有新買超 / 賣超，避免只抓到昨天持倉的隔日衝。

    注意：
    如果某個標的是今天第一次完全陌生出現，且 API4 也沒有更新，
    任何非全市場掃描都無法事先知道它；這裡處理的是「近期有動作 / 尚有庫存」的高機率標的池。
    """
    pairs = set()
    reason_count = {"尚有淨庫存": 0, "近期有歷史活動": 0}

    if history_cache_df is None or history_cache_df.empty:
        return pairs, reason_count

    required_cols = {
        "標的股", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
    }
    if not required_cols.issubset(set(history_cache_df.columns)):
        return pairs, reason_count

    target_date = normalize_moneydj_search_repair_target_date(target_date)
    target_dt = parse_date(target_date) or datetime.today()
    recent_days = max(int(MONEYDJ_SEARCH_REPAIR_RECENT_HISTORY_DAYS or 0), 0)
    recent_floor_dt = target_dt - timedelta(days=recent_days)

    df = history_cache_df[[
        "標的股", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
    ]].copy().fillna("")

    df["標的股"] = df["標的股"].map(lambda x: normalize_underlying_code_for_group(x) or str(x).strip())
    df["券商代號"] = df["券商代號"].astype(str).str.strip()

    for col in ["買進股數", "賣出股數", "買進金額", "賣出金額"]:
        df[col] = df[col].map(_safe_int_for_repair)

    df = df[(df["標的股"] != "") & (df["券商代號"] != "")].copy()

    if df.empty:
        return pairs, reason_count

    active_target_broker_codes = {str(code).strip() for _, code in TARGET_PATTERNS.items()}
    broker_codes_in_scope = {str(code).strip() for _, code in FALLBACK.values()}
    if broker_codes_in_scope:
        df = df[df["券商代號"].isin(broker_codes_in_scope)].copy()

    if df.empty:
        return pairs, reason_count

    for (underlying_code, broker_code), g in df.groupby(["標的股", "券商代號"], dropna=False):
        latest_dt = None
        net_shares = 0
        latest_total_amount = 0

        for row in g.itertuples(index=False):
            rd = row._asdict()
            dt = parse_date(rd.get("日期", ""))
            buy_shares = _safe_int_for_repair(rd.get("買進股數", 0))
            sell_shares = _safe_int_for_repair(rd.get("賣出股數", 0))
            buy_amount = _safe_int_for_repair(rd.get("買進金額", 0))
            sell_amount = _safe_int_for_repair(rd.get("賣出金額", 0))

            net_shares += buy_shares - sell_shares

            if dt and (latest_dt is None or dt > latest_dt):
                latest_dt = dt
                latest_total_amount = abs(buy_amount) + abs(sell_amount)

        if not latest_dt:
            continue

        if MONEYDJ_SEARCH_REPAIR_INCLUDE_OPEN_POSITION and net_shares > 0:
            pairs.add((str(underlying_code).strip(), str(broker_code).strip()))
            reason_count["尚有淨庫存"] += 1
            continue

        if recent_days > 0 and latest_dt.date() >= recent_floor_dt.date() and latest_total_amount > 0:
            pairs.add((str(underlying_code).strip(), str(broker_code).strip()))
            reason_count["近期有歷史活動"] += 1

    return pairs, reason_count


def build_moneydj_search_repair_discovery_candidates(warrants, broker_map, history_cache_df, target_date=None):
    """
    MoneyDJ Search 補漏用的「先找標的，再展開權證」候選池。

    當 API4 還停在昨天時，API5 無法用「分點 + 日期」直接反查今日新標的，
    所以這裡先從歷史快取找出該分點近期 / 有庫存的標的股，
    再展開該標的所有認購權證給 API5 檢查今天是否有新資料。
    """
    if not moneydj_search_repair_is_active():
        return []

    if not MONEYDJ_SEARCH_REPAIR_HISTORY_UNDERLYING_EXPAND_ENABLED:
        print("  ⚠️ MONEYDJ_SEARCH_REPAIR_HISTORY_UNDERLYING_EXPAND_ENABLED=0，略過歷史活動標的補漏池。")
        return []

    pairs, reason_count = collect_moneydj_search_repair_underlying_broker_pairs_from_history(
        history_cache_df,
        target_date=target_date,
    )

    if not pairs:
        print("  ⚠️ MoneyDJ Search 補漏：歷史快取沒有可展開的近期活動 / 淨庫存標的。")
        return []

    underlying_codes = {u for u, _ in pairs if str(u).strip()}
    reason_text = "，".join(f"{k}:{v:,}" for k, v in reason_count.items() if v) or "-"

    print(
        f"  🔎 MoneyDJ Search 補漏：從歷史快取找出可能今日有動作的標的 "
        f"{len(underlying_codes):,} 檔｜標的×分點 {len(pairs):,} 組｜{reason_text}"
    )

    candidates = build_underlying_expanded_candidates(
        warrants,
        broker_map,
        underlying_codes,
        source_label="MoneyDJ補漏：歷史活動標的",
        active_pairs=pairs,
    )
    candidates = filter_candidates_by_broker_map(candidates, broker_map)

    print(
        f"  ✅ MoneyDJ Search 補漏：歷史活動標的展開候選 {len(candidates):,} 組"
    )
    return candidates


def build_moneydj_search_repair_fetch_keys(history_cache_df, candidates, target_date=None):
    """
    MoneyDJ Search 補漏候選挑選邏輯。

    不做手動指定分點 / 手動指定權證；只要偵測到 OpenAPI 落後，
    依本次 RUN_MODE 的追蹤分點候選池自動挑出需要補抓的 key。

    補抓優先序：
    1. API4 直接近期活動候選與活動標的擴展候選
    2. 快取中仍有淨庫存的候選
    3. 快取中最近 MONEYDJ_SEARCH_REPAIR_RECENT_HISTORY_DAYS 天有活動的候選

    這樣可避免全市場無差別重打 API5，同時會補到五分點與全分點範圍內最容易漏掉的資料。
    """
    global MONEYDJ_SEARCH_REPAIR_FETCH_KEYS

    MONEYDJ_SEARCH_REPAIR_FETCH_KEYS = set()

    if not moneydj_search_repair_is_active():
        return set()

    target_date = normalize_moneydj_search_repair_target_date(target_date)
    target_dt = parse_date(target_date) or datetime.today()
    recent_days = max(int(MONEYDJ_SEARCH_REPAIR_RECENT_HISTORY_DAYS or 0), 0)
    recent_floor_dt = target_dt - timedelta(days=recent_days)
    max_fetch = int(MONEYDJ_SEARCH_REPAIR_MAX_FETCH_KEYS or 0)

    candidate_keys = {candidate_key_from_tuple(c) for c in candidates} if candidates else set()

    if not candidate_keys:
        return set()

    history_info = _history_latest_and_net_by_candidate(history_cache_df, candidate_keys=candidate_keys)
    priority_rows = []

    direct_or_expanded_keys = (PRESCAN_REFRESH_KEYS | PRESCAN_MISSING_FETCH_KEYS) & candidate_keys
    discovery_keys = MONEYDJ_SEARCH_REPAIR_DISCOVERY_FETCH_KEYS & candidate_keys

    for key in direct_or_expanded_keys:
        info = history_info.get(key, {})
        latest_dt = info.get("latest_dt")
        latest_ord = latest_dt.toordinal() if latest_dt else 0
        priority_rows.append((0, -latest_ord, -int(info.get("latest_total_amount", 0) or 0), key, "API4直接/活動標的候選"))

    # API4 停在前一日時，這批 key 來自「歷史近期有動作 / 尚有庫存標的」展開。
    # 即使某檔權證以前完全沒有快取，也要強制打 API5，才能補到今天新買超。
    if MONEYDJ_SEARCH_REPAIR_FORCE_HISTORY_UNDERLYING_FETCH:
        for key in discovery_keys:
            info = history_info.get(key, {})
            latest_dt = info.get("latest_dt")
            latest_ord = latest_dt.toordinal() if latest_dt else 0
            priority_rows.append((1, -latest_ord, -int(info.get("latest_total_amount", 0) or 0), key, "歷史活動標的展開候選"))

    for key, info in history_info.items():
        latest_dt = info.get("latest_dt")
        if not latest_dt:
            continue

        # 若快取已經有報表目標日，不需要補漏。
        if latest_dt.date() >= target_dt.date():
            continue

        if MONEYDJ_SEARCH_REPAIR_INCLUDE_OPEN_POSITION and info.get("has_open_position"):
            priority_rows.append((2, -latest_dt.toordinal(), -abs(int(info.get("net_shares", 0) or 0)), key, "尚有淨庫存"))
            continue

        if recent_days > 0 and latest_dt.date() >= recent_floor_dt.date():
            priority_rows.append((3, -latest_dt.toordinal(), -int(info.get("latest_total_amount", 0) or 0), key, "近期有歷史活動"))

    if MONEYDJ_SEARCH_REPAIR_FULL_POOL_FORCE_FETCH:
        full_pool_keys = sorted(candidate_keys, key=lambda x: (str(x[1]), str(x[0])))
        full_pool_max = max(int(MONEYDJ_SEARCH_REPAIR_FULL_POOL_MAX_FETCH or 0), 0)
        if full_pool_max > 0:
            full_pool_keys = full_pool_keys[:full_pool_max]
        for key in full_pool_keys:
            info = history_info.get(key, {})
            latest_dt = info.get("latest_dt")
            latest_ord = latest_dt.toordinal() if latest_dt else 0
            priority_rows.append((4, -latest_ord, -int(info.get("latest_total_amount", 0) or 0), key, "全候選池強制掃描"))

    priority_rows = sorted(priority_rows, key=lambda x: (x[0], x[1], x[2], str(x[3])))

    selected = []
    seen = set()
    reason_count = {}

    for _, _, _, key, reason in priority_rows:
        if key in seen:
            continue
        seen.add(key)
        selected.append(key)
        reason_count[reason] = reason_count.get(reason, 0) + 1
        if max_fetch > 0 and len(selected) >= max_fetch:
            break

    MONEYDJ_SEARCH_REPAIR_FETCH_KEYS = set(selected)

    reason_text = "，".join(f"{k}:{v:,}" for k, v in reason_count.items()) or "-"
    cap_text = f"｜上限 {max_fetch:,} 組" if max_fetch > 0 else ""
    print(
        f"  🔎 MoneyDJ Search 補漏候選：本次候選 {len(candidate_keys):,} 組｜"
        f"需強制補抓 {len(MONEYDJ_SEARCH_REPAIR_FETCH_KEYS):,} 組{cap_text}｜{reason_text}"
    )

    if MONEYDJ_SEARCH_REPAIR_FETCH_KEYS:
        print(
            "  ✅ MoneyDJ Search 補漏模式已套用：這些候選會略過一般增量判斷，強制重抓 API5。"
        )
    else:
        print(
            "  ⚠️ MoneyDJ Search 補漏模式已啟用，但目前沒有符合條件的補漏候選；"
            "流程會照一般增量資料繼續。"
        )

    return MONEYDJ_SEARCH_REPAIR_FETCH_KEYS


def build_price_prefetch_context_from_items(items):
    """
    用既有分點歷史快取重建正式流程會用到的事件，僅供價格預抓使用。

    注意：這裡不寫入任何結果工作表，只重用正式流程的 A/B/C/D/E 金額強度分類，
    讓價格預抓覆蓋正式流程真正會用到的標的股價格、TOP15 剩餘部位權證價格，
    以及 RUN_MODE=2 的近10日分點明細權證價格。
    """
    item_map = {}

    for item in items:
        item_map[(item["broker_code"], item["warrant_code"])] = item

    daily_records = build_daily_records(items)
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [
        amount_events.get(code, [])
        for code in AMOUNT_CLASS_CODES
    ]

    return item_map, a_events, b_events, c_events, d_events, e_events

def run_auto_price_prefetch_from_history(history_cache_df, candidate_keys=None, target_date=None, reason="今日分點資料尚未出現"):
    """
    當 API4 預掃描沒有看到今日目標分點資料時，自動改跑價格預抓。

    這個模式只會：
    1. 讀既有分點歷史快取
    2. 重建正式流程會用到的事件集合
    3. 呼叫所有正式流程需要抓價格的函式
    4. 更新 price_cache.csv / 快取_價格
    5. 寫入 price_prefetch_state.csv，避免同一天重複慢抓

    不會建立 Excel、不會同步 A/B/C/D/E、TOP15、近10日分點明細結果，因此不會把尚未完整的當日分點資料寫到結果表。
    """
    target_date = normalize_price_prefetch_target_date(target_date)
    start_ts = time.time()
    scope = get_result_data_scope()

    if history_cache_df is None or history_cache_df.empty:
        print("  ⚠️ 尚無原始分點歷史快取，無法進行價格預抓。")
        write_price_prefetch_state({
            "日期": target_date,
            "資料範圍": scope,
            "模式": "AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING",
            "狀態": "no_items",
            "原因": "尚無原始分點歷史快取",
            "候選組合數": len(candidate_keys or []),
            "快取歷史筆數": 0,
            "還原項目數": 0,
            "執行時間": f"{time.time() - start_ts:.2f}",
        })
        return True

    if candidate_keys:
        items = items_from_history_cache(history_cache_df, candidate_filter=candidate_keys)
    else:
        items = items_from_history_cache(history_cache_df)

    if not items:
        print("  ⚠️ 原始分點歷史快取中沒有可還原的候選項目，無法進行價格預抓。")
        write_price_prefetch_state({
            "日期": target_date,
            "資料範圍": scope,
            "模式": "AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING",
            "狀態": "no_items",
            "原因": "無可還原項目",
            "候選組合數": len(candidate_keys or []),
            "快取歷史筆數": len(history_cache_df),
            "還原項目數": 0,
            "執行時間": f"{time.time() - start_ts:.2f}",
        })
        return True

    print(f"【價格預抓】使用既有分點歷史快取還原 {len(items):,} 組資料，只更新價格快取，不建立結果報表。")

    item_map, a_events, b_events, c_events, d_events, e_events = build_price_prefetch_context_from_items(items)

    print(
        f"  ✅ 價格預抓事件重建完成："
        f"A:{len(a_events):,}｜B:{len(b_events):,}｜C:{len(c_events):,}｜D:{len(d_events):,}｜E:{len(e_events):,}"
    )

    if a_events or b_events or c_events or d_events or e_events:
        price_cache = fetch_all_prices(a_events, b_events, c_events, d_events, e_events)
        # TOP15 固定資料集會在內部補抓仍有剩餘部位的權證最新價。
        build_top15_position_detail_and_consensus_rows(
            a_events,
            b_events,
            c_events,
            d_events,
            e_events,
            item_map,
            price_cache,
            target_date=target_date,
        )
    else:
        price_cache = load_price_cache()
        print("  ⚠️ 快取資料目前無 A/B/C/D/E 事件，略過事件價格預抓，只檢查其他價格需求。")

    # 不論 RUN_MODE 為何，價格預抓都先補近10日分點明細會用到的標的股價格。
    # 這是為了讓盤後現股收盤價可以先進 price_cache，避免之後分點資料更新時「現股10日」仍停在前一交易日。
    price_cache = ensure_broker_10d_underlying_prices(price_cache, items, target_date=target_date)

    if RUN_MODE == 2:
        price_cache = ensure_broker_10d_warrant_prices(price_cache, items, target_date=target_date)
    else:
        print("  ✅ RUN_MODE=1：近10日分點明細工作表不更新，因此價格預抓略過該表權證價。")

    # 完整價格預抓：不只補事件 / TOP15 / 近10日圖卡目前會用到的價格，
    # 也把既有分點歷史快取中所有權證與標的股的最新價格先補進 price_cache。
    price_cache = ensure_price_prefetch_all_item_prices(price_cache, items, target_date=target_date)

    price_summary = price_cache_target_date_summary(price_cache, target_date)
    elapsed = time.time() - start_ts
    write_price_prefetch_state({
        "日期": target_date,
        "資料範圍": scope,
        "模式": "AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING",
        "狀態": "done",
        "原因": reason,
        "候選組合數": len(candidate_keys or []),
        "快取歷史筆數": len(history_cache_df),
        "還原項目數": len(items),
        "A事件數": len(a_events),
        "B事件數": len(b_events),
        "C事件數": len(c_events),
        "D事件數": len(d_events),
        "價格快取代號數": len(price_cache or {}),
        "價格最新日期": price_summary.get("latest_date", ""),
        "目標日價格代號數": price_summary.get("target_date_code_count", 0),
        "執行時間": f"{elapsed:.2f}",
    })

    print(f"  ✅ 價格預抓完成，已記錄今日狀態，下次今日分點資料仍未出現時會快速略過。耗時 {elapsed:.2f} 秒")
    return True


def maybe_auto_price_prefetch_before_api5(candidates, program_start):
    """
    在正式 API5 大量更新前判斷是否要改跑價格預抓。

    判斷依據使用剛剛 API4 預掃描結果：
    - 若 API4 已看到今日目標分點資料 → 正常跑正式流程。
    - 若 API4 尚未看到今日目標分點資料 → 不產生報表，改成價格預抓。
    - 若今日已經預抓過且 API4 仍未看到今日資料 → 快速結束。
    """
    if not AUTO_PRICE_PREFETCH_WHEN_BROKER_DATA_MISSING:
        return False

    target_date = normalize_price_prefetch_target_date()

    if moneydj_search_repair_is_active():
        print(
            "  ✅ 已啟用 MoneyDJ Search 補漏模式：略過價格預抓快速結束，"
            "繼續進入 API5 補漏與正式報表流程。"
        )
        return False

    if has_today_broker_data_from_prescan(target_date):
        return False

    latest_label = PRESCAN_LATEST_ACTIVITY_DATE.strftime("%Y/%m/%d") if PRESCAN_LATEST_ACTIVITY_DATE else "-"
    print(
        f"  ⚠️ API4 尚未看到 {target_date} 的目標分點資料｜"
        f"目前最新活動日：{latest_label}。"
    )

    candidate_keys = {candidate_key_from_tuple(c) for c in candidates} if candidates else None
    history_cache_df = None

    if PRICE_PREFETCH_SKIP_IF_DONE_TODAY and is_price_prefetch_done_for_today(target_date):
        if PRICE_PREFETCH_RETRY_UNTIL_TARGET_PRICE:
            history_cache_df = load_history_cache()
            has_required_10d_underlying_prices, price_summary = price_cache_has_required_10d_underlying_target_prices(
                history_cache_df,
                candidate_keys=candidate_keys,
                target_date=target_date,
            )

            print(
                f"  🔎 檢查近10日現股價格：需 {price_summary.get('required_underlying_count', 0):,} 檔｜"
                f"已有目標日 {price_summary.get('target_date_underlying_count', 0):,} 檔｜"
                f"缺 {price_summary.get('missing_underlying_count', 0):,} 檔"
            )

            if has_required_10d_underlying_prices:
                print(
                    "  ✅ 今日已完成價格預抓，且近10日現股所需標的股已有目標日收盤價；"
                    "分點資料仍未出現，略過正式報表與價格重抓，快速結束。"
                )
                write_prescan_status(
                    "broker_data_not_ready",
                    target_date=target_date,
                    reason="分點資料尚未出現；價格預抓已完成，快速結束",
                    candidate_count=len(candidates or []),
                    candidate_cache_saved=False,
                )
                elapsed = time.time() - program_start
                print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
                return True

            missing_sample = str(price_summary.get("missing_sample", "") or "").strip()
            if missing_sample:
                missing_sample = f"｜缺少樣本：{missing_sample}"

            print(
                f"  🔄 今日曾完成價格預抓，但近10日現股所需標的股尚未全數取得 {target_date} 收盤價 "
                f"（價格最新日期：{price_summary.get('latest_date', '-') or '-'}｜"
                f"缺少標的數：{price_summary.get('missing_underlying_count', 0)}{missing_sample}），"
                "本次再嘗試預抓盤後最新價格。"
            )
        else:
            print("  ✅ 今日已完成價格預抓，且分點資料仍未出現；略過正式報表與價格重抓，快速結束。")
            write_prescan_status(
                "broker_data_not_ready",
                target_date=target_date,
                reason="分點資料尚未出現；價格預抓已完成，快速結束",
                candidate_count=len(candidates or []),
                candidate_cache_saved=False,
            )
            elapsed = time.time() - program_start
            print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
            return True

    print("  🔄 自動進入價格預抓模式：只更新 price_cache，不寫入今日結果表。")
    if history_cache_df is None:
        history_cache_df = load_history_cache()
    run_auto_price_prefetch_from_history(
        history_cache_df,
        candidate_keys=candidate_keys,
        target_date=target_date,
        reason="API4 尚未看到今日目標分點資料",
    )
    write_prescan_status(
        "broker_data_not_ready",
        target_date=target_date,
        reason="API4 尚未看到今日目標分點資料；已完成價格預抓",
        candidate_count=len(candidates or []),
        candidate_cache_saved=False,
    )
    elapsed = time.time() - program_start
    print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
    return True


def run_automatic_cache_maintenance(warrants=None):
    """同步清理本機 CSV、Google Sheet 與 Supabase 的超期／失效快取。"""
    if not CACHE_AUTO_PRUNE_ENABLED or not USE_CACHE:
        return

    print("\n【快取維護】自動清理超過保留天數、已移除分點與失效權證...")

    if supabase_enabled() and SUPABASE_AUTO_PRUNE_ENABLED:
        prune_supabase_cache_tables()

        # Supabase 已是大型快取唯一來源時，不再為了清理而額外把所有大型表下載一次。
        # 後續正式流程讀取快取時，會自然把已清理資料寫入 runner 內本機 CSV。
        if SUPABASE_CACHE_SKIP_GSHEET_SYNC:
            print("  ✅ Supabase 大型快取已直接在資料庫端清理；略過額外全表下載與 Google Sheet 重寫。")
            return

    # 價格快取：最近 PRICE_RETENTION_TRADING_DAYS 個交易日。
    price_df = read_cache_csv(PRICE_CACHE_PATH)
    if price_df is not None and not price_df.empty:
        price_clean, stats = prune_price_cache_dataframe(price_df)
        if stats.get("removed", 0) > 0:
            write_cache_csv(
                price_clean,
                PRICE_CACHE_PATH,
                sync_supabase=False,
                force_gsheet_sync=CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC,
            )
            print(
                f"  🧹 價格快取已持久化清理：{stats['before']:,} → {stats['after']:,} 筆｜"
                f"起始日 {stats.get('cutoff') or '-'}"
            )

    # 分點歷史：完整分點清單 + 最近 HISTORY_RETENTION_TRADING_DAYS 個交易日。
    history_df = read_cache_csv(HISTORY_CACHE_PATH)
    if history_df is not None and not history_df.empty:
        history_clean, stats = prune_history_cache_dataframe(history_df)
        if stats.get("removed", 0) > 0:
            write_cache_csv(
                history_clean,
                HISTORY_CACHE_PATH,
                sync_supabase=False,
                force_gsheet_sync=CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC,
            )
            print(
                f"  🧹 分點歷史已持久化清理：{stats['before']:,} → {stats['after']:,} 筆｜"
                f"起始日 {stats.get('cutoff') or '-'}"
            )

    # 候選組合：兩種 scope 都移除不在程式碼中的分點與已失效權證。
    valid_codes = CURRENT_LIVE_WARRANT_CODES if LIVE_WARRANT_SNAPSHOT_READY else None
    for candidate_path in (CANDIDATES_CACHE_ALL_PATH, CANDIDATES_CACHE_SELECTED5_PATH):
        candidate_df = read_cache_csv(candidate_path)
        if candidate_df is None or candidate_df.empty:
            continue
        candidate_clean, stats = prune_candidates_dataframe(
            candidate_df,
            path=candidate_path,
            valid_warrant_codes=valid_codes,
        )
        if stats.get("removed", 0) > 0:
            write_cache_csv(
                candidate_clean,
                candidate_path,
                sync_supabase=False,
                force_gsheet_sync=CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC,
            )
            print(
                f"  🧹 {cache_sheet_name_from_path(candidate_path)} 已持久化清理："
                f"{stats['before']:,} → {stats['after']:,} 組"
            )

    # 目前執行 scope 的分點代號快取。
    broker_df = read_cache_csv(BROKER_MAP_CACHE_PATH)
    if broker_df is not None and not broker_df.empty:
        scope = "selected5" if RUN_MODE == 1 else "all"
        broker_clean, stats = prune_broker_map_dataframe(broker_df, scope=scope)
        if stats.get("removed", 0) > 0:
            write_cache_csv(
                broker_clean,
                BROKER_MAP_CACHE_PATH,
                sync_supabase=False,
                force_gsheet_sync=CACHE_PRUNE_FORCE_GSHEET_FULL_SYNC,
            )
            print(
                f"  🧹 分點代號快取已持久化清理：{stats['before']:,} → {stats['after']:,} 筆"
            )

    print("  ✅ 快取自動維護完成")



# ══════════════════════════════════════════════════════════════════════
# WORKFLOW_MODE=longterm：長期留單補價格與修正勝率
# ══════════════════════════════════════════════════════════════════════

def longterm_safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return default
        return float(s)
    except Exception:
        return default


def longterm_target_date_from_history(history_df):
    raw = str(LONGTERM_TARGET_DATE or PRICE_PREFETCH_TARGET_DATE or TOP15_TARGET_DATE or "").strip()
    if raw:
        dt = parse_date(raw)
        if dt:
            return dt.strftime("%Y/%m/%d")

    if history_df is not None and not history_df.empty and "日期" in history_df.columns:
        df = history_df.copy().fillna("")
        # 優先使用有實際買賣活動的最新日期。
        # 若假日或資料源異常產生 0 買 0 賣的日期列，不應拿它當 longterm 統計日期。
        activity_cols = [c for c in ["買進股數", "賣出股數", "買進金額", "賣出金額"] if c in df.columns]
        if activity_cols:
            activity_sum = pd.Series(0.0, index=df.index)
            for col in activity_cols:
                # Google Sheet 讀回來的數字可能是 "1,234" 字串，
                # 不能直接 pd.to_numeric，否則會變 NaN，導致誤判沒有活動。
                activity_values = df[col].map(longterm_safe_float)
                activity_sum = activity_sum + activity_values.fillna(0).abs()
            active_df = df[activity_sum > 0].copy()
        else:
            active_df = df

        dates = []
        for value in active_df["日期"].tolist():
            dt = parse_date(value)
            if dt:
                dates.append(dt)
        if dates:
            return max(dates).strftime("%Y/%m/%d")

    return datetime.today().strftime("%Y/%m/%d")


def longterm_history_range_info(history_df):
    if history_df is None or history_df.empty or "日期" not in history_df.columns:
        return "", "", 0

    dates = []
    for value in history_df["日期"].tolist():
        dt = parse_date(value)
        if dt:
            dates.append(dt)

    if not dates:
        return "", "", 0

    return min(dates).strftime("%Y/%m/%d"), max(dates).strftime("%Y/%m/%d"), len(set(d.date() for d in dates))



def longterm_event_class_from_amount(total_buy_amount, max_single_warrant_buy_amount):
    """依 ABCDE 規則判斷 longterm 留單的原始事件類型。"""
    try:
        total_buy_amount = float(total_buy_amount or 0)
        max_single_warrant_buy_amount = float(max_single_warrant_buy_amount or 0)
    except Exception:
        return ""

    if max_single_warrant_buy_amount < AMOUNT_THRESH:
        return ""

    for code, _label, low, high in AMOUNT_CLASS_SPECS:
        if total_buy_amount >= low and (high is None or total_buy_amount < high):
            return code
    return ""


def build_longterm_event_code_map(df):
    """
    從快取_分點歷史重建「同一分點 × 同一標的 × 同一天」的 ABCDE 事件分類。

    longterm 的未出清留單是以權證買進批次為單位；但修正勝率要對應原本勝率統計的
    A/B/C/D/E 分類，因此這裡先把買進日當天同分點、同標的的買進金額合併分類，
    再把每個未出清買進批次標上事件代碼。
    """
    if df is None or df.empty:
        return {}

    required = {"分點", "券商代號", "標的股", "日期", "買進股數", "買進金額"}
    if not required.issubset(set(df.columns)):
        return {}

    work = df.copy().fillna("")
    work["標的股_norm"] = work["標的股"].map(normalize_underlying_code_for_group)
    work["買進股數_num"] = work["買進股數"].map(longterm_safe_float).fillna(0.0)
    work["買進金額_num"] = work["買進金額"].map(longterm_safe_float).fillna(0.0)
    work = work[(work["買進股數_num"] > 0) & (work["買進金額_num"] > 0)].copy()
    if work.empty:
        return {}

    event_map = {}
    group_cols = ["分點", "券商代號", "標的股_norm", "日期"]
    for key, g in work.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_code, underlying_code, trade_date = key
        total_amount = float(g["買進金額_num"].sum())
        max_amount = float(g["買進金額_num"].max())
        code = longterm_event_class_from_amount(total_amount, max_amount)
        if not code:
            continue
        event_map[(
            str(broker_label).strip(),
            str(broker_code).strip(),
            str(underlying_code).strip(),
            normalize_date_str(trade_date),
        )] = {
            "事件代碼": code,
            "事件類型": f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}",
            "事件總買進金額": total_amount,
            "事件最大單檔買進金額": max_amount,
        }
    return event_map

def rebuild_longterm_open_lots_from_history(history_df, target_date):
    """
    從快取_分點歷史直接用 FIFO 重建目前未出清部位。

    單位是「分點 × 權證」的實際買進批次，不依賴 A/B/C/D/E 事件，
    因此可以抓出所有長期未賣出的留單。賣出依日期 FIFO 扣買進股數與成本，
    同日買賣不互扣，避免把事件日內的買賣誤當出清。
    """
    if history_df is None or history_df.empty:
        return []

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()

    needed_cols = {
        "權證代號", "權證名稱", "標的股", "標的名稱", "分點", "分點名稱", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
    }
    if not needed_cols.issubset(set(history_df.columns)):
        missing = sorted(needed_cols - set(history_df.columns))
        print(f"  ⚠️ 長期留單分析失敗：快取_分點歷史缺少欄位：{missing}")
        return []

    df = history_df.copy().fillna("")
    df = fix_known_underlying_info_dataframe(df, "權證名稱", "標的股", "標的名稱")
    df["日期"] = df["日期"].map(normalize_date_str)
    # 不使用底線開頭欄名，例如 _dt。pandas itertuples() 會把底線欄名改名，
    # 導致 row._asdict().get("_dt") 取不到值，最後所有列都被跳過。
    df["dt_parsed"] = pd.to_datetime(df["日期"].astype(str).str.replace("/", "-", regex=False), errors="coerce")
    df = df[df["dt_parsed"].notna()].copy()
    df = df[df["dt_parsed"].dt.date <= target_dt.date()].copy()

    if df.empty:
        return []

    for col in ["買進股數", "賣出股數", "買進金額", "賣出金額"]:
        # Google Sheet 讀回可能含千分位逗號，直接 pd.to_numeric 會變 NaN。
        df[col] = df[col].map(longterm_safe_float).fillna(0.0)

    event_code_map = build_longterm_event_code_map(df)

    open_lots = []
    group_cols = ["分點", "分點名稱", "券商代號", "權證代號", "權證名稱", "標的股", "標的名稱"]

    for key, g in df.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_name, broker_code, warrant_code, warrant_name, underlying_code, underlying_name = key
        lots = []
        g = g.sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

        for rd in g.to_dict("records"):
            trade_dt = rd.get("dt_parsed")
            if pd.isna(trade_dt):
                continue
            trade_date = trade_dt.strftime("%Y/%m/%d")

            buy_qty = float(rd.get("買進股數", 0) or 0)
            buy_amount = float(rd.get("買進金額", 0) or 0)
            if buy_qty > 0 and buy_amount > 0:
                underlying_norm = normalize_underlying_code_for_group(underlying_code)
                event_info = event_code_map.get((
                    str(broker_label).strip(),
                    str(broker_code).strip(),
                    str(underlying_norm).strip(),
                    trade_date,
                ), {})
                lots.append({
                    "分點": str(broker_label).strip(),
                    "分點名稱": str(broker_name).strip(),
                    "券商代號": str(broker_code).strip(),
                    "權證代號": str(warrant_code).strip(),
                    "權證名稱": str(warrant_name).strip(),
                    "標的股": underlying_norm,
                    "標的名稱": str(underlying_name).strip(),
                    "事件代碼": event_info.get("事件代碼", ""),
                    "事件類型": event_info.get("事件類型", "未達ABCDE門檻"),
                    "事件總買進金額": event_info.get("事件總買進金額", ""),
                    "事件最大單檔買進金額": event_info.get("事件最大單檔買進金額", ""),
                    "買進日": trade_date,
                    "買進股數": buy_qty,
                    "買進金額": buy_amount,
                    "剩餘股數": buy_qty,
                    "剩餘成本": buy_amount,
                    "第一筆日期": trade_date,
                    "最後筆日期": trade_date,
                })

            sell_qty_left = float(rd.get("賣出股數", 0) or 0)
            if sell_qty_left <= 0:
                continue

            for lot in lots:
                if sell_qty_left <= 0:
                    break
                buy_dt = parse_date(lot.get("買進日"))
                trade_day = trade_dt.to_pydatetime().date() if hasattr(trade_dt, "to_pydatetime") else trade_dt.date()
                if not buy_dt or trade_day <= buy_dt.date():
                    continue

                remain_qty = float(lot.get("剩餘股數", 0) or 0)
                remain_cost = float(lot.get("剩餘成本", 0) or 0)
                original_qty = float(lot.get("買進股數", 0) or 0)
                original_cost = float(lot.get("買進金額", 0) or 0)
                if remain_qty <= 0 or remain_cost <= 0 or original_qty <= 0 or original_cost <= 0:
                    continue

                avg_cost = original_cost / original_qty
                alloc_qty = min(sell_qty_left, remain_qty)
                alloc_cost = min(remain_cost, alloc_qty * avg_cost)
                lot["剩餘股數"] = max(remain_qty - alloc_qty, 0)
                lot["剩餘成本"] = max(remain_cost - alloc_cost, 0)
                lot["最後筆日期"] = trade_date
                sell_qty_left -= alloc_qty

        for lot in lots:
            remain_qty = float(lot.get("剩餘股數", 0) or 0)
            remain_cost = float(lot.get("剩餘成本", 0) or 0)
            if remain_qty <= 0 or remain_cost <= 0:
                continue
            buy_dt = parse_date(lot.get("買進日"))
            if not buy_dt:
                continue
            holding_days = max((target_dt.date() - buy_dt.date()).days, 0)
            lot["持有天數"] = holding_days
            lot["剩餘均價"] = remain_cost / remain_qty if remain_qty > 0 else None
            open_lots.append(lot)

    print(f"  ✅ FIFO 未出清部位重建完成：{len(open_lots):,} 筆")
    return open_lots


def ensure_longterm_warrant_prices(price_cache, open_lots, target_date):
    """只針對長期未出清權證補最新可用價格，並同步回 price_cache。"""
    if not open_lots:
        return price_cache

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()
    target_dt = min(target_dt, datetime.today())
    start_dt = target_dt - timedelta(days=max(int(LONGTERM_PRICE_LOOKBACK_DAYS or 420), 30))

    needed_codes = sorted({
        normalize_price_code(lot.get("權證代號", ""))
        for lot in open_lots
        if normalize_price_code(lot.get("權證代號", ""))
    })

    if not needed_codes:
        return price_cache

    persistent_price_cache = load_price_cache()
    fetch_plan = []

    for code in needed_codes:
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        current_prices = get_price_series_from_cache(price_cache, code)
        merged_prices = merge_price_dicts(cached_prices, current_prices)
        if merged_prices:
            add_price_aliases(price_cache, code, merged_prices)
            persistent_price_cache[normalize_price_code(code)] = merged_prices

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_dt.strftime("%Y/%m/%d"))
        latest_dt = parse_date(latest_date) if latest_date else None
        need_fetch = latest_price is None
        if latest_dt and (target_dt - latest_dt).days > LONGTERM_PRICE_STALE_DAYS:
            need_fetch = True
        if need_fetch:
            fetch_plan.append(code)

    print(f"  長期留單需檢查權證價格：{len(needed_codes):,} 檔")
    print(f"  長期留單需補抓權證價格：{len(fetch_plan):,} 檔")

    if fetch_plan:
        changed_price_codes = set()

        def fetch_one(code):
            return code, fetch_twse_prices(code, start_dt, target_dt)

        done = 0
        with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as ex:
            futures = {ex.submit(fetch_one, code): code for code in fetch_plan}
            for future in as_completed(futures):
                done += 1
                code = futures[future]
                try:
                    code, fetched_prices = future.result()
                except Exception:
                    fetched_prices = {}

                old_prices = get_cached_prices_for_code(persistent_price_cache, code)
                merged_prices = merge_price_dicts(old_prices, fetched_prices)
                if merged_prices:
                    norm_code = normalize_price_code(code)
                    if norm_code:
                        persistent_price_cache[norm_code] = merged_prices
                        if fetched_prices:
                            changed_price_codes.add(norm_code)
                    add_price_aliases(price_cache, code, merged_prices)

                if done % 50 == 0:
                    print(f"  [{done:,}/{len(fetch_plan):,}] 長期留單權證價格補抓中...")

        save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)

    return price_cache


def attach_longterm_price_and_pnl(open_lots, price_cache, target_date):
    rows = []
    target_dt = parse_date(target_date) or datetime.today()

    for lot in open_lots:
        code = normalize_price_code(lot.get("權證代號", ""))
        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_date)
        remain_qty = longterm_safe_float(lot.get("剩餘股數"))
        remain_cost = longterm_safe_float(lot.get("剩餘成本"))
        buy_dt = parse_date(lot.get("買進日"))
        holding_days = max((target_dt.date() - buy_dt.date()).days, 0) if buy_dt else int(lot.get("持有天數", 0) or 0)

        market_value = None
        pnl = None
        return_pct = None
        result = "缺價"
        price_status = "缺最新權證價格"

        if latest_price is not None and remain_qty > 0:
            market_value = remain_qty * float(latest_price)
            pnl = market_value - remain_cost
            return_pct = (pnl / remain_cost * 100.0) if remain_cost > 0 else None
            if float(latest_price) <= LONGTERM_ZERO_PRICE_THRESHOLD:
                result = "疑似龜苓膏"
                price_status = "價格接近歸零"
            elif pnl > 0:
                result = "賺錢"
                price_status = "有價格"
            elif pnl < 0:
                result = "賠錢"
                price_status = "有價格"
            else:
                result = "打平"
                price_status = "有價格"

        rows.append({
            "統計日期": target_date,
            "分點": lot.get("分點", ""),
            "分點名稱": lot.get("分點名稱", ""),
            "券商代號": lot.get("券商代號", ""),
            "事件代碼": lot.get("事件代碼", ""),
            "事件類型": lot.get("事件類型", ""),
            "事件總買進金額": lot.get("事件總買進金額", ""),
            "事件最大單檔買進金額": lot.get("事件最大單檔買進金額", ""),
            "標的股": lot.get("標的股", ""),
            "標的名稱": lot.get("標的名稱", ""),
            "權證代號": lot.get("權證代號", ""),
            "權證名稱": lot.get("權證名稱", ""),
            "買進日": lot.get("買進日", ""),
            "持有天數": holding_days,
            "買進股數": round(longterm_safe_float(lot.get("買進股數")), 0),
            "買進金額": round(longterm_safe_float(lot.get("買進金額")), 0),
            "剩餘股數": round(remain_qty, 0),
            "剩餘成本": round(remain_cost, 0),
            "剩餘均價": "" if remain_qty <= 0 else round(remain_cost / remain_qty, 4),
            "最新權證價格": "" if latest_price is None else latest_price,
            "最新價格日期": latest_date or "",
            "估算市值": "" if market_value is None else round(market_value, 0),
            "估算損益": "" if pnl is None else round(pnl, 0),
            "估算報酬%": "" if return_pct is None else round(return_pct, 2),
            "結果": result,
            "價格狀態": price_status,
        })

    return rows


def filter_longterm_rows_by_days(rows, days):
    return [row for row in rows if int(longterm_safe_float(row.get("持有天數"), 0)) >= int(days)]


def summarize_longterm_rows(rows, days):
    if not rows:
        return []

    df = pd.DataFrame(rows).fillna("")
    numeric_cols = ["剩餘成本", "估算市值", "估算損益", "持有天數"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    summary_rows = []
    for broker, g in df.groupby("分點", dropna=False, sort=False):
        total = len(g)
        priced = int((g["結果"].astype(str) != "缺價").sum())
        missing = int((g["結果"].astype(str) == "缺價").sum())
        win = int((g["結果"].astype(str) == "賺錢").sum())
        loss = int((g["結果"].astype(str) == "賠錢").sum())
        flat = int((g["結果"].astype(str) == "打平").sum())
        zero = int((g["結果"].astype(str) == "疑似龜苓膏").sum())
        known_risk = loss + zero
        total_cost = float(g["剩餘成本"].fillna(0).sum()) if "剩餘成本" in g.columns else 0
        missing_cost = float(g.loc[g["結果"].astype(str) == "缺價", "剩餘成本"].fillna(0).sum()) if "剩餘成本" in g.columns else 0
        pnl_sum = float(g["估算損益"].fillna(0).sum()) if "估算損益" in g.columns else 0
        avg_days = float(g["持有天數"].dropna().mean()) if "持有天數" in g.columns and not g["持有天數"].dropna().empty else 0

        summary_rows.append({
            "門檻天數": days,
            "分點": broker,
            "留單筆數": total,
            "有價格筆數": priced,
            "缺價筆數": missing,
            "賺錢筆數": win,
            "賠錢筆數": loss,
            "打平筆數": flat,
            "疑似龜苓膏筆數": zero,
            "已知風險筆數": known_risk,
            "有價格賠錢比例": "" if priced <= 0 else round(known_risk / priced * 100, 2),
            "缺價比例": round(missing / total * 100, 2) if total else 0,
            "留單成本": round(total_cost, 0),
            "缺價成本": round(missing_cost, 0),
            "缺價成本占比": "" if total_cost <= 0 else round(missing_cost / total_cost * 100, 2),
            "已估損益合計": round(pnl_sum, 0),
            "平均持有天數": round(avg_days, 2) if avg_days else "",
        })

    summary_rows.sort(key=lambda r: (-(longterm_safe_float(r.get("留單成本"))), str(r.get("分點", ""))))
    return summary_rows


def build_longterm_overview(threshold_rows_map, history_df, target_date):
    hist_start, hist_end, hist_trade_days = longterm_history_range_info(history_df)
    overview = []
    for days in LONGTERM_OPEN_DAYS:
        rows = threshold_rows_map.get(days, [])
        total = len(rows)
        priced = sum(1 for r in rows if str(r.get("結果", "")) != "缺價")
        missing = sum(1 for r in rows if str(r.get("結果", "")) == "缺價")
        win = sum(1 for r in rows if str(r.get("結果", "")) == "賺錢")
        loss = sum(1 for r in rows if str(r.get("結果", "")) == "賠錢")
        zero = sum(1 for r in rows if str(r.get("結果", "")) == "疑似龜苓膏")
        flat = sum(1 for r in rows if str(r.get("結果", "")) == "打平")
        total_cost = sum(longterm_safe_float(r.get("剩餘成本")) for r in rows)
        missing_cost = sum(longterm_safe_float(r.get("剩餘成本")) for r in rows if str(r.get("結果", "")) == "缺價")
        overview.append({
            "統計日期": target_date,
            "門檻天數": days,
            "留單筆數": total,
            "有價格筆數": priced,
            "缺價筆數": missing,
            "賺錢筆數": win,
            "賠錢筆數": loss,
            "打平筆數": flat,
            "疑似龜苓膏筆數": zero,
            "有價格賠錢比例": "" if priced <= 0 else round((loss + zero) / priced * 100, 2),
            "缺價比例": "" if total <= 0 else round(missing / total * 100, 2),
            "留單成本": round(total_cost, 0),
            "缺價成本": round(missing_cost, 0),
            "缺價成本占比": "" if total_cost <= 0 else round(missing_cost / total_cost * 100, 2),
            "歷史快取起始日": hist_start,
            "歷史快取最後日": hist_end,
            "歷史快取交易日數": hist_trade_days,
            "HISTORY_RETENTION_TRADING_DAYS": HISTORY_RETENTION_TRADING_DAYS,
        })
    return overview


def build_longterm_risk_summary_by_broker_event(rows):
    """把 longterm 明細整理成 分點 × A/B/C/D/E/ALL 的風險筆數。"""
    empty = {}
    if not rows:
        return empty

    df = pd.DataFrame(rows).fillna("")
    if df.empty or "分點" not in df.columns:
        return empty

    df["事件代碼"] = df.get("事件代碼", "").astype(str).str.strip()
    df = df[df["事件代碼"].isin(AMOUNT_CLASS_CODES)].copy()
    if df.empty:
        return empty

    out = {}

    def summarize_group(g):
        total = len(g)
        result = g["結果"].astype(str) if "結果" in g.columns else pd.Series([], dtype=str)
        loss = int((result == "賠錢").sum())
        zero = int((result == "疑似龜苓膏").sum())
        missing = int((result == "缺價").sum())
        win = int((result == "賺錢").sum())
        flat = int((result == "打平").sum())
        cost = sum(longterm_safe_float(v) for v in g.get("剩餘成本", []))
        missing_cost = sum(longterm_safe_float(v) for v in g.loc[result == "缺價", "剩餘成本"]) if "剩餘成本" in g.columns and len(result) else 0
        return {
            "長期留單筆數": int(total),
            "長期賺錢筆數": int(win),
            "長期賠錢筆數": int(loss),
            "長期打平筆數": int(flat),
            "長期疑似龜苓膏筆數": int(zero),
            "長期已知虧損/龜苓膏筆數": int(loss + zero),
            "長期缺價筆數": int(missing),
            "長期留單成本": round(float(cost), 0),
            "長期缺價成本": round(float(missing_cost), 0),
        }

    for (broker, code), g in df.groupby(["分點", "事件代碼"], dropna=False, sort=False):
        out[(str(broker).strip(), str(code).strip())] = summarize_group(g)

    for broker, g in df.groupby("分點", dropna=False, sort=False):
        out[(str(broker).strip(), "ALL")] = summarize_group(g)

    return out


def build_longterm_adjusted_winrate_rows(summary_map, threshold_rows_map):
    """
    產生可放進「勝率統計」的長期留單修正勝率。

    這裡會依 分點 × A/B/C/D/E/ALL 計算：
    - 原始勝率：原本已出清勝率。
    - 修正勝率：把長期留單中已知虧損與疑似龜苓膏視為額外敗筆。
    - 保守勝率：再把長期缺價留單也視為風險敗筆。
    """
    rows = []
    brokers = list(TARGET_PATTERNS.keys())
    for broker in summary_map.keys():
        if broker not in brokers:
            brokers.append(broker)

    event_type_names = {
        code: f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}"
        for code in AMOUNT_CLASS_CODES
    }
    event_type_names["ALL"] = "全部-A+B+C+D+E合併"

    for days in LONGTERM_OPEN_DAYS:
        risk_map = build_longterm_risk_summary_by_broker_event(threshold_rows_map.get(days, []))
        for broker in sorted(set(brokers)):
            for code in AMOUNT_CLASS_CODES + ["ALL"]:
                original = (summary_map.get(broker, {}) or {}).get(code, calc_empty_summary(broker, event_type_names.get(code, code)))
                risk = risk_map.get((broker, code), {})
                closed_count = int(original.get("已出清筆數") or 0)
                win_count = int(original.get("勝筆數") or 0)
                loss_count = int(original.get("敗筆數") or 0)
                flat_count = int(original.get("平手筆數") or 0)
                long_loss = int(longterm_safe_float(risk.get("長期已知虧損/龜苓膏筆數"), 0))
                long_missing = int(longterm_safe_float(risk.get("長期缺價筆數"), 0))
                revised_den = closed_count + long_loss
                conservative_den = closed_count + long_loss + long_missing
                rows.append({
                    "門檻天數": days,
                    "分點": broker,
                    "事件代碼": code,
                    "事件類型": event_type_names.get(code, code),
                    "原始事件數": int(original.get("事件數") or 0),
                    "原始已出清筆數": closed_count,
                    "原始未出清筆數": int(original.get("未出清筆數") or 0),
                    "原始勝筆數": win_count,
                    "原始敗筆數": loss_count,
                    "原始平手筆數": flat_count,
                    "原始勝率": original.get("勝率"),
                    "長期留單筆數": int(longterm_safe_float(risk.get("長期留單筆數"), 0)),
                    "長期賺錢筆數": int(longterm_safe_float(risk.get("長期賺錢筆數"), 0)),
                    "長期賠錢筆數": int(longterm_safe_float(risk.get("長期賠錢筆數"), 0)),
                    "長期疑似龜苓膏筆數": int(longterm_safe_float(risk.get("長期疑似龜苓膏筆數"), 0)),
                    "長期已知虧損/龜苓膏筆數": long_loss,
                    "長期缺價筆數": long_missing,
                    "長期留單成本": risk.get("長期留單成本", 0),
                    "長期缺價成本": risk.get("長期缺價成本", 0),
                    "修正勝率_已知虧損視為敗": "" if revised_den <= 0 else round(win_count / revised_den * 100, 2),
                    "保守勝率_已知虧損加缺價視為敗": "" if conservative_den <= 0 else round(win_count / conservative_den * 100, 2),
                    "說明": "修正勝率以同分點同事件類型的長期留單已知虧損/龜苓膏視為額外敗筆；保守勝率再把缺價留單也視為風險敗筆。",
                })
    return rows

def dataframe_to_sheet(wb, title, rows, freeze=True, max_rows=0):
    ws = wb.create_sheet(safe_worksheet_title(title))
    if max_rows and rows and len(rows) > max_rows:
        rows = rows[:max_rows]
    if not rows:
        ws.append(["無資料"])
        style_sheet(ws, [16])
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    if freeze:
        ws.freeze_panes = "A2"
    widths = []
    for h in headers:
        if "名稱" in h or "說明" in h:
            widths.append(28)
        elif "權證代號" in h or "券商代號" in h or "標的股" in h:
            widths.append(12)
        elif "日期" in h:
            widths.append(12)
        elif "金額" in h or "成本" in h or "市值" in h or "損益" in h:
            widths.append(16)
        else:
            widths.append(14)
    style_sheet(ws, widths)
    return ws


def build_longterm_workbook(output_path, overview_rows, threshold_summary_map, threshold_rows_map, adjusted_rows):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "長期留單總覽"
    dataframe_to_sheet(wb, "_tmp_remove", [])
    try:
        del wb["_tmp_remove"]
    except Exception:
        pass
    # 直接寫入 active sheet，避免 Workbook 預設空白表殘留。
    if overview_rows:
        headers = list(overview_rows[0].keys())
        ws0.append(headers)
        for row in overview_rows:
            ws0.append([row.get(h, "") for h in headers])
        ws0.freeze_panes = "A2"
        style_sheet(ws0, [14] * len(headers))
    else:
        ws0.append(["無資料"])
        style_sheet(ws0, [16])

    dataframe_to_sheet(wb, "長期留單修正勝率", adjusted_rows)

    for days in LONGTERM_OPEN_DAYS:
        dataframe_to_sheet(wb, f"長期留單分點彙總_{days}日", threshold_summary_map.get(days, []))
        detail_rows = threshold_rows_map.get(days, [])
        detail_rows = sorted(
            detail_rows,
            key=lambda r: (-(longterm_safe_float(r.get("剩餘成本"))), str(r.get("分點", "")), str(r.get("權證代號", ""))),
        )
        dataframe_to_sheet(
            wb,
            f"長期留單明細_{days}日",
            detail_rows,
            max_rows=max(int(LONGTERM_MAX_DETAIL_ROWS or 0), 0),
        )

    wb.save(output_path)



def _longterm_pct_text(value):
    if value is None or str(value).strip() == "":
        return "-"
    try:
        return f"{float(value):.2f}%"
    except Exception:
        return str(value).strip() or "-"


def _longterm_extract_event_code(event_type):
    s = str(event_type or "").strip()
    if not s:
        return ""
    if s.startswith("全部"):
        return "ALL"
    m = re.match(r"^([A-E])(?:[-－]|$)", s)
    if m:
        return m.group(1)
    return "ALL" if "全部" in s else s


def _longterm_build_inline_winrate_map(adjusted_rows):
    """
    取 LONGTERM_WINRATE_SHEET_DAYS 對應門檻的修正勝率，準備回填到勝率統計原表格。
    Google Sheet 只放一個「修正勝率」欄位；完整 120 / 150 / 180 日明細仍保留在 Excel artifact。
    """
    selected_days = int(LONGTERM_WINRATE_SHEET_DAYS or (LONGTERM_OPEN_DAYS[0] if LONGTERM_OPEN_DAYS else 120))
    rows = [r for r in (adjusted_rows or []) if int(longterm_safe_float(r.get("門檻天數"), -1)) == selected_days]
    if not rows and adjusted_rows:
        available_days = sorted({int(longterm_safe_float(r.get("門檻天數"), 0)) for r in adjusted_rows if longterm_safe_float(r.get("門檻天數"), 0) > 0})
        if available_days:
            selected_days = available_days[0]
            rows = [r for r in adjusted_rows if int(longterm_safe_float(r.get("門檻天數"), -1)) == selected_days]

    out = {}
    for rec in rows:
        broker = str(rec.get("分點", "")).strip()
        code = str(rec.get("事件代碼", "")).strip() or _longterm_extract_event_code(rec.get("事件類型", ""))
        if not broker or not code:
            continue
        out[(broker, code)] = _longterm_pct_text(rec.get("修正勝率_已知虧損視為敗"))
    return out, selected_days


def _longterm_find_winrate_header_info(values):
    """找出勝率統計內既有的勝率欄位，並判斷旁邊是否已有修正勝率欄。"""
    for row_idx, row in enumerate(values or [], start=1):
        cells = [str(x).strip() for x in row]
        if "分點" not in cells or "事件類型" not in cells or "勝率" not in cells:
            continue
        winrate_idx = cells.index("勝率")
        has_adjusted_next = (winrate_idx + 1 < len(cells) and str(cells[winrate_idx + 1]).strip().startswith("修正勝率"))
        return {
            "header_row": row_idx,
            "winrate_col_index0": winrate_idx,
            "adjusted_col_index0": winrate_idx + 1,
            "has_adjusted_next": has_adjusted_next,
        }
    return None


def _longterm_clear_old_bottom_winrate_block(ws, values):
    """清掉舊版曾寫在勝率統計底部的長期留單修正勝率區塊。"""
    marker = "長期留單修正勝率"
    marker_row = None
    for idx, row in enumerate(values or [], start=1):
        first = str(row[0]).strip() if row else ""
        if first.startswith(marker):
            marker_row = idx
            break
    if marker_row is None:
        return False
    try:
        current_row_count = int(getattr(ws, "row_count", len(values)) or len(values))
        clear_range = f"A{marker_row}:Z{current_row_count}"
        gsheet_api_call("清除舊版勝率統計底部長期修正勝率區塊", ws.batch_clear, [clear_range])
        print(f"  🧹 已清除舊版勝率統計底部長期修正勝率區塊：第 {marker_row} 列起")
        return True
    except Exception as e:
        print(f"  ⚠️ 清除舊版勝率統計底部長期修正勝率區塊失敗：{type(e).__name__}: {e}")
        return False


def upload_longterm_adjusted_winrate_to_gsheet(adjusted_rows, target_date):
    """
    longterm 模式只把修正勝率回填到既有「勝率統計」表格的「勝率」旁邊。

    不新增長期留單明細 / 分點彙總工作表，完整明細只保留在 Excel artifact。
    若勝率旁邊尚未有修正勝率欄，第一次 longterm 會插入一欄；之後重跑只更新該欄。
    """
    if not LONGTERM_UPDATE_WINRATE_SHEET:
        print("  ✅ LONGTERM_UPDATE_WINRATE_SHEET=0，略過更新 Google Sheet 勝率統計。")
        return False
    if not GSHEET_RESULT_ENABLED or not gsheet_enabled():
        print("  ⚠️ 未設定 GCP_SERVICE_KEY，略過 Google Sheet 長期修正勝率同步。")
        return False

    title = "勝率統計"
    ws = get_or_create_worksheet(title, rows=200, cols=30)
    if ws is None:
        print("  ⚠️ 找不到或無法建立 Google Sheet：勝率統計，略過長期修正勝率同步。")
        return False

    try:
        existing_values = ws.get_all_values() or []
    except Exception:
        existing_values = []

    if not existing_values:
        print("  ⚠️ 勝率統計工作表目前沒有資料，無法回填修正勝率。")
        return False

    _longterm_clear_old_bottom_winrate_block(ws, existing_values)
    try:
        existing_values = ws.get_all_values() or []
    except Exception:
        existing_values = existing_values or []

    info = _longterm_find_winrate_header_info(existing_values)
    if not info:
        print("  ⚠️ 找不到勝率統計表頭中的『勝率』欄，無法回填修正勝率。")
        return False

    winrate_idx0 = int(info["winrate_col_index0"])
    adjusted_idx0 = int(info["adjusted_col_index0"])
    adjusted_col_1based = adjusted_idx0 + 1

    try:
        sheet_id = int(ws.id)
    except Exception:
        sheet_id = None

    if not info.get("has_adjusted_next"):
        if sheet_id is None:
            print("  ⚠️ 無法取得勝率統計 sheetId，無法插入修正勝率欄。")
            return False
        requests = [{
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": adjusted_idx0,
                    "endIndex": adjusted_idx0 + 1,
                },
                "inheritFromBefore": True,
            }
        }]
        _gsheet_batch_update(requests)
        print(f"  ✅ 已在勝率欄旁新增欄位：{LONGTERM_WINRATE_SHEET_HEADER}")
        try:
            existing_values = ws.get_all_values() or []
        except Exception:
            existing_values = existing_values or []
    else:
        print(f"  ✅ 勝率欄旁已存在修正勝率欄，直接更新：{LONGTERM_WINRATE_SHEET_HEADER}")

    winrate_map, selected_days = _longterm_build_inline_winrate_map(adjusted_rows)
    if not winrate_map:
        print("  ⚠️ 沒有可回填的長期修正勝率資料。")
        return False

    max_row = max(len(existing_values), 1)
    column_values = [[""] for _ in range(max_row)]
    updated_count = 0
    header_count = 0
    current_header = None

    for row_idx, row in enumerate(existing_values, start=1):
        cells = [str(x).strip() for x in row]
        if "分點" in cells and "事件類型" in cells and "勝率" in cells:
            # 每個分點區塊都有自己的表頭；同一欄都要顯示修正勝率。
            current_header = {h: i for i, h in enumerate(cells) if h}
            column_values[row_idx - 1] = [LONGTERM_WINRATE_SHEET_HEADER]
            header_count += 1
            continue

        if not current_header:
            continue

        broker_col = current_header.get("分點")
        event_col = current_header.get("事件類型")
        if broker_col is None or event_col is None:
            continue
        if broker_col >= len(cells) or event_col >= len(cells):
            continue

        broker = str(cells[broker_col]).strip()
        event_type = str(cells[event_col]).strip()
        if not broker or broker.startswith("分點：") or not event_type:
            continue
        if broker == "分點" or event_type == "事件類型":
            continue

        event_code = _longterm_extract_event_code(event_type)
        value = winrate_map.get((broker, event_code))
        if value is None:
            # 若某分點 / 類型沒有對應修正資料，維持空白，不改其他欄位。
            continue
        column_values[row_idx - 1] = [value]
        updated_count += 1

    col_letter = get_column_letter(adjusted_col_1based)
    try:
        gsheet_api_call(
            "回填勝率統計修正勝率欄",
            ws.update,
            values=column_values,
            range_name=f"{col_letter}1:{col_letter}{max_row}",
            value_input_option="USER_ENTERED",
        )

        # 只針對新增 / 更新的修正勝率欄做最小格式處理，不重套整張表。
        try:
            requests = [{
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": max_row,
                        "startColumnIndex": adjusted_idx0,
                        "endColumnIndex": adjusted_idx0 + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "horizontalAlignment": "CENTER",
                            "verticalAlignment": "MIDDLE",
                        }
                    },
                    "fields": "userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment",
                }
            }]
            _gsheet_batch_update(requests)
        except Exception:
            pass

        print(
            f"  ☁️ 已更新 Google Sheet：勝率統計｜回填 {LONGTERM_WINRATE_SHEET_HEADER} "
            f"{updated_count:,} 筆｜門檻 {selected_days} 日｜目標日期 {target_date}"
        )
        return True
    except Exception as e:
        print(f"  ⚠️ 回填 Google Sheet 勝率統計修正勝率失敗：{type(e).__name__}: {e}")
        return False


def run_longterm_workflow(warrants, broker_map, output_path, program_start):
    print("\n【Longterm】長期留單補價格與修正勝率分析...")
    history_df = load_history_cache()
    if history_df is None or history_df.empty:
        print("  ⚠️ 長期留單分析停止：快取_分點歷史為空。請先跑 daily 或 repair 建立歷史快取。")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    target_date = longterm_target_date_from_history(history_df)
    print(f"  ✅ 長期留單統計日期：{target_date}")
    hist_start, hist_end, hist_trade_days = longterm_history_range_info(history_df)
    print(f"  ✅ 歷史快取範圍：{hist_start or '-'} ～ {hist_end or '-'}｜交易日數 {hist_trade_days:,}｜保留設定 {HISTORY_RETENTION_TRADING_DAYS} 個交易日")

    open_lots = rebuild_longterm_open_lots_from_history(history_df, target_date)
    if not open_lots:
        print("  ⚠️ 長期留單分析停止：沒有重建出未出清部位。")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    min_days = min(LONGTERM_OPEN_DAYS)
    long_open_lots = [lot for lot in open_lots if int(lot.get("持有天數", 0) or 0) >= min_days]
    print(f"  ✅ 未出清買進批次：{len(open_lots):,} 筆｜{min_days} 日以上：{len(long_open_lots):,} 筆")

    price_cache = load_price_cache()
    price_cache = ensure_longterm_warrant_prices(price_cache, long_open_lots, target_date)
    priced_rows_all = attach_longterm_price_and_pnl(long_open_lots, price_cache, target_date)

    threshold_rows_map = {days: filter_longterm_rows_by_days(priced_rows_all, days) for days in LONGTERM_OPEN_DAYS}
    threshold_summary_map = {days: summarize_longterm_rows(threshold_rows_map.get(days, []), days) for days in LONGTERM_OPEN_DAYS}
    overview_rows = build_longterm_overview(threshold_rows_map, history_df, target_date)

    # 用現有 A/B/C/D/E 事件統計提供「原始已出清勝率」基準，再把長期風險做修正勝率。
    items = items_from_history_cache(history_df)
    item_map = {(item["broker_code"], item["warrant_code"]): item for item in items}
    daily_records = build_daily_records(items)
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [amount_events.get(code, []) for code in AMOUNT_CLASS_CODES]
    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)
    summary_map, _ = make_summary_map(stat_records)
    adjusted_rows = build_longterm_adjusted_winrate_rows(summary_map, threshold_rows_map)

    longterm_output_path = os.path.join(
        OUTPUT_DIR,
        f"warrant_longterm_open_positions_{datetime.today().strftime('%Y%m%d')}.xlsx",
    )
    build_longterm_workbook(
        longterm_output_path,
        overview_rows,
        threshold_summary_map,
        threshold_rows_map,
        adjusted_rows,
    )

    if LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET:
        print("  ⚠️ LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET=1：將完整長期留單 Excel 同步到 Google Sheet，可能占用大量格數。")
        upload_excel_to_google_sheet(longterm_output_path)
    else:
        print("  ✅ 長期留單明細 / 分點彙總僅輸出 Excel，不同步到 Google Sheet，避免占用格數。")

    upload_longterm_adjusted_winrate_to_gsheet(adjusted_rows, target_date)

    for row in overview_rows:
        print(
            f"  ✅ {row.get('門檻天數')}日以上：留單 {int(row.get('留單筆數') or 0):,} 筆｜"
            f"有價格 {int(row.get('有價格筆數') or 0):,}｜缺價 {int(row.get('缺價筆數') or 0):,}｜"
            f"賠錢 {int(row.get('賠錢筆數') or 0):,}｜龜苓膏 {int(row.get('疑似龜苓膏筆數') or 0):,}"
        )

    elapsed = time.time() - program_start
    print(f"\n{'=' * 70}")
    print("✅ 長期留單分析完成！")
    print(f"📄 {longterm_output_path}")
    print(f"⏱️ 總執行時間：{elapsed:.2f} 秒")

# ══════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════

def main():
    _GROUP_OUTCOME_SALE_ROWS_CACHE.clear()
    MONEYDJ_SEARCH_REPAIR_DISCOVERY_FETCH_KEYS.clear()
    program_start = time.time()

    configure_run_mode()

    today_fn = datetime.today().strftime("%Y%m%d")
    output_path = os.path.join(OUTPUT_DIR, f"warrant_backtest_ABCDE_{today_fn}.xlsx")

    print(f"\n認購權證特定分點買超回測 ABCDE 版 | {today_fn}")
    print("新制分類：同一分點 × 同一標的 × 同一天 = 1 筆事件")
    print(f"進入條件：事件內至少 1 檔權證單日買進金額 >= {AMOUNT_THRESH // 10000}萬")
    print("分類依據：同一事件內所有權證單日買進金額合計")
    print("A：100–159萬｜B：160–249萬｜C：250–499萬｜D：500–999萬｜E：1000萬以上")
    print(f"工作流模式：WORKFLOW_MODE={WORKFLOW_MODE}（daily=每日自動，longterm=長期留單，repair=完整修補）")
    print(f"執行模式：RUN_MODE={RUN_MODE}（1=精選分點全市場追蹤，2=完整分點清單）")
    print(f"分點數：{len(TARGET_PATTERNS)} 個")
    print(f"追蹤分點：{', '.join(TARGET_PATTERNS.keys())}")
    print(f"加速模式：FAST_SKIP_RECENT_PRESCAN={FAST_SKIP_RECENT_PRESCAN}，FETCH_GROUP_WARRANT_PRICES={FETCH_GROUP_WARRANT_PRICES}")
    print("=" * 70)

    warrants = get_all_call_warrants()

    if warrants:
        run_automatic_cache_maintenance(warrants)

    if not warrants:
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    broker_map = filter_broker_map_for_active_targets(find_broker_codes(warrants))

    if not broker_map:
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    if workflow_is_longterm():
        run_longterm_workflow(warrants, broker_map, output_path, program_start)
        return

    if workflow_is_daily():
        # 若今天已有 prescan success，直接交給 prescan_all() 驗證並讀回 keys；
        # 避免輕量探測樣本剛好今日無交易而誤判停止。
        if find_valid_prescan_success_record() is None:
            probe_result = light_probe_today_broker_data(warrants, broker_map)
            if probe_result is False:
                print("  ⚠️ 輕量探測未看到今日分點資料：不跑全市場 prescan，改為價格預抓後停止。")
                prefetch_candidates = filter_candidates_by_broker_map(load_candidates_cache(), broker_map)
                if prefetch_candidates:
                    print(f"  ✅ 價格預抓範圍改用既有候選快取：{len(prefetch_candidates):,} 組")
                else:
                    print("  ⚠️ 沒有既有候選快取可縮小範圍，價格預抓將回退使用歷史快取全範圍。")
                if maybe_auto_price_prefetch_before_api5(prefetch_candidates, program_start):
                    return
        else:
            print("  ✅ 已找到今日 prescan success 狀態，略過輕量探測，稍後直接驗證並讀取候選快取。")

    candidates = prescan_all(warrants, broker_map)

    activate_moneydj_search_repair_if_needed()
    history_cache_for_repair_pool = None

    # OpenAPI / API4 落後時，RUN_MODE=1 額外把「所有認購權證 × 精選分點」放入候選池。
    # 後面仍會用歷史快取 / 淨庫存 / 近期活動條件挑出要強制補抓的 key，
    # 不會因為候選池擴大就全部無差別重打 API5。
    if moneydj_search_repair_is_active() and RUN_MODE == 1 and MONEYDJ_SEARCH_REPAIR_SELECTED_FULL_POOL:
        selected_repair_candidates = build_selected_full_market_candidates(warrants, broker_map)
        before_count = len(candidates or [])
        candidates = merge_candidates(candidates or [], selected_repair_candidates)
        candidates = filter_candidates_by_broker_map(candidates, broker_map)
        print(
            f"  🔎 MoneyDJ Search 補漏：RUN_MODE=1 候選池擴充 "
            f"{before_count:,} → {len(candidates):,} 組"
        )

    # API4 / OpenAPI 落後時，不能直接用「分點 + 日期」查今日新標的。
    # 因此先從歷史快取找出近期有動作 / 尚有庫存的標的股，
    # 再展開同標的所有權證，讓 API5 去檢查今天是否有新買賣。
    if moneydj_search_repair_is_active() and MONEYDJ_SEARCH_REPAIR_HISTORY_UNDERLYING_EXPAND_ENABLED:
        history_cache_for_repair_pool = load_history_cache()
        repair_discovery_candidates = build_moneydj_search_repair_discovery_candidates(
            warrants,
            broker_map,
            history_cache_for_repair_pool,
            target_date=MONEYDJ_SEARCH_REPAIR_TARGET_DATE,
        )
        if repair_discovery_candidates:
            before_count = len(candidates or [])
            candidates = merge_candidates(candidates or [], repair_discovery_candidates)
            candidates = filter_candidates_by_broker_map(candidates, broker_map)
            MONEYDJ_SEARCH_REPAIR_DISCOVERY_FETCH_KEYS.update(
                candidate_key_from_tuple(c) for c in repair_discovery_candidates
            )
            print(
                f"  🔎 MoneyDJ Search 補漏：歷史活動標的候選池合併 "
                f"{before_count:,} → {len(candidates):,} 組｜"
                f"強制 API5 檢查 {len(MONEYDJ_SEARCH_REPAIR_DISCOVERY_FETCH_KEYS):,} 組"
            )

    # 候選快取分模式保存，但 API5 歷史快取是共用的。
    # 因此在正式使用 candidate_keys 過濾歷史資料之前，先把共用歷史中屬於目前追蹤分點、
    # 且仍在上市清單內的「權證代號 + 券商代號」補回目前模式候選池。
    # 這可直接修正：全分點已抓到資料，但精選五分點因 candidates_cache_selected5.csv 缺 key，
    # 導致 A/B/C/D/E 事件無法從共用歷史重建的問題。
    if history_cache_for_repair_pool is None:
        history_cache_for_repair_pool = load_history_cache()

    history_backfill_candidates = build_candidates_from_history_cache(
        warrants,
        broker_map,
        history_cache_for_repair_pool,
    )

    if history_backfill_candidates:
        before_keys = {
            candidate_key_from_tuple(c)
            for c in (candidates or [])
        }
        candidates = merge_candidates(candidates or [], history_backfill_candidates)
        candidates = filter_candidates_by_broker_map(candidates, broker_map)
        after_keys = {
            candidate_key_from_tuple(c)
            for c in candidates
        }
        added_count = len(after_keys - before_keys)

        if added_count > 0:
            save_candidates_cache(candidates)
            print(
                f"  ✅ 已從共用分點歷史補回目前模式候選：新增 {added_count:,} 組｜"
                f"候選池 {len(before_keys):,} → {len(after_keys):,} 組"
            )

    if maybe_auto_price_prefetch_before_api5(candidates, program_start):
        return

    if not candidates:
        print("⚠️ 預篩後無候選")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    print(f"\n【Step 3b】處理 {len(candidates)} 組候選...")

    candidate_keys = {candidate_key_from_tuple(c) for c in candidates}
    history_cache_df = history_cache_for_repair_pool if history_cache_for_repair_pool is not None else load_history_cache()
    history_was_empty = history_cache_df is None or history_cache_df.empty
    history_keys = history_cache_keys(history_cache_df)
    history_latest_map = history_cache_latest_dates(history_cache_df)

    if moneydj_search_repair_is_active():
        build_moneydj_search_repair_fetch_keys(
            history_cache_df,
            candidates,
            target_date=MONEYDJ_SEARCH_REPAIR_TARGET_DATE,
        )

    if CACHE_INCREMENTAL_UPDATE_ENABLED and not history_was_empty:
        before_prune_count = len(candidates)
        keep_keys = (history_keys & candidate_keys) | PRESCAN_MISSING_FETCH_KEYS | MONEYDJ_SEARCH_REPAIR_FETCH_KEYS
        candidates = [c for c in candidates if candidate_key_from_tuple(c) in keep_keys]
        candidate_keys = {candidate_key_from_tuple(c) for c in candidates}
        pruned_count = before_prune_count - len(candidates)
        if pruned_count > 0:
            print(f"  ✅ 增量模式已略過舊候選快取中的無歷史資料空候選：{pruned_count:,} 組")

    cached_items = items_from_history_cache(history_cache_df, candidate_filter=candidate_keys)

    if cached_items:
        print(f"  ✅ 已從原始分點資料快取還原 {len(cached_items)} 組資料")

    candidates_to_fetch = []

    for c in candidates:
        key = candidate_key_from_tuple(c)

        if key in MONEYDJ_SEARCH_REPAIR_FETCH_KEYS:
            candidates_to_fetch.append(c)
            continue

        if should_fetch_candidate_incremental(key, history_keys, history_latest_map, PRESCAN_REFRESH_KEYS, PRESCAN_MISSING_FETCH_KEYS):
            candidates_to_fetch.append(c)

    print(f"  ✅ 快取已有候選：{len(history_keys & candidate_keys)} 組")
    print(f"  ✅ API4 直接近期活動候選：{len(PRESCAN_REFRESH_KEYS & candidate_keys)} 組")
    print(f"  ✅ 本次允許缺快取補抓候選：{len(PRESCAN_MISSING_FETCH_KEYS & candidate_keys)} 組")
    if moneydj_search_repair_is_active():
        print(f"  ✅ MoneyDJ Search 強制補漏候選：{len(MONEYDJ_SEARCH_REPAIR_FETCH_KEYS & candidate_keys)} 組")
    print(f"  ✅ 增量更新模式：CACHE_INCREMENTAL_UPDATE_ENABLED={CACHE_INCREMENTAL_UPDATE_ENABLED}，目標日期={get_incremental_refresh_target_dt().strftime('%Y/%m/%d')}")
    print(f"  ✅ 本次需要 API5 更新：{len(candidates_to_fetch)} 組")

    fetched_items = []
    done = 0

    if candidates_to_fetch:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futures = {
                ex.submit(process_candidate, *c): c
                for c in candidates_to_fetch
            }

            for future in as_completed(futures):
                done += 1

                try:
                    item = future.result()
                except:
                    item = None

                if item:
                    fetched_items.append(item)

                if done % 100 == 0:
                    print(f"  [{done:,}/{len(candidates_to_fetch):,}] 已更新，成功取得 {len(fetched_items)} 組資料")
    else:
        print("  ✅ 所有候選組合皆可使用快取，略過 API5 歷史資料重抓")

    if fetched_items:
        history_cache_df = merge_items_into_history_cache(history_cache_df, fetched_items)
        history_cache_df = save_history_cache(history_cache_df, fetched_items=fetched_items, previous_history_empty=history_was_empty)

    items = items_from_history_cache(history_cache_df, candidate_filter=candidate_keys)

    if not items and fetched_items:
        items = fetched_items

    if not items:
        print("⚠️ 無任何候選資料")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    item_map = {}

    for item in items:
        item_map[(item["broker_code"], item["warrant_code"])] = item

    daily_records = build_daily_records(items)

    print("【Step 3c】建立 A/B/C/D/E 金額強度分類事件：同分點 × 同標的 × 同一天...")
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [
        amount_events.get(code, [])
        for code in AMOUNT_CLASS_CODES
    ]

    print(
        f"  ✅ 金額強度事件："
        f"A:{len(a_events):,}｜B:{len(b_events):,}｜C:{len(c_events):,}｜D:{len(d_events):,}｜E:{len(e_events):,}"
    )

    if not a_events and not b_events and not c_events and not d_events and not e_events:
        print("⚠️ A/B/C/D/E 皆無事件")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    price_cache = fetch_all_prices(a_events, b_events, c_events, d_events, e_events)
    top15_detail_rows, top15_consensus_rows = build_top15_position_detail_and_consensus_rows(
        a_events, b_events, c_events, d_events, e_events, item_map, price_cache
    )
    warrant_consensus_7d_rows = build_7d_warrant_consensus_top15_rows(items)

    if RUN_MODE == 2:
        price_cache = ensure_broker_10d_underlying_prices(price_cache, items)
        price_cache = ensure_broker_10d_warrant_prices(price_cache, items)
        broker_10d_detail_rows = build_10d_broker_underlying_detail_rows(items, price_cache)
        broker_10d_winrate_rank_rows = build_10d_broker_winrate_rank_rows(broker_10d_detail_rows)
    else:
        print("  ✅ RUN_MODE=1 精選分點模式：略過近10日分點買賣明細與分點勝率排行工作表，避免動到 Google Sheet 既有資料。")
        broker_10d_detail_rows = None
        broker_10d_winrate_rank_rows = None

    build_excel(
        a_events, b_events, c_events, d_events, e_events,
        item_map, price_cache, items, output_path,
        top15_detail_rows, top15_consensus_rows, warrant_consensus_7d_rows, broker_10d_detail_rows, broker_10d_winrate_rank_rows
    )
    upload_excel_to_google_sheet(output_path)

    elapsed = time.time() - program_start
    minutes = int(elapsed // 60)
    seconds = elapsed % 60

    print(f"\n{'=' * 70}")
    print("✅ 完成！")
    print(f"📄 {output_path}")
    print(f"⏱️ 總執行時間：{elapsed:.2f} 秒")
    print(f"⏱️ 約為：{minutes} 分 {seconds:.2f} 秒")


if __name__ == "__main__":
    main()
