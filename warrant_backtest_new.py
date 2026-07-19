"""
認購權證特定分點買超回測系統
=================================
新制金額強度分類：同一分點 + 同一標的 + 同一天為 1 筆事件。
進入條件：該事件內至少 1 檔權證單日買進金額 >= 100萬。
分類依據：同一分點 + 同一標的 + 同一天的所有權證買進金額合計。
A：單日累積買進金額 100–159萬
B：單日累積買進金額 160–249萬
C：單日累積買進金額 250–499萬
D：單日累積買進金額 500–999萬
E：單日累積買進金額 >= 1000萬

事件規則：同一分點 + 同一標的 + 同一天只會產生 A / B / C / D / E 其中一類；不同分點交易同一標的互不排除。每次符合條件的事件各自成一筆，後續賣出依剩餘股數 FIFO 扣減；金額僅用於成本與損益。

每日流程只輸出與同步以下 8 張工作表：
1. A_基礎買超
2. B_明顯買超
3. C_強勢買超
4. D_大額布局
5. E_超大額布局
6. 每日賣出明細
7. 快取_TOP15共識淨買超
8. 快取_TOP15部位明細

每日流程仍只計算與同步上述 8 張工作表；其他既有 Google Sheet 工作表一律保留。
完整修補模式會重新產生完整報表，缺少的工作表才建立，既有工作表只增量插入缺少資料。

資料來源：FinMind API（FINMIND_API_0714）
執行：python warrant_backtest.py
依賴：pip install requests pandas openpyxl pyarrow
"""

import json, re, time, os
import threading
from bisect import bisect_right
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from functools import lru_cache
from io import BytesIO, StringIO

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


os.environ.setdefault("TZ", "Asia/Taipei")
if hasattr(time, "tzset"):
    time.tzset()


# ══════════════════════════════════════════════════════════════════════
# 設定
# ══════════════════════════════════════════════════════════════════════

DEFAULT_OUTPUT_DIR = "output" if os.getenv("GITHUB_ACTIONS", "").strip().lower() == "true" else r"C:\Users\chen1_ukw0m7r\Downloads"
OUTPUT_DIR = os.getenv("OUTPUT_DIR", DEFAULT_OUTPUT_DIR)
AMOUNT_THRESH = 1_000_000
PROGRAM_BUILD_ID = "FINMIND-BATCH-PRICE-PERFORMANCE-FIXED-20260719-V6"

# 權證／標的身分配對防錯：
# 1. 標的名稱永遠以 TaiwanStockInfo 的「股號→股名」主檔為準。
# 2. 權證名稱以日期區間、標的股與權證名稱前綴交叉驗證，避免權證代號重用後抓到舊名稱。
# 3. 若 FinMind Summary 的 target_stock_id 與權證名稱明確指向不同標的，
#    只在名稱前綴屬於高信心匹配時才更正，避免短別名誤判。
WARRANT_IDENTITY_RECONCILE_ENABLED = os.getenv("WARRANT_IDENTITY_RECONCILE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
WARRANT_IDENTITY_OVERRIDE_MIN_PREFIX = max(int(os.getenv("WARRANT_IDENTITY_OVERRIDE_MIN_PREFIX", "3")), 2)
WARRANT_IDENTITY_REPAIR_HISTORY = os.getenv("WARRANT_IDENTITY_REPAIR_HISTORY", "1").strip().lower() not in ("0", "false", "no")
GSHEET_REPAIR_SECURITY_IDENTITY_ROWS = os.getenv("GSHEET_REPAIR_SECURITY_IDENTITY_ROWS", "1").strip().lower() not in ("0", "false", "no")

_CURRENT_STOCK_CODE_TO_NAME = {}
_CURRENT_STOCK_NAME_TO_CODE = {}
_CURRENT_UNDERLYING_RESOLVER = []
_CURRENT_WARRANT_INTERVAL_RECORDS = []

# 權證身分配對效能快取：只快取索引與既有判斷結果，不改變任何身分判斷規則。
# 1. 股名 resolver 依第一個字建立候選索引，仍保持原 resolver 排序。
# 2. 相同 resolver + 權證名稱只解析一次。
# 3. 權證日期對照依 warrants 物件與日期快取，同一天只建立一次。
_UNDERLYING_RESOLVER_RUNTIME_LOCK = threading.Lock()
_UNDERLYING_RESOLVER_RUNTIME_REF = None
_UNDERLYING_RESOLVER_FIRST_CHAR_INDEX = {}
_UNDERLYING_RESOLUTION_RESULT_CACHE = {}
_UNDERLYING_RESOLUTION_RESULT_CACHE_MAXSIZE = 65536

_WARRANT_LOOKUP_CACHE_LOCK = threading.Lock()
_WARRANT_LOOKUP_CACHE_WARRANTS_REF = None
_WARRANT_LOOKUP_CACHE = {}

# 新制金額強度分類：
# 事件單位為「同一分點 + 同一標的 + 同一天」。
# 進入條件為事件內至少一檔權證單日買進金額 >= AMOUNT_THRESH；
# 分類依據為同標的單日累積買進金額。
AMOUNT_CLASS_SPECS = [
    ("A", "基礎買超", 1_000_000, 1_600_000),
    ("B", "明顯買超", 1_600_000, 2_500_000),
    ("C", "強勢買超", 2_500_000, 5_000_000),
    ("D", "大額布局", 5_000_000, 10_000_000),
    ("E", "超大額布局", 10_000_000, None),
]
AMOUNT_CLASS_CODES = [spec[0] for spec in AMOUNT_CLASS_SPECS]
AMOUNT_CLASS_LABELS = {code: label for code, label, _, _ in AMOUNT_CLASS_SPECS}
AMOUNT_CLASS_SHEET_NAMES = {
    f"{code}_{label}"
    for code, label, _, _ in AMOUNT_CLASS_SPECS
}
SELL_DETAIL_DAYS = int(os.getenv("SELL_DETAIL_DAYS", "3"))

# 快取與效能設定。FinMind 原始日檔、標準化歷史與價格皆使用本機 Parquet；
# Google Sheet 僅同步最終 8 張結果表。
USE_CACHE = os.getenv("USE_CACHE", "1").strip().lower() not in ("0", "false", "no")
FORCE_FULL_CACHE_REFRESH = os.getenv("FORCE_FULL_CACHE_REFRESH", "0").strip().lower() in ("1", "true", "yes")
PRICE_WORKERS = int(os.getenv("PRICE_WORKERS", "80"))
FETCH_GROUP_WARRANT_PRICES = os.getenv("FETCH_GROUP_WARRANT_PRICES", "0").strip().lower() in ("1", "true", "yes")

CACHE_DIR = os.getenv("CACHE_DIR", os.path.join(OUTPUT_DIR, "warrant_cache"))
CACHE_ENCODING = "utf-8-sig"
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CACHE_DIR, exist_ok=True)

WARRANTS_CACHE_PATH = os.path.join(CACHE_DIR, "warrants_cache.csv")
BROKER_MAP_CACHE_PATH = os.path.join(CACHE_DIR, "broker_map_cache.csv")
HISTORY_CACHE_PATH = os.path.join(CACHE_DIR, "broker_warrant_history_cache.csv")
PRICE_CACHE_PATH = os.path.join(CACHE_DIR, "price_cache.csv")

# FinMind 正式資料來源。
FINMIND_TOKEN = os.getenv("FINMIND_API_0714", "").strip()
FINMIND_API_BASE = os.getenv("FINMIND_API_BASE", "https://api.finmindtrade.com/api/v4").rstrip("/")
FINMIND_DATA_URL = f"{FINMIND_API_BASE}/data"
FINMIND_STORAGE_OBJECTS_URL = f"{FINMIND_API_BASE}/storage_objects"
FINMIND_REQUEST_TIMEOUT_CONNECT = float(os.getenv("FINMIND_REQUEST_TIMEOUT_CONNECT", "8"))
FINMIND_REQUEST_TIMEOUT_READ = float(os.getenv("FINMIND_REQUEST_TIMEOUT_READ", "120"))
FINMIND_MAX_RETRIES = max(int(os.getenv("FINMIND_MAX_RETRIES", "4")), 1)
FINMIND_RETRY_BASE_SECONDS = max(float(os.getenv("FINMIND_RETRY_BASE_SECONDS", "2")), 0.2)
FINMIND_METADATA_CACHE_HOURS = max(float(os.getenv("FINMIND_METADATA_CACHE_HOURS", "18")), 0.0)
FINMIND_HISTORY_WORKERS = max(int(os.getenv("FINMIND_HISTORY_WORKERS", "6")), 1)
FINMIND_FORCE_REFRESH_TARGET_DATE = os.getenv("FINMIND_FORCE_REFRESH_TARGET_DATE", "1").strip().lower() not in ("0", "false", "no")
FINMIND_INITIAL_BACKFILL_ON_EMPTY = os.getenv("FINMIND_INITIAL_BACKFILL_ON_EMPTY", "1").strip().lower() not in ("0", "false", "no")
FINMIND_DAILY_CACHE_DIR = os.path.join(CACHE_DIR, "finmind_warrant_daily")
FINMIND_METADATA_CACHE_DIR = os.path.join(CACHE_DIR, "finmind_metadata")
FINMIND_STATE_PATH = os.path.join(CACHE_DIR, "finmind_update_state.json")
FINMIND_CACHE_SCHEMA_VERSION = os.getenv("FINMIND_CACHE_SCHEMA_VERSION", "finmind-v2-date-aware").strip() or "finmind-v2-date-aware"
FINMIND_RAW_DAILY_RETENTION_DAYS = max(int(os.getenv("FINMIND_RAW_DAILY_RETENTION_DAYS", "7")), 1)
# 完整修補模式不可因最新交易日資料尚未發布而提前結束。
# repair 會往前探測最近已發布的全市場權證分點日檔，並以該日作為修補基準日；
# daily 仍維持最新日未發布就快速結束，避免排程空抓。
FINMIND_REPAIR_ALLOW_TARGET_FALLBACK = os.getenv(
    "FINMIND_REPAIR_ALLOW_TARGET_FALLBACK",
    "1",
).strip().lower() not in ("0", "false", "no")
FINMIND_REPAIR_TARGET_FALLBACK_TRADING_DAYS = max(
    int(os.getenv("FINMIND_REPAIR_TARGET_FALLBACK_TRADING_DAYS", "15")),
    1,
)
CACHE_WRITE_CSV_COMPAT = os.getenv("CACHE_WRITE_CSV_COMPAT", "0").strip().lower() in ("1", "true", "yes")
GSHEET_CLEAN_DELETED_BROKER_ROWS = os.getenv("GSHEET_CLEAN_DELETED_BROKER_ROWS", "1").strip().lower() not in ("0", "false", "no")

os.makedirs(FINMIND_DAILY_CACHE_DIR, exist_ok=True)
os.makedirs(FINMIND_METADATA_CACHE_DIR, exist_ok=True)

_FINMIND_DAY_LOCKS = {}
_FINMIND_DAY_LOCKS_GUARD = threading.Lock()
_FINMIND_TARGET_DATE_OK = False
_FINMIND_TARGET_DATE = ""

# 保留給 longterm 指定日期相容使用。
PRICE_PREFETCH_TARGET_DATE = os.getenv("PRICE_PREFETCH_TARGET_DATE", "").strip()

# 長期留單補價／修正勝率設定。
LONGTERM_OPEN_DAYS = [
    int(x.strip())
    for x in re.split(r"[,;；、\n\r\t]+", os.getenv("LONGTERM_OPEN_DAYS", "120,150,180"))
    if x.strip().isdigit()
]
if not LONGTERM_OPEN_DAYS:
    LONGTERM_OPEN_DAYS = [120, 150, 180]
LONGTERM_OPEN_DAYS = sorted(set(max(int(x), 1) for x in LONGTERM_OPEN_DAYS))
LONGTERM_PRICE_LOOKBACK_DAYS = int(os.getenv("LONGTERM_PRICE_LOOKBACK_DAYS", "420"))
LONGTERM_PRICE_STALE_DAYS = int(os.getenv("LONGTERM_PRICE_STALE_DAYS", "10"))
LONGTERM_ZERO_PRICE_THRESHOLD = float(os.getenv("LONGTERM_ZERO_PRICE_THRESHOLD", "0.05"))
LONGTERM_TARGET_DATE = os.getenv("LONGTERM_TARGET_DATE", "").strip()
LONGTERM_MAX_DETAIL_ROWS = int(os.getenv("LONGTERM_MAX_DETAIL_ROWS", "0"))
LONGTERM_UPDATE_WINRATE_SHEET = os.getenv("LONGTERM_UPDATE_WINRATE_SHEET", os.getenv("LONGTERM_UPLOAD_TO_GSHEET", "1")).strip().lower() not in ("0", "false", "no")
LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET = os.getenv("LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET", "0").strip().lower() in ("1", "true", "yes")
LONGTERM_WINRATE_SHEET_DAYS = int(os.getenv("LONGTERM_WINRATE_SHEET_DAYS", str(LONGTERM_OPEN_DAYS[0])) or "120")
LONGTERM_WINRATE_SHEET_HEADER = os.getenv("LONGTERM_WINRATE_SHEET_HEADER", "修正勝率").strip() or "修正勝率"

# daily：每日增量；longterm：長期留單；repair：重建保留區間。
WORKFLOW_MODE = os.getenv("WORKFLOW_MODE", "daily").strip().lower() or "daily"
if WORKFLOW_MODE not in ("daily", "longterm", "repair"):
    print(f"  ⚠️ WORKFLOW_MODE={WORKFLOW_MODE} 不支援，改用 daily。")
    WORKFLOW_MODE = "daily"

# TOP15 固定資料集。
TOP15_CACHE_ENABLED = os.getenv("TOP15_CACHE_ENABLED", os.getenv("TOP15_RETURN_CACHE_ENABLED", "1")).strip().lower() not in ("0", "false", "no")
TOP15_POSITION_DETAIL_SHEET = os.getenv("TOP15_POSITION_DETAIL_SHEET", "快取_TOP15部位明細")
TOP15_CONSENSUS_SHEET = os.getenv("TOP15_CONSENSUS_SHEET", "快取_TOP15共識淨買超")
DAILY_RESULT_SHEET_TITLES = {
    *AMOUNT_CLASS_SHEET_NAMES,
    "每日賣出明細",
    TOP15_CONSENSUS_SHEET,
    TOP15_POSITION_DETAIL_SHEET,
}
TOP15_LOOKBACK_TRADING_DAYS = int(os.getenv("TOP15_LOOKBACK_TRADING_DAYS", os.getenv("TOP15_RETURN_LOOKBACK_TRADING_DAYS", "40")))
TOP15_PRICE_LOOKBACK_DAYS = int(os.getenv("TOP15_PRICE_LOOKBACK_DAYS", os.getenv("TOP15_RETURN_PRICE_LOOKBACK_DAYS", "75")))
TOP15_PRICE_STALE_DAYS = int(os.getenv("TOP15_PRICE_STALE_DAYS", os.getenv("TOP15_RETURN_PRICE_STALE_DAYS", "0")))
TOP15_REQUIRE_TARGET_DATE_PRICE = os.getenv("TOP15_REQUIRE_TARGET_DATE_PRICE", "1").strip().lower() not in ("0", "false", "no")
TOP15_MIN_BROKER_REMAINING_COST = max(float(os.getenv("TOP15_MIN_BROKER_REMAINING_COST", "100000")), 0.0)
TOP15_FAIL_ON_MISSING_PRICE = os.getenv("TOP15_FAIL_ON_MISSING_PRICE", "1").strip().lower() not in ("0", "false", "no")
TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN = os.getenv("TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN", "1").strip().lower() not in ("0", "false", "no")
# TOP15 統計日沒有成交價時，改向權證資訊揭露平台查詢流動量提供者最佳委買價。
# 當日成交價永遠優先；只有統計日無成交價時才使用 LP 委買價估值。
TOP15_LP_QUOTE_FALLBACK_ENABLED = os.getenv("TOP15_LP_QUOTE_FALLBACK_ENABLED", "1").strip().lower() not in ("0", "false", "no")
TOP15_LP_QUOTE_TIMEOUT_SECONDS = max(float(os.getenv("TOP15_LP_QUOTE_TIMEOUT_SECONDS", "15")), 3.0)
TOP15_LP_QUOTE_WORKERS = max(int(os.getenv("TOP15_LP_QUOTE_WORKERS", "10")), 1)
# LP 頁面若沒有顯示資料日期，預設不接受，避免通用頁面或錯誤端點的數字被誤認為統計日報價。
# 只有未來確認正式端點可安全使用時，才手動設定 TOP15_LP_ALLOW_IMPLICIT_LATEST_DATE=1。
TOP15_LP_ALLOW_IMPLICIT_LATEST_DATE = os.getenv(
    "TOP15_LP_ALLOW_IMPLICIT_LATEST_DATE", "0"
).strip().lower() not in ("0", "false", "no")
# LP 仍取不到時，使用該權證在統計日前最近一次有效成交價作最後備援，
# 並以 * 明確標示這不是統計日實際成交價。
TOP15_LAST_TRADE_FALLBACK_ENABLED = os.getenv(
    "TOP15_LAST_TRADE_FALLBACK_ENABLED", "1"
).strip().lower() not in ("0", "false", "no")
TOP15_LAST_TRADE_FALLBACK_MAX_DAYS = max(
    int(os.getenv("TOP15_LAST_TRADE_FALLBACK_MAX_DAYS", str(TOP15_PRICE_LOOKBACK_DAYS))),
    0,
)
TOP15_TRADED_PRICE_COVERAGE_NOTE_THRESHOLD_PCT = min(
    max(float(os.getenv("TOP15_TRADED_PRICE_COVERAGE_NOTE_THRESHOLD_PCT", "60")), 0.0),
    100.0,
)
TOP15_LP_QUOTE_URLS_ENV = os.getenv("TOP15_LP_QUOTE_URLS", "").strip()
TOP15_LP_QUOTE_URLS = [
    x.strip()
    for x in re.split(r"[,;；、\n\r\t]+", TOP15_LP_QUOTE_URLS_ENV)
    if x.strip()
] or [
    "https://warrants.twse.com.tw/Query.aspx",
    "https://warrants.twse.com.tw/query",
    "https://warrants.sfi.org.tw/Query.aspx",
]
TOP15_TARGET_DATE = os.getenv("TOP15_TARGET_DATE", "").strip()

# 完整修補模式會重建下列延伸報表；每日流程不計算、不碰觸這些既有工作表。
RECENT_RANKING_DAYS = int(os.getenv("RECENT_RANKING_DAYS", "62"))
WARRANT_CONSENSUS_7D_ENABLED = os.getenv("WARRANT_CONSENSUS_7D_ENABLED", "1").strip().lower() not in ("0", "false", "no")
WARRANT_CONSENSUS_7D_SHEET = os.getenv("WARRANT_CONSENSUS_7D_SHEET", "快取_近7日權證分點共識TOP15")
WARRANT_CONSENSUS_7D_DAYS = int(os.getenv("WARRANT_CONSENSUS_7D_DAYS", "7"))
WARRANT_CONSENSUS_14D_DAYS = int(os.getenv("WARRANT_CONSENSUS_14D_DAYS", "14"))
WARRANT_CONSENSUS_21D_DAYS = int(os.getenv("WARRANT_CONSENSUS_21D_DAYS", "21"))
WARRANT_CONSENSUS_7D_TOP_N = int(os.getenv("WARRANT_CONSENSUS_7D_TOP_N", "15"))
WARRANT_CONSENSUS_SELECTED_BROKERS_DEFAULT = [
    "元大南屯", "華南永昌台中", "新光", "統一三多", "永豐金竹科",
    "福邦證券", "群益金鼎新竹", "凱基士林", "元大內湖民權",
    "群益金鼎古亭", "兆豐板橋", "富邦敦南", "永豐金內湖",
]
WARRANT_CONSENSUS_SELECTED_BROKERS_ENV = os.getenv("WARRANT_CONSENSUS_SELECTED_BROKERS", "").strip()
BROKER_10D_DETAIL_ENABLED = os.getenv("BROKER_10D_DETAIL_ENABLED", "1").strip().lower() not in ("0", "false", "no")
BROKER_10D_DETAIL_SHEET = os.getenv("BROKER_10D_DETAIL_SHEET", "快取_近10日分點買賣明細")
BROKER_10D_WINRATE_RANK_ENABLED = os.getenv("BROKER_10D_WINRATE_RANK_ENABLED", "1").strip().lower() not in ("0", "false", "no")
BROKER_10D_WINRATE_RANK_SHEET = os.getenv("BROKER_10D_WINRATE_RANK_SHEET", "快取_近10日分點勝率排行")
BROKER_10D_DETAIL_DAYS = int(os.getenv("BROKER_10D_DETAIL_DAYS", "10"))

FULL_REPAIR_RESULT_SHEET_TITLES = {
    *DAILY_RESULT_SHEET_TITLES,
    "勝率統計",
    "ABCDE組合勝率",
    "近兩月買賣金額排行",
    "近兩月分點數排行",
    "券商查詢",
    "券商查詢資料",
    "股票ABCDE查詢",
    "股票ABCDE查詢資料",
    "價格抓取狀態",
    "顏色說明",
    WARRANT_CONSENSUS_7D_SHEET,
    BROKER_10D_DETAIL_SHEET,
    BROKER_10D_WINRATE_RANK_SHEET,
}

# RUN_MODE=1：精選五分點；RUN_MODE=2：完整分點清單。
RUN_MODE = int(os.getenv("RUN_MODE", os.getenv("BROKER_RUN_MODE", "1")) or "1")
SELECTED_TARGET_LABELS_DEFAULT = [
    "華南永昌台中",
    "元大南屯",
    "富邦敦南",
    "永豐金內湖",
    "新光",
]
SELECTED_TARGET_LABELS_ENV = os.getenv("SELECTED_TARGET_LABELS", "").strip()

# 快取保留與清理。
CACHE_AUTO_PRUNE_ENABLED = os.getenv("CACHE_AUTO_PRUNE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
HISTORY_RETENTION_TRADING_DAYS = max(int(os.getenv("HISTORY_RETENTION_TRADING_DAYS", "200")), 1)
PRICE_RETENTION_TRADING_DAYS = max(int(os.getenv("PRICE_RETENTION_TRADING_DAYS", "200")), 1)

TARGET_PATTERNS = {
    "富邦公益":       r"富邦.*公益",
    "富邦敦南":       r"富邦.*敦南",
    "富邦仁愛":       r"富邦.*仁愛",
    "新光":           r"^新光$",
    "永豐金內湖":     r"永豐.*內湖",
    "永豐金竹北":     r"永豐.*竹北",
    "永豐金竹科":     r"永豐.*竹科",
    "永豐金市政":     r"永豐.*市政",
    "永豐金信義":     r"永豐.*信義",
    "華南永昌台中":   r"華南.*台中",
    "華南永昌淡水":   r"華南.*淡水",
    "福邦證券":       r"^福邦證券",
    "群益東大":       r"群益.*東大",
    "群益金鼎古亭":   r"群益.*古亭",
    "群益金鼎新竹":   r"群益.*新竹",
    "元大內湖民權":   r"元大.*(內湖.*民權|民權)",
    "元大南屯":       r"元大.*南屯",
    "元大汐止":       r"元大.*汐止",
    "元大虎尾":       r"元大.*虎尾",
    "元大彰化民生":   r"元大.*彰化民生",
    "兆豐板橋":       r"兆豐.*板橋",
    "凱基士林":       r"凱基.*士林",
    "凱基中山":       r"凱基.*中山",
    "國票敦北法人":   r"國票.*(敦北|法人)",
    "統一三多":       r"統一.*三多",
    "第一金中壢":     r"第一金.*中壢",
}

FALLBACK = {
    "富邦公益":       ("富邦-公益",       "961F"),
    "富邦敦南":       ("富邦-敦南",       "9663"),
    "富邦仁愛":       ("富邦-仁愛",       "9676"),
    "新光":           ("新光",           "8560"),
    "永豐金內湖":     ("永豐金-內湖",     "9A9g"),
    "永豐金竹北":     ("永豐金-竹北",     "9A9P"),
    "永豐金竹科":     ("永豐金-竹科",     "9A9X"),
    "永豐金市政":     ("永豐金-市政",     "9A9W"),
    "永豐金信義":     ("永豐金-信義",     "9A9R"),
    "華南永昌台中":   ("華南永昌-台中",   "9302"),
    "華南永昌淡水":   ("華南永昌-淡水",   "9316"),
    "福邦證券":       ("福邦證券",       "6480"),
    "群益東大":       ("群益金鼎-東大",   "9135"),
    "群益金鼎古亭":   ("群益金鼎-古亭",   "918C"),
    "群益金鼎新竹":   ("群益金鼎-新竹",   "9186"),
    "元大內湖民權":   ("元大-內湖民權",   "9867"),
    "元大南屯":       ("元大-南屯",       "9853"),
    "元大汐止":       ("元大-汐止",       "989Q"),
    "元大虎尾":       ("元大-虎尾",       "980l"),
    "元大彰化民生":   ("元大-彰化民生",   "989J"),
    "兆豐板橋":       ("兆豐-板橋",       "700B"),
    "凱基士林":       ("凱基-士林",       "9238"),
    "凱基中山":       ("凱基-中山",       "9229"),
    "國票敦北法人":   ("國票-敦北法人",   "779c"),
    "統一三多":       ("統一-三多",       "585Q"),
    "第一金中壢":     ("第一金-中壢",      "538Y"),
}

FULL_TARGET_PATTERNS = dict(TARGET_PATTERNS)
FULL_FALLBACK = dict(FALLBACK)
LIVE_WARRANT_SNAPSHOT_READY = False




def parse_selected_target_labels():
    if SELECTED_TARGET_LABELS_ENV:
        labels = [
            x.strip()
            for x in re.split(r"[,;；、\n\r\t]+", SELECTED_TARGET_LABELS_ENV)
            if x.strip()
        ]
    else:
        labels = list(SELECTED_TARGET_LABELS_DEFAULT)

    out = []
    for label in labels:
        if label not in out:
            out.append(label)

    return out



def configure_run_mode():
    """依 RUN_MODE 切換完整分點或精選五分點範圍。"""
    global RUN_MODE, TARGET_PATTERNS, FALLBACK

    try:
        RUN_MODE = int(os.getenv("RUN_MODE", os.getenv("BROKER_RUN_MODE", str(RUN_MODE))) or "1")
    except Exception:
        RUN_MODE = 1

    if RUN_MODE not in (1, 2):
        print(f"  ⚠️ RUN_MODE={RUN_MODE} 不支援，改用 RUN_MODE=1 精選分點模式。")
        RUN_MODE = 1

    if RUN_MODE == 1:
        selected_labels = parse_selected_target_labels()
        missing = [label for label in selected_labels if label not in FULL_TARGET_PATTERNS]
        if missing:
            print(f"  ⚠️ SELECTED_TARGET_LABELS 中不存在的分點已略過：{missing}")

        active_labels = [label for label in selected_labels if label in FULL_TARGET_PATTERNS]
        if not active_labels:
            print("  ⚠️ 精選分點清單為空，改用預設 5 間分點。")
            active_labels = [
                label for label in SELECTED_TARGET_LABELS_DEFAULT
                if label in FULL_TARGET_PATTERNS
            ]

        TARGET_PATTERNS = {label: FULL_TARGET_PATTERNS[label] for label in active_labels}
        FALLBACK = {
            label: FULL_FALLBACK[label]
            for label in active_labels
            if label in FULL_FALLBACK
        }
        print("  ✅ RUN_MODE=1：精選分點全市場追蹤模式")
        print(f"  ✅ 精選分點：{', '.join(TARGET_PATTERNS.keys())}")
    else:
        TARGET_PATTERNS = dict(FULL_TARGET_PATTERNS)
        FALLBACK = dict(FULL_FALLBACK)
        print("  ✅ RUN_MODE=2：完整分點清單模式")
        print(f"  ✅ 分點數：{len(TARGET_PATTERNS)}")



def filter_broker_map_for_active_targets(broker_map):
    if not broker_map:
        return {}

    active_labels = set(TARGET_PATTERNS.keys())
    return {
        label: value
        for label, value in broker_map.items()
        if label in active_labels
    }





_THREAD_LOCAL = threading.local()


def get_thread_session():
    """
    每個執行緒各自建立並重用 requests.Session。

    目的：
    1. 避免 api4_get / api5_get 每次呼叫都重新建立連線。
    2. 在 ThreadPoolExecutor 多執行緒抓資料時，每個 thread 使用自己的 Session，
       避免多執行緒共用同一個 Session 造成不穩定。
    3. 不改變任何抓資料邏輯，只改善大量 API 請求時的連線重用效率。
    """
    session = getattr(_THREAD_LOCAL, "session", None)

    if session is None:
        session = requests.Session()
        adapter = HTTPAdapter(pool_connections=100, pool_maxsize=100, max_retries=1)
        session.mount("http://", adapter)
        session.mount("https://", adapter)
        _THREAD_LOCAL.session = session

    return session




# Excel 顏色（柔和舒適版）
RED    = PatternFill("solid", fgColor="F4CCCC")
GREEN  = PatternFill("solid", fgColor="D9EAD3")
BLUE   = PatternFill("solid", fgColor="D9EAF7")
ORANGE = PatternFill("solid", fgColor="FCE5CD")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GRAY   = PatternFill("solid", fgColor="E7E6E6")
WHITE  = PatternFill("solid", fgColor="FFFFFF")

# 出清且獲利時使用的外框（不改底色，只用外框凸顯）
PROFIT_EXIT_SIDE = Side(style="thick", color="C00000")
LOSS_EXIT_SIDE   = Side(style="thick", color="38761D")


# ══════════════════════════════════════════════════════════════════════
# 工具函式
# ══════════════════════════════════════════════════════════════════════

@lru_cache(maxsize=65536)
def _parse_date_cached_text(text):
    try:
        s = str(text).strip()
        if not s or s == "-":
            return None
        # pandas Timestamp / datetime 轉成字串時可能帶時間，只取日期部分。
        s = s.split(" ", 1)[0].replace("-", "/")
        parts = s.split("/")
        if len(parts) != 3:
            return None
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        return datetime(y, m, d)
    except Exception:
        return None


def parse_date(date_str):
    if date_str is None:
        return None
    if isinstance(date_str, datetime):
        return datetime(date_str.year, date_str.month, date_str.day)
    try:
        if isinstance(date_str, pd.Timestamp):
            if pd.isna(date_str):
                return None
            return datetime(date_str.year, date_str.month, date_str.day)
    except Exception:
        pass
    return _parse_date_cached_text(str(date_str).strip())


@lru_cache(maxsize=65536)
def _normalize_date_str_cached_value(date_str):
    dt = parse_date(date_str)
    return dt.strftime("%Y/%m/%d") if dt else str(date_str).strip()


def normalize_date_str(date_str):
    try:
        hash(date_str)
    except TypeError:
        dt = parse_date(date_str)
        return dt.strftime("%Y/%m/%d") if dt else str(date_str).strip()
    return _normalize_date_str_cached_value(date_str)


def add_months(dt, months):
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    return datetime(year, month, 1)


def iter_month_starts(start_dt, end_dt):
    start_dt = datetime(start_dt.year, start_dt.month, 1)
    end_dt   = datetime(end_dt.year, end_dt.month, 1)

    cur = start_dt
    while cur <= end_dt:
        yield cur.strftime("%Y%m01")
        cur = add_months(cur, 1)


def fmt_pct(v):
    return "-" if v is None else f"{v:+.2f}%"


def fmt_num(v):
    return "-" if v is None else v


def fmt_amount(v):
    if v is None:
        return "-"
    try:
        return f"{int(round(float(v))):,}"
    except Exception:
        return str(v)


def calc_result_tag(return_pct):
    if return_pct is None:
        return "未出清"
    if return_pct > 0:
        return "勝"
    if return_pct < 0:
        return "敗"
    return "平手"


def match_target(name):
    for label, pat in TARGET_PATTERNS.items():
        if re.search(pat, name):
            return label
    return ""
















def safe_price_float(x):
    try:
        s = str(x).replace(",", "").replace("--", "").replace("X", "").strip()

        if s in ["", "-", "---", "除權息", "None", "nan", "null"]:
            return None

        v = float(s)

        # 權證 / 股價不應該用 0 當有效收盤價。
        # 測試時發現部分權證會回傳 0.0，不能拿來計算 D+。
        if v <= 0:
            return None

        return v
    except Exception:
        return None


def merge_price_dicts(*dicts):
    merged = {}

    for prices in dicts:
        if not prices:
            continue

        for d, p in prices.items():
            if p is not None and p > 0:
                merged[d] = p

    return merged


def normalize_price_code(code):
    s = str(code).strip()

    if s.endswith(".0"):
        s = s[:-2]

    s = "".join(ch for ch in s if ch.isdigit())

    if not s:
        return ""

    # 權證通常為 6 碼；股票通常為 4 碼。
    # 若是 5 碼權證，很可能是 Excel / pandas 吃掉前導 0，補回 6 碼。
    if len(s) == 5:
        return s.zfill(6)

    return s















def configured_broker_labels_for_scope(scope="all"):
    scope = str(scope or "all").strip().lower()
    if scope == "selected5":
        return {
            label for label in parse_selected_target_labels()
            if label in FULL_TARGET_PATTERNS
        }
    return set(FULL_TARGET_PATTERNS.keys())


def configured_broker_codes_for_scope(scope="all"):
    labels = configured_broker_labels_for_scope(scope)
    return {
        str(FULL_FALLBACK[label][1]).strip()
        for label in labels
        if label in FULL_FALLBACK and str(FULL_FALLBACK[label][1]).strip()
    }


def normalize_broker_code_for_compare(value):
    """將券商代號轉成穩定比較格式，避免大小寫或空白造成誤判。"""
    return str(value or "").strip().upper()


def configured_broker_pair_maps_for_scope(scope="all"):
    """
    回傳目前有效分點的名稱／券商代號雙向對照。

    只有仍存在於 TARGET_PATTERNS 範圍內，且 FALLBACK 有券商代號的分點，
    才會被視為有效分點。這可讓 Google Sheet 舊資料在分點被刪除後同步移除；
    若只是分點標籤改名但券商代號相同，則會自動改成目前的新標籤。
    """
    labels = configured_broker_labels_for_scope(scope)
    label_to_code = {}
    code_to_label = {}

    for label in labels:
        if label not in FULL_FALLBACK:
            continue

        canonical_code = str(FULL_FALLBACK[label][1]).strip()
        normalized_code = normalize_broker_code_for_compare(canonical_code)
        if not normalized_code:
            continue

        canonical_label = str(label).strip()
        label_to_code[canonical_label] = canonical_code
        code_to_label[normalized_code] = (canonical_label, canonical_code)

    return label_to_code, code_to_label




def recent_trading_date_cutoff_from_series(series, keep_days):
    keep_days = max(int(keep_days or 1), 1)
    parsed = pd.to_datetime(series.astype(str).str.replace("/", "-", regex=False), errors="coerce")
    valid_dates = sorted({d.date() for d in parsed.dropna().tolist()}, reverse=True)
    if not valid_dates:
        return None
    return valid_dates[min(keep_days - 1, len(valid_dates) - 1)]


def prune_price_cache_dataframe(df):
    stats = {"before": 0, "after": 0, "removed": 0, "cutoff": ""}
    if df is None or df.empty:
        return pd.DataFrame(columns=["代號", "日期", "收盤價"]), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    required = {"代號", "日期", "收盤價"}
    if not required.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    out["_parsed_date"] = pd.to_datetime(
        out["日期"].astype(str).str.replace("/", "-", regex=False),
        errors="coerce",
    )
    out = out[out["_parsed_date"].notna()].copy()
    cutoff = recent_trading_date_cutoff_from_series(out["日期"], PRICE_RETENTION_TRADING_DAYS)
    if cutoff is not None:
        out = out[out["_parsed_date"].dt.date >= cutoff].copy()
        stats["cutoff"] = cutoff.strftime("%Y/%m/%d")

    out["代號"] = out["代號"].map(normalize_price_code)
    out = out[out["代號"].astype(str).str.strip() != ""].copy()
    out["日期"] = out["_parsed_date"].dt.strftime("%Y/%m/%d")
    out["收盤價"] = pd.to_numeric(out["收盤價"], errors="coerce")
    out = out[out["收盤價"].fillna(0) > 0].copy()
    out = out.drop(columns=["_parsed_date"], errors="ignore")
    out = out.drop_duplicates(subset=["代號", "日期"], keep="last")
    out = out.sort_values(["代號", "日期"]).reset_index(drop=True)
    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out[["代號", "日期", "收盤價"]], stats


def prune_history_cache_dataframe(df):
    stats = {
        "before": 0,
        "after": 0,
        "removed": 0,
        "removed_broker": 0,
        "removed_date": 0,
        "cutoff": "",
    }
    if df is None or df.empty:
        return pd.DataFrame(), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    required = {"分點", "券商代號", "日期"}
    if not required.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    allowed_labels = configured_broker_labels_for_scope("all")
    allowed_codes = configured_broker_codes_for_scope("all")
    label_ok = out["分點"].astype(str).str.strip().isin(allowed_labels)
    code_ok = out["券商代號"].astype(str).str.strip().isin(allowed_codes)
    broker_mask = label_ok | code_ok
    stats["removed_broker"] = int((~broker_mask).sum())
    out = out[broker_mask].copy()

    out["_parsed_date"] = pd.to_datetime(
        out["日期"].astype(str).str.replace("/", "-", regex=False),
        errors="coerce",
    )
    invalid_date_count = int(out["_parsed_date"].isna().sum())
    out = out[out["_parsed_date"].notna()].copy()

    cutoff = recent_trading_date_cutoff_from_series(out["日期"], HISTORY_RETENTION_TRADING_DAYS)
    before_date_filter = len(out)
    if cutoff is not None:
        out = out[out["_parsed_date"].dt.date >= cutoff].copy()
        stats["cutoff"] = cutoff.strftime("%Y/%m/%d")
    stats["removed_date"] = invalid_date_count + max(before_date_filter - len(out), 0)

    out["日期"] = out["_parsed_date"].dt.strftime("%Y/%m/%d")
    out = out.drop(columns=["_parsed_date"], errors="ignore")
    if {"權證代號", "券商代號", "日期"}.issubset(set(out.columns)):
        out = out.drop_duplicates(
            subset=["權證代號", "券商代號", "日期"],
            keep="last",
        )
        out = out.sort_values(["權證代號", "券商代號", "日期"]).reset_index(drop=True)

    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out, stats




def prune_broker_map_dataframe(df, scope=None):
    stats = {"before": 0, "after": 0, "removed": 0}
    if df is None or df.empty:
        return pd.DataFrame(), stats

    out = df.copy().fillna("")
    stats["before"] = len(out)
    if not {"分點", "券商代號"}.issubset(set(out.columns)):
        stats["after"] = len(out)
        return out, stats

    scope = scope or ("selected5" if RUN_MODE == 1 else "all")
    labels = configured_broker_labels_for_scope(scope)
    codes = configured_broker_codes_for_scope(scope)
    mask = out["分點"].astype(str).str.strip().isin(labels) | out["券商代號"].astype(str).str.strip().isin(codes)
    out = out[mask].copy()
    out = out.drop_duplicates(subset=["分點"], keep="last").reset_index(drop=True)
    stats["after"] = len(out)
    stats["removed"] = max(stats["before"] - stats["after"], 0)
    return out, stats

def price_code_variants(code):
    code = normalize_price_code(code)

    if not code:
        return []

    variants = [code]

    no_zero = code.lstrip("0")
    if no_zero and no_zero != code:
        variants.append(no_zero)

    out = []
    for v in variants:
        if v not in out:
            out.append(v)

    return out


def yahoo_symbol_variants(code):
    symbols = []

    for c in price_code_variants(code):
        symbols.append(f"{c}.TW")
        symbols.append(f"{c}.TWO")

    out = []
    for s in symbols:
        if s not in out:
            out.append(s)

    return out


def fetch_twse_stock_day_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_dt in iter_month_starts(start_dt, end_dt):
        try:
            session = get_thread_session()
            rp = session.get(
                f"https://www.twse.com.tw/exchangeReport/STOCK_DAY"
                f"?response=json&date={month_dt}&stockNo={code}",
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=8
            )
            data = rp.json()

            for row in data.get("data", []):
                try:
                    parts = str(row[0]).split("/")
                    dk = f"{int(parts[0]) + 1911}/{int(parts[1]):02d}/{int(parts[2]):02d}"
                    close_price = safe_price_float(row[6])

                    if close_price is not None:
                        prices[dk] = close_price
                except Exception:
                    pass
        except Exception:
            pass

    return prices


def fetch_tpex_new_trading_stock_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_start in iter_month_starts(start_dt, end_dt):
        try:
            month_dt = datetime.strptime(month_start, "%Y%m%d")
            date_str = month_dt.strftime("%Y/%m/01")

            urls = [
                (
                    "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock"
                    f"?code={code}&date={date_str}&response=json"
                ),
                (
                    "https://www.tpex.org.tw/www/zh-tw/afterTrading/tradingStock"
                    f"?code={code}&date={date_str}&type=EW&response=json"
                ),
            ]

            for url in urls:
                try:
                    session = get_thread_session()
                    rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
                    data = rp.json()

                    rows = []

                    if isinstance(data, dict):
                        if isinstance(data.get("tables"), list):
                            for table in data.get("tables", []):
                                if isinstance(table, dict) and isinstance(table.get("data"), list):
                                    rows.extend(table.get("data", []))

                        if isinstance(data.get("data"), list):
                            rows.extend(data.get("data", []))

                        if isinstance(data.get("aaData"), list):
                            rows.extend(data.get("aaData", []))

                    for row in rows:
                        try:
                            if not isinstance(row, (list, tuple)) or not row:
                                continue

                            raw_date = str(row[0]).strip().replace("-", "/")
                            parts = raw_date.split("/")

                            if len(parts) != 3:
                                continue

                            y = int(parts[0])

                            if y < 1911:
                                y += 1911

                            dk = f"{y}/{int(parts[1]):02d}/{int(parts[2]):02d}"

                            close_price = None

                            # 測試結果顯示 TPEx 新版 tradingStock 是 70xxxx 權證與上櫃股最穩來源。
                            # 常見收盤價欄位落在 6 / 5 / 4 / 7 / 3，逐一嘗試。
                            for idx in [6, 5, 4, 7, 3]:
                                if idx < len(row):
                                    v = safe_price_float(row[idx])

                                    if v is not None:
                                        close_price = v
                                        break

                            if close_price is not None:
                                prices[dk] = close_price
                        except Exception:
                            pass

                except Exception:
                    pass

        except Exception:
            pass

    return prices


def fetch_tpex_old_st43_prices(code, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}
    code = normalize_price_code(code)

    if not code:
        return prices

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    for month_start in iter_month_starts(start_dt, end_dt):
        try:
            month_dt = datetime.strptime(month_start, "%Y%m%d")
            roc_month = f"{month_dt.year - 1911}/{month_dt.month:02d}"

            url = (
                "https://www.tpex.org.tw/web/stock/aftertrading/daily_trading_info/"
                f"st43_result.php?l=zh-tw&d={roc_month}&stkno={code}"
            )

            session = get_thread_session()
            rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            data = rp.json()

            rows = data.get("aaData", []) if isinstance(data, dict) else []

            for row in rows:
                try:
                    if not isinstance(row, (list, tuple)) or not row:
                        continue

                    raw_date = str(row[0]).strip().replace("-", "/")
                    parts = raw_date.split("/")

                    if len(parts) != 3:
                        continue

                    y = int(parts[0])

                    if y < 1911:
                        y += 1911

                    dk = f"{y}/{int(parts[1]):02d}/{int(parts[2]):02d}"

                    close_price = None

                    for idx in [6, 5, 4, 7, 3]:
                        if idx < len(row):
                            v = safe_price_float(row[idx])

                            if v is not None:
                                close_price = v
                                break

                    if close_price is not None:
                        prices[dk] = close_price
                except Exception:
                    pass
        except Exception:
            pass

    return prices


def fetch_yahoo_chart_prices(symbol, start_dt=None, end_dt=None, host="query1"):
    today = datetime.today()
    prices = {}

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    period1 = int(start_dt.timestamp())
    period2 = int((end_dt + timedelta(days=1)).timestamp())

    url = (
        f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}"
        f"?period1={period1}&period2={period2}&interval=1d&events=history"
    )

    try:
        session = get_thread_session()
        rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
        data = rp.json()

        result = data.get("chart", {}).get("result", [])

        if not result:
            return prices

        result = result[0]
        timestamps = result.get("timestamp", [])
        quote = result.get("indicators", {}).get("quote", [{}])[0]
        closes = quote.get("close", [])

        for ts, close_price in zip(timestamps, closes):
            v = safe_price_float(close_price)

            if v is None:
                continue

            dt = datetime.fromtimestamp(int(ts))
            prices[dt.strftime("%Y/%m/%d")] = v
    except Exception:
        pass

    return prices


def fetch_yahoo_range_prices(symbol):
    prices = {}

    # 只在官方來源不足時才會進入 Yahoo，這裡用 5y + max 做快速備援。
    for range_value in ["5y", "max"]:
        url = (
            f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}"
            f"?range={range_value}&interval=1d&events=history"
        )

        try:
            session = get_thread_session()
            rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            data = rp.json()

            result = data.get("chart", {}).get("result", [])

            if not result:
                continue

            result = result[0]
            timestamps = result.get("timestamp", [])
            quote = result.get("indicators", {}).get("quote", [{}])[0]
            closes = quote.get("close", [])

            for ts, close_price in zip(timestamps, closes):
                v = safe_price_float(close_price)

                if v is None:
                    continue

                dt = datetime.fromtimestamp(int(ts))
                prices[dt.strftime("%Y/%m/%d")] = v

            if prices:
                break
        except Exception:
            pass

    return prices


def fetch_yahoo_download_prices(symbol, start_dt=None, end_dt=None):
    today = datetime.today()
    prices = {}

    if start_dt is None:
        start_dt = add_months(datetime(today.year, today.month, 1), -13)
    if end_dt is None:
        end_dt = today

    if end_dt > today:
        end_dt = today

    if start_dt > end_dt:
        start_dt = end_dt

    period1 = int(start_dt.timestamp())
    period2 = int((end_dt + timedelta(days=1)).timestamp())

    url = (
        f"https://query1.finance.yahoo.com/v7/finance/download/{symbol}"
        f"?period1={period1}&period2={period2}&interval=1d&events=history&includeAdjustedClose=true"
    )

    try:
        session = get_thread_session()
        rp = session.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)

        if rp.status_code != 200 or "Date" not in rp.text:
            return prices

        df = pd.read_csv(StringIO(rp.text))

        if "Date" not in df.columns or "Close" not in df.columns:
            return prices

        for _, row in df.iterrows():
            dt = parse_date(row["Date"])
            close_price = safe_price_float(row["Close"])

            if dt and close_price is not None:
                prices[dt.strftime("%Y/%m/%d")] = close_price
    except Exception:
        pass

    return prices


def fetch_yahoo_prices(code, start_dt=None, end_dt=None):
    """
    Yahoo 備援：
    1. 同時測補零版與去零版，例如 064390.TW / 64390.TW。
    2. 同時測 .TW / .TWO。
    3. 優先 query1 chart，其次 query2、range、download。
    """
    prices = {}

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_chart_prices(symbol, start_dt, end_dt, host="query1")

        if p:
            return p

        p = fetch_yahoo_chart_prices(symbol, start_dt, end_dt, host="query2")

        if p:
            return p

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_range_prices(symbol)

        if p:
            return p

    for symbol in yahoo_symbol_variants(code):
        p = fetch_yahoo_download_prices(symbol, start_dt, end_dt)

        if p:
            return p

    return prices


def prices_need_yahoo_fallback(prices, start_dt=None, end_dt=None):
    """
    判斷官方來源價格是否不足，需要 Yahoo 備援。

    只用 len(prices) < 2 不夠，因為有些來源雖然有幾筆，
    但覆蓋不到 D+1 ~ D+20 的日期，Excel 還是會出現大量 權:- / 標:-。
    """
    if not prices:
        return True

    valid_dates = sorted([d for d, p in prices.items() if p is not None and p > 0])

    if not valid_dates:
        return True

    # 如果測試區間超過 30 天，但有效價格少於 10 筆，代表覆蓋率明顯不足。
    if start_dt and end_dt:
        try:
            span_days = (end_dt - start_dt).days
            if span_days >= 30 and len(valid_dates) < 10:
                return True
        except Exception:
            pass

    # 如果最後一筆價格離需要的結束日太遠，也要補 Yahoo。
    if end_dt:
        try:
            latest_dt = parse_date(valid_dates[-1])
            target_end = min(end_dt, datetime.today())

            if latest_dt and (target_end - latest_dt).days > 10:
                return True
        except Exception:
            pass

    return False


def fetch_twse_prices(code, start_dt=None, end_dt=None):
    """
    統一價格抓取函式（保留原函式名稱，避免改動其他流程）：

    速度與準確率策略：
    1. 先用官方資料源：
       - 上市股票 / 上市權證：TWSE STOCK_DAY
       - 上櫃股票 / 上櫃權證：TPEx tradingStock
    2. 若官方來源覆蓋率不足，再補 Yahoo。
    3. 價格 <= 0 一律不採用。
    4. 權證 5 碼會自動補回 6 碼，避免前導 0 被吃掉。
    """
    code = normalize_price_code(code)

    if not code:
        return {}

    prices = {}

    # 先根據代號型態決定優先順序，減少不必要請求。
    # 70xxxx 權證大多走 TPEx；0xxxxx 權證大多走 TWSE。
    if len(code) == 6 and code.startswith("7"):
        prices = merge_price_dicts(
            prices,
            fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_old_st43_prices(code, start_dt, end_dt)
            )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_twse_stock_day_prices(code, start_dt, end_dt)
            )

    elif len(code) == 6 and code.startswith("0"):
        prices = merge_price_dicts(
            prices,
            fetch_twse_stock_day_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
            )

    else:
        prices = merge_price_dicts(
            prices,
            fetch_twse_stock_day_prices(code, start_dt, end_dt)
        )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_new_trading_stock_prices(code, start_dt, end_dt)
            )

        if prices_need_yahoo_fallback(prices, start_dt, end_dt):
            prices = merge_price_dicts(
                prices,
                fetch_tpex_old_st43_prices(code, start_dt, end_dt)
            )

    # 官方來源覆蓋率不足時才用 Yahoo，兼顧速度與完整度。
    if prices_need_yahoo_fallback(prices, start_dt, end_dt):
        prices = merge_price_dicts(
            prices,
            fetch_yahoo_prices(code, start_dt, end_dt)
        )

    return prices


# 單日全市場收盤價：TWSE／TPEx 各一個請求；逐檔月檔 API 只保留作補漏備援。
_MARKET_CLOSE_PRICE_CACHE_LOCK = threading.Lock()
_MARKET_CLOSE_PRICE_CACHE = {}
_MARKET_CLOSE_SEEN_CODE_CACHE = {}
_MARKET_CLOSE_DAY_LOCKS = {}
_MARKET_CLOSE_DAY_LOCKS_GUARD = threading.Lock()

# 價格序列日期索引：以 dict 物件身分為主鍵，同一序列只排序一次。
_SORTED_PRICE_DATES_CACHE_LOCK = threading.Lock()
_SORTED_PRICE_DATES_CACHE = {}
_SORTED_PRICE_DATES_CACHE_MAXSIZE = 8192


def _clean_market_cell_text(value):
    text = re.sub(r"<[^>]*>", "", str(value or ""))
    return text.replace("&nbsp;", " ").replace("\u3000", " ").strip()


def _normalize_market_header(value):
    return re.sub(r"[\s\u3000]+", "", _clean_market_cell_text(value)).lower()


def _market_header_index(headers, kind):
    normalized = [_normalize_market_header(header) for header in (headers or [])]
    if kind == "code":
        exact = {
            "證券代號", "股票代號", "商品代號", "代號", "證券代碼", "股票代碼",
            "securitycode", "stockid", "stock_id", "code",
        }
        contains = ("證券代號", "股票代號", "商品代號")
    else:
        exact = {
            "收盤價", "收盤", "收市價", "成交價", "close", "closingprice",
            "closeprice", "closing_price",
        }
        contains = ("收盤價", "收市價")

    for idx, header in enumerate(normalized):
        if header in exact:
            return idx
    for idx, header in enumerate(normalized):
        if any(token in header for token in contains):
            return idx
    return None


def _extract_market_close_prices_from_payload(payload):
    """從 TWSE／TPEx 多種 JSON 表格格式擷取「代號、收盤價」，並保留所有看見的代號。"""
    prices = {}
    seen_codes = set()
    visited = set()

    def add_record(record):
        if not isinstance(record, dict):
            return
        headers = list(record.keys())
        code_idx = _market_header_index(headers, "code")
        close_idx = _market_header_index(headers, "close")
        if code_idx is None:
            return
        values = list(record.values())
        code = normalize_price_code(_clean_market_cell_text(values[code_idx]))
        if not code:
            return
        seen_codes.add(code)
        if close_idx is None:
            return
        price = safe_price_float(_clean_market_cell_text(values[close_idx]))
        if price is not None:
            prices[code] = price

    def add_table(headers, rows):
        if not isinstance(headers, list) or not isinstance(rows, list):
            return
        code_idx = _market_header_index(headers, "code")
        close_idx = _market_header_index(headers, "close")
        if code_idx is None:
            return
        for row in rows:
            if isinstance(row, dict):
                add_record(row)
                continue
            if not isinstance(row, (list, tuple)) or code_idx >= len(row):
                continue
            code = normalize_price_code(_clean_market_cell_text(row[code_idx]))
            if not code:
                continue
            seen_codes.add(code)
            if close_idx is None or close_idx >= len(row):
                continue
            price = safe_price_float(_clean_market_cell_text(row[close_idx]))
            if price is not None:
                prices[code] = price

    def visit(obj):
        if isinstance(obj, (dict, list)):
            obj_id = id(obj)
            if obj_id in visited:
                return
            visited.add(obj_id)

        if isinstance(obj, dict):
            # 單筆 dict record。
            add_record(obj)

            # 標準 fields/data 與 TWSE fields1/data1、fields2/data2...。
            if isinstance(obj.get("fields"), list) and isinstance(obj.get("data"), list):
                add_table(obj.get("fields"), obj.get("data"))
            for key, headers in obj.items():
                match = re.fullmatch(r"fields(\d*)", str(key or ""), flags=re.IGNORECASE)
                if not match or not isinstance(headers, list):
                    continue
                rows = obj.get(f"data{match.group(1)}")
                if isinstance(rows, list):
                    add_table(headers, rows)

            # TPEx tables 陣列及其他巢狀容器。
            for value in obj.values():
                if isinstance(value, (dict, list)):
                    visit(value)

        elif isinstance(obj, list):
            for value in obj:
                if isinstance(value, dict):
                    add_record(value)
                if isinstance(value, (dict, list)):
                    visit(value)

    visit(payload)
    return prices, seen_codes


def _market_close_day_lock(date_key):
    with _MARKET_CLOSE_DAY_LOCKS_GUARD:
        lock = _MARKET_CLOSE_DAY_LOCKS.get(date_key)
        if lock is None:
            lock = threading.Lock()
            _MARKET_CLOSE_DAY_LOCKS[date_key] = lock
        return lock


def _fetch_market_close_snapshot_for_date(target_date):
    """同一天最多抓一次 TWSE＋TPEx；回傳有效價格與該日市場回應曾出現的代號。"""
    target_dt = parse_date(target_date)
    if not target_dt:
        return {}, set()
    target_key = target_dt.strftime("%Y/%m/%d")

    with _MARKET_CLOSE_PRICE_CACHE_LOCK:
        if target_key in _MARKET_CLOSE_PRICE_CACHE:
            return (
                dict(_MARKET_CLOSE_PRICE_CACHE[target_key]),
                set(_MARKET_CLOSE_SEEN_CODE_CACHE.get(target_key, set())),
            )

    with _market_close_day_lock(target_key):
        with _MARKET_CLOSE_PRICE_CACHE_LOCK:
            if target_key in _MARKET_CLOSE_PRICE_CACHE:
                return (
                    dict(_MARKET_CLOSE_PRICE_CACHE[target_key]),
                    set(_MARKET_CLOSE_SEEN_CODE_CACHE.get(target_key, set())),
                )

        endpoints = [
            (
                "TWSE",
                "https://www.twse.com.tw/exchangeReport/MI_INDEX",
                {
                    "response": "json",
                    "date": target_dt.strftime("%Y%m%d"),
                    "type": "ALL",
                },
            ),
            (
                "TPEx",
                "https://www.tpex.org.tw/www/zh-tw/afterTrading/otc",
                {
                    "date": target_dt.strftime("%Y/%m/%d"),
                    "response": "json",
                },
            ),
        ]

        def fetch_one(endpoint):
            market_name, url, params = endpoint
            session = get_thread_session()
            response = session.get(
                url,
                params=params,
                headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json, */*"},
                timeout=(5, 25),
            )
            response.raise_for_status()
            try:
                payload = response.json()
            except Exception:
                payload = json.loads(response.content.decode("utf-8"))
            market_prices, seen_codes = _extract_market_close_prices_from_payload(payload)
            return market_name, market_prices, seen_codes

        all_prices = {}
        all_seen_codes = set()
        market_results = []
        successful_markets = 0

        with ThreadPoolExecutor(max_workers=2) as executor:
            futures = {executor.submit(fetch_one, endpoint): endpoint[0] for endpoint in endpoints}
            for future in as_completed(futures):
                market_name = futures[future]
                try:
                    market_name, market_prices, seen_codes = future.result()
                    all_prices.update(market_prices)
                    all_seen_codes.update(seen_codes)
                    market_results.append(
                        f"{market_name}:{len(market_prices):,}價/{len(seen_codes):,}檔"
                    )
                    if seen_codes:
                        successful_markets += 1
                except Exception as exc:
                    market_results.append(f"{market_name}:失敗")
                    print(
                        f"  ⚠️ {market_name} 全市場收盤價批次抓取失敗：{target_key}｜"
                        f"{type(exc).__name__}: {exc}"
                    )

        # 兩市場皆成功才固定快取；部分成功仍回傳，讓逐檔補漏接手缺口。
        if successful_markets == len(endpoints):
            with _MARKET_CLOSE_PRICE_CACHE_LOCK:
                _MARKET_CLOSE_PRICE_CACHE[target_key] = dict(all_prices)
                _MARKET_CLOSE_SEEN_CODE_CACHE[target_key] = set(all_seen_codes)

        print(
            f"  ✅ 單日全市場收盤價批次：{target_key}｜"
            f"{'｜'.join(sorted(market_results))}｜合併 {len(all_prices):,} 價"
        )
        return dict(all_prices), set(all_seen_codes)


def fetch_market_close_prices_for_date(target_date):
    prices, _seen_codes = _fetch_market_close_snapshot_for_date(target_date)
    return prices


def _normalize_price_fetch_plan(fetch_plan):
    normalized = {}
    for raw_code, raw_range in (fetch_plan or {}).items():
        code = normalize_price_code(raw_code)
        if not code or not raw_range or len(raw_range) < 2:
            continue
        start_dt = parse_date(raw_range[0])
        end_dt = parse_date(raw_range[1])
        if not start_dt or not end_dt:
            continue
        end_dt = min(end_dt, datetime.today())
        if start_dt > end_dt:
            start_dt = end_dt
        normalized[code] = [start_dt, end_dt]
    return normalized


def _price_plan_trading_dates(fetch_plan):
    fetch_plan = _normalize_price_fetch_plan(fetch_plan)
    if not fetch_plan:
        return []
    min_day = min(value[0].date() for value in fetch_plan.values())
    max_day = max(value[1].date() for value in fetch_plan.values())

    official_dates = [
        day for day in get_finmind_trading_dates()
        if min_day <= day <= max_day
    ]
    selected = set(official_dates)

    # 交易日主檔可能尚未包含今天；只對主檔最後日期之後補平日候選。
    fallback_start = (max(official_dates) + timedelta(days=1)) if official_dates else min_day
    current = fallback_start
    while current <= max_day:
        if current.weekday() < 5:
            selected.add(current)
        current += timedelta(days=1)

    return sorted(selected)


def _latest_valid_price_date(prices):
    valid_dates = _sorted_valid_price_dates(prices)
    return valid_dates[-1] if valid_dates else ""


def fetch_price_plan_batch_first(
    price_cache,
    persistent_price_cache,
    fetch_plan,
    label="價格",
    progress_every=10,
):
    """
    先依交易日抓 TWSE／TPEx 全市場收盤價，再只對未出現在批次資料或仍完全缺價的代號逐檔補漏。
    """
    fetch_plan = _normalize_price_fetch_plan(fetch_plan)
    changed_codes = set()
    if not fetch_plan:
        return changed_codes

    trading_dates = _price_plan_trading_dates(fetch_plan)
    batch_seen_codes = set()
    latest_expected_by_code = {}

    print(
        f"  {label}全市場批次價：{len(trading_dates):,} 個交易日｜"
        f"最多 {len(trading_dates) * 2:,} 個市場請求"
    )

    for date_index, trade_day in enumerate(trading_dates, start=1):
        target_key = trade_day.strftime("%Y/%m/%d")
        wanted_codes = {
            code
            for code, (start_dt, end_dt) in fetch_plan.items()
            if start_dt.date() <= trade_day <= end_dt.date()
        }
        if not wanted_codes:
            continue

        market_prices, seen_codes = _fetch_market_close_snapshot_for_date(target_key)
        batch_seen_codes.update(wanted_codes & seen_codes)
        for code in wanted_codes:
            latest_expected_by_code[code] = target_key
            price = safe_price_float(market_prices.get(code))
            if price is None:
                continue
            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            old_price = safe_price_float(old_prices.get(target_key))
            merged_prices = merge_price_dicts(old_prices, {target_key: price})
            persistent_price_cache[code] = merged_prices
            add_price_aliases(price_cache, code, merged_prices)
            if old_price != price:
                changed_codes.add(code)

        if progress_every and date_index % max(int(progress_every), 1) == 0:
            print(f"  [{date_index}/{len(trading_dates)}] {label}全市場批次價處理中...")

    fallback_plan = {}
    for code, date_range in fetch_plan.items():
        prices = get_cached_prices_for_code(persistent_price_cache, code)
        latest_date = _latest_valid_price_date(prices)
        expected_date = latest_expected_by_code.get(code, "")
        need_fallback = not prices or not latest_date or code not in batch_seen_codes

        # 股票通常每天都有收盤價；若批次漏掉需求區間最後交易日，逐檔補一次。
        # 權證若該日無成交但全市場回應已看見代號，不重複逐檔打 API。
        if len(code) == 4 and expected_date and latest_date != expected_date:
            need_fallback = True

        if need_fallback:
            fallback_plan[code] = date_range
        elif prices:
            add_price_aliases(price_cache, code, prices)

    print(
        f"  {label}批次價命中：{len(fetch_plan) - len(fallback_plan):,}/{len(fetch_plan):,} 檔｜"
        f"逐檔補漏：{len(fallback_plan):,} 檔"
    )

    if not fallback_plan:
        return changed_codes

    def fetch_one(code):
        start_dt, end_dt = fallback_plan[code]
        return code, fetch_twse_prices(code, start_dt, end_dt)

    done = 0
    with ThreadPoolExecutor(max_workers=PRICE_WORKERS) as executor:
        futures = {executor.submit(fetch_one, code): code for code in fallback_plan}
        for future in as_completed(futures):
            done += 1
            code = futures[future]
            try:
                code, fetched_prices = future.result()
            except Exception:
                fetched_prices = {}

            old_prices = get_cached_prices_for_code(persistent_price_cache, code)
            merged_prices = merge_price_dicts(old_prices, fetched_prices)
            if merged_prices:
                persistent_price_cache[code] = merged_prices
                add_price_aliases(price_cache, code, merged_prices)
                if fetched_prices:
                    changed_codes.add(code)
            else:
                add_price_aliases(price_cache, code, old_prices)

            if done % 20 == 0:
                print(f"  [{done}/{len(fallback_plan)}] {label}逐檔補漏中...")

    return changed_codes


def _price_dates_cache_signature(prices):
    try:
        length = len(prices)
        if length == 0:
            return (0, None, None, None, None)
        first_key = next(iter(prices))
        last_key = next(reversed(prices))
        return (
            length,
            first_key,
            last_key,
            prices.get(first_key),
            prices.get(last_key),
        )
    except Exception:
        return (len(prices) if prices else 0, None, None, None, None)


def _sorted_valid_price_dates(prices):
    if not prices:
        return []

    cache_key = id(prices)
    signature = _price_dates_cache_signature(prices)
    with _SORTED_PRICE_DATES_CACHE_LOCK:
        cached = _SORTED_PRICE_DATES_CACHE.get(cache_key)
        if cached is not None and cached[0] is prices and cached[1] == signature:
            return cached[2]

    valid_dates = sorted(
        normalize_date_str(d)
        for d, p in prices.items()
        if parse_date(d) and safe_price_float(p) is not None
    )

    with _SORTED_PRICE_DATES_CACHE_LOCK:
        if len(_SORTED_PRICE_DATES_CACHE) >= _SORTED_PRICE_DATES_CACHE_MAXSIZE:
            _SORTED_PRICE_DATES_CACHE.clear()
        _SORTED_PRICE_DATES_CACHE[cache_key] = (prices, signature, valid_dates)

    return valid_dates


def get_price_nearest(prices, date):
    date = normalize_date_str(date)
    if date in prices and safe_price_float(prices.get(date)) is not None:
        return prices[date]

    valid_dates = _sorted_valid_price_dates(prices)
    idx = bisect_right(valid_dates, date) - 1
    if idx < 0:
        return None
    return prices.get(valid_dates[idx])







# ══════════════════════════════════════════════════════════════════════
# Google Sheet 快取 / 結果同步工具（GitHub Actions 部署用）
# ══════════════════════════════════════════════════════════════════════

GOOGLE_SHEET_NAME = os.getenv("GOOGLE_SHEET_NAME", "權證分點資料_NEW")
GOOGLE_SHEET_ID = os.getenv("GOOGLE_SHEET_ID", os.getenv("GSHEET_ID", "")).strip()
GSHEET_CACHE_ENABLED = os.getenv("GSHEET_CACHE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_RESULT_ENABLED = os.getenv("GSHEET_RESULT_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_CHUNK_ROWS = int(os.getenv("GSHEET_CHUNK_ROWS", "3000"))

# Google Sheet 結果表自動封存 / 保留設定：
# 1. 每日賣出明細預設只在主試算表保留最近 60 個實際交易日，可改成 30 / 60 / 90 或其他正整數。
# 2. TOP15 兩張快取表預設各資料範圍保留最近 5 個統計日期。
# 3. 超出保留範圍的舊資料會先寫入獨立的年度封存試算表，封存成功後才從主表移除。
# 4. 封存試算表依年份自動建立；單一年度封存檔接近儲存格上限時，會自動切換到 _02、_03...。
# 5. 同步前會把所有工作表縮到實際有內容的列數 / 欄數；後續每張寫入表也會直接 resize 成實際大小。
GSHEET_RESULT_ARCHIVE_ENABLED = os.getenv("GSHEET_RESULT_ARCHIVE_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_ARCHIVE_NAME_PREFIX = os.getenv(
    "GSHEET_ARCHIVE_NAME_PREFIX",
    f"{GOOGLE_SHEET_NAME}_歷史封存",
).strip() or f"{GOOGLE_SHEET_NAME}_歷史封存"
GSHEET_ARCHIVE_SHARE_EMAILS = os.getenv("GSHEET_ARCHIVE_SHARE_EMAILS", "").strip()
# 若服務帳號需要在指定的 Shared Drive / 資料夾內建立封存檔，可填 Google Drive 資料夾 ID。
GSHEET_ARCHIVE_FOLDER_ID = os.getenv("GSHEET_ARCHIVE_FOLDER_ID", "").strip()
# 服務帳號通常沒有可用的個人 Drive 儲存空間；若 gc.create() 回傳
#「The user's Drive storage quota has been exceeded」，請先用自己的 Google 帳號建立一份
# 封存試算表並分享給服務帳號，再把試算表 ID 填入下列環境變數。
# 可填單一 ID，也可用逗號 / 分號分隔多個 ID；第一份接近格數上限時會自動換到下一份。
GSHEET_ARCHIVE_SPREADSHEET_IDS_RAW = os.getenv(
    "GSHEET_ARCHIVE_SPREADSHEET_IDS",
    os.getenv("GSHEET_ARCHIVE_SPREADSHEET_ID", ""),
).strip()
GSHEET_ARCHIVE_SPREADSHEET_IDS = [
    item.strip()
    for item in re.split(r"[,;；、\n\r\t]+", GSHEET_ARCHIVE_SPREADSHEET_IDS_RAW)
    if item.strip()
]
# 未提供既有封存試算表 ID 時，是否允許服務帳號自行建立新檔。
# 若 Drive quota 不足，程式會在第一次 403 後自動停用本次執行的建立動作，不再連續嘗試 _02～_99。
GSHEET_ARCHIVE_ALLOW_CREATE = os.getenv("GSHEET_ARCHIVE_ALLOW_CREATE", "1").strip().lower() not in ("0", "false", "no")
GSHEET_ARCHIVE_MAX_CELLS = max(int(os.getenv("GSHEET_ARCHIVE_MAX_CELLS", "8500000")), 100000)
GSHEET_ARCHIVE_YEARLY_SPLIT = os.getenv("GSHEET_ARCHIVE_YEARLY_SPLIT", "1").strip().lower() not in ("0", "false", "no")
GSHEET_TOP15_KEEP_STAT_DATES = max(int(os.getenv("GSHEET_TOP15_KEEP_STAT_DATES", "5")), 1)
GSHEET_DAILY_SELL_KEEP_TRADING_DAYS = max(int(os.getenv("GSHEET_DAILY_SELL_KEEP_TRADING_DAYS", "30")), 1)
GSHEET_COMPACT_BLANK_GRID_ENABLED = os.getenv("GSHEET_COMPACT_BLANK_GRID_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_COMPACT_MIN_ROWS = max(int(os.getenv("GSHEET_COMPACT_MIN_ROWS", "1")), 1)
GSHEET_COMPACT_MIN_COLS = max(int(os.getenv("GSHEET_COMPACT_MIN_COLS", "1")), 1)

# Google Sheets API 有「每分鐘寫入請求」限制。
# 結果工作表很多、又要同步格式時，如果沒有節流與 429 重試，
# 會出現後面工作表建立 / 寫入失敗，甚至因先刪後建導致工作表消失。
GSHEET_WRITE_SLEEP_SECONDS = float(os.getenv("GSHEET_WRITE_SLEEP_SECONDS", "0.05"))
GSHEET_WRITE_RATE_PER_MINUTE = max(float(os.getenv("GSHEET_WRITE_RATE_PER_MINUTE", "45")), 1.0)
GSHEET_WRITE_BURST = max(float(os.getenv("GSHEET_WRITE_BURST", "5")), 1.0)
GSHEET_MAX_RETRIES = int(os.getenv("GSHEET_MAX_RETRIES", "6"))
GSHEET_RETRY_BASE_SECONDS = float(os.getenv("GSHEET_RETRY_BASE_SECONDS", "12"))

_GSHEET_CLIENT = None
_GSHEET_SPREADSHEET = None
_GSHEET_LAST_WRITE_TS = 0.0
_GSHEET_TOKEN_BUCKET_TOKENS = GSHEET_WRITE_BURST
_GSHEET_TOKEN_BUCKET_UPDATED_TS = time.monotonic()
_GSHEET_TOKEN_BUCKET_LOCK = threading.Lock()
_GSHEET_ARCHIVE_SPREADSHEETS = {}
_GSHEET_ARCHIVE_PERMISSION_SYNCED = set()
_GSHEET_ARCHIVE_CREATE_BLOCKED = False
_GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED = False

CACHE_SHEET_NAME_MAP = {
    "warrants_cache.csv": "快取_權證清單",
    "broker_map_cache.csv": "快取_分點代號",
    "candidates_cache.csv": "快取_候選組合",
    "candidates_cache_selected5.csv": "快取_候選組合_精選5",
    "broker_warrant_history_cache.csv": "快取_分點歷史",
    "price_cache.csv": "快取_價格",
    "price_prefetch_state.csv": "快取_價格預抓狀態",
    "prescan_status.csv": "快取_候選掃描狀態",
    "prescan_refresh_keys.csv": "快取_候選掃描Keys",
}


# ══════════════════════════════════════════════════════════════════════
# Supabase 快取同步工具（大型快取分流用）
# ══════════════════════════════════════════════════════════════════════

SUPABASE_CACHE_ENABLED = os.getenv("SUPABASE_CACHE_ENABLED", "0").strip().lower() in ("1", "true", "yes")
SUPABASE_DB_URL = os.getenv("SUPABASE_DB_URL", "").strip()
SUPABASE_SCHEMA = os.getenv("SUPABASE_SCHEMA", "warrant_cache").strip() or "warrant_cache"
SUPABASE_BATCH_SIZE = int(os.getenv("SUPABASE_BATCH_SIZE", "5000"))
# 啟用 Supabase 後，大型快取預設不再同步到 Google Sheet，避免 1,000 萬格上限。
# Excel / Google Sheet 結果表仍照原本流程同步，不受影響。

_SUPABASE_AVAILABLE = None
_SUPABASE_EMPTY_CACHE_KEYS = set()


def supabase_enabled():
    return bool(SUPABASE_CACHE_ENABLED and SUPABASE_DB_URL)


def supabase_import_psycopg():
    global _SUPABASE_AVAILABLE

    if not supabase_enabled():
        return None

    if _SUPABASE_AVAILABLE is False:
        return None

    try:
        import psycopg
        _SUPABASE_AVAILABLE = True
        return psycopg
    except Exception as e:
        _SUPABASE_AVAILABLE = False
        print(f"  ⚠️ Supabase 快取停用：無法載入 psycopg，原因：{type(e).__name__}: {e}")
        return None


def get_supabase_conn():
    psycopg = supabase_import_psycopg()
    if psycopg is None:
        return None

    try:
        return psycopg.connect(
            SUPABASE_DB_URL,
            connect_timeout=20,
            autocommit=True,
            prepare_threshold=None,
        )
    except Exception as e:
        print(f"  ⚠️ Supabase 連線失敗：{type(e).__name__}: {e}")
        return None


def supabase_cache_kind_and_scope(path):
    base = os.path.basename(str(path))

    if base == "price_cache.csv":
        return "price", ""
    if base == "broker_warrant_history_cache.csv":
        return "history", ""
    if base == "warrants_cache.csv":
        return "warrants", ""
    if base == "broker_map_cache.csv":
        return "broker_map", "selected5" if RUN_MODE == 1 else "all"

    return "", ""


def supabase_cache_supported(path):
    kind, _ = supabase_cache_kind_and_scope(path)
    return bool(kind)


def supabase_cache_identifier(path):
    kind, scope = supabase_cache_kind_and_scope(path)
    return f"{kind}:{scope}" if scope else kind




def _supabase_date_for_db(value):
    dt = parse_date(value)
    return dt.strftime("%Y-%m-%d") if dt else None




def _supabase_int(value):
    try:
        if value is None:
            return 0
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return 0
        return int(float(s))
    except Exception:
        return 0


def _supabase_float(value):
    try:
        if value is None:
            return None
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return None
        v = float(s)
        return v if v > 0 else None
    except Exception:
        return None


def _supabase_text(value):
    if value is None:
        return ""
    return strip_gsheet_text_prefix(str(value)).strip()


def _supabase_table(name):
    # schema 與 table 名稱由程式固定產生，不吃使用者輸入值。
    return f"{SUPABASE_SCHEMA}.{name}"




def _supabase_sync_rows(sql, rows, label="", pre_statements=None):
    conn = get_supabase_conn()
    if conn is None:
        return False

    pre_statements = pre_statements or []
    rows = list(rows or [])

    try:
        batch_size = max(int(SUPABASE_BATCH_SIZE or 5000), 1)
        with conn.transaction():
            with conn.cursor() as cur:
                for pre_sql, pre_params in pre_statements:
                    cur.execute(pre_sql, pre_params or ())

                total = len(rows)
                for start in range(0, total, batch_size):
                    batch = rows[start:start + batch_size]
                    cur.executemany(sql, batch)
                    if total >= batch_size * 2:
                        print(f"  🗄️ Supabase 寫入中：{label} {min(start + len(batch), total):,}/{total:,}")
        return True
    except Exception as e:
        print(f"  ⚠️ Supabase 快取同步失敗：{label}，原因：{type(e).__name__}: {e}")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass








def write_cache_to_supabase(df, path, full_df=None):
    if not supabase_enabled() or not supabase_cache_supported(path):
        return False

    kind, scope = supabase_cache_kind_and_scope(path)
    cache_id = supabase_cache_identifier(path)
    df2 = (df.copy() if df is not None else pd.DataFrame()).fillna("")
    full_df2 = (full_df.copy() if full_df is not None else df2.copy()).fillna("")
    pre_statements = []

    if kind == "history":
        full_df2 = fix_known_underlying_info_dataframe(full_df2, "權證名稱", "標的股", "標的名稱")
        full_df2, _ = prune_history_cache_dataframe(full_df2)
        df2 = fix_known_underlying_info_dataframe(df2, "權證名稱", "標的股", "標的名稱")
        df2, _ = prune_history_cache_dataframe(df2)
    elif kind == "warrants":
        full_df2 = fix_known_underlying_info_dataframe(full_df2, "名稱", "標的股", "標的名稱")
        df2 = full_df2.copy()
    elif kind == "price":
        full_df2, _ = prune_price_cache_dataframe(full_df2)
        df2, _ = prune_price_cache_dataframe(df2)
    elif kind == "broker_map":
        full_df2, _ = prune_broker_map_dataframe(full_df2, scope=scope)
        df2 = full_df2.copy()

    if kind == "price":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            code = normalize_price_code(row.get("代號", ""))
            trade_date = _supabase_date_for_db(row.get("日期", ""))
            close_price = _supabase_float(row.get("收盤價", ""))
            if code and trade_date and close_price is not None:
                rows.append((code, trade_date, close_price))

        if not full_df2.empty and "日期" in full_df2.columns:
            cutoff = recent_trading_date_cutoff_from_series(full_df2["日期"], PRICE_RETENTION_TRADING_DAYS)
            if cutoff:
                pre_statements.append((
                    f"delete from {_supabase_table('price_cache')} where trade_date < %s",
                    (cutoff,),
                ))

        sql = f"""
            insert into {_supabase_table('price_cache')} as pc (code, trade_date, close_price, updated_at)
            values (%s, %s, %s, now())
            on conflict (code, trade_date) do update set
                close_price = excluded.close_price,
                updated_at = now()
            where pc.close_price is distinct from excluded.close_price
        """
        label = "快取_價格"

    elif kind == "history":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("權證代號", ""))
            broker_code = _supabase_text(row.get("券商代號", ""))
            trade_date = _supabase_date_for_db(row.get("日期", ""))
            if not warrant_code or not broker_code or not trade_date:
                continue
            rows.append((
                warrant_code,
                _supabase_text(row.get("權證名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
                _supabase_text(row.get("分點", "")),
                _supabase_text(row.get("分點名稱", "")),
                broker_code,
                trade_date,
                _supabase_int(row.get("買進股數", 0)),
                _supabase_int(row.get("賣出股數", 0)),
                _supabase_int(row.get("買進金額", 0)),
                _supabase_int(row.get("賣出金額", 0)),
                _supabase_int(row.get("買超股數", 0)),
                _supabase_int(row.get("買超金額", 0)),
            ))

        if not full_df2.empty and "日期" in full_df2.columns:
            cutoff = recent_trading_date_cutoff_from_series(full_df2["日期"], HISTORY_RETENTION_TRADING_DAYS)
            if cutoff:
                pre_statements.append((
                    f"delete from {_supabase_table('broker_warrant_history')} where trade_date < %s",
                    (cutoff,),
                ))

        allowed_labels = sorted(configured_broker_labels_for_scope("all"))
        allowed_codes = sorted(configured_broker_codes_for_scope("all"))
        pre_statements.append((
            f"""
            delete from {_supabase_table('broker_warrant_history')}
            where not (
                broker_label = any(%s::text[])
                or broker_code = any(%s::text[])
            )
            """,
            (allowed_labels, allowed_codes),
        ))

        sql = f"""
            insert into {_supabase_table('broker_warrant_history')} as h
            (warrant_code, warrant_name, underlying_code, underlying_name,
             broker_label, broker_name, broker_code, trade_date,
             buy_shares, sell_shares, buy_amount, sell_amount,
             net_buy_shares, net_buy_amount, updated_at)
            values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now())
            on conflict (warrant_code, broker_code, trade_date) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                broker_label = excluded.broker_label,
                broker_name = excluded.broker_name,
                buy_shares = excluded.buy_shares,
                sell_shares = excluded.sell_shares,
                buy_amount = excluded.buy_amount,
                sell_amount = excluded.sell_amount,
                net_buy_shares = excluded.net_buy_shares,
                net_buy_amount = excluded.net_buy_amount,
                updated_at = now()
            where (
                h.warrant_name,
                h.underlying_code,
                h.underlying_name,
                h.broker_label,
                h.broker_name,
                h.buy_shares,
                h.sell_shares,
                h.buy_amount,
                h.sell_amount,
                h.net_buy_shares,
                h.net_buy_amount
            ) is distinct from (
                excluded.warrant_name,
                excluded.underlying_code,
                excluded.underlying_name,
                excluded.broker_label,
                excluded.broker_name,
                excluded.buy_shares,
                excluded.sell_shares,
                excluded.buy_amount,
                excluded.sell_amount,
                excluded.net_buy_shares,
                excluded.net_buy_amount
            )
        """
        label = "快取_分點歷史"

    elif kind == "candidates":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("權證代號", ""))
            broker_code = _supabase_text(row.get("券商代號", ""))
            if not warrant_code or not broker_code:
                continue
            rows.append((
                scope,
                warrant_code,
                _supabase_text(row.get("權證名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
                _supabase_text(row.get("分點", "")),
                _supabase_text(row.get("分點名稱", "")),
                broker_code,
            ))
        sql = f"""
            insert into {_supabase_table('candidates')} as c
            (scope, warrant_code, warrant_name, underlying_code, underlying_name,
             broker_label, broker_name, broker_code, updated_at)
            values (%s, %s, %s, %s, %s, %s, %s, %s, now())
            on conflict (scope, warrant_code, broker_code) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                broker_label = excluded.broker_label,
                broker_name = excluded.broker_name,
                updated_at = now()
            where (
                c.warrant_name,
                c.underlying_code,
                c.underlying_name,
                c.broker_label,
                c.broker_name
            ) is distinct from (
                excluded.warrant_name,
                excluded.underlying_code,
                excluded.underlying_name,
                excluded.broker_label,
                excluded.broker_name
            )
        """
        label = f"快取_候選組合({scope})"

    elif kind == "warrants":
        rows = []
        seen_codes = set()
        for row in df2.itertuples(index=False):
            row = row._asdict()
            warrant_code = _supabase_text(row.get("代號", ""))
            if not warrant_code or warrant_code in seen_codes:
                continue
            seen_codes.add(warrant_code)
            rows.append((
                warrant_code,
                _supabase_text(row.get("名稱", "")),
                normalize_underlying_code_for_group(row.get("標的股", "")) or _supabase_text(row.get("標的股", "")),
                _supabase_text(row.get("標的名稱", "")),
            ))
        pre_statements.append((f"delete from {_supabase_table('warrants')}", ()))
        sql = f"""
            insert into {_supabase_table('warrants')}
            (warrant_code, warrant_name, underlying_code, underlying_name, updated_at)
            values (%s, %s, %s, %s, now())
            on conflict (warrant_code) do update set
                warrant_name = excluded.warrant_name,
                underlying_code = excluded.underlying_code,
                underlying_name = excluded.underlying_name,
                updated_at = now()
        """
        label = "快取_權證清單"

    elif kind == "broker_map":
        rows = []
        for row in df2.itertuples(index=False):
            row = row._asdict()
            broker_label = _supabase_text(row.get("分點", ""))
            if not broker_label:
                continue
            rows.append((
                scope,
                broker_label,
                _supabase_text(row.get("分點名稱", "")),
                _supabase_text(row.get("券商代號", "")),
            ))
        pre_statements.append((
            f"delete from {_supabase_table('broker_map')} where scope = %s",
            (scope,),
        ))
        sql = f"""
            insert into {_supabase_table('broker_map')}
            (scope, broker_label, broker_name, broker_code, updated_at)
            values (%s, %s, %s, %s, now())
            on conflict (scope, broker_label) do update set
                broker_name = excluded.broker_name,
                broker_code = excluded.broker_code,
                updated_at = now()
        """
        label = f"快取_分點代號({scope})"
    else:
        return False

    ok = _supabase_sync_rows(sql, rows, label=label, pre_statements=pre_statements)
    if ok:
        _SUPABASE_EMPTY_CACHE_KEYS.discard(cache_id)
        print(f"  🗄️ 已同步快取到 Supabase：{label}，本次寫入 {len(rows):,} 筆")
    return ok

def get_gcp_service_key():
    return os.getenv("GCP_SERVICE_KEY", os.getenv("GCP_SERVICE_KEY_NEW", "")).strip()


def gsheet_enabled():
    return bool(get_gcp_service_key())


def is_gsheet_quota_error(exc):
    msg = str(exc)
    return (
        "429" in msg
        or "Quota exceeded" in msg
        or "RESOURCE_EXHAUSTED" in msg
        or "Write requests per minute" in msg
    )


def gsheet_write_sleep():
    """
    Google Sheets token bucket 節流。

    相較每次固定等待 1.25 秒，token bucket 可讓少量批次請求快速通過，
    長時間大量寫入仍維持每分鐘安全速率；GSHEET_WRITE_SLEEP_SECONDS 只保留為最小間隔。
    """
    global _GSHEET_LAST_WRITE_TS
    global _GSHEET_TOKEN_BUCKET_TOKENS
    global _GSHEET_TOKEN_BUCKET_UPDATED_TS

    refill_per_second = GSHEET_WRITE_RATE_PER_MINUTE / 60.0
    while True:
        wait_seconds = 0.0
        with _GSHEET_TOKEN_BUCKET_LOCK:
            now_mono = time.monotonic()
            elapsed = max(now_mono - _GSHEET_TOKEN_BUCKET_UPDATED_TS, 0.0)
            _GSHEET_TOKEN_BUCKET_TOKENS = min(
                GSHEET_WRITE_BURST,
                _GSHEET_TOKEN_BUCKET_TOKENS + elapsed * refill_per_second,
            )
            _GSHEET_TOKEN_BUCKET_UPDATED_TS = now_mono

            if _GSHEET_TOKEN_BUCKET_TOKENS >= 1.0:
                _GSHEET_TOKEN_BUCKET_TOKENS -= 1.0
                break

            wait_seconds = max((1.0 - _GSHEET_TOKEN_BUCKET_TOKENS) / refill_per_second, 0.05)

        time.sleep(wait_seconds)

    if GSHEET_WRITE_SLEEP_SECONDS > 0:
        now = time.time()
        elapsed_wall = now - _GSHEET_LAST_WRITE_TS
        if elapsed_wall < GSHEET_WRITE_SLEEP_SECONDS:
            time.sleep(GSHEET_WRITE_SLEEP_SECONDS - elapsed_wall)
    _GSHEET_LAST_WRITE_TS = time.time()


def gsheet_api_call(description, func, *args, **kwargs):
    """
    所有 Google Sheet 寫入動作統一走這裡：
    1. 先節流
    2. 遇到 429 自動等待重試
    3. 不讓暫時 quota 造成後續工作表消失
    """
    last_error = None

    for attempt in range(1, GSHEET_MAX_RETRIES + 1):
        try:
            gsheet_write_sleep()
            return func(*args, **kwargs)
        except Exception as e:
            last_error = e

            if not is_gsheet_quota_error(e):
                raise

            wait_seconds = min(90, GSHEET_RETRY_BASE_SECONDS * attempt)
            print(
                f"  ⚠️ Google Sheet 觸發寫入配額限制，{description} 第 {attempt}/{GSHEET_MAX_RETRIES} 次重試，"
                f"等待 {wait_seconds:.0f} 秒..."
            )
            time.sleep(wait_seconds)

    raise last_error


def get_gsheet_client():
    global _GSHEET_CLIENT

    if _GSHEET_CLIENT is not None:
        return _GSHEET_CLIENT

    service_key = get_gcp_service_key()

    if not service_key:
        return None

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        info = json.loads(service_key)
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_info(info, scopes=scopes)
        _GSHEET_CLIENT = gspread.authorize(creds)
        return _GSHEET_CLIENT
    except Exception as e:
        print(f"  ⚠️ Google Sheet 金鑰初始化失敗：{type(e).__name__}: {e}")
        return None


def get_gsheet_spreadsheet():
    global _GSHEET_SPREADSHEET

    if _GSHEET_SPREADSHEET is not None:
        return _GSHEET_SPREADSHEET

    gc = get_gsheet_client()

    if gc is None:
        return None

    try:
        if GOOGLE_SHEET_ID:
            _GSHEET_SPREADSHEET = gc.open_by_key(GOOGLE_SHEET_ID)
        else:
            _GSHEET_SPREADSHEET = gc.open(GOOGLE_SHEET_NAME)
        return _GSHEET_SPREADSHEET
    except Exception as e:
        try:
            _GSHEET_SPREADSHEET = gc.create(GOOGLE_SHEET_NAME)
            print(f"  ✅ 已建立 Google Sheet：{GOOGLE_SHEET_NAME}")
            return _GSHEET_SPREADSHEET
        except Exception as e2:
            print(f"  ⚠️ Google Sheet 開啟/建立失敗：{type(e).__name__}: {e} / {type(e2).__name__}: {e2}")
            return None



def spreadsheet_grid_cell_count(spreadsheet):
    """計算整份 Google 試算表目前配置的格數（包含空白格）。"""
    if spreadsheet is None:
        return 0

    total = 0
    try:
        for ws in spreadsheet.worksheets():
            total += max(int(getattr(ws, "row_count", 1) or 1), 1) * max(
                int(getattr(ws, "col_count", 1) or 1),
                1,
            )
    except Exception:
        return total

    return total


def _worksheet_values_with_formulas(ws):
    """
    讀取工作表實際內容，優先保留公式字串。

    使用 FORMULA 可避免公式結果剛好是空字串時，被誤判成多餘空白列而刪除。
    """
    if ws is None:
        return []

    try:
        return ws.get_all_values(value_render_option="FORMULA") or []
    except TypeError:
        try:
            return ws.get_all_values() or []
        except Exception:
            return []
    except Exception:
        try:
            return ws.get_all_values() or []
        except Exception:
            return []


def worksheet_used_grid_size(ws):
    """回傳工作表最後一個有值或有公式的實際列數 / 欄數。"""
    values = _worksheet_values_with_formulas(ws)
    last_row = 0
    last_col = 0

    for row_idx, row in enumerate(values, start=1):
        row_last_col = 0
        for col_idx, value in enumerate(list(row), start=1):
            if str(value).strip() != "":
                row_last_col = col_idx

        if row_last_col > 0:
            last_row = row_idx
            last_col = max(last_col, row_last_col)

    return (
        max(last_row, GSHEET_COMPACT_MIN_ROWS),
        max(last_col, GSHEET_COMPACT_MIN_COLS),
    )


def compact_spreadsheet_blank_grid(spreadsheet=None, label="主試算表"):
    """
    把每張工作表縮到實際有內容 / 公式的列數與欄數。

    Google Sheets 的 1,000 萬格上限會把空白預留列與欄也算進去，因此不能只 clear，
    必須真的 resize 才能釋放格數。
    """
    if not GSHEET_COMPACT_BLANK_GRID_ENABLED:
        return False

    spreadsheet = spreadsheet or get_gsheet_spreadsheet()
    if spreadsheet is None:
        return False

    before_cells = spreadsheet_grid_cell_count(spreadsheet)
    changed_count = 0

    try:
        worksheets = spreadsheet.worksheets()
    except Exception as e:
        print(f"  ⚠️ Google Sheet 空白格清理失敗：{label}，原因：{type(e).__name__}: {e}")
        return False

    for ws in worksheets:
        try:
            used_rows, used_cols = worksheet_used_grid_size(ws)
            current_rows = max(int(getattr(ws, "row_count", 1) or 1), 1)
            current_cols = max(int(getattr(ws, "col_count", 1) or 1), 1)

            if current_rows == used_rows and current_cols == used_cols:
                continue

            gsheet_api_call(
                f"縮減空白列欄 {label}/{ws.title}",
                ws.resize,
                rows=used_rows,
                cols=used_cols,
            )
            changed_count += 1
            print(
                f"  🧹 已縮減工作表空白格：{ws.title}｜"
                f"{current_rows:,}×{current_cols:,} → {used_rows:,}×{used_cols:,}"
            )
        except Exception as e:
            print(f"  ⚠️ 工作表空白格縮減失敗：{getattr(ws, 'title', '-')}，原因：{type(e).__name__}: {e}")

    after_cells = spreadsheet_grid_cell_count(spreadsheet)
    if changed_count > 0:
        print(
            f"  🧹 Google Sheet 空白列欄清理完成：{label}｜"
            f"{before_cells:,} 格 → {after_cells:,} 格｜釋放 {max(before_cells - after_cells, 0):,} 格"
        )

    return changed_count > 0

def safe_worksheet_title(title):
    title = str(title).strip()
    bad_chars = [":", "\\", "/", "?", "*", "[", "]"]
    for ch in bad_chars:
        title = title.replace(ch, "_")
    return title[:100] if title else "工作表"


def cache_sheet_name_from_path(path):
    base = os.path.basename(str(path))
    return CACHE_SHEET_NAME_MAP.get(base, safe_worksheet_title(f"快取_{os.path.splitext(base)[0]}"))


def get_or_create_worksheet(title, rows=100, cols=20):
    sh = get_gsheet_spreadsheet()

    if sh is None:
        return None

    title = safe_worksheet_title(title)

    try:
        return sh.worksheet(title)
    except Exception:
        try:
            return gsheet_api_call(
                f"建立工作表 {title}",
                sh.add_worksheet,
                title=title,
                rows=max(int(rows), 1),
                cols=max(int(cols), 1),
            )
        except Exception as e:
            print(f"  ⚠️ 建立工作表失敗：{title}，原因：{type(e).__name__}: {e}")
            return None


def get_or_recreate_result_worksheet(title, rows=100, cols=20):
    """
    取得結果工作表。

    重要修正：
    以前為了清掉舊格式，會先刪除再重建結果工作表。
    但 Google Sheets API 有每分鐘寫入限制，一旦刪除後重建遇到 429，
    會導致「券商查詢」、「ABCDE組合勝率」等工作表直接消失。

    這版改成不刪除工作表：
    1. 已存在就沿用
    2. 不存在才建立
    3. 寫入前會清除舊格式與資料，再重新寫入與套樣式
    """
    sh = get_gsheet_spreadsheet()

    if sh is None:
        return None

    title = safe_worksheet_title(title)
    rows = max(int(rows), 1)
    cols = max(int(cols), 1)

    try:
        return sh.worksheet(title)
    except Exception:
        try:
            return gsheet_api_call(
                f"建立結果工作表 {title}",
                sh.add_worksheet,
                title=title,
                rows=rows,
                cols=cols,
            )
        except Exception as e:
            print(f"  ⚠️ 建立結果工作表失敗：{title}，原因：{type(e).__name__}: {e}")
            return None


GSHEET_TEXT_HEADER_KEYWORDS = (
    "權證代號",
    "權證代碼",
    "權證清單",
    "券商代號",
    "券商代碼",
    "股票代號",
    "股票代碼",
    "標的股",
    "標的代號",
    "標的代碼",
    "證券代號",
    "證券代碼",
    "商品代號",
    "商品代碼",
    "代號",
)


def clean_gsheet_value(value):
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        return value

    return str(value)


def is_gsheet_text_header(header):
    header = str(header).strip()
    return any(keyword in header for keyword in GSHEET_TEXT_HEADER_KEYWORDS)


def _strip_trailing_excel_decimal_for_code(s):
    s = str(s).strip()

    if s.endswith(".0"):
        head = s[:-2]
        if head.isdigit():
            return head

    return s


def normalize_gsheet_code_text_value(header, value):
    """
    Google Sheet 寫入 / 讀回時的代號欄位修正。

    目的：
    1. 權證代號若被 Google Sheet 轉成 5 碼，例如 30004，補回 030004。
    2. 權證清單內的每一個 5 碼權證代號也補回 6 碼。
    3. 股票代號 / 標的股 / 券商代號只維持文字，不任意補零，避免破壞原代號。
    4. 公式欄位不可處理，避免 =IFERROR(...) 被改成純文字。
    """
    if value is None:
        return ""

    s = str(value).strip()

    if s == "":
        return ""

    if s.startswith("="):
        return s

    had_prefix = s.startswith("'")
    if had_prefix:
        s = s[1:]

    header = str(header).strip()
    s = _strip_trailing_excel_decimal_for_code(s)

    if "權證清單" in header:
        # 權證清單常見格式：30004 海華國票...；60234 文華統一...
        # 若 Google Sheet 已經把 030004 轉成 30004，這裡將每段開頭的 5 碼權證補回 6 碼。
        def repl(match):
            token = match.group(0)
            return token.zfill(6) if token.isdigit() and len(token) == 5 else token

        s = re.sub(r"(?<!\d)\d{5}(?!\d)", repl, s)
        return ("'" + s) if had_prefix else s

    if "權證代號" in header or "權證代碼" in header:
        if s.isdigit() and len(s) == 5:
            s = s.zfill(6)
        return ("'" + s) if had_prefix else s

    # 快取_權證清單 的欄位名稱是「代號」，價格快取也可能是「代號」。
    # 只有 5 碼純數字才視為權證前導 0 被吃掉，補回 6 碼；4 碼股票不補。
    if header in ("代號", "證券代號", "證券代碼", "商品代號", "商品代碼"):
        if s.isdigit() and len(s) == 5:
            s = s.zfill(6)
        return ("'" + s) if had_prefix else s

    return ("'" + s) if had_prefix else s


def add_gsheet_text_prefix(value):
    """
    讓 Google Sheet 用文字格式寫入代號欄位。

    若使用 USER_ENTERED 寫入 064390，Google Sheet 可能會自動轉成數字 64390；
    對權證代號 / 券商代號欄位加上前導單引號，可保留開頭 0。
    Google Sheet 顯示時不會顯示這個單引號，只會把儲存格視為文字。

    注意：如果儲存格內容是公式，不能加單引號，否則 Google Sheet 會把公式當成純文字顯示。
    例如「券商查詢」工作表的前 15 名查詢欄位就是公式欄位，必須保持 =IFERROR(...) 可計算。
    """
    if value is None:
        return ""

    s = str(value).strip()

    if s == "":
        return ""

    if s.startswith("="):
        return s

    if s.startswith("'"):
        return s

    return "'" + s


def strip_gsheet_text_prefix(value):
    if value is None:
        return ""

    s = str(value)

    if s.startswith("'"):
        return s[1:]

    return s


def normalize_gsheet_values_for_text_columns(values):
    """
    依據表頭自動判斷權證代號 / 券商代號 / 代號欄位，
    寫入 Google Sheet 前強制改成文字，避免前導 0 被吃掉。

    支援：
    1. 快取工作表：表頭通常在第 1 列。
    2. 一般結果工作表：表頭多數也在第 1 列。
    3. 例如 ABCDE組合勝率 這種表頭在第 4 列的工作表，會掃前 10 列找表頭。
    """
    if not values:
        return values

    out = [list(row) for row in values]
    header_rows = []

    scan_limit = min(len(out), 10)
    for row_idx in range(scan_limit):
        row = out[row_idx]
        text_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_text_header(header):
                text_cols.append(col_idx)

        if text_cols:
            header_rows.append((row_idx, text_cols))

    if not header_rows:
        return out

    for header_row_idx, text_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(out)

        for row_idx in range(header_row_idx + 1, end_row):
            for col_idx in text_cols:
                if col_idx >= len(out[row_idx]):
                    continue

                cell_value = out[row_idx][col_idx]
                if isinstance(cell_value, str) and cell_value.strip().startswith("="):
                    continue

                header = out[header_row_idx][col_idx] if col_idx < len(out[header_row_idx]) else ""
                cell_value = normalize_gsheet_code_text_value(header, cell_value)
                out[row_idx][col_idx] = add_gsheet_text_prefix(cell_value)

    return out


def collect_gsheet_text_column_ranges(values):
    """
    根據工作表前 10 列的表頭，找出需要強制使用「純文字」格式的欄位範圍。

    這裡除了「權證代號 / 券商代號 / 代號」之外，也包含「權證清單」、
    「股票代號」、「標的股」等欄位，避免 Google Sheet 把 0 開頭的代號自動轉成數字。
    """
    ranges = []

    if not values:
        return ranges

    scan_limit = min(len(values), 10)
    header_rows = []

    for row_idx in range(scan_limit):
        row = list(values[row_idx])
        text_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_text_header(header):
                text_cols.append(col_idx)

        if text_cols:
            header_rows.append((row_idx, text_cols))

    for header_row_idx, text_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(values)

        if end_row <= header_row_idx + 1:
            continue

        for col_idx in text_cols:
            ranges.append({
                "start_row": header_row_idx + 1,
                "end_row": end_row,
                "start_col": col_idx,
                "end_col": col_idx + 1,
            })

    return ranges


def apply_text_format_to_gsheet(gws, values):
    """
    將 Google Sheet 中的代號相關欄位套用純文字格式。

    write_values_to_worksheet() 已經會在代號欄位前加單引號，這裡再補上
    Google Sheets 的 TEXT numberFormat，雙重避免權證代號、股票代號、券商代號
    或權證清單中的 0 開頭代號被吃掉。

    注意：「券商查詢」工作表的資料列是公式查詢結果，不可把公式欄位套成純文字，
    否則 Google Sheet 會顯示 =IFERROR(...) 文字而不是計算結果。
    權證代號與權證清單的文字格式會保留在「券商查詢資料」隱藏工作表中。
    """
    if gws is None or not values:
        return

    if str(getattr(gws, "title", "")).strip() == "券商查詢":
        return

    requests = []

    for r in collect_gsheet_text_column_ranges(values):
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": gws.id,
                    "startRowIndex": r["start_row"],
                    "endRowIndex": r["end_row"],
                    "startColumnIndex": r["start_col"],
                    "endColumnIndex": r["end_col"],
                },
                "cell": {
                    "userEnteredFormat": {
                        "numberFormat": {
                            "type": "TEXT"
                        }
                    }
                },
                "fields": "userEnteredFormat.numberFormat",
            }
        })

    _gsheet_batch_update(requests)



GSHEET_COMMA_NUMBER_HEADER_KEYWORDS = (
    "排名",
    "金額",
    "股數",
    "張數",
    "筆數",
    "事件數",
    "權證檔數",
    "買進分點數",
    "涵蓋權證數",
    "持有天數",
    "價格筆數",
    "成本",
    "市值",
    "損益",
)

GSHEET_COMMA_NUMBER_EXCLUDE_KEYWORDS = (
    "代號",
    "日期",
    "名稱",
    "清單",
    "類型",
    "狀態",
    "均價",
    "勝率",
    "占比",
    "比例",
    "報酬%",
    "獲利%",
)

GSHEET_DECIMAL_NUMBER_HEADER_KEYWORDS = (
    "均價",
    "收盤價",
    "最新權證價格",
    "權證價格",
    "股價",
    "報酬率",
    "報酬%",
    "獲利%",
    "勝率",
    "占比",
    "比例",
    "平均",
)

GSHEET_DECIMAL_NUMBER_EXCLUDE_KEYWORDS = (
    "日期",
    "代號",
    "名稱",
    "清單",
    "類型",
    "狀態",
    "文字",
    "說明",
    "來源",
)


def is_gsheet_percent_header(header):
    header = str(header).strip()

    if not header:
        return False

    if "文字" in header or "日期" in header or "代號" in header:
        return False

    return ("%" in header) or ("勝率" in header) or ("占比" in header) or ("比例" in header)


def is_gsheet_decimal_number_header(header):
    """
    判斷 Google Sheet 中需要數字格式但不一定要整數千分位的欄位。

    主要修正：
    1. 買進均價 / 減碼均價 / 出清均價不可被誤套日期格式。
    2. 減碼獲利% / 出清獲利% 不可被誤套日期格式。
    3. 報酬率、勝率、占比等欄位要保持數字或百分比顯示。
    """
    header = str(header).strip()

    if not header:
        return False

    if header.startswith("D+"):
        return False

    if is_gsheet_text_header(header):
        return False

    for keyword in GSHEET_DECIMAL_NUMBER_EXCLUDE_KEYWORDS:
        if keyword in header:
            return False

    return any(keyword in header for keyword in GSHEET_DECIMAL_NUMBER_HEADER_KEYWORDS)


def is_gsheet_numeric_format_header(header):
    return is_gsheet_comma_number_header(header) or is_gsheet_decimal_number_header(header)


def is_gsheet_comma_number_header(header):
    """
    判斷 Google Sheet 結果工作表中，哪些欄位要用千分位逗號顯示。

    注意：
    1. 權證代號 / 券商代號 / 代號欄位不能套數字格式，否則開頭 0 會被吃掉。
    2. 日期、名稱、清單、百分比、均價等欄位不套用千分位。
    3. 這個函式只用在「結果工作表」同步，不改快取工作表邏輯。
    """
    header = str(header).strip()

    if not header:
        return False

    if header.startswith("D+"):
        return False

    if is_gsheet_text_header(header):
        return False

    for keyword in GSHEET_COMMA_NUMBER_EXCLUDE_KEYWORDS:
        if keyword in header:
            return False

    return any(keyword in header for keyword in GSHEET_COMMA_NUMBER_HEADER_KEYWORDS)


def _parse_comma_number_for_gsheet(value):
    """
    將 1,234 / 12,345.67 這類字串轉成數字，讓 Google Sheet 可以搭配 numberFormat 顯示逗號。
    公式、空值、百分比、文字說明都保持原樣。
    """
    if value is None:
        return ""

    if isinstance(value, bool):
        return value

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return value

    s = str(value).strip()

    if s == "" or s == "-":
        return value

    if s.startswith("="):
        return value

    if "%" in s or "\n" in s or "；" in s:
        return value

    raw = s.replace(",", "").strip()

    if raw.startswith("+"):
        raw = raw[1:]

    if raw.startswith("-"):
        sign = "-"
        num_part = raw[1:]
    else:
        sign = ""
        num_part = raw

    if not num_part:
        return value

    if num_part.replace(".", "", 1).isdigit():
        try:
            if "." in num_part:
                return float(sign + num_part)
            return int(sign + num_part)
        except Exception:
            return value

    return value


def normalize_result_values_for_comma_numbers(values):
    """
    Google Sheet 結果同步前，將需要千分位的欄位轉成數字。

    原因：若直接把 "1,234" 用 USER_ENTERED 寫入 Google Sheet，可能被解析成數字但顯示為 1234，
    或在部分語系下變成文字。這裡先依表頭把金額 / 股數 / 張數 / 筆數等欄位轉成數字，
    後續再用 Google Sheets numberFormat 套 #,##0，確保畫面會顯示逗號。
    """
    if not values:
        return values

    out = [list(row) for row in values]
    header_rows = []

    scan_limit = min(len(out), 10)
    for row_idx in range(scan_limit):
        row = out[row_idx]
        number_cols = []

        for col_idx, header in enumerate(row):
            if is_gsheet_numeric_format_header(header):
                number_cols.append(col_idx)

        if number_cols:
            header_rows.append((row_idx, number_cols))

    if not header_rows:
        return out

    for header_row_idx, number_cols in header_rows:
        next_header_rows = [idx for idx, _ in header_rows if idx > header_row_idx]
        end_row = min(next_header_rows) if next_header_rows else len(out)

        for row_idx in range(header_row_idx + 1, end_row):
            for col_idx in number_cols:
                if col_idx >= len(out[row_idx]):
                    continue

                out[row_idx][col_idx] = _parse_comma_number_for_gsheet(out[row_idx][col_idx])

    return out


def _gsheet_number_format_for_header(header):
    header = str(header).strip()

    # 重要：Excel / Google Sheet 用 USER_ENTERED 寫入「+7.00%」這種字串時，
    # 儲存格底層值會變成 0.07。這類欄位必須套用真正的 PERCENT 格式，
    # 才會顯示成 +7.00%，不能用 NUMBER + 文字百分號，否則會被顯示成 +0.07%。
    #
    # 但「報酬率」這種沒有 % 符號的 TOP15 快取欄位，程式內多為 70.30 這種百分點數值，
    # 所以不歸類為 PERCENT，仍以一般數字顯示，避免變成 7,030%。
    if is_gsheet_percent_header(header):
        # 勝率 / 占比 / 比例不是報酬率，不應該顯示 + 號。
        if "勝率" in header or "占比" in header or "比例" in header:
            return {
                "type": "PERCENT",
                "pattern": '0.00%',
            }

        return {
            "type": "PERCENT",
            "pattern": '+0.00%;-0.00%;0.00%',
        }

    if is_gsheet_decimal_number_header(header):
        return {
            "type": "NUMBER",
            "pattern": "#,##0.00",
        }

    return {
        "type": "NUMBER",
        "pattern": "#,##0",
    }


def _gsheet_number_pattern_for_header(header):
    # 保留舊函式名稱供既有呼叫相容；實際格式型別請用 _gsheet_number_format_for_header()。
    return _gsheet_number_format_for_header(header).get("pattern", "#,##0")


def _format_source_rows_from_values_or_xlsx(ws_xlsx, values=None):
    """
    Google Sheet 格式套用時的表頭來源。

    upsert 模式會在 Google Sheet 寫入前新增「資料範圍」欄位，
    因此不能再只看原本 Excel 的欄位位置，否則日期格式會往左/往右套錯，
    造成買進均價、買超張數、獲利% 被顯示成 1900 年代日期。
    """
    if values:
        rows = [list(row) for row in values]
        max_row = max(len(rows), 1)
        max_col = max(max((len(row) for row in rows), default=1), 1)
        return rows, max_row, max_col

    rows = []
    if ws_xlsx is not None:
        scan_limit = min(ws_xlsx.max_row, 10)
        for row_idx in range(1, scan_limit + 1):
            rows.append([ws_xlsx.cell(row_idx, col_idx).value for col_idx in range(1, ws_xlsx.max_column + 1)])
        return rows, ws_xlsx.max_row, ws_xlsx.max_column

    return [], 1, 1


def apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=None):
    """
    將 Google Sheet 結果工作表的金額 / 股數 / 張數 / 成本 / 均價 / 獲利% 等欄位套用正確數字格式。

    重點修正：
    - upsert 後的實際 Google Sheet 表頭可能比原 Excel 多「資料範圍」欄。
    - 這裡改用實際寫入的 values 找欄位位置，避免格式位移。
    - 所有金額 / 成本 / 市值 / 損益使用千分位逗號顯示。
    - 均價、報酬率、獲利%、勝率等欄位也強制套數字/百分比格式，避免被日期格式污染。
    """
    if gws is None:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    source_rows, max_row, max_col = _format_source_rows_from_values_or_xlsx(ws_xlsx, values)
    header_rows = []
    scan_limit = min(len(source_rows), 10)

    for row_idx in range(1, scan_limit + 1):
        row = source_rows[row_idx - 1]
        number_cols = []

        for col_idx in range(1, max_col + 1):
            header = row[col_idx - 1] if col_idx - 1 < len(row) else ""

            if is_gsheet_numeric_format_header(header):
                number_cols.append((col_idx, _gsheet_number_pattern_for_header(header)))

        if number_cols:
            header_rows.append((row_idx, number_cols))

    if not header_rows:
        return

    requests = []

    for idx, (header_row_idx, number_cols) in enumerate(header_rows):
        if idx + 1 < len(header_rows):
            end_row = header_rows[idx + 1][0] - 1
        else:
            end_row = max_row

        start_data_row = header_row_idx + 1

        if start_data_row > end_row:
            continue

        for col_idx, pattern in number_cols:
            header = ""
            try:
                row = source_rows[header_row_idx - 1]
                header = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            except Exception:
                header = ""

            number_format = _gsheet_number_format_for_header(header)

            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_data_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": col_idx - 1,
                        "endColumnIndex": col_idx,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": number_format
                        }
                    },
                    "fields": "userEnteredFormat.numberFormat",
                }
            })

    _gsheet_batch_update(requests)


GSHEET_DATE_HEADER_KEYWORDS = (
    "日期",
    "買進日",
    "賣出日",
    "事件日",
    "起始日",
    "結束日",
    "減碼日",
    "出清日",
    "最近買進日",
    "第一筆日期",
    "最後筆日期",
)


def is_gsheet_date_header(header):
    """
    判斷 Google Sheet 中哪些欄位應該以日期格式顯示。

    主要修正「券商查詢」工作表的「最近買進日」：
    Google Sheet 公式 INDEX/MATCH 從資料表抓到日期時，常會以日期序號顯示，
    例如 46160。這裡統一把日期欄位套成 yyyy/mm/dd，讓畫面顯示正常日期。
    """
    header = str(header).strip()

    if not header:
        return False

    if "天數" in header:
        return False

    return any(keyword in header for keyword in GSHEET_DATE_HEADER_KEYWORDS)


def apply_date_format_to_gsheet(ws_xlsx, gws, values=None):
    """
    將日期相關欄位套用 Google Sheets 日期格式 yyyy/mm/dd。

    upsert 模式會新增「資料範圍」欄位，因此日期欄位置必須以實際寫入 values 的表頭判斷，
    不可只看原本 Excel 欄位位置。這可避免買進均價 / 張數 / 獲利% 被誤套成日期格式。
    """
    if gws is None:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    source_rows, max_row, max_col = _format_source_rows_from_values_or_xlsx(ws_xlsx, values)
    header_rows = []
    scan_limit = min(len(source_rows), 10)

    for row_idx in range(1, scan_limit + 1):
        row = source_rows[row_idx - 1]
        date_cols = []

        for col_idx in range(1, max_col + 1):
            header = row[col_idx - 1] if col_idx - 1 < len(row) else ""

            if is_gsheet_date_header(header):
                date_cols.append(col_idx)

        if date_cols:
            header_rows.append((row_idx, date_cols))

    if not header_rows:
        return

    requests = []

    for idx, (header_row_idx, date_cols) in enumerate(header_rows):
        if idx + 1 < len(header_rows):
            end_row = header_rows[idx + 1][0] - 1
        else:
            end_row = max_row

        start_data_row = header_row_idx + 1

        if start_data_row > end_row:
            continue

        for col_idx in date_cols:
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": start_data_row - 1,
                        "endRowIndex": end_row,
                        "startColumnIndex": col_idx - 1,
                        "endColumnIndex": col_idx,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "numberFormat": {
                                "type": "DATE",
                                "pattern": "yyyy/mm/dd",
                            }
                        }
                    },
                    "fields": "userEnteredFormat.numberFormat",
                }
            })

    _gsheet_batch_update(requests)


def _gsheet_pixel_width_for_header(header):
    header = str(header).strip()

    if not header:
        return None

    if header == "資料範圍":
        return 105

    # 有權證資訊的欄位加寬，避免新增「資料範圍」後欄寬位移造成內容被擠在一起。
    if "權證清單" in header or "權證集合" in header:
        return 520
    if "分點明細_JSON" in header:
        return 700
    if "權證名稱" in header:
        return 190
    if "權證代號" in header or "權證代碼" in header:
        return 105
    if "權證檔數" in header:
        return 95

    if "標的名稱" in header or header == "股票名稱":
        return 140
    if "標的股" in header or "標的代號" in header:
        return 95
    if "分點名稱" in header or header == "分點":
        return 140
    if "券商代號" in header:
        return 95

    if "日期" in header or header in ("買進日", "事件日", "起始日", "結束日", "減碼日", "出清日"):
        return 105

    if any(k in header for k in ("金額", "成本", "市值", "損益")):
        return 125

    if any(k in header for k in ("均價", "報酬率", "獲利%", "勝率", "占比", "比例")):
        return 105

    if any(k in header for k in ("股數", "張數", "筆數", "天數", "排名")):
        return 85

    return None


def apply_header_widths_to_gsheet(gws, values=None):
    """
    依照實際 Google Sheet 表頭重新調整欄寬。

    因為 upsert 會新增「資料範圍」欄位，不能只沿用原 Excel 欄寬；
    否則有權證名稱 / 權證清單的欄位會被擠到太窄。
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    header_row_idx = _find_simple_header_row(values)
    if header_row_idx is None or header_row_idx >= len(values):
        return

    headers = list(values[header_row_idx])
    requests = []

    for col_idx, header in enumerate(headers):
        pixel_size = _gsheet_pixel_width_for_header(header)
        if not pixel_size:
            continue

        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": col_idx,
                    "endIndex": col_idx + 1,
                },
                "properties": {
                    "pixelSize": pixel_size,
                },
                "fields": "pixelSize",
            }
        })

    _gsheet_batch_update(requests)


def reset_worksheet_before_value_write(ws, row_count, col_count):
    """
    寫入值之前先清掉舊格式、舊資料驗證與舊合併範圍。

    這可以解決「券商查詢」公式被舊的純文字格式吃掉，
    同時避免因刪除 / 重建工作表造成 429 後工作表消失。
    """
    if ws is None:
        return

    try:
        sheet_id = int(ws.id)
    except Exception:
        return

    requests = [
        {
            "unmergeCells": {
                "range": {
                    "sheetId": sheet_id,
                }
            }
        },
        {
            "updateCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": max(int(row_count), 1),
                    "startColumnIndex": 0,
                    "endColumnIndex": max(int(col_count), 1),
                },
                "fields": "userEnteredFormat,dataValidation",
            }
        },
    ]

    _gsheet_batch_update(requests)


def gsheet_values_batch_update(ws, data, description, chunk_size=5000):
    """以 values.batchUpdate 在單一請求內寫入多個 range；超大 payload 才分大批次。"""
    if ws is None or not data:
        return
    chunk_size = max(int(chunk_size or 5000), 1)
    for start in range(0, len(data), chunk_size):
        chunk = data[start:start + chunk_size]
        gsheet_api_call(
            f"{description} {start + 1}-{start + len(chunk)}",
            ws.batch_update,
            chunk,
            value_input_option="USER_ENTERED",
        )


def write_values_to_worksheet(ws, values):
    if ws is None:
        return False

    if not values:
        values = [[""]]

    row_count = max(len(values), 1)
    col_count = max(max((len(row) for row in values), default=1), 1)

    normalized_values = []
    for row in values:
        row = list(row)
        if len(row) < col_count:
            row = row + [""] * (col_count - len(row))
        normalized_values.append([clean_gsheet_value(v) for v in row])

    normalized_values = normalize_gsheet_values_for_text_columns(normalized_values)

    try:
        # 不再使用 delete/recreate。先 resize，再清格式，最後寫入值。
        # 這樣遇到 429 時不會把既有工作表刪掉。
        gsheet_api_call(
            f"調整工作表大小 {ws.title}",
            ws.resize,
            rows=max(row_count, 1),
            cols=max(col_count, 1),
        )

        reset_worksheet_before_value_write(ws, row_count, col_count)

        batch_ranges = []
        for start in range(0, len(normalized_values), GSHEET_CHUNK_ROWS):
            chunk = normalized_values[start:start + GSHEET_CHUNK_ROWS]
            batch_ranges.append({
                "range": f"A{start + 1}",
                "values": chunk,
            })
        gsheet_values_batch_update(
            ws,
            batch_ranges,
            f"批次寫入工作表資料 {ws.title}",
        )

        apply_text_format_to_gsheet(ws, normalized_values)

        return True
    except Exception as e:
        print(f"  ⚠️ Google Sheet 寫入失敗：{ws.title}，原因：{type(e).__name__}: {e}")
        return False


def read_cache_from_gsheet(path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return pd.DataFrame()

    title = cache_sheet_name_from_path(path)

    try:
        sh = get_gsheet_spreadsheet()
        if sh is None:
            return pd.DataFrame()

        ws = sh.worksheet(title)
        values = ws.get_all_values()

        if not values or len(values) < 2:
            return pd.DataFrame()

        headers = [str(h).strip() for h in values[0]]
        rows = values[1:]

        if not headers or all(h == "" for h in headers):
            return pd.DataFrame()

        fixed_rows = []
        n_cols = len(headers)
        for row in rows:
            row = list(row)
            if len(row) < n_cols:
                row = row + [""] * (n_cols - len(row))
            elif len(row) > n_cols:
                row = row[:n_cols]
            fixed_rows.append(row)

        df = pd.DataFrame(fixed_rows, columns=headers).fillna("")

        for col in df.columns:
            if is_gsheet_text_header(col):
                df[col] = df[col].map(strip_gsheet_text_prefix)
                df[col] = df[col].map(lambda v, _col=col: normalize_gsheet_code_text_value(_col, v))

        print(f"  ☁️ 已從 Google Sheet 讀取快取：{title}，共 {len(df):,} 筆")
        return df
    except Exception as exc:
        print(f"  ⚠️ Google Sheet 快取讀取失敗：{title}，原因：{type(exc).__name__}: {exc}")
        return pd.DataFrame()


def write_cache_to_gsheet(df, path):
    if not GSHEET_CACHE_ENABLED or not gsheet_enabled():
        return

    if df is None:
        return

    try:
        title = cache_sheet_name_from_path(path)
        df2 = df.copy().fillna("")
        values = [list(df2.columns)] + df2.astype(str).values.tolist()
        ws = get_or_create_worksheet(title, rows=max(len(values), 100), cols=max(len(df2.columns), 20))

        if write_values_to_worksheet(ws, values):
            print(f"  ☁️ 已同步快取到 Google Sheet：{title}，共 {len(df2):,} 筆")
    except Exception as e:
        print(f"  ⚠️ 快取同步到 Google Sheet 失敗：{path}，原因：{type(e).__name__}: {e}")



def _hex_to_gsheet_color(hex_value):
    if not hex_value:
        return None

    s = str(hex_value).strip()

    if not s or s in ("00000000", "FFFFFFFF"):
        # 00000000 在 openpyxl 常代表無填色，不是真的黑色背景。
        return None

    if len(s) == 8:
        s = s[-6:]

    if len(s) != 6:
        return None

    try:
        return {
            "red": int(s[0:2], 16) / 255,
            "green": int(s[2:4], 16) / 255,
            "blue": int(s[4:6], 16) / 255,
        }
    except Exception:
        return None


def _openpyxl_color_to_gsheet(color_obj):
    if color_obj is None:
        return None

    try:
        if color_obj.type == "rgb" and color_obj.rgb:
            return _hex_to_gsheet_color(color_obj.rgb)
    except Exception:
        pass

    return None


def _openpyxl_fill_to_gsheet(cell):
    try:
        fill = cell.fill
        if fill is None or fill.fill_type is None:
            return None

        color = _openpyxl_color_to_gsheet(fill.fgColor)
        return color
    except Exception:
        return None


def _openpyxl_font_to_gsheet(cell):
    text_format = {}

    try:
        font = cell.font

        if font is None:
            return text_format

        if font.bold is not None:
            text_format["bold"] = bool(font.bold)

        if font.italic is not None:
            text_format["italic"] = bool(font.italic)

        if font.sz:
            text_format["fontSize"] = int(float(font.sz))

        color = _openpyxl_color_to_gsheet(font.color)
        if color:
            text_format["foregroundColor"] = color
    except Exception:
        pass

    return text_format


def _openpyxl_alignment_to_gsheet(cell):
    fmt = {}

    try:
        alignment = cell.alignment

        horizontal_map = {
            "center": "CENTER",
            "left": "LEFT",
            "right": "RIGHT",
        }
        vertical_map = {
            "center": "MIDDLE",
            "top": "TOP",
            "bottom": "BOTTOM",
        }

        if alignment.horizontal in horizontal_map:
            fmt["horizontalAlignment"] = horizontal_map[alignment.horizontal]

        if alignment.vertical in vertical_map:
            fmt["verticalAlignment"] = vertical_map[alignment.vertical]

        if alignment.wrap_text:
            fmt["wrapStrategy"] = "WRAP"
    except Exception:
        pass

    return fmt


def _openpyxl_number_format_to_gsheet(cell):
    try:
        nf = str(cell.number_format or "").strip()

        if not nf or nf == "General":
            return None

        if "%" in nf:
            return {"type": "PERCENT", "pattern": nf}

        if "#" in nf or "0" in nf:
            return {"type": "NUMBER", "pattern": nf}
    except Exception:
        pass

    return None


def _cell_gsheet_format(cell):
    fmt = {}

    bg = _openpyxl_fill_to_gsheet(cell)
    if bg:
        fmt["backgroundColor"] = bg

    text_format = _openpyxl_font_to_gsheet(cell)
    if text_format:
        fmt["textFormat"] = text_format

    align_format = _openpyxl_alignment_to_gsheet(cell)
    fmt.update(align_format)

    number_format = _openpyxl_number_format_to_gsheet(cell)
    if number_format:
        fmt["numberFormat"] = number_format

    return fmt


def _format_fields(fmt):
    fields = []

    if "backgroundColor" in fmt:
        fields.append("userEnteredFormat.backgroundColor")
    if "textFormat" in fmt:
        for key in fmt["textFormat"].keys():
            fields.append(f"userEnteredFormat.textFormat.{key}")
    if "horizontalAlignment" in fmt:
        fields.append("userEnteredFormat.horizontalAlignment")
    if "verticalAlignment" in fmt:
        fields.append("userEnteredFormat.verticalAlignment")
    if "wrapStrategy" in fmt:
        fields.append("userEnteredFormat.wrapStrategy")
    if "numberFormat" in fmt:
        fields.append("userEnteredFormat.numberFormat")

    return ",".join(fields)


def _gsheet_batch_update(requests, chunk_size=1000):
    if not requests:
        return

    sh = get_gsheet_spreadsheet()

    if sh is None:
        return

    for start in range(0, len(requests), chunk_size):
        chunk = requests[start:start + chunk_size]
        try:
            gsheet_api_call(
                f"套用 Google Sheet 格式 batchUpdate {start + 1}-{start + len(chunk)}",
                sh.batch_update,
                {"requests": chunk},
            )
        except Exception as e:
            print(f"  ⚠️ Google Sheet 格式套用失敗：{type(e).__name__}: {e}")
            return


def _openpyxl_freeze_to_grid_properties(ws_xlsx):
    frozen_rows = 0
    frozen_cols = 0

    try:
        pane = ws_xlsx.freeze_panes

        if pane:
            from openpyxl.utils.cell import coordinate_to_tuple
            row, col = coordinate_to_tuple(str(pane))
            frozen_rows = max(row - 1, 0)
            frozen_cols = max(col - 1, 0)
    except Exception:
        pass

    return frozen_rows, frozen_cols


def _excel_width_to_pixels(width):
    try:
        return max(20, int(float(width) * 7 + 5))
    except Exception:
        return None


def _excel_height_to_pixels(height):
    try:
        return max(18, int(float(height) * 1.333))
    except Exception:
        return None


def normalize_formula_for_gsheet(value):
    if not isinstance(value, str):
        return value

    if not value.startswith("="):
        return value

    # Google Sheets 對中文工作表名稱建議加單引號，避免公式解析失敗。
    value = value.replace("INDEX(券商查詢資料!", "INDEX('券商查詢資料'!")
    value = value.replace(",券商查詢資料!", ",'券商查詢資料'!")
    value = value.replace("MATCH($B$2&\"|\"&$A", "MATCH($B$2&\"|\"&$A")
    return value


def apply_excel_style_to_gsheet(ws_xlsx, gws):
    """
    將 openpyxl 產生的 Excel 樣式轉成 Google Sheets 格式。

    會同步：
    1. 背景色、字體粗細 / 字色 / 字級
    2. 文字置中、換行、數字格式
    3. 欄寬、列高、凍結列 / 欄
    4. 合併儲存格
    5. 隱藏工作表
    6. 券商查詢 B2 下拉選單

    Google Sheets API 與 Excel 格式模型不同，因此外框只保留主要視覺效果，
    但 A/B/C/D/E 的紅綠藍橘狀態色、標頭色與查詢頁互動功能會完整保留。
    """
    if gws is None:
        return

    sheet_id = int(gws.id)
    requests = []

    frozen_rows, frozen_cols = _openpyxl_freeze_to_grid_properties(ws_xlsx)

    requests.append({
        "updateSheetProperties": {
            "properties": {
                "sheetId": sheet_id,
                "hidden": bool(ws_xlsx.sheet_state == "hidden"),
                "gridProperties": {
                    "frozenRowCount": frozen_rows,
                    "frozenColumnCount": frozen_cols,
                },
            },
            "fields": "hidden,gridProperties.frozenRowCount,gridProperties.frozenColumnCount",
        }
    })

    # 先清除舊合併範圍，避免重複執行時 mergeCells 失敗。
    requests.append({
        "unmergeCells": {
            "range": {
                "sheetId": sheet_id,
            }
        }
    })

    # 合併儲存格。
    for merged_range in ws_xlsx.merged_cells.ranges:
        try:
            requests.append({
                "mergeCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": merged_range.min_row - 1,
                        "endRowIndex": merged_range.max_row,
                        "startColumnIndex": merged_range.min_col - 1,
                        "endColumnIndex": merged_range.max_col,
                    },
                    "mergeType": "MERGE_ALL",
                }
            })
        except Exception:
            pass

    # 欄寬。
    for col_idx in range(1, ws_xlsx.max_column + 1):
        letter = get_column_letter(col_idx)
        width = ws_xlsx.column_dimensions[letter].width
        pixel_size = _excel_width_to_pixels(width) if width else None

        if pixel_size:
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": col_idx - 1,
                        "endIndex": col_idx,
                    },
                    "properties": {
                        "pixelSize": pixel_size,
                    },
                    "fields": "pixelSize",
                }
            })

    # 列高。
    for row_idx in range(1, ws_xlsx.max_row + 1):
        height = ws_xlsx.row_dimensions[row_idx].height
        pixel_size = _excel_height_to_pixels(height) if height else None

        if pixel_size:
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "ROWS",
                        "startIndex": row_idx - 1,
                        "endIndex": row_idx,
                    },
                    "properties": {
                        "pixelSize": pixel_size,
                    },
                    "fields": "pixelSize",
                }
            })

    # 逐列壓縮相同格式的連續儲存格，減少 batchUpdate 請求數量。
    import json as _json

    for row_idx in range(1, ws_xlsx.max_row + 1):
        run_start = None
        run_fmt = None
        run_key = None

        for col_idx in range(1, ws_xlsx.max_column + 2):
            if col_idx <= ws_xlsx.max_column:
                cell = ws_xlsx.cell(row_idx, col_idx)
                fmt = _cell_gsheet_format(cell)
                key = _json.dumps(fmt, sort_keys=True, ensure_ascii=False) if fmt else None
            else:
                fmt = None
                key = None

            if key and run_key is None:
                run_start = col_idx
                run_fmt = fmt
                run_key = key
            elif key and key == run_key:
                continue
            else:
                if run_key and run_start is not None and run_fmt:
                    fields = _format_fields(run_fmt)
                    if fields:
                        requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": row_idx - 1,
                                    "endRowIndex": row_idx,
                                    "startColumnIndex": run_start - 1,
                                    "endColumnIndex": col_idx - 1,
                                },
                                "cell": {
                                    "userEnteredFormat": run_fmt,
                                },
                                "fields": fields,
                            }
                        })

                if key:
                    run_start = col_idx
                    run_fmt = fmt
                    run_key = key
                else:
                    run_start = None
                    run_fmt = None
                    run_key = None

    # Google Sheets 版券商查詢：B2 下拉選單。
    if ws_xlsx.title == "券商查詢":
        try:
            # 使用 ONE_OF_LIST 直接寫入券商清單，比 ONE_OF_RANGE 更穩定，
            # 可避免部分環境下跨工作表範圍驗證未顯示下拉箭頭。
            sh = get_gsheet_spreadsheet()
            data_ws = sh.worksheet("券商查詢資料") if sh else None
            broker_values = []

            if data_ws:
                for value in data_ws.col_values(16)[1:]:
                    value = str(value).strip()
                    if value and value not in broker_values:
                        broker_values.append(value)

            if broker_values:
                requests.append({
                    "setDataValidation": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "endRowIndex": 2,
                            "startColumnIndex": 1,
                            "endColumnIndex": 2,
                        },
                        "rule": {
                            "condition": {
                                "type": "ONE_OF_LIST",
                                "values": [
                                    {"userEnteredValue": broker}
                                    for broker in broker_values
                                ],
                            },
                            "showCustomUi": True,
                            "strict": True,
                        },
                    }
                })
        except Exception as e:
            print(f"  ⚠️ 券商查詢下拉選單建立失敗：{type(e).__name__}: {e}")

    _gsheet_batch_update(requests)



# ══════════════════════════════════════════════════════════════════════
# Google Sheet 結果同步：補充 / 更新模式
# ══════════════════════════════════════════════════════════════════════

GSHEET_RESULT_UPSERT_ENABLED = os.getenv("GSHEET_RESULT_UPSERT_ENABLED", "1").strip().lower() not in ("0", "false", "no")
GSHEET_LEGACY_SCOPE_LABEL = os.getenv("GSHEET_LEGACY_SCOPE_LABEL", "未標記舊資料").strip() or "未標記舊資料"

GSHEET_RESULT_UPSERT_TITLES = {
    *AMOUNT_CLASS_SHEET_NAMES,
    "每日賣出明細",
    "近兩月買賣金額排行",
    "近兩月分點數排行",
    TOP15_POSITION_DETAIL_SHEET,
    TOP15_CONSENSUS_SHEET,
    BROKER_10D_DETAIL_SHEET,
    BROKER_10D_WINRATE_RANK_SHEET,
    WARRANT_CONSENSUS_7D_SHEET,
}

# 這兩張是「目前近兩月排名快照」，不是逐日歷史明細。
# 同一資料範圍每次都應以本次結果完整替換，否則跌出排名或已刪除的分點會殘留。
GSHEET_RESULT_REPLACE_CURRENT_SCOPE_TITLES = set()

GSHEET_RESULT_OVERWRITE_TITLES = {
    "券商查詢",
    "券商查詢資料",
    "股票ABCDE查詢",
    "股票ABCDE查詢資料",
    "價格抓取狀態",
    "顏色說明",
}

# RUN_MODE=1 精選五分點模式只用來產出每日圖卡與精選資料，
# 不應覆蓋全分點模式才需要維護的查詢頁與勝率統計頁。
# 這些工作表只在 RUN_MODE=2 全分點模式同步到 Google Sheet。
GSHEET_RUN_MODE1_SKIP_RESULT_TITLES = {
    "券商查詢",
    "券商查詢資料",
    "股票ABCDE查詢",
    "股票ABCDE查詢資料",
    "勝率統計",
    "ABCDE組合勝率",
}


def should_skip_result_sheet_in_run_mode(title):
    title = safe_worksheet_title(title)
    return RUN_MODE == 1 and title in GSHEET_RUN_MODE1_SKIP_RESULT_TITLES


def get_result_data_scope():
    """
    Google Sheet 結果同步用的資料範圍標籤。

    目的：
    - RUN_MODE=1 跑精選五分點時，只更新「資料範圍=精選五分點」的資料。
    - RUN_MODE=2 跑完整分點時，只更新「資料範圍=全分點」的資料。
    - 同一張 Google Sheet 內不同資料範圍可以並存，避免五分點覆蓋全分點資料。
    """
    return "精選五分點" if RUN_MODE == 1 else "全分點"


def result_record_broker_scope(record):
    """依 Google Sheet 資料範圍判斷該列應套用哪一份有效分點清單。"""
    data_scope = str(
        strip_gsheet_text_prefix(record.get("資料範圍", ""))
    ).strip()

    if data_scope == "精選五分點":
        return "selected5"

    # 全分點與舊版未標記資料，都用完整分點清單清理。
    return "all"


def normalize_or_remove_deleted_broker_result_record(record):
    """
    清理 Google Sheet 已存在的結果列。

    規則：
    1. 有「分點／券商代號」的資料，若該分點已從目前設定移除，整列刪除。
    2. 若只是分點標籤改名，但券商代號仍存在，自動改成目前的新標籤。
    3. 「買進分點」為單一分點欄位時，也會套用相同清理。
    4. 沒有任何分點識別欄位的彙總資料保持不變。
    """
    rec = dict(record)
    scope = result_record_broker_scope(rec)
    label_to_code, code_to_label = configured_broker_pair_maps_for_scope(scope)
    active_labels = set(label_to_code.keys())

    broker_label = str(
        strip_gsheet_text_prefix(rec.get("分點", ""))
    ).strip()
    broker_code = normalize_broker_code_for_compare(
        strip_gsheet_text_prefix(rec.get("券商代號", ""))
    )

    if broker_label or broker_code:
        if broker_code:
            canonical_identity = code_to_label.get(broker_code)
            if canonical_identity is None:
                return None

            canonical_label, canonical_code = canonical_identity
            if "分點" in rec:
                rec["分點"] = canonical_label
            if "券商代號" in rec:
                rec["券商代號"] = canonical_code
            return rec

        canonical_label = _resolve_active_broker_label(broker_label, scope=scope)
        if not canonical_label:
            return None

        if "分點" in rec:
            rec["分點"] = canonical_label
        if "券商代號" in rec:
            rec["券商代號"] = label_to_code.get(canonical_label, "")
        return rec

    buy_broker_label = str(
        strip_gsheet_text_prefix(rec.get("買進分點", ""))
    ).strip()
    if buy_broker_label:
        canonical_label = _resolve_active_broker_label(buy_broker_label, scope=scope)
        if not canonical_label:
            return None
        if "買進分點" in rec:
            rec["買進分點"] = canonical_label
        return rec

    return rec


def should_upsert_result_sheet(title):
    title = safe_worksheet_title(title)

    if not GSHEET_RESULT_UPSERT_ENABLED:
        return False

    if title in GSHEET_RESULT_OVERWRITE_TITLES:
        return False

    if title in GSHEET_RESULT_UPSERT_TITLES:
        return True

    return False


def read_existing_worksheet_values(title):
    try:
        sh = get_gsheet_spreadsheet()
        if sh is None:
            return []
        ws = sh.worksheet(safe_worksheet_title(title))
        values = ws.get_all_values()
        return values or []
    except Exception as exc:
        print(f"  ⚠️ Google Sheet 結果表讀取失敗：{safe_worksheet_title(title)}，原因：{type(exc).__name__}: {exc}")
        return []


def _find_simple_header_row(values):
    if not values:
        return None

    scan_limit = min(len(values), 10)
    key_headers = {
        "統計日期", "事件類型", "日期", "排名", "權證代號", "標的股", "分點",
        "買進金額", "賣出金額", "淨買超成本", "排名類型",
    }

    for idx in range(scan_limit):
        row = [str(x).strip() for x in values[idx]]
        non_empty = [x for x in row if x]
        if not non_empty:
            continue
        if len(set(row) & key_headers) >= 2:
            return idx

    return 0 if values and any(str(x).strip() for x in values[0]) else None


def _values_to_records(values, header_row_idx=0, default_scope=None):
    if not values or header_row_idx is None or header_row_idx >= len(values):
        return [], []

    headers = [str(h).strip() for h in values[header_row_idx]]
    if not headers or all(h == "" for h in headers):
        return [], []

    # 去掉尾端完全空白表頭，避免 Google Sheet 舊資料多出空欄造成 key 錯亂。
    while headers and headers[-1] == "":
        headers.pop()

    records = []
    for raw_row in values[header_row_idx + 1:]:
        row = list(raw_row)
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = row[:len(headers)]

        if not any(str(v).strip() for v in row):
            continue

        rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(headers)}
        if default_scope is not None:
            rec["資料範圍"] = str(rec.get("資料範圍", "") or default_scope).strip()
        records.append(rec)

    return headers, records


def _records_to_values(headers, records):
    out = [list(headers)]
    for rec in records:
        out.append([rec.get(h, "") for h in headers])
    return out


def _ensure_scope_header(headers):
    headers = [str(h).strip() for h in headers]
    if "資料範圍" not in headers:
        return ["資料範圍"] + headers
    return headers


def _sheet_upsert_key_columns(title, headers):
    title = safe_worksheet_title(title)
    hset = set(headers)

    def keep(cols):
        return [c for c in cols if c in hset]

    if title in AMOUNT_CLASS_SHEET_NAMES:
        return keep(["資料範圍", "事件類型", "分點", "券商代號", "標的股", "事件日", "權證清單"])

    if title == TOP15_CONSENSUS_SHEET:
        return keep(["資料範圍", "統計日期", "標的股"])

    if title == TOP15_POSITION_DETAIL_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號", "標的股", "權證代號", "事件", "事件日", "買進日"])

    if title == WARRANT_CONSENSUS_7D_SHEET:
        # 近7／14／21日共識皆為「標的層級」排名；統計天數必須納入 key，避免三個期間互相覆蓋。
        return keep(["資料範圍", "統計日期", "統計天數", "排名類型", "標的股"])

    if title == BROKER_10D_DETAIL_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號", "標的股"])

    if title == BROKER_10D_WINRATE_RANK_SHEET:
        return keep(["資料範圍", "統計日期", "分點", "券商代號"])


    if title == "每日賣出明細":
        return keep(["資料範圍", "日期", "分點", "券商代號", "標的股", "權證代號", "事件", "狀態", "事件日"])

    if title == "近兩月買賣金額排行":
        return keep(["資料範圍", "權證代號", "買進分點"])

    if title == "近兩月分點數排行":
        return keep(["資料範圍", "標的股", "買進分點清單"])

    # 通用備援：至少要有資料範圍，加上一些穩定欄位。
    cols = keep([
        "資料範圍", "統計日期", "日期", "事件類型", "分點", "券商代號",
        "標的股", "權證代號", "買進日", "事件日", "起始日", "結束日",
        "排名類型",
    ])
    return cols if len(cols) >= 2 else []



def normalize_underlying_code_for_group(value, fallback_text=""):
    """
    近7日共識與 Google Sheet upsert 用的標的股代號正規化。

    修正目的：
    同一標的可能因 Google Sheet / pandas 讀寫變成 2408、'2408、2408.0，
    若直接拿原字串當 group key，就會造成本週 TOP15 同標的重複出現。
    這裡統一轉成穩定代號後再合併與排序。
    """
    candidates = []
    for raw in (value, fallback_text):
        if raw is None:
            continue
        s = strip_gsheet_text_prefix(str(raw)).strip()
        if not s:
            continue
        candidates.append(s)

    for s in candidates:
        s = s.replace("，", ",").replace(",", "").strip()
        if s.endswith(".0") and s[:-2].isdigit():
            s = s[:-2]

        m = re.match(r"^(\d{4,6})(?:\s|$)", s)
        if m:
            return m.group(1)

        if s.isdigit() and 4 <= len(s) <= 6:
            return s

        digits = "".join(ch for ch in s if ch.isdigit())
        if 4 <= len(digits) <= 6:
            return digits

    return ""

def _record_key(rec, key_cols):
    parts = []
    for col in key_cols:
        value = rec.get(col, "")
        if col in ("日期", "統計日期", "買進日", "事件日", "起始日", "結束日", "第一筆日期", "最後筆日期"):
            value = normalize_date_str(strip_gsheet_text_prefix(value))
        elif col in ("標的股", "標的代號", "標的代碼"):
            value = normalize_underlying_code_for_group(value)
        else:
            value = strip_gsheet_text_prefix(value)
        parts.append(str(value).strip())
    return tuple(parts)



def _result_retention_policy(title):
    """回傳需要裁切的日期欄與保留的不同日期數。"""
    title = safe_worksheet_title(title)

    if title == "每日賣出明細":
        return {
            "date_col": "日期",
            "keep_count": GSHEET_DAILY_SELL_KEEP_TRADING_DAYS,
            "label": "交易日",
        }

    if title in (TOP15_POSITION_DETAIL_SHEET, TOP15_CONSENSUS_SHEET):
        return {
            "date_col": "統計日期",
            "keep_count": GSHEET_TOP15_KEEP_STAT_DATES,
            "label": "統計日期",
        }

    return None


def _parse_result_record_date(value):
    """解析結果表日期；同時支援 yyyy/mm/dd 與 Google Sheet 日期序號。"""
    raw = strip_gsheet_text_prefix(value)
    dt = parse_date(raw)
    if dt is not None:
        return dt

    try:
        serial = float(str(raw).replace(",", "").strip())
        if 20000 <= serial <= 100000:
            return datetime(1899, 12, 30) + timedelta(days=serial)
    except Exception:
        pass

    return None


def _split_records_by_recent_dates(records, date_col, keep_count):
    """
    依「資料範圍」分開保留最近 N 個不同日期。

    精選五分點與全分點可能不是同一天更新，因此不能用全表共同日期裁切，
    否則更新較頻繁的一邊會把另一邊的日期提前擠出主表。
    """
    keep_count = max(int(keep_count or 1), 1)
    dates_by_scope = {}

    for rec in records:
        scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip() or GSHEET_LEGACY_SCOPE_LABEL
        dt = _parse_result_record_date(rec.get(date_col, ""))
        if dt:
            dates_by_scope.setdefault(scope, set()).add(dt.date())

    keep_dates_by_scope = {
        scope: set(sorted(date_set, reverse=True)[:keep_count])
        for scope, date_set in dates_by_scope.items()
    }

    retained = []
    archive = []

    for rec in records:
        scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip() or GSHEET_LEGACY_SCOPE_LABEL
        dt = _parse_result_record_date(rec.get(date_col, ""))

        # 日期無法辨識時留在主表，避免誤刪舊版或人工資料。
        if dt is None:
            retained.append(rec)
            continue

        if dt.date() in keep_dates_by_scope.get(scope, set()):
            retained.append(rec)
        else:
            archive.append(rec)

    return retained, archive


def _archive_year_label(rec, date_col):
    if not GSHEET_ARCHIVE_YEARLY_SPLIT:
        return "全部"

    dt = _parse_result_record_date(rec.get(date_col, ""))
    return str(dt.year) if dt else "未分類"


def _archive_spreadsheet_name(year_label, part_no=1):
    base = f"{GSHEET_ARCHIVE_NAME_PREFIX}_{year_label}"
    return base if int(part_no) <= 1 else f"{base}_{int(part_no):02d}"


def _explicit_archive_share_emails():
    if not GSHEET_ARCHIVE_SHARE_EMAILS:
        return []

    out = []
    for email in re.split(r"[,;；、\n\r\t]+", GSHEET_ARCHIVE_SHARE_EMAILS):
        email = str(email).strip()
        if email and "@" in email and email not in out:
            out.append(email)
    return out


def _sync_archive_permissions(archive_sh):
    """
    新建立的封存檔由服務帳號擁有時，盡量複製主試算表既有使用者 / 群組權限；
    也可透過 GSHEET_ARCHIVE_SHARE_EMAILS 明確指定要分享的信箱。
    """
    if archive_sh is None:
        return

    archive_id = str(getattr(archive_sh, "id", "") or "")
    if archive_id and archive_id in _GSHEET_ARCHIVE_PERMISSION_SYNCED:
        return

    share_targets = {}

    try:
        primary_sh = get_gsheet_spreadsheet()
        if primary_sh is not None and hasattr(primary_sh, "list_permissions"):
            for permission in primary_sh.list_permissions() or []:
                ptype = str(permission.get("type", "")).strip()
                role = str(permission.get("role", "")).strip()
                email = str(permission.get("emailAddress", "")).strip()

                if ptype not in ("user", "group") or not email:
                    continue
                if role not in ("owner", "reader", "writer", "commenter"):
                    continue

                # 主檔 owner 在封存檔改授予 writer；commenter 則降成 reader。
                if role == "owner":
                    mapped_role = "writer"
                elif role == "commenter":
                    mapped_role = "reader"
                else:
                    mapped_role = role
                share_targets[(ptype, email)] = mapped_role
    except Exception:
        pass

    for email in _explicit_archive_share_emails():
        share_targets[("user", email)] = "writer"

    for (ptype, email), role in share_targets.items():
        try:
            archive_sh.share(
                email,
                perm_type=ptype,
                role=role,
                notify=False,
            )
        except Exception:
            pass

    if archive_id:
        _GSHEET_ARCHIVE_PERMISSION_SYNCED.add(archive_id)


def is_gsheet_drive_storage_quota_error(exc):
    """判斷是否為 Google Drive 永久性儲存空間不足，而不是可藉由換檔名重試的錯誤。"""
    msg = str(exc or "").strip().lower()
    return (
        "drive storage quota has been exceeded" in msg
        or "storage quota has been exceeded" in msg
        or "storagequota" in msg
        or ("403" in msg and "drive" in msg and "quota" in msg)
    )


def _configured_archive_spreadsheet_id(part_no):
    try:
        idx = int(part_no) - 1
    except Exception:
        return ""

    if 0 <= idx < len(GSHEET_ARCHIVE_SPREADSHEET_IDS):
        return str(GSHEET_ARCHIVE_SPREADSHEET_IDS[idx]).strip()
    return ""


def _archive_worksheet_title(title, year_label):
    """
    使用固定封存試算表 ID 時，同一份檔案可能保存多個年度，
    因此把年度放進工作表名稱，避免 2026 與 2027 資料混在同一張分頁。
    """
    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        return safe_worksheet_title(f"{title}_{year_label}")
    return safe_worksheet_title(title)


def _mark_archive_create_blocked(exc):
    global _GSHEET_ARCHIVE_CREATE_BLOCKED
    global _GSHEET_ARCHIVE_CREATE_BLOCK_REASON
    global _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED

    _GSHEET_ARCHIVE_CREATE_BLOCKED = True
    _GSHEET_ARCHIVE_CREATE_BLOCK_REASON = str(exc or "Google Drive storage quota exceeded").strip()

    if not _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED:
        print(
            "  ⚠️ Google Drive 儲存空間不足，本次執行已停止自動建立新的封存試算表；"
            "不會再嘗試 _02、_03……。請先用自己的 Google 帳號建立封存試算表、"
            "分享給服務帳號，並設定 GSHEET_ARCHIVE_SPREADSHEET_ID。"
        )
        _GSHEET_ARCHIVE_CREATE_BLOCK_LOGGED = True


def _open_archive_spreadsheet(year_label, part_no=1, create_if_missing=False):
    key = (str(year_label), int(part_no))
    if key in _GSHEET_ARCHIVE_SPREADSHEETS:
        return _GSHEET_ARCHIVE_SPREADSHEETS[key]

    gc = get_gsheet_client()
    if gc is None:
        return None

    # 優先使用使用者自行建立、由使用者帳號持有的封存試算表。
    # 這不會消耗服務帳號的個人 Drive 配額。
    configured_id = _configured_archive_spreadsheet_id(part_no)
    if configured_id:
        try:
            sh = gc.open_by_key(configured_id)
            _GSHEET_ARCHIVE_SPREADSHEETS[key] = sh
            _sync_archive_permissions(sh)
            return sh
        except Exception as e:
            print(
                f"  ⚠️ 開啟指定 Google Sheet 封存檔失敗：分卷={int(part_no):02d}｜"
                f"ID={configured_id}｜原因：{type(e).__name__}: {e}"
            )
            return None

    # 已設定固定封存檔清單時，只使用清單內的檔案；不再由服務帳號另外建立新檔。
    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        return None

    name = _archive_spreadsheet_name(year_label, part_no)
    sh = None

    try:
        sh = gc.open(name)
    except Exception:
        if not create_if_missing or not GSHEET_ARCHIVE_ALLOW_CREATE:
            return None

        # Drive quota 屬於永久性錯誤，本次執行第一次遇到後便停止建立，避免連續刷出 _02～_99。
        if _GSHEET_ARCHIVE_CREATE_BLOCKED:
            return None

        try:
            if GSHEET_ARCHIVE_FOLDER_ID:
                try:
                    sh = gc.create(name, folder_id=GSHEET_ARCHIVE_FOLDER_ID)
                except TypeError:
                    # 相容較舊版 gspread；舊版 create() 沒有 folder_id 參數。
                    sh = gc.create(name)
            else:
                sh = gc.create(name)
            print(f"  📦 已建立 Google Sheet 年度封存檔：{name}")
        except Exception as e:
            if is_gsheet_drive_storage_quota_error(e):
                _mark_archive_create_blocked(e)
            else:
                print(f"  ⚠️ 建立 Google Sheet 封存檔失敗：{name}，原因：{type(e).__name__}: {e}")
            return None

    _GSHEET_ARCHIVE_SPREADSHEETS[key] = sh
    _sync_archive_permissions(sh)
    return sh


def _is_blank_archive_worksheet(ws):
    try:
        values = _worksheet_values_with_formulas(ws)
        return not any(str(v).strip() for row in values for v in row)
    except Exception:
        return False


def _prepare_archive_target_worksheet(archive_sh, title, target_rows, target_cols):
    """
    檢查指定封存檔是否還容得下這張工作表；若可容納，回傳可寫入的 worksheet。
    """
    if archive_sh is None:
        return None

    title = safe_worksheet_title(title)
    target_rows = max(int(target_rows), 1)
    target_cols = max(int(target_cols), 1)

    existing_ws = None
    try:
        existing_ws = archive_sh.worksheet(title)
    except Exception:
        existing_ws = None

    total_cells = spreadsheet_grid_cell_count(archive_sh)
    current_target_cells = 0
    blank_default_ws = None

    if existing_ws is not None:
        current_target_cells = max(int(existing_ws.row_count), 1) * max(int(existing_ws.col_count), 1)
    else:
        try:
            sheets = archive_sh.worksheets()
            if len(sheets) == 1 and _is_blank_archive_worksheet(sheets[0]):
                blank_default_ws = sheets[0]
                current_target_cells = max(int(blank_default_ws.row_count), 1) * max(int(blank_default_ws.col_count), 1)
        except Exception:
            blank_default_ws = None

    projected_cells = total_cells - current_target_cells + target_rows * target_cols
    if projected_cells > GSHEET_ARCHIVE_MAX_CELLS:
        return None

    if existing_ws is not None:
        return existing_ws

    if blank_default_ws is not None:
        try:
            gsheet_api_call(
                f"重新命名封存工作表 {title}",
                blank_default_ws.update_title,
                title,
            )
            return blank_default_ws
        except Exception:
            pass

    try:
        return gsheet_api_call(
            f"建立封存工作表 {title}",
            archive_sh.add_worksheet,
            title=title,
            rows=target_rows,
            cols=target_cols,
        )
    except Exception as e:
        print(f"  ⚠️ 建立封存工作表失敗：{title}，原因：{type(e).__name__}: {e}")
        return None


def _read_archive_worksheet_values(archive_sh, title):
    if archive_sh is None:
        return []

    try:
        ws = archive_sh.worksheet(safe_worksheet_title(title))
        return ws.get_all_values() or []
    except Exception:
        return []


def _merge_records_for_archive(title, incoming_headers, incoming_records, old_values):
    old_headers, old_records = _values_to_records(
        old_values,
        header_row_idx=0,
        default_scope=GSHEET_LEGACY_SCOPE_LABEL,
    )
    old_headers = _ensure_scope_header(old_headers) if old_headers else []

    headers = []
    for h in list(incoming_headers) + list(old_headers):
        h = str(h).strip()
        if h and h not in headers:
            headers.append(h)

    key_cols = _sheet_upsert_key_columns(title, headers)
    if not key_cols or "資料範圍" not in key_cols:
        # 理論上目前三張封存表一定有穩定 key；若未來欄位改名，至少保留原順序追加。
        return headers, old_records + list(incoming_records)

    merged_map = {}
    order = []

    for rec in list(old_records) + list(incoming_records):
        if not rec.get("資料範圍"):
            rec["資料範圍"] = GSHEET_LEGACY_SCOPE_LABEL

        key = _record_key(rec, key_cols)
        if not any(key):
            continue

        if key not in merged_map:
            order.append(key)
        merged_map[key] = rec

    return headers, [merged_map[key] for key in order]


def _write_values_to_archive_worksheet(ws, values):
    """封存檔使用的精簡寫入：完整覆蓋、保留代號前導 0，並把格數縮到實際大小。"""
    if ws is None:
        return False

    if not values:
        values = [[""]]

    row_count = max(len(values), 1)
    col_count = max(max((len(row) for row in values), default=1), 1)
    normalized_values = []

    for row in values:
        row = list(row)
        if len(row) < col_count:
            row += [""] * (col_count - len(row))
        normalized_values.append([clean_gsheet_value(v) for v in row])

    normalized_values = normalize_gsheet_values_for_text_columns(normalized_values)

    try:
        gsheet_api_call(
            f"調整封存工作表大小 {ws.title}",
            ws.resize,
            rows=row_count,
            cols=col_count,
        )
        gsheet_api_call(f"清除封存工作表 {ws.title}", ws.clear)

        batch_ranges = []
        for start in range(0, len(normalized_values), GSHEET_CHUNK_ROWS):
            chunk = normalized_values[start:start + GSHEET_CHUNK_ROWS]
            batch_ranges.append({
                "range": f"A{start + 1}",
                "values": chunk,
            })
        gsheet_values_batch_update(
            ws,
            batch_ranges,
            f"批次寫入封存工作表 {ws.title}",
        )

        # 再 resize 一次，確保 clear / update 後沒有殘留多餘空白列欄。
        gsheet_api_call(
            f"最終縮減封存工作表 {ws.title}",
            ws.resize,
            rows=row_count,
            cols=col_count,
        )
        return True
    except Exception as e:
        print(f"  ⚠️ Google Sheet 封存寫入失敗：{ws.title}，原因：{type(e).__name__}: {e}")
        return False


def _archive_records_to_year_part(title, headers, records, year_label):
    """
    把同一年度資料寫入可容納的封存分卷。

    - 有設定 GSHEET_ARCHIVE_SPREADSHEET_IDS：依序使用既有試算表 ID。
    - 未設定 ID：沿用原本依名稱開啟 / 建立年度封存檔。
    - 一旦確認 Drive quota 不足，立即停止，不再嘗試 _02～_99。
    """
    if not records:
        return True

    max_parts = len(GSHEET_ARCHIVE_SPREADSHEET_IDS) if GSHEET_ARCHIVE_SPREADSHEET_IDS else 99
    archive_ws_title = _archive_worksheet_title(title, year_label)

    for part_no in range(1, max_parts + 1):
        archive_sh = _open_archive_spreadsheet(
            year_label,
            part_no=part_no,
            create_if_missing=True,
        )

        if archive_sh is None:
            if _GSHEET_ARCHIVE_CREATE_BLOCKED:
                break
            continue

        old_values = _read_archive_worksheet_values(archive_sh, archive_ws_title)
        merged_headers, merged_records = _merge_records_for_archive(
            title,
            headers,
            records,
            old_values,
        )
        merged_values = _records_to_values(merged_headers, merged_records)
        target_rows = max(len(merged_values), 1)
        target_cols = max(max((len(row) for row in merged_values), default=1), 1)

        ws = _prepare_archive_target_worksheet(
            archive_sh,
            archive_ws_title,
            target_rows=target_rows,
            target_cols=target_cols,
        )
        if ws is None:
            # 這一卷預估超過安全格數，改試下一個既有 ID 或下一個可建立分卷。
            continue

        if _write_values_to_archive_worksheet(ws, merged_values):
            archive_url = str(getattr(archive_sh, "url", "") or "").strip()
            print(
                f"  📦 Google Sheet 舊資料封存完成：{title}｜年度={year_label}｜"
                f"分卷={part_no:02d}｜工作表={archive_ws_title}｜本次 {len(records):,} 筆｜"
                f"封存表合計 {len(merged_records):,} 筆"
                + (f"｜{archive_url}" if archive_url else "")
            )
            return True

    if GSHEET_ARCHIVE_SPREADSHEET_IDS:
        print(
            f"  ⚠️ Google Sheet 舊資料封存失敗：{title}｜年度={year_label}，"
            "已設定的封存試算表 ID 均無法使用或已達安全格數。"
        )
    elif not _GSHEET_ARCHIVE_CREATE_BLOCKED:
        print(f"  ⚠️ Google Sheet 舊資料封存失敗：{title}｜年度={year_label}，找不到可用封存分卷。")
    return False


def archive_result_records(title, headers, records, date_col):
    """
    將主表裁切出的舊資料依年度寫入獨立 Google Sheet。

    只有全部年度都封存成功時才回傳 True；任一年度失敗時，主表會保留全部資料避免遺失。
    """
    if not records:
        return True

    if not GSHEET_RESULT_ARCHIVE_ENABLED:
        print(f"  ⚠️ 已停用 Google Sheet 自動封存，{title} 舊資料暫不從主表移除。")
        return False

    grouped = {}
    for rec in records:
        year_label = _archive_year_label(rec, date_col)
        grouped.setdefault(year_label, []).append(rec)

    all_ok = True
    for year_label in sorted(grouped.keys()):
        ok = _archive_records_to_year_part(
            title,
            headers,
            grouped[year_label],
            year_label,
        )
        all_ok = all_ok and ok

    return all_ok


def apply_result_retention_and_archive(title, headers, merged_records):
    policy = _result_retention_policy(title)
    if not policy:
        return merged_records, 0

    retained_records, archive_records = _split_records_by_recent_dates(
        merged_records,
        date_col=policy["date_col"],
        keep_count=policy["keep_count"],
    )

    if not archive_records:
        return retained_records, 0

    archive_ok = archive_result_records(
        title,
        headers,
        archive_records,
        date_col=policy["date_col"],
    )

    if not archive_ok:
        print(
            f"  ⚠️ {title} 有 {len(archive_records):,} 筆舊資料尚未完成封存，"
            "本次仍保留在主試算表，避免資料遺失。"
        )
        return merged_records, 0

    print(
        f"  🧹 Google Sheet 主表保留策略：{title}｜"
        f"每個資料範圍保留最近 {policy['keep_count']} 個{policy['label']}｜"
        f"主表 {len(retained_records):,} 筆｜已移至封存 {len(archive_records):,} 筆"
    )
    return retained_records, len(archive_records)


def merge_result_values_for_gsheet(
    title,
    new_values,
    data_scope=None,
    extra_scope_values=None,
):
    """
    將本次 Excel 結果與 Google Sheet 既有結果做 upsert 合併。

    規則：
    - 主 Excel 資料使用 data_scope；未指定時沿用目前 RUN_MODE。
    - extra_scope_values 若提供有效表頭，固定以「精選五分點」併入同一次 upsert；即使資料列為空，也視為本次已更新。
    - upsert key 本來就包含「資料範圍」，兩個 scope 不會互相覆蓋。
    - 舊資料若沒有「資料範圍」，會保留為「未標記舊資料」。
    - 每日賣出明細與 TOP15 快取超過主表保留日期的資料，會先封存。
    """
    if not should_upsert_result_sheet(title):
        return new_values

    header_row_idx = _find_simple_header_row(new_values)
    if header_row_idx is None or header_row_idx != 0:
        # 多段式報表仍維持原本覆蓋模式，避免破壞版面與公式。
        return new_values

    current_scope = str(data_scope or get_result_data_scope()).strip()
    extra_scope = "精選五分點"

    new_headers, new_records = _values_to_records(
        new_values,
        header_row_idx=0,
        default_scope=current_scope,
    )
    if not new_headers:
        return new_values

    new_headers = _ensure_scope_header(new_headers)

    def clean_incoming_records(records, scope, source_label):
        cleaned = []
        removed_count = 0

        for rec in records:
            rec = dict(rec)
            rec["資料範圍"] = scope
            cleaned_rec = normalize_or_remove_deleted_broker_result_record(rec)
            if cleaned_rec is None:
                removed_count += 1
                continue
            cleaned_rec["資料範圍"] = scope
            cleaned.append(cleaned_rec)

        if removed_count > 0:
            print(
                f"  🧹 {source_label}已排除失效分點資料："
                f"{safe_worksheet_title(title)}｜{removed_count:,} 筆"
            )

        return cleaned

    new_records = clean_incoming_records(new_records, current_scope, "本次主範圍結果")

    extra_headers = []
    extra_records = []
    extra_scope_supplied = False

    if extra_scope_values:
        extra_header_row_idx = _find_simple_header_row(extra_scope_values)

        if extra_header_row_idx == 0:
            # 即使只有表頭、沒有資料列，也代表本次已更新精選五分點範圍。
            extra_scope_supplied = True

            extra_headers, parsed_extra_records = _values_to_records(
                extra_scope_values,
                header_row_idx=0,
                default_scope=extra_scope,
            )
            extra_headers = _ensure_scope_header(extra_headers) if extra_headers else []
            extra_records = clean_incoming_records(
                parsed_extra_records,
                extra_scope,
                "本次精選五分點結果",
            )

    incoming_scopes = {current_scope}

    if extra_scope_supplied:
        incoming_scopes.add(extra_scope)

    old_values = read_existing_worksheet_values(title)
    old_headers, old_records = _values_to_records(
        old_values,
        header_row_idx=0,
        default_scope=GSHEET_LEGACY_SCOPE_LABEL,
    )
    old_headers = _ensure_scope_header(old_headers) if old_headers else []

    cleaned_old_records = []
    removed_deleted_broker_count = 0

    for rec in old_records:
        cleaned_rec = normalize_or_remove_deleted_broker_result_record(rec)
        if cleaned_rec is None:
            removed_deleted_broker_count += 1
            continue
        cleaned_old_records.append(cleaned_rec)

    old_records = cleaned_old_records

    # 近兩月兩張排行是目前狀態快照；本次有帶入的所有 scope 都完整替換。
    replaced_scope_counts = {}
    if safe_worksheet_title(title) in GSHEET_RESULT_REPLACE_CURRENT_SCOPE_TITLES:
        kept_old_records = []
        for rec in old_records:
            rec_scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip()
            if rec_scope in incoming_scopes:
                replaced_scope_counts[rec_scope] = replaced_scope_counts.get(rec_scope, 0) + 1
                continue
            kept_old_records.append(rec)
        old_records = kept_old_records

    if removed_deleted_broker_count > 0:
        print(
            f"  🧹 Google Sheet 已移除失效分點舊資料："
            f"{safe_worksheet_title(title)}｜{removed_deleted_broker_count:,} 筆"
        )

    for scope_name, replaced_count in sorted(replaced_scope_counts.items()):
        print(
            f"  ♻️ Google Sheet 排名快照完整替換："
            f"{safe_worksheet_title(title)}｜資料範圍={scope_name}｜"
            f"移除上次快照 {replaced_count:,} 筆"
        )

    headers = []
    for h in new_headers + extra_headers + old_headers:
        h = str(h).strip()
        if h and h not in headers:
            headers.append(h)

    if not headers:
        return new_values

    key_cols = _sheet_upsert_key_columns(title, headers)
    if not key_cols or "資料範圍" not in key_cols:
        return new_values

    old_map = {}
    old_order = []
    for rec in old_records:
        if not rec.get("資料範圍"):
            rec["資料範圍"] = GSHEET_LEGACY_SCOPE_LABEL
        key = _record_key(rec, key_cols)
        if not any(key):
            continue
        if key not in old_map:
            old_order.append(key)
        old_map[key] = rec

    new_map = {}
    new_order = []

    for rec in list(new_records) + list(extra_records):
        key = _record_key(rec, key_cols)
        if not any(key):
            continue
        if key not in new_map:
            new_order.append(key)
        new_map[key] = rec

    merged_records = []
    used_keys = set()

    # 本次全分點與精選五分點資料都排在前面。
    for key in new_order:
        merged_records.append(new_map[key])
        used_keys.add(key)

    # 沒有被本次 key 覆蓋的歷史資料保留。
    for key in old_order:
        if key in used_keys:
            continue
        merged_records.append(old_map[key])
        used_keys.add(key)

    retained_records, archived_count = apply_result_retention_and_archive(
        title,
        headers,
        merged_records,
    )
    merged_values = _records_to_values(headers, retained_records)

    scope_parts = [f"{current_scope}:{len(new_records):,}筆"]
    if extra_scope_supplied:
        scope_parts.append(f"{extra_scope}:{len(extra_records):,}筆")

    print(
        f"  ☁️ Google Sheet upsert：{safe_worksheet_title(title)}｜"
        f"本次 {'｜'.join(scope_parts)}｜舊資料 {len(old_records):,} 筆｜"
        f"合併後 {len(merged_records):,} 筆｜主表保留 {len(retained_records):,} 筆｜"
        f"封存 {archived_count:,} 筆"
    )

    return merged_values



def apply_safe_result_table_style_to_gsheet(gws, values=None):
    """
    針對 upsert 結果表套用安全版基本樣式。

    為什麼不用 apply_excel_style_to_gsheet：
    upsert 會在最左邊新增「資料範圍」欄位，而且會把舊資料與本次資料合併。
    如果繼續照原 Excel 欄位位置套樣式，日期格式會套到錯欄，
    例如「買進均價」會被顯示成 1899/12/30。

    這個函式只根據實際寫入 Google Sheet 的 values 表頭套安全樣式：
    - 凍結表頭列
    - 表頭粗體、淡黃色底、置中
    - 資料列垂直置中、換行
    - 不套任何日期或數字格式；日期 / 數字格式由後面的專用函式處理
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    header_row_idx = _find_simple_header_row(values)
    if header_row_idx is None:
        return

    max_cols = max(max((len(row) for row in values), default=1), 1)
    max_rows = max(len(values), 1)

    requests = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": header_row_idx + 1,
                    },
                },
                "fields": "gridProperties.frozenRowCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row_idx,
                    "endRowIndex": header_row_idx + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 1.0, "green": 0.949, "blue": 0.8},
                        "textFormat": {"bold": True},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "WRAP",
                    }
                },
                "fields": "userEnteredFormat.backgroundColor,userEnteredFormat.textFormat.bold,userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy",
            }
        },
    ]

    if max_rows > header_row_idx + 1:
        requests.append({
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row_idx + 1,
                    "endRowIndex": max_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": max_cols,
                },
                "cell": {
                    "userEnteredFormat": {
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "CLIP",
                    }
                },
                "fields": "userEnteredFormat.verticalAlignment,userEnteredFormat.wrapStrategy",
            }
        })
        requests.append({
            "updateDimensionProperties": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "ROWS",
                    "startIndex": header_row_idx + 1,
                    "endIndex": max_rows,
                },
                "properties": {
                    "pixelSize": 30,
                },
                "fields": "pixelSize",
            }
        })

    _gsheet_batch_update(requests)


def clear_all_number_formats_for_written_range(gws, values=None):
    """
    寫入結果後，先把實際資料範圍內的 numberFormat 全部清掉。

    這是為了處理舊版已經把「買進均價 / 減碼均價 / 出清獲利%」誤套成日期格式的工作表。
    單純寫入值不一定會移除舊格式，因此必須先清 numberFormat，再由日期 / 數字專用函式重套。
    """
    if gws is None or not values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    max_rows = max(len(values), 1)
    max_cols = max(max((len(row) for row in values), default=1), 1)

    requests = [{
        "repeatCell": {
            "range": {
                "sheetId": sheet_id,
                "startRowIndex": 0,
                "endRowIndex": max_rows,
                "startColumnIndex": 0,
                "endColumnIndex": max_cols,
            },
            "cell": {
                "userEnteredFormat": {}
            },
            "fields": "userEnteredFormat.numberFormat",
        }
    }]

    # 這裡只清 numberFormat，不動其他樣式。
    # 後續 TEXT / DATE / NUMBER 專用函式會依實際表頭重新套正確格式。
    _gsheet_batch_update(requests)



def _openpyxl_border_to_gsheet(cell):
    """
    只把 Excel 裡真正有設定的外框轉到 Google Sheet。

    用途：保留「減碼獲利% / 出清獲利%」的粗紅 / 粗綠外框。
    一般沒有外框的儲存格不處理，避免新增資料範圍欄位後破壞既有版面。
    """
    try:
        border = cell.border
    except Exception:
        return None

    if border is None:
        return None

    def side_to_gsheet(side):
        try:
            style = str(side.style or "").strip()
        except Exception:
            style = ""

        if not style:
            return None

        style_map = {
            "hair": "DOTTED",
            "dotted": "DOTTED",
            "dashDot": "DASHED",
            "dashDotDot": "DASHED",
            "dashed": "DASHED",
            "mediumDashDot": "DASHED",
            "mediumDashDotDot": "DASHED",
            "mediumDashed": "DASHED",
            "thin": "SOLID",
            "medium": "SOLID_MEDIUM",
            "thick": "SOLID_THICK",
            "double": "DOUBLE",
        }

        out = {"style": style_map.get(style, "SOLID")}
        color = _openpyxl_color_to_gsheet(getattr(side, "color", None))
        if color:
            out["color"] = color
        return out

    out = {}
    for attr, gname in [
        ("top", "top"),
        ("bottom", "bottom"),
        ("left", "left"),
        ("right", "right"),
    ]:
        side_fmt = side_to_gsheet(getattr(border, attr, None))
        if side_fmt:
            out[gname] = side_fmt

    return out or None


def _cell_gsheet_visual_format(cell):
    """
    只同步視覺樣式，不同步 numberFormat。

    目的：
    1. 保留 A/B/C/D/E 原本 D+1～D+20 的紅綠藍橘配色。
    2. 保留獲利欄位的粗紅 / 粗綠外框。
    3. 不把 Excel 舊欄位位置的日期格式帶進 Google Sheet，避免「買進均價」再次變成 1899/12/30。
    """
    fmt = {}

    bg = _openpyxl_fill_to_gsheet(cell)
    if bg:
        fmt["backgroundColor"] = bg

    text_format = _openpyxl_font_to_gsheet(cell)
    if text_format:
        fmt["textFormat"] = text_format

    align_format = _openpyxl_alignment_to_gsheet(cell)
    fmt.update(align_format)

    borders = _openpyxl_border_to_gsheet(cell)
    if borders:
        fmt["borders"] = borders

    return fmt


def _format_fields_visual(fmt):
    fields = []

    if "backgroundColor" in fmt:
        fields.append("userEnteredFormat.backgroundColor")
    if "textFormat" in fmt:
        for key in fmt["textFormat"].keys():
            fields.append(f"userEnteredFormat.textFormat.{key}")
    if "horizontalAlignment" in fmt:
        fields.append("userEnteredFormat.horizontalAlignment")
    if "verticalAlignment" in fmt:
        fields.append("userEnteredFormat.verticalAlignment")
    if "wrapStrategy" in fmt:
        fields.append("userEnteredFormat.wrapStrategy")
    if "borders" in fmt:
        fields.append("userEnteredFormat.borders")

    return ",".join(fields)


def _source_record_key_to_excel_row_map(title, source_values, final_headers, data_scopes=None):
    """
    建立 upsert 後資料列 key -> 原 Excel row number 對照。

    upsert 會把「資料範圍」插到最左邊，但原本 Excel 沒有這欄。
    這個 mapping 用 key 找回原 Excel 列，讓 Google Sheet 可以照原 Excel 的狀態色套回去，
    而不是用錯位的欄位位置硬套樣式。
    """
    header_row_idx = _find_simple_header_row(source_values)
    if header_row_idx is None:
        return {}, []

    source_headers = [str(h).strip() for h in source_values[header_row_idx]]
    while source_headers and source_headers[-1] == "":
        source_headers.pop()

    key_cols = _sheet_upsert_key_columns(title, final_headers)
    if not key_cols:
        return {}, source_headers

    scopes = [str(x).strip() for x in (data_scopes or [get_result_data_scope()]) if str(x).strip()]
    if not scopes:
        scopes = [get_result_data_scope()]
    out = {}

    for offset, raw_row in enumerate(source_values[header_row_idx + 1:], start=header_row_idx + 2):
        row = list(raw_row)
        if len(row) < len(source_headers):
            row = row + [""] * (len(source_headers) - len(row))
        elif len(row) > len(source_headers):
            row = row[:len(source_headers)]

        if not any(str(v).strip() for v in row):
            continue

        base_rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(source_headers)}

        # 精選五分點資料是全分點資料的子集合；同一份完整 Excel 列即可同時作為
        # 全分點與精選五分點列的視覺樣式來源，避免為了樣式再同步第二輪。
        for scope in scopes:
            rec = dict(base_rec)
            rec["資料範圍"] = scope
            key = _record_key(rec, key_cols)
            if any(key):
                out[key] = offset

    return out, source_headers


def _dplus_status_format_from_text(value):
    """
    依 D+ 欄位文字補回基本紅綠色。

    這只是保護機制：
    - 本次新資料會優先從原 Excel 樣式精準套回紅 / 綠 / 藍 / 橘。
    - 舊資料沒有原 Excel 樣式可參照時，至少依文字中的正負報酬補回紅綠底色。
    """
    s = str(value or "").strip()
    if not s:
        return None

    nums = []
    for m in re.finditer(r"([+-]?\d+(?:\.\d+)?)%", s):
        try:
            nums.append(float(m.group(1)))
        except Exception:
            pass

    if not nums:
        return None

    # 只要有正報酬，就以台股習慣用紅色；否則用綠色。
    fill = RED if any(v > 0 for v in nums) else GREEN
    color = _openpyxl_color_to_gsheet(fill.fgColor)
    if not color:
        return None

    return {
        "backgroundColor": color,
        "horizontalAlignment": "CENTER",
        "verticalAlignment": "MIDDLE",
        "wrapStrategy": "WRAP",
        "textFormat": {"foregroundColor": {"red": 0, "green": 0, "blue": 0}},
    }


def apply_upsert_original_excel_visual_style_to_gsheet(
    ws_xlsx,
    gws,
    source_values=None,
    final_values=None,
    title="",
    data_scope=None,
    extra_scopes=None,
):
    """
    upsert 工作表專用：在不套錯日期 / 數字格式的前提下，把原本 Excel 的配色套回 Google Sheet。

    修正重點：
    - 不再因為新增「資料範圍」欄位而放棄原本配色。
    - 透過表頭名稱與 upsert key 對齊原 Excel 列，避免欄位位移。
    - 只套背景色、字型、對齊與外框；numberFormat 仍交給後面的日期 / 數字專用函式處理。
    """
    if gws is None or ws_xlsx is None or not source_values or not final_values:
        return

    try:
        sheet_id = int(gws.id)
    except Exception:
        return

    title = safe_worksheet_title(title)
    final_header_row_idx = _find_simple_header_row(final_values)
    if final_header_row_idx is None:
        return

    final_headers = [str(h).strip() for h in final_values[final_header_row_idx]]
    while final_headers and final_headers[-1] == "":
        final_headers.pop()

    current_scope = str(data_scope or get_result_data_scope()).strip()
    source_scopes = [current_scope]
    for scope in extra_scopes or []:
        scope = str(scope).strip()
        if scope and scope not in source_scopes:
            source_scopes.append(scope)

    source_key_to_row, source_headers = _source_record_key_to_excel_row_map(
        title,
        source_values,
        final_headers,
        data_scopes=source_scopes,
    )
    if not final_headers:
        return

    source_col_by_header = {h: idx + 1 for idx, h in enumerate(source_headers) if h}
    key_cols = _sheet_upsert_key_columns(title, final_headers)
    if not key_cols:
        return

    requests = []

    # 先處理表頭：如果原 Excel 表頭有黃底等樣式，依欄名套回；資料範圍欄則維持安全樣式。
    for final_col_idx, header in enumerate(final_headers, start=1):
        source_col_idx = source_col_by_header.get(header)
        if not source_col_idx:
            continue

        src_cell = ws_xlsx.cell(1, source_col_idx)
        fmt = _cell_gsheet_visual_format(src_cell)
        fields = _format_fields_visual(fmt)
        if fmt and fields:
            requests.append({
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": final_header_row_idx,
                        "endRowIndex": final_header_row_idx + 1,
                        "startColumnIndex": final_col_idx - 1,
                        "endColumnIndex": final_col_idx,
                    },
                    "cell": {"userEnteredFormat": fmt},
                    "fields": fields,
                }
            })

    final_records = []
    for raw_row in final_values[final_header_row_idx + 1:]:
        row = list(raw_row)
        if len(row) < len(final_headers):
            row = row + [""] * (len(final_headers) - len(row))
        elif len(row) > len(final_headers):
            row = row[:len(final_headers)]

        rec = {h: row[i] if i < len(row) else "" for i, h in enumerate(final_headers)}
        final_records.append((row, rec))

    import json as _json

    for row_offset, (row_values, rec) in enumerate(final_records, start=final_header_row_idx + 2):
        rec_scope = str(strip_gsheet_text_prefix(rec.get("資料範圍", ""))).strip()
        key = _record_key(rec, key_cols)
        src_excel_row = source_key_to_row.get(key)

        run_start = None
        run_fmt = None
        run_key = None

        for final_col_idx in range(1, len(final_headers) + 2):
            fmt = None
            if final_col_idx <= len(final_headers):
                header = final_headers[final_col_idx - 1]
                source_col_idx = source_col_by_header.get(header)

                if src_excel_row and source_col_idx and rec_scope in source_scopes:
                    src_cell = ws_xlsx.cell(src_excel_row, source_col_idx)
                    fmt = _cell_gsheet_visual_format(src_cell)
                elif str(header).startswith("D+"):
                    # 舊資料或沒有對到原 Excel key 的資料，至少補回 D+ 欄位紅綠底色。
                    cell_value = row_values[final_col_idx - 1] if final_col_idx - 1 < len(row_values) else ""
                    fmt = _dplus_status_format_from_text(cell_value)

            key_json = _json.dumps(fmt, sort_keys=True, ensure_ascii=False) if fmt else None

            if key_json and run_key is None:
                run_start = final_col_idx
                run_fmt = fmt
                run_key = key_json
            elif key_json and key_json == run_key:
                continue
            else:
                if run_key and run_start is not None and run_fmt:
                    fields = _format_fields_visual(run_fmt)
                    if fields:
                        requests.append({
                            "repeatCell": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": row_offset - 1,
                                    "endRowIndex": row_offset,
                                    "startColumnIndex": run_start - 1,
                                    "endColumnIndex": final_col_idx - 1,
                                },
                                "cell": {"userEnteredFormat": run_fmt},
                                "fields": fields,
                            }
                        })

                if key_json:
                    run_start = final_col_idx
                    run_fmt = fmt
                    run_key = key_json
                else:
                    run_start = None
                    run_fmt = None
                    run_key = None

    _gsheet_batch_update(requests)

def worksheet_values_for_gsheet(ws_xlsx):
    values = []

    for row in ws_xlsx.iter_rows(values_only=False):
        row_values = []
        for cell in row:
            value = normalize_formula_for_gsheet(cell.value)
            row_values.append(clean_gsheet_value(value))
        values.append(row_values)

    return values or [[""]]


def read_excel_values_by_title(xlsx_path, allowed_titles=None):
    """只讀本機 Excel，整理成 {工作表名稱: values}，不呼叫 Google Sheet API。"""
    from openpyxl import load_workbook

    allowed = None
    if allowed_titles is not None:
        allowed = {safe_worksheet_title(x) for x in allowed_titles}

    wb = load_workbook(xlsx_path, data_only=False)
    result = {}

    for ws_xlsx in wb.worksheets:
        title = safe_worksheet_title(ws_xlsx.title)
        if allowed is not None and title not in allowed:
            continue
        result[title] = worksheet_values_for_gsheet(ws_xlsx)

    return result




def _normalize_warrant_code_for_identity(value):
    """將 Google Sheet／Excel 內的權證代號正規化後用於日期化身分查找。"""
    code = normalize_security_code_text(strip_gsheet_text_prefix(value))
    if code.isdigit() and len(code) == 5:
        code = code.zfill(6)
    return code


def _result_record_identity_date(record):
    """從結果列挑選最適合判斷權證代號重用區間的日期。"""
    for header in (
        "統計日期", "日期", "買進日", "事件日", "起始日", "結束日",
        "第一筆日期", "最後筆日期",
    ):
        value = record.get(header, "")
        if parse_date(value):
            return normalize_date_str(value)
    return ""


def _split_warrant_label_segments(value):
    """將「代號 名稱；代號 名稱」拆成保留分隔符的片段。"""
    text = str(strip_gsheet_text_prefix(value) or "")
    return re.split(r"([；;\n]+)", text)


def _warrant_meta_from_label_segment(segment, warrant_lookup):
    text = str(segment or "").strip()
    if not text:
        return None, ""
    match = re.match(r"^'?([0-9A-Za-z]{5,8})(?:\s+|$)(.*)$", text)
    if not match:
        return None, ""
    code = _normalize_warrant_code_for_identity(match.group(1))
    return warrant_lookup.get(code), code


def _repair_warrant_label_text(value, warrant_lookup):
    """只更正可由權證代號精準確認的名稱，不猜測無代號文字。"""
    parts = _split_warrant_label_segments(value)
    changed = False
    metas = []
    for idx in range(0, len(parts), 2):
        segment = parts[idx]
        meta, code = _warrant_meta_from_label_segment(segment, warrant_lookup)
        if not meta or not code:
            continue
        name = str(meta.get("名稱", "")).strip()
        if not name:
            continue
        replacement = f"{code} {name}".strip()
        if str(segment).strip() != replacement:
            leading = str(segment)[: len(str(segment)) - len(str(segment).lstrip())]
            trailing = str(segment)[len(str(segment).rstrip()):]
            parts[idx] = leading + replacement + trailing
            changed = True
        metas.append(meta)
    return "".join(parts), metas, changed


def repair_result_record_security_identity(record):
    """
    依「權證代號＋結果列日期」修正結果列的權證名稱與標的股／名稱。

    安全原則：
    1. 沒有權證代號就不以股名猜測。
    2. 多檔權證只有在全部指向同一標的時才更正標的股／名稱。
    3. 只修正身分欄位，不改金額、股數、FIFO、報酬率或事件分類。
    """
    rec = dict(record or {})
    if not WARRANT_IDENTITY_RECONCILE_ENABLED or not _CURRENT_WARRANT_INTERVAL_RECORDS:
        return rec, set()

    identity_date = _result_record_identity_date(rec)
    warrant_lookup = _warrant_lookup(_CURRENT_WARRANT_INTERVAL_RECORDS, identity_date or None)
    if not warrant_lookup:
        return rec, set()

    changed_fields = set()
    metas = []

    direct_code = _normalize_warrant_code_for_identity(rec.get("權證代號", ""))
    if direct_code:
        meta = warrant_lookup.get(direct_code)
        if meta:
            metas.append(meta)
            canonical_name = str(meta.get("名稱", "")).strip()
            if "權證代號" in rec and str(strip_gsheet_text_prefix(rec.get("權證代號", ""))).strip() != direct_code:
                rec["權證代號"] = direct_code
                changed_fields.add("權證代號")
            if "權證名稱" in rec and canonical_name and str(rec.get("權證名稱", "")).strip() != canonical_name:
                rec["權證名稱"] = canonical_name
                changed_fields.add("權證名稱")

    for header in ("權證清單", "最大單筆權證"):
        if header not in rec or not str(rec.get(header, "")).strip():
            continue
        repaired_text, label_metas, text_changed = _repair_warrant_label_text(rec.get(header, ""), warrant_lookup)
        metas.extend(label_metas)
        if text_changed:
            rec[header] = repaired_text
            changed_fields.add(header)

    # 去除同一權證因同時出現在直接欄位與清單欄位造成的重複。
    unique_metas = {}
    for meta in metas:
        code = _normalize_warrant_code_for_identity(meta.get("代號", ""))
        if code:
            unique_metas[code] = meta
    metas = list(unique_metas.values())

    underlying_pairs = {
        (
            normalize_security_code_text(meta.get("標的股", "")),
            str(meta.get("標的名稱", "")).strip(),
        )
        for meta in metas
        if normalize_security_code_text(meta.get("標的股", ""))
    }
    if len(underlying_pairs) == 1:
        underlying_code, underlying_name = next(iter(underlying_pairs))
        master_name = str(_CURRENT_STOCK_CODE_TO_NAME.get(underlying_code, underlying_name)).strip()
        if "標的股" in rec and underlying_code and normalize_security_code_text(rec.get("標的股", "")) != underlying_code:
            rec["標的股"] = underlying_code
            changed_fields.add("標的股")
        if "標的名稱" in rec and master_name and str(rec.get("標的名稱", "")).strip() != master_name:
            rec["標的名稱"] = master_name
            changed_fields.add("標的名稱")

    return rec, changed_fields


def repair_existing_worksheet_security_identities(ws, title, existing_values):
    """
    對既有 Google Sheet 做「精準儲存格更正」，不清表、不重建、不覆寫整列。

    這一步是必要的，因為純 insert 模式不會覆蓋同 key 的舊錯配資料；若不修正，
    舊的「1301 晶豪科」之類列會永久留在工作表並繼續污染產圖。
    """
    if (
        not GSHEET_REPAIR_SECURITY_IDENTITY_ROWS
        or ws is None
        or not existing_values
        or not _CURRENT_WARRANT_INTERVAL_RECORDS
    ):
        return existing_values, 0
    header_idx = _find_simple_header_row(existing_values)
    if header_idx != 0:
        return existing_values, 0

    headers = [str(h).strip() for h in existing_values[0]]
    while headers and headers[-1] == "":
        headers.pop()
    if not headers:
        return existing_values, 0
    header_to_col = {header: idx for idx, header in enumerate(headers) if header}
    identity_headers = {
        "權證代號", "權證名稱", "權證清單", "最大單筆權證", "標的股", "標的名稱"
    }
    if not (set(header_to_col) & identity_headers):
        return existing_values, 0

    repaired_values = [list(row) for row in existing_values]
    cell_updates = []
    repaired_rows = 0
    for row_no in range(2, len(repaired_values) + 1):
        row = repaired_values[row_no - 1]
        padded = row + [""] * max(len(headers) - len(row), 0)
        record = {header: padded[col_idx] for header, col_idx in header_to_col.items()}
        repaired_record, changed_fields = repair_result_record_security_identity(record)
        changed_fields = [field for field in changed_fields if field in header_to_col]
        if not changed_fields:
            continue
        repaired_rows += 1
        for field in changed_fields:
            col_idx = header_to_col[field]
            new_value = repaired_record.get(field, "")
            while len(row) <= col_idx:
                row.append("")
            row[col_idx] = new_value
            cell_updates.append({
                "range": f"{get_column_letter(col_idx + 1)}{row_no}",
                "values": [[clean_gsheet_value(new_value)]],
            })

    if cell_updates:
        gsheet_values_batch_update(
            ws,
            cell_updates,
            f"修正權證與標的身分 {safe_worksheet_title(title)}",
            chunk_size=5000,
        )
        print(
            f"  🧭 Google Sheet 身分欄位精準更正：{safe_worksheet_title(title)}｜"
            f"{repaired_rows:,} 列｜{len(cell_updates):,} 格｜未覆寫其他欄位"
        )
    return repaired_values, repaired_rows


def _prepare_incremental_result_values(title, new_values, data_scope=None, extra_scope_values=None):
    """整理本次可增量插入的簡單表格資料，不讀取或覆寫既有工作表。"""
    if not should_upsert_result_sheet(title):
        return None

    header_row_idx = _find_simple_header_row(new_values)
    if header_row_idx != 0:
        return None

    current_scope = str(data_scope or get_result_data_scope()).strip()
    headers, records = _values_to_records(new_values, header_row_idx=0, default_scope=current_scope)
    if not headers:
        return None
    headers = _ensure_scope_header(headers)

    cleaned_records = []
    for rec in records:
        rec = dict(rec)
        rec["資料範圍"] = current_scope
        cleaned = normalize_or_remove_deleted_broker_result_record(rec)
        if cleaned is not None:
            cleaned, _identity_changes = repair_result_record_security_identity(cleaned)
            cleaned["資料範圍"] = current_scope
            cleaned_records.append(cleaned)

    extra_headers = []
    if extra_scope_values:
        extra_idx = _find_simple_header_row(extra_scope_values)
        if extra_idx == 0:
            extra_headers, extra_records = _values_to_records(
                extra_scope_values,
                header_row_idx=0,
                default_scope="精選五分點",
            )
            extra_headers = _ensure_scope_header(extra_headers) if extra_headers else []
            for rec in extra_records:
                rec = dict(rec)
                rec["資料範圍"] = "精選五分點"
                cleaned = normalize_or_remove_deleted_broker_result_record(rec)
                if cleaned is not None:
                    cleaned, _identity_changes = repair_result_record_security_identity(cleaned)
                    cleaned["資料範圍"] = "精選五分點"
                    cleaned_records.append(cleaned)

    combined_headers = []
    for header in headers + extra_headers:
        header = str(header).strip()
        if header and header not in combined_headers:
            combined_headers.append(header)

    if not combined_headers:
        return None
    return combined_headers, cleaned_records


def _existing_result_sheet(primary_sh, title):
    title = safe_worksheet_title(title)
    try:
        return primary_sh.worksheet(title), False
    except Exception:
        ws = gsheet_api_call(
            f"第一次建立結果工作表 {title}",
            primary_sh.add_worksheet,
            title=title,
            rows=100,
            cols=20,
        )
        return ws, True


def ensure_worksheet_grid_capacity(ws, required_rows=None, required_cols=None):
    """
    只在既有 Google Sheet 工作表格數不足時向外擴充。

    - 不縮小工作表。
    - 不清空、不覆寫任何既有資料。
    - 增量同步新增欄位前先呼叫，避免寫入 AB1 等超出目前 grid limits 的範圍。
    """
    if ws is None:
        return False

    try:
        current_rows = max(int(getattr(ws, "row_count", 1) or 1), 1)
        current_cols = max(int(getattr(ws, "col_count", 1) or 1), 1)
        target_rows = max(current_rows, int(required_rows or 1), 1)
        target_cols = max(current_cols, int(required_cols or 1), 1)

        if target_rows == current_rows and target_cols == current_cols:
            return True

        gsheet_api_call(
            f"擴充工作表格數 {ws.title}",
            ws.resize,
            rows=target_rows,
            cols=target_cols,
        )
        print(
            f"  ↔️ Google Sheet 工作表自動擴充：{ws.title}｜"
            f"{current_rows:,}×{current_cols:,} → {target_rows:,}×{target_cols:,}"
        )
        return True
    except Exception as exc:
        print(
            f"  ⚠️ Google Sheet 工作表擴充失敗：{getattr(ws, 'title', '-')}，"
            f"原因：{type(exc).__name__}: {exc}"
        )
        return False

def _insert_rows_below_header(ws, rows):
    """由後往前分批插入第 2 列，維持本次資料原排序且不覆寫既有列。"""
    if not rows:
        return True

    normalized = normalize_gsheet_values_for_text_columns(
        [[clean_gsheet_value(value) for value in row] for row in rows]
    )
    chunks = [
        normalized[start:start + GSHEET_CHUNK_ROWS]
        for start in range(0, len(normalized), GSHEET_CHUNK_ROWS)
    ]

    try:
        for chunk in reversed(chunks):
            gsheet_api_call(
                f"增量插入工作表資料 {ws.title}",
                ws.insert_rows,
                chunk,
                row=2,
                value_input_option="USER_ENTERED",
                inherit_from_before=False,
            )
        return True
    except Exception as exc:
        print(f"  ⚠️ Google Sheet 增量插入失敗：{ws.title}，原因：{type(exc).__name__}: {exc}")
        return False


def insert_missing_result_rows_to_worksheet(ws, title, new_values, data_scope=None, extra_scope_values=None):
    """
    已存在工作表的同步方式：保留全部既有資料，只插入唯一鍵尚未存在的新列。
    不 clear、不 resize 縮小、不整表 update；唯一例外是精準修正已確認錯配的
    權證名稱與標的股／名稱儲存格，其他金額、股數、報酬率與事件欄位完全不動。
    """
    prepared = _prepare_incremental_result_values(
        title,
        new_values,
        data_scope=data_scope,
        extra_scope_values=extra_scope_values,
    )
    if prepared is None:
        print(f"  ✅ 既有版面型工作表保留不重寫：{safe_worksheet_title(title)}")
        return 0

    incoming_headers, incoming_records = prepared
    existing_values = _worksheet_values_with_formulas(ws)
    existing_values, _identity_repaired_rows = repair_existing_worksheet_security_identities(
        ws, title, existing_values
    )
    existing_header_idx = _find_simple_header_row(existing_values)
    if existing_header_idx != 0:
        print(f"  ⚠️ {safe_worksheet_title(title)} 無法確認第一列表頭，為避免破壞既有工作表，本次不插入。")
        return 0

    existing_headers, existing_records = _values_to_records(
        existing_values,
        header_row_idx=0,
        default_scope="全分點",
    )
    existing_headers = [str(h).strip() for h in existing_headers if str(h).strip()]
    if not existing_headers:
        print(f"  ⚠️ {safe_worksheet_title(title)} 既有表頭為空，本次不覆寫。")
        return 0

    # 只在新版本多出欄位時，將欄位加到表頭尾端；既有欄位順序完全不動。
    missing_headers = [h for h in incoming_headers if h not in existing_headers]
    final_headers = list(existing_headers) + missing_headers

    # 增量同步可能比既有工作表多出新欄位；必須先擴充 grid，才能寫入 AB1 等範圍。
    # 這裡只向外擴充，不縮小、不清空，也不覆寫任何既有資料。
    required_rows = max(
        int(getattr(ws, "row_count", 1) or 1),
        len(existing_values),
        1,
    )
    if not ensure_worksheet_grid_capacity(
        ws,
        required_rows=required_rows,
        required_cols=max(len(final_headers), 1),
    ):
        return 0

    if missing_headers:
        start_col = len(existing_headers) + 1
        start_letter = get_column_letter(start_col)
        gsheet_api_call(
            f"補上新欄位 {ws.title}",
            ws.update,
            values=[missing_headers],
            range_name=f"{start_letter}1",
            value_input_option="USER_ENTERED",
        )
        existing_headers = final_headers

    key_cols = _sheet_upsert_key_columns(title, final_headers)
    if not key_cols:
        print(f"  ⚠️ {safe_worksheet_title(title)} 找不到安全唯一鍵，本次不插入，避免重複資料。")
        return 0

    # 舊版沒有資料範圍欄時，以其既有欄位計算 key，不強迫改寫舊資料。
    if "資料範圍" not in final_headers:
        key_cols = [col for col in key_cols if col != "資料範圍"]
    if not key_cols:
        return 0

    existing_keys = set()
    for rec in existing_records:
        if "資料範圍" in final_headers and not str(rec.get("資料範圍", "")).strip():
            rec["資料範圍"] = "全分點"
        key = _record_key(rec, key_cols)
        if any(key):
            existing_keys.add(key)

    rows_to_insert = []
    incoming_seen = set()
    for rec in incoming_records:
        key = _record_key(rec, key_cols)
        if not any(key) or key in existing_keys or key in incoming_seen:
            continue
        incoming_seen.add(key)
        rows_to_insert.append([rec.get(header, "") for header in final_headers])

    if not rows_to_insert:
        print(f"  ✅ Google Sheet 無缺少資料：{safe_worksheet_title(title)}，既有資料完全保留。")
        return 0

    if not _insert_rows_below_header(ws, rows_to_insert):
        return 0

    print(
        f"  ☁️ Google Sheet 增量插入：{safe_worksheet_title(title)}｜"
        f"新增 {len(rows_to_insert):,} 列｜既有 {len(existing_records):,} 列保留"
    )
    return len(rows_to_insert)

def upload_excel_to_google_sheet(
    xlsx_path,
    data_scope=None,
    allowed_titles=None,
    extra_scope_values=None,
):
    """
    Google Sheet 同步規則：
    1. 工作表不存在時才第一次建立並完整寫入。
    2. 工作表已存在時只 insert 唯一鍵尚不存在的新資料。
    3. 絕不刪除、清空、重建或整張覆寫任何既有工作表。
    4. 唯一允許刪除的是 cleanup_deleted_broker_rows_in_existing_worksheets()
       明確判定為已從目前程式分點清單移除的舊分點資料列。
    """
    if not GSHEET_RESULT_ENABLED or not gsheet_enabled():
        print("  ⚠️ 未設定 GCP_SERVICE_KEY，略過 Google Sheet 結果同步")
        return

    try:
        from openpyxl import load_workbook

        current_scope = str(data_scope or get_result_data_scope()).strip()
        allowed = None
        if allowed_titles is not None:
            allowed = {safe_worksheet_title(x) for x in allowed_titles}
        extra_scope_values = extra_scope_values or {}

        wb = load_workbook(xlsx_path, data_only=False)
        primary_sh = get_gsheet_spreadsheet()
        if primary_sh is None:
            return

        # 只清除已刪除分點的舊資料列；不刪任何工作表。
        cleanup_deleted_broker_rows_in_existing_worksheets()
        print(
            f"  ☁️ Google Sheet 目標試算表：{GOOGLE_SHEET_NAME}"
            + (f"｜ID={GOOGLE_SHEET_ID}" if GOOGLE_SHEET_ID else "")
        )
        print("  ⚙️ 同步模式：工作表只建立一次；之後只增量 insert 缺少資料。")

        for ws_xlsx in wb.worksheets:
            title = safe_worksheet_title(ws_xlsx.title)
            if allowed is not None and title not in allowed:
                continue
            if should_skip_result_sheet_in_run_mode(title):
                print(f"  ✅ RUN_MODE=1 略過全分點專用工作表：{title}")
                continue

            raw_values = worksheet_values_for_gsheet(ws_xlsx)
            selected_values = extra_scope_values.get(title)
            gws, created = _existing_result_sheet(primary_sh, title)
            if gws is None:
                continue

            if created:
                prepared = _prepare_incremental_result_values(
                    title,
                    raw_values,
                    data_scope=current_scope,
                    extra_scope_values=selected_values,
                )
                if prepared is not None:
                    headers, records = prepared
                    values = _records_to_values(headers, records)
                else:
                    values = raw_values

                values = normalize_result_values_for_comma_numbers(values)
                if write_values_to_worksheet(gws, values):
                    if should_upsert_result_sheet(title):
                        clear_all_number_formats_for_written_range(gws, values=values)
                        apply_safe_result_table_style_to_gsheet(gws, values=values)
                        apply_text_format_to_gsheet(gws, values)
                        apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=values)
                        apply_date_format_to_gsheet(ws_xlsx, gws, values=values)
                        apply_header_widths_to_gsheet(gws, values=values)
                    else:
                        apply_excel_style_to_gsheet(ws_xlsx, gws)
                        apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=values)
                        apply_date_format_to_gsheet(ws_xlsx, gws, values=values)
                        apply_header_widths_to_gsheet(gws, values=values)
                    print(f"  ☁️ 第一次建立並寫入工作表：{title}")
                continue

            inserted = insert_missing_result_rows_to_worksheet(
                gws,
                title,
                raw_values,
                data_scope=current_scope,
                extra_scope_values=selected_values,
            )
            if inserted > 0:
                # 新列透過 insert_rows 繼承鄰近資料列格式；只補文字／數字／日期格式，絕不重寫資料。
                current_values = _worksheet_values_with_formulas(gws)
                apply_text_format_to_gsheet(gws, current_values)
                apply_comma_number_format_to_gsheet(ws_xlsx, gws, values=current_values)
                apply_date_format_to_gsheet(ws_xlsx, gws, values=current_values)
                apply_header_widths_to_gsheet(gws, values=current_values)

        print(f"  🧮 Google Sheet 同步後配置格數：{spreadsheet_grid_cell_count(primary_sh):,}")

    except Exception as e:
        print(f"  ⚠️ Excel 同步 Google Sheet 失敗：{type(e).__name__}: {e}")




# ══════════════════════════════════════════════════════════════════════
# 快取工具：避免每次重爬舊資料
# ══════════════════════════════════════════════════════════════════════

def cache_enabled():
    return USE_CACHE and not FORCE_FULL_CACHE_REFRESH


def read_cache_csv(path):
    if not cache_enabled():
        return pd.DataFrame()

    parquet_path = f"{path}.parquet"
    if os.path.exists(parquet_path):
        try:
            return pd.read_parquet(parquet_path).fillna("")
        except Exception as exc:
            print(f"  ⚠️ Parquet 快取讀取失敗：{parquet_path}｜{type(exc).__name__}: {exc}")

    # 相容既有 CSV；成功讀取後會在下一次寫入時自動遷移成 Parquet。
    if os.path.exists(path):
        try:
            return pd.read_csv(path, dtype=str, encoding=CACHE_ENCODING).fillna("")
        except Exception as exc:
            print(f"  ⚠️ CSV 快取讀取失敗：{path}｜{type(exc).__name__}: {exc}")

    # 只保留一次性遷移能力；正式架構不再把原始大型快取存進 Google Sheet。
    if GSHEET_CACHE_ENABLED:
        df_from_gsheet = read_cache_from_gsheet(path)
        if df_from_gsheet is not None and not df_from_gsheet.empty:
            try:
                _atomic_write_parquet(df_from_gsheet, parquet_path)
            except Exception:
                pass
            return df_from_gsheet.fillna("")

    return pd.DataFrame()



def write_cache_csv(
    df,
    path,
    sync_gsheet=False,
    supabase_df=None,
    sync_supabase=True,
    force_gsheet_sync=False,
    force_supabase_sync=False,
):
    if not USE_CACHE or df is None:
        return

    os.makedirs(os.path.dirname(path), exist_ok=True)
    parquet_path = f"{path}.parquet"
    try:
        _atomic_write_parquet(df, parquet_path)
        if CACHE_WRITE_CSV_COMPAT:
            df.to_csv(path, index=False, encoding=CACHE_ENCODING)
        elif os.path.exists(path):
            try:
                os.remove(path)
            except Exception:
                pass
    except Exception as exc:
        print(f"  ⚠️ 快取寫入失敗：{parquet_path}｜{type(exc).__name__}: {exc}")
        return

    if sync_supabase and supabase_enabled():
        upload_df = df if supabase_df is None else supabase_df
        write_cache_to_supabase(upload_df, path, full_df=df)

    # FinMind 版不再將原始快取同步到 Google Sheet；Google Sheet 僅保留最終 8 張結果表。
    if sync_gsheet and GSHEET_CACHE_ENABLED:
        write_cache_to_gsheet(df, path)


def load_price_cache():
    """讀取價格持久化快取，並只保留最近 PRICE_RETENTION_TRADING_DAYS 個交易日。"""
    df = read_cache_csv(PRICE_CACHE_PATH)

    if df.empty:
        return {}

    required_cols = ["代號", "日期", "收盤價"]
    for col in required_cols:
        if col not in df.columns:
            print(f"  ⚠️ 價格快取欄位不完整，缺少：{col}")
            return {}

    df, prune_stats = prune_price_cache_dataframe(df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 價格快取記憶體清理：移除 {prune_stats['removed']:,} 筆｜"
            f"保留最近 {PRICE_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    price_cache = {}

    for row in df.itertuples(index=False):
        row = row._asdict()
        code = normalize_price_code(row.get("代號", ""))

        if not code:
            continue

        date_str = normalize_date_str(row.get("日期", ""))
        dt = parse_date(date_str)

        if not dt:
            continue

        price = safe_price_float(row.get("收盤價", ""))

        if price is None:
            continue

        price_cache.setdefault(code, {})[dt.strftime("%Y/%m/%d")] = price

    return price_cache




def save_price_cache(price_cache, changed_codes=None):
    """保存價格快取並裁切為最近指定交易日；不再同步大型原始快取到 Google Sheet。"""
    if not USE_CACHE or not price_cache:
        return

    rows = []
    for code, prices in price_cache.items():
        norm_code = normalize_price_code(code)
        if not norm_code or not prices:
            continue
        for date_str, raw_price in prices.items():
            dt = parse_date(date_str)
            price = safe_price_float(raw_price)
            if dt and price is not None:
                rows.append({
                    "代號": norm_code,
                    "日期": dt.strftime("%Y/%m/%d"),
                    "收盤價": price,
                })

    if not rows:
        return

    df = pd.DataFrame(rows, columns=["代號", "日期", "收盤價"])
    df, prune_stats = prune_price_cache_dataframe(df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 價格快取自動清理：移除 {prune_stats['removed']:,} 筆｜"
            f"保留最近 {PRICE_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    normalized_changed_codes = {
        normalize_price_code(code)
        for code in (changed_codes or [])
        if normalize_price_code(code)
    }
    if normalized_changed_codes:
        supabase_df = df[df["代號"].astype(str).isin(normalized_changed_codes)].copy()
    else:
        supabase_df = df

    write_cache_csv(
        df,
        PRICE_CACHE_PATH,
        sync_gsheet=False,
        supabase_df=supabase_df,
    )
    print(f"  💾 已更新價格快取：{PRICE_CACHE_PATH}，共 {len(df):,} 筆")


def get_cached_prices_for_code(price_cache, code):
    """
    從價格快取中取出指定代號的價格。

    同時支援補零版與去零版查找，最後回傳單一合併後 dict。
    """
    out = {}
    norm_code = normalize_price_code(code)

    if not norm_code:
        return out

    lookup_codes = []
    for c in price_code_variants(norm_code):
        if c and c not in lookup_codes:
            lookup_codes.append(c)

        no_zero = c.lstrip("0")
        if no_zero and no_zero not in lookup_codes:
            lookup_codes.append(no_zero)

        norm_c = normalize_price_code(c)
        if norm_c and norm_c not in lookup_codes:
            lookup_codes.append(norm_c)

    for c in lookup_codes:
        cached = price_cache.get(c)

        if not cached:
            continue

        out = merge_price_dicts(out, cached)

    return out


def add_price_aliases(price_cache, code, prices):
    """
    在記憶體 price_cache 中建立補零 / 去零別名，
    避免 Excel 或 pandas 吃掉前導 0 時查不到。
    """
    if not prices:
        prices = {}

    norm_code = normalize_price_code(code)

    if not norm_code:
        return

    price_cache[norm_code] = prices

    no_zero = norm_code.lstrip("0")

    if no_zero:
        price_cache[no_zero] = prices

    raw_code = str(code).strip()

    if raw_code:
        price_cache[raw_code] = prices





def candidate_key_from_values(warrant_code, broker_code):
    return (str(warrant_code).strip(), str(broker_code).strip())



def save_warrants_cache(warrants):
    if not USE_CACHE or not warrants:
        return

    df = pd.DataFrame(warrants)
    wanted_cols = ["代號", "名稱", "標的股", "標的名稱", "上市日", "最後交易日"]
    for col in wanted_cols:
        if col not in df.columns:
            df[col] = ""

    df = fix_known_underlying_info_dataframe(df, "名稱", "標的股", "標的名稱")
    df["代號"] = df["代號"].astype(str).str.strip().str.upper()
    df["上市日"] = df["上市日"].map(lambda x: normalize_date_str(x) if parse_date(x) else "")
    df["最後交易日"] = df["最後交易日"].map(lambda x: normalize_date_str(x) if parse_date(x) else "")
    df = df[df["代號"] != ""].drop_duplicates(
        subset=["代號", "上市日", "最後交易日"],
        keep="last",
    )
    df = df.sort_values(["代號", "上市日", "最後交易日"]).reset_index(drop=True)

    write_cache_csv(df[wanted_cols], WARRANTS_CACHE_PATH, force_supabase_sync=FORCE_FULL_CACHE_REFRESH)
    print(f"  💾 已更新權證對照快取：{WARRANTS_CACHE_PATH}，保留 {len(df):,} 筆日期區間紀錄")



def load_warrants_cache():
    df = read_cache_csv(WARRANTS_CACHE_PATH)
    if df.empty:
        return []

    required_cols = ["代號", "名稱", "標的股", "標的名稱"]
    if any(col not in df.columns for col in required_cols):
        return []
    for col in ("上市日", "最後交易日"):
        if col not in df.columns:
            df[col] = ""

    warrants = []
    for row in df.itertuples(index=False):
        rec = row._asdict()
        code = str(rec.get("代號", "")).strip().upper()
        name = str(rec.get("名稱", "")).strip() or code
        if not code:
            continue

        underlying_code, underlying_name = correct_underlying_info_by_warrant_name(
            name,
            rec.get("標的股", ""),
            rec.get("標的名稱", ""),
        )
        warrants.append({
            "代號": code,
            "名稱": name,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "上市日": normalize_date_str(rec.get("上市日", "")) if parse_date(rec.get("上市日", "")) else "",
            "最後交易日": normalize_date_str(rec.get("最後交易日", "")) if parse_date(rec.get("最後交易日", "")) else "",
        })
    return warrants



def save_broker_map_cache(broker_map):
    if not USE_CACHE or not broker_map:
        return

    rows = []
    for label, (name, code) in broker_map.items():
        rows.append({
            "分點": label,
            "分點名稱": name,
            "券商代號": code,
        })

    df = pd.DataFrame(rows)
    write_cache_csv(df, BROKER_MAP_CACHE_PATH)
    print(f"  💾 已更新分點代號快取：{BROKER_MAP_CACHE_PATH}")


def load_broker_map_cache():
    df = read_cache_csv(BROKER_MAP_CACHE_PATH)

    if df.empty:
        return {}

    required_cols = ["分點", "分點名稱", "券商代號"]
    for col in required_cols:
        if col not in df.columns:
            return {}

    broker_map = {}

    for row in df.itertuples(index=False):
        row = row._asdict()
        label = str(row["分點"]).strip()
        name = str(row["分點名稱"]).strip()
        code = str(row["券商代號"]).strip()

        if label and name and code and label in TARGET_PATTERNS:
            broker_map[label] = (name, code)

    missing = [k for k in TARGET_PATTERNS if k not in broker_map]
    if missing:
        print(f"  ⚠️ 分點代號快取不完整，缺少：{missing}")
        return {}

    return broker_map






# ══════════════════════════════════════════════════════════════════════
# daily / repair 工作流：prescan 狀態與 refresh keys 快取
# ══════════════════════════════════════════════════════════════════════



def workflow_is_repair():
    return WORKFLOW_MODE == "repair"


def workflow_is_longterm():
    return WORKFLOW_MODE == "longterm"


























def load_history_cache():
    df = read_cache_csv(HISTORY_CACHE_PATH)

    if df.empty:
        return pd.DataFrame()

    required_cols = [
        "權證代號", "權證名稱", "標的股", "標的名稱",
        "分點", "分點名稱", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
        "買超股數", "買超金額",
    ]

    for col in required_cols:
        if col not in df.columns:
            print(f"  ⚠️ 原始分點資料快取欄位不完整，缺少：{col}")
            return pd.DataFrame()

    out_df = df[required_cols].copy()
    out_df = fix_known_underlying_info_dataframe(out_df, "權證名稱", "標的股", "標的名稱")
    out_df["日期"] = out_df["日期"].map(normalize_date_str)
    out_df = out_df.drop_duplicates(
        subset=["權證代號", "券商代號", "日期"],
        keep="last"
    ).reset_index(drop=True)

    out_df, prune_stats = prune_history_cache_dataframe(out_df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 分點歷史記憶體清理：移除 {prune_stats['removed']:,} 筆｜"
            f"超期/異常日期 {prune_stats.get('removed_date', 0):,}｜"
            f"已移除分點 {prune_stats.get('removed_broker', 0):,}｜"
            f"保留最近 {HISTORY_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    return out_df
















def save_history_cache(history_df, fetched_items=None, previous_history_empty=False):
    """保存 FinMind 標準化歷史快取；大型原始資料不再同步 Google Sheet。"""
    if not USE_CACHE or history_df is None:
        return history_df

    history_df, prune_stats = prune_history_cache_dataframe(history_df)
    if prune_stats.get("removed", 0) > 0:
        print(
            f"  🧹 分點歷史自動清理：移除 {prune_stats['removed']:,} 筆｜"
            f"超期/異常日期 {prune_stats.get('removed_date', 0):,}｜"
            f"已移除分點 {prune_stats.get('removed_broker', 0):,}｜"
            f"保留最近 {HISTORY_RETENTION_TRADING_DAYS} 個交易日｜起始日 {prune_stats.get('cutoff') or '-'}"
        )

    # FinMind 日檔與標準化歷史皆以本機 Parquet 為主；Google Sheet 只保留最終 8 張結果。
    write_cache_csv(
        history_df,
        HISTORY_CACHE_PATH,
        sync_gsheet=False,
        force_gsheet_sync=False,
    )
    print(f"  💾 已更新 FinMind 分點歷史快取：{HISTORY_CACHE_PATH}，共 {len(history_df):,} 筆")
    return history_df


def items_from_history_cache(history_df, candidate_filter=None):
    items = []

    if history_df is None or history_df.empty:
        return items

    df = history_df.copy().fillna("")

    if candidate_filter:
        mask = pd.Series(
            [candidate_key_from_values(w, b) in candidate_filter for w, b in zip(df["權證代號"], df["券商代號"])],
            index=df.index
        )
        df = df[mask].copy()

    if df.empty:
        return items

    numeric_cols = ["買進股數", "賣出股數", "買進金額", "賣出金額", "買超股數", "買超金額"]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    group_cols = ["權證代號", "權證名稱", "標的股", "標的名稱", "分點", "分點名稱", "券商代號"]

    for key, g in df.groupby(group_cols, dropna=False):
        warrant_code, warrant_name, underlying_code, underlying_name, broker_label, broker_name, broker_code = key

        item_df = g[[
            "日期", "買進股數", "賣出股數",
            "買進金額", "賣出金額", "買超股數", "買超金額"
        ]].copy()

        item_df["日期"] = item_df["日期"].map(normalize_date_str)
        item_df = item_df.sort_values("日期").reset_index(drop=True)

        item = {
            "warrant_code": str(warrant_code).strip(),
            "warrant_name": str(warrant_name).strip(),
            "underlying_code": str(underlying_code).strip(),
            "underlying_name": str(underlying_name).strip(),
            "broker_label": str(broker_label).strip(),
            "broker_name": str(broker_name).strip(),
            "broker_code": str(broker_code).strip(),
            "df": item_df,
        }

        items.append(item)

    return items



# ══════════════════════════════════════════════════════════════════════
# Step 1：取所有認購權證 + 標的股代號
# ══════════════════════════════════════════════════════════════════════



def normalize_stock_name_text(s):
    return str(s).strip().replace(" ", "").replace("　", "")


# 特殊標的優先對照：
# 有些 ETF 權證名稱會使用市場慣用簡稱，例如「台灣50台新5B購17」。
# 若直接丟進一般股票名稱 alias，比對時可能被「台灣大哥大」的短 alias「台灣」吃到，
# 導致 0050 台灣50 被誤判成 3045 台灣大哥大。
# 因此這裡先處理已知 ETF / 特殊標的，且一定要放在一般 resolver 比對前面。
SPECIAL_UNDERLYING_PREFIX_RULES = [
    {
        "prefixes": [
            "台灣50",
            "臺灣50",
            "元大台灣50",
            "元大臺灣50",
        ],
        "exclude_prefixes": [
            "台灣50正",
            "臺灣50正",
            "元大台灣50正",
            "元大臺灣50正",
            "台灣50反",
            "臺灣50反",
            "元大台灣50反",
            "元大臺灣50反",
        ],
        "code": "0050",
        "name": "元大台灣50",
    },
]


def get_special_underlying_info_from_warrant_name(warrant_name):
    wname = normalize_stock_name_text(warrant_name)

    if not wname:
        return "", ""

    for rule in SPECIAL_UNDERLYING_PREFIX_RULES:
        prefixes = [normalize_stock_name_text(x) for x in rule.get("prefixes", []) if str(x).strip()]
        exclude_prefixes = [normalize_stock_name_text(x) for x in rule.get("exclude_prefixes", []) if str(x).strip()]

        if exclude_prefixes and any(wname.startswith(prefix) for prefix in exclude_prefixes):
            continue

        if prefixes and any(wname.startswith(prefix) for prefix in prefixes):
            return str(rule.get("code", "")).strip(), str(rule.get("name", "")).strip()

    return "", ""


def correct_underlying_info_by_warrant_name(warrant_name, underlying_code="", underlying_name=""):
    special_code, special_name = get_special_underlying_info_from_warrant_name(warrant_name)

    if special_code:
        return special_code, special_name

    return str(underlying_code or "").strip(), str(underlying_name or "").strip()


def fix_known_underlying_info_dataframe(df, warrant_name_col, underlying_code_col, underlying_name_col):
    """
    針對已知 ETF / 特殊標的修正既有快取中的標的股代號。

    這裡只處理明確命中的特殊規則，不改一般股票 / 權證辨識邏輯。
    目的：避免舊快取或 Supabase 讀回的「台灣50 → 3045」繼續污染後續候選與歷史資料。
    """
    if df is None or df.empty:
        return df

    needed = {warrant_name_col, underlying_code_col, underlying_name_col}
    if not needed.issubset(set(df.columns)):
        return df

    out = df.copy()
    names = out[warrant_name_col].astype(str).map(normalize_stock_name_text)

    for rule in SPECIAL_UNDERLYING_PREFIX_RULES:
        prefixes = [normalize_stock_name_text(x) for x in rule.get("prefixes", []) if str(x).strip()]
        exclude_prefixes = [normalize_stock_name_text(x) for x in rule.get("exclude_prefixes", []) if str(x).strip()]

        if not prefixes:
            continue

        mask = False
        for prefix in prefixes:
            mask = mask | names.str.startswith(prefix, na=False)

        for prefix in exclude_prefixes:
            mask = mask & ~names.str.startswith(prefix, na=False)

        if mask.any():
            out.loc[mask, underlying_code_col] = str(rule.get("code", "")).strip()
            out.loc[mask, underlying_name_col] = str(rule.get("name", "")).strip()

    return out









def normalize_security_code_text(value):
    s = str(value or "").strip().upper()
    if s.endswith(".0") and s[:-2].isalnum():
        s = s[:-2]
    return re.sub(r"\s+", "", s)


def _looks_like_warrant_security_name(name):
    name = normalize_stock_name_text(name)
    return bool(name and ("購" in name or "售" in name) and len(name) >= 4)


def build_canonical_stock_master(stock_df):
    """
    建立唯一且穩定的「股號→股名」主檔。

    TaiwanStockInfo 可能保留同代號的多筆歷史紀錄；優先選 date 最新的一列，
    同日仍有重複時才用名稱出現次數與原始順序決定，避免 dict(zip) 任意取最後列。
    """
    code_to_name = {}
    name_to_code = {}
    if stock_df is None or stock_df.empty:
        return code_to_name, name_to_code
    if not {"stock_id", "stock_name"}.issubset(stock_df.columns):
        return code_to_name, name_to_code

    columns = ["stock_id", "stock_name"]
    if "date" in stock_df.columns:
        columns.append("date")
    work = stock_df[columns].copy().fillna("")
    if "date" not in work.columns:
        work["date"] = ""

    candidates = defaultdict(list)
    for row_no, (raw_code, raw_name, raw_date) in enumerate(
        work[["stock_id", "stock_name", "date"]].itertuples(index=False, name=None)
    ):
        code = normalize_security_code_text(raw_code)
        name = str(raw_name or "").strip()
        if not code or not name:
            continue
        if not re.fullmatch(r"[0-9A-Z]{4,12}", code):
            continue
        if _looks_like_warrant_security_name(name):
            continue
        parsed_date = parse_date(raw_date) or datetime.min
        candidates[code].append((parsed_date, row_no, name))

    for code, rows in candidates.items():
        latest_date = max(row[0] for row in rows)
        latest_rows = [row for row in rows if row[0] == latest_date]
        counts = Counter(row[2] for row in latest_rows)
        chosen = max(
            counts,
            key=lambda name: (
                counts[name],
                max(row[1] for row in latest_rows if row[2] == name),
                len(normalize_stock_name_text(name)),
            ),
        )
        code_to_name[code] = chosen

    name_codes = defaultdict(set)
    for code, name in code_to_name.items():
        norm_name = normalize_stock_name_text(name)
        if norm_name:
            name_codes[norm_name].add(code)
    for norm_name, codes in name_codes.items():
        if len(codes) == 1:
            name_to_code[norm_name] = next(iter(codes))

    return code_to_name, name_to_code



def make_stock_aliases(stock_name, exact_stock_names=None):
    """建立安全的股名別名，避免短別名撞到另一檔正式股名。"""
    name = normalize_stock_name_text(stock_name)
    aliases = set()
    if not name:
        return aliases

    exact_stock_names = exact_stock_names or set()
    aliases.add(name)
    suffixes = [
        "半導體", "科技", "電子", "光電", "精密", "材料", "生技", "醫療",
        "資訊", "電腦", "通信", "通訊", "電機", "機械", "工業", "實業",
        "企業", "國際", "控股", "投控", "控", "建設", "營造", "食品",
        "鋼鐵", "化學", "化工", "紡織", "玻璃", "塑膠", "水泥",
    ]

    stripped = name
    changed = True
    while changed:
        changed = False
        for suffix in suffixes:
            if stripped.endswith(suffix) and len(stripped) > len(suffix) + 1:
                candidate = stripped[:-len(suffix)]
                stripped = candidate
                if candidate not in exact_stock_names or candidate == name:
                    aliases.add(candidate)
                changed = True
                break

    for length in range(min(4, len(name)), 1, -1):
        candidate = name[:length]
        if candidate in exact_stock_names and candidate != name:
            continue
        aliases.add(candidate)

    dangerous_aliases = {"台灣", "臺灣"}
    return {alias for alias in aliases if len(alias) >= 2 and alias not in dangerous_aliases}


def build_underlying_resolver_from_stock_master(code_to_name):
    exact_names = {
        normalize_stock_name_text(name)
        for name in (code_to_name or {}).values()
        if normalize_stock_name_text(name)
    }
    candidates = []
    seen = set()

    def add_candidate(prefix, code, name, exact):
        prefix_norm = normalize_stock_name_text(prefix)
        name_norm = normalize_stock_name_text(name)
        key = (prefix_norm, code, name_norm)
        if not prefix_norm or key in seen:
            return
        seen.add(key)
        candidates.append({
            "prefix": prefix_norm,
            "prefix_len": len(prefix_norm),
            "is_exact_alias": bool(exact),
            "stock_code": code,
            "stock_name": name,
            "stock_name_len": len(name_norm),
        })

    for code, name in (code_to_name or {}).items():
        name_norm = normalize_stock_name_text(name)
        add_candidate(name_norm, code, name, True)
        for alias in make_stock_aliases(name, exact_names):
            alias_norm = normalize_stock_name_text(alias)
            add_candidate(alias_norm, code, name, alias_norm == name_norm)

    candidates.sort(
        key=lambda rec: (
            rec["prefix_len"],
            1 if rec["is_exact_alias"] else 0,
            -rec["stock_name_len"],
        ),
        reverse=True,
    )
    return candidates


def _reset_underlying_resolver_runtime_cache(resolver=None):
    """切換股名 resolver 時重建第一字索引並清空名稱解析快取。"""
    global _UNDERLYING_RESOLVER_RUNTIME_REF
    global _UNDERLYING_RESOLVER_FIRST_CHAR_INDEX
    global _UNDERLYING_RESOLUTION_RESULT_CACHE

    resolver = resolver if resolver is not None else _CURRENT_UNDERLYING_RESOLVER
    first_char_index = defaultdict(list)
    for rec in resolver or []:
        prefix = str(rec.get("prefix", "") or "")
        if prefix:
            first_char_index[prefix[0]].append(rec)

    with _UNDERLYING_RESOLVER_RUNTIME_LOCK:
        _UNDERLYING_RESOLVER_RUNTIME_REF = resolver
        _UNDERLYING_RESOLVER_FIRST_CHAR_INDEX = dict(first_char_index)
        _UNDERLYING_RESOLUTION_RESULT_CACHE = {}


def _ensure_underlying_resolver_runtime_cache(resolver):
    """確保目前 resolver 的索引存在；以物件身分判斷，不改 resolver 內容。"""
    if resolver is _UNDERLYING_RESOLVER_RUNTIME_REF:
        return
    _reset_underlying_resolver_runtime_cache(resolver)


def resolve_underlying_from_warrant_name(warrant_name, resolver=None):
    wname = normalize_stock_name_text(warrant_name)
    if not wname:
        return None

    special_code, special_name = get_special_underlying_info_from_warrant_name(warrant_name)
    if special_code:
        return {
            "stock_code": special_code,
            "stock_name": special_name,
            "prefix": normalize_stock_name_text(special_name),
            "prefix_len": len(normalize_stock_name_text(special_name)),
            "is_exact_alias": True,
            "is_special": True,
        }

    resolver = resolver if resolver is not None else _CURRENT_UNDERLYING_RESOLVER
    _ensure_underlying_resolver_runtime_cache(resolver)

    with _UNDERLYING_RESOLVER_RUNTIME_LOCK:
        if wname in _UNDERLYING_RESOLUTION_RESULT_CACHE:
            return _UNDERLYING_RESOLUTION_RESULT_CACHE[wname]
        candidates = _UNDERLYING_RESOLVER_FIRST_CHAR_INDEX.get(wname[0], ())

    result = None
    # candidates 仍維持 build_underlying_resolver_from_stock_master() 的原排序，
    # 因此命中結果與原本逐筆掃完整 resolver 完全相同。
    for rec in candidates:
        prefix = rec.get("prefix", "")
        if prefix and wname.startswith(prefix):
            result = rec
            break

    with _UNDERLYING_RESOLVER_RUNTIME_LOCK:
        if len(_UNDERLYING_RESOLUTION_RESULT_CACHE) >= _UNDERLYING_RESOLUTION_RESULT_CACHE_MAXSIZE:
            _UNDERLYING_RESOLUTION_RESULT_CACHE.clear()
        _UNDERLYING_RESOLUTION_RESULT_CACHE[wname] = result

    return result


def reconcile_underlying_identity(
    warrant_name,
    underlying_code="",
    underlying_name="",
    code_to_name=None,
    resolver=None,
):
    """
    交叉驗證標的身分，並保證回傳的股號與股名來自同一筆主檔。

    Summary target_stock_id 仍是主要來源；只有權證名稱前綴屬於高信心匹配、
    且 Summary 所指標的名稱明顯不符合權證名稱時，才改用名稱解析結果。
    """
    code_to_name = code_to_name if code_to_name is not None else _CURRENT_STOCK_CODE_TO_NAME
    resolver = resolver if resolver is not None else _CURRENT_UNDERLYING_RESOLVER
    raw_code = normalize_security_code_text(underlying_code)
    raw_name = str(underlying_name or "").strip()
    master_name = str((code_to_name or {}).get(raw_code, "")).strip()
    wname_norm = normalize_stock_name_text(warrant_name)
    resolved = resolve_underlying_from_warrant_name(warrant_name, resolver)

    chosen_code = raw_code
    source = "summary"
    if WARRANT_IDENTITY_RECONCILE_ENABLED and resolved:
        resolved_code = normalize_security_code_text(resolved.get("stock_code", ""))
        resolved_name = str(resolved.get("stock_name", "")).strip()
        prefix_len = int(resolved.get("prefix_len", 0) or 0)
        high_confidence = bool(
            resolved.get("is_special")
            or resolved.get("is_exact_alias")
            or prefix_len >= WARRANT_IDENTITY_OVERRIDE_MIN_PREFIX
        )

        if not chosen_code and resolved_code:
            chosen_code = resolved_code
            source = "warrant_name"
        elif resolved_code and resolved_code != chosen_code and high_confidence:
            summary_name_norm = normalize_stock_name_text(master_name or raw_name)
            summary_name_matches = bool(summary_name_norm and wname_norm.startswith(summary_name_norm))
            if not summary_name_matches:
                chosen_code = resolved_code
                source = "warrant_name_override"
        elif resolved_code == chosen_code:
            source = "summary_verified"

    chosen_name = str((code_to_name or {}).get(chosen_code, "")).strip()
    if not chosen_name and resolved and normalize_security_code_text(resolved.get("stock_code", "")) == chosen_code:
        chosen_name = str(resolved.get("stock_name", "")).strip()
    if not chosen_name and chosen_code == raw_code:
        chosen_name = raw_name

    return chosen_code, chosen_name, source


def build_cached_warrant_index(cached_warrants):
    """將舊權證快取依代號預先分組；只改查找方式，不改候選順序。"""
    by_code = defaultdict(list)
    for rec in cached_warrants or []:
        code = normalize_security_code_text(rec.get("代號", ""))
        if code:
            by_code[code].append(rec)
    return dict(by_code)


def _cached_warrant_name_candidates(cached_warrants, code, list_date, end_date):
    out = []
    code = normalize_security_code_text(code)
    list_dt = parse_date(list_date)
    end_dt = parse_date(end_date)

    # get_all_call_warrants_live() 會傳入按代號建立的 dict 索引；
    # 仍相容舊呼叫傳入 list，避免影響其他既有流程。
    if isinstance(cached_warrants, dict):
        candidate_records = cached_warrants.get(code, ())
    else:
        candidate_records = (
            rec for rec in (cached_warrants or [])
            if normalize_security_code_text(rec.get("代號", "")) == code
        )

    for rec in candidate_records:
        name = str(rec.get("名稱", "")).strip()
        if not name:
            continue
        rec_start = parse_date(rec.get("上市日", ""))
        rec_end = parse_date(rec.get("最後交易日", ""))
        exact = bool(rec_start and rec_end and list_dt and end_dt and rec_start.date() == list_dt.date() and rec_end.date() == end_dt.date())
        overlap = False
        if rec_start and rec_end and list_dt and end_dt:
            overlap = rec_start.date() <= end_dt.date() and rec_end.date() >= list_dt.date()
        priority = 500 if exact else 350 if overlap else 120
        out.append((name, priority, "cache_exact" if exact else "cache_overlap" if overlap else "cache_code"))
    return out


def select_warrant_name_for_interval(
    code,
    underlying_code,
    list_date,
    end_date,
    info_name_candidates,
    cached_warrants,
    resolver=None,
    enforce_underlying_match=True,
):
    """日期化挑選權證名稱，避免同一權證代號重用時拿到另一期間的名稱。"""
    code = normalize_security_code_text(code)
    underlying_code = normalize_security_code_text(underlying_code)
    candidates = []
    for name in info_name_candidates.get(code, []):
        candidates.append((str(name).strip(), 300, "finmind_info"))
    candidates.extend(_cached_warrant_name_candidates(cached_warrants, code, list_date, end_date))

    dedup = {}
    for name, source_priority, source in candidates:
        if not name:
            continue
        key = normalize_stock_name_text(name)
        previous = dedup.get(key)
        if previous is None or source_priority > previous[1]:
            dedup[key] = (name, source_priority, source)

    scored = []
    for name, source_priority, source in dedup.values():
        score = source_priority
        if name != code:
            score += 20
        if "購" in name:
            score += 20
        resolved = resolve_underlying_from_warrant_name(name, resolver)
        if resolved:
            resolved_code = normalize_security_code_text(resolved.get("stock_code", ""))
            high_confidence = bool(
                resolved.get("is_special")
                or resolved.get("is_exact_alias")
                or int(resolved.get("prefix_len", 0) or 0) >= WARRANT_IDENTITY_OVERRIDE_MIN_PREFIX
            )
            if resolved_code == underlying_code:
                score += 1200
            elif enforce_underlying_match and high_confidence and resolved_code:
                score -= 1200
        scored.append((score, name, source, resolved))

    if not scored:
        return code, "code_only"
    scored.sort(key=lambda item: (item[0], len(item[1])), reverse=True)
    best_score, best_name, best_source, best_resolved = scored[0]
    if best_resolved:
        resolved_code = normalize_security_code_text(best_resolved.get("stock_code", ""))
        high_confidence = bool(
            best_resolved.get("is_special")
            or best_resolved.get("is_exact_alias")
            or int(best_resolved.get("prefix_len", 0) or 0) >= WARRANT_IDENTITY_OVERRIDE_MIN_PREFIX
        )
        if enforce_underlying_match and high_confidence and resolved_code and resolved_code != underlying_code:
            return code, "code_only_mismatch_guard"
    return best_name, best_source


def _warrant_record_identity_score(record, resolver=None):
    score = 0
    code = normalize_security_code_text(record.get("標的股", ""))
    name = str(record.get("名稱", "")).strip()
    if code and _CURRENT_STOCK_CODE_TO_NAME.get(code):
        score += 100
    if name and name != record.get("代號", ""):
        score += 20
    resolved = resolve_underlying_from_warrant_name(name, resolver)
    if resolved:
        if normalize_security_code_text(resolved.get("stock_code", "")) == code:
            score += 1000
        else:
            score -= 1000
    return score


def dedupe_warrant_interval_records(records, resolver=None):
    """同代號、同上市／到期區間若有衝突，保留身分交叉驗證分數最高的一筆。"""
    best = {}
    order = []
    for rec in records or []:
        key = (
            normalize_security_code_text(rec.get("代號", "")),
            normalize_date_str(rec.get("上市日", "")),
            normalize_date_str(rec.get("最後交易日", "")),
        )
        score = _warrant_record_identity_score(rec, resolver)
        if key not in best:
            order.append(key)
            best[key] = (score, rec)
        elif score >= best[key][0]:
            best[key] = (score, rec)
    return [best[key][1] for key in order]


def repair_history_metadata_from_warrants(history_df, warrants):
    """
    用「權證代號＋交易日期」批次回填權證名稱與標的股／名稱。

    依日期分組後以 Series.map + .loc 批次比較與回寫，避免逐列 .at。
    """
    stats = {"rows": 0, "warrant_name": 0, "underlying_code": 0, "underlying_name": 0}
    if not WARRANT_IDENTITY_REPAIR_HISTORY or history_df is None or history_df.empty or not warrants:
        return history_df, stats
    required = {"權證代號", "權證名稱", "標的股", "標的名稱", "日期"}
    if not required.issubset(history_df.columns):
        return history_df, stats

    out = history_df.copy()
    out["日期"] = out["日期"].map(normalize_date_str)
    changed_row_mask = pd.Series(False, index=out.index)

    field_specs = (
        ("權證名稱", "名稱", "warrant_name", lambda value: str(value or "").strip()),
        ("標的股", "標的股", "underlying_code", normalize_security_code_text),
        ("標的名稱", "標的名稱", "underlying_name", lambda value: str(value or "").strip()),
    )

    for date_value, indices in out.groupby("日期", sort=False).groups.items():
        lookup = _warrant_lookup(warrants, date_value)
        if not lookup:
            continue

        group_index = pd.Index(indices)
        codes = out.loc[group_index, "權證代號"].map(normalize_security_code_text)
        metas = codes.map(lookup)
        has_meta = metas.map(lambda meta: isinstance(meta, dict))
        if not bool(has_meta.any()):
            continue

        for column, meta_key, stat_key, normalizer in field_specs:
            new_values = metas.map(
                lambda meta, _key=meta_key, _normalizer=normalizer:
                    _normalizer(meta.get(_key, "")) if isinstance(meta, dict) else ""
            )
            old_values = out.loc[group_index, column].fillna("").astype(str).str.strip()
            change_mask = has_meta & new_values.ne("") & old_values.ne(new_values)
            if not bool(change_mask.any()):
                continue

            changed_indices = group_index[change_mask.to_numpy()]
            out.loc[changed_indices, column] = new_values.loc[changed_indices].to_numpy()
            changed_row_mask.loc[changed_indices] = True
            stats[stat_key] += int(change_mask.sum())

    stats["rows"] = int(changed_row_mask.sum())
    return out, stats



def _finmind_headers():
    if not FINMIND_TOKEN:
        raise RuntimeError("缺少 FINMIND_API_0714，請在 GitHub Secrets 或環境變數中設定 FinMind Token。")
    return {
        "Authorization": f"Bearer {FINMIND_TOKEN}",
        "User-Agent": "warrant-backtest-finmind/1.0",
        "Accept": "application/json, application/octet-stream, */*",
    }


def _finmind_request(method, url, *, params=None, stream=False, description="FinMind API"):
    session = get_thread_session()
    last_error = None

    for attempt in range(1, FINMIND_MAX_RETRIES + 1):
        try:
            response = session.request(
                method,
                url,
                params=params,
                headers=_finmind_headers(),
                timeout=(FINMIND_REQUEST_TIMEOUT_CONNECT, FINMIND_REQUEST_TIMEOUT_READ),
                stream=stream,
            )

            if response.status_code == 402:
                raise RuntimeError("FinMind API 使用額度已達上限（HTTP 402）")
            if response.status_code in (401, 403):
                raise RuntimeError(f"FinMind Token 無效或權限不足（HTTP {response.status_code}）")
            if response.status_code in (400, 404):
                return response

            response.raise_for_status()
            return response
        except Exception as exc:
            last_error = exc
            if attempt >= FINMIND_MAX_RETRIES:
                break
            wait_seconds = min(20.0, FINMIND_RETRY_BASE_SECONDS * (2 ** (attempt - 1)))
            print(
                f"  ⚠️ {description} 第 {attempt}/{FINMIND_MAX_RETRIES} 次失敗："
                f"{type(exc).__name__}: {exc}｜{wait_seconds:.1f} 秒後重試"
            )
            time.sleep(wait_seconds)

    raise RuntimeError(f"{description} 失敗：{type(last_error).__name__}: {last_error}")


def _atomic_write_parquet(df, path):
    path = str(path)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = f"{path}.tmp.{os.getpid()}.{threading.get_ident()}"
    try:
        df.to_parquet(tmp_path, index=False)
        os.replace(tmp_path, path)
    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass


def _read_parquet_safe(path):
    try:
        if os.path.exists(path) and os.path.getsize(path) > 0:
            return pd.read_parquet(path)
    except Exception as exc:
        print(f"  ⚠️ Parquet 快取讀取失敗，將重新下載：{path}｜{type(exc).__name__}: {exc}")
    return pd.DataFrame()


def _metadata_cache_path(dataset):
    return os.path.join(FINMIND_METADATA_CACHE_DIR, f"{dataset}.parquet")


def fetch_finmind_dataset(dataset, *, params=None, cache_hours=None, force=False):
    cache_hours = FINMIND_METADATA_CACHE_HOURS if cache_hours is None else max(float(cache_hours), 0.0)
    cache_path = _metadata_cache_path(dataset)

    if not force and os.path.exists(cache_path):
        age_hours = max((time.time() - os.path.getmtime(cache_path)) / 3600.0, 0.0)
        if age_hours <= cache_hours:
            cached = _read_parquet_safe(cache_path)
            if not cached.empty:
                return cached

    query = {"dataset": dataset}
    if params:
        query.update({k: v for k, v in params.items() if v not in (None, "")})

    try:
        response = _finmind_request(
            "GET",
            FINMIND_DATA_URL,
            params=query,
            description=f"FinMind {dataset}",
        )
        payload = response.json()
        status = int(payload.get("status", response.status_code) or response.status_code)
        if status != 200:
            raise RuntimeError(payload.get("msg") or f"status={status}")
        df = pd.DataFrame(payload.get("data") or [])
        if df.empty:
            raise RuntimeError("回傳空資料")
        _atomic_write_parquet(df, cache_path)
        return df
    except Exception as exc:
        stale = _read_parquet_safe(cache_path)
        if not stale.empty:
            print(
                f"  ⚠️ FinMind {dataset} 更新失敗，沿用既有中繼資料快取："
                f"{type(exc).__name__}: {exc}"
            )
            return stale
        raise


def get_finmind_trading_dates():
    try:
        df = fetch_finmind_dataset("TaiwanStockTradingDate", cache_hours=6)
        if "date" not in df.columns:
            return []
        dates = pd.to_datetime(df["date"], errors="coerce").dropna().dt.date.tolist()
        return sorted(set(dates))
    except Exception as exc:
        print(f"  ⚠️ FinMind 交易日清單取得失敗：{type(exc).__name__}: {exc}")
        return []


def latest_finmind_trading_date_on_or_before(target=None):
    target_dt = parse_date(target) if target else datetime.today()
    target_day = target_dt.date() if target_dt else datetime.today().date()
    dates = get_finmind_trading_dates()
    available = [d for d in dates if d <= target_day]
    if available:
        return max(available).strftime("%Y/%m/%d")
    return resolve_latest_trading_date_on_or_before(target_dt or datetime.today())


def resolve_finmind_refresh_target(requested_target_date, preloaded_raw=None, preloaded_status=None):
    """
    依工作流模式決定實際刷新基準日。

    daily：
    - 只接受 requested_target_date。
    - 最新交易日尚未發布時，交由主流程快速結束。

    repair：
    - requested_target_date 已發布時直接使用。
    - 若尚未發布／請求失敗，往前探測最近已發布的實際交易日。
    - 找到後仍會由 refresh_history_from_finmind() 依完整修補天數補抓歷史，
      不會因「今天沒有資料」或「今天已經處理」而跳過修補。

    回傳：
        (effective_date, raw_df, status, used_fallback)
    """
    requested_key = normalize_date_str(requested_target_date)

    if preloaded_status is None:
        preloaded_raw, preloaded_status = download_finmind_warrant_day(
            requested_key,
            force_refresh=FINMIND_FORCE_REFRESH_TARGET_DATE,
        )

    if preloaded_status == "ok":
        return requested_key, preloaded_raw, "ok", False

    if not workflow_is_repair() or not FINMIND_REPAIR_ALLOW_TARGET_FALLBACK:
        return requested_key, preloaded_raw, preloaded_status, False

    requested_dt = parse_date(requested_key) or datetime.today()
    trading_dates = get_finmind_trading_dates()
    candidates = sorted(
        {d for d in trading_dates if d < requested_dt.date()},
        reverse=True,
    )[:FINMIND_REPAIR_TARGET_FALLBACK_TRADING_DAYS]

    # 交易日清單暫時無法取得時，以曆日回退作最後備援；
    # download_finmind_warrant_day() 仍會驗證是否為有效 Parquet。
    if not candidates:
        candidates = [
            (requested_dt - timedelta(days=offset)).date()
            for offset in range(1, FINMIND_REPAIR_TARGET_FALLBACK_TRADING_DAYS + 8)
        ]

    print(
        f"  🔧 完整修補模式：{requested_key} 尚未取得（status={preloaded_status}），"
        f"往前探測最近已發布交易日，最多 {FINMIND_REPAIR_TARGET_FALLBACK_TRADING_DAYS} 個交易日。"
    )

    seen = set()
    for candidate_day in candidates:
        candidate_key = candidate_day.strftime("%Y/%m/%d")
        if candidate_key in seen:
            continue
        seen.add(candidate_key)

        raw_df, status = download_finmind_warrant_day(
            candidate_key,
            force_refresh=False,
        )
        if status == "ok":
            print(
                f"  ✅ 完整修補模式改用最近已發布交易日："
                f"{requested_key} → {candidate_key}｜原始列數 {len(raw_df):,}"
            )
            return candidate_key, raw_df, "ok", True

        print(f"  ↪️ 修補基準日探測：{candidate_key}｜status={status}")

    return requested_key, preloaded_raw, preloaded_status, False


def _finmind_day_cache_path(date_value):
    dt = parse_date(date_value)
    if not dt:
        raise ValueError(f"無效日期：{date_value}")
    return os.path.join(FINMIND_DAILY_CACHE_DIR, f"{dt.strftime('%Y-%m-%d')}.parquet")


def _finmind_day_lock(date_value):
    key = normalize_date_str(date_value)
    with _FINMIND_DAY_LOCKS_GUARD:
        lock = _FINMIND_DAY_LOCKS.get(key)
        if lock is None:
            lock = threading.Lock()
            _FINMIND_DAY_LOCKS[key] = lock
        return lock


def _response_is_parquet(response):
    content = response.content or b""
    ctype = str(response.headers.get("Content-Type", "")).lower()
    return (
        "parquet" in ctype
        or (len(content) >= 8 and content[:4] == b"PAR1" and content[-4:] == b"PAR1")
    )


def download_finmind_warrant_day(date_value, *, force_refresh=False):
    """
    下載指定交易日的全市場權證分點 Parquet。

    回傳 (DataFrame, status)：
    - status="ok"：完整取得；即使目標分點為 0 筆，仍屬完整市場資料。
    - status="missing"：該日資料尚未發布或不是有效資料日。
    - status="error"：請求失敗；不得據此清空 Google Sheet。
    """
    date_dt = parse_date(date_value)
    if not date_dt:
        return pd.DataFrame(), "error"
    date_iso = date_dt.strftime("%Y-%m-%d")
    cache_path = _finmind_day_cache_path(date_iso)

    with _finmind_day_lock(date_iso):
        if not force_refresh:
            cached = _read_parquet_safe(cache_path)
            if not cached.empty:
                return cached, "ok"

        try:
            response = _finmind_request(
                "GET",
                FINMIND_STORAGE_OBJECTS_URL,
                params={
                    "dataset": "TaiwanStockWarrantTradingDailyReport",
                    "date": date_iso,
                },
                description=f"FinMind 全市場權證分點 {date_iso}",
            )

            if response.status_code in (400, 404):
                return pd.DataFrame(), "missing"

            if not _response_is_parquet(response):
                try:
                    payload = response.json()
                except Exception:
                    payload = {}
                status = int(payload.get("status", response.status_code) or response.status_code)
                msg = str(payload.get("msg", "")).strip()
                if status in (400, 404) or any(x in msg.lower() for x in ("not found", "no data", "尚未", "不存在")):
                    return pd.DataFrame(), "missing"
                raise RuntimeError(msg or "回傳內容不是 Parquet")

            df = pd.read_parquet(BytesIO(response.content))
            required = {"securities_trader", "price", "buy", "sell", "securities_trader_id", "stock_id", "date"}
            if not required.issubset(set(df.columns)):
                missing_cols = sorted(required - set(df.columns))
                raise RuntimeError(f"欄位不完整：{missing_cols}")

            _atomic_write_parquet(df, cache_path)
            return df, "ok"
        except Exception as exc:
            cached = _read_parquet_safe(cache_path)
            if not cached.empty:
                print(
                    f"  ⚠️ {date_iso} 即時下載失敗，沿用該日既有 Parquet："
                    f"{type(exc).__name__}: {exc}"
                )
                return cached, "ok"
            print(f"  ⚠️ FinMind 全市場權證分點取得失敗：{date_iso}｜{type(exc).__name__}: {exc}")
            return pd.DataFrame(), "error"



def _reset_warrant_lookup_cache(warrants=None):
    """切換權證區間清單時清空日期化 lookup；不修改權證資料。"""
    global _WARRANT_LOOKUP_CACHE_WARRANTS_REF
    global _WARRANT_LOOKUP_CACHE
    with _WARRANT_LOOKUP_CACHE_LOCK:
        _WARRANT_LOOKUP_CACHE_WARRANTS_REF = warrants
        _WARRANT_LOOKUP_CACHE = {}


def _warrant_lookup(warrants, date_value=None):
    """依交易日期選出當時有效的權證代號→標的對照，正確處理代號重用。"""
    global _WARRANT_LOOKUP_CACHE_WARRANTS_REF
    global _WARRANT_LOOKUP_CACHE

    target_dt = parse_date(date_value) if date_value else None
    cache_key = target_dt.strftime("%Y/%m/%d") if target_dt else ""

    with _WARRANT_LOOKUP_CACHE_LOCK:
        if warrants is not _WARRANT_LOOKUP_CACHE_WARRANTS_REF:
            _WARRANT_LOOKUP_CACHE_WARRANTS_REF = warrants
            _WARRANT_LOOKUP_CACHE = {}
        cached = _WARRANT_LOOKUP_CACHE.get(cache_key)
        if cached is not None:
            return cached

    selected = {}
    selected_start = {}

    for warrant in warrants or []:
        code = str(warrant.get("代號", "")).strip().upper()
        if not code:
            continue

        start_dt = parse_date(warrant.get("上市日", ""))
        end_dt = parse_date(warrant.get("最後交易日", ""))
        if target_dt:
            if start_dt and target_dt.date() < start_dt.date():
                continue
            if end_dt and target_dt.date() > end_dt.date():
                continue

        start_key = start_dt or datetime.min
        if code not in selected or start_key >= selected_start[code]:
            selected[code] = warrant
            selected_start[code] = start_key

    with _WARRANT_LOOKUP_CACHE_LOCK:
        # 若計算期間 warrants 未被另一流程切換，才寫入目前快取。
        if warrants is _WARRANT_LOOKUP_CACHE_WARRANTS_REF:
            _WARRANT_LOOKUP_CACHE[cache_key] = selected

    return selected




def normalize_finmind_warrant_day(raw_df, warrants, broker_map, date_value):
    columns = [
        "權證代號", "權證名稱", "標的股", "標的名稱",
        "分點", "分點名稱", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
        "買超股數", "買超金額",
    ]
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=columns)

    warrant_map = _warrant_lookup(warrants, date_value)
    broker_by_code = {}
    for label, (broker_name, broker_code) in (broker_map or {}).items():
        norm = normalize_broker_code_for_compare(broker_code)
        if norm:
            broker_by_code[norm] = (
                str(label).strip(),
                str(broker_name).strip(),
                str(broker_code).strip(),
            )

    if not warrant_map or not broker_by_code:
        return pd.DataFrame(columns=columns)

    df = raw_df.copy()
    df["stock_id"] = df["stock_id"].astype(str).str.strip().str.upper()
    df["securities_trader_id"] = (
        df["securities_trader_id"].astype(str).str.strip().str.upper()
    )
    df = df[
        df["stock_id"].isin(warrant_map)
        & df["securities_trader_id"].isin(broker_by_code)
    ].copy()
    if df.empty:
        return pd.DataFrame(columns=columns)

    df["price"] = pd.to_numeric(df["price"], errors="coerce").fillna(0.0)
    df["buy"] = pd.to_numeric(df["buy"], errors="coerce").fillna(0).astype("int64")
    df["sell"] = pd.to_numeric(df["sell"], errors="coerce").fillna(0).astype("int64")
    df = df[(df["buy"] != 0) | (df["sell"] != 0)].copy()
    if df.empty:
        return pd.DataFrame(columns=columns)

    # FinMind 是價格檔位明細；必須先逐列 price × 股數，再彙總成分點日資料。
    df["_buy_amount"] = (df["price"] * df["buy"]).round().astype("int64")
    df["_sell_amount"] = (df["price"] * df["sell"]).round().astype("int64")
    parsed_dates = pd.to_datetime(df["date"], errors="coerce")
    fallback_date = normalize_date_str(date_value)
    df["日期"] = parsed_dates.dt.strftime("%Y/%m/%d").fillna(fallback_date)
    df["broker_name_finmind"] = df["securities_trader"].fillna("").astype(str).str.strip()

    grouped = (
        df.groupby(["stock_id", "securities_trader_id", "日期"], as_index=False, sort=False)
        .agg(
            買進股數=("buy", "sum"),
            賣出股數=("sell", "sum"),
            買進金額=("_buy_amount", "sum"),
            賣出金額=("_sell_amount", "sum"),
            broker_name_finmind=("broker_name_finmind", "last"),
        )
    )

    records = []
    for (
        wcode, broker_code_raw, date_str, buy_qty, sell_qty,
        buy_amount, sell_amount, finmind_broker_name,
    ) in grouped.itertuples(index=False, name=None):
        wcode = str(wcode).strip().upper()
        bnorm = normalize_broker_code_for_compare(broker_code_raw)
        warrant = warrant_map.get(wcode)
        broker = broker_by_code.get(bnorm)
        if warrant is None or broker is None:
            continue

        label, configured_name, canonical_code = broker
        broker_name = str(finmind_broker_name or configured_name).strip() or configured_name
        buy_qty = int(buy_qty or 0)
        sell_qty = int(sell_qty or 0)
        buy_amount = int(buy_amount or 0)
        sell_amount = int(sell_amount or 0)
        records.append({
            "權證代號": str(warrant.get("代號", wcode)).strip(),
            "權證名稱": str(warrant.get("名稱", "")).strip(),
            "標的股": str(warrant.get("標的股", "")).strip(),
            "標的名稱": str(warrant.get("標的名稱", "")).strip(),
            "分點": label,
            "分點名稱": broker_name,
            "券商代號": canonical_code,
            "日期": str(date_str),
            "買進股數": buy_qty,
            "賣出股數": sell_qty,
            "買進金額": buy_amount,
            "賣出金額": sell_amount,
            "買超股數": buy_qty - sell_qty,
            "買超金額": buy_amount - sell_amount,
        })

    if not records:
        return pd.DataFrame(columns=columns)
    out = pd.DataFrame.from_records(records, columns=columns)
    out = out.sort_values(["日期", "分點", "標的股", "權證代號"]).reset_index(drop=True)
    return out




def _merge_complete_finmind_days(history_df, normalized_by_date, broker_map):
    """一次合併多個完整日快照，避免 repair 逐日 concat／排序造成 O(n²) 放大。"""
    normalized_by_date = normalized_by_date or {}
    if history_df is None or history_df.empty:
        base = pd.DataFrame()
    else:
        base = history_df.copy()

    success_dates = {normalize_date_str(date_value) for date_value in normalized_by_date}
    active_codes = {
        normalize_broker_code_for_compare(code)
        for _, code in (broker_map or {}).values()
        if normalize_broker_code_for_compare(code)
    }

    if not base.empty and success_dates and {"日期", "券商代號"}.issubset(base.columns):
        date_series = base["日期"].map(normalize_date_str)
        broker_series = base["券商代號"].map(normalize_broker_code_for_compare)
        base = base[~(date_series.isin(success_dates) & broker_series.isin(active_codes))].copy()

    frames = []
    if not base.empty:
        frames.append(base)
    for date_value in sorted(normalized_by_date):
        frame = normalized_by_date.get(date_value)
        if frame is not None and not frame.empty:
            frames.append(frame)

    if not frames:
        return base

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined["日期"] = combined["日期"].map(normalize_date_str)
    combined = combined.drop_duplicates(
        subset=["權證代號", "券商代號", "日期"],
        keep="last",
    )
    return combined.sort_values(["權證代號", "券商代號", "日期"]).reset_index(drop=True)


def prune_finmind_daily_raw_cache(keep_date=None):
    """只保留最近少量原始全市場日檔，避免 GitHub Actions 快取膨脹。"""
    try:
        entries = []
        for filename in os.listdir(FINMIND_DAILY_CACHE_DIR):
            match = re.fullmatch(r"(\d{4}-\d{2}-\d{2})\.parquet", filename)
            if not match:
                continue
            dt = parse_date(match.group(1))
            if dt:
                entries.append((dt, os.path.join(FINMIND_DAILY_CACHE_DIR, filename)))

        entries.sort(key=lambda item: item[0], reverse=True)
        keep_paths = {path for _dt, path in entries[:FINMIND_RAW_DAILY_RETENTION_DAYS]}
        if keep_date:
            try:
                keep_paths.add(_finmind_day_cache_path(keep_date))
            except Exception:
                pass

        removed = 0
        for _dt, raw_path in entries:
            if raw_path in keep_paths:
                continue
            try:
                os.remove(raw_path)
                removed += 1
            except FileNotFoundError:
                pass
        if removed:
            print(f"  🧹 FinMind 原始全市場日檔：移除 {removed:,} 檔，只保留最近 {FINMIND_RAW_DAILY_RETENTION_DAYS} 檔")
    except Exception as exc:
        print(f"  ⚠️ FinMind 原始日檔清理失敗：{type(exc).__name__}: {exc}")


def _load_finmind_state():
    try:
        if os.path.exists(FINMIND_STATE_PATH):
            with open(FINMIND_STATE_PATH, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
                if isinstance(payload, dict):
                    return payload
    except Exception:
        pass
    return {"completed_dates": []}


def _save_finmind_state(state):
    os.makedirs(os.path.dirname(FINMIND_STATE_PATH), exist_ok=True)
    tmp = f"{FINMIND_STATE_PATH}.tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(state, fh, ensure_ascii=False, indent=2, sort_keys=True)
    os.replace(tmp, FINMIND_STATE_PATH)



def _run_finmind_refresh_dates(target_date, history_df, force_full=False):
    trading_dates = get_finmind_trading_dates()
    target_dt = parse_date(target_date) or datetime.today()
    eligible = [d for d in trading_dates if d <= target_dt.date()]
    if not eligible:
        eligible = [target_dt.date()]

    recent_count = max(int(os.getenv("DAILY_UPDATE_DAYS", "3")), 1)
    repair_count = min(
        max(int(os.getenv("INITIAL_HISTORY_DAYS", str(HISTORY_RETENTION_TRADING_DAYS))), 1),
        HISTORY_RETENTION_TRADING_DAYS,
    )

    history_empty = history_df is None or history_df.empty
    if force_full or workflow_is_repair() or FORCE_FULL_CACHE_REFRESH or (history_empty and FINMIND_INITIAL_BACKFILL_ON_EMPTY):
        count = repair_count
    else:
        count = recent_count

    selected = eligible[-count:]
    target_day = target_dt.date()
    if target_day not in selected:
        selected.append(target_day)
    return [d.strftime("%Y/%m/%d") for d in sorted(set(selected))]




def refresh_history_from_finmind(warrants, broker_map, history_df, target_date, preloaded_target_df=None):
    global _FINMIND_TARGET_DATE_OK, _FINMIND_TARGET_DATE

    state = _load_finmind_state()
    schema_reset = state.get("schema_version") != FINMIND_CACHE_SCHEMA_VERSION
    if schema_reset:
        print(
            f"  ♻️ 偵測到舊資料源／舊快取格式，將以 FinMind 完整重建最近 "
            f"{HISTORY_RETENTION_TRADING_DAYS} 個交易日，不沿用舊原始分點歷史。"
        )
        history_df = pd.DataFrame()

    refresh_dates = _run_finmind_refresh_dates(
        target_date,
        history_df,
        force_full=schema_reset,
    )
    target_key = normalize_date_str(target_date)
    previous_empty = history_df is None or history_df.empty
    combined = history_df.copy() if history_df is not None else pd.DataFrame()
    completed_dates = set(str(x) for x in state.get("completed_dates", []) if str(x).strip())
    success_dates = []
    failed_dates = []
    normalized_by_date = {}
    target_status = "error"

    print(
        f"【Step 3】FinMind 全市場權證分點批次更新：{len(refresh_dates):,} 個交易日｜"
        f"workers={min(FINMIND_HISTORY_WORKERS, len(refresh_dates))}"
    )

    def load_one(date_key):
        if date_key == target_key and preloaded_target_df is not None:
            raw = preloaded_target_df
            status = "ok"
        else:
            force = bool(date_key == target_key and FINMIND_FORCE_REFRESH_TARGET_DATE)
            raw, status = download_finmind_warrant_day(date_key, force_refresh=force)
        normalized = (
            normalize_finmind_warrant_day(raw, warrants, broker_map, date_key)
            if status == "ok"
            else pd.DataFrame()
        )
        return date_key, raw, normalized, status

    max_workers = min(FINMIND_HISTORY_WORKERS, max(len(refresh_dates), 1))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(load_one, date_key): date_key for date_key in refresh_dates}
        for future in as_completed(futures):
            date_key = futures[future]
            try:
                result_date, raw, normalized, status = future.result()
            except Exception as exc:
                print(f"  ⚠️ FinMind 日資料處理失敗：{date_key}｜{type(exc).__name__}: {exc}")
                failed_dates.append(date_key)
                continue

            if result_date == target_key:
                target_status = status
                _FINMIND_TARGET_DATE = target_key
                _FINMIND_TARGET_DATE_OK = status == "ok"

            if status != "ok":
                failed_dates.append(result_date)
                continue

            normalized_by_date[result_date] = normalized
            completed_dates.add(result_date)
            success_dates.append(result_date)
            print(
                f"  ✅ {result_date}：市場原始 {len(raw):,} 列｜"
                f"追蹤分點彙總 {len(normalized):,} 列"
            )

    if success_dates:
        combined = _merge_complete_finmind_days(combined, normalized_by_date, broker_map)

    state["completed_dates"] = sorted(completed_dates)[-max(HISTORY_RETENTION_TRADING_DAYS * 2, 400):]
    state["last_target_date"] = target_key
    state["last_target_status"] = target_status
    state["updated_at"] = datetime.today().strftime("%Y-%m-%d %H:%M:%S")
    if not failed_dates and len(success_dates) == len(refresh_dates):
        state["schema_version"] = FINMIND_CACHE_SCHEMA_VERSION
    else:
        state.pop("schema_version", None)
    _save_finmind_state(state)

    combined, identity_stats = repair_history_metadata_from_warrants(combined, warrants)
    if identity_stats.get("rows", 0) > 0:
        print(
            f"  🧭 歷史快取身分配對修正：{identity_stats['rows']:,} 列｜"
            f"權證名稱 {identity_stats['warrant_name']:,}｜"
            f"標的股 {identity_stats['underlying_code']:,}｜"
            f"標的名稱 {identity_stats['underlying_name']:,}"
        )

    if success_dates or identity_stats.get("rows", 0) > 0:
        combined = save_history_cache(combined, fetched_items=None, previous_history_empty=previous_empty)
    prune_finmind_daily_raw_cache(keep_date=target_key)

    return combined, {
        "target_status": target_status,
        "success_dates": success_dates,
        "failed_dates": failed_dates,
        "schema_reset": schema_reset,
    }



def _normalized_broker_label_key(value):
    """分點名稱比對用正規化；只移除常見分隔符，不改變中文字義。"""
    text = str(strip_gsheet_text_prefix(value or "")).strip()
    return re.sub(r"[\s\-－_—–‧・．.()（）]+", "", text)


def _resolve_active_broker_label(label, scope="all"):
    """
    將舊版／帶連字號的分點名稱解析成目前有效標籤。

    回傳空字串代表該名稱不屬於目前有效分點，呼叫端可安全移除該資料列。
    """
    raw_label = str(strip_gsheet_text_prefix(label or "")).strip()
    if not raw_label:
        return ""

    label_to_code, _ = configured_broker_pair_maps_for_scope(scope)
    active_labels = list(label_to_code.keys())
    if raw_label in label_to_code:
        return raw_label

    normalized = _normalized_broker_label_key(raw_label)
    if normalized:
        for active_label in active_labels:
            if normalized == _normalized_broker_label_key(active_label):
                return active_label

        fallback_source = FULL_FALLBACK
        for active_label in active_labels:
            fallback_name = ""
            if active_label in fallback_source:
                fallback_name = str(fallback_source[active_label][0]).strip()
            if fallback_name and normalized == _normalized_broker_label_key(fallback_name):
                return active_label

    pattern_source = FULL_TARGET_PATTERNS
    if str(scope or "all").strip().lower() == "selected5":
        allowed = configured_broker_labels_for_scope("selected5")
    else:
        allowed = configured_broker_labels_for_scope("all")

    matches = []
    for active_label in active_labels:
        if active_label not in allowed:
            continue
        pattern = pattern_source.get(active_label)
        if pattern and re.search(pattern, raw_label):
            matches.append(active_label)

    return matches[0] if len(matches) == 1 else ""


def _canonical_gsheet_header(value):
    header = str(strip_gsheet_text_prefix(value or "")).strip()
    aliases = {
        "券商分點": "分點",
        "分公司": "分點",
        "券商代碼": "券商代號",
        "分點代號": "券商代號",
        "買進券商": "買進分點",
    }
    return aliases.get(header, header)


def _contiguous_row_ranges(row_numbers):
    """把 1-based 列號整理成連續區間，供 Google Sheet 由後往前刪列。"""
    rows = sorted({int(x) for x in row_numbers if int(x) >= 1})
    if not rows:
        return []

    ranges = []
    start = prev = rows[0]
    for row_no in rows[1:]:
        if row_no == prev + 1:
            prev = row_no
            continue
        ranges.append((start, prev))
        start = prev = row_no
    ranges.append((start, prev))
    return ranges


def cleanup_deleted_broker_rows_in_existing_worksheets():
    """
    保留所有既有 Google Sheet 工作表，只清除其中已失效分點的資料列。

    安全規則：
    1. 絕不呼叫 del_worksheet，也不刪除或重建任何工作表。
    2. 只有找到明確的「分點／券商代號／買進分點」表頭後才處理後續資料列。
    3. 券商代號仍有效但名稱已改名時，只更新名稱，不刪資料。
    4. 無法辨識表頭、人工說明列、彙總列與沒有分點欄位的工作表完全不動。
    5. 刪資料列時由後往前刪除，避免列號位移造成誤刪。
    """
    if not GSHEET_CLEAN_DELETED_BROKER_ROWS or not gsheet_enabled():
        return

    sh = get_gsheet_spreadsheet()
    if sh is None:
        return

    relevant_headers = {"分點", "券商代號", "買進分點", "資料範圍"}
    total_deleted = 0
    total_renamed = 0
    touched_sheets = 0

    for ws in list(sh.worksheets()):
        values = _worksheet_values_with_formulas(ws)
        if not values:
            continue

        current_headers = None
        rows_to_delete = []
        cell_updates = []

        for row_no, row in enumerate(values, start=1):
            canonical_row = [_canonical_gsheet_header(value) for value in row]
            non_empty_headers = {value for value in canonical_row if value}

            # 任何位置出現明確分點欄位時，視為新的表頭區段。
            if non_empty_headers & {"分點", "券商代號", "買進分點"}:
                current_headers = {
                    header: col_idx
                    for col_idx, header in enumerate(canonical_row)
                    if header
                }
                continue

            if not current_headers:
                continue

            record = {}
            for header, col_idx in current_headers.items():
                if header not in relevant_headers:
                    continue
                record[header] = row[col_idx] if col_idx < len(row) else ""

            broker_label = str(strip_gsheet_text_prefix(record.get("分點", ""))).strip()
            broker_code = str(strip_gsheet_text_prefix(record.get("券商代號", ""))).strip()
            buy_broker = str(strip_gsheet_text_prefix(record.get("買進分點", ""))).strip()

            # 空白列或沒有分點識別資料的彙總／說明列完全不動。
            if not (broker_label or broker_code or buy_broker):
                continue

            cleaned = normalize_or_remove_deleted_broker_result_record(record)
            if cleaned is None:
                rows_to_delete.append(row_no)
                continue

            for header in ("分點", "券商代號", "買進分點"):
                if header not in current_headers or header not in cleaned:
                    continue
                col_idx = current_headers[header]
                old_value = row[col_idx] if col_idx < len(row) else ""
                new_value = cleaned.get(header, "")
                if str(strip_gsheet_text_prefix(old_value)).strip() == str(new_value).strip():
                    continue
                cell_updates.append({
                    "range": f"{get_column_letter(col_idx + 1)}{row_no}",
                    "values": [[clean_gsheet_value(new_value)]],
                })

        if not rows_to_delete and not cell_updates:
            continue

        touched_sheets += 1

        # 先修正仍有效但已改名的分點；之後刪列時內容會自然跟著位移。
        if cell_updates:
            for start_idx in range(0, len(cell_updates), 500):
                chunk = cell_updates[start_idx:start_idx + 500]
                gsheet_api_call(
                    f"更新有效分點名稱 {ws.title}",
                    ws.batch_update,
                    chunk,
                    value_input_option="USER_ENTERED",
                )
            total_renamed += len(cell_updates)

        # 必須由後往前刪除，否則前面刪列後會改變後續列號。
        for start_row, end_row in reversed(_contiguous_row_ranges(rows_to_delete)):
            gsheet_api_call(
                f"清除失效分點舊資料 {ws.title} R{start_row}:R{end_row}",
                ws.delete_rows,
                start_row,
                end_row,
            )

        total_deleted += len(rows_to_delete)
        print(
            f"  🧹 Google Sheet 舊分點資料清理：{ws.title}｜"
            f"刪除 {len(rows_to_delete):,} 列｜更正名稱 {len(cell_updates):,} 格｜工作表保留"
        )

    if touched_sheets:
        print(
            f"  ✅ Google Sheet 舊分點資料清理完成："
            f"處理 {touched_sheets:,} 張工作表｜刪除 {total_deleted:,} 列｜"
            f"更正名稱 {total_renamed:,} 格｜未刪除任何工作表"
        )
    else:
        print("  ✅ Google Sheet 舊分點資料檢查完成：沒有需要清除的失效分點列，未刪除任何工作表。")


def get_all_call_warrants_live(cached_warrants=None):
    global LIVE_WARRANT_SNAPSHOT_READY, CURRENT_LIVE_WARRANT_CODES
    global _CURRENT_STOCK_CODE_TO_NAME, _CURRENT_STOCK_NAME_TO_CODE, _CURRENT_UNDERLYING_RESOLVER
    global _CURRENT_WARRANT_INTERVAL_RECORDS

    print("【Step 1】FinMind 取得認購權證清單與日期化標的對照...")
    with ThreadPoolExecutor(max_workers=3) as ex:
        future_info = ex.submit(fetch_finmind_dataset, "TaiwanStockInfoWithWarrant")
        future_summary = ex.submit(fetch_finmind_dataset, "TaiwanStockInfoWithWarrantSummary")
        future_stock = ex.submit(fetch_finmind_dataset, "TaiwanStockInfo")
        info_df = future_info.result()
        summary_df = future_summary.result()
        stock_df = future_stock.result()

    required_summary = {"stock_id", "target_stock_id", "type", "date", "end_date"}
    if info_df.empty or summary_df.empty or not required_summary.issubset(summary_df.columns):
        LIVE_WARRANT_SNAPSHOT_READY = False
        return []

    # 先建立唯一股號／股名主檔，禁止再用 dict(zip) 的最後一列任意覆蓋。
    # TaiwanStockInfo 為主；TaiwanStockInfoWithWarrant 中非權證的標的／指數名稱只做補充。
    stock_master_parts = []
    if stock_df is not None and not stock_df.empty:
        stock_master_parts.append(stock_df.copy())
    if info_df is not None and not info_df.empty and {"stock_id", "stock_name"}.issubset(info_df.columns):
        info_master = info_df.copy()
        if "date" not in info_master.columns:
            info_master["date"] = ""
        stock_master_parts.append(info_master)
    stock_master_df = pd.concat(stock_master_parts, ignore_index=True, sort=False) if stock_master_parts else pd.DataFrame()
    stock_code_to_name, stock_name_to_code = build_canonical_stock_master(stock_master_df)
    _CURRENT_STOCK_CODE_TO_NAME = stock_code_to_name
    _CURRENT_STOCK_NAME_TO_CODE = stock_name_to_code
    _CURRENT_UNDERLYING_RESOLVER = build_underlying_resolver_from_stock_master(stock_code_to_name)
    _reset_underlying_resolver_runtime_cache(_CURRENT_UNDERLYING_RESOLVER)
    cached_warrant_index = build_cached_warrant_index(cached_warrants)

    info = info_df.copy().fillna("")
    info["stock_id"] = info["stock_id"].map(normalize_security_code_text)
    info_name_candidates = defaultdict(list)
    if "stock_name" in info.columns:
        for code, name in info[["stock_id", "stock_name"]].itertuples(index=False, name=None):
            name = str(name or "").strip()
            if code and name and name not in info_name_candidates[code]:
                info_name_candidates[code].append(name)

    summary = summary_df.copy().fillna("")
    summary["stock_id"] = summary["stock_id"].map(normalize_security_code_text)
    summary["target_stock_id"] = summary["target_stock_id"].map(normalize_security_code_text)
    summary["_list_date"] = pd.to_datetime(summary["date"], errors="coerce")
    summary["_end_date"] = pd.to_datetime(summary["end_date"], errors="coerce")

    today = pd.Timestamp(datetime.today().date())
    lookback_days = max(
        int(os.getenv("WARRANT_METADATA_LOOKBACK_DAYS", str(max(HISTORY_RETENTION_TRADING_DAYS * 2, 420)))),
        365,
    )
    cutoff = today - pd.Timedelta(days=lookback_days)
    call_mask = summary["type"].astype(str).str.contains("認購", na=False)
    relevant_mask = (
        summary["_list_date"].notna()
        & summary["_end_date"].notna()
        & (summary["_list_date"] <= today)
        & (summary["_end_date"] >= cutoff)
    )
    summary = summary[call_mask & relevant_mask].copy()
    # target_stock_id 必須保留在去重鍵中；同區間若 FinMind 有衝突列，後面再用身分分數選正確者。
    summary = summary.sort_values(["stock_id", "_list_date", "_end_date", "target_stock_id"]).drop_duplicates(
        ["stock_id", "target_stock_id", "_list_date", "_end_date"],
        keep="last",
    )

    warrants = []
    active_codes = set()
    identity_override_count = 0
    name_guard_count = 0

    for code, summary_underlying, list_ts, end_ts in summary[
        ["stock_id", "target_stock_id", "_list_date", "_end_date"]
    ].itertuples(index=False, name=None):
        code = normalize_security_code_text(code)
        summary_underlying = normalize_security_code_text(summary_underlying)
        if not code or not summary_underlying or pd.isna(list_ts) or pd.isna(end_ts):
            continue

        list_date = pd.Timestamp(list_ts).strftime("%Y/%m/%d")
        end_date = pd.Timestamp(end_ts).strftime("%Y/%m/%d")
        provisional_name, name_source = select_warrant_name_for_interval(
            code,
            summary_underlying,
            list_date,
            end_date,
            info_name_candidates,
            cached_warrant_index,
            resolver=_CURRENT_UNDERLYING_RESOLVER,
            enforce_underlying_match=False,
        )
        final_underlying, final_underlying_name, identity_source = reconcile_underlying_identity(
            provisional_name,
            summary_underlying,
            stock_code_to_name.get(summary_underlying, ""),
            code_to_name=stock_code_to_name,
            resolver=_CURRENT_UNDERLYING_RESOLVER,
        )
        if identity_source == "warrant_name_override":
            identity_override_count += 1

        # 以更正後標的再挑一次名稱；可處理 Summary 錯標但權證名稱正確的情況。
        final_name, final_name_source = select_warrant_name_for_interval(
            code,
            final_underlying,
            list_date,
            end_date,
            info_name_candidates,
            cached_warrant_index,
            resolver=_CURRENT_UNDERLYING_RESOLVER,
        )
        if final_name_source.startswith("code_only"):
            name_guard_count += 1

        final_underlying_name = str(stock_code_to_name.get(final_underlying, final_underlying_name)).strip()
        warrants.append({
            "代號": code,
            "名稱": final_name,
            "標的股": final_underlying,
            "標的名稱": final_underlying_name,
            "上市日": list_date,
            "最後交易日": end_date,
            "_名稱來源": final_name_source or name_source,
            "_標的來源": identity_source,
        })

    warrants = dedupe_warrant_interval_records(warrants, resolver=_CURRENT_UNDERLYING_RESOLVER)
    _CURRENT_WARRANT_INTERVAL_RECORDS = [dict(rec) for rec in warrants]
    _reset_warrant_lookup_cache(_CURRENT_WARRANT_INTERVAL_RECORDS)
    for rec in warrants:
        rec.pop("_名稱來源", None)
        rec.pop("_標的來源", None)
        if parse_date(rec.get("上市日", "")) <= datetime.today() <= parse_date(rec.get("最後交易日", "")):
            active_codes.add(normalize_security_code_text(rec.get("代號", "")))

    LIVE_WARRANT_SNAPSHOT_READY = bool(warrants)
    CURRENT_LIVE_WARRANT_CODES = active_codes
    print(
        f"  ✅ FinMind 認購權證日期區間紀錄：{len(warrants):,} 筆｜"
        f"今日有效代號：{len(active_codes):,} 支｜"
        f"標的交叉更正：{identity_override_count:,} 筆｜"
        f"權證名稱防錯降級：{name_guard_count:,} 筆"
    )
    return warrants






def get_all_call_warrants():
    global CURRENT_LIVE_WARRANT_CODES, LIVE_WARRANT_SNAPSHOT_READY
    global _CURRENT_WARRANT_INTERVAL_RECORDS

    cached_warrants = load_warrants_cache()
    try:
        warrants = get_all_call_warrants_live(cached_warrants)
    except Exception as exc:
        warrants = []
        print(f"  ⚠️ FinMind 權證清單更新失敗：{type(exc).__name__}: {exc}")

    if warrants:
        _CURRENT_WARRANT_INTERVAL_RECORDS = [dict(rec) for rec in warrants]
        save_warrants_cache(warrants)
        return warrants

    if cached_warrants:
        _CURRENT_WARRANT_INTERVAL_RECORDS = [dict(rec) for rec in cached_warrants]
        LIVE_WARRANT_SNAPSHOT_READY = False
        today = datetime.today()
        CURRENT_LIVE_WARRANT_CODES = {
            str(w.get("代號", "")).strip().upper()
            for w in cached_warrants
            if str(w.get("代號", "")).strip()
            and (not parse_date(w.get("上市日", "")) or parse_date(w.get("上市日", "")) <= today)
            and (not parse_date(w.get("最後交易日", "")) or parse_date(w.get("最後交易日", "")) >= today)
        }
        print(
            f"  ⚠️ 沿用既有權證日期對照快取：{len(cached_warrants):,} 筆；"
            "本次不允許以空結果清除 Google Sheet。"
        )
        return cached_warrants

    return []




# ══════════════════════════════════════════════════════════════════════
# Step 2：找目標分點券商代號
# ══════════════════════════════════════════════════════════════════════


def find_broker_codes_live(warrants=None):
    print("【Step 2】FinMind 取得目標分點券商代號...")
    try:
        df = fetch_finmind_dataset("TaiwanSecuritiesTraderInfo")
    except Exception as exc:
        print(f"  ⚠️ FinMind 券商資訊取得失敗：{type(exc).__name__}: {exc}")
        df = pd.DataFrame()

    found = {}
    if not df.empty and {"securities_trader_id", "securities_trader"}.issubset(df.columns):
        work = df[["securities_trader_id", "securities_trader"]].copy().fillna("")
        records = []
        for raw_code, raw_name in work.itertuples(index=False, name=None):
            code = str(raw_code).strip()
            name = str(raw_name).strip()
            norm = normalize_broker_code_for_compare(code)
            if norm:
                records.append((norm, code, name))

        # 優先以目前設定的券商代號精準配對，名稱 regex 只作備援。
        by_code = {norm: (name, code) for norm, code, name in records}
        for label in TARGET_PATTERNS:
            fallback_name, fallback_code = FALLBACK.get(label, (label, ""))
            hit = by_code.get(normalize_broker_code_for_compare(fallback_code))
            if hit:
                found[label] = (hit[0] or fallback_name, fallback_code)

        for _norm, raw_code, name in records:
            label = match_target(name)
            if label and label not in found:
                canonical_code = FALLBACK.get(label, (name, raw_code))[1]
                found[label] = (name, canonical_code)

    for label in TARGET_PATTERNS:
        if label not in found and label in FALLBACK:
            found[label] = FALLBACK[label]
            print(f"    ✅ {label:14s} → {FALLBACK[label][0]} ({FALLBACK[label][1]}) [設定備援]")
        elif label in found:
            print(f"    ✅ {label:14s} → {found[label][0]} ({found[label][1]})")

    return found






def find_broker_codes(warrants=None):
    cached = filter_broker_map_for_active_targets(load_broker_map_cache())
    try:
        live = filter_broker_map_for_active_targets(find_broker_codes_live(warrants))
    except Exception:
        live = {}
    if live:
        save_broker_map_cache(live)
        return live
    if cached:
        print("  ⚠️ 沿用既有分點代號快取。")
        return cached
    return {}



# ══════════════════════════════════════════════════════════════════════
# Step 3a：預篩有目標分點出現的 (權證, 分點) 組合
# ══════════════════════════════════════════════════════════════════════














# ══════════════════════════════════════════════════════════════════════
# 新制說明：不再建立舊版「A_單檔大買」逐筆事件。
# 事件統一由 build_amount_class_events() 依「同分點 × 同標的 × 同一天」產生 A/B/C/D/E。
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# Step 3b：抓候選組合歷史資料
# ══════════════════════════════════════════════════════════════════════



# ══════════════════════════════════════════════════════════════════════
# 建立每日資料，用於新制 A/B/C/D/E 金額強度分類
# ══════════════════════════════════════════════════════════════════════

def build_daily_records(items):
    daily_records = []

    for item in items:
        if not item["underlying_code"]:
            continue

        for row in item["df"].itertuples(index=False):
            row_dict = row._asdict()
            daily_records.append({
                "日期": row_dict["日期"],
                "分點": item["broker_label"],
                "分點名稱": item["broker_name"],
                "券商代號": item["broker_code"],
                "權證代號": item["warrant_code"],
                "權證名稱": item["warrant_name"],
                "標的股": item["underlying_code"],
                "標的名稱": item.get("underlying_name", ""),
                "買進股數": int(row_dict["買進股數"]),
                "賣出股數": int(row_dict["賣出股數"]),
                "買進金額": int(row_dict["買進金額"]),
                "賣出金額": int(row_dict["賣出金額"]),
                "買超股數": int(row_dict["買超股數"]),
                "買超金額": int(row_dict["買超金額"]),
            })

    return daily_records


def normalize_warrant_code_for_unique(warrant_code):
    """
    用於新制金額強度分類與 FIFO 扣減的權證代號正規化。

    """
    return str(warrant_code or "").strip()


def make_broker_warrant_key(broker_code, warrant_code):
    """
    建立新制金額強度分類與 FIFO 扣減使用的唯一鍵。

    正確單位為「券商代號 + 權證代號」：
    - 新制事件以同分點、同標的、同一天合併。
    - FIFO 扣減時仍需用券商代號與權證代號分開追蹤。
    """
    broker_code = str(broker_code or "").strip()
    warrant_code = normalize_warrant_code_for_unique(warrant_code)

    if not broker_code or not warrant_code:
        return None

    return broker_code, warrant_code


def classify_amount_class(total_amount):
    """
    依同標的單日累積買進金額回傳 A / B / C / D / E 分類。
    """
    try:
        total_amount = int(total_amount or 0)
    except Exception:
        total_amount = 0

    for code, label, lower, upper in AMOUNT_CLASS_SPECS:
        if total_amount >= lower and (upper is None or total_amount < upper):
            return code, label

    return "", ""


def split_amount_class_events(events):
    out = {code: [] for code in AMOUNT_CLASS_CODES}

    for ev in events or []:
        code = str(ev.get("事件代碼", "")).strip()
        if code in out:
            out[code].append(ev)

    return tuple(out[code] for code in AMOUNT_CLASS_CODES)


def iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events=None):
    event_groups = [
        ("A", a_events or []),
        ("B", b_events or []),
        ("C", c_events or []),
        ("D", d_events or []),
    ]

    if e_events is not None:
        event_groups.append(("E", e_events or []))

    return event_groups




def build_amount_class_events(daily_records, item_map):
    """
    新制 A/B/C/D/E 金額強度分類。

    事件單位：同一分點 + 同一標的 + 同一天。
    進入條件：該事件內至少一檔權證單日買進金額 >= AMOUNT_THRESH。
    分類依據：同一分點 + 同一標的 + 同一天的所有權證買進金額合計。

    注意：
    1. 這裡使用「買進金額」而不是「買超金額」，避免同日賣出把實際買入力道扣掉。
    2. 單筆 100 萬負責過濾小單；單日累積金額負責判斷資金強度。
    3. 同一天同分點同標的若買多檔權證，只會合併成一筆事件，不會拆成多筆分類。
    """
    events = []

    if not daily_records:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    df = pd.DataFrame(daily_records)

    if df.empty:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    df = df[(df["買進金額"] > 0) & (df["買進股數"] > 0)].copy()

    if df.empty:
        return {code: [] for code in AMOUNT_CLASS_CODES}

    group_cols = ["分點", "分點名稱", "券商代號", "標的股", "標的名稱", "日期"]

    for key, g in df.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_name, broker_code, underlying_code, underlying_name, date = key
        date = normalize_date_str(date)

        warrant_rows = g.groupby(["權證代號", "權證名稱"], as_index=False).agg({
            "買進金額": "sum",
            "買進股數": "sum",
        })

        lots = []
        max_single_amount = 0
        max_single_warrant = ""

        for wr in warrant_rows.itertuples(index=False):
            wr = wr._asdict()
            warrant_code = normalize_warrant_code_for_unique(wr["權證代號"])
            warrant_key = make_broker_warrant_key(broker_code, warrant_code)

            if not warrant_key:
                continue

            buy_amount = int(wr["買進金額"] or 0)
            buy_shares = int(wr["買進股數"] or 0)

            if buy_amount <= 0 or buy_shares <= 0:
                continue

            if buy_amount > max_single_amount:
                max_single_amount = buy_amount
                max_single_warrant = f'{wr["權證代號"]} {wr["權證名稱"]}'

            lots.append({
                "買進日": date,
                "權證代號": wr["權證代號"],
                "權證名稱": wr["權證名稱"],
                "金額": buy_amount,
                "股數": buy_shares,
            })

        if not lots:
            continue

        # 最低進入條件：同日同分點同標的內，至少有一檔權證單筆買進金額 >= 100 萬。
        if max_single_amount < AMOUNT_THRESH:
            continue

        total_amount = int(sum(lot["金額"] for lot in lots))
        total_shares = int(sum(lot["股數"] for lot in lots))
        warrant_count = len(set(normalize_warrant_code_for_unique(lot["權證代號"]) for lot in lots))

        if total_amount < AMOUNT_THRESH or total_shares <= 0:
            continue

        event_code, event_label = classify_amount_class(total_amount)

        if not event_code:
            continue

        event = {
            "事件類型": f"{event_code}-{event_label}",
            "事件代碼": event_code,
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "起始日": date,
            "結束日": date,
            "事件日": date,
            "涵蓋權證數": warrant_count,
            "權證清單": "；".join([f'{lot["權證代號"]} {lot["權證名稱"]}' for lot in lots]),
            "最大單筆金額": max_single_amount,
            "最大單筆權證": max_single_warrant,
            "單日累積買進金額": total_amount,
            "買超金額": total_amount,
            "買超股數": total_shares,
            "買超張數": total_shares // 1000,
            "lots": lots,
        }

        events.append(event)

    events = simulate_group_outcomes_fifo(events, item_map)
    return {code: group for code, group in zip(AMOUNT_CLASS_CODES, split_amount_class_events(events))}








# ══════════════════════════════════════════════════════════════════════
# B / C 群組事件出清推估
# ══════════════════════════════════════════════════════════════════════


_GROUP_OUTCOME_SALE_ROWS_CACHE = {}


def get_group_sale_rows_for_warrant(item_map, broker_code, warrant_code):
    """
    新制金額強度事件 FIFO 扣減使用的賣出資料快取。

    同一個「券商代號 + 權證代號」只整理一次每日賣出資料；
    後續由 simulate_group_outcomes_fifo() 將每一筆賣出依事件時間 FIFO 分配，
    避免同一筆賣出被不同事件重複使用。
    """
    key = (str(broker_code).strip(), str(warrant_code).strip())

    if key in _GROUP_OUTCOME_SALE_ROWS_CACHE:
        return _GROUP_OUTCOME_SALE_ROWS_CACHE[key]

    item = item_map.get(key)
    if not item:
        item = item_map.get((key[0], normalize_price_code(key[1])))

    rows = []

    if item:
        df = item.get("df", pd.DataFrame())

        if df is not None and not df.empty:
            needed_cols = ["日期", "賣出股數", "賣出金額"]

            if all(col in df.columns for col in needed_cols):
                for date, sell_s, sell_a in df[needed_cols].itertuples(index=False, name=None):
                    try:
                        sell_s = int(sell_s)
                        sell_a = int(sell_a)
                    except Exception:
                        continue

                    if sell_s > 0:
                        rows.append({
                            "日期": normalize_date_str(date),
                            "權證代號": key[1],
                            "賣出股數": sell_s,
                            "賣出金額": sell_a,
                        })

    rows = sorted(rows, key=lambda x: (x["日期"], x["權證代號"]))
    _GROUP_OUTCOME_SALE_ROWS_CACHE[key] = rows
    return rows


def simulate_group_outcomes_fifo(events, item_map):
    """
    對同一類別的新制金額強度事件進行「跨事件、逐筆 FIFO」扣減。

    規則：
    1. 每個符合條件的事件各自保留成一筆資料，不合併。
    2. 同一分點 + 同一權證的賣出，只能使用一次。
    3. 賣出依事件成立時間由早到晚扣；舊事件未扣完前，不得先扣新事件。
    4. 同日買賣不互扣：賣出日必須晚於該事件結束日。
    5. 減碼 / 出清只看事件 lots 的剩餘股數；金額只計算已實現報酬。
    """
    if not events:
        return events

    queues = {}

    for event_seq, event in enumerate(events):
        event_start_date = normalize_date_str(
            event.get("起始日") or event.get("事件日") or event.get("結束日") or ""
        )
        event_end_date = normalize_date_str(
            event.get("結束日") or event.get("事件日") or event.get("起始日") or ""
        )
        broker_code = str(event.get("券商代號", "")).strip()

        event["減碼日"] = None
        event["減碼賣出金額"] = None
        event["減碼獲利%"] = None
        event["出清日"] = None
        event["出清賣出金額"] = None
        event["出清獲利%"] = None
        event["持有天數"] = None
        event["狀態"] = "持有"
        event["累計賣出股數"] = 0
        event["已實現賣出金額"] = 0.0
        event["已實現成本"] = 0.0
        event["_fifo_allocations"] = []

        original_total = 0
        valid_lots = []

        for lot_seq, lot in enumerate(event.get("lots", [])):
            try:
                qty = int(lot.get("股數", 0) or 0)
                amount = float(lot.get("金額", 0) or 0)
            except Exception:
                continue

            warrant_code = normalize_warrant_code_for_unique(lot.get("權證代號", ""))
            if qty <= 0 or amount <= 0 or not broker_code or not warrant_code:
                continue

            buy_date = normalize_date_str(lot.get("買進日") or event_end_date)
            avg_cost = amount / qty if qty else 0

            lot["買進日"] = buy_date
            lot["股數"] = qty
            lot["金額"] = amount
            lot["均價"] = avg_cost
            lot["剩餘股數"] = qty
            lot["累計賣出股數"] = 0
            lot["已實現賣出金額"] = 0.0
            lot["已實現成本"] = 0.0

            valid_lots.append(lot)
            original_total += qty

            key = (broker_code, warrant_code)
            queues.setdefault(key, []).append({
                "event": event,
                "event_seq": event_seq,
                "lot": lot,
                "lot_seq": lot_seq,
                "event_start_date": event_start_date,
                "event_end_date": event_end_date,
                "buy_date": buy_date,
            })

        event["lots"] = valid_lots
        event["原始股數"] = original_total
        event["剩餘股數"] = original_total

    # 同一分點 + 權證依舊事件優先，再依 lot 買進日排序。
    for key, queue in queues.items():
        queue.sort(key=lambda ref: (
            ref["event_end_date"],
            ref["buy_date"],
            ref["event_seq"],
            ref["lot_seq"],
        ))

        broker_code, warrant_code = key
        sales = get_group_sale_rows_for_warrant(item_map, broker_code, warrant_code)

        for sale in sales:
            sell_date = normalize_date_str(sale.get("日期", ""))
            sell_qty = int(sale.get("賣出股數", 0) or 0)
            sell_amount = float(sale.get("賣出金額", 0) or 0)

            if not sell_date or sell_qty <= 0:
                continue

            sell_price = sell_amount / sell_qty if sell_qty > 0 else 0
            sell_left = sell_qty

            for ref in queue:
                if sell_left <= 0:
                    break

                # 事件視窗內的買賣已反映在事件成立金額 / 股數中；只扣事件結束日之後的賣出。
                if sell_date <= ref["event_end_date"]:
                    continue

                lot = ref["lot"]
                remaining = int(lot.get("剩餘股數", 0) or 0)
                if remaining <= 0:
                    continue

                alloc = min(sell_left, remaining)
                if alloc <= 0:
                    continue

                alloc_revenue = alloc * sell_price
                alloc_cost = alloc * float(lot.get("均價", 0) or 0)

                lot["剩餘股數"] = remaining - alloc
                lot["累計賣出股數"] = int(lot.get("累計賣出股數", 0) or 0) + alloc
                lot["已實現賣出金額"] = float(lot.get("已實現賣出金額", 0) or 0) + alloc_revenue
                lot["已實現成本"] = float(lot.get("已實現成本", 0) or 0) + alloc_cost

                event = ref["event"]
                event["_fifo_allocations"].append({
                    "日期": sell_date,
                    "股數": alloc,
                    "賣出金額": alloc_revenue,
                    "成本": alloc_cost,
                })

                sell_left -= alloc

    # 依每筆事件自己的剩餘股數判斷持有 / 減碼 / 出清。
    for event in events:
        original_total = int(sum(int(lot.get("股數", 0) or 0) for lot in event.get("lots", [])))
        remaining_total = int(sum(int(lot.get("剩餘股數", 0) or 0) for lot in event.get("lots", [])))
        sold_total = max(original_total - remaining_total, 0)

        allocations_by_date = {}
        for alloc in event.pop("_fifo_allocations", []):
            d = normalize_date_str(alloc.get("日期", ""))
            if not d:
                continue
            rec = allocations_by_date.setdefault(d, {
                "股數": 0,
                "賣出金額": 0.0,
                "成本": 0.0,
            })
            rec["股數"] += int(alloc.get("股數", 0) or 0)
            rec["賣出金額"] += float(alloc.get("賣出金額", 0) or 0)
            rec["成本"] += float(alloc.get("成本", 0) or 0)

        realized_revenue = sum(rec["賣出金額"] for rec in allocations_by_date.values())
        realized_cost = sum(rec["成本"] for rec in allocations_by_date.values())

        event["原始股數"] = original_total
        event["剩餘股數"] = remaining_total
        event["累計賣出股數"] = sold_total
        event["已實現賣出金額"] = realized_revenue
        event["已實現成本"] = realized_cost

        if sold_total <= 0:
            event["狀態"] = "持有"
            continue

        running_remaining = original_total
        cumulative_revenue = 0.0
        cumulative_cost = 0.0

        for sell_date in sorted(allocations_by_date.keys()):
            rec = allocations_by_date[sell_date]
            running_remaining = max(running_remaining - int(rec["股數"]), 0)
            cumulative_revenue += rec["賣出金額"]
            cumulative_cost += rec["成本"]

            if running_remaining > 0 and event.get("減碼日") is None:
                event["減碼日"] = sell_date
                event["減碼賣出金額"] = round(rec["賣出金額"], 0)
                event["減碼獲利%"] = round(
                    (rec["賣出金額"] - rec["成本"]) / rec["成本"] * 100,
                    2,
                ) if rec["成本"] else None

            if running_remaining <= 0:
                event["出清日"] = sell_date
                event["出清賣出金額"] = round(cumulative_revenue, 0)

                total_cost = float(sum(float(lot.get("金額", 0) or 0) for lot in event.get("lots", [])))
                event["出清獲利%"] = round(
                    (cumulative_revenue - total_cost) / total_cost * 100,
                    2,
                ) if total_cost else None

                start_dt = parse_date(event.get("起始日") or event.get("事件日"))
                exit_dt = parse_date(sell_date)
                if start_dt and exit_dt:
                    event["持有天數"] = (exit_dt - start_dt).days
                break

        # 最終狀態仍只依剩餘股數判斷。
        if remaining_total <= 0:
            event["剩餘股數"] = 0
            event["狀態"] = "出清"
        else:
            event["狀態"] = "減碼"

    return events




# ══════════════════════════════════════════════════════════════════════
# 舊版 B/C/D 事件建立邏輯已移除。
# 目前只使用 build_amount_class_events() 的新制 A/B/C/D/E 單日金額強度分類。
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# Step 4：抓收盤價
# ══════════════════════════════════════════════════════════════════════

def fetch_all_prices(
    a_events,
    b_events,
    c_events,
    d_events,
    e_events=None,
    persistent_price_cache=None,
    defer_save=False,
):
    print("【Step 4】抓收盤價...")

    code_ranges = {}

    def update_code_range(code, start_dt, end_dt):
        if not code or start_dt is None or end_dt is None:
            return

        code = normalize_price_code(code)

        if not code:
            return

        if end_dt < start_dt:
            end_dt = start_dt

        if code not in code_ranges:
            code_ranges[code] = [start_dt, end_dt]
        else:
            code_ranges[code][0] = min(code_ranges[code][0], start_dt)
            code_ranges[code][1] = max(code_ranges[code][1], end_dt)

    for _, group_events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in group_events:
            dt = parse_date(ev.get("事件日") or ev.get("結束日") or ev.get("起始日"))
            if dt:
                start_dt = dt - timedelta(days=60)
                end_dt = dt + timedelta(days=160)
                update_code_range(ev.get("標的股"), start_dt, end_dt)

                # 新制 A/B/C/D/E 的 D+ 欄位目前只看標的股；若未來需要群組事件權證明細價格，
                # 可設定 FETCH_GROUP_WARRANT_PRICES=1。
                if FETCH_GROUP_WARRANT_PRICES:
                    for lot in ev.get("lots", []):
                        update_code_range(lot.get("權證代號"), start_dt, end_dt)

    all_codes = list(code_ranges.keys())
    price_cache = {}
    total = len(all_codes)
    changed_price_codes = set()

    if persistent_price_cache is None:
        persistent_price_cache = load_price_cache()
        print(f"  價格快取讀取：{len(persistent_price_cache):,} 個代號")

    if total == 0:
        print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
        return _finish_price_ensure(
            price_cache,
            persistent_price_cache,
            changed_price_codes,
            defer_save=defer_save,
        )

    today = datetime.today()
    fetch_plan = {}

    for code in all_codes:
        start_dt, end_dt = code_ranges[code]

        if end_dt > today:
            end_dt = today

        if start_dt > end_dt:
            start_dt = end_dt

        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)

        if cached_prices:
            add_price_aliases(price_cache, code, cached_prices)

            valid_dates = sorted([d for d, p in cached_prices.items() if p is not None and p > 0])
            latest_dt = parse_date(valid_dates[-1]) if valid_dates else None

            # 快取不只看最後日期，也要檢查覆蓋率品質。
            # 若快取筆數太少、有效日期不足，或覆蓋不到需求區間，仍重新補抓完整區間，避免 D+ 欄位大量出現 -。
            if prices_need_yahoo_fallback(cached_prices, start_dt, end_dt):
                fetch_plan[code] = [start_dt, end_dt]
            elif latest_dt and latest_dt < end_dt:
                fetch_start_dt = latest_dt + timedelta(days=1)

                if fetch_start_dt <= end_dt:
                    fetch_plan[code] = [fetch_start_dt, end_dt]
            elif not latest_dt:
                fetch_plan[code] = [start_dt, end_dt]
        else:
            add_price_aliases(price_cache, code, {})
            fetch_plan[code] = [start_dt, end_dt]

    print(f"  價格代號去重後：{total:,} 檔")
    print(f"  價格快取命中：{total - len(fetch_plan):,} 檔")
    print(f"  本次需補抓價格：{len(fetch_plan):,} 檔")

    if not fetch_plan:
        print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
        return _finish_price_ensure(
            price_cache,
            persistent_price_cache,
            changed_price_codes,
            defer_save=defer_save,
        )

    changed_price_codes.update(
        fetch_price_plan_batch_first(
            price_cache,
            persistent_price_cache,
            fetch_plan,
            label="收盤價",
        )
    )

    print(f"  ✅ 共 {len(price_cache)} 支股票/權證收盤價")
    return _finish_price_ensure(
        price_cache,
        persistent_price_cache,
        changed_price_codes,
        defer_save=defer_save,
    )

# ══════════════════════════════════════════════════════════════════════
# D+1 ~ D+20
# ══════════════════════════════════════════════════════════════════════

def get_price_series_from_cache(price_cache, code):
    if not code:
        return {}

    code1 = str(code).strip()
    code2 = normalize_price_code(code1)
    code3 = code2.lstrip("0") if code2 else ""

    for c in [code1, code2, code3]:
        if c and c in price_cache:
            return price_cache.get(c, {})

    return {}


def price_dates_after(prices, base_date):
    base_date = normalize_date_str(base_date)
    return [d for d in sorted(prices.keys()) if d > base_date]


def build_dplus_dates(base_date, primary_prices, secondary_prices=None, limit=20):
    """
    建立 D+1 ~ D+20 的交易日序列。

    原本用「權證價格日期 + 標的價格日期」聯集，容易造成 D+ 日期亂跳；
    現在改成：
    1. 優先使用標的股價格日期，因為標的股交易日最完整。
    2. 標的股不足時，用權證價格日期補。
    3. 最後才用兩者聯集。
    """
    base_date = normalize_date_str(base_date)

    primary_dates = price_dates_after(primary_prices or {}, base_date)
    secondary_dates = price_dates_after(secondary_prices or {}, base_date)

    if len(primary_dates) >= limit:
        return primary_dates[:limit]

    merged = sorted(set(primary_dates + secondary_dates))
    return merged[:limit]


def get_price_on_or_before(prices, date):
    return get_price_nearest(prices, date)


def calc_pct_by_base(current_price, base_price):
    if current_price is None or base_price is None:
        return None

    try:
        current_price = float(current_price)
        base_price = float(base_price)

        if base_price <= 0 or current_price <= 0:
            return None

        return round((current_price - base_price) / base_price * 100, 2)
    except Exception:
        return None






def make_group_day_cells(ev, price_cache):
    day_values = []
    day_status = []

    u_prices = get_price_series_from_cache(price_cache, ev["標的股"])
    base_date = normalize_date_str(ev["結束日"])

    buy_u = get_price_on_or_before(u_prices, base_date) if u_prices else None

    future_dates = build_dplus_dates(
        base_date,
        primary_prices=u_prices,
        secondary_prices=None,
        limit=20
    )

    stopped = False

    for i, check_date in enumerate(future_dates[:20]):
        check_date = normalize_date_str(check_date)

        u_p = get_price_on_or_before(u_prices, check_date) if u_prices else None
        u_chg = calc_pct_by_base(u_p, buy_u)

        cell_text = f"標:{fmt_pct(u_chg)}"

        if ev["出清日"] == check_date:
            status = "exit"
        elif ev["減碼日"] == check_date:
            status = "reduce"
        elif u_chg is not None and u_chg > 0:
            status = "win"
        else:
            status = "lose"

        day_values.append(cell_text)
        day_status.append(status)

        if ev["出清日"] and check_date >= ev["出清日"]:
            for _ in range(20 - i - 1):
                day_values.append("")
                day_status.append("none")
            stopped = True
            break

    if not stopped:
        while len(day_values) < 20:
            day_values.append("")
            day_status.append("none")

    return day_values, day_status


def fmt_price_value(v):
    if v is None:
        return "-"

    try:
        v = float(v)
        if v <= 0:
            return "-"
        if v.is_integer():
            return int(v)
        return round(v, 2)
    except Exception:
        return "-"


def get_ma_status_cells(price_cache, underlying_code, base_date):
    """
    依照事件選出日計算標的股技術狀態：
    1. 標的股股價：選出日當天或之前最近一個交易日收盤價
    2. 5MA：若標的股股價 > 當天 5 日均線，顯示 ✓，否則空白
    3. 20MA：若當天 20 日均線 > 前一交易日 20 日均線，顯示 ✓，否則空白
    """
    prices = get_price_series_from_cache(price_cache, underlying_code)

    if not prices:
        return "-", "", ""

    base_date = normalize_date_str(base_date)
    valid_rows = []

    for d, p in prices.items():
        dt = parse_date(d)
        price = safe_price_float(p)

        if not dt or price is None:
            continue

        if normalize_date_str(d) <= base_date:
            valid_rows.append((normalize_date_str(d), price))

    valid_rows = sorted(valid_rows, key=lambda x: x[0])

    if not valid_rows:
        return "-", "", ""

    current_price = valid_rows[-1][1]
    close_values = [p for _, p in valid_rows]

    ma5_mark = ""
    ma20_mark = ""

    if len(close_values) >= 5:
        ma5 = sum(close_values[-5:]) / 5
        if current_price > ma5:
            ma5_mark = "✓"

    if len(close_values) >= 21:
        ma20_today = sum(close_values[-20:]) / 20
        ma20_prev = sum(close_values[-21:-1]) / 20
        if ma20_today > ma20_prev:
            ma20_mark = "✓"

    return fmt_price_value(current_price), ma5_mark, ma20_mark



# ══════════════════════════════════════════════════════════════════════
# TOP15 圖片用：固定部位明細與共識淨買超資料集
# ══════════════════════════════════════════════════════════════════════

def get_latest_price_info_on_or_before(price_cache, code, target_date):
    """使用 bisect 取得 target_date 當天或之前最近一筆有效收盤價。"""
    prices = get_price_series_from_cache(price_cache, code)
    if not prices:
        return None, ""

    target_str = normalize_date_str(target_date)
    valid_dates = _sorted_valid_price_dates(prices)
    idx = bisect_right(valid_dates, target_str) - 1
    if idx < 0:
        return None, ""
    date_key = valid_dates[idx]
    return safe_price_float(prices.get(date_key)), date_key



def _top15_find_item(item_map, broker_code, warrant_code):
    """用券商代號與權證代號的常見正規化形式尋找原始分點歷史項目。"""
    item_map = item_map or {}
    broker_code = str(broker_code or "").strip()
    warrant_code = str(warrant_code or "").strip()

    if not broker_code or not warrant_code:
        return None

    broker_candidates = []
    for value in [broker_code, broker_code.upper(), broker_code.lower()]:
        if value and value not in broker_candidates:
            broker_candidates.append(value)

    warrant_candidates = []
    for value in [
        warrant_code,
        normalize_warrant_code_for_unique(warrant_code),
        normalize_price_code(warrant_code),
    ]:
        value = str(value or "").strip()
        if value and value not in warrant_candidates:
            warrant_candidates.append(value)

    for bc in broker_candidates:
        for wc in warrant_candidates:
            item = item_map.get((bc, wc))
            if item:
                return item

    normalized_broker = broker_code.upper()
    normalized_warrant = normalize_warrant_code_for_unique(warrant_code)
    normalized_price_warrant = normalize_price_code(warrant_code)

    for (raw_broker, raw_warrant), item in item_map.items():
        if str(raw_broker or "").strip().upper() != normalized_broker:
            continue
        raw_warrant_text = str(raw_warrant or "").strip()
        if (
            normalize_warrant_code_for_unique(raw_warrant_text) == normalized_warrant
            or normalize_price_code(raw_warrant_text) == normalized_price_warrant
        ):
            return item

    return None


def _top15_observed_market_dates(item_map=None, price_cache=None, target_date=None):
    """
    從既有價格快取與分點歷史收集實際出現過的市場交易日。

    不用星期一到星期五硬推交易日，避免國定假日被誤算；價格日期優先，
    分點歷史日期作為補充。回傳日期皆不晚於 target_date。
    """
    target_dt = parse_date(target_date) or datetime.today()
    target_text = target_dt.strftime("%Y/%m/%d")
    dates = set()

    seen_series_ids = set()
    for prices in (price_cache or {}).values():
        if not isinstance(prices, dict):
            continue
        series_id = id(prices)
        if series_id in seen_series_ids:
            continue
        seen_series_ids.add(series_id)

        for raw_date, raw_price in prices.items():
            dt = parse_date(raw_date)
            if not dt or dt > target_dt:
                continue
            if safe_price_float(raw_price) is None:
                continue
            dates.add(dt.strftime("%Y/%m/%d"))

    seen_item_ids = set()
    for item in (item_map or {}).values():
        item_id = id(item)
        if item_id in seen_item_ids:
            continue
        seen_item_ids.add(item_id)

        df = item.get("df", pd.DataFrame()) if isinstance(item, dict) else pd.DataFrame()
        if df is None or df.empty or "日期" not in df.columns:
            continue

        for raw_date in df["日期"].tolist():
            dt = parse_date(raw_date)
            if dt and dt <= target_dt:
                dates.add(dt.strftime("%Y/%m/%d"))

    return sorted(d for d in dates if d <= target_text)


def _top15_merge_lot_metadata(existing, incoming):
    """合併指向同一筆原始每日買進流水的事件資訊，避免 C／D 等重疊事件重複加總。"""
    if existing is None:
        existing = dict(incoming)
        existing["_事件集合"] = {str(incoming.get("事件", "")).strip()} if str(incoming.get("事件", "")).strip() else set()
        existing["_事件類型集合"] = {str(incoming.get("事件類型", "")).strip()} if str(incoming.get("事件類型", "")).strip() else set()
        existing["_事件日集合"] = {normalize_date_str(incoming.get("事件日", ""))} if parse_date(incoming.get("事件日", "")) else set()
        existing["_來源集合"] = {str(incoming.get("來源", "")).strip()} if str(incoming.get("來源", "")).strip() else set()
        return existing

    event_code = str(incoming.get("事件", "")).strip()
    event_type = str(incoming.get("事件類型", "")).strip()
    event_date = normalize_date_str(incoming.get("事件日", ""))
    source_text = str(incoming.get("來源", "")).strip()

    if event_code:
        existing.setdefault("_事件集合", set()).add(event_code)
    if event_type:
        existing.setdefault("_事件類型集合", set()).add(event_type)
    if parse_date(event_date):
        existing.setdefault("_事件日集合", set()).add(event_date)
    if source_text:
        existing.setdefault("_來源集合", set()).add(source_text)

    # 同一分點、同一權證、同一買進日的原始流水只能算一次。
    # 若備援欄位與原始歷史數值略有差異，採較完整的較大值，不做加總。
    existing["原始股數"] = max(float(existing.get("原始股數", 0) or 0), float(incoming.get("原始股數", 0) or 0))
    existing["原始成本"] = max(float(existing.get("原始成本", 0) or 0), float(incoming.get("原始成本", 0) or 0))
    existing["剩餘股數"] = existing["原始股數"]
    existing["剩餘成本"] = existing["原始成本"]

    for field in ["分點", "分點名稱", "券商代號", "標的股", "標的名稱", "權證代號", "權證名稱"]:
        if not str(existing.get(field, "")).strip() and str(incoming.get(field, "")).strip():
            existing[field] = incoming.get(field, "")

    return existing


def _top15_finalize_lot_metadata(lot):
    lot = dict(lot)
    event_codes = sorted(x for x in lot.pop("_事件集合", set()) if x)
    event_types = sorted(x for x in lot.pop("_事件類型集合", set()) if x)
    event_dates = sorted(x for x in lot.pop("_事件日集合", set()) if parse_date(x))
    sources = sorted(x for x in lot.pop("_來源集合", set()) if x)

    lot["事件"] = "/".join(event_codes)
    lot["事件類型"] = "／".join(event_types)
    lot["事件日"] = event_dates[-1] if event_dates else normalize_date_str(lot.get("事件日", ""))
    lot["來源"] = "；".join(sources)
    return lot


def collect_top15_return_recent_dates(
    a_events,
    b_events,
    c_events,
    d_events,
    e_events=None,
    lookback_days=None,
    item_map=None,
    price_cache=None,
    target_date=None,
):
    """取得統計日前最近 N 個實際市場交易日，而不是最近 N 個事件發生日。"""
    if lookback_days is None:
        lookback_days = TOP15_LOOKBACK_TRADING_DAYS

    target_dt = parse_date(target_date) or datetime.today()
    dates = set(_top15_observed_market_dates(item_map, price_cache, target_dt))

    # 事件日期只作為快取資料不足時的補充，不再作為 40 日曆的唯一來源。
    for _, group_events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in group_events:
            dt = parse_date(ev.get("事件日") or ev.get("結束日") or ev.get("起始日") or ev.get("買進日"))
            if dt and dt <= target_dt:
                dates.add(dt.strftime("%Y/%m/%d"))

    recent_dates = sorted(dates, reverse=True)[:max(int(lookback_days), 1)]
    if len(recent_dates) < max(int(lookback_days), 1):
        print(
            f"  ⚠️ TOP15交易日曆僅取得 {len(recent_dates)} 個實際交易日，"
            f"少於設定的 {int(lookback_days)} 日；不以平日硬補國定假日。"
        )
    return recent_dates


def _top15_return_event_date(ev, is_a=False):
    if is_a:
        return normalize_date_str(ev.get("買進日") or ev.get("事件日") or "")
    return normalize_date_str(ev.get("事件日") or ev.get("結束日") or ev.get("起始日") or "")


def collect_top15_return_position_lots(a_events, b_events, c_events, d_events, e_events=None, recent_dates=None, item_map=None):
    """
    將 40 個市場交易日範圍內的 A/B/C/D/E 事件標記成 TOP15 事件 lot。

    同一筆「券商代號＋權證代號＋實際買進日」原始流水只建立一次；
    即使事件來源重疊，也只合併事件資訊，不重複增加股數與成本。
    """
    date_set = set(recent_dates or [])
    lot_map = {}
    item_map = item_map or {}

    def add_lot(
        event_code, event_type, broker_label, broker_name, broker_code,
        underlying_code, underlying_name, event_date, buy_date,
        warrant_code, warrant_name, buy_amount, buy_qty, source_text
    ):
        event_date = normalize_date_str(event_date)
        buy_date = normalize_date_str(buy_date or event_date)
        if not event_date or event_date not in date_set:
            return False

        warrant_code = normalize_warrant_code_for_unique(warrant_code)
        buy_amount = float(buy_amount or 0)
        buy_qty = float(buy_qty or 0)
        broker_code_text = str(broker_code or "").strip()
        if not broker_code_text or not warrant_code or buy_amount <= 0 or buy_qty <= 0:
            return False

        incoming = {
            "事件": str(event_code or "").strip(),
            "事件類型": str(event_type or "").strip(),
            "事件日": event_date,
            "買進日": buy_date,
            "分點": str(broker_label or "").strip(),
            "分點名稱": str(broker_name or "").strip(),
            "券商代號": broker_code_text,
            "標的股": str(underlying_code or "").strip(),
            "標的名稱": str(underlying_name or "").strip(),
            "權證代號": warrant_code,
            "權證名稱": str(warrant_name or "").strip(),
            "原始股數": buy_qty,
            "原始成本": buy_amount,
            "剩餘股數": buy_qty,
            "剩餘成本": buy_amount,
            "來源": source_text,
        }
        dedup_key = (broker_code_text.upper(), warrant_code, buy_date)
        lot_map[dedup_key] = _top15_merge_lot_metadata(lot_map.get(dedup_key), incoming)
        return True

    def add_group_lots_from_history(event_code, event_type, ev, lot, event_date, start_date, end_date):
        event_date = normalize_date_str(event_date)
        if not event_date or event_date not in date_set:
            return False

        warrant_code = normalize_warrant_code_for_unique(lot.get("權證代號", ""))
        if not warrant_code:
            return False

        start_date = normalize_date_str(start_date or event_date)
        end_date = normalize_date_str(end_date or start_date)
        item = _top15_find_item(item_map, ev.get("券商代號", ""), warrant_code)
        if not item:
            return False

        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            return False

        added = False
        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2.sort_values("日期").reset_index(drop=True)

        for row in df2.itertuples(index=False):
            row_dict = row._asdict()
            buy_date = normalize_date_str(row_dict.get("日期", ""))
            if not buy_date or buy_date < start_date or buy_date > end_date:
                continue

            buy_qty = float(row_dict.get("買進股數", 0) or 0)
            buy_amount = float(row_dict.get("買進金額", 0) or 0)
            if buy_qty <= 0 or buy_amount <= 0:
                continue

            added = add_lot(
                event_code, event_type,
                ev.get("分點", ""), ev.get("分點名稱", ""), ev.get("券商代號", ""),
                ev.get("標的股", ""), ev.get("標的名稱", ""),
                event_date, buy_date, warrant_code, lot.get("權證名稱", ""),
                buy_amount, buy_qty,
                f'{event_code} | {buy_date} | {warrant_code} {lot.get("權證名稱", "")}',
            ) or added

        return added

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            event_date = _top15_return_event_date(ev, is_a=False)
            event_type = ev.get("事件類型", f"{event_code}-事件")

            for lot in ev.get("lots", []):
                if event_code == "D":
                    buy_start_date = ev.get("起始日") or lot.get("買進日") or event_date
                    buy_end_date = ev.get("結束日") or event_date
                else:
                    buy_start_date = lot.get("買進日") or ev.get("事件日") or ev.get("結束日") or event_date
                    buy_end_date = buy_start_date

                used_history = add_group_lots_from_history(
                    event_code, event_type, ev, lot, event_date, buy_start_date, buy_end_date
                )
                if used_history:
                    continue

                add_lot(
                    event_code, event_type,
                    ev.get("分點", ""), ev.get("分點名稱", ""), ev.get("券商代號", ""),
                    ev.get("標的股", ""), ev.get("標的名稱", ""),
                    event_date, lot.get("買進日") or event_date,
                    lot.get("權證代號", ""), lot.get("權證名稱", ""),
                    lot.get("金額", 0), lot.get("股數", 0),
                    f'{event_code} | {lot.get("權證代號", "")} {lot.get("權證名稱", "")}',
                )

    return [
        _top15_finalize_lot_metadata(lot)
        for _, lot in sorted(lot_map.items(), key=lambda x: x[0])
    ]


def apply_sales_to_top15_return_lots(position_lots, item_map, target_date, window_start=None):
    """
    用完整可用分點歷史重建 FIFO，再只輸出仍未出清的 TOP15 事件 lot。

    每日順序固定為：
    1. 當日賣出先扣前一日以前的庫存。
    2. 尚未配對的賣出再抵銷當日買進。
    3. 當日剩餘買進才建立新 lot。

    因此 40 日以前的舊庫存會先被扣除，同日賣出也不會再被忽略。
    """
    if not position_lots:
        return position_lots

    target_dt = parse_date(target_date) or datetime.today()
    target_text = target_dt.strftime("%Y/%m/%d")
    window_start_text = normalize_date_str(window_start) if parse_date(window_start) else ""

    templates_by_key = {}
    for raw_lot in position_lots:
        lot = dict(raw_lot)
        broker_code = str(lot.get("券商代號", "")).strip()
        warrant_code = normalize_warrant_code_for_unique(lot.get("權證代號", ""))
        buy_date = normalize_date_str(lot.get("買進日") or lot.get("事件日", ""))
        if not broker_code or not warrant_code or not parse_date(buy_date):
            continue
        key = (broker_code.upper(), warrant_code)
        templates_by_key.setdefault(key, {})[buy_date] = lot

    rebuilt_event_lots = []

    for key, templates_by_date in templates_by_key.items():
        broker_code_upper, warrant_code = key
        sample_template = next(iter(templates_by_date.values()))
        broker_code = str(sample_template.get("券商代號", "")).strip()
        item = _top15_find_item(item_map, broker_code, warrant_code)

        if not item:
            for template in templates_by_date.values():
                fallback = dict(template)
                fallback["當日抵銷股數"] = 0.0
                fallback["歷史FIFO扣除股數"] = max(
                    float(fallback.get("原始股數", 0) or 0) - float(fallback.get("剩餘股數", 0) or 0),
                    0.0,
                )
                fallback["完整FIFO歷史起日"] = ""
                fallback["期初庫存股數"] = 0.0
                fallback["未配對賣出股數"] = 0.0
                fallback["FIFO完整狀態"] = "缺少分點歷史"
                rebuilt_event_lots.append(fallback)
            continue

        df = item.get("df", pd.DataFrame())
        if df is None or df.empty or "日期" not in df.columns:
            for template in templates_by_date.values():
                fallback = dict(template)
                fallback["當日抵銷股數"] = 0.0
                fallback["歷史FIFO扣除股數"] = 0.0
                fallback["完整FIFO歷史起日"] = ""
                fallback["期初庫存股數"] = 0.0
                fallback["未配對賣出股數"] = 0.0
                fallback["FIFO完整狀態"] = "分點歷史為空"
                rebuilt_event_lots.append(fallback)
            continue

        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2[df2["日期"].map(lambda x: bool(parse_date(x)) and x <= target_text)].copy()

        for col in ["買進股數", "賣出股數", "買進金額", "賣出金額"]:
            if col not in df2.columns:
                df2[col] = 0
            df2[col] = pd.to_numeric(df2[col], errors="coerce").fillna(0.0)

        daily_map = {}
        if not df2.empty:
            grouped = df2.groupby("日期", as_index=False)[["買進股數", "賣出股數", "買進金額", "賣出金額"]].sum()
            for row in grouped.itertuples(index=False):
                row_dict = row._asdict()
                date_text = normalize_date_str(row_dict.get("日期", ""))
                daily_map[date_text] = {
                    "買進股數": float(row_dict.get("買進股數", 0) or 0),
                    "賣出股數": float(row_dict.get("賣出股數", 0) or 0),
                    "買進金額": float(row_dict.get("買進金額", 0) or 0),
                    "賣出金額": float(row_dict.get("賣出金額", 0) or 0),
                    "合成事件列": False,
                }

        # 若舊快取缺少事件當日原始流水，才用事件 lot 建立合成買進列，並在稽核欄標記。
        for buy_date, template in templates_by_date.items():
            if buy_date not in daily_map or float(daily_map[buy_date].get("買進股數", 0) or 0) <= 0:
                daily_map[buy_date] = {
                    "買進股數": float(template.get("原始股數", 0) or 0),
                    "賣出股數": float(daily_map.get(buy_date, {}).get("賣出股數", 0) or 0),
                    "買進金額": float(template.get("原始成本", 0) or 0),
                    "賣出金額": float(daily_map.get(buy_date, {}).get("賣出金額", 0) or 0),
                    "合成事件列": True,
                }

        all_dates = sorted(d for d in daily_map if parse_date(d) and d <= target_text)
        history_start = all_dates[0] if all_dates else ""
        queue = []
        unmatched_sell_total = 0.0
        opening_inventory_qty = 0.0
        opening_captured = not bool(window_start_text)

        for date_text in all_dates:
            if not opening_captured and date_text >= window_start_text:
                opening_inventory_qty = sum(float(q.get("剩餘股數", 0) or 0) for q in queue)
                opening_captured = True

            day = daily_map[date_text]
            sell_qty_left = max(float(day.get("賣出股數", 0) or 0), 0.0)

            # 先扣前一日以前已存在的 FIFO 庫存。
            for qlot in queue:
                if sell_qty_left <= 0:
                    break
                remaining_qty = float(qlot.get("剩餘股數", 0) or 0)
                remaining_cost = float(qlot.get("剩餘成本", 0) or 0)
                if remaining_qty <= 0:
                    continue

                alloc_qty = min(sell_qty_left, remaining_qty)
                avg_cost = remaining_cost / remaining_qty if remaining_qty > 0 else 0.0
                alloc_cost = min(remaining_cost, alloc_qty * avg_cost)
                qlot["剩餘股數"] = max(remaining_qty - alloc_qty, 0.0)
                qlot["剩餘成本"] = max(remaining_cost - alloc_cost, 0.0)
                if qlot.get("_是TOP15事件lot"):
                    qlot["歷史FIFO扣除股數"] = float(qlot.get("歷史FIFO扣除股數", 0) or 0) + alloc_qty
                sell_qty_left -= alloc_qty

            buy_qty = max(float(day.get("買進股數", 0) or 0), 0.0)
            buy_amount = max(float(day.get("買進金額", 0) or 0), 0.0)

            # 前一日庫存不足時，剩餘賣出抵銷當日買進。
            same_day_offset_qty = min(sell_qty_left, buy_qty)
            residual_buy_qty = max(buy_qty - same_day_offset_qty, 0.0)
            residual_buy_cost = (
                buy_amount * residual_buy_qty / buy_qty
                if buy_qty > 0 and buy_amount > 0 and residual_buy_qty > 0
                else 0.0
            )
            sell_qty_left -= same_day_offset_qty

            if sell_qty_left > 0:
                unmatched_sell_total += sell_qty_left

            if residual_buy_qty <= 0 or residual_buy_cost <= 0:
                continue

            template = templates_by_date.get(date_text)
            is_event_lot = template is not None

            if template:
                qlot = dict(template)
                qlot["原始股數"] = residual_buy_qty
                qlot["原始成本"] = residual_buy_cost
                qlot["剩餘股數"] = residual_buy_qty
                qlot["剩餘成本"] = residual_buy_cost
                qlot["買進日"] = date_text
                qlot["當日抵銷股數"] = same_day_offset_qty
                qlot["歷史FIFO扣除股數"] = 0.0
                qlot["_合成事件列"] = bool(day.get("合成事件列"))
            else:
                qlot = {
                    "買進日": date_text,
                    "原始股數": residual_buy_qty,
                    "原始成本": residual_buy_cost,
                    "剩餘股數": residual_buy_qty,
                    "剩餘成本": residual_buy_cost,
                    "當日抵銷股數": 0.0,
                    "歷史FIFO扣除股數": 0.0,
                    "_合成事件列": False,
                }

            qlot["_是TOP15事件lot"] = is_event_lot
            queue.append(qlot)

        if not opening_captured:
            opening_inventory_qty = sum(float(q.get("剩餘股數", 0) or 0) for q in queue)

        synthetic_used = any(bool(q.get("_合成事件列")) for q in queue if q.get("_是TOP15事件lot"))
        if synthetic_used:
            fifo_status = "歷史缺少事件買進列"
        elif unmatched_sell_total > 0:
            fifo_status = "歷史起點前可能有庫存"
        else:
            fifo_status = "OK"

        for qlot in queue:
            if not qlot.get("_是TOP15事件lot"):
                continue
            if float(qlot.get("剩餘股數", 0) or 0) <= 0 or float(qlot.get("剩餘成本", 0) or 0) <= 0:
                continue

            qlot.pop("_是TOP15事件lot", None)
            qlot.pop("_合成事件列", None)
            qlot["完整FIFO歷史起日"] = history_start
            qlot["期初庫存股數"] = opening_inventory_qty
            qlot["未配對賣出股數"] = unmatched_sell_total
            qlot["FIFO完整狀態"] = fifo_status
            rebuilt_event_lots.append(qlot)

    rebuilt_event_lots.sort(
        key=lambda x: (
            str(x.get("券商代號", "")).upper(),
            str(x.get("權證代號", "")),
            normalize_date_str(x.get("買進日", "")),
        )
    )
    return rebuilt_event_lots



_TOP15_LP_QUOTE_CACHE = {}
_TOP15_LP_QUOTE_CACHE_LOCK = threading.Lock()
_TOP15_LP_USER_AGENT = "Mozilla/5.0"


def _top15_lp_normalize_field_name(value):
    s = str(value or "").strip().upper()
    s = s.replace("％", "%")
    return re.sub(r"[\s_\-－—–/\\()（）\[\]【】:：｜|]+", "", s)


def _top15_lp_parse_positive_price(value):
    try:
        if value is None:
            return None
        s = str(value).replace(",", "").replace("元", "").strip()
        s = re.sub(r"[^0-9.+-]", "", s)
        if not s:
            return None
        v = float(s)
        return v if v > 0 else None
    except Exception:
        return None


def _top15_lp_extract_date_from_text(text):
    s = str(text or "")
    patterns = [
        r"(?:資料日期|報價日期|交易日期|日期)\s*[:：]?\s*(\d{4})[/-](\d{1,2})[/-](\d{1,2})",
        r"(?:資料日期|報價日期|交易日期|日期)\s*[:：]?\s*(\d{3})[/-](\d{1,2})[/-](\d{1,2})",
    ]
    for idx, pattern in enumerate(patterns):
        match = re.search(pattern, s)
        if not match:
            continue
        year, month, day = map(int, match.groups())
        if idx == 1 or year < 1911:
            year += 1911
        try:
            return f"{year:04d}/{month:02d}/{day:02d}"
        except Exception:
            continue
    return ""


def _top15_lp_price_field_score(field_name):
    """挑選 LP 最佳一檔委買價格，排除委買量、隱含波動率等非價格欄。"""
    name = _top15_lp_normalize_field_name(field_name)
    if not name:
        return -1
    if any(token in name for token in ("IV", "隱含", "波動", "委買量", "買進量", "買量", "張數", "筆數", "數量")):
        return -1

    exact_priority = [
        "最佳一檔委託買進價格", "最佳一檔委買價格", "最佳委託買進價格",
        "最佳委買價格", "流動量提供者委買價格", "流動量提供者買進價格",
        "委託買進價格", "買1價", "買一價", "委買價", "買進價", "買價",
    ]
    for rank, candidate in enumerate(_top15_lp_normalize_field_name(x) for x in exact_priority):
        if name == candidate:
            return 1000 - rank

    score = 0
    if "流動量提供者" in name or name.startswith("LP"):
        score += 100
    if "最佳" in name or "一檔" in name or "買1" in name or "買一" in name:
        score += 50
    if "委買" in name or "委託買進" in name or "買進價格" in name or name.endswith("買價"):
        score += 30
    if "價" in name or "價格" in name:
        score += 10
    return score if score >= 40 else -1


def _top15_lp_code_matches(value, warrant_code):
    raw = str(value or "").strip()
    wanted = normalize_price_code(warrant_code)
    if not raw or not wanted:
        return False
    candidates = set(re.findall(r"(?<!\d)\d{4,6}(?!\d)", raw))
    for candidate in candidates:
        if normalize_price_code(candidate) == wanted:
            return True
    return normalize_price_code(raw) == wanted


def _top15_lp_extract_from_record(record, warrant_code):
    if not isinstance(record, dict) or not record:
        return None

    code_fields = []
    for key, value in record.items():
        key_name = _top15_lp_normalize_field_name(key)
        if any(token in key_name for token in ("權證代號", "證券代號", "商品代號", "權證代碼", "代號", "CODE")):
            code_fields.append(value)

    if code_fields and not any(_top15_lp_code_matches(value, warrant_code) for value in code_fields):
        return None

    candidates = []
    for key, value in record.items():
        score = _top15_lp_price_field_score(key)
        if score < 0:
            continue
        price = _top15_lp_parse_positive_price(value)
        if price is not None:
            candidates.append((score, price, str(key)))

    if not candidates:
        return None
    candidates.sort(key=lambda item: item[0], reverse=True)
    score, price, field = candidates[0]
    return {"price": price, "field": field}


def _top15_lp_walk_json(payload, warrant_code):
    if isinstance(payload, dict):
        hit = _top15_lp_extract_from_record(payload, warrant_code)
        if hit:
            return hit
        for value in payload.values():
            hit = _top15_lp_walk_json(value, warrant_code)
            if hit:
                return hit
    elif isinstance(payload, list):
        for value in payload:
            hit = _top15_lp_walk_json(value, warrant_code)
            if hit:
                return hit
    return None


def _top15_lp_parse_html_tables(text, warrant_code):
    try:
        tables = pd.read_html(StringIO(text))
    except Exception:
        tables = []

    for table in tables:
        if table is None or table.empty:
            continue
        df = table.copy()
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [
                " ".join(str(x) for x in col if str(x).strip() and str(x).lower() != "nan").strip()
                for col in df.columns
            ]
        else:
            df.columns = [str(x).strip() for x in df.columns]

        rows = df.fillna("").to_dict("records")
        matched_rows = [row for row in rows if _top15_lp_code_matches(" ".join(str(v) for v in row.values()), warrant_code)]
        search_rows = matched_rows or (rows if len(rows) == 1 else [])
        for row in search_rows:
            hit = _top15_lp_extract_from_record(row, warrant_code)
            if hit:
                return hit
    return None


def _top15_lp_parse_response(response, warrant_code, target_date):
    text = response.text or ""
    quote_date = _top15_lp_extract_date_from_text(text)
    try:
        payload = response.json()
    except Exception:
        payload = None

    hit = _top15_lp_walk_json(payload, warrant_code) if payload is not None else None
    if not hit:
        hit = _top15_lp_parse_html_tables(text, warrant_code)
    if not hit:
        pattern = (
            r"(?:最佳一檔委託買進價格|最佳一檔委買價格|最佳委買價格|流動量提供者委買價格|委買價|買1價|買一價)"
            r"[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)"
        )
        match = re.search(pattern, text, flags=re.I)
        if match:
            price = _top15_lp_parse_positive_price(match.group(1))
            if price is not None:
                hit = {"price": price, "field": "頁面委買價"}

    if not hit:
        return None

    # fetch_top15_lp_quote_map() 只會在統計日等於最新交易日時呼叫本函式。
    # 部分平台回應有委買價但沒有獨立日期欄；舊版在這裡直接 return None，
    # 造成 LP 明明抓到價格仍被當成缺價。現在可透過設定允許以統計日補上日期。
    if not quote_date:
        if not TOP15_LP_ALLOW_IMPLICIT_LATEST_DATE:
            return None
        quote_date = normalize_date_str(target_date)

    if normalize_date_str(quote_date) != normalize_date_str(target_date):
        return None
    hit["date"] = normalize_date_str(quote_date)
    return hit


def _top15_lp_accept_disclaimer(session, base_url):
    try:
        root = re.match(r"^(https?://[^/]+)", base_url)
        root_url = root.group(1) + "/" if root else base_url
        response = session.get(
            root_url,
            headers={"User-Agent": _TOP15_LP_USER_AGENT},
            timeout=(5, TOP15_LP_QUOTE_TIMEOUT_SECONDS),
        )
        if response.status_code >= 400 or "BTNConfirm" not in response.text:
            return

        hidden = {}
        for tag in re.findall(r"<input\b[^>]*>", response.text, flags=re.I):
            name_match = re.search(r"\bname=[\"']([^\"']+)[\"']", tag, flags=re.I)
            if not name_match:
                continue
            value_match = re.search(r"\bvalue=[\"']([^\"']*)[\"']", tag, flags=re.I)
            hidden[name_match.group(1)] = value_match.group(1) if value_match else ""

        hidden["agree"] = "on"
        hidden["BTNConfirm"] = "確認"
        session.post(
            response.url,
            data=hidden,
            headers={"User-Agent": _TOP15_LP_USER_AGENT, "Referer": response.url},
            timeout=(5, TOP15_LP_QUOTE_TIMEOUT_SECONDS),
        )
    except Exception:
        return


def fetch_top15_lp_quote_price(warrant_code, target_date):
    """統計日無成交價時，取得權證流動量提供者最佳委買價。"""
    if not TOP15_LP_QUOTE_FALLBACK_ENABLED or not TOP15_LP_QUOTE_URLS:
        return None

    code = normalize_price_code(warrant_code)
    target_date = normalize_date_str(target_date)
    if not code or not target_date:
        return None

    cache_key = (code, target_date)
    with _TOP15_LP_QUOTE_CACHE_LOCK:
        if cache_key in _TOP15_LP_QUOTE_CACHE:
            cached = _TOP15_LP_QUOTE_CACHE[cache_key]
            return dict(cached) if isinstance(cached, dict) else None

    params = {"stockNo": code, "duration1": 0, "duration2": 730, "Period": 14}
    headers = {
        "User-Agent": _TOP15_LP_USER_AGENT,
        "Accept": "application/json,text/html,application/xhtml+xml,*/*",
    }
    result = None

    for url in TOP15_LP_QUOTE_URLS:
        session = requests.Session()
        try:
            _top15_lp_accept_disclaimer(session, url)
            responses = []
            try:
                responses.append(session.get(
                    url, params=params, headers=headers,
                    timeout=(5, TOP15_LP_QUOTE_TIMEOUT_SECONDS), allow_redirects=True,
                ))
            except Exception:
                pass
            try:
                responses.append(session.post(
                    url, data=params, headers=headers,
                    timeout=(5, TOP15_LP_QUOTE_TIMEOUT_SECONDS), allow_redirects=True,
                ))
            except Exception:
                pass

            for response in responses:
                if response is None or response.status_code >= 400:
                    continue
                hit = _top15_lp_parse_response(response, code, target_date)
                if hit:
                    result = {
                        "price": float(hit["price"]),
                        "date": hit.get("date") or target_date,
                        "source": "權證資訊揭露平台LP委買價",
                        "field": hit.get("field", ""),
                        "url": response.url,
                    }
                    break
            if result:
                break
        finally:
            try:
                session.close()
            except Exception:
                pass

    with _TOP15_LP_QUOTE_CACHE_LOCK:
        _TOP15_LP_QUOTE_CACHE[cache_key] = dict(result) if result else False
    return result


def fetch_top15_lp_quote_map(warrant_codes, target_date):
    codes = sorted({normalize_price_code(code) for code in warrant_codes if normalize_price_code(code)})
    if not TOP15_LP_QUOTE_FALLBACK_ENABLED or not TOP15_LP_QUOTE_URLS or not codes:
        return {}

    normalized_target_date = normalize_date_str(target_date)
    latest_market_date = normalize_date_str(resolve_latest_trading_date_on_or_before(datetime.today()))
    if normalized_target_date != latest_market_date:
        print(f"  TOP15 LP 報價略過：統計日 {normalized_target_date} 不是最新交易日 {latest_market_date}")
        return {}

    print(f"  TOP15統計日無成交價，查詢權證資訊揭露平台 LP 報價：{len(codes):,} 檔")
    result = {}
    done = 0
    with ThreadPoolExecutor(max_workers=min(TOP15_LP_QUOTE_WORKERS, len(codes))) as executor:
        futures = {executor.submit(fetch_top15_lp_quote_price, code, target_date): code for code in codes}
        for future in as_completed(futures):
            code = futures[future]
            done += 1
            try:
                quote = future.result()
            except Exception:
                quote = None
            if quote and top15_safe_float(quote.get("price"), 0) > 0:
                result[code] = quote
            if done % 20 == 0:
                print(f"  [{done}/{len(codes)}] TOP15 LP 報價查詢中...")

    print(f"  TOP15 LP 報價查詢完成：取得 {len(result):,}/{len(codes):,} 檔")
    return result


def ensure_top15_return_warrant_prices(
    price_cache,
    position_lots,
    target_date,
    persistent_price_cache=None,
    defer_save=False,
):
    """
    TOP15 固定資料集需要權證目前價格。

    原本 fetch_all_prices() 為了加速，B/C/D 預設只抓標的股價格，
    因此這裡會針對目前仍有剩餘部位的權證補抓最新價格。
    可由主流程傳入共用 persistent_price_cache 並延後儲存，確保整次流程只讀一次、只寫一次。
    """
    if not position_lots:
        return _finish_price_ensure(
            price_cache, persistent_price_cache, set(), defer_save=defer_save
        )

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()

    target_dt = min(target_dt, datetime.today())
    start_dt = target_dt - timedelta(days=max(TOP15_PRICE_LOOKBACK_DAYS, 10))

    needed_codes = sorted({
        normalize_price_code(lot.get("權證代號", ""))
        for lot in position_lots
        if float(lot.get("剩餘成本", 0) or 0) > 0
        and normalize_price_code(lot.get("權證代號", ""))
    })

    if not needed_codes:
        return _finish_price_ensure(
            price_cache, persistent_price_cache, set(), defer_save=defer_save
        )

    if persistent_price_cache is None:
        persistent_price_cache = load_price_cache()

    fetch_plan = []
    fetch_start_by_code = {}
    changed_price_codes = set()

    for code in needed_codes:
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        current_prices = get_price_series_from_cache(price_cache, code)
        merged_prices = merge_price_dicts(cached_prices, current_prices)

        if merged_prices:
            add_price_aliases(price_cache, code, merged_prices)
            persistent_price_cache[normalize_price_code(code)] = merged_prices

        latest_price, latest_date = get_latest_price_info_on_or_before(
            price_cache,
            code,
            target_dt.strftime("%Y/%m/%d"),
        )
        latest_dt = parse_date(latest_date) if latest_date else None

        target_date_text = target_dt.strftime("%Y/%m/%d")
        latest_date_text = normalize_date_str(latest_date) if latest_date else ""

        need_fetch = latest_price is None
        if TOP15_REQUIRE_TARGET_DATE_PRICE:
            # 統計日報酬率只能使用統計日當日價格；快取只有較早日期時必須補抓。
            if latest_date_text != target_date_text:
                need_fetch = True
        elif latest_dt and (target_dt - latest_dt).days > TOP15_PRICE_STALE_DAYS:
            need_fetch = True

        if need_fetch:
            fetch_plan.append(code)
            # 已有 75 日內可用歷史價時，只補抓統計日前短窗口；
            # 完全沒有可用備援價時才抓完整 TOP15_PRICE_LOOKBACK_DAYS。
            has_recent_fallback = bool(
                latest_dt
                and latest_price is not None
                and (target_dt - latest_dt).days <= TOP15_LAST_TRADE_FALLBACK_MAX_DAYS
            )
            short_days = max(int(os.getenv("TOP15_TARGET_PRICE_REFETCH_DAYS", "12")), 5)
            fetch_start_by_code[code] = (
                target_dt - timedelta(days=short_days)
                if has_recent_fallback
                else start_dt
            )

    print(f"  TOP15固定資料集需檢查權證價格：{len(needed_codes):,} 檔")
    print(f"  TOP15固定資料集需補抓權證價格：{len(fetch_plan):,} 檔")

    if not fetch_plan:
        return _finish_price_ensure(
            price_cache,
            persistent_price_cache,
            changed_price_codes,
            defer_save=defer_save,
        )

    top15_fetch_plan = {
        code: [fetch_start_by_code.get(code, start_dt), target_dt]
        for code in fetch_plan
    }
    changed_price_codes.update(
        fetch_price_plan_batch_first(
            price_cache,
            persistent_price_cache,
            top15_fetch_plan,
            label="TOP15固定資料集權證價格",
        )
    )

    return _finish_price_ensure(
        price_cache,
        persistent_price_cache,
        changed_price_codes,
        defer_save=defer_save,
    )


_TRADING_DATE_RESOLUTION_CACHE = {}
_OFFICIAL_TRADING_DATE_STATUS_CACHE = {}


def _market_rows_contain_records(rows):
    if not isinstance(rows, list):
        return False

    for row in rows:
        if isinstance(row, dict) and row:
            return True
        if isinstance(row, (list, tuple)) and len(row) >= 2:
            return True

    return False


def _market_payload_has_trading_rows(payload):
    """判斷 TWSE／TPEx 單日行情回應是否含有實際交易資料列。"""
    if not isinstance(payload, (dict, list)):
        return False

    if isinstance(payload, dict):
        for key, value in payload.items():
            key_text = str(key or "").strip().lower()

            if re.fullmatch(r"data\d*", key_text) and _market_rows_contain_records(value):
                return True

            if key_text in {"tables", "table"} and isinstance(value, list):
                for table in value:
                    if not isinstance(table, dict):
                        continue
                    rows = table.get("data") or table.get("aaData") or table.get("rows")
                    if _market_rows_contain_records(rows):
                        return True

            if isinstance(value, (dict, list)) and _market_payload_has_trading_rows(value):
                return True

        return False

    for value in payload:
        if isinstance(value, (dict, list)) and _market_payload_has_trading_rows(value):
            return True

    return False


def _official_market_has_trading_data(target_dt):
    """
    用 TWSE／TPEx 官方盤後行情確認指定日期是否為已有完整資料的交易日。

    回傳：
    - True：至少一個官方市場已確認有當日交易資料。
    - False：官方市場成功回應，但指定日期沒有交易資料。
    - None：官方來源皆連線或解析失敗，交由本機快取備援。
    """
    if not target_dt:
        return None

    target_dt = datetime(target_dt.year, target_dt.month, target_dt.day)
    target_key = target_dt.strftime("%Y/%m/%d")

    if target_key in _OFFICIAL_TRADING_DATE_STATUS_CACHE:
        return _OFFICIAL_TRADING_DATE_STATUS_CACHE[target_key]

    if target_dt.weekday() >= 5:
        _OFFICIAL_TRADING_DATE_STATUS_CACHE[target_key] = False
        return False

    endpoints = [
        (
            "https://www.twse.com.tw/exchangeReport/MI_INDEX",
            {
                "response": "json",
                "date": target_dt.strftime("%Y%m%d"),
                "type": "ALL",
            },
        ),
        (
            "https://www.tpex.org.tw/www/zh-tw/afterTrading/otc",
            {
                "date": target_dt.strftime("%Y/%m/%d"),
                "response": "json",
            },
        ),
    ]

    successful_results = []

    for url, params in endpoints:
        try:
            session = get_thread_session()
            response = session.get(
                url,
                params=params,
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=(5, 15),
            )
            response.raise_for_status()

            try:
                payload = response.json()
            except Exception:
                payload = json.loads(response.content.decode("utf-8"))

            has_rows = _market_payload_has_trading_rows(payload)
            successful_results.append(bool(has_rows))

            if has_rows:
                _OFFICIAL_TRADING_DATE_STATUS_CACHE[target_key] = True
                return True
        except Exception:
            continue

    if successful_results:
        _OFFICIAL_TRADING_DATE_STATUS_CACHE[target_key] = False
        return False

    return None



def _latest_cached_trading_date_on_or_before(target_dt):
    """官方行情無法使用時，從既有價格／FinMind 分點歷史快取找最近資料日。"""
    if not target_dt:
        return None

    candidates = []
    for path_value in (PRICE_CACHE_PATH, HISTORY_CACHE_PATH):
        try:
            df = read_cache_csv(path_value)
        except Exception:
            continue
        if df is None or df.empty or "日期" not in df.columns:
            continue

        parsed = pd.to_datetime(
            df["日期"].astype(str).str.replace("/", "-", regex=False),
            errors="coerce",
        ).dropna()
        parsed = parsed[parsed <= pd.Timestamp(target_dt)]
        if not parsed.empty:
            latest = parsed.max().to_pydatetime()
            candidates.append(datetime(latest.year, latest.month, latest.day))

    return max(candidates) if candidates else None



def resolve_latest_trading_date_on_or_before(target_date=None):
    """
    將指定日期解析成「該日或之前最近一個已有官方盤後資料的交易日」。

    - 假日／週末執行會自動回退到最近交易日。
    - 當日盤後資料尚未發布時，會使用前一個已完成交易日。
    - 官方來源暫時失敗時，改由既有價格／分點歷史快取判斷。
    """
    if isinstance(target_date, datetime):
        requested_dt = target_date
    elif hasattr(target_date, "to_pydatetime"):
        try:
            requested_dt = target_date.to_pydatetime()
        except Exception:
            requested_dt = parse_date(target_date) or datetime.today()
    else:
        requested_dt = parse_date(target_date) or datetime.today()

    today_dt = datetime.today()

    if requested_dt.date() > today_dt.date():
        requested_dt = today_dt

    requested_dt = datetime(requested_dt.year, requested_dt.month, requested_dt.day)
    requested_key = requested_dt.strftime("%Y/%m/%d")

    if requested_key in _TRADING_DATE_RESOLUTION_CACHE:
        return _TRADING_DATE_RESOLUTION_CACHE[requested_key]

    resolved_dt = None

    # 31 個日曆日足以涵蓋春節等連續休市區間。
    for offset in range(32):
        candidate_dt = requested_dt - timedelta(days=offset)
        if candidate_dt.weekday() >= 5:
            continue

        market_status = _official_market_has_trading_data(candidate_dt)
        if market_status is True:
            resolved_dt = candidate_dt
            break

    if resolved_dt is None:
        resolved_dt = _latest_cached_trading_date_on_or_before(requested_dt)

    if resolved_dt is None:
        # 官方與本機快取都暫時不可用時，至少排除星期六、星期日。
        resolved_dt = requested_dt
        while resolved_dt.weekday() >= 5:
            resolved_dt -= timedelta(days=1)

    resolved_text = resolved_dt.strftime("%Y/%m/%d")
    _TRADING_DATE_RESOLUTION_CACHE[requested_key] = resolved_text

    if resolved_text != requested_key:
        print(f"  📅 統計日自動回退：{requested_key} → {resolved_text}（最近可用交易日）")

    return resolved_text


def normalize_top15_target_date(target_date=None):
    raw = str(target_date or TOP15_TARGET_DATE or "").strip()

    if raw:
        dt = parse_date(raw)
        if not dt:
            raise RuntimeError(f"TOP15_TARGET_DATE 格式錯誤，請使用 YYYY/MM/DD 或 YYYY-MM-DD：{raw}")
        return resolve_latest_trading_date_on_or_before(dt)

    return resolve_latest_trading_date_on_or_before(datetime.today())


def top15_safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        s = str(value).replace(",", "").strip()
        if s in ("", "-", "None", "nan", "null"):
            return default
        return float(s)
    except Exception:
        return default


def top15_fmt_amount_wan(value):
    try:
        return f"{float(value) / 10000:.1f}萬"
    except Exception:
        return "0.0萬"


def build_top15_position_detail_and_consensus_rows(
    a_events,
    b_events,
    c_events,
    d_events,
    e_events=None,
    item_map=None,
    price_cache=None,
    target_date=None,
    data_scope=None,
    allow_price_fetch=True,
    persistent_price_cache=None,
    defer_price_save=False,
    price_changed_codes=None,
):
    """
    建立 TOP15 圖片用固定資料集。

    產出兩份資料：
    1. 快取_TOP15部位明細：一列代表一筆實際買進 lot，包含原始成本、賣出扣減、剩餘成本、價格與報酬率。
    2. 快取_TOP15共識淨買超：由部位明細加總而成，一列代表一檔標的股，圖片程式可直接讀取排名。

    嚴格規則：
    - 價格日期若距離估值日超過 TOP15_PRICE_STALE_DAYS，視為缺價格。
    - 若剩餘部位缺少有效權證價格，預設保留淨買超成本，但該筆部位不納入報酬率估算。
    - TOP15 總表完全由部位明細加總，不再另外重算。
    """
    item_map = item_map or {}
    if price_cache is None:
        price_cache = {}
    scope = str(data_scope or get_result_data_scope()).strip()

    if not TOP15_CACHE_ENABLED:
        return [], []

    print("【Step 4b】建立 TOP15 圖片用固定資料集...")

    target_date = normalize_top15_target_date(target_date)
    target_dt = parse_date(target_date)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    recent_dates = collect_top15_return_recent_dates(
        a_events,
        b_events,
        c_events,
        d_events,
        e_events,
        TOP15_LOOKBACK_TRADING_DAYS,
        item_map=item_map,
        price_cache=price_cache,
        target_date=target_date,
    )

    if not recent_dates:
        print("  ⚠️ TOP15固定資料集：沒有可用事件日期")
        return [], []

    period_text = f"{min(recent_dates)} ～ {max(recent_dates)}"

    position_lots = collect_top15_return_position_lots(
        a_events, b_events, c_events, d_events, e_events, recent_dates, item_map=item_map
    )

    if not position_lots:
        print("  ⚠️ TOP15固定資料集：沒有可用買超 lot")
        return [], []

    position_lots = apply_sales_to_top15_return_lots(
        position_lots,
        item_map,
        target_date,
        window_start=min(recent_dates),
    )
    position_lots = [
        lot for lot in position_lots
        if top15_safe_float(lot.get("剩餘成本", 0)) > 0 and top15_safe_float(lot.get("剩餘股數", 0)) > 0
    ]

    if not position_lots:
        print("  ⚠️ TOP15固定資料集：扣除賣出後沒有剩餘部位")
        return [], []

    if allow_price_fetch:
        if defer_price_save:
            price_cache, changed_codes = ensure_top15_return_warrant_prices(
                price_cache,
                position_lots,
                target_date,
                persistent_price_cache=persistent_price_cache,
                defer_save=True,
            )
            if price_changed_codes is not None:
                price_changed_codes.update(changed_codes)
        else:
            ensure_top15_return_warrant_prices(
                price_cache,
                position_lots,
                target_date,
                persistent_price_cache=persistent_price_cache,
                defer_save=False,
            )

    # 僅針對統計日沒有成交收盤價的權證查詢 LP；有當日成交價者不查平台。
    lp_quote_needed_codes = set()
    for lot in position_lots:
        warrant_code = normalize_price_code(lot.get("權證代號", ""))
        if not warrant_code:
            continue
        traded_price, traded_price_date = get_latest_price_info_on_or_before(
            price_cache,
            warrant_code,
            target_date,
        )
        if traded_price is None or normalize_date_str(traded_price_date) != target_date:
            lp_quote_needed_codes.add(warrant_code)

    lp_quote_map = fetch_top15_lp_quote_map(lp_quote_needed_codes, target_date)

    detail_rows = []
    validation_errors = []
    missing_price_rows = []

    for lot in position_lots:
        original_qty = top15_safe_float(lot.get("原始股數", 0))
        original_cost = top15_safe_float(lot.get("原始成本", 0))
        remaining_qty = top15_safe_float(lot.get("剩餘股數", 0))
        remaining_cost = top15_safe_float(lot.get("剩餘成本", 0))

        if original_qty <= 0 or original_cost <= 0 or remaining_qty <= 0 or remaining_cost <= 0:
            continue

        sold_qty = max(original_qty - remaining_qty, 0)
        sold_cost = max(original_cost - remaining_cost, 0)
        original_avg = original_cost / original_qty if original_qty else None
        remaining_avg = remaining_cost / remaining_qty if remaining_qty else None

        if remaining_qty - original_qty > 0.0001:
            validation_errors.append(
                f"剩餘股數大於原始股數：{lot.get('分點')} {lot.get('權證代號')} 原始={original_qty} 剩餘={remaining_qty}"
            )

        if remaining_cost - original_cost > 1:
            validation_errors.append(
                f"剩餘成本大於原始成本：{lot.get('分點')} {lot.get('權證代號')} 原始={original_cost} 剩餘={remaining_cost}"
            )

        warrant_code = normalize_price_code(lot.get("權證代號", ""))
        latest_price, latest_price_date = get_latest_price_info_on_or_before(
            price_cache,
            warrant_code,
            target_date,
        )

        price_status = "OK"
        valuation_price_source = ""
        latest_dt = parse_date(latest_price_date) if latest_price_date else None
        has_target_date_trade_price = bool(
            latest_price is not None
            and latest_price_date
            and normalize_date_str(latest_price_date) == target_date
        )

        if has_target_date_trade_price:
            valuation_price_source = "當日成交價"
        else:
            lp_quote = lp_quote_map.get(warrant_code)
            if lp_quote and top15_safe_float(lp_quote.get("price"), 0) > 0:
                latest_price = top15_safe_float(lp_quote.get("price"), None)
                latest_price_date = normalize_date_str(lp_quote.get("date") or target_date)
                latest_dt = parse_date(latest_price_date) if latest_price_date else None
                valuation_price_source = "LP流動量提供者委買價"
                price_status = "OK"
            elif latest_price is None:
                price_status = "缺價格"
            else:
                fallback_age_days = (target_dt - latest_dt).days if latest_dt and target_dt else None
                can_use_last_trade_fallback = bool(
                    TOP15_LAST_TRADE_FALLBACK_ENABLED
                    and latest_price is not None
                    and fallback_age_days is not None
                    and fallback_age_days >= 0
                    and (
                        TOP15_LAST_TRADE_FALLBACK_MAX_DAYS <= 0
                        or fallback_age_days <= TOP15_LAST_TRADE_FALLBACK_MAX_DAYS
                    )
                )

                if can_use_last_trade_fallback:
                    valuation_price_source = "最近成交價備援"
                    price_status = "OK"
                elif TOP15_REQUIRE_TARGET_DATE_PRICE:
                    # 超過備援天數時仍維持嚴格缺價，不拿過舊價格估值。
                    price_status = "非統計日價格"
                    latest_price = None
                elif latest_dt and target_dt and (target_dt - latest_dt).days > TOP15_PRICE_STALE_DAYS:
                    price_status = "價格過舊"
                    latest_price = None
                else:
                    valuation_price_source = "最近成交價"

        if latest_price is None:
            market_value = ""
            unrealized_pnl = ""
            return_pct = ""
            return_text = "-"
            if price_status == "缺價格":
                price_status = "未造市不計報酬率"
            elif price_status == "非統計日價格":
                price_status = "無統計日價格不計報酬率"
            elif price_status == "價格過舊":
                price_status = "價格過舊不計報酬率"

            missing_price_rows.append(
                f"TOP15剩餘部位不計報酬率：{lot.get('分點')} {lot.get('標的股')} {lot.get('權證代號')} {lot.get('權證名稱')}，剩餘成本={round(remaining_cost, 0)}，原因={price_status}"
            )

            if not TOP15_EXCLUDE_MISSING_PRICE_FROM_RETURN:
                validation_errors.append(
                    f"TOP15剩餘部位缺價格：{lot.get('分點')} {lot.get('標的股')} {lot.get('權證代號')} {lot.get('權證名稱')}，剩餘成本={round(remaining_cost, 0)}"
                )
        else:
            market_value_float = remaining_qty * latest_price
            unrealized_pnl_float = market_value_float - remaining_cost
            return_pct_float = unrealized_pnl_float / remaining_cost * 100 if remaining_cost > 0 else None
            market_value = round(market_value_float, 0)
            unrealized_pnl = round(unrealized_pnl_float, 0)
            return_pct = "" if return_pct_float is None else round(return_pct_float, 2)
            detail_valuation_symbol = (
                "*"
                if return_pct_float is not None and valuation_price_source != "當日成交價"
                else ""
            )
            return_text = (
                "-"
                if return_pct_float is None
                else f"{return_pct_float:+.2f}%{detail_valuation_symbol}"
            )

        if latest_price is None:
            detail_valuation_symbol = ""

        detail_rows.append({
            "資料範圍": scope,
            "統計日期": target_date,
            "統計期間": period_text,
            "有效交易日數": len(recent_dates),
            "交易日區間起日": min(recent_dates),
            "交易日區間迄日": max(recent_dates),
            "事件是否位於40日內": "是",
            "完整FIFO歷史起日": normalize_date_str(lot.get("完整FIFO歷史起日", "")),
            "期初庫存股數": round(top15_safe_float(lot.get("期初庫存股數", 0)), 0),
            "同日抵銷股數": round(top15_safe_float(lot.get("當日抵銷股數", 0)), 0),
            "歷史FIFO扣除股數": round(top15_safe_float(lot.get("歷史FIFO扣除股數", 0)), 0),
            "未配對賣出股數": round(top15_safe_float(lot.get("未配對賣出股數", 0)), 0),
            "FIFO完整狀態": str(lot.get("FIFO完整狀態", "OK")).strip() or "OK",
            "分點": str(lot.get("分點", "")).strip(),
            "分點名稱": str(lot.get("分點名稱", "")).strip(),
            "券商代號": str(lot.get("券商代號", "")).strip(),
            "標的股": str(lot.get("標的股", "")).strip(),
            "標的名稱": str(lot.get("標的名稱", "")).strip(),
            "事件": str(lot.get("事件", "")).strip(),
            "事件類型": str(lot.get("事件類型", "")).strip(),
            "事件日": normalize_date_str(lot.get("事件日", "")),
            "買進日": normalize_date_str(lot.get("買進日", "")),
            "權證代號": str(lot.get("權證代號", "")).strip(),
            "權證名稱": str(lot.get("權證名稱", "")).strip(),
            "原始股數": round(original_qty, 0),
            "原始成本": round(original_cost, 0),
            "原始均價": "" if original_avg is None else round(original_avg, 4),
            "已扣賣出股數": round(sold_qty, 0),
            "已扣賣出成本": round(sold_cost, 0),
            "剩餘股數": round(remaining_qty, 0),
            "剩餘成本": round(remaining_cost, 0),
            "剩餘均價": "" if remaining_avg is None else round(remaining_avg, 4),
            "最新權證價格": "" if latest_price is None else latest_price,
            "最新價格日期": latest_price_date,
            "估值價格來源": valuation_price_source,
            "估值符號": detail_valuation_symbol,
            "目前市值": market_value,
            "未實現損益": unrealized_pnl,
            "報酬率": return_pct,
            "報酬率文字": return_text,
            "價格狀態": price_status,
            "完成狀態": "DONE",
            "來源": str(lot.get("來源", "")).strip(),
            "run_id": run_id,
            "更新時間": update_time,
        })

    if missing_price_rows:
        preview = "\n".join(missing_price_rows[:20])
        extra = "" if len(missing_price_rows) <= 20 else f"\n... 其餘 {len(missing_price_rows) - 20} 筆略"
        print(
            "  ⚠️ TOP15固定資料集：部分權證因多日未造市 / 無有效價格，不納入報酬率估算，但仍保留淨買超成本：\n"
            + preview
            + extra
        )

    if validation_errors and TOP15_FAIL_ON_MISSING_PRICE:
        preview = "\n".join(validation_errors[:20])
        extra = "" if len(validation_errors) <= 20 else f"\n... 其餘 {len(validation_errors) - 20} 筆略"
        raise RuntimeError(
            "TOP15固定資料集驗證失敗，為避免淨買超成本錯誤，本次 RUN 已中止：\n"
            + preview
            + extra
        )

    consensus_rows = build_top15_consensus_rows_from_detail(detail_rows, run_id, update_time)

    print(
        f"  ✅ TOP15固定資料集完成：部位明細 {len(detail_rows):,} 筆，"
        f"共識淨買超 {len(consensus_rows):,} 檔標的"
    )
    return detail_rows, consensus_rows


def build_top15_consensus_rows_from_detail(detail_rows, run_id, update_time):
    """
    由「快取_TOP15部位明細」加總出「快取_TOP15共識淨買超」。
    這裡不再讀 A/B/C/D/E 或分點歷史，確保圖片用資料只來自同一份固定明細。

    零碎部位規則：
    - 先依「標的股＋分點」完整加總剩餘部位。
    - 分點合計剩餘成本低於 TOP15_MIN_BROKER_REMAINING_COST 時，
      僅保留在部位明細，不納入共識 TOP15 的金額、排名、報酬率與參與分點數。
    """
    if not detail_rows:
        return []

    scope = str(
        detail_rows[0].get("資料範圍", "")
        or get_result_data_scope()
    ).strip()

    # 第一階段先依「標的股＋分點」彙總，避免逐 lot 套門檻時誤刪
    # 兩筆各 6 萬、合計其實達 12 萬的有效分點部位。
    broker_agg = {}

    for row in detail_rows:
        underlying = str(row.get("標的股", "")).strip()
        if not underlying:
            continue

        broker = str(row.get("分點", "")).strip()
        broker_name = str(row.get("分點名稱", "")).strip()
        broker_code = str(row.get("券商代號", "")).strip()
        broker_key = (underlying, broker, broker_name, broker_code)

        broker_rec = broker_agg.setdefault(broker_key, {
            "資料範圍": row.get("資料範圍", scope),
            "統計日期": row.get("統計日期", ""),
            "統計期間": row.get("統計期間", ""),
            "有效交易日數": row.get("有效交易日數", ""),
            "標的股": underlying,
            "標的名稱": str(row.get("標的名稱", "")).strip(),
            "分點": broker,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "淨買超成本": 0.0,
            "可估成本": 0.0,
            "成交價成本": 0.0,
            "LP報價成本": 0.0,
            "最近成交價備援成本": 0.0,
            "缺價格成本": 0.0,
            "目前市值": 0.0,
            "未實現損益": 0.0,
            "事件集合": set(),
            "權證集合": set(),
            "權證清單": [],
            "最新價格日期集合": set(),
            "FIFO狀態集合": set(),
        })

        remaining_cost = top15_safe_float(row.get("剩餘成本", 0))
        market_value = top15_safe_float(row.get("目前市值", 0), 0.0) if row.get("目前市值", "") != "" else 0.0
        pnl = top15_safe_float(row.get("未實現損益", 0), 0.0) if row.get("未實現損益", "") != "" else 0.0
        price_status = str(row.get("價格狀態", "")).strip()
        valuation_price_source = str(row.get("估值價格來源", "")).strip()
        event_code = str(row.get("事件", "")).strip()
        warrant_code = str(row.get("權證代號", "")).strip()
        warrant_name = str(row.get("權證名稱", "")).strip()
        latest_price_date = str(row.get("最新價格日期", "")).strip()
        fifo_status = str(row.get("FIFO完整狀態", "OK")).strip() or "OK"

        broker_rec["淨買超成本"] += remaining_cost
        if price_status == "OK":
            broker_rec["可估成本"] += remaining_cost
            broker_rec["目前市值"] += market_value
            broker_rec["未實現損益"] += pnl
            if valuation_price_source == "當日成交價":
                broker_rec["成交價成本"] += remaining_cost
            elif valuation_price_source == "LP流動量提供者委買價":
                broker_rec["LP報價成本"] += remaining_cost
            elif valuation_price_source == "最近成交價備援":
                broker_rec["最近成交價備援成本"] += remaining_cost
        else:
            broker_rec["缺價格成本"] += remaining_cost

        if event_code:
            broker_rec["事件集合"].add(event_code)
        if warrant_code:
            broker_rec["權證集合"].add(warrant_code)
            warrant_label = f"{warrant_code} {warrant_name}".strip()
            if warrant_label and warrant_label not in broker_rec["權證清單"]:
                broker_rec["權證清單"].append(warrant_label)
        if latest_price_date:
            broker_rec["最新價格日期集合"].add(latest_price_date)
        if fifo_status and fifo_status != "OK":
            broker_rec["FIFO狀態集合"].add(fifo_status)

    # 第二階段套用「分點＋標的」合計門檻，再建立標的股總表。
    agg = {}
    filtered_broker_count = 0
    filtered_broker_cost = 0.0

    for broker_rec in broker_agg.values():
        broker_cost = float(broker_rec.get("淨買超成本", 0) or 0)
        if broker_cost < TOP15_MIN_BROKER_REMAINING_COST:
            filtered_broker_count += 1
            filtered_broker_cost += broker_cost
            continue

        underlying = broker_rec["標的股"]
        rec = agg.setdefault(underlying, {
            "資料範圍": broker_rec.get("資料範圍", scope),
            "統計日期": broker_rec.get("統計日期", ""),
            "統計期間": broker_rec.get("統計期間", ""),
            "有效交易日數": broker_rec.get("有效交易日數", ""),
            "標的股": underlying,
            "標的名稱": broker_rec.get("標的名稱", ""),
            "淨買超成本": 0.0,
            "可估成本": 0.0,
            "成交價成本": 0.0,
            "LP報價成本": 0.0,
            "最近成交價備援成本": 0.0,
            "缺價格成本": 0.0,
            "目前市值": 0.0,
            "未實現損益": 0.0,
            "分點": {},
            "事件集合": set(),
            "權證集合": set(),
            "權證清單": [],
            "最新價格日期集合": set(),
            "FIFO狀態集合": set(),
            "資料狀態": "OK",
        })

        rec["淨買超成本"] += broker_cost
        rec["可估成本"] += float(broker_rec.get("可估成本", 0) or 0)
        rec["成交價成本"] += float(broker_rec.get("成交價成本", 0) or 0)
        rec["LP報價成本"] += float(broker_rec.get("LP報價成本", 0) or 0)
        rec["最近成交價備援成本"] += float(
            broker_rec.get("最近成交價備援成本", 0) or 0
        )
        rec["缺價格成本"] += float(broker_rec.get("缺價格成本", 0) or 0)
        rec["目前市值"] += float(broker_rec.get("目前市值", 0) or 0)
        rec["未實現損益"] += float(broker_rec.get("未實現損益", 0) or 0)
        rec["事件集合"].update(broker_rec["事件集合"])
        rec["權證集合"].update(broker_rec["權證集合"])
        rec["最新價格日期集合"].update(broker_rec["最新價格日期集合"])
        rec["FIFO狀態集合"].update(broker_rec.get("FIFO狀態集合", set()))
        for warrant_label in broker_rec["權證清單"]:
            if warrant_label not in rec["權證清單"]:
                rec["權證清單"].append(warrant_label)

        if float(broker_rec.get("缺價格成本", 0) or 0) > 0:
            rec["資料狀態"] = "部分報酬率未估"
        if broker_rec.get("FIFO狀態集合"):
            fifo_warning = "／".join(sorted(broker_rec["FIFO狀態集合"]))
            rec["資料狀態"] = (
                f"{rec['資料狀態']}；FIFO:{fifo_warning}"
                if rec["資料狀態"] != "OK"
                else f"FIFO:{fifo_warning}"
            )

        broker_identity = (
            broker_rec["分點"],
            broker_rec["分點名稱"],
            broker_rec["券商代號"],
        )
        rec["分點"][broker_identity] = broker_rec

    if filtered_broker_count > 0:
        print(
            f"  ℹ️ TOP15零碎分點部位已排除：{filtered_broker_count:,} 筆｜"
            f"合計 {filtered_broker_cost:,.0f} 元｜"
            f"門檻 {TOP15_MIN_BROKER_REMAINING_COST:,.0f} 元"
        )

    rows = []
    sorted_records = sorted(
        agg.values(),
        key=lambda r: (float(r.get("淨買超成本", 0) or 0), float(r.get("可估成本", 0) or 0)),
        reverse=True,
    )

    for rank, rec in enumerate(sorted_records, 1):
        total_cost = float(rec.get("淨買超成本", 0) or 0)
        estimated_cost = float(rec.get("可估成本", 0) or 0)
        traded_price_cost = float(rec.get("成交價成本", 0) or 0)
        lp_quote_cost = float(rec.get("LP報價成本", 0) or 0)
        last_trade_fallback_cost = float(rec.get("最近成交價備援成本", 0) or 0)
        missing_cost = float(rec.get("缺價格成本", 0) or 0)
        market_value = float(rec.get("目前市值", 0) or 0)
        pnl = float(rec.get("未實現損益", 0) or 0)
        return_pct = round(pnl / estimated_cost * 100, 2) if estimated_cost > 0 else None
        coverage_pct = round(estimated_cost / total_cost * 100, 2) if total_cost > 0 else None
        traded_price_coverage_pct = round(traded_price_cost / total_cost * 100, 2) if total_cost > 0 else None
        valuation_symbol = (
            "*"
            if return_pct is not None
            and traded_price_coverage_pct is not None
            and traded_price_coverage_pct < TOP15_TRADED_PRICE_COVERAGE_NOTE_THRESHOLD_PCT
            else ""
        )

        broker_rows = []
        broker_json = []

        for broker_rec in sorted(rec["分點"].values(), key=lambda x: x["淨買超成本"], reverse=True):
            b_cost = float(broker_rec.get("淨買超成本", 0) or 0)
            b_estimated_cost = float(broker_rec.get("可估成本", 0) or 0)
            b_traded_price_cost = float(broker_rec.get("成交價成本", 0) or 0)
            b_lp_quote_cost = float(broker_rec.get("LP報價成本", 0) or 0)
            b_last_trade_fallback_cost = float(
                broker_rec.get("最近成交價備援成本", 0) or 0
            )
            b_pnl = float(broker_rec.get("未實現損益", 0) or 0)
            b_return_pct = round(b_pnl / b_estimated_cost * 100, 2) if b_estimated_cost > 0 else None
            b_traded_price_coverage_pct = round(b_traded_price_cost / b_cost * 100, 2) if b_cost > 0 else None
            b_valuation_symbol = (
                "*"
                if b_return_pct is not None
                and b_traded_price_coverage_pct is not None
                and b_traded_price_coverage_pct < TOP15_TRADED_PRICE_COVERAGE_NOTE_THRESHOLD_PCT
                else ""
            )
            b_events = "/".join(sorted(x for x in broker_rec["事件集合"] if x))
            b_warrant_count = len(broker_rec["權證集合"])
            b_return_text = "-" if b_return_pct is None else f"{b_return_pct:+.2f}%{b_valuation_symbol}"

            b_missing_cost = float(broker_rec.get("缺價格成本", 0) or 0)
            b_coverage_text = ""
            if b_missing_cost > 0 and b_cost > 0:
                b_coverage_text = f"｜估值{b_estimated_cost / b_cost * 100:.0f}%"

            broker_rows.append(
                f"{broker_rec['分點']} {top15_fmt_amount_wan(b_cost)}（{b_return_text}{b_coverage_text}｜{b_events}｜{b_warrant_count}檔）"
            )
            broker_json.append({
                "分點": broker_rec["分點"],
                "分點名稱": broker_rec["分點名稱"],
                "券商代號": broker_rec["券商代號"],
                "淨買超成本": round(b_cost, 0),
                "可估成本": round(b_estimated_cost, 0),
                "成交價成本": round(b_traded_price_cost, 0),
                "LP報價成本": round(b_lp_quote_cost, 0),
                "最近成交價備援成本": round(b_last_trade_fallback_cost, 0),
                "缺價格成本": round(float(broker_rec.get("缺價格成本", 0) or 0), 0),
                "成交價覆蓋率": "" if b_traded_price_coverage_pct is None else b_traded_price_coverage_pct,
                "估值符號": b_valuation_symbol,
                "目前市值": round(float(broker_rec.get("目前市值", 0) or 0), 0),
                "未實現損益": round(b_pnl, 0),
                "報酬率": "" if b_return_pct is None else b_return_pct,
                "事件": b_events,
                "權證檔數": b_warrant_count,
            })

        rows.append({
            "資料範圍": rec.get("資料範圍", scope),
            "統計日期": rec.get("統計日期", ""),
            "統計期間": rec.get("統計期間", ""),
            "有效交易日數": rec.get("有效交易日數", ""),
            "排名": rank,
            "標的股": rec.get("標的股", ""),
            "標的名稱": rec.get("標的名稱", ""),
            "淨買超成本": round(total_cost, 0),
            "可估成本": round(estimated_cost, 0),
            "成交價成本": round(traded_price_cost, 0),
            "LP報價成本": round(lp_quote_cost, 0),
            "最近成交價備援成本": round(last_trade_fallback_cost, 0),
            "缺價格成本": round(missing_cost, 0),
            "成交價覆蓋率": "" if traded_price_coverage_pct is None else traded_price_coverage_pct,
            "估值符號": valuation_symbol,
            "目前市值": round(market_value, 0),
            "未實現損益": round(pnl, 0),
            "報酬率": "" if return_pct is None else return_pct,
            "報酬率文字": "-" if return_pct is None else (
                f"{return_pct:+.2f}%{valuation_symbol}"
                if missing_cost <= 0
                else f"{return_pct:+.2f}%{valuation_symbol}（部分估）"
            ),
            "價格覆蓋率": "" if coverage_pct is None else coverage_pct,
            "價格覆蓋率文字": "-" if coverage_pct is None else f"{coverage_pct:.2f}%",
            "參與分點數": len(rec["分點"]),
            "參與分點明細": "\n".join(broker_rows),
            "事件": "/".join(sorted(x for x in rec["事件集合"] if x)),
            "權證檔數": len(rec["權證集合"]),
            "權證清單": "；".join(rec["權證清單"]),
            "最新價格日期": max(rec["最新價格日期集合"]) if rec["最新價格日期集合"] else "",
            "資料狀態": rec.get("資料狀態", "OK"),
            "完成狀態": "DONE",
            "更新時間": update_time,
            "run_id": run_id,
            "分點明細_JSON": json.dumps(broker_json, ensure_ascii=False),
        })

    return rows

def write_top15_position_detail_sheet(wb, detail_rows):
    """寫入 TOP15 部位明細固定資料集。"""
    ws = wb.create_sheet(TOP15_POSITION_DETAIL_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "有效交易日數",
        "交易日區間起日", "交易日區間迄日", "事件是否位於40日內",
        "完整FIFO歷史起日", "期初庫存股數", "同日抵銷股數",
        "歷史FIFO扣除股數", "未配對賣出股數", "FIFO完整狀態",
        "分點", "分點名稱", "券商代號",
        "標的股", "標的名稱",
        "事件", "事件類型", "事件日", "買進日",
        "權證代號", "權證名稱",
        "原始股數", "原始成本", "原始均價",
        "已扣賣出股數", "已扣賣出成本",
        "剩餘股數", "剩餘成本", "剩餘均價",
        "最新權證價格", "最新價格日期", "估值價格來源", "估值符號",
        "目前市值", "未實現損益", "報酬率", "報酬率文字",
        "價格狀態", "完成狀態", "來源", "run_id", "更新時間",
    ]

    ws.append(headers)

    for row in detail_rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [12, 12, 24, 12, 12, 12, 14, 14, 12, 12, 14, 14, 18, 14, 18, 12, 10, 12, 8, 22, 12, 12, 12, 24, 12, 14, 10, 14, 14, 12, 14, 10, 12, 12, 22, 10, 14, 14, 10, 12, 10, 10, 44, 16, 20]
    _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="價格狀態")


def write_top15_consensus_cache_sheet(wb, consensus_rows):
    """寫入 TOP15 共識淨買超固定資料集，圖片程式應直接讀這張表。"""
    ws = wb.create_sheet(TOP15_CONSENSUS_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "有效交易日數", "排名",
        "標的股", "標的名稱",
        "淨買超成本", "可估成本", "成交價成本", "LP報價成本",
        "最近成交價備援成本", "缺價格成本",
        "成交價覆蓋率", "估值符號",
        "目前市值", "未實現損益", "報酬率", "報酬率文字",
        "價格覆蓋率", "價格覆蓋率文字",
        "參與分點數", "參與分點明細",
        "事件", "權證檔數", "權證清單",
        "最新價格日期", "資料狀態", "完成狀態",
        "更新時間", "run_id", "分點明細_JSON",
    ]

    ws.append(headers)

    for row in consensus_rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [12, 12, 24, 12, 8, 10, 12, 14, 14, 14, 14, 14, 14, 12, 10, 14, 14, 10, 12, 12, 14, 12, 48, 10, 10, 60, 12, 10, 10, 20, 16, 80]
    _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="資料狀態")


def _style_top15_cache_sheet(ws, col_widths, return_col_name="報酬率", status_col_name="資料狀態"):
    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    return_col_idx = header_map.get(return_col_name)
    status_col_idx = header_map.get(status_col_name)

    for row in ws.iter_rows(min_row=2):
        pct = None
        status_text = ""

        if return_col_idx:
            try:
                value = row[return_col_idx - 1].value
                if value not in (None, "", "-"):
                    pct = float(value)
            except Exception:
                pct = None

        if status_col_idx:
            status_text = str(row[status_col_idx - 1].value or "").strip()

        if status_text and status_text != "OK":
            row_fill = ORANGE
        elif pct is not None and pct > 0:
            row_fill = RED
        elif pct is not None and pct < 0:
            row_fill = GREEN
        else:
            row_fill = WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 30

    ws.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════
# Excel 樣式
# ══════════════════════════════════════════════════════════════════════

def style_sheet(ws, col_widths, status_rows=None, header_row=1):
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[header_row].height = 24

    fill_map = {
        "win": RED,
        "lose": GREEN,
        "reduce": BLUE,
        "exit": ORANGE,
        "none": PatternFill(),
    }

    if status_rows:
        for r_idx, status_row in enumerate(status_rows, header_row + 1):
            for c_idx, status in enumerate(status_row, 1):
                cell = ws.cell(r_idx, c_idx)
                cell.fill = fill_map.get(status, PatternFill())
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[r_idx].height = 36
    else:
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = f"A{header_row + 1}"


def apply_exit_profit_result_outline(ws, row_idx, col_idx, return_pct):
    """
    針對實際「減碼獲利% / 出清獲利%」欄位加外框：
    獲利% > 0：粗紅色外框
    獲利% < 0：粗綠色外框

    不再框 D+N 橘色出清格，避免把「收盤價漲跌」誤解成實際出清損益。
    """
    if return_pct is None:
        return

    try:
        pct = float(return_pct)
    except Exception:
        return

    if pct > 0:
        side = PROFIT_EXIT_SIDE
    elif pct < 0:
        side = LOSS_EXIT_SIDE
    else:
        return

    cell = ws.cell(row_idx, col_idx)
    cell.border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side,
    )


# ══════════════════════════════════════════════════════════════════════
# 建立 Excel：新制 A / B / C / D / E / 勝率統計
# ══════════════════════════════════════════════════════════════════════

def write_group_sheet(wb, sheet_name, events, price_cache, is_c=False):
    ws = wb.create_sheet(sheet_name)

    day_cols = [f"D+{i}" for i in range(1, 21)]

    if is_c:
        headers = [
            "事件類型", "分點", "標的股",
            "起始日", "結束日", "標的股股價", "5MA", "20MA",
            "涵蓋權證數", "權證清單",
            "最大單筆金額", "最大單筆權證",
            "單日累積買進金額", "買超張數",
            "減碼日", "減碼賣出金額", "減碼獲利%",
            "出清日", "出清賣出金額", "出清獲利%",
            "持有天數",
        ] + day_cols
        fixed_len = 21
    else:
        headers = [
            "事件類型", "分點", "標的股",
            "事件日", "標的股股價", "5MA", "20MA",
            "涵蓋權證數", "權證清單",
            "最大單筆金額", "最大單筆權證",
            "單日累積買進金額", "買超張數",
            "減碼日", "減碼賣出金額", "減碼獲利%",
            "出清日", "出清賣出金額", "出清獲利%",
            "持有天數",
        ] + day_cols
        fixed_len = 20

    ws.append(headers)
    status_rows = []
    exit_profit_result_cells = []

    sort_date_field = "結束日" if is_c else "事件日"
    sorted_events = sorted(
        events,
        key=lambda x: (
            -((parse_date(x.get(sort_date_field) or x.get("事件日") or x.get("起始日")) or datetime.min).toordinal()),
            str(x.get("分點", "")),
            str(x.get("標的股", "")),
        )
    )

    for ev in sorted_events:
        day_values, day_status = make_group_day_cells(ev, price_cache)
        ma_base_date = ev.get("結束日") if is_c else ev.get("事件日")
        underlying_price, ma5_mark, ma20_mark = get_ma_status_cells(
            price_cache,
            ev.get("標的股"),
            ma_base_date,
        )

        if is_c:
            row = [
                ev["事件類型"],
                ev["分點"],
                ev["標的股"],
                ev["起始日"],
                ev["結束日"],
                underlying_price,
                ma5_mark,
                ma20_mark,
                ev["涵蓋權證數"],
                ev["權證清單"],
                fmt_amount(ev.get("最大單筆金額")),
                ev.get("最大單筆權證", "-"),
                fmt_amount(ev.get("單日累積買進金額", ev.get("買超金額"))),
                ev["買超張數"],
                ev["減碼日"] or "-",
                fmt_amount(ev["減碼賣出金額"]),
                fmt_pct(ev["減碼獲利%"]),
                ev["出清日"] or "-",
                fmt_amount(ev["出清賣出金額"]),
                fmt_pct(ev["出清獲利%"]),
                fmt_num(ev["持有天數"]),
            ] + day_values
        else:
            row = [
                ev["事件類型"],
                ev["分點"],
                ev["標的股"],
                ev["事件日"],
                underlying_price,
                ma5_mark,
                ma20_mark,
                ev["涵蓋權證數"],
                ev["權證清單"],
                fmt_amount(ev.get("最大單筆金額")),
                ev.get("最大單筆權證", "-"),
                fmt_amount(ev.get("單日累積買進金額", ev.get("買超金額"))),
                ev["買超張數"],
                ev["減碼日"] or "-",
                fmt_amount(ev["減碼賣出金額"]),
                fmt_pct(ev["減碼獲利%"]),
                ev["出清日"] or "-",
                fmt_amount(ev["出清賣出金額"]),
                fmt_pct(ev["出清獲利%"]),
                fmt_num(ev["持有天數"]),
            ] + day_values

        ws.append(row)
        current_row_idx = ws.max_row
        status_rows.append(["none"] * fixed_len + day_status)

        if ev["減碼獲利%"] is not None:
            # 新制 A/B/C/D/E 工作表第 16 欄為「減碼獲利%」；跨日起訖格式第 17 欄為「減碼獲利%」
            reduce_profit_col_idx = 17 if is_c else 16
            exit_profit_result_cells.append((current_row_idx, reduce_profit_col_idx, ev["減碼獲利%"]))

        if ev["出清獲利%"] is not None:
            # 新制 A/B/C/D/E 工作表第 19 欄為「出清獲利%」；跨日起訖格式第 20 欄為「出清獲利%」
            exit_profit_col_idx = 20 if is_c else 19
            exit_profit_result_cells.append((current_row_idx, exit_profit_col_idx, ev["出清獲利%"]))

    if is_c:
        col_widths = [20, 14, 8, 12, 12, 12, 8, 8, 12, 45, 14, 24, 16, 10, 12, 14, 12, 12, 14, 12, 10] + [14] * 20
    else:
        col_widths = [20, 14, 8, 12, 12, 8, 8, 12, 45, 14, 24, 16, 10, 12, 14, 12, 12, 14, 12, 10] + [14] * 20

    style_sheet(ws, col_widths, status_rows)

    for row_idx, col_idx, return_pct in exit_profit_result_cells:
        apply_exit_profit_result_outline(ws, row_idx, col_idx, return_pct)


def _safe_stat_amount(value):
    try:
        if value is None:
            return 0
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return 0
        return int(round(float(s)))
    except Exception:
        return 0


def collect_stat_records(a_events, b_events, c_events, d_events, e_events=None):
    records = []

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            return_pct = ev["出清獲利%"]
            closed = return_pct is not None

            records.append({
                "分點": ev["分點"],
                "事件類型": ev.get("事件類型", f"{event_code}-{AMOUNT_CLASS_LABELS.get(event_code, '事件')}"),
                "事件代碼": event_code,
                "是否出清": closed,
                "結果": calc_result_tag(return_pct),
                "持有天數": ev["持有天數"],
                "報酬%": return_pct,
                "買進金額": _safe_stat_amount(ev.get("買超金額", ev.get("單日累積買進金額", 0))),
            })

    return records

def calc_summary_for_group(g, broker, event_type):
    total_events = len(g)
    closed_g = g[g["是否出清"] == True]
    open_g = g[g["是否出清"] == False]

    closed_count = len(closed_g)
    open_count = len(open_g)

    win_count = int((closed_g["結果"] == "勝").sum()) if closed_count > 0 else 0
    loss_count = int((closed_g["結果"] == "敗").sum()) if closed_count > 0 else 0
    flat_count = int((closed_g["結果"] == "平手").sum()) if closed_count > 0 else 0

    win_rate = round(win_count / closed_count * 100, 2) if closed_count > 0 else None

    avg_holding_days = None
    if closed_count > 0:
        holding_series = pd.to_numeric(closed_g["持有天數"], errors="coerce").dropna()
        if len(holding_series) > 0:
            avg_holding_days = round(float(holding_series.mean()), 2)

    avg_return = None
    weighted_return = None
    max_return = None
    min_return = None
    total_entry_amount = 0
    closed_entry_amount = 0
    estimated_pnl_amount = None
    avg_entry_amount = None

    if "買進金額" in g.columns:
        total_entry_amount = int(pd.to_numeric(g["買進金額"], errors="coerce").fillna(0).sum())

    if closed_count > 0:
        return_series = pd.to_numeric(closed_g["報酬%"], errors="coerce").dropna()
        if len(return_series) > 0:
            avg_return = round(float(return_series.mean()), 2)
            max_return = round(float(return_series.max()), 2)
            min_return = round(float(return_series.min()), 2)

        if "買進金額" in closed_g.columns:
            weighted_df = closed_g.copy()
            weighted_df["報酬%"] = pd.to_numeric(weighted_df["報酬%"], errors="coerce")
            weighted_df["買進金額"] = pd.to_numeric(weighted_df["買進金額"], errors="coerce").fillna(0)
            weighted_df = weighted_df.dropna(subset=["報酬%"] )
            weighted_df = weighted_df[weighted_df["買進金額"] > 0]

            if not weighted_df.empty:
                closed_entry_amount = int(round(float(weighted_df["買進金額"].sum())))
                weighted_pnl = float((weighted_df["買進金額"] * weighted_df["報酬%"] / 100.0).sum())
                estimated_pnl_amount = int(round(weighted_pnl))
                weighted_return = round(weighted_pnl / closed_entry_amount * 100.0, 2) if closed_entry_amount > 0 else None
                avg_entry_amount = int(round(closed_entry_amount / len(weighted_df))) if len(weighted_df) > 0 else None

    return {
        "分點": broker,
        "事件類型": event_type,
        "事件數": total_events,
        "已出清筆數": closed_count,
        "未出清筆數": open_count,
        "勝筆數": win_count,
        "敗筆數": loss_count,
        "平手筆數": flat_count,
        "勝率": win_rate,
        "平均持有天數": avg_holding_days,
        "平均報酬%": avg_return,
        "加權報酬%": weighted_return,
        "總買進金額": total_entry_amount,
        "已出清買進金額": closed_entry_amount,
        "估算損益金額": estimated_pnl_amount,
        "平均單筆買進金額": avg_entry_amount,
        "最高報酬%": max_return,
        "最低報酬%": min_return,
    }


def calc_empty_summary(broker, event_type):
    return {
        "分點": broker,
        "事件類型": event_type,
        "事件數": 0,
        "已出清筆數": 0,
        "未出清筆數": 0,
        "勝筆數": 0,
        "敗筆數": 0,
        "平手筆數": 0,
        "勝率": None,
        "平均持有天數": None,
        "平均報酬%": None,
        "加權報酬%": None,
        "總買進金額": 0,
        "已出清買進金額": 0,
        "估算損益金額": None,
        "平均單筆買進金額": None,
        "最高報酬%": None,
        "最低報酬%": None,
    }


WINRATE_STATS_MIN_TOTAL_EVENTS = int(os.getenv("WINRATE_STATS_MIN_TOTAL_EVENTS", "10"))


def _winrate_sort_value(value):
    try:
        if value is None:
            return -1.0
        return float(value)
    except Exception:
        return -1.0


def sort_brokers_by_winrate_summary(summary_map, broker_order):
    """
    勝率統計工作表的分點排序。

    排序邏輯：
    1. 先看「全部-A+B+C+D+E合併」的事件數是否達到門檻。
       - ABCDE 總事件數 >= WINRATE_STATS_MIN_TOTAL_EVENTS 的分點排前面。
       - ABCDE 總事件數 < WINRATE_STATS_MIN_TOTAL_EVENTS 的分點全部排最後。
    2. 同一組內依「全部-A+B+C+D+E合併」勝率由高到低排序。
    3. 若總勝率一樣，依序比較 A / B / C / D / E 類勝率。
    4. 若勝率都一樣，再以 ABCDE 總事件數多到少排序。
    5. 最後保留原本分點順序作為穩定排序依據。
    """
    original_index = {broker: idx for idx, broker in enumerate(broker_order)}

    def key_func(broker):
        broker_summary = summary_map.get(broker, {})
        all_summary = broker_summary.get("ALL", {})
        total_events = int(all_summary.get("事件數") or 0)
        qualified_rank = 0 if total_events >= WINRATE_STATS_MIN_TOTAL_EVENTS else 1

        return (
            qualified_rank,
            -_winrate_sort_value(all_summary.get("勝率")),
            -_winrate_sort_value(broker_summary.get("A", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("B", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("C", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("D", {}).get("勝率")),
            -_winrate_sort_value(broker_summary.get("E", {}).get("勝率")),
            -total_events,
            original_index.get(broker, 999999),
        )

    return sorted(broker_order, key=key_func)


def make_summary_map(stat_records):
    if not stat_records:
        stat_df = pd.DataFrame(columns=["分點", "事件代碼", "事件類型", "是否出清", "結果", "持有天數", "報酬%", "買進金額"])
    else:
        stat_df = pd.DataFrame(stat_records)

    broker_order = list(TARGET_PATTERNS.keys())

    if not stat_df.empty:
        for broker in sorted(stat_df["分點"].dropna().unique()):
            if broker not in broker_order:
                broker_order.append(broker)

    summary_map = {}

    event_types = {
        code: f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}"
        for code in AMOUNT_CLASS_CODES
    }
    event_types["ALL"] = "全部-A+B+C+D+E合併"

    for broker in broker_order:
        summary_map[broker] = {}

        broker_g = stat_df[stat_df["分點"] == broker] if not stat_df.empty else pd.DataFrame()

        for code in AMOUNT_CLASS_CODES:
            if not broker_g.empty:
                g = broker_g[broker_g["事件代碼"] == code]
            else:
                g = pd.DataFrame()

            if len(g) > 0:
                summary_map[broker][code] = calc_summary_for_group(g, broker, event_types[code])
            else:
                summary_map[broker][code] = calc_empty_summary(broker, event_types[code])

        if len(broker_g) > 0:
            summary_map[broker]["ALL"] = calc_summary_for_group(broker_g, broker, event_types["ALL"])
        else:
            summary_map[broker]["ALL"] = calc_empty_summary(broker, event_types["ALL"])

    broker_order = sort_brokers_by_winrate_summary(summary_map, broker_order)

    return summary_map, broker_order























def apply_global_amount_comma_format(wb):
    """
    全工作表套用金額千分位格式。

    只要欄位名稱包含「金額」，例如：
    買進金額、賣出金額、淨買進金額、近20日淨買進金額、買超金額，
    就統一顯示為 1,234,567。

    不處理「均價」、「比例」、「獲利%」等欄位，避免小數或百分比被誤改。
    """
    for ws in wb.worksheets:
        # 掃前 10 列，因為大部分工作表標題在第 1 列，
        # 「券商查詢」的欄位標題在第 5 列。
        for header_row in range(1, min(ws.max_row, 10) + 1):
            amount_cols = []

            for col_idx in range(1, ws.max_column + 1):
                header_value = ws.cell(header_row, col_idx).value
                header = str(header_value).strip() if header_value is not None else ""

                if not header:
                    continue

                if "金額" in header:
                    # 排除非金額欄位，避免誤格式化
                    if "%" in header or "比例" in header or "均價" in header:
                        continue

                    amount_cols.append(col_idx)

            if not amount_cols:
                continue

            for col_idx in amount_cols:
                for row_idx in range(header_row + 1, ws.max_row + 1):
                    cell = ws.cell(row_idx, col_idx)

                    if cell.value is None or cell.value == "":
                        continue

                    # 公式欄位不動，只套顯示格式
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.number_format = '#,##0'
                        continue

                    if str(cell.value).strip() == "-":
                        continue

                    try:
                        raw = str(cell.value).replace(",", "").strip()

                        if raw.startswith("-"):
                            numeric_part = raw[1:]
                        else:
                            numeric_part = raw

                        if numeric_part.replace(".", "", 1).isdigit():
                            num = float(raw)

                            if num.is_integer():
                                cell.value = int(num)
                                cell.number_format = '#,##0'
                            else:
                                cell.value = num
                                cell.number_format = '#,##0.00'

                    except Exception:
                        pass






def build_event_warrant_source_map(a_events, b_events, c_events, d_events, e_events=None):
    """
    建立「分點 + 權證代號 -> A/B/C/D/E 事件來源」對照表。

    用途：
    1. 每日賣出明細直接來自原始分點歷史資料 items，不再用 A/B/C/D/E 工作表的減碼日推估。
    2. 若賣出的權證屬於任一新制 A/B/C/D/E 金額強度事件，也能標示它原本歸屬哪一類事件。
    3. 同一權證在不同分點互不混用；同一分點與權證有多次事件時，以事件日 ≤ 賣出日且尚未出清者中最舊的事件為準。
    4. 一併保存出清日，讓每日賣出明細可排除已在該筆賣出日前出清的舊事件。
    """
    source_map = {}

    def put_source(broker, warrant_code, event_code, event_type, event_date, exit_date, event_source):
        broker = str(broker).strip()
        warrant_code = normalize_warrant_code_for_unique(warrant_code)
        if not warrant_code:
            return
        source_map.setdefault((broker, warrant_code), []).append({
            "事件": event_code,
            "事件類型": event_type,
            "事件日": normalize_date_str(event_date),
            "出清日": normalize_date_str(exit_date) if exit_date else "",
            "事件來源": event_source,
        })

    for event_code, events in iter_amount_class_event_groups(a_events, b_events, c_events, d_events, e_events):
        for ev in events:
            event_type = ev.get("事件類型", "")
            event_date = ev.get("事件日") or ev.get("結束日") or ev.get("起始日")

            for lot in ev.get("lots", []):
                put_source(
                    ev.get("分點", ""),
                    lot.get("權證代號", ""),
                    event_code,
                    event_type,
                    event_date,
                    ev.get("出清日"),
                    f'{event_code} | {lot.get("權證代號", "")} {lot.get("權證名稱", "")}',
                )

    return source_map




def _fmt_daily_sell_return_pct(value):
    if value is None:
        return "-"
    try:
        return f"{float(value):+.2f}%"
    except Exception:
        return "-"


def _daily_sell_fifo_return_map_for_item(item):
    """
    用同一分點＋同一權證的 FinMind 歷史資料，替每日賣出明細計算賣出報酬率。

    重點：
    1. 不再只依賴 A/B/C/D/E 事件，未歸類權證也能從快取_分點歷史回頭找買進成本。
    2. 權證不可當沖，同一天先扣舊庫存賣出，再加入當日買進。
    3. 成本優先使用 FIFO；若歷史快取缺少足夠庫存，才用賣出前所有歷史買進均價估算不足部分。
    4. 若完全沒有歷史買進成本，報酬率維持 '-'，並標示「缺歷史買進成本」。
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty or "日期" not in df.columns:
        return {}

    df2 = df.copy()
    df2["日期"] = df2["日期"].map(normalize_date_str)
    df2["dt_parsed"] = df2["日期"].map(parse_date)
    df2 = df2.dropna(subset=["dt_parsed"]).sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []
    out = {}
    historical_buy_qty = 0.0
    historical_buy_amount = 0.0

    for row in df2.itertuples(index=False):
        row_dict = row._asdict()
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0), 0.0)
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0), 0.0)
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0), 0.0)
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0), 0.0)

        # 權證不能當沖：同一天賣出先扣舊庫存，不吃同日買進。
        if sell_amount > 0:
            hist_avg_cost = historical_buy_amount / historical_buy_qty if historical_buy_qty > 0 and historical_buy_amount > 0 else None
            sell_price = sell_amount / sell_qty if sell_qty > 0 else None

            matched_qty = 0.0
            matched_cost = 0.0
            estimated_qty = 0.0
            estimated_cost = 0.0
            unmatched_qty = 0.0
            unmatched_amount = 0.0

            if sell_qty > 0 and sell_price is not None:
                sell_left = sell_qty

                for lot in lots:
                    if sell_left <= 0:
                        break

                    remain_qty = top15_safe_float(lot.get("剩餘股數", 0), 0.0)
                    if remain_qty <= 0:
                        continue

                    alloc_qty = min(sell_left, remain_qty)
                    lot_avg = top15_safe_float(lot.get("均價", 0), 0.0)
                    if alloc_qty <= 0 or lot_avg <= 0:
                        continue

                    lot["剩餘股數"] = remain_qty - alloc_qty
                    sell_left -= alloc_qty
                    matched_qty += alloc_qty
                    matched_cost += alloc_qty * lot_avg

                if sell_left > 0:
                    if hist_avg_cost is not None and hist_avg_cost > 0:
                        estimated_qty = sell_left
                        estimated_cost = sell_left * hist_avg_cost
                    else:
                        unmatched_qty = sell_left
                        unmatched_amount = sell_left * sell_price

                total_cost = matched_cost + estimated_cost
                pnl = sell_amount - total_cost if total_cost > 0 else None
                return_pct = pnl / total_cost * 100 if total_cost > 0 else None

                if total_cost > 0:
                    if estimated_qty > 0 and matched_qty > 0:
                        cost_status = "部分FIFO＋歷史均價估"
                    elif estimated_qty > 0:
                        cost_status = "歷史均價估"
                    elif unmatched_qty > 0:
                        cost_status = "部分成本不足"
                    else:
                        cost_status = "FIFO"
                else:
                    cost_status = "缺歷史買進成本"
            else:
                total_cost = None
                pnl = None
                return_pct = None
                cost_status = "賣出股數異常"

            out[date_str] = {
                "歷史買進股數": round(historical_buy_qty, 0),
                "歷史買進金額": round(historical_buy_amount, 0),
                "歷史平均成本": round(hist_avg_cost, 4) if hist_avg_cost is not None else "",
                "FIFO對應股數": round(matched_qty, 0),
                "估算對應股數": round(estimated_qty, 0),
                "成本不足股數": round(unmatched_qty, 0),
                "成本不足金額": round(unmatched_amount, 0),
                "賣出成本": round(total_cost, 0) if total_cost is not None else "",
                "賣出損益": round(pnl, 0) if pnl is not None else "",
                "報酬率": return_pct,
                "報酬率文字": _fmt_daily_sell_return_pct(return_pct),
                "成本狀態": cost_status,
            }

        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
            })
            historical_buy_qty += buy_qty
            historical_buy_amount += buy_amount

    return out

def write_daily_sell_detail_sheet(wb, items, a_events, b_events, c_events, d_events, e_events=None):
    """
    新增「每日賣出明細」工作表。

    速度修正：
    只輸出最近 SELL_DETAIL_DAYS 天的賣出明細，避免每次把 250 天全歷史賣出資料整張寫入 Google Sheet。
    原始歷史快取與 A/B/C/D/E、近兩月排行、券商查詢仍照舊使用完整 items，不受影響。

    重要修正：
    1. 今日賣超圖片不應再用 A/B/C/D/E 工作表的「減碼日 / 出清日」去反推賣出金額。
       A 類如果只賣 1 張，不能用「減碼均價 × 原始買進張數」估算，否則會把金額放大。
    2. 本表直接使用原始分點歷史資料 item["df"] 的「賣出股數 / 賣出金額」，
       因此會與官方分點資料一致。
    3. B / C / D 事件中的其中一檔權證若發生賣出，也會被列出，不會被群組事件的
       第一個減碼日 / 出清日限制而漏掉。
    4. 賣出報酬率不再只依賴 A/B/C/D/E 事件；未歸類權證會回到同分點 + 同權證歷史買進資料，
       用 FIFO 成本估算賣出損益與報酬率。
    """
    ws = wb.create_sheet("每日賣出明細")

    headers = [
        "日期",
        "分點",
        "分點名稱",
        "券商代號",
        "事件",
        "狀態",
        "標的股",
        "標的名稱",
        "權證代號",
        "權證名稱",
        "賣出張數",
        "賣出股數",
        "賣出金額",
        "賣出均價",
        "報酬率",
        "賣出成本",
        "賣出損益",
        "歷史買進張數",
        "歷史買進股數",
        "歷史買進金額",
        "歷史平均成本",
        "成本狀態",
        "事件日",
        "事件來源",
    ]
    ws.append(headers)

    source_map = build_event_warrant_source_map(a_events, b_events, c_events, d_events, e_events)
    rows = []
    cutoff_dt = datetime.today() - timedelta(days=max(SELL_DETAIL_DAYS - 1, 0))
    cutoff_dt = datetime(cutoff_dt.year, cutoff_dt.month, cutoff_dt.day)

    for item in items:
        df = item.get("df", pd.DataFrame())

        if df is None or df.empty:
            continue

        position = 0
        df2 = df.copy()
        df2["日期"] = df2["日期"].map(normalize_date_str)
        df2 = df2.sort_values("日期").reset_index(drop=True)
        sell_return_map = _daily_sell_fifo_return_map_for_item(item)

        for row in df2.itertuples(index=False):
            row_dict = row._asdict()
            date = normalize_date_str(row_dict.get("日期", ""))
            trade_dt = parse_date(date)
            buy_s = int(row_dict.get("買進股數", 0) or 0)
            sell_s = int(row_dict.get("賣出股數", 0) or 0)
            sell_a = int(row_dict.get("賣出金額", 0) or 0)

            before_position = position
            within_sell_detail_range = trade_dt is not None and trade_dt >= cutoff_dt

            if within_sell_detail_range and sell_s > 0 and sell_a > 0:
                if before_position > 0 and sell_s >= before_position:
                    status = "出清"
                elif before_position > 0:
                    status = "減碼"
                else:
                    status = "賣超"

                warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
                candidates_src = source_map.get(
                    (str(item.get("broker_label", "")).strip(), warrant_code), []
                )
                # 挑選規則：事件日 <= 賣出日，且（未出清 或 出清日 >= 賣出日）。
                # 符合多筆時取事件日最舊者，與 FIFO「舊事件先扣」的歸屬一致。
                source = {}
                best_dt = None
                for cand in candidates_src:
                    ev_dt = parse_date(cand.get("事件日", ""))
                    ex_dt = parse_date(cand.get("出清日", ""))
                    if not ev_dt or not trade_dt or ev_dt > trade_dt:
                        continue
                    if ex_dt and ex_dt < trade_dt:
                        continue
                    if best_dt is None or ev_dt < best_dt:
                        best_dt = ev_dt
                        source = cand

                sell_avg = round(sell_a / sell_s, 4) if sell_s > 0 else ""
                sell_return = sell_return_map.get(date, {})
                hist_buy_qty = top15_safe_float(sell_return.get("歷史買進股數", 0), 0.0)

                rows.append({
                    "日期": date,
                    "分點": item.get("broker_label", ""),
                    "分點名稱": item.get("broker_name", ""),
                    "券商代號": item.get("broker_code", ""),
                    "事件": source.get("事件", "未歸類"),
                    "狀態": status,
                    "標的股": item.get("underlying_code", ""),
                    "標的名稱": item.get("underlying_name", ""),
                    "權證代號": item.get("warrant_code", ""),
                    "權證名稱": item.get("warrant_name", ""),
                    "賣出張數": sell_s // 1000,
                    "賣出股數": sell_s,
                    "賣出金額": sell_a,
                    "賣出均價": sell_avg,
                    "報酬率": sell_return.get("報酬率文字", "-"),
                    "賣出成本": sell_return.get("賣出成本", ""),
                    "賣出損益": sell_return.get("賣出損益", ""),
                    "歷史買進張數": int(hist_buy_qty // 1000) if hist_buy_qty > 0 else 0,
                    "歷史買進股數": sell_return.get("歷史買進股數", 0),
                    "歷史買進金額": sell_return.get("歷史買進金額", 0),
                    "歷史平均成本": sell_return.get("歷史平均成本", ""),
                    "成本狀態": sell_return.get("成本狀態", "缺歷史買進成本"),
                    "事件日": source.get("事件日", ""),
                    "事件來源": source.get("事件來源", ""),
                })

            # 權證不可當沖：同日賣出先扣舊庫存，再把當日買進加入庫存。
            if sell_s > 0:
                position = max(position - sell_s, 0)

            if buy_s > 0:
                position += buy_s

    rows = sorted(
        rows,
        key=lambda r: (
            -((parse_date(r.get("日期")) or datetime.min).toordinal()),
            str(r.get("分點", "")),
            str(r.get("標的股", "")),
            -int(r.get("賣出金額", 0) or 0),
        )
    )

    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    col_widths = [12, 14, 18, 12, 8, 8, 10, 12, 12, 24, 10, 12, 16, 10, 10, 14, 14, 12, 14, 16, 12, 20, 12, 40]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    fill_map = {
        "減碼": BLUE,
        "出清": ORANGE,
        "賣超": GREEN,
    }

    for row in ws.iter_rows(min_row=2):
        status = str(row[5].value or "").strip()
        row_fill = fill_map.get(status, WHITE)

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 28

    ws.freeze_panes = "A2"



# ══════════════════════════════════════════════════════════════════════
# 近 10 日分點買賣明細快取：單一分點 + 標的股層級
# ══════════════════════════════════════════════════════════════════════







def _finish_price_ensure(
    price_cache,
    persistent_price_cache,
    changed_codes,
    defer_save=False,
):
    """統一處理三個價格補抓函式的回傳與延後儲存。"""
    changed_codes = {
        normalize_price_code(code)
        for code in (changed_codes or set())
        if normalize_price_code(code)
    }

    if not defer_save and persistent_price_cache is not None and changed_codes:
        save_price_cache(
            persistent_price_cache,
            changed_codes=changed_codes,
        )

    if defer_save:
        return price_cache, changed_codes

    return price_cache
















# ══════════════════════════════════════════════════════════════════════
# 近 10 日分點勝率排行（僅 RUN_MODE=2 更新）
# ══════════════════════════════════════════════════════════════════════







# ══════════════════════════════════════════════════════════════════════
# 近 7／14／21 日精選分點權證共識買賣超 TOP15（僅 RUN_MODE=2 更新）
# ══════════════════════════════════════════════════════════════════════











# ══════════════════════════════════════════════════════════════════════
# 完整修補模式延伸報表
# ══════════════════════════════════════════════════════════════════════

def write_stats_sheet(wb, a_events, b_events, c_events, d_events, e_events=None):
    ws = wb.create_sheet("勝率統計")

    headers = [
        "分點",
        "事件類型",
        "事件數",
        "已出清筆數",
        "未出清筆數",
        "勝筆數",
        "敗筆數",
        "平手筆數",
        "勝率",
        "平均持有天數",
        "平均報酬%",
        "加權報酬%",
        "總買進金額",
        "已出清買進金額",
        "估算損益金額",
        "平均單筆買進金額",
        "最高報酬%",
        "最低報酬%",
    ]

    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)
    summary_map, broker_order = make_summary_map(stat_records)

    thin_gray = Side(style="thin", color="B7B7B7")
    medium_gray = Side(style="medium", color="999999")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    broker_border = Border(left=medium_gray, right=medium_gray, top=medium_gray, bottom=medium_gray)

    current_row = 1

    for broker in broker_order:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(current_row, 1)
        title_cell.value = f"分點：{broker}"
        title_cell.font = Font(bold=True, color="000000", size=12)
        title_cell.fill = GRAY
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = broker_border
        ws.row_dimensions[current_row].height = 22

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(current_row, col_idx)
            cell.fill = GRAY
            cell.border = broker_border

        current_row += 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[current_row].height = 24

        current_row += 1

        for code in AMOUNT_CLASS_CODES + ["ALL"]:
            row = summary_map[broker][code]

            values = [
                row["分點"],
                row["事件類型"],
                row["事件數"],
                row["已出清筆數"],
                row["未出清筆數"],
                row["勝筆數"],
                row["敗筆數"],
                row["平手筆數"],
                "-" if row["勝率"] is None else f'{row["勝率"]:.2f}%',
                "-" if row["平均持有天數"] is None else row["平均持有天數"],
                "-" if row["平均報酬%"] is None else f'{row["平均報酬%"]:+.2f}%',
                "-" if row["加權報酬%"] is None else f'{row["加權報酬%"]:+.2f}%',
                fmt_amount(row.get("總買進金額")),
                fmt_amount(row.get("已出清買進金額")),
                "-" if row.get("估算損益金額") is None else fmt_amount(row.get("估算損益金額")),
                "-" if row.get("平均單筆買進金額") is None else fmt_amount(row.get("平均單筆買進金額")),
                "-" if row["最高報酬%"] is None else f'{row["最高報酬%"]:+.2f}%',
                "-" if row["最低報酬%"] is None else f'{row["最低報酬%"]:+.2f}%',
            ]

            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(current_row, col_idx)
                cell.value = value
                cell.font = Font(color="000000", bold=True if code == "ALL" else False)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = normal_border

                if code == "ALL":
                    cell.fill = PatternFill("solid", fgColor="EAF2F8")
                else:
                    cell.fill = WHITE

            ws.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1

    col_widths = [16, 24, 10, 12, 12, 10, 10, 10, 10, 14, 12, 12, 14, 16, 14, 16, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A1"

def write_combo_winrate_sheet(wb, a_events, b_events, c_events, d_events, e_events=None):
    """
    新增「ABCDE組合勝率」工作表。

    統計邏輯：
    1. 以分點為單位。
    2. 組合包含 AB / AC / AD / BC / BD / CD / ABC / ABD / ACD / BCD / ABCDE。
    3. 只有該分點同時具備該組合內所有事件類型，才列入該組合勝率。
    4. 勝率只用「已出清」事件計算，未出清不列入勝敗。

    排版邏輯：
    參考「勝率統計」工作表，每個分點獨立區塊顯示，方便閱讀與截圖。
    """
    ws = wb.create_sheet("ABCDE組合勝率")

    headers = [
        "分點",
        "組合",
        "包含事件",
        "是否同時出現",
        "A事件數",
        "B事件數",
        "C事件數",
        "D事件數",
        "E事件數",
        "組合事件數",
        "已出清筆數",
        "未出清筆數",
        "勝筆數",
        "敗筆數",
        "平手筆數",
        "勝率",
        "平均持有天數",
        "平均報酬%",
        "最高報酬%",
        "最低報酬%",
    ]

    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)

    if not stat_records:
        stat_df = pd.DataFrame(columns=[
            "分點", "事件代碼", "事件類型", "是否出清",
            "結果", "持有天數", "報酬%", "買進金額"
        ])
    else:
        stat_df = pd.DataFrame(stat_records)

    broker_order = list(TARGET_PATTERNS.keys())

    if not stat_df.empty:
        for broker in sorted(stat_df["分點"].dropna().unique()):
            if broker not in broker_order:
                broker_order.append(broker)

    combo_defs = []
    for mask in range(1, 1 << len(AMOUNT_CLASS_CODES)):
        codes = [AMOUNT_CLASS_CODES[idx] for idx in range(len(AMOUNT_CLASS_CODES)) if mask & (1 << idx)]
        if len(codes) >= 2:
            combo_defs.append(("".join(codes), codes))

    thin_gray = Side(style="thin", color="B7B7B7")
    medium_gray = Side(style="medium", color="999999")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    broker_border = Border(left=medium_gray, right=medium_gray, top=medium_gray, bottom=medium_gray)

    current_row = 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    title_cell = ws.cell(current_row, 1)
    title_cell.value = "ABCDE 所有組合勝率統計（依分點）"
    title_cell.font = Font(bold=True, color="000000", size=14)
    title_cell.fill = YELLOW
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = normal_border
    ws.row_dimensions[current_row].height = 28

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(current_row, col_idx)
        cell.fill = YELLOW
        cell.border = normal_border

    current_row += 1

    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    note_cell = ws.cell(current_row, 1)
    note_cell.value = "統計邏輯：只有該分點同時具備該組合內所有事件類型，才列入該組合勝率；勝率只用「已出清」事件計算，未出清不列入勝敗。"
    note_cell.font = Font(color="666666")
    note_cell.fill = WHITE
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.border = normal_border
    ws.row_dimensions[current_row].height = 24

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(current_row, col_idx)
        cell.fill = WHITE
        cell.border = normal_border

    current_row += 2

    for broker in broker_order:
        if stat_df.empty:
            broker_g = pd.DataFrame(columns=stat_df.columns)
        else:
            broker_g = stat_df[stat_df["分點"] == broker].copy()

        event_counts = {}
        for code in AMOUNT_CLASS_CODES:
            if broker_g.empty:
                event_counts[code] = 0
            else:
                event_counts[code] = int((broker_g["事件代碼"] == code).sum())

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        title_cell = ws.cell(current_row, 1)
        title_cell.value = f"分點：{broker}"
        title_cell.font = Font(bold=True, color="000000", size=12)
        title_cell.fill = GRAY
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        title_cell.border = broker_border
        ws.row_dimensions[current_row].height = 22

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(current_row, col_idx)
            cell.fill = GRAY
            cell.border = broker_border

        current_row += 1

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(current_row, col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[current_row].height = 24

        current_row += 1

        for combo_name, combo_codes in combo_defs:
            has_all = all(event_counts.get(code, 0) > 0 for code in combo_codes)
            include_text = " + ".join(combo_codes)

            if has_all and not broker_g.empty:
                combo_g = broker_g[broker_g["事件代碼"].isin(combo_codes)].copy()

                combo_event_count = len(combo_g)
                closed_g = combo_g[combo_g["是否出清"] == True]
                open_g = combo_g[combo_g["是否出清"] == False]

                closed_count = len(closed_g)
                open_count = len(open_g)

                win_count = int((closed_g["結果"] == "勝").sum()) if closed_count > 0 else 0
                loss_count = int((closed_g["結果"] == "敗").sum()) if closed_count > 0 else 0
                flat_count = int((closed_g["結果"] == "平手").sum()) if closed_count > 0 else 0

                win_rate = round(win_count / closed_count * 100, 2) if closed_count > 0 else None

                avg_holding_days = None
                if closed_count > 0:
                    holding_series = pd.to_numeric(closed_g["持有天數"], errors="coerce").dropna()
                    if len(holding_series) > 0:
                        avg_holding_days = round(float(holding_series.mean()), 2)

                avg_return = None
                max_return = None
                min_return = None
                if closed_count > 0:
                    return_series = pd.to_numeric(closed_g["報酬%"], errors="coerce").dropna()
                    if len(return_series) > 0:
                        avg_return = round(float(return_series.mean()), 2)
                        max_return = round(float(return_series.max()), 2)
                        min_return = round(float(return_series.min()), 2)

                row_values = [
                    broker,
                    combo_name,
                    include_text,
                    "是",
                    event_counts["A"],
                    event_counts["B"],
                    event_counts["C"],
                    event_counts["D"],
                    event_counts["E"],
                    combo_event_count,
                    closed_count,
                    open_count,
                    win_count,
                    loss_count,
                    flat_count,
                    "-" if win_rate is None else f"{win_rate:.2f}%",
                    "-" if avg_holding_days is None else avg_holding_days,
                    "-" if avg_return is None else f"{avg_return:+.2f}%",
                    "-" if max_return is None else f"{max_return:+.2f}%",
                    "-" if min_return is None else f"{min_return:+.2f}%",
                ]
            else:
                row_values = [
                    broker,
                    combo_name,
                    include_text,
                    "否",
                    event_counts["A"],
                    event_counts["B"],
                    event_counts["C"],
                    event_counts["D"],
                    event_counts["E"],
                    0,
                    0,
                    0,
                    0,
                    0,
                    0,
                    "-",
                    "-",
                    "-",
                    "-",
                    "-",
                ]

            for col_idx, value in enumerate(row_values, 1):
                cell = ws.cell(current_row, col_idx)
                cell.value = value
                cell.font = Font(color="000000")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = normal_border

                if col_idx == 4:
                    if value == "是":
                        cell.fill = PatternFill("solid", fgColor="EAF2F8")
                    else:
                        cell.fill = GRAY
                else:
                    cell.fill = WHITE

            ws.row_dimensions[current_row].height = 22
            current_row += 1

        current_row += 1

    col_widths = [16, 8, 14, 14, 10, 10, 10, 10, 10, 12, 12, 12, 10, 10, 10, 10, 14, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

def fmt_ratio_value(numerator, denominator):
    try:
        numerator = float(numerator)
        denominator = float(denominator)
        if denominator == 0:
            return "-"
        return f"{(numerator / denominator * 100):.2f}%"
    except Exception:
        return "-"

def collect_recent_trade_date_sets(items, cutoff_dt):
    trade_dates = set()

    for item in items:
        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_dates.add(normalize_date_str(row_dict["日期"]))

    sorted_dates = sorted(trade_dates)
    last_5_dates = set(sorted_dates[-5:])
    last_20_dates = set(sorted_dates[-20:])

    return last_5_dates, last_20_dates

def write_recent_warrant_amount_ranking_sheet(wb, items):
    ws = wb.create_sheet("近兩月買賣金額排行")

    headers = [
        "排名",
        "權證代號",
        "權證名稱",
        "標的股",
        "標的名稱",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "近20日淨買進金額",
        "近5日淨買進金額",
        "近20日占比",
        "近5日占比",
        "買進分點",
        "最近買進日",
    ]

    ws.append(headers)

    cutoff_dt = datetime.today() - timedelta(days=RECENT_RANKING_DAYS)
    last_5_dates, last_20_dates = collect_recent_trade_date_sets(items, cutoff_dt)
    ranking_map = {}

    for item in items:
        warrant_code = item["warrant_code"]
        warrant_name = item["warrant_name"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")
        broker_label = item["broker_label"]

        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_date_str = normalize_date_str(row_dict["日期"])
            buy_amount = int(row_dict["買進金額"])
            sell_amount = int(row_dict["賣出金額"])

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            key = (warrant_code, warrant_name, underlying_code, underlying_name)

            if key not in ranking_map:
                ranking_map[key] = {
                    "權證代號": warrant_code,
                    "權證名稱": warrant_name,
                    "標的股": underlying_code,
                    "標的名稱": underlying_name,
                    "買進金額": 0,
                    "賣出金額": 0,
                    "淨買進金額": 0,
                    "近20日淨買進金額": 0,
                    "近5日淨買進金額": 0,
                    "分點買進金額": {},
                    "分點賣出金額": {},
                    "最近買進日": "",
                }

            rec = ranking_map[key]
            net_amount = buy_amount - sell_amount

            rec["買進金額"] += buy_amount
            rec["賣出金額"] += sell_amount
            rec["淨買進金額"] += net_amount

            if trade_date_str in last_20_dates:
                rec["近20日淨買進金額"] += net_amount

            if trade_date_str in last_5_dates:
                rec["近5日淨買進金額"] += net_amount

            if buy_amount > 0:
                rec["分點買進金額"][broker_label] = rec["分點買進金額"].get(broker_label, 0) + buy_amount
                if not rec["最近買進日"] or trade_date_str > rec["最近買進日"]:
                    rec["最近買進日"] = trade_date_str

            if sell_amount > 0:
                rec["分點賣出金額"][broker_label] = rec["分點賣出金額"].get(broker_label, 0) + sell_amount

    ranking_rows = [
        rec for rec in ranking_map.values()
        if rec["淨買進金額"] > 0
    ]

    ranking_rows = sorted(
        ranking_rows,
        key=lambda x: x["淨買進金額"],
        reverse=True
    )[:20]

    for rank, rec in enumerate(ranking_rows, 1):
        broker_net_rows = []

        for broker, buy_amount in rec["分點買進金額"].items():
            sell_amount = rec["分點賣出金額"].get(broker, 0)
            net_amount = buy_amount - sell_amount

            if net_amount > 0:
                broker_net_rows.append((broker, net_amount))

        broker_net_rows = sorted(
            broker_net_rows,
            key=lambda x: x[1],
            reverse=True
        )

        if broker_net_rows:
            broker_text = "；".join([f"{broker}({fmt_amount(amount)})" for broker, amount in broker_net_rows])
        else:
            broker_text = "-"

        ws.append([
            rank,
            rec["權證代號"],
            rec["權證名稱"],
            rec["標的股"],
            rec["標的名稱"] or "-",
            fmt_amount(rec["買進金額"]),
            fmt_amount(rec["賣出金額"]),
            fmt_amount(rec["淨買進金額"]),
            fmt_amount(rec["近20日淨買進金額"]),
            fmt_amount(rec["近5日淨買進金額"]),
            fmt_ratio_value(rec["近20日淨買進金額"], rec["淨買進金額"]),
            fmt_ratio_value(rec["近5日淨買進金額"], rec["淨買進金額"]),
            broker_text,
            rec["最近買進日"] or "-",
        ])

    col_widths = [8, 12, 24, 10, 12, 16, 16, 18, 18, 18, 12, 12, 70, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"

def write_underlying_broker_count_ranking_sheet(wb, items):
    ws = wb.create_sheet("近兩月分點數排行")

    headers = [
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "買進分點數",
        "買進分點清單",
        "近兩月買進金額",
        "近兩月賣出金額",
        "近兩月淨買進金額",
        "近20日淨買進金額",
        "近5日淨買進金額",
        "近20日占比",
        "近5日占比",
        "最近買進日",
    ]

    ws.append(headers)

    cutoff_dt = datetime.today() - timedelta(days=RECENT_RANKING_DAYS)
    last_5_dates, last_20_dates = collect_recent_trade_date_sets(items, cutoff_dt)
    underlying_map = {}

    for item in items:
        warrant_code = item["warrant_code"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")
        broker_label = item["broker_label"]

        if not underlying_code:
            continue

        df = item["df"]

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            trade_dt = parse_date(row_dict["日期"])

            if not trade_dt:
                continue

            if trade_dt < cutoff_dt:
                continue

            trade_date_str = normalize_date_str(row_dict["日期"])
            buy_amount = int(row_dict["買進金額"])
            sell_amount = int(row_dict["賣出金額"])
            buy_shares = int(row_dict["買進股數"])
            sell_shares = int(row_dict["賣出股數"])

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            if underlying_code not in underlying_map:
                underlying_map[underlying_code] = {
                    "標的股": underlying_code,
                    "標的名稱": underlying_name,
                    "權證集合": set(),
                    "分點資料": {},
                }

            rec = underlying_map[underlying_code]
            if not rec["標的名稱"] and underlying_name:
                rec["標的名稱"] = underlying_name

            broker_rec = rec["分點資料"].setdefault(broker_label, {
                "買進金額": 0,
                "賣出金額": 0,
                "淨買進金額": 0,
                "近20日淨買進金額": 0,
                "近5日淨買進金額": 0,
                "買進股數": 0,
                "賣出股數": 0,
                "淨買進股數": 0,
                "最近買進日": "",
                "權證集合": set(),
            })

            net_amount = buy_amount - sell_amount
            net_shares = buy_shares - sell_shares

            broker_rec["權證集合"].add(warrant_code)
            broker_rec["買進金額"] += buy_amount
            broker_rec["賣出金額"] += sell_amount
            broker_rec["淨買進金額"] += net_amount
            broker_rec["買進股數"] += buy_shares
            broker_rec["賣出股數"] += sell_shares
            broker_rec["淨買進股數"] += net_shares

            if trade_date_str in last_20_dates:
                broker_rec["近20日淨買進金額"] += net_amount

            if trade_date_str in last_5_dates:
                broker_rec["近5日淨買進金額"] += net_amount

            if buy_amount > 0:
                if not broker_rec["最近買進日"] or trade_date_str > broker_rec["最近買進日"]:
                    broker_rec["最近買進日"] = trade_date_str

    ranking_rows = []

    for rec in underlying_map.values():
        active_broker_rows = []
        active_warrants = set()
        active_buy_amount = 0
        active_sell_amount = 0
        active_net_amount = 0
        active_20_net_amount = 0
        active_5_net_amount = 0
        latest_buy_date = ""

        for broker, broker_rec in rec["分點資料"].items():
            # 分點若已出清，淨買進股數 <= 0，不列入分點數排行與金額統計。
            if broker_rec["淨買進股數"] <= 0:
                continue

            # 若金額面也沒有正向淨買進，避免已賣出獲利但零庫存的分點被列入。
            if broker_rec["淨買進金額"] <= 0:
                continue

            active_broker_rows.append((broker, broker_rec["淨買進金額"]))
            active_warrants.update(broker_rec["權證集合"])
            active_buy_amount += broker_rec["買進金額"]
            active_sell_amount += broker_rec["賣出金額"]
            active_net_amount += broker_rec["淨買進金額"]
            active_20_net_amount += broker_rec["近20日淨買進金額"]
            active_5_net_amount += broker_rec["近5日淨買進金額"]

            if broker_rec["最近買進日"] and (not latest_buy_date or broker_rec["最近買進日"] > latest_buy_date):
                latest_buy_date = broker_rec["最近買進日"]

        if active_net_amount <= 0:
            continue

        if not active_broker_rows:
            continue

        active_broker_rows = sorted(
            active_broker_rows,
            key=lambda x: x[1],
            reverse=True
        )

        rec["權證檔數"] = len(active_warrants)
        rec["買進分點數"] = len(active_broker_rows)
        rec["買進分點清單"] = "；".join([f"{broker}({fmt_amount(amount)})" for broker, amount in active_broker_rows])
        rec["近兩月買進金額"] = active_buy_amount
        rec["近兩月賣出金額"] = active_sell_amount
        rec["近兩月淨買進金額"] = active_net_amount
        rec["近20日淨買進金額"] = active_20_net_amount
        rec["近5日淨買進金額"] = active_5_net_amount
        rec["最近買進日"] = latest_buy_date

        ranking_rows.append(rec)

    ranking_rows = sorted(
        ranking_rows,
        key=lambda x: (x["買進分點數"], x["近兩月淨買進金額"]),
        reverse=True
    )[:20]

    for rank, rec in enumerate(ranking_rows, 1):
        ws.append([
            rank,
            rec["標的股"],
            rec["標的名稱"] or "-",
            rec["權證檔數"],
            rec["買進分點數"],
            rec["買進分點清單"],
            fmt_amount(rec["近兩月買進金額"]),
            fmt_amount(rec["近兩月賣出金額"]),
            fmt_amount(rec["近兩月淨買進金額"]),
            fmt_amount(rec["近20日淨買進金額"]),
            fmt_amount(rec["近5日淨買進金額"]),
            fmt_ratio_value(rec["近20日淨買進金額"], rec["近兩月淨買進金額"]),
            fmt_ratio_value(rec["近5日淨買進金額"], rec["近兩月淨買進金額"]),
            rec["最近買進日"] or "-",
        ])

    col_widths = [8, 10, 12, 10, 12, 70, 16, 16, 18, 18, 18, 12, 12, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"

def write_broker_query_sheet(wb, items):
    """
    新增「券商查詢」工作表：
    可用下拉選單選擇券商，並顯示該券商目前「標的股」相關權證總淨買進金額前 15 名。
    不是單一權證排名，而是同一券商買進同一標的股底下所有相關認購權證的合計排名。

    排名邏輯改回：淨買進金額由高到低排序。
    淨買進金額 = 買進金額 - 賣出金額，代表扣掉賣出後仍留在裡面的推估金額。
    這版不使用 FILTER / SORTBY，改為先在隱藏工作表預先整理每個券商前 15 名，
    再用 INDEX / MATCH 查詢，提高 Excel 相容性。
    """
    data_ws = wb.create_sheet("券商查詢資料")
    data_ws.sheet_state = "hidden"

    data_headers = [
        "分點",
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "權證清單",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "買進張數",
        "賣出張數",
        "淨買進張數",
        "最近買進日",
        "查詢鍵",
    ]
    data_ws.append(data_headers)

    broker_underlying_map = {}

    for item in items:
        broker_label = item["broker_label"]
        warrant_code = item["warrant_code"]
        warrant_name = item["warrant_name"]
        underlying_code = item["underlying_code"]
        underlying_name = item.get("underlying_name", "")

        if not underlying_code:
            continue

        df = item["df"]

        buy_amount = int(df["買進金額"].sum()) if not df.empty else 0
        sell_amount = int(df["賣出金額"].sum()) if not df.empty else 0
        buy_shares = int(df["買進股數"].sum()) if not df.empty else 0
        sell_shares = int(df["賣出股數"].sum()) if not df.empty else 0

        if buy_amount <= 0 and sell_amount <= 0:
            continue

        net_amount = buy_amount - sell_amount
        net_shares = buy_shares - sell_shares

        latest_buy_date = ""
        if not df.empty:
            buy_dates = df[df["買進金額"] > 0]["日期"].dropna().tolist()
            if buy_dates:
                latest_buy_date = max([normalize_date_str(d) for d in buy_dates])

        key = (broker_label, underlying_code)

        if key not in broker_underlying_map:
            broker_underlying_map[key] = {
                "分點": broker_label,
                "標的股": underlying_code,
                "標的名稱": underlying_name,
                "權證集合": set(),
                "權證清單": [],
                "買進金額": 0,
                "賣出金額": 0,
                "淨買進金額": 0,
                "買進股數": 0,
                "賣出股數": 0,
                "淨買進股數": 0,
                "最近買進日": "",
            }

        rec = broker_underlying_map[key]

        if not rec["標的名稱"] and underlying_name:
            rec["標的名稱"] = underlying_name

        rec["權證集合"].add(warrant_code)

        warrant_label = f"{warrant_code} {warrant_name}"
        if warrant_label not in rec["權證清單"]:
            rec["權證清單"].append(warrant_label)

        rec["買進金額"] += buy_amount
        rec["賣出金額"] += sell_amount
        rec["淨買進金額"] += net_amount
        rec["買進股數"] += buy_shares
        rec["賣出股數"] += sell_shares
        rec["淨買進股數"] += net_shares

        if latest_buy_date and (not rec["最近買進日"] or latest_buy_date > rec["最近買進日"]):
            rec["最近買進日"] = latest_buy_date

    broker_map = {}

    for rec in broker_underlying_map.values():
        # 券商查詢主排序改回「淨買進金額」由高到低，
        # 但列入條件必須同時符合：
        # 1. 買進金額 > 0：代表這個券商確實有買這個標的底下的權證
        # 2. 淨買進金額 > 0：避免買很多但賣更多、實際已轉為淨賣出的標的進榜
        # 3. 淨買進股數 > 0：避免金額面為正但張數面已經接近或完全出清
        #
        # 這樣比較符合「主力還留著沒賣的總金額」這個籌碼意圖，
        # 同時也不會出現淨買進金額為負的標的排在前面。
        if rec["買進金額"] <= 0:
            continue

        if rec["淨買進金額"] <= 0:
            continue

        if rec["淨買進股數"] <= 0:
            continue

        broker_label = rec["分點"]
        broker_map.setdefault(broker_label, []).append(rec)

    broker_order = list(TARGET_PATTERNS.keys())
    active_brokers = sorted(broker_map.keys())

    broker_list = []
    for broker in broker_order:
        if broker in active_brokers and broker not in broker_list:
            broker_list.append(broker)

    for broker in active_brokers:
        if broker not in broker_list:
            broker_list.append(broker)

    for broker in broker_list:
        rows = sorted(
            broker_map.get(broker, []),
            key=lambda x: (x["淨買進金額"], x["最近買進日"], x["買進金額"], x["標的股"]),
            reverse=True
        )[:15]

        for rank, rec in enumerate(rows, 1):
            warrant_list = sorted(rec["權證清單"])

            data_ws.append([
                broker,
                rank,
                rec["標的股"],
                rec["標的名稱"] or "-",
                len(rec["權證集合"]),
                "；".join(warrant_list),
                rec["買進金額"],
                rec["賣出金額"],
                rec["淨買進金額"],
                rec["買進股數"] // 1000,
                rec["賣出股數"] // 1000,
                rec["淨買進股數"] // 1000,
                rec["最近買進日"],
                f"{broker}|{rank}",
            ])

    broker_start_col = 16  # P 欄
    data_ws.cell(1, broker_start_col).value = "券商清單"

    for idx, broker in enumerate(broker_list, 2):
        data_ws.cell(idx, broker_start_col).value = broker

    for col_idx, width in enumerate([16, 8, 10, 12, 10, 70, 16, 16, 18, 12, 12, 12, 14, 24], 1):
        data_ws.column_dimensions[get_column_letter(col_idx)].width = width

    data_ws.column_dimensions[get_column_letter(broker_start_col)].width = 18

    ws = wb.create_sheet("券商查詢")

    ws["A1"] = "券商標的股淨買進金額前 15 名查詢"
    ws.merge_cells("A1:M1")
    ws["A1"].font = Font(bold=True, size=14, color="000000")
    ws["A1"].fill = YELLOW
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "選擇券商"
    ws["A2"].font = Font(bold=True, color="000000")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = GRAY

    ws["B2"] = broker_list[0] if broker_list else ""
    ws["B2"].font = Font(bold=True, color="000000")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B2"].fill = PatternFill("solid", fgColor="EAF2F8")

    ws["D2"] = "排序邏輯：同一券商買進同一標的股底下所有相關權證合計；先排除淨買進金額<=0或淨買進張數<=0，再依淨買進金額由高到低排序。"
    ws.merge_cells("D2:M2")
    ws["D2"].font = Font(color="666666")
    ws["D2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if broker_list:
        broker_formula = f"'券商查詢資料'!$P$2:$P${len(broker_list) + 1}"
        dv = DataValidation(type="list", formula1=broker_formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws["B2"])

    headers = [
        "排名",
        "標的股",
        "標的名稱",
        "權證檔數",
        "權證清單",
        "買進金額",
        "賣出金額",
        "淨買進金額",
        "買進張數",
        "賣出張數",
        "淨買進張數",
        "最近買進日",
    ]

    header_row = 5

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    max_data_row = max(data_ws.max_row, 2)

    for i in range(15):
        row_idx = header_row + 1 + i
        rank = i + 1

        ws.cell(row_idx, 1).value = rank

        for col_idx in range(2, 13):
            data_col_idx = col_idx + 1
            data_col_letter = get_column_letter(data_col_idx)

            formula = (
                f'=IFERROR(INDEX(券商查詢資料!${data_col_letter}$2:${data_col_letter}${max_data_row},'
                f'MATCH($B$2&"|"&$A{row_idx},券商查詢資料!$N$2:$N${max_data_row},0)),"")'
            )
            ws.cell(row_idx, col_idx).value = formula

    col_widths = [8, 10, 12, 10, 70, 16, 16, 18, 12, 12, 12, 14]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=header_row + 15, min_col=1, max_col=12):
        for cell in row:
            cell.border = normal_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row > header_row:
                cell.font = Font(color="000000")

    for row_idx in range(header_row + 1, header_row + 16):
        ws.row_dimensions[row_idx].height = 32

    ws.row_dimensions[2].height = 24
    ws.freeze_panes = "A6"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

def collect_stock_abcde_query_rows(a_events, b_events, c_events, d_events, e_events=None):
    """整理股票查詢頁使用的全部 A/B/C/D/E 事件紀錄。"""
    rows = []

    for event_code, events in iter_amount_class_event_groups(
        a_events,
        b_events,
        c_events,
        d_events,
        e_events,
    ):
        for ev in events or []:
            event_date = normalize_date_str(
                ev.get("事件日")
                or ev.get("結束日")
                or ev.get("起始日")
                or ""
            )
            reduce_date = normalize_date_str(ev.get("減碼日", "")) if ev.get("減碼日") else ""
            exit_date = normalize_date_str(ev.get("出清日", "")) if ev.get("出清日") else ""
            exit_return = ev.get("出清獲利%")

            if exit_date:
                current_status = "已出清"
            elif reduce_date:
                current_status = "已減碼未出清"
            else:
                current_status = "目前持有"

            rows.append({
                "事件類型": ev.get(
                    "事件類型",
                    f"{event_code}-{AMOUNT_CLASS_LABELS.get(event_code, '事件')}",
                ),
                "事件代碼": event_code,
                "標的股": str(ev.get("標的股", "")).strip(),
                "標的名稱": str(ev.get("標的名稱", "")).strip(),
                "分點": str(ev.get("分點", "")).strip(),
                "分點名稱": str(ev.get("分點名稱", "")).strip(),
                "券商代號": str(ev.get("券商代號", "")).strip(),
                "事件日": event_date,
                "目前狀態": current_status,
                "結果": calc_result_tag(exit_return),
                "單日累積買進金額": _safe_stat_amount(
                    ev.get("單日累積買進金額", ev.get("買超金額", 0))
                ),
                "買超張數": ev.get("買超張數", 0),
                "涵蓋權證數": ev.get("涵蓋權證數", 0),
                "權證清單": ev.get("權證清單", ""),
                "最大單筆金額": _safe_stat_amount(ev.get("最大單筆金額", 0)),
                "最大單筆權證": ev.get("最大單筆權證", ""),
                "減碼日": reduce_date or "-",
                "減碼賣出金額": _safe_stat_amount(ev.get("減碼賣出金額", 0)) if ev.get("減碼賣出金額") is not None else "-",
                "減碼獲利%": fmt_pct(ev.get("減碼獲利%")),
                "出清日": exit_date or "-",
                "出清賣出金額": _safe_stat_amount(ev.get("出清賣出金額", 0)) if ev.get("出清賣出金額") is not None else "-",
                "出清獲利%": fmt_pct(exit_return),
                "持有天數": fmt_num(ev.get("持有天數")),
            })

    rows.sort(
        key=lambda row: (
            -((parse_date(row.get("事件日", "")) or datetime.min).toordinal()),
            str(row.get("標的股", "")),
            str(row.get("事件代碼", "")),
            str(row.get("分點", "")),
        )
    )
    return rows

def write_stock_abcde_query_sheet(wb, a_events, b_events, c_events, d_events, e_events=None):
    """
    建立股票 A/B/C/D/E 查詢頁。

    使用方式：
    - 在「股票ABCDE查詢」B2 輸入股號或股名。
    - 查詢結果會列出目前快取保存範圍內，該標的曾出現過的全部 A/B/C/D/E 事件。
    - RUN_MODE=1 不同步這兩張工作表，避免精選五分點結果覆蓋全分點查詢資料。
    """
    data_title = "股票ABCDE查詢資料"
    query_title = "股票ABCDE查詢"
    rows = collect_stock_abcde_query_rows(
        a_events,
        b_events,
        c_events,
        d_events,
        e_events,
    )

    headers = [
        "事件類型",
        "事件代碼",
        "標的股",
        "標的名稱",
        "分點",
        "分點名稱",
        "券商代號",
        "事件日",
        "目前狀態",
        "結果",
        "單日累積買進金額",
        "買超張數",
        "涵蓋權證數",
        "權證清單",
        "最大單筆金額",
        "最大單筆權證",
        "減碼日",
        "減碼賣出金額",
        "減碼獲利%",
        "出清日",
        "出清賣出金額",
        "出清獲利%",
        "持有天數",
    ]

    data_ws = wb.create_sheet(data_title)
    data_ws.append(headers)
    for row in rows:
        data_ws.append([row.get(header, "") for header in headers])

    data_widths = [
        18, 10, 10, 14, 16, 18, 12, 12, 14, 10,
        18, 12, 12, 70, 16, 28, 12, 16, 14, 12,
        16, 14, 12,
    ]
    style_sheet(data_ws, data_widths)
    data_ws.freeze_panes = "A2"
    data_ws.auto_filter.ref = data_ws.dimensions
    data_ws.sheet_state = "hidden"

    query_ws = wb.create_sheet(query_title)
    query_ws["A1"] = "股票 A/B/C/D/E 現有與歷史事件查詢"
    query_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    query_ws["A1"].font = Font(bold=True, size=14, color="000000")
    query_ws["A1"].fill = YELLOW
    query_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    query_ws.row_dimensions[1].height = 28

    query_ws["A2"] = "輸入股名或股號"
    query_ws["A2"].font = Font(bold=True, color="000000")
    query_ws["A2"].fill = GRAY
    query_ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    query_ws["B2"] = ""
    query_ws["B2"].font = Font(bold=True, color="000000")
    query_ws["B2"].fill = PatternFill("solid", fgColor="FFF2CC")
    query_ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

    query_ws["D2"] = (
        "可輸入完整或部分股號／股名，例如 2330、台積電。"
        "查詢範圍是目前程式快取仍保留並重建出的全部 A～E 事件，包含目前持有、已減碼與已出清紀錄。"
    )
    query_ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=len(headers))
    query_ws["D2"].font = Font(color="666666")
    query_ws["D2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    query_ws.row_dimensions[2].height = 34

    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = query_ws.cell(header_row, col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_data_row = max(data_ws.max_row, 2)
    last_col_letter = get_column_letter(len(headers))
    stock_code_col = get_column_letter(headers.index("標的股") + 1)
    stock_name_col = get_column_letter(headers.index("標的名稱") + 1)

    query_ws["A6"] = (
        f'=IF($B$2="","",IFERROR('
        f'FILTER(\'{data_title}\'!$A$2:${last_col_letter}${max_data_row},'
        f'(ISNUMBER(SEARCH($B$2,\'{data_title}\'!${stock_code_col}$2:${stock_code_col}${max_data_row})))+'
        f'(ISNUMBER(SEARCH($B$2,\'{data_title}\'!${stock_name_col}$2:${stock_name_col}${max_data_row})))), '
        f'"查無符合資料"))'
    )

    # 預留 FILTER 溢出空間：Google Sheet 同步時會依 Excel 實際尺寸 resize。
    # 使用空字串而不是 None，確保工作簿儲存並重新 load_workbook() 後仍保留工作表尺寸；
    # 同步到 Google Sheet 時仍會寫成空白值，不會阻擋 FILTER 陣列結果展開。
    reserve_rows = max(data_ws.max_row + 10, 200)
    for row_idx in range(7, 7 + reserve_rows):
        query_ws.cell(row_idx, 1).value = ""
    query_ws.cell(6 + reserve_rows, len(headers)).value = ""

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for row_idx in [1, 2, header_row, 6]:
        for col_idx in range(1, len(headers) + 1):
            cell = query_ws.cell(row_idx, col_idx)
            cell.border = normal_border
            if row_idx in (header_row, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, width in enumerate(data_widths, 1):
        query_ws.column_dimensions[get_column_letter(col_idx)].width = width

    query_ws.freeze_panes = "A6"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

def write_price_status_sheet(wb, price_cache):
    ws = wb.create_sheet("價格抓取狀態")

    headers = [
        "代號",
        "價格筆數",
        "第一筆日期",
        "最後筆日期",
        "狀態",
    ]

    ws.append(headers)

    rows = []

    seen = set()

    for code, prices in price_cache.items():
        code = str(code).strip()

        if not code or code in seen:
            continue

        seen.add(code)

        valid_dates = sorted([d for d, p in prices.items() if p is not None and p > 0])

        if valid_dates:
            rows.append([
                code,
                len(valid_dates),
                valid_dates[0],
                valid_dates[-1],
                "OK",
            ])
        else:
            rows.append([
                code,
                0,
                "-",
                "-",
                "NO DATA",
            ])

    rows = sorted(rows, key=lambda x: (x[4], x[0]))

    for row in rows:
        ws.append(row)

    col_widths = [12, 12, 14, 14, 12]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border

            if cell.column == 5 and cell.value == "NO DATA":
                cell.fill = GREEN

        ws.row_dimensions[row[0].row].height = 20

    ws.freeze_panes = "A2"

def write_color_legend_sheet(wb):
    ws = wb.create_sheet("顏色說明")

    headers = ["色塊", "顏色用途", "出現位置", "說明"]
    ws.append(headers)

    legend_rows = [
        {
            "fill": YELLOW,
            "用途": "標頭 / 欄位名稱",
            "位置": "所有工作表",
            "說明": "用於表格欄位名稱與重要標題列。",
        },
        {
            "fill": RED,
            "用途": "勝 / 上漲",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表該追蹤日為正報酬或上漲狀態；依台股習慣使用紅色表示上漲。",
        },
        {
            "fill": GREEN,
            "用途": "敗 / 下跌或未上漲",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表該追蹤日為負報酬、下跌或未形成正報酬；依台股習慣使用綠色表示下跌。",
        },
        {
            "fill": BLUE,
            "用途": "減碼",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表程式偵測到該分點後續有賣出，但尚未完全出清。",
        },
        {
            "fill": ORANGE,
            "用途": "出清",
            "位置": "A、B、C、D 工作表的 D+1～D+20 欄位",
            "說明": "代表依 FIFO 推估，該筆事件已經被完全賣出。",
        },
        {
            "fill": WHITE,
            "border": "profit",
            "用途": "粗紅色外框",
            "位置": "A、B、C、D 工作表的「減碼獲利% / 出清獲利%」欄位",
            "說明": "代表該筆事件已減碼或出清，且實際獲利% > 0。",
        },
        {
            "fill": WHITE,
            "border": "loss",
            "用途": "粗綠色外框",
            "位置": "A、B、C、D 工作表的「減碼獲利% / 出清獲利%」欄位",
            "說明": "代表該筆事件已減碼或出清，且實際獲利% < 0。",
        },
        {
            "fill": GRAY,
            "用途": "分點區隔列",
            "位置": "勝率統計工作表",
            "說明": "用於區隔不同分點，讓每個分點的統計區塊更清楚。",
        },
        {
            "fill": PatternFill("solid", fgColor="EAF2F8"),
            "用途": "全部-A+B+C+D+E合併",
            "位置": "勝率統計工作表",
            "說明": "代表該分點 A、B、C、D 四類事件合併後的統計列。",
        },
        {
            "fill": WHITE,
            "用途": "一般資料列",
            "位置": "勝率統計工作表",
            "說明": "一般事件類型統計列，沒有特殊狀態標記。",
        },
    ]

    for row_info in legend_rows:
        ws.append(["", row_info["用途"], row_info["位置"], row_info["說明"]])
        row_idx = ws.max_row
        ws.cell(row_idx, 1).fill = row_info["fill"]

        if row_info.get("border"):
            side = PROFIT_EXIT_SIDE if row_info.get("border") == "profit" else LOSS_EXIT_SIDE
            ws.cell(row_idx, 1).border = Border(
                left=side,
                right=side,
                top=side,
                bottom=side,
            )

    col_widths = [12, 24, 34, 70]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows(min_row=2):
        legend_name = str(row[1].value).strip()
        is_profit_outline_legend = legend_name == "粗紅色外框"
        is_loss_outline_legend = legend_name == "粗綠色外框"

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            if is_profit_outline_legend and cell.column == 1:
                cell.border = Border(
                    left=PROFIT_EXIT_SIDE,
                    right=PROFIT_EXIT_SIDE,
                    top=PROFIT_EXIT_SIDE,
                    bottom=PROFIT_EXIT_SIDE,
                )
            elif is_loss_outline_legend and cell.column == 1:
                cell.border = Border(
                    left=LOSS_EXIT_SIDE,
                    right=LOSS_EXIT_SIDE,
                    top=LOSS_EXIT_SIDE,
                    bottom=LOSS_EXIT_SIDE,
                )
            else:
                cell.border = normal_border

        ws.row_dimensions[row[0].row].height = 36

    ws.freeze_panes = "A2"

def _fmt_pct_text(value, signed=True):
    if value is None:
        return "-"
    try:
        v = float(value)
    except Exception:
        return "-"
    return f"{v:+.2f}%" if signed else f"{v:.2f}%"

def _near10_window_dates(target_date=None, window_days=None):
    # 主流程已確認 FinMind 目標交易日；明確傳入日期時直接使用，避免再呼叫外部價格來源解析交易日。
    target_dt = parse_date(target_date) if target_date else None
    if target_dt:
        target_date = target_dt.strftime("%Y/%m/%d")
    else:
        target_date = normalize_top15_target_date(target_date)
        target_dt = parse_date(target_date)

    if not target_dt:
        target_dt = datetime.today()
        target_date = target_dt.strftime("%Y/%m/%d")

    if window_days is None:
        window_days = BROKER_10D_DETAIL_DAYS

    window_days = max(int(window_days), 1)
    start_dt = target_dt - timedelta(days=window_days - 1)
    start_date = start_dt.strftime("%Y/%m/%d")
    period_text = f"{start_date} ～ {target_date}"
    return target_date, target_dt, start_date, start_dt, window_days, period_text

def _sell_return_summary_for_item(item, start_dt, target_dt, fallback_price=None):
    """
    用同一分點 + 同一權證的 FinMind 歷史資料估算近10日賣出實現報酬。

    規則：
    - 權證不可當沖：同一天先賣舊庫存，再把當日買進加入庫存。
    - 賣出成本優先用 FIFO 持倉成本估算。
    - 若近10日賣出找不到足夠舊庫存成本，改用可取得的成本備援估算，避免賣超報酬率空白：
      1. 優先用該權證歷史已出現買進均價
      2. 其次用最新權證價格
      3. 最後用當筆賣出均價，讓報酬率保守落在 0%
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty:
        return {
            "revenue": 0.0,
            "cost": 0.0,
            "unmatched_amount": 0.0,
            "unmatched_qty": 0.0,
        }

    df = df.copy()
    if "日期" not in df.columns:
        return {
            "revenue": 0.0,
            "cost": 0.0,
            "unmatched_amount": 0.0,
            "unmatched_qty": 0.0,
        }

    df["dt_parsed"] = df["日期"].map(parse_date)
    df = df.dropna(subset=["dt_parsed"]).sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []
    revenue = 0.0
    cost = 0.0
    unmatched_amount = 0.0
    unmatched_qty = 0.0

    historical_buy_amount = 0.0
    historical_buy_qty = 0.0

    try:
        fallback_price = float(fallback_price) if fallback_price is not None else None
    except Exception:
        fallback_price = None

    if fallback_price is not None and fallback_price <= 0:
        fallback_price = None

    def fallback_unit_cost(sell_price):
        if historical_buy_qty > 0 and historical_buy_amount > 0:
            return historical_buy_amount / historical_buy_qty
        if fallback_price is not None and fallback_price > 0:
            return fallback_price
        if sell_price and sell_price > 0:
            return sell_price
        return None

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        dt = row_dict.get("dt_parsed")
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

        in_window = bool(dt and start_dt <= dt <= target_dt)

        # 權證不能當沖：同一天先處理賣出，只能扣舊庫存。
        if sell_amount > 0:
            if sell_qty <= 0:
                # API 異常時仍保留數據，不讓近10日賣超報酬率變空白。
                if in_window:
                    revenue += sell_amount
                    cost += sell_amount
                sell_qty = 0
            else:
                sell_price = sell_amount / sell_qty
                sell_left = sell_qty
                allocated_revenue = 0.0
                allocated_cost = 0.0

                for lot in lots:
                    if sell_left <= 0:
                        break
                    if lot.get("剩餘股數", 0) <= 0:
                        continue

                    alloc = min(sell_left, lot["剩餘股數"])
                    if alloc <= 0:
                        continue

                    lot["剩餘股數"] -= alloc
                    sell_left -= alloc
                    allocated_revenue += alloc * sell_price
                    allocated_cost += alloc * lot["均價"]

                if sell_left > 0:
                    fallback_cost_price = fallback_unit_cost(sell_price)
                    if fallback_cost_price is not None and fallback_cost_price > 0:
                        allocated_revenue += sell_left * sell_price
                        allocated_cost += sell_left * fallback_cost_price
                    else:
                        unmatched_qty += sell_left
                        unmatched_amount += sell_left * sell_price

                if in_window:
                    revenue += allocated_revenue
                    cost += allocated_cost

        # 同日買進放到賣出後面，避免當沖錯扣。
        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
            })
            historical_buy_qty += buy_qty
            historical_buy_amount += buy_amount

    return {
        "revenue": revenue,
        "cost": cost,
        "unmatched_amount": unmatched_amount,
        "unmatched_qty": unmatched_qty,
    }

def _recent_buy_position_summary_for_item(item, start_dt, target_dt, latest_price=None):
    """
    計算近10日買進 lot 在統計日仍然留下的真實 FIFO 持倉。

    這個函式用完整歷史跑到統計日：
    - 每天先賣出扣 FIFO 舊庫存，再加入當天買進。
    - 只把「買進日落在近10日視窗內」且統計日尚未被賣掉的剩餘 lot 納入買超持倉報酬。
    - 買超報酬 = (有最新價格的剩餘股數 × 最新權證價格 - 有最新價格的剩餘成本) / 有最新價格的剩餘成本。
    - 缺最新權證價格的剩餘部位不納入報酬率，改在明細備註統計省略檔數。
    """
    df = item.get("df", pd.DataFrame())
    if df is None or df.empty:
        return {
            "remaining_qty": 0.0,
            "remaining_cost": 0.0,
            "market_value": 0.0,
            "missing_price": False,
        }

    df = df.copy()
    if "日期" not in df.columns:
        return {
            "remaining_qty": 0.0,
            "remaining_cost": 0.0,
            "market_value": 0.0,
            "missing_price": False,
        }

    df["dt_parsed"] = df["日期"].map(parse_date)
    df = df.dropna(subset=["dt_parsed"])
    df = df[df["dt_parsed"] <= target_dt].sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

    lots = []

    for row in df.itertuples(index=False):
        row_dict = row._asdict()
        dt = row_dict.get("dt_parsed")
        date_str = normalize_date_str(row_dict.get("日期", ""))
        buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
        sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
        buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
        sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

        # 權證不能當沖：同一天先賣出，只扣舊庫存。
        if sell_qty > 0 and sell_amount > 0:
            sell_left = sell_qty
            for lot in lots:
                if sell_left <= 0:
                    break
                if lot.get("剩餘股數", 0) <= 0:
                    continue

                alloc = min(sell_left, lot["剩餘股數"])
                if alloc <= 0:
                    continue

                lot["剩餘股數"] -= alloc
                sell_left -= alloc

        if buy_qty > 0 and buy_amount > 0:
            lots.append({
                "買進日": date_str,
                "股數": buy_qty,
                "剩餘股數": buy_qty,
                "金額": buy_amount,
                "均價": buy_amount / buy_qty if buy_qty else 0,
                "近10日買進": bool(dt and start_dt <= dt <= target_dt),
            })

    remaining_qty = 0.0
    remaining_cost = 0.0

    for lot in lots:
        if not lot.get("近10日買進"):
            continue
        qty = top15_safe_float(lot.get("剩餘股數", 0))
        avg = top15_safe_float(lot.get("均價", 0))
        if qty <= 0 or avg <= 0:
            continue
        remaining_qty += qty
        remaining_cost += qty * avg

    try:
        latest_price = float(latest_price) if latest_price is not None else None
    except Exception:
        latest_price = None

    if latest_price is not None and latest_price > 0 and remaining_qty > 0:
        market_value = remaining_qty * latest_price
        missing_price = False
    else:
        market_value = 0.0
        missing_price = remaining_cost > 0

    return {
        "remaining_qty": remaining_qty,
        "remaining_cost": remaining_cost,
        "market_value": market_value,
        "missing_price": missing_price,
    }

def build_10d_broker_underlying_detail_rows(items, price_cache, target_date=None):
    """
    建立「快取_近10日分點買賣明細」。

    統計單位：資料範圍 + 統計日期 + 單一分點 + 標的股。
    同一分點同一標的底下的所有權證會先完整合併，再計算買賣金額、買超 / 賣超報酬與勝率。
    """
    if not BROKER_10D_DETAIL_ENABLED:
        return None

    print("【Step 4e】建立近10日分點買賣明細快取...")

    if not items:
        print("  ⚠️ 近10日分點買賣明細：沒有 items 資料")
        return []

    target_date, target_dt, start_date, start_dt, window_days, period_text = _near10_window_dates(target_date)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = get_result_data_scope()

    agg = {}

    for item in items:
        df = item.get("df", pd.DataFrame())
        if df is None or df.empty:
            continue

        warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
        if not warrant_code:
            continue

        warrant_name = str(item.get("warrant_name", "")).strip()
        underlying_code = str(item.get("underlying_code", "")).strip()
        underlying_name = str(item.get("underlying_name", "")).strip()
        broker_label = str(item.get("broker_label", "")).strip()
        broker_name = str(item.get("broker_name", "")).strip()
        broker_code = str(item.get("broker_code", "")).strip()

        if not underlying_code or not broker_label:
            continue

        key = (broker_label, broker_name, broker_code, underlying_code)
        rec = agg.setdefault(key, {
            "資料範圍": scope,
            "統計日期": add_gsheet_text_prefix(target_date),
            "統計期間": period_text,
            "統計天數": window_days,
            "第一筆日期": "",
            "最後筆日期": "",
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "標的股": underlying_code,
            "標的名稱": underlying_name,
            "近10日買進股數": 0.0,
            "近10日買進金額": 0.0,
            "近10日賣出股數": 0.0,
            "近10日賣出金額": 0.0,
            "日期集合": set(),
            "權證": {},
            "買超報酬加權分子": 0.0,
            "買超報酬權重": 0.0,
            "買超剩餘股數": 0.0,
            "買超剩餘成本": 0.0,
            "買超目前市值": 0.0,
            "買超報酬有效權證數": 0,
            "買超報酬缺價權證數": 0,
            "買超缺價省略權證清單": [],
            "賣超實現賣出金額": 0.0,
            "賣超實現成本": 0.0,
            "賣超成本不足金額": 0.0,
            "賣超成本不足股數": 0.0,
            "更新時間": update_time,
            "run_id": run_id,
        })

        if underlying_name and not rec.get("標的名稱"):
            rec["標的名稱"] = underlying_name

        item_buy_qty = 0.0
        item_buy_amount = 0.0
        item_sell_qty = 0.0
        item_sell_amount = 0.0
        item_dates = set()

        for row in df.itertuples(index=False):
            row_dict = row._asdict()
            date_str = normalize_date_str(row_dict.get("日期", ""))
            dt = parse_date(date_str)

            if not dt or dt < start_dt or dt > target_dt:
                continue

            buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
            sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))
            buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
            sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))

            if buy_qty <= 0 and sell_qty <= 0 and buy_amount <= 0 and sell_amount <= 0:
                continue

            rec["近10日買進股數"] += buy_qty
            rec["近10日買進金額"] += buy_amount
            rec["近10日賣出股數"] += sell_qty
            rec["近10日賣出金額"] += sell_amount
            rec["日期集合"].add(date_str)

            item_buy_qty += buy_qty
            item_buy_amount += buy_amount
            item_sell_qty += sell_qty
            item_sell_amount += sell_amount
            item_dates.add(date_str)

        if item_buy_amount <= 0 and item_sell_amount <= 0:
            continue

        warrant_rec = rec["權證"].setdefault(warrant_code, {
            "權證代號": warrant_code,
            "權證名稱": warrant_name,
            "買進金額": 0.0,
            "賣出金額": 0.0,
            "買進股數": 0.0,
            "賣出股數": 0.0,
            "日期集合": set(),
            "最新權證價格": None,
            "最新權證價格日": "",
        })
        warrant_rec["買進金額"] += item_buy_amount
        warrant_rec["賣出金額"] += item_sell_amount
        warrant_rec["買進股數"] += item_buy_qty
        warrant_rec["賣出股數"] += item_sell_qty
        warrant_rec["日期集合"].update(item_dates)

        latest_price, latest_price_date = get_latest_price_info_on_or_before(price_cache, warrant_code, target_date)
        if latest_price is not None:
            warrant_rec["最新權證價格"] = latest_price
            warrant_rec["最新權證價格日"] = latest_price_date

        if item_buy_qty > 0 and item_buy_amount > 0:
            position_summary = _recent_buy_position_summary_for_item(
                item,
                start_dt,
                target_dt,
                latest_price=latest_price,
            )
            remaining_cost = top15_safe_float(position_summary.get("remaining_cost", 0.0))
            remaining_qty = top15_safe_float(position_summary.get("remaining_qty", 0.0))
            market_value = top15_safe_float(position_summary.get("market_value", 0.0))

            if remaining_cost > 0:
                if position_summary.get("missing_price"):
                    # 缺最新權證價格的部位直接從買超平均報酬分子 / 分母排除，
                    # 避免用 0 市值把整體報酬率不合理壓低。
                    rec["買超報酬缺價權證數"] += 1
                    rec.setdefault("買超缺價省略權證清單", []).append(f"{warrant_code} {warrant_name}".strip())
                    warrant_rec["買超報酬省略原因"] = "缺最新權證價格，未納入買超平均報酬"
                else:
                    rec["買超剩餘成本"] += remaining_cost
                    rec["買超剩餘股數"] += remaining_qty
                    rec["買超目前市值"] += market_value
                    rec["買超報酬權重"] += remaining_cost
                    rec["買超報酬加權分子"] += market_value - remaining_cost
                    rec["買超報酬有效權證數"] += 1

            warrant_rec["近10日買進剩餘股數"] = remaining_qty
            warrant_rec["近10日買進剩餘成本"] = remaining_cost
            warrant_rec["近10日買進目前市值"] = market_value

        if item_sell_qty > 0 and item_sell_amount > 0:
            sell_summary = _sell_return_summary_for_item(
                item,
                start_dt,
                target_dt,
                fallback_price=latest_price,
            )
            rec["賣超實現賣出金額"] += sell_summary.get("revenue", 0.0)
            rec["賣超實現成本"] += sell_summary.get("cost", 0.0)
            rec["賣超成本不足金額"] += sell_summary.get("unmatched_amount", 0.0)
            rec["賣超成本不足股數"] += sell_summary.get("unmatched_qty", 0.0)

    rows = []

    for rec in agg.values():
        dates = sorted(rec.get("日期集合", set()))
        if not dates:
            continue

        buy_amount = float(rec.get("近10日買進金額", 0) or 0)
        sell_amount = float(rec.get("近10日賣出金額", 0) or 0)
        buy_qty = float(rec.get("近10日買進股數", 0) or 0)
        sell_qty = float(rec.get("近10日賣出股數", 0) or 0)
        net_buy_amount = buy_amount - sell_amount
        net_sell_amount = sell_amount - buy_amount
        net_buy_qty = buy_qty - sell_qty
        net_sell_qty = sell_qty - buy_qty

        buy_return_pct = None
        buy_return_weight = float(rec.get("買超報酬權重", 0) or 0)
        buy_return_numerator = float(rec.get("買超報酬加權分子", 0) or 0)

        if buy_return_weight > 0:
            # 買超平均報酬只使用「有最新價格」的剩餘部位；缺價部位已在上方省略。
            buy_return_pct = buy_return_numerator / buy_return_weight * 100
        elif buy_amount > 0:
            # 近10日有買進但沒有任何可估剩餘部位時，仍給出 0.00%，避免勝率報酬空白。
            buy_return_pct = 0.0

        sell_return_pct = None
        sell_realized_cost = float(rec.get("賣超實現成本", 0) or 0)
        sell_realized_revenue = float(rec.get("賣超實現賣出金額", 0) or 0)

        if sell_realized_cost > 0:
            sell_return_pct = (sell_realized_revenue - sell_realized_cost) / sell_realized_cost * 100
        elif sell_amount > 0:
            # API 賣出資料異常或成本仍不足時，保守帶入 0.00%，確保賣超報酬率一定有值。
            sell_return_pct = 0.0

        if net_buy_amount > 0:
            direction = "買超"
            win_return_pct = buy_return_pct
        elif net_sell_amount > 0:
            direction = "賣超"
            win_return_pct = sell_return_pct
        else:
            direction = "買賣平衡"
            win_return_pct = buy_return_pct if buy_return_pct is not None else sell_return_pct

        if win_return_pct is None:
            win_return_pct = 0.0

        # 分點層級加權報酬使用「主要方向淨額」作為權重：
        # - 買超標的：使用近10日淨買超金額
        # - 賣超標的：使用近10日淨賣超金額
        # - 買賣平衡：退回使用買進 / 賣出金額較大者
        # 這樣可以避免單純平均讓小金額標的過度影響整體分點表現。
        primary_return_weight = net_buy_amount if direction == "買超" else net_sell_amount if direction == "賣超" else max(buy_amount, sell_amount)
        if primary_return_weight <= 0:
            primary_return_weight = max(buy_amount, sell_amount, 0.0)

        broker_buy_return_weight = net_buy_amount if net_buy_amount > 0 else 0.0
        broker_sell_return_weight = net_sell_amount if net_sell_amount > 0 else 0.0

        if win_return_pct > 0:
            result = "勝"
        else:
            # 使用者指定 0% 要算賠錢，所以 <= 0 都是敗。
            result = "敗"

        notes = []
        missing_price_count = int(rec.get("買超報酬缺價權證數", 0) or 0)

        if missing_price_count > 0:
            notes.append(f"買超報酬已省略缺價權證 {missing_price_count} 檔")

        if buy_amount > 0 and buy_return_weight <= 0:
            if missing_price_count > 0:
                notes.append("買超剩餘部位全數缺價或無可估價格，買超報酬以 0.00% 保守帶入")
            else:
                notes.append("近10日買進目前無可估剩餘部位，買超報酬以 0.00% 帶入")

        if sell_amount > 0 and sell_realized_cost <= 0:
            notes.append("賣超成本不足，賣超報酬以 0.00% 帶入")

        note_text = "；".join(notes)

        underlying_10d_return_pct = None
        underlying_start_price = None
        underlying_end_price = None
        underlying_start_price_date = ""
        underlying_end_price_date = ""
        underlying_code_for_price = str(rec.get("標的股", "") or "").strip()

        if underlying_code_for_price:
            underlying_start_price, underlying_start_price_date = get_latest_price_info_on_or_before(
                price_cache,
                underlying_code_for_price,
                start_date,
            )
            underlying_end_price, underlying_end_price_date = get_latest_price_info_on_or_before(
                price_cache,
                underlying_code_for_price,
                target_date,
            )

            try:
                if (
                    underlying_start_price is not None
                    and underlying_end_price is not None
                    and float(underlying_start_price) > 0
                    and float(underlying_end_price) > 0
                ):
                    underlying_10d_return_pct = (float(underlying_end_price) - float(underlying_start_price)) / float(underlying_start_price) * 100
            except Exception:
                underlying_10d_return_pct = None

        warrant_rows = []
        warrant_json = []
        for w in sorted(
            rec.get("權證", {}).values(),
            key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
            reverse=True,
        ):
            warrant_rows.append(
                f"{w.get('權證代號', '')} {w.get('權證名稱', '')}"
                f"｜買{_fmt_wan_text(w.get('買進金額', 0))}／賣{_fmt_wan_text(w.get('賣出金額', 0))}"
            )
            warrant_json.append({
                "權證代號": w.get("權證代號", ""),
                "權證名稱": w.get("權證名稱", ""),
                "買進金額": round(float(w.get("買進金額", 0) or 0), 0),
                "賣出金額": round(float(w.get("賣出金額", 0) or 0), 0),
                "買進股數": round(float(w.get("買進股數", 0) or 0), 0),
                "賣出股數": round(float(w.get("賣出股數", 0) or 0), 0),
                "最新權證價格": w.get("最新權證價格"),
                "最新權證價格日": w.get("最新權證價格日", ""),
                "近10日買進剩餘股數": round(float(w.get("近10日買進剩餘股數", 0) or 0), 0),
                "近10日買進剩餘成本": round(float(w.get("近10日買進剩餘成本", 0) or 0), 0),
                "近10日買進目前市值": round(float(w.get("近10日買進目前市值", 0) or 0), 0),
                "買超報酬省略原因": w.get("買超報酬省略原因", ""),
                "日期數": len(w.get("日期集合", set())),
            })

        rows.append({
            "資料範圍": rec.get("資料範圍", scope),
            "統計日期": add_gsheet_text_prefix(target_date),
            "統計期間": period_text,
            "統計天數": window_days,
            "有效日期數": len(dates),
            "第一筆日期": add_gsheet_text_prefix(dates[0]),
            "最後筆日期": add_gsheet_text_prefix(dates[-1]),
            "分點": rec.get("分點", ""),
            "分點名稱": rec.get("分點名稱", ""),
            "券商代號": rec.get("券商代號", ""),
            "標的股": rec.get("標的股", ""),
            "標的名稱": rec.get("標的名稱", ""),
            "標的10日漲跌幅%": _fmt_pct_text(underlying_10d_return_pct, signed=True) if underlying_10d_return_pct is not None else "-",
            "現股10日報酬率%": _fmt_pct_text(underlying_10d_return_pct, signed=True) if underlying_10d_return_pct is not None else "-",
            "標的10日起始價": round(float(underlying_start_price), 4) if underlying_start_price is not None else "",
            "標的10日收盤價": round(float(underlying_end_price), 4) if underlying_end_price is not None else "",
            "標的10日起始價格日": add_gsheet_text_prefix(underlying_start_price_date) if underlying_start_price_date else "",
            "標的10日收盤價格日": add_gsheet_text_prefix(underlying_end_price_date) if underlying_end_price_date else "",
            "買賣方向": direction,
            "近10日買進股數": round(buy_qty, 0),
            "近10日買進金額": round(buy_amount, 0),
            "近10日賣出股數": round(sell_qty, 0),
            "近10日賣出金額": round(sell_amount, 0),
            "近10日淨買超股數": round(net_buy_qty, 0),
            "近10日淨買超金額": round(net_buy_amount, 0),
            "近10日淨賣超股數": round(net_sell_qty, 0),
            "近10日淨賣超金額": round(net_sell_amount, 0),
            "涉及權證檔數": len(rec.get("權證", {})),
            "權證清單": "；".join(warrant_rows),
            "買超平均報酬%": _fmt_pct_text(buy_return_pct, signed=True),
            "買超剩餘股數": round(float(rec.get("買超剩餘股數", 0) or 0), 0),
            "買超剩餘成本": round(float(rec.get("買超剩餘成本", 0) or 0), 0),
            "買超目前市值": round(float(rec.get("買超目前市值", 0) or 0), 0),
            "買超報酬有效權證數": int(rec.get("買超報酬有效權證數", 0) or 0),
            "買超報酬缺價權證數": int(rec.get("買超報酬缺價權證數", 0) or 0),
            "賣超平均報酬%": _fmt_pct_text(sell_return_pct, signed=True),
            "賣超實現賣出金額": round(float(rec.get("賣超實現賣出金額", 0) or 0), 0),
            "賣超實現成本": round(float(rec.get("賣超實現成本", 0) or 0), 0),
            "賣超成本不足金額": round(float(rec.get("賣超成本不足金額", 0) or 0), 0),
            "用於勝率報酬%": _fmt_pct_text(win_return_pct, signed=True),
            "判定": result,
            "備註": note_text,
            "分點近10日勝率": "-",
            "分點近10日勝筆數": 0,
            "分點近10日敗筆數": 0,
            "分點近10日加權平均報酬%": "-",
            "分點近10日買超加權平均報酬%": "-",
            "分點近10日賣超加權平均報酬%": "-",
            "分點近10日加權平均勝報酬%": "-",
            "分點近10日加權平均敗報酬%": "-",
            "分點近10日盈虧比": "-",
            "分點近10日加權期望值%": "-",
            "分點近10日加權報酬權重金額": 0,
            "更新時間": rec.get("更新時間", update_time),
            "run_id": rec.get("run_id", run_id),
            "權證明細_JSON": json.dumps(warrant_json, ensure_ascii=False),
            "_分點統計報酬率數值": win_return_pct,
            "_分點統計權重": primary_return_weight,
            "_分點統計買超報酬率數值": buy_return_pct,
            "_分點統計買超權重": broker_buy_return_weight,
            "_分點統計賣超報酬率數值": sell_return_pct,
            "_分點統計賣超權重": broker_sell_return_weight,
        })

    broker_stats = {}
    for row in rows:
        broker_key = (row.get("分點", ""), row.get("券商代號", ""))
        stat = broker_stats.setdefault(broker_key, {
            "win": 0,
            "loss": 0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
        })

        if row.get("判定") == "勝":
            stat["win"] += 1
        elif row.get("判定") == "敗":
            stat["loss"] += 1

        ret = row.get("_分點統計報酬率數值")
        weight = top15_safe_float(row.get("_分點統計權重", 0), 0.0)

        if ret is not None and weight > 0:
            ret = top15_safe_float(ret, None)
            if ret is not None:
                stat["return_weighted_numerator"] += ret * weight
                stat["return_weight"] += weight

                if ret > 0:
                    stat["win_return_weighted_numerator"] += ret * weight
                    stat["win_return_weight"] += weight
                else:
                    stat["loss_return_weighted_numerator"] += ret * weight
                    stat["loss_return_weight"] += weight

        buy_ret = row.get("_分點統計買超報酬率數值")
        buy_weight = top15_safe_float(row.get("_分點統計買超權重", 0), 0.0)
        if buy_ret is not None and buy_weight > 0:
            buy_ret = top15_safe_float(buy_ret, None)
            if buy_ret is not None:
                stat["buy_return_weighted_numerator"] += buy_ret * buy_weight
                stat["buy_return_weight"] += buy_weight

        sell_ret = row.get("_分點統計賣超報酬率數值")
        sell_weight = top15_safe_float(row.get("_分點統計賣超權重", 0), 0.0)
        if sell_ret is not None and sell_weight > 0:
            sell_ret = top15_safe_float(sell_ret, None)
            if sell_ret is not None:
                stat["sell_return_weighted_numerator"] += sell_ret * sell_weight
                stat["sell_return_weight"] += sell_weight

    for row in rows:
        broker_key = (row.get("分點", ""), row.get("券商代號", ""))
        stat = broker_stats.get(broker_key, {
            "win": 0,
            "loss": 0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
        })

        total = stat["win"] + stat["loss"]
        win_rate_value = stat["win"] / total * 100 if total else None

        avg_return = (
            stat["return_weighted_numerator"] / stat["return_weight"]
            if stat["return_weight"] > 0 else None
        )
        avg_buy_return = (
            stat["buy_return_weighted_numerator"] / stat["buy_return_weight"]
            if stat["buy_return_weight"] > 0 else None
        )
        avg_sell_return = (
            stat["sell_return_weighted_numerator"] / stat["sell_return_weight"]
            if stat["sell_return_weight"] > 0 else None
        )
        avg_win_return = (
            stat["win_return_weighted_numerator"] / stat["win_return_weight"]
            if stat["win_return_weight"] > 0 else None
        )
        avg_loss_return = (
            stat["loss_return_weighted_numerator"] / stat["loss_return_weight"]
            if stat["loss_return_weight"] > 0 else None
        )

        profit_loss_ratio = None
        if avg_win_return is not None and avg_win_return > 0 and avg_loss_return is not None and avg_loss_return < 0:
            profit_loss_ratio = avg_win_return / abs(avg_loss_return)

        expectancy = None
        if total > 0 and (avg_win_return is not None or avg_loss_return is not None):
            win_rate_decimal = stat["win"] / total
            loss_rate_decimal = stat["loss"] / total
            expectancy = win_rate_decimal * (avg_win_return or 0.0) + loss_rate_decimal * (avg_loss_return or 0.0)
        elif avg_return is not None:
            expectancy = avg_return

        row["分點近10日勝筆數"] = stat["win"]
        row["分點近10日敗筆數"] = stat["loss"]
        row["分點近10日勝率"] = _fmt_pct_text(win_rate_value, signed=False) if win_rate_value is not None else "-"
        row["分點近10日加權平均報酬%"] = _fmt_pct_text(avg_return, signed=True)
        row["分點近10日買超加權平均報酬%"] = _fmt_pct_text(avg_buy_return, signed=True)
        row["分點近10日賣超加權平均報酬%"] = _fmt_pct_text(avg_sell_return, signed=True)
        row["分點近10日加權平均勝報酬%"] = _fmt_pct_text(avg_win_return, signed=True)
        row["分點近10日加權平均敗報酬%"] = _fmt_pct_text(avg_loss_return, signed=True)
        row["分點近10日盈虧比"] = f"{profit_loss_ratio:.2f}" if profit_loss_ratio is not None else "-"
        row["分點近10日加權期望值%"] = _fmt_pct_text(expectancy, signed=True)
        row["分點近10日加權報酬權重金額"] = round(float(stat.get("return_weight", 0.0) or 0.0), 0)

    rows = sorted(
        rows,
        key=lambda r: (
            str(r.get("分點", "")),
            -abs(float(r.get("近10日淨買超金額", 0) or 0)),
            str(r.get("標的股", "")),
        )
    )

    print(f"  ✅ 近10日分點買賣明細完成：{len(rows):,} 筆，統計期間 {period_text}")
    return rows

def write_10d_broker_underlying_detail_sheet(wb, rows):
    """寫入近10日分點買賣明細。rows=None 代表不建立工作表。"""
    if rows is None:
        return

    ws = wb.create_sheet(BROKER_10D_DETAIL_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期",
        "分點", "分點名稱", "券商代號", "標的股", "標的名稱",
        "標的10日漲跌幅%", "現股10日報酬率%", "標的10日起始價", "標的10日收盤價", "標的10日起始價格日", "標的10日收盤價格日",
        "買賣方向",
        "近10日買進股數", "近10日買進金額", "近10日賣出股數", "近10日賣出金額",
        "近10日淨買超股數", "近10日淨買超金額", "近10日淨賣超股數", "近10日淨賣超金額",
        "涉及權證檔數", "權證清單",
        "買超平均報酬%", "買超剩餘股數", "買超剩餘成本", "買超目前市值", "買超報酬有效權證數", "買超報酬缺價權證數",
        "賣超平均報酬%", "賣超實現賣出金額", "賣超實現成本", "賣超成本不足金額",
        "用於勝率報酬%", "判定", "備註", "分點近10日勝率", "分點近10日勝筆數", "分點近10日敗筆數",
        "分點近10日加權平均報酬%", "分點近10日買超加權平均報酬%", "分點近10日賣超加權平均報酬%",
        "分點近10日加權平均勝報酬%", "分點近10日加權平均敗報酬%", "分點近10日盈虧比",
        "分點近10日加權期望值%", "分點近10日加權報酬權重金額",
        "更新時間", "run_id", "權證明細_JSON",
    ]

    ws.append(headers)

    for row in rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [
        12, 12, 24, 10, 12, 12, 12,
        14, 18, 12, 10, 14,
        16, 16, 14, 14, 14, 14,
        10,
        14, 16, 14, 16,
        16, 18, 16, 18,
        12, 72,
        14, 14, 16, 16, 14, 14,
        14, 16, 16, 16,
        14, 10, 36, 14, 14, 14,
        16, 20, 20, 18, 18, 12, 16, 18,
        20, 18, 90,
    ]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 26

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    direction_col = header_map.get("買賣方向")

    for row in ws.iter_rows(min_row=2):
        direction = str(row[direction_col - 1].value or "").strip() if direction_col else ""
        row_fill = RED if direction == "買超" else GREEN if direction == "賣超" else WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 32

    ws.freeze_panes = "A2"

def _parse_pct_text_to_float(value):
    try:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        s = str(value).strip()
        if not s or s == "-":
            return None
        s = s.replace("%", "").replace("+", "").replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None

def build_10d_broker_winrate_rank_rows(broker_10d_detail_rows):
    """
    建立「快取_近10日分點勝率排行」。

    重要規則：
    1. 只在 RUN_MODE=2 完整分點清單模式執行。
    2. RUN_MODE=1 精選分點模式直接回傳 None，build_excel 不會建立該工作表，
       因此 upload_excel_to_google_sheet() 也不會更新 / 清空 Google Sheet 上原本的同名工作表。
    3. 統計來源沿用「快取_近10日分點買賣明細」已完成的分點 + 標的層級結果，
       不重新改動 A/B/C/D/E、TOP15 或每日賣出明細邏輯。
    4. 勝率排序：勝率高到低，其次統計筆數多到少、加權報酬高到低、總交易金額高到低。
    """
    if not BROKER_10D_WINRATE_RANK_ENABLED:
        return None

    if RUN_MODE != 2:
        print("  ✅ RUN_MODE=1 精選分點模式：略過近10日分點勝率排行工作表，避免動到 Google Sheet 既有資料。")
        return None

    if broker_10d_detail_rows is None:
        return None

    print("【Step 4f】建立近10日分點勝率排行（RUN_MODE=2 專用）...")

    if not broker_10d_detail_rows:
        print("  ⚠️ 近10日分點勝率排行：沒有近10日分點明細資料")
        return []

    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = get_result_data_scope()

    agg = {}

    for row in broker_10d_detail_rows:
        broker_label = str(row.get("分點", "")).strip()
        broker_name = str(row.get("分點名稱", "")).strip()
        broker_code = str(row.get("券商代號", "")).strip()

        if not broker_label:
            continue

        key = (broker_label, broker_code)
        rec = agg.setdefault(key, {
            "資料範圍": row.get("資料範圍", scope),
            "統計日期": row.get("統計日期", ""),
            "統計期間": row.get("統計期間", ""),
            "統計天數": row.get("統計天數", ""),
            "分點": broker_label,
            "分點名稱": broker_name,
            "券商代號": broker_code,
            "win": 0,
            "loss": 0,
            "buy_amount": 0.0,
            "sell_amount": 0.0,
            "net_buy_amount": 0.0,
            "net_sell_amount": 0.0,
            "return_weighted_numerator": 0.0,
            "return_weight": 0.0,
            "buy_return_weighted_numerator": 0.0,
            "buy_return_weight": 0.0,
            "sell_return_weighted_numerator": 0.0,
            "sell_return_weight": 0.0,
            "win_return_weighted_numerator": 0.0,
            "win_return_weight": 0.0,
            "loss_return_weighted_numerator": 0.0,
            "loss_return_weight": 0.0,
            "underlying_codes": set(),
            "warrant_count": 0,
            "detail_rows": [],
            "date_set": set(),
            "first_date": "",
            "last_date": "",
        })

        if broker_name and not rec.get("分點名稱"):
            rec["分點名稱"] = broker_name
        if broker_code and not rec.get("券商代號"):
            rec["券商代號"] = broker_code

        result = str(row.get("判定", "")).strip()
        if result == "勝":
            rec["win"] += 1
        elif result == "敗":
            rec["loss"] += 1

        buy_amount = top15_safe_float(row.get("近10日買進金額", 0), 0.0)
        sell_amount = top15_safe_float(row.get("近10日賣出金額", 0), 0.0)
        net_buy_amount = max(top15_safe_float(row.get("近10日淨買超金額", 0), 0.0), 0.0)
        net_sell_amount = max(top15_safe_float(row.get("近10日淨賣超金額", 0), 0.0), 0.0)
        warrant_count = int(top15_safe_float(row.get("涉及權證檔數", 0), 0.0) or 0)

        rec["buy_amount"] += buy_amount
        rec["sell_amount"] += sell_amount
        rec["net_buy_amount"] += net_buy_amount
        rec["net_sell_amount"] += net_sell_amount
        rec["warrant_count"] += warrant_count

        underlying_code = str(row.get("標的股", "")).strip()
        if underlying_code:
            rec["underlying_codes"].add(underlying_code)

        for d in [row.get("第一筆日期", ""), row.get("最後筆日期", "")]:
            d = normalize_date_str(strip_gsheet_text_prefix(d))
            if parse_date(d):
                rec["date_set"].add(d)

        ret = row.get("_分點統計報酬率數值")
        if ret is None:
            ret = _parse_pct_text_to_float(row.get("用於勝率報酬%", ""))
        weight = top15_safe_float(row.get("_分點統計權重", 0), 0.0)
        if weight <= 0:
            direction = str(row.get("買賣方向", "")).strip()
            if direction == "買超":
                weight = net_buy_amount
            elif direction == "賣超":
                weight = net_sell_amount
            else:
                weight = max(buy_amount, sell_amount, 0.0)

        if ret is not None and weight > 0:
            ret = top15_safe_float(ret, None)
            if ret is not None:
                rec["return_weighted_numerator"] += ret * weight
                rec["return_weight"] += weight
                if ret > 0:
                    rec["win_return_weighted_numerator"] += ret * weight
                    rec["win_return_weight"] += weight
                else:
                    rec["loss_return_weighted_numerator"] += ret * weight
                    rec["loss_return_weight"] += weight

        buy_ret = row.get("_分點統計買超報酬率數值")
        if buy_ret is None:
            buy_ret = _parse_pct_text_to_float(row.get("買超平均報酬%", ""))
        buy_weight = top15_safe_float(row.get("_分點統計買超權重", 0), 0.0)
        if buy_weight <= 0:
            buy_weight = net_buy_amount
        if buy_ret is not None and buy_weight > 0:
            buy_ret = top15_safe_float(buy_ret, None)
            if buy_ret is not None:
                rec["buy_return_weighted_numerator"] += buy_ret * buy_weight
                rec["buy_return_weight"] += buy_weight

        sell_ret = row.get("_分點統計賣超報酬率數值")
        if sell_ret is None:
            sell_ret = _parse_pct_text_to_float(row.get("賣超平均報酬%", ""))
        sell_weight = top15_safe_float(row.get("_分點統計賣超權重", 0), 0.0)
        if sell_weight <= 0:
            sell_weight = net_sell_amount
        if sell_ret is not None and sell_weight > 0:
            sell_ret = top15_safe_float(sell_ret, None)
            if sell_ret is not None:
                rec["sell_return_weighted_numerator"] += sell_ret * sell_weight
                rec["sell_return_weight"] += sell_weight

        detail_amount = max(buy_amount + sell_amount, abs(net_buy_amount), abs(net_sell_amount), 0.0)
        rec["detail_rows"].append({
            "標的股": row.get("標的股", ""),
            "標的名稱": row.get("標的名稱", ""),
            "買賣方向": row.get("買賣方向", ""),
            "判定": result,
            "用於勝率報酬%": row.get("用於勝率報酬%", ""),
            "近10日買進金額": round(buy_amount, 0),
            "近10日賣出金額": round(sell_amount, 0),
            "近10日淨買超金額": round(net_buy_amount, 0),
            "近10日淨賣超金額": round(net_sell_amount, 0),
            "涉及權證檔數": warrant_count,
            "權證清單": row.get("權證清單", ""),
            "排序金額": detail_amount,
        })

    rows = []

    for rec in agg.values():
        total = int(rec.get("win", 0) or 0) + int(rec.get("loss", 0) or 0)
        if total <= 0:
            continue

        win_rate_value = rec["win"] / total * 100 if total else None
        avg_return = rec["return_weighted_numerator"] / rec["return_weight"] if rec["return_weight"] > 0 else None
        avg_buy_return = rec["buy_return_weighted_numerator"] / rec["buy_return_weight"] if rec["buy_return_weight"] > 0 else None
        avg_sell_return = rec["sell_return_weighted_numerator"] / rec["sell_return_weight"] if rec["sell_return_weight"] > 0 else None
        avg_win_return = rec["win_return_weighted_numerator"] / rec["win_return_weight"] if rec["win_return_weight"] > 0 else None
        avg_loss_return = rec["loss_return_weighted_numerator"] / rec["loss_return_weight"] if rec["loss_return_weight"] > 0 else None

        profit_loss_ratio = None
        if avg_win_return is not None and avg_win_return > 0 and avg_loss_return is not None and avg_loss_return < 0:
            profit_loss_ratio = avg_win_return / abs(avg_loss_return)

        expectancy = None
        if total > 0 and (avg_win_return is not None or avg_loss_return is not None):
            win_rate_decimal = rec["win"] / total
            loss_rate_decimal = rec["loss"] / total
            expectancy = win_rate_decimal * (avg_win_return or 0.0) + loss_rate_decimal * (avg_loss_return or 0.0)
        elif avg_return is not None:
            expectancy = avg_return

        date_values = sorted(rec.get("date_set", set()))
        first_date = date_values[0] if date_values else ""
        last_date = date_values[-1] if date_values else ""

        details_sorted = sorted(
            rec.get("detail_rows", []),
            key=lambda x: float(x.get("排序金額", 0) or 0),
            reverse=True,
        )
        main_underlying_rows = []
        detail_json = []
        for d in details_sorted:
            amount = max(
                float(d.get("近10日買進金額", 0) or 0),
                float(d.get("近10日賣出金額", 0) or 0),
                float(d.get("近10日淨買超金額", 0) or 0),
                float(d.get("近10日淨賣超金額", 0) or 0),
            )
            main_underlying_rows.append(
                f"{d.get('標的股', '')} {d.get('標的名稱', '')}".strip()
                + f"｜{d.get('買賣方向', '')}｜{d.get('判定', '')}｜{d.get('用於勝率報酬%', '')}｜{_fmt_wan_text(amount)}"
            )
            detail_json.append({k: v for k, v in d.items() if k != "排序金額"})

        total_trade_amount = float(rec.get("buy_amount", 0.0) or 0.0) + float(rec.get("sell_amount", 0.0) or 0.0)

        rows.append({
            "資料範圍": rec.get("資料範圍", scope),
            "統計日期": rec.get("統計日期", ""),
            "統計期間": rec.get("統計期間", ""),
            "統計天數": rec.get("統計天數", ""),
            "有效日期數": len(date_values),
            "第一筆日期": add_gsheet_text_prefix(first_date) if first_date else "",
            "最後筆日期": add_gsheet_text_prefix(last_date) if last_date else "",
            "排名": 0,
            "分點": rec.get("分點", ""),
            "分點名稱": rec.get("分點名稱", ""),
            "券商代號": rec.get("券商代號", ""),
            "近10日勝率": _fmt_pct_text(win_rate_value, signed=False),
            "近10日勝筆數": int(rec.get("win", 0) or 0),
            "近10日敗筆數": int(rec.get("loss", 0) or 0),
            "近10日統計筆數": total,
            "近10日加權平均報酬%": _fmt_pct_text(avg_return, signed=True),
            "近10日買超加權平均報酬%": _fmt_pct_text(avg_buy_return, signed=True),
            "近10日賣超加權平均報酬%": _fmt_pct_text(avg_sell_return, signed=True),
            "近10日加權平均勝報酬%": _fmt_pct_text(avg_win_return, signed=True),
            "近10日加權平均敗報酬%": _fmt_pct_text(avg_loss_return, signed=True),
            "近10日盈虧比": f"{profit_loss_ratio:.2f}" if profit_loss_ratio is not None else "-",
            "近10日加權期望值%": _fmt_pct_text(expectancy, signed=True),
            "近10日加權報酬權重金額": round(float(rec.get("return_weight", 0.0) or 0.0), 0),
            "近10日買進金額": round(float(rec.get("buy_amount", 0.0) or 0.0), 0),
            "近10日賣出金額": round(float(rec.get("sell_amount", 0.0) or 0.0), 0),
            "近10日總交易金額": round(total_trade_amount, 0),
            "近10日淨買超金額": round(float(rec.get("net_buy_amount", 0.0) or 0.0), 0),
            "近10日淨賣超金額": round(float(rec.get("net_sell_amount", 0.0) or 0.0), 0),
            "涉及標的數": len(rec.get("underlying_codes", set())),
            "涉及權證檔數": int(rec.get("warrant_count", 0) or 0),
            "主要交易標的": "；".join(main_underlying_rows[:10]),
            "排序說明": "勝率高到低，其次統計筆數、加權報酬、總交易金額",
            "更新時間": update_time,
            "run_id": run_id,
            "標的明細_JSON": json.dumps(detail_json, ensure_ascii=False),
            "_勝率數值": win_rate_value if win_rate_value is not None else -999999,
            "_加權報酬數值": avg_return if avg_return is not None else -999999,
            "_總交易金額": total_trade_amount,
        })

    rows = sorted(
        rows,
        key=lambda r: (
            -float(r.get("_勝率數值", -999999) or -999999),
            -int(r.get("近10日統計筆數", 0) or 0),
            -float(r.get("_加權報酬數值", -999999) or -999999),
            -float(r.get("_總交易金額", 0) or 0),
            str(r.get("分點", "")),
        )
    )

    for idx, row in enumerate(rows, start=1):
        row["排名"] = idx

    print(f"  ✅ 近10日分點勝率排行完成：{len(rows):,} 個分點")
    return rows

def write_10d_broker_winrate_rank_sheet(wb, rows):
    """寫入近10日分點勝率排行。rows=None 代表不建立工作表。"""
    if rows is None:
        return

    ws = wb.create_sheet(BROKER_10D_WINRATE_RANK_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期", "排名",
        "分點", "分點名稱", "券商代號",
        "近10日勝率", "近10日勝筆數", "近10日敗筆數", "近10日統計筆數",
        "近10日加權平均報酬%", "近10日買超加權平均報酬%", "近10日賣超加權平均報酬%",
        "近10日加權平均勝報酬%", "近10日加權平均敗報酬%", "近10日盈虧比", "近10日加權期望值%",
        "近10日加權報酬權重金額", "近10日買進金額", "近10日賣出金額", "近10日總交易金額",
        "近10日淨買超金額", "近10日淨賣超金額", "涉及標的數", "涉及權證檔數",
        "主要交易標的", "排序說明", "更新時間", "run_id", "標的明細_JSON",
    ]

    ws.append(headers)

    for row in rows or []:
        ws.append([row.get(h, "") for h in headers])

    col_widths = [
        12, 12, 24, 10, 12, 12, 12, 8,
        14, 18, 12,
        12, 12, 12, 12,
        18, 22, 22,
        20, 20, 12, 18,
        18, 16, 16, 16,
        18, 18, 12, 12,
        80, 44, 20, 18, 90,
    ]

    thin_gray = Side(style="thin", color="B7B7B7")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = YELLOW
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = normal_border

    ws.row_dimensions[1].height = 26

    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1])}
    win_rate_col = header_map.get("近10日勝率")

    for row in ws.iter_rows(min_row=2):
        win_rate_value = _parse_pct_text_to_float(row[win_rate_col - 1].value) if win_rate_col else None
        if win_rate_value is not None and win_rate_value >= 70:
            row_fill = RED
        elif win_rate_value is not None and win_rate_value < 50:
            row_fill = GREEN
        else:
            row_fill = WHITE

        for cell in row:
            cell.font = Font(color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
            cell.fill = row_fill

        ws.row_dimensions[row[0].row].height = 32

    ws.freeze_panes = "A2"

def _fmt_wan_text(value):
    try:
        return f"{float(value) / 10000:.1f}萬"
    except Exception:
        return "0.0萬"

def parse_warrant_consensus_selected_brokers():
    """取得近 7／14／21 日共識排名專用的精選分點清單。"""
    if WARRANT_CONSENSUS_SELECTED_BROKERS_ENV:
        labels = [
            x.strip()
            for x in re.split(r"[,;；、\n\r\t]+", WARRANT_CONSENSUS_SELECTED_BROKERS_ENV)
            if x.strip()
        ]
    else:
        labels = list(WARRANT_CONSENSUS_SELECTED_BROKERS_DEFAULT)

    out = []
    for label in labels:
        if label in FULL_TARGET_PATTERNS and label not in out:
            out.append(label)

    return out

def build_7d_warrant_consensus_top15_rows(items, target_date=None):
    """
    建立「快取_近7日權證分點共識TOP15」。

    重要規則：
    1. 只在 RUN_MODE=2 完整分點清單模式執行，確保 13 個指定分點的歷史資料都有被更新。
    2. 實際排名只統計 parse_warrant_consensus_selected_brokers() 回傳的精選分點。
    3. 同一張工作表依序放入近 7 日、近 14 日、近 21 日三組排名資料。
    4. 各期間都以「標的股」層級完整合併同標的全部權證，再各自排序取 TOP15。
    5. 共識買超 TOP15：買進金額 - 賣出金額 > 0，依淨買超金額排序。
    6. 共識賣超 TOP15：賣出金額 - 買進金額 > 0，依淨賣超金額排序。
    7. 近 7／14／21 日沿用原本邏輯，皆以日曆日區間計算。
    """
    if not WARRANT_CONSENSUS_7D_ENABLED:
        return None

    if RUN_MODE != 2:
        print("  ✅ RUN_MODE=1 精選分點模式：略過近7／14／21日精選分點共識TOP15工作表，避免動到 Google Sheet 既有資料。")
        return None

    selected_brokers = parse_warrant_consensus_selected_brokers()
    selected_broker_set = set(selected_brokers)
    selected_broker_codes = {
        str(FULL_FALLBACK[label][1]).strip().lower()
        for label in selected_brokers
        if label in FULL_FALLBACK
    }

    print(
        "【Step 4c】建立近7／14／21日精選分點共識買賣超 TOP15"
        f"（標的層級，共 {len(selected_brokers)} 個分點）..."
    )
    print(f"  ✅ 統計分點：{', '.join(selected_brokers)}")

    if not items:
        print("  ⚠️ 近7／14／21日精選分點共識TOP15：沒有 items 資料")
        return []

    # 主流程已確認 FinMind 目標交易日；明確傳入日期時直接使用，避免再次探測其他資料源。
    target_dt = parse_date(target_date) if target_date else None
    if target_dt:
        target_date = target_dt.strftime("%Y/%m/%d")
    else:
        target_date = normalize_top15_target_date(target_date)
        target_dt = parse_date(target_date)

    if not target_dt:
        target_dt = datetime.today()
        target_date = target_dt.strftime("%Y/%m/%d")

    window_days_list = []
    for days in [
        WARRANT_CONSENSUS_7D_DAYS,
        WARRANT_CONSENSUS_14D_DAYS,
        WARRANT_CONSENSUS_21D_DAYS,
    ]:
        days = max(int(days), 1)
        if days not in window_days_list:
            window_days_list.append(days)

    top_n = max(int(WARRANT_CONSENSUS_7D_TOP_N), 1)
    update_time = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    scope = f"精選{len(selected_brokers)}分點"

    def build_rows_for_window(window_days):
        start_dt = target_dt - timedelta(days=window_days - 1)
        start_date = start_dt.strftime("%Y/%m/%d")
        period_text = f"{start_date} ～ {target_date}"

        # group_key = 標的股。
        # 同標的底下所有權證、所有指定精選分點先完整加總，再做 TOP15 排名。
        agg = {}

        for item in items:
            broker_label = str(item.get("broker_label", "")).strip()
            broker_code = str(item.get("broker_code", "")).strip()
            broker_code_key = broker_code.lower()

            if broker_label not in selected_broker_set and broker_code_key not in selected_broker_codes:
                continue

            df = item.get("df", pd.DataFrame())

            if df is None or df.empty:
                continue

            warrant_code = normalize_warrant_code_for_unique(item.get("warrant_code", ""))
            if not warrant_code:
                continue

            warrant_name = str(item.get("warrant_name", "")).strip()
            raw_underlying_code = str(item.get("underlying_code", "")).strip()
            underlying_name = str(item.get("underlying_name", "")).strip()
            underlying_code = normalize_underlying_code_for_group(raw_underlying_code, underlying_name or warrant_name)
            broker_name = str(item.get("broker_name", "")).strip()

            if not underlying_code:
                continue

            rec = agg.setdefault(underlying_code, {
                "標的股": underlying_code,
                "標的名稱": underlying_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "分點": {},
                "權證": {},
                "日期集合": set(),
            })

            if underlying_name and not rec.get("標的名稱"):
                rec["標的名稱"] = underlying_name

            warrant_rec = rec["權證"].setdefault(warrant_code, {
                "權證代號": warrant_code,
                "權證名稱": warrant_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "日期集合": set(),
            })
            if warrant_name and not warrant_rec.get("權證名稱"):
                warrant_rec["權證名稱"] = warrant_name

            broker_key = (broker_label, broker_name, broker_code)
            broker_rec = rec["分點"].setdefault(broker_key, {
                "分點": broker_label,
                "分點名稱": broker_name,
                "券商代號": broker_code,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "權證": {},
                "日期集合": set(),
            })
            broker_warrant_rec = broker_rec["權證"].setdefault(warrant_code, {
                "權證代號": warrant_code,
                "權證名稱": warrant_name,
                "買進金額": 0.0,
                "賣出金額": 0.0,
                "買進股數": 0.0,
                "賣出股數": 0.0,
                "日期集合": set(),
            })

            for row in df.itertuples(index=False):
                row_dict = row._asdict()
                date_str = normalize_date_str(row_dict.get("日期", ""))
                dt = parse_date(date_str)

                if not dt or dt < start_dt or dt > target_dt:
                    continue

                buy_amount = top15_safe_float(row_dict.get("買進金額", 0))
                sell_amount = top15_safe_float(row_dict.get("賣出金額", 0))
                buy_qty = top15_safe_float(row_dict.get("買進股數", 0))
                sell_qty = top15_safe_float(row_dict.get("賣出股數", 0))

                if buy_amount <= 0 and sell_amount <= 0 and buy_qty <= 0 and sell_qty <= 0:
                    continue

                rec["日期集合"].add(date_str)
                rec["買進金額"] += buy_amount
                rec["賣出金額"] += sell_amount
                rec["買進股數"] += buy_qty
                rec["賣出股數"] += sell_qty

                warrant_rec["日期集合"].add(date_str)
                warrant_rec["買進金額"] += buy_amount
                warrant_rec["賣出金額"] += sell_amount
                warrant_rec["買進股數"] += buy_qty
                warrant_rec["賣出股數"] += sell_qty

                broker_rec["日期集合"].add(date_str)
                broker_rec["買進金額"] += buy_amount
                broker_rec["賣出金額"] += sell_amount
                broker_rec["買進股數"] += buy_qty
                broker_rec["賣出股數"] += sell_qty

                broker_warrant_rec["日期集合"].add(date_str)
                broker_warrant_rec["買進金額"] += buy_amount
                broker_warrant_rec["賣出金額"] += sell_amount
                broker_warrant_rec["買進股數"] += buy_qty
                broker_warrant_rec["賣出股數"] += sell_qty

        records = []

        for rec in agg.values():
            buy_amount = float(rec.get("買進金額", 0) or 0)
            sell_amount = float(rec.get("賣出金額", 0) or 0)
            buy_qty = float(rec.get("買進股數", 0) or 0)
            sell_qty = float(rec.get("賣出股數", 0) or 0)
            net_buy = buy_amount - sell_amount
            net_sell = sell_amount - buy_amount

            if buy_amount <= 0 and sell_amount <= 0:
                continue

            records.append({
                **rec,
                "買進金額": buy_amount,
                "賣出金額": sell_amount,
                "買進股數": buy_qty,
                "賣出股數": sell_qty,
                "淨買超金額": net_buy,
                "淨賣超金額": net_sell,
            })

        # 排名前再做一次標的層級保護性合併。
        merged_records = {}
        for rec in records:
            key = normalize_underlying_code_for_group(rec.get("標的股", ""), rec.get("標的名稱", ""))
            if not key:
                continue

            if key not in merged_records:
                rec["標的股"] = key
                merged_records[key] = rec
                continue

            dst = merged_records[key]
            for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數", "淨買超金額", "淨賣超金額"]:
                dst[numeric_col] = float(dst.get(numeric_col, 0) or 0) + float(rec.get(numeric_col, 0) or 0)

            if not dst.get("標的名稱") and rec.get("標的名稱"):
                dst["標的名稱"] = rec.get("標的名稱")

            dst.setdefault("日期集合", set()).update(rec.get("日期集合", set()))

            for warrant_code, warrant_rec in rec.get("權證", {}).items():
                if warrant_code in dst.get("權證", {}):
                    dw = dst["權證"][warrant_code]
                    for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                        dw[numeric_col] = float(dw.get(numeric_col, 0) or 0) + float(warrant_rec.get(numeric_col, 0) or 0)
                    dw.setdefault("日期集合", set()).update(warrant_rec.get("日期集合", set()))
                else:
                    dst.setdefault("權證", {})[warrant_code] = warrant_rec

            for broker_key, broker_rec in rec.get("分點", {}).items():
                if broker_key not in dst.get("分點", {}):
                    dst.setdefault("分點", {})[broker_key] = broker_rec
                    continue

                db = dst["分點"][broker_key]
                for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                    db[numeric_col] = float(db.get(numeric_col, 0) or 0) + float(broker_rec.get(numeric_col, 0) or 0)
                db.setdefault("日期集合", set()).update(broker_rec.get("日期集合", set()))

                for warrant_code, bw in broker_rec.get("權證", {}).items():
                    if warrant_code in db.get("權證", {}):
                        dbw = db["權證"][warrant_code]
                        for numeric_col in ["買進金額", "賣出金額", "買進股數", "賣出股數"]:
                            dbw[numeric_col] = float(dbw.get(numeric_col, 0) or 0) + float(bw.get(numeric_col, 0) or 0)
                        dbw.setdefault("日期集合", set()).update(bw.get("日期集合", set()))
                    else:
                        db.setdefault("權證", {})[warrant_code] = bw

        records = list(merged_records.values())

        buy_top = [r for r in records if float(r.get("淨買超金額", 0) or 0) > 0]
        sell_top = [r for r in records if float(r.get("淨賣超金額", 0) or 0) > 0]

        buy_top = sorted(
            buy_top,
            key=lambda r: (float(r.get("淨買超金額", 0) or 0), float(r.get("買進金額", 0) or 0)),
            reverse=True,
        )[:top_n]
        sell_top = sorted(
            sell_top,
            key=lambda r: (float(r.get("淨賣超金額", 0) or 0), float(r.get("賣出金額", 0) or 0)),
            reverse=True,
        )[:top_n]

        def make_rank_rows(rank_type, records_for_rank):
            rows = []
            is_buy = rank_type == "共識買超"

            for rank, rec in enumerate(records_for_rank, 1):
                broker_rows = []
                broker_json = []
                same_direction_count = 0
                opposite_direction_count = 0

                for broker_rec in sorted(
                    rec["分點"].values(),
                    key=lambda x: (
                        (float(x.get("買進金額", 0) or 0) - float(x.get("賣出金額", 0) or 0))
                        if is_buy else
                        (float(x.get("賣出金額", 0) or 0) - float(x.get("買進金額", 0) or 0))
                    ),
                    reverse=True,
                ):
                    b_buy = float(broker_rec.get("買進金額", 0) or 0)
                    b_sell = float(broker_rec.get("賣出金額", 0) or 0)
                    b_net_buy = b_buy - b_sell
                    b_net_sell = b_sell - b_buy
                    direction_amount = b_net_buy if is_buy else b_net_sell

                    if direction_amount > 0:
                        same_direction_count += 1
                        broker_rows.append(f"{broker_rec['分點']} {_fmt_wan_text(direction_amount)}")
                    elif direction_amount < 0:
                        opposite_direction_count += 1

                    broker_warrants = []
                    for bw in sorted(
                        broker_rec.get("權證", {}).values(),
                        key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
                        reverse=True,
                    ):
                        broker_warrants.append({
                            "權證代號": bw.get("權證代號", ""),
                            "權證名稱": bw.get("權證名稱", ""),
                            "買進金額": round(float(bw.get("買進金額", 0) or 0), 0),
                            "賣出金額": round(float(bw.get("賣出金額", 0) or 0), 0),
                            "買進股數": round(float(bw.get("買進股數", 0) or 0), 0),
                            "賣出股數": round(float(bw.get("賣出股數", 0) or 0), 0),
                            "日期數": len(bw.get("日期集合", set())),
                        })

                    broker_json.append({
                        "分點": broker_rec["分點"],
                        "分點名稱": broker_rec["分點名稱"],
                        "券商代號": broker_rec["券商代號"],
                        "買進金額": round(b_buy, 0),
                        "賣出金額": round(b_sell, 0),
                        "淨買超金額": round(b_net_buy, 0),
                        "淨賣超金額": round(b_net_sell, 0),
                        "權證檔數": len(broker_rec.get("權證", {})),
                        "日期數": len(broker_rec.get("日期集合", set())),
                        "權證明細": broker_warrants,
                    })

                rank_amount = float(rec.get("淨買超金額", 0) or 0) if is_buy else float(rec.get("淨賣超金額", 0) or 0)
                dates = sorted(rec.get("日期集合", set()))
                warrant_values = sorted(
                    rec.get("權證", {}).values(),
                    key=lambda x: (float(x.get("買進金額", 0) or 0) + float(x.get("賣出金額", 0) or 0)),
                    reverse=True,
                )
                warrant_list = "；".join([
                    f"{w.get('權證代號', '')} {w.get('權證名稱', '')}"
                    f"｜買{_fmt_wan_text(w.get('買進金額', 0))}／賣{_fmt_wan_text(w.get('賣出金額', 0))}"
                    for w in warrant_values
                ])
                top_warrant = warrant_values[0] if warrant_values else {}

                rows.append({
                    "資料範圍": scope,
                    "統計日期": target_date,
                    "統計期間": period_text,
                    "統計天數": window_days,
                    "有效日期數": len(dates),
                    "第一筆日期": dates[0] if dates else "",
                    "最後筆日期": dates[-1] if dates else "",
                    "排名類型": rank_type,
                    "排名": rank,
                    # 保留舊欄位，避免圖片端或舊公式依欄名讀取時壞掉；但實際排名已是標的層級。
                    "權證代號": top_warrant.get("權證代號", ""),
                    "權證名稱": top_warrant.get("權證名稱", "同標的合計"),
                    "標的股": rec.get("標的股", ""),
                    "標的名稱": rec.get("標的名稱", ""),
                    "權證檔數": len(rec.get("權證", {})),
                    "權證清單": warrant_list,
                    "排名金額": round(rank_amount, 0),
                    "買進金額": round(float(rec.get("買進金額", 0) or 0), 0),
                    "賣出金額": round(float(rec.get("賣出金額", 0) or 0), 0),
                    "淨買超金額": round(float(rec.get("淨買超金額", 0) or 0), 0),
                    "淨賣超金額": round(float(rec.get("淨賣超金額", 0) or 0), 0),
                    "買進股數": round(float(rec.get("買進股數", 0) or 0), 0),
                    "賣出股數": round(float(rec.get("賣出股數", 0) or 0), 0),
                    "參與分點數": len(rec.get("分點", {})),
                    "同向分點數": same_direction_count,
                    "反向分點數": opposite_direction_count,
                    "主要同向分點": "；".join(broker_rows[:8]),
                    "完成狀態": "DONE",
                    "更新時間": update_time,
                    "run_id": run_id,
                    "分點明細_JSON": json.dumps(broker_json, ensure_ascii=False),
                })

            return rows

        rows = []
        rows.extend(make_rank_rows("共識買超", buy_top))
        rows.extend(make_rank_rows("共識賣超", sell_top))

        print(
            f"  ✅ 近{window_days}日精選分點共識TOP15完成："
            f"共識買超 {len(buy_top):,} 檔標的，共識賣超 {len(sell_top):,} 檔標的，"
            f"統計期間 {period_text}"
        )
        return rows

    all_rows = []
    for window_days in window_days_list:
        all_rows.extend(build_rows_for_window(window_days))

    return all_rows

def write_7d_warrant_consensus_top15_sheet(wb, rows):
    """
    寫入近 7／14／21 日精選分點共識買賣超 TOP15。

    rows=None 代表不建立工作表；同一張工作表內由上往下建立三張獨立排名表：
    近 7 日、近 14 日、近 21 日。每張排名表都有自己的標題、統計期間與欄位表頭。
    """
    if rows is None:
        return

    ws = wb.create_sheet(WARRANT_CONSENSUS_7D_SHEET)

    headers = [
        "資料範圍", "統計日期", "統計期間", "統計天數", "有效日期數", "第一筆日期", "最後筆日期",
        "排名類型", "排名",
        "權證代號", "權證名稱", "標的股", "標的名稱", "權證檔數", "權證清單",
        "排名金額", "買進金額", "賣出金額", "淨買超金額", "淨賣超金額",
        "買進股數", "賣出股數",
        "參與分點數", "同向分點數", "反向分點數", "主要同向分點",
        "完成狀態", "更新時間", "run_id", "分點明細_JSON",
    ]

    window_order = []
    for days in [
        WARRANT_CONSENSUS_7D_DAYS,
        WARRANT_CONSENSUS_14D_DAYS,
        WARRANT_CONSENSUS_21D_DAYS,
    ]:
        days = max(int(days), 1)
        if days not in window_order:
            window_order.append(days)

    rows_by_window = {days: [] for days in window_order}
    for row in rows or []:
        try:
            days = int(float(row.get("統計天數", 0) or 0))
        except Exception:
            days = 0
        rows_by_window.setdefault(days, []).append(row)

    col_widths = [12, 12, 24, 10, 12, 12, 12, 12, 8, 12, 24, 10, 12, 10, 72, 14, 14, 14, 14, 14, 14, 14, 12, 12, 12, 56, 10, 20, 16, 90]
    thin_gray = Side(style="thin", color="B7B7B7")
    medium_brown = Side(style="medium", color="7F6000")
    normal_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    section_top_border = Border(left=thin_gray, right=thin_gray, top=medium_brown, bottom=thin_gray)

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    selected_brokers = parse_warrant_consensus_selected_brokers()
    selected_brokers_text = "、".join(selected_brokers)
    max_col = len(headers)
    first_header_row = None

    for section_index, days in enumerate(window_order):
        section_rows = rows_by_window.get(days, [])

        if section_index > 0:
            ws.append([""] * max_col)
            ws.row_dimensions[ws.max_row].height = 9

        title_row = 1 if section_index == 0 else ws.max_row + 1
        ws.cell(title_row, 1, f"近{days}日精選{len(selected_brokers)}分點權證共識買賣超 TOP15")
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)
        title_cell = ws.cell(title_row, 1)
        title_cell.font = Font(bold=True, size=14, color="000000")
        title_cell.fill = YELLOW
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = Border(top=medium_brown, bottom=medium_brown)
        ws.row_dimensions[title_row].height = 28

        period_text = str(section_rows[0].get("統計期間", "")).strip() if section_rows else ""
        subtitle_parts = []
        if period_text:
            subtitle_parts.append(f"統計期間：{period_text}")
        subtitle_parts.append(f"統計分點：{selected_brokers_text}")

        subtitle_row = ws.max_row + 1
        ws.cell(subtitle_row, 1, "｜".join(subtitle_parts))
        ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=max_col)
        subtitle_cell = ws.cell(subtitle_row, 1)
        subtitle_cell.font = Font(color="000000")
        subtitle_cell.fill = BLUE
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        subtitle_cell.border = normal_border
        ws.row_dimensions[subtitle_row].height = 32

        header_row = ws.max_row + 1
        ws.append(headers)
        if first_header_row is None:
            first_header_row = header_row

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="000000")
            cell.fill = YELLOW
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = normal_border
        ws.row_dimensions[header_row].height = 24

        if not section_rows:
            empty_row = ws.max_row + 1
            ws.cell(empty_row, 1, f"近{days}日無符合排名資料")
            ws.merge_cells(start_row=empty_row, start_column=1, end_row=empty_row, end_column=max_col)
            empty_cell = ws.cell(empty_row, 1)
            empty_cell.font = Font(color="666666")
            empty_cell.fill = WHITE
            empty_cell.alignment = Alignment(horizontal="center", vertical="center")
            empty_cell.border = normal_border
            ws.row_dimensions[empty_row].height = 24
            continue

        previous_rank_type = None
        for record in section_rows:
            ws.append([record.get(h, "") for h in headers])
            current_row = ws.max_row
            rank_type = str(record.get("排名類型", "")).strip()
            row_fill = RED if rank_type == "共識買超" else GREEN if rank_type == "共識賣超" else WHITE
            use_section_top = rank_type != previous_rank_type

            for cell in ws[current_row]:
                cell.font = Font(
                    bold=use_section_top,
                    color="000000",
                )
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = section_top_border if use_section_top else normal_border
                cell.fill = row_fill

            # 金額與股數欄使用千分位，維持原本報表的閱讀格式。
            for col_name in [
                "排名金額", "買進金額", "賣出金額", "淨買超金額", "淨賣超金額",
                "買進股數", "賣出股數",
            ]:
                col_idx = headers.index(col_name) + 1
                ws.cell(current_row, col_idx).number_format = '#,##0'

            ws.row_dimensions[current_row].height = 30
            previous_rank_type = rank_type

    ws.freeze_panes = f"A{(first_header_row or 1) + 1}"

def build_excel(a_events, b_events, c_events, d_events, e_events, item_map, price_cache, items, output_path, top15_detail_rows=None, top15_consensus_rows=None, warrant_consensus_7d_rows=None, broker_10d_detail_rows=None, broker_10d_winrate_rank_rows=None):
    if workflow_is_repair():
        print("【Step 5】完整修補模式：建立完整結果工作簿...")
    else:
        print("【Step 5】建立每日 8 張結果工作表...")

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # 每日固定 8 張工作表。
    write_group_sheet(wb, "A_基礎買超", a_events, price_cache, is_c=False)
    write_group_sheet(wb, "B_明顯買超", b_events, price_cache, is_c=False)
    write_group_sheet(wb, "C_強勢買超", c_events, price_cache, is_c=False)
    write_group_sheet(wb, "D_大額布局", d_events, price_cache, is_c=False)
    write_group_sheet(wb, "E_超大額布局", e_events, price_cache, is_c=False)
    write_daily_sell_detail_sheet(wb, items, a_events, b_events, c_events, d_events, e_events)
    write_top15_consensus_cache_sheet(wb, top15_consensus_rows or [])
    write_top15_position_detail_sheet(wb, top15_detail_rows or [])

    # 完整修補模式才計算／建立延伸報表；每日流程完全不碰這些工作表。
    if workflow_is_repair():
        write_stats_sheet(wb, a_events, b_events, c_events, d_events, e_events)
        write_combo_winrate_sheet(wb, a_events, b_events, c_events, d_events, e_events)
        write_recent_warrant_amount_ranking_sheet(wb, items)
        write_underlying_broker_count_ranking_sheet(wb, items)
        write_broker_query_sheet(wb, items)
        write_stock_abcde_query_sheet(wb, a_events, b_events, c_events, d_events, e_events)
        write_price_status_sheet(wb, price_cache)
        write_color_legend_sheet(wb)

        if warrant_consensus_7d_rows is not None:
            write_7d_warrant_consensus_top15_sheet(wb, warrant_consensus_7d_rows)
        if broker_10d_detail_rows is not None:
            write_10d_broker_underlying_detail_sheet(wb, broker_10d_detail_rows)
        if broker_10d_winrate_rank_rows is not None:
            write_10d_broker_winrate_rank_sheet(wb, broker_10d_winrate_rank_rows)

    apply_global_amount_comma_format(wb)

    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)

    print(
        f"  ✅ 已存：{output_path} "
        f"（工作表 {len(wb.sheetnames)} 張｜A:{len(a_events)} 筆，B:{len(b_events)} 筆，"
        f"C:{len(c_events)} 筆，D:{len(d_events)} 筆，E:{len(e_events)} 筆）"
    )







# ══════════════════════════════════════════════════════════════════════
# 價格預抓工具：供 longterm 與既有價格流程共用
# ══════════════════════════════════════════════════════════════════════









































def build_price_prefetch_context_from_items(items):
    """
    用既有分點歷史快取重建每日 8 張結果工作表會用到的事件，僅供價格預抓使用。

    注意：這裡不寫入任何結果工作表，只重用正式流程的 A/B/C/D/E 金額強度分類，
    讓價格預抓覆蓋 A～E 與近兩個月（約 40 個交易日）TOP15 真正需要的標的股／權證價格。
    """
    item_map = {}

    for item in items:
        item_map[(item["broker_code"], item["warrant_code"])] = item

    daily_records = build_daily_records(items)
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [
        amount_events.get(code, [])
        for code in AMOUNT_CLASS_CODES
    ]

    return item_map, a_events, b_events, c_events, d_events, e_events


def filter_items_for_selected_scope(items):
    """從已抓好的完整追蹤分點資料中篩出精選五分點，不重新呼叫 API。"""
    selected_labels = set(parse_selected_target_labels())
    return [
        item for item in (items or [])
        if str(item.get("broker_label", "")).strip() in selected_labels
    ]


def build_selected_scope_excel(
    selected_items,
    selected_events,
    selected_item_map,
    price_cache,
    output_path,
    top15_detail_rows=None,
    top15_consensus_rows=None,
):
    """只建立每日流程需要的 8 張精選五分點結果表；資料全部來自同次完整追蹤分點結果。"""
    a_events, b_events, c_events, d_events, e_events = selected_events
    wb = Workbook()
    wb.remove(wb.active)

    write_group_sheet(wb, "A_基礎買超", a_events, price_cache, is_c=False)
    write_group_sheet(wb, "B_明顯買超", b_events, price_cache, is_c=False)
    write_group_sheet(wb, "C_強勢買超", c_events, price_cache, is_c=False)
    write_group_sheet(wb, "D_大額布局", d_events, price_cache, is_c=False)
    write_group_sheet(wb, "E_超大額布局", e_events, price_cache, is_c=False)
    write_daily_sell_detail_sheet(wb, selected_items, a_events, b_events, c_events, d_events, e_events)
    write_top15_consensus_cache_sheet(wb, top15_consensus_rows or [])
    write_top15_position_detail_sheet(wb, top15_detail_rows or [])
    apply_global_amount_comma_format(wb)

    output_dir = os.path.dirname(os.path.abspath(output_path))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return output_path







def run_automatic_cache_maintenance(warrants=None):
    """清理 FinMind 本機快取中的超期日期與已移除分點。"""
    if not CACHE_AUTO_PRUNE_ENABLED or not USE_CACHE:
        return

    print("\n【快取維護】清理超期日期、已移除分點與失效資料...")

    price_df = read_cache_csv(PRICE_CACHE_PATH)
    if price_df is not None and not price_df.empty:
        price_clean, stats = prune_price_cache_dataframe(price_df)
        if stats.get("removed", 0) > 0:
            write_cache_csv(price_clean, PRICE_CACHE_PATH, sync_gsheet=False, sync_supabase=False)
            print(
                f"  🧹 價格快取：{stats['before']:,} → {stats['after']:,} 筆｜"
                f"起始日 {stats.get('cutoff') or '-'}"
            )

    history_df = read_cache_csv(HISTORY_CACHE_PATH)
    if history_df is not None and not history_df.empty:
        history_clean, stats = prune_history_cache_dataframe(history_df)
        if stats.get("removed", 0) > 0:
            write_cache_csv(history_clean, HISTORY_CACHE_PATH, sync_gsheet=False, sync_supabase=False)
            print(
                f"  🧹 分點歷史：{stats['before']:,} → {stats['after']:,} 筆｜"
                f"起始日 {stats.get('cutoff') or '-'}"
            )

    broker_df = read_cache_csv(BROKER_MAP_CACHE_PATH)
    if broker_df is not None and not broker_df.empty:
        scope = "selected5" if RUN_MODE == 1 else "all"
        broker_clean, stats = prune_broker_map_dataframe(broker_df, scope=scope)
        if stats.get("removed", 0) > 0:
            write_cache_csv(broker_clean, BROKER_MAP_CACHE_PATH, sync_gsheet=False, sync_supabase=False)
            print(f"  🧹 分點代號快取：{stats['before']:,} → {stats['after']:,} 筆")

    print("  ✅ 快取自動維護完成")




# ══════════════════════════════════════════════════════════════════════
# WORKFLOW_MODE=longterm：長期留單補價格與修正勝率
# ══════════════════════════════════════════════════════════════════════

def longterm_safe_float(value, default=0.0):
    try:
        if value is None:
            return default
        s = str(value).replace(",", "").strip()
        if not s or s in ("-", "None", "nan", "null"):
            return default
        return float(s)
    except Exception:
        return default


def longterm_target_date_from_history(history_df):
    raw = str(LONGTERM_TARGET_DATE or PRICE_PREFETCH_TARGET_DATE or TOP15_TARGET_DATE or "").strip()
    if raw:
        dt = parse_date(raw)
        if dt:
            return dt.strftime("%Y/%m/%d")

    if history_df is not None and not history_df.empty and "日期" in history_df.columns:
        df = history_df.copy().fillna("")
        # 優先使用有實際買賣活動的最新日期。
        # 若假日或資料源異常產生 0 買 0 賣的日期列，不應拿它當 longterm 統計日期。
        activity_cols = [c for c in ["買進股數", "賣出股數", "買進金額", "賣出金額"] if c in df.columns]
        if activity_cols:
            activity_sum = pd.Series(0.0, index=df.index)
            for col in activity_cols:
                # Google Sheet 讀回來的數字可能是 "1,234" 字串，
                # 不能直接 pd.to_numeric，否則會變 NaN，導致誤判沒有活動。
                activity_values = df[col].map(longterm_safe_float)
                activity_sum = activity_sum + activity_values.fillna(0).abs()
            active_df = df[activity_sum > 0].copy()
        else:
            active_df = df

        dates = []
        for value in active_df["日期"].tolist():
            dt = parse_date(value)
            if dt:
                dates.append(dt)
        if dates:
            return max(dates).strftime("%Y/%m/%d")

    return datetime.today().strftime("%Y/%m/%d")


def longterm_history_range_info(history_df):
    if history_df is None or history_df.empty or "日期" not in history_df.columns:
        return "", "", 0

    dates = []
    for value in history_df["日期"].tolist():
        dt = parse_date(value)
        if dt:
            dates.append(dt)

    if not dates:
        return "", "", 0

    return min(dates).strftime("%Y/%m/%d"), max(dates).strftime("%Y/%m/%d"), len(set(d.date() for d in dates))



def longterm_event_class_from_amount(total_buy_amount, max_single_warrant_buy_amount):
    """依 ABCDE 規則判斷 longterm 留單的原始事件類型。"""
    try:
        total_buy_amount = float(total_buy_amount or 0)
        max_single_warrant_buy_amount = float(max_single_warrant_buy_amount or 0)
    except Exception:
        return ""

    if max_single_warrant_buy_amount < AMOUNT_THRESH:
        return ""

    for code, _label, low, high in AMOUNT_CLASS_SPECS:
        if total_buy_amount >= low and (high is None or total_buy_amount < high):
            return code
    return ""


def build_longterm_event_code_map(df):
    """
    從快取_分點歷史重建「同一分點 × 同一標的 × 同一天」的 ABCDE 事件分類。

    longterm 的未出清留單是以權證買進批次為單位；但修正勝率要對應原本勝率統計的
    A/B/C/D/E 分類，因此這裡先把買進日當天同分點、同標的的買進金額合併分類，
    再把每個未出清買進批次標上事件代碼。
    """
    if df is None or df.empty:
        return {}

    required = {"分點", "券商代號", "標的股", "日期", "買進股數", "買進金額"}
    if not required.issubset(set(df.columns)):
        return {}

    work = df.copy().fillna("")
    work["標的股_norm"] = work["標的股"].map(normalize_underlying_code_for_group)
    work["買進股數_num"] = work["買進股數"].map(longterm_safe_float).fillna(0.0)
    work["買進金額_num"] = work["買進金額"].map(longterm_safe_float).fillna(0.0)
    work = work[(work["買進股數_num"] > 0) & (work["買進金額_num"] > 0)].copy()
    if work.empty:
        return {}

    event_map = {}
    group_cols = ["分點", "券商代號", "標的股_norm", "日期"]
    for key, g in work.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_code, underlying_code, trade_date = key
        total_amount = float(g["買進金額_num"].sum())
        max_amount = float(g["買進金額_num"].max())
        code = longterm_event_class_from_amount(total_amount, max_amount)
        if not code:
            continue
        event_map[(
            str(broker_label).strip(),
            str(broker_code).strip(),
            str(underlying_code).strip(),
            normalize_date_str(trade_date),
        )] = {
            "事件代碼": code,
            "事件類型": f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}",
            "事件總買進金額": total_amount,
            "事件最大單檔買進金額": max_amount,
        }
    return event_map

def rebuild_longterm_open_lots_from_history(history_df, target_date):
    """
    從快取_分點歷史直接用 FIFO 重建目前未出清部位。

    單位是「分點 × 權證」的實際買進批次，不依賴 A/B/C/D/E 事件，
    因此可以抓出所有長期未賣出的留單。賣出依日期 FIFO 扣買進股數與成本，
    同日買賣不互扣，避免把事件日內的買賣誤當出清。
    """
    if history_df is None or history_df.empty:
        return []

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()

    needed_cols = {
        "權證代號", "權證名稱", "標的股", "標的名稱", "分點", "分點名稱", "券商代號", "日期",
        "買進股數", "賣出股數", "買進金額", "賣出金額",
    }
    if not needed_cols.issubset(set(history_df.columns)):
        missing = sorted(needed_cols - set(history_df.columns))
        print(f"  ⚠️ 長期留單分析失敗：快取_分點歷史缺少欄位：{missing}")
        return []

    df = history_df.copy().fillna("")
    df = fix_known_underlying_info_dataframe(df, "權證名稱", "標的股", "標的名稱")
    df["日期"] = df["日期"].map(normalize_date_str)
    # 不使用底線開頭欄名，例如 _dt。pandas itertuples() 會把底線欄名改名，
    # 導致 row._asdict().get("_dt") 取不到值，最後所有列都被跳過。
    df["dt_parsed"] = pd.to_datetime(df["日期"].astype(str).str.replace("/", "-", regex=False), errors="coerce")
    df = df[df["dt_parsed"].notna()].copy()
    df = df[df["dt_parsed"].dt.date <= target_dt.date()].copy()

    if df.empty:
        return []

    for col in ["買進股數", "賣出股數", "買進金額", "賣出金額"]:
        # Google Sheet 讀回可能含千分位逗號，直接 pd.to_numeric 會變 NaN。
        df[col] = df[col].map(longterm_safe_float).fillna(0.0)

    event_code_map = build_longterm_event_code_map(df)

    open_lots = []
    group_cols = ["分點", "分點名稱", "券商代號", "權證代號", "權證名稱", "標的股", "標的名稱"]

    for key, g in df.groupby(group_cols, dropna=False, sort=False):
        broker_label, broker_name, broker_code, warrant_code, warrant_name, underlying_code, underlying_name = key
        lots = []
        g = g.sort_values(["dt_parsed", "日期"]).reset_index(drop=True)

        for rd in g.to_dict("records"):
            trade_dt = rd.get("dt_parsed")
            if pd.isna(trade_dt):
                continue
            trade_date = trade_dt.strftime("%Y/%m/%d")

            buy_qty = float(rd.get("買進股數", 0) or 0)
            buy_amount = float(rd.get("買進金額", 0) or 0)
            if buy_qty > 0 and buy_amount > 0:
                underlying_norm = normalize_underlying_code_for_group(underlying_code)
                event_info = event_code_map.get((
                    str(broker_label).strip(),
                    str(broker_code).strip(),
                    str(underlying_norm).strip(),
                    trade_date,
                ), {})
                lots.append({
                    "分點": str(broker_label).strip(),
                    "分點名稱": str(broker_name).strip(),
                    "券商代號": str(broker_code).strip(),
                    "權證代號": str(warrant_code).strip(),
                    "權證名稱": str(warrant_name).strip(),
                    "標的股": underlying_norm,
                    "標的名稱": str(underlying_name).strip(),
                    "事件代碼": event_info.get("事件代碼", ""),
                    "事件類型": event_info.get("事件類型", "未達ABCDE門檻"),
                    "事件總買進金額": event_info.get("事件總買進金額", ""),
                    "事件最大單檔買進金額": event_info.get("事件最大單檔買進金額", ""),
                    "買進日": trade_date,
                    "買進股數": buy_qty,
                    "買進金額": buy_amount,
                    "剩餘股數": buy_qty,
                    "剩餘成本": buy_amount,
                    "第一筆日期": trade_date,
                    "最後筆日期": trade_date,
                })

            sell_qty_left = float(rd.get("賣出股數", 0) or 0)
            if sell_qty_left <= 0:
                continue

            for lot in lots:
                if sell_qty_left <= 0:
                    break
                buy_dt = parse_date(lot.get("買進日"))
                trade_day = trade_dt.to_pydatetime().date() if hasattr(trade_dt, "to_pydatetime") else trade_dt.date()
                if not buy_dt or trade_day <= buy_dt.date():
                    continue

                remain_qty = float(lot.get("剩餘股數", 0) or 0)
                remain_cost = float(lot.get("剩餘成本", 0) or 0)
                original_qty = float(lot.get("買進股數", 0) or 0)
                original_cost = float(lot.get("買進金額", 0) or 0)
                if remain_qty <= 0 or remain_cost <= 0 or original_qty <= 0 or original_cost <= 0:
                    continue

                avg_cost = original_cost / original_qty
                alloc_qty = min(sell_qty_left, remain_qty)
                alloc_cost = min(remain_cost, alloc_qty * avg_cost)
                lot["剩餘股數"] = max(remain_qty - alloc_qty, 0)
                lot["剩餘成本"] = max(remain_cost - alloc_cost, 0)
                lot["最後筆日期"] = trade_date
                sell_qty_left -= alloc_qty

        for lot in lots:
            remain_qty = float(lot.get("剩餘股數", 0) or 0)
            remain_cost = float(lot.get("剩餘成本", 0) or 0)
            if remain_qty <= 0 or remain_cost <= 0:
                continue
            buy_dt = parse_date(lot.get("買進日"))
            if not buy_dt:
                continue
            holding_days = max((target_dt.date() - buy_dt.date()).days, 0)
            lot["持有天數"] = holding_days
            lot["剩餘均價"] = remain_cost / remain_qty if remain_qty > 0 else None
            open_lots.append(lot)

    print(f"  ✅ FIFO 未出清部位重建完成：{len(open_lots):,} 筆")
    return open_lots


def ensure_longterm_warrant_prices(price_cache, open_lots, target_date):
    """只針對長期未出清權證補最新可用價格，並同步回 price_cache。"""
    if not open_lots:
        return price_cache

    target_dt = parse_date(target_date)
    if not target_dt:
        target_dt = datetime.today()
    target_dt = min(target_dt, datetime.today())
    start_dt = target_dt - timedelta(days=max(int(LONGTERM_PRICE_LOOKBACK_DAYS or 420), 30))

    needed_codes = sorted({
        normalize_price_code(lot.get("權證代號", ""))
        for lot in open_lots
        if normalize_price_code(lot.get("權證代號", ""))
    })

    if not needed_codes:
        return price_cache

    persistent_price_cache = load_price_cache()
    fetch_plan = []

    for code in needed_codes:
        cached_prices = get_cached_prices_for_code(persistent_price_cache, code)
        current_prices = get_price_series_from_cache(price_cache, code)
        merged_prices = merge_price_dicts(cached_prices, current_prices)
        if merged_prices:
            add_price_aliases(price_cache, code, merged_prices)
            persistent_price_cache[normalize_price_code(code)] = merged_prices

        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_dt.strftime("%Y/%m/%d"))
        latest_dt = parse_date(latest_date) if latest_date else None
        need_fetch = latest_price is None
        if latest_dt and (target_dt - latest_dt).days > LONGTERM_PRICE_STALE_DAYS:
            need_fetch = True
        if need_fetch:
            fetch_plan.append(code)

    print(f"  長期留單需檢查權證價格：{len(needed_codes):,} 檔")
    print(f"  長期留單需補抓權證價格：{len(fetch_plan):,} 檔")

    if fetch_plan:
        longterm_fetch_plan = {
            code: [start_dt, target_dt]
            for code in fetch_plan
        }
        changed_price_codes = fetch_price_plan_batch_first(
            price_cache,
            persistent_price_cache,
            longterm_fetch_plan,
            label="長期留單權證價格",
        )
        save_price_cache(persistent_price_cache, changed_codes=changed_price_codes)

    return price_cache


def attach_longterm_price_and_pnl(open_lots, price_cache, target_date):
    rows = []
    target_dt = parse_date(target_date) or datetime.today()

    for lot in open_lots:
        code = normalize_price_code(lot.get("權證代號", ""))
        latest_price, latest_date = get_latest_price_info_on_or_before(price_cache, code, target_date)
        remain_qty = longterm_safe_float(lot.get("剩餘股數"))
        remain_cost = longterm_safe_float(lot.get("剩餘成本"))
        buy_dt = parse_date(lot.get("買進日"))
        holding_days = max((target_dt.date() - buy_dt.date()).days, 0) if buy_dt else int(lot.get("持有天數", 0) or 0)

        market_value = None
        pnl = None
        return_pct = None
        result = "缺價"
        price_status = "缺最新權證價格"

        if latest_price is not None and remain_qty > 0:
            market_value = remain_qty * float(latest_price)
            pnl = market_value - remain_cost
            return_pct = (pnl / remain_cost * 100.0) if remain_cost > 0 else None
            if float(latest_price) <= LONGTERM_ZERO_PRICE_THRESHOLD:
                result = "疑似龜苓膏"
                price_status = "價格接近歸零"
            elif pnl > 0:
                result = "賺錢"
                price_status = "有價格"
            elif pnl < 0:
                result = "賠錢"
                price_status = "有價格"
            else:
                result = "打平"
                price_status = "有價格"

        rows.append({
            "統計日期": target_date,
            "分點": lot.get("分點", ""),
            "分點名稱": lot.get("分點名稱", ""),
            "券商代號": lot.get("券商代號", ""),
            "事件代碼": lot.get("事件代碼", ""),
            "事件類型": lot.get("事件類型", ""),
            "事件總買進金額": lot.get("事件總買進金額", ""),
            "事件最大單檔買進金額": lot.get("事件最大單檔買進金額", ""),
            "標的股": lot.get("標的股", ""),
            "標的名稱": lot.get("標的名稱", ""),
            "權證代號": lot.get("權證代號", ""),
            "權證名稱": lot.get("權證名稱", ""),
            "買進日": lot.get("買進日", ""),
            "持有天數": holding_days,
            "買進股數": round(longterm_safe_float(lot.get("買進股數")), 0),
            "買進金額": round(longterm_safe_float(lot.get("買進金額")), 0),
            "剩餘股數": round(remain_qty, 0),
            "剩餘成本": round(remain_cost, 0),
            "剩餘均價": "" if remain_qty <= 0 else round(remain_cost / remain_qty, 4),
            "最新權證價格": "" if latest_price is None else latest_price,
            "最新價格日期": latest_date or "",
            "估算市值": "" if market_value is None else round(market_value, 0),
            "估算損益": "" if pnl is None else round(pnl, 0),
            "估算報酬%": "" if return_pct is None else round(return_pct, 2),
            "結果": result,
            "價格狀態": price_status,
        })

    return rows


def filter_longterm_rows_by_days(rows, days):
    return [row for row in rows if int(longterm_safe_float(row.get("持有天數"), 0)) >= int(days)]


def summarize_longterm_rows(rows, days):
    if not rows:
        return []

    df = pd.DataFrame(rows).fillna("")
    numeric_cols = ["剩餘成本", "估算市值", "估算損益", "持有天數"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    summary_rows = []
    for broker, g in df.groupby("分點", dropna=False, sort=False):
        total = len(g)
        priced = int((g["結果"].astype(str) != "缺價").sum())
        missing = int((g["結果"].astype(str) == "缺價").sum())
        win = int((g["結果"].astype(str) == "賺錢").sum())
        loss = int((g["結果"].astype(str) == "賠錢").sum())
        flat = int((g["結果"].astype(str) == "打平").sum())
        zero = int((g["結果"].astype(str) == "疑似龜苓膏").sum())
        known_risk = loss + zero
        total_cost = float(g["剩餘成本"].fillna(0).sum()) if "剩餘成本" in g.columns else 0
        missing_cost = float(g.loc[g["結果"].astype(str) == "缺價", "剩餘成本"].fillna(0).sum()) if "剩餘成本" in g.columns else 0
        pnl_sum = float(g["估算損益"].fillna(0).sum()) if "估算損益" in g.columns else 0
        avg_days = float(g["持有天數"].dropna().mean()) if "持有天數" in g.columns and not g["持有天數"].dropna().empty else 0

        summary_rows.append({
            "門檻天數": days,
            "分點": broker,
            "留單筆數": total,
            "有價格筆數": priced,
            "缺價筆數": missing,
            "賺錢筆數": win,
            "賠錢筆數": loss,
            "打平筆數": flat,
            "疑似龜苓膏筆數": zero,
            "已知風險筆數": known_risk,
            "有價格賠錢比例": "" if priced <= 0 else round(known_risk / priced * 100, 2),
            "缺價比例": round(missing / total * 100, 2) if total else 0,
            "留單成本": round(total_cost, 0),
            "缺價成本": round(missing_cost, 0),
            "缺價成本占比": "" if total_cost <= 0 else round(missing_cost / total_cost * 100, 2),
            "已估損益合計": round(pnl_sum, 0),
            "平均持有天數": round(avg_days, 2) if avg_days else "",
        })

    summary_rows.sort(key=lambda r: (-(longterm_safe_float(r.get("留單成本"))), str(r.get("分點", ""))))
    return summary_rows


def build_longterm_overview(threshold_rows_map, history_df, target_date):
    hist_start, hist_end, hist_trade_days = longterm_history_range_info(history_df)
    overview = []
    for days in LONGTERM_OPEN_DAYS:
        rows = threshold_rows_map.get(days, [])
        total = len(rows)
        priced = sum(1 for r in rows if str(r.get("結果", "")) != "缺價")
        missing = sum(1 for r in rows if str(r.get("結果", "")) == "缺價")
        win = sum(1 for r in rows if str(r.get("結果", "")) == "賺錢")
        loss = sum(1 for r in rows if str(r.get("結果", "")) == "賠錢")
        zero = sum(1 for r in rows if str(r.get("結果", "")) == "疑似龜苓膏")
        flat = sum(1 for r in rows if str(r.get("結果", "")) == "打平")
        total_cost = sum(longterm_safe_float(r.get("剩餘成本")) for r in rows)
        missing_cost = sum(longterm_safe_float(r.get("剩餘成本")) for r in rows if str(r.get("結果", "")) == "缺價")
        overview.append({
            "統計日期": target_date,
            "門檻天數": days,
            "留單筆數": total,
            "有價格筆數": priced,
            "缺價筆數": missing,
            "賺錢筆數": win,
            "賠錢筆數": loss,
            "打平筆數": flat,
            "疑似龜苓膏筆數": zero,
            "有價格賠錢比例": "" if priced <= 0 else round((loss + zero) / priced * 100, 2),
            "缺價比例": "" if total <= 0 else round(missing / total * 100, 2),
            "留單成本": round(total_cost, 0),
            "缺價成本": round(missing_cost, 0),
            "缺價成本占比": "" if total_cost <= 0 else round(missing_cost / total_cost * 100, 2),
            "歷史快取起始日": hist_start,
            "歷史快取最後日": hist_end,
            "歷史快取交易日數": hist_trade_days,
            "HISTORY_RETENTION_TRADING_DAYS": HISTORY_RETENTION_TRADING_DAYS,
        })
    return overview


def build_longterm_risk_summary_by_broker_event(rows):
    """把 longterm 明細整理成 分點 × A/B/C/D/E/ALL 的風險筆數。"""
    empty = {}
    if not rows:
        return empty

    df = pd.DataFrame(rows).fillna("")
    if df.empty or "分點" not in df.columns:
        return empty

    df["事件代碼"] = df.get("事件代碼", "").astype(str).str.strip()
    df = df[df["事件代碼"].isin(AMOUNT_CLASS_CODES)].copy()
    if df.empty:
        return empty

    out = {}

    def summarize_group(g):
        total = len(g)
        result = g["結果"].astype(str) if "結果" in g.columns else pd.Series([], dtype=str)
        loss = int((result == "賠錢").sum())
        zero = int((result == "疑似龜苓膏").sum())
        missing = int((result == "缺價").sum())
        win = int((result == "賺錢").sum())
        flat = int((result == "打平").sum())
        cost = sum(longterm_safe_float(v) for v in g.get("剩餘成本", []))
        missing_cost = sum(longterm_safe_float(v) for v in g.loc[result == "缺價", "剩餘成本"]) if "剩餘成本" in g.columns and len(result) else 0
        return {
            "長期留單筆數": int(total),
            "長期賺錢筆數": int(win),
            "長期賠錢筆數": int(loss),
            "長期打平筆數": int(flat),
            "長期疑似龜苓膏筆數": int(zero),
            "長期已知虧損/龜苓膏筆數": int(loss + zero),
            "長期缺價筆數": int(missing),
            "長期留單成本": round(float(cost), 0),
            "長期缺價成本": round(float(missing_cost), 0),
        }

    for (broker, code), g in df.groupby(["分點", "事件代碼"], dropna=False, sort=False):
        out[(str(broker).strip(), str(code).strip())] = summarize_group(g)

    for broker, g in df.groupby("分點", dropna=False, sort=False):
        out[(str(broker).strip(), "ALL")] = summarize_group(g)

    return out


def build_longterm_adjusted_winrate_rows(summary_map, threshold_rows_map):
    """
    產生可放進「勝率統計」的長期留單修正勝率。

    這裡會依 分點 × A/B/C/D/E/ALL 計算：
    - 原始勝率：原本已出清勝率。
    - 修正勝率：把長期留單中已知虧損與疑似龜苓膏視為額外敗筆。
    - 保守勝率：再把長期缺價留單也視為風險敗筆。
    """
    rows = []
    brokers = list(TARGET_PATTERNS.keys())
    for broker in summary_map.keys():
        if broker not in brokers:
            brokers.append(broker)

    event_type_names = {
        code: f"{code}-{AMOUNT_CLASS_LABELS.get(code, '事件')}"
        for code in AMOUNT_CLASS_CODES
    }
    event_type_names["ALL"] = "全部-A+B+C+D+E合併"

    for days in LONGTERM_OPEN_DAYS:
        risk_map = build_longterm_risk_summary_by_broker_event(threshold_rows_map.get(days, []))
        for broker in sorted(set(brokers)):
            for code in AMOUNT_CLASS_CODES + ["ALL"]:
                original = (summary_map.get(broker, {}) or {}).get(code, calc_empty_summary(broker, event_type_names.get(code, code)))
                risk = risk_map.get((broker, code), {})
                closed_count = int(original.get("已出清筆數") or 0)
                win_count = int(original.get("勝筆數") or 0)
                loss_count = int(original.get("敗筆數") or 0)
                flat_count = int(original.get("平手筆數") or 0)
                long_loss = int(longterm_safe_float(risk.get("長期已知虧損/龜苓膏筆數"), 0))
                long_missing = int(longterm_safe_float(risk.get("長期缺價筆數"), 0))
                revised_den = closed_count + long_loss
                conservative_den = closed_count + long_loss + long_missing
                rows.append({
                    "門檻天數": days,
                    "分點": broker,
                    "事件代碼": code,
                    "事件類型": event_type_names.get(code, code),
                    "原始事件數": int(original.get("事件數") or 0),
                    "原始已出清筆數": closed_count,
                    "原始未出清筆數": int(original.get("未出清筆數") or 0),
                    "原始勝筆數": win_count,
                    "原始敗筆數": loss_count,
                    "原始平手筆數": flat_count,
                    "原始勝率": original.get("勝率"),
                    "長期留單筆數": int(longterm_safe_float(risk.get("長期留單筆數"), 0)),
                    "長期賺錢筆數": int(longterm_safe_float(risk.get("長期賺錢筆數"), 0)),
                    "長期賠錢筆數": int(longterm_safe_float(risk.get("長期賠錢筆數"), 0)),
                    "長期疑似龜苓膏筆數": int(longterm_safe_float(risk.get("長期疑似龜苓膏筆數"), 0)),
                    "長期已知虧損/龜苓膏筆數": long_loss,
                    "長期缺價筆數": long_missing,
                    "長期留單成本": risk.get("長期留單成本", 0),
                    "長期缺價成本": risk.get("長期缺價成本", 0),
                    "修正勝率_已知虧損視為敗": "" if revised_den <= 0 else round(win_count / revised_den * 100, 2),
                    "保守勝率_已知虧損加缺價視為敗": "" if conservative_den <= 0 else round(win_count / conservative_den * 100, 2),
                    "說明": "修正勝率以同分點同事件類型的長期留單已知虧損/龜苓膏視為額外敗筆；保守勝率再把缺價留單也視為風險敗筆。",
                })
    return rows

def dataframe_to_sheet(wb, title, rows, freeze=True, max_rows=0):
    ws = wb.create_sheet(safe_worksheet_title(title))
    if max_rows and rows and len(rows) > max_rows:
        rows = rows[:max_rows]
    if not rows:
        ws.append(["無資料"])
        style_sheet(ws, [16])
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    if freeze:
        ws.freeze_panes = "A2"
    widths = []
    for h in headers:
        if "名稱" in h or "說明" in h:
            widths.append(28)
        elif "權證代號" in h or "券商代號" in h or "標的股" in h:
            widths.append(12)
        elif "日期" in h:
            widths.append(12)
        elif "金額" in h or "成本" in h or "市值" in h or "損益" in h:
            widths.append(16)
        else:
            widths.append(14)
    style_sheet(ws, widths)
    return ws


def build_longterm_workbook(output_path, overview_rows, threshold_summary_map, threshold_rows_map, adjusted_rows):
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "長期留單總覽"
    dataframe_to_sheet(wb, "_tmp_remove", [])
    try:
        del wb["_tmp_remove"]
    except Exception:
        pass
    # 直接寫入 active sheet，避免 Workbook 預設空白表殘留。
    if overview_rows:
        headers = list(overview_rows[0].keys())
        ws0.append(headers)
        for row in overview_rows:
            ws0.append([row.get(h, "") for h in headers])
        ws0.freeze_panes = "A2"
        style_sheet(ws0, [14] * len(headers))
    else:
        ws0.append(["無資料"])
        style_sheet(ws0, [16])

    dataframe_to_sheet(wb, "長期留單修正勝率", adjusted_rows)

    for days in LONGTERM_OPEN_DAYS:
        dataframe_to_sheet(wb, f"長期留單分點彙總_{days}日", threshold_summary_map.get(days, []))
        detail_rows = threshold_rows_map.get(days, [])
        detail_rows = sorted(
            detail_rows,
            key=lambda r: (-(longterm_safe_float(r.get("剩餘成本"))), str(r.get("分點", "")), str(r.get("權證代號", ""))),
        )
        dataframe_to_sheet(
            wb,
            f"長期留單明細_{days}日",
            detail_rows,
            max_rows=max(int(LONGTERM_MAX_DETAIL_ROWS or 0), 0),
        )

    wb.save(output_path)



def _longterm_pct_text(value):
    if value is None or str(value).strip() == "":
        return "-"
    try:
        return f"{float(value):.2f}%"
    except Exception:
        return str(value).strip() or "-"


def _longterm_extract_event_code(event_type):
    s = str(event_type or "").strip()
    if not s:
        return ""
    if s.startswith("全部"):
        return "ALL"
    m = re.match(r"^([A-E])(?:[-－]|$)", s)
    if m:
        return m.group(1)
    return "ALL" if "全部" in s else s


def _longterm_build_inline_winrate_map(adjusted_rows):
    """
    取 LONGTERM_WINRATE_SHEET_DAYS 對應門檻的修正勝率，準備回填到勝率統計原表格。
    Google Sheet 只放一個「修正勝率」欄位；完整 120 / 150 / 180 日明細仍保留在 Excel artifact。
    """
    selected_days = int(LONGTERM_WINRATE_SHEET_DAYS or (LONGTERM_OPEN_DAYS[0] if LONGTERM_OPEN_DAYS else 120))
    rows = [r for r in (adjusted_rows or []) if int(longterm_safe_float(r.get("門檻天數"), -1)) == selected_days]
    if not rows and adjusted_rows:
        available_days = sorted({int(longterm_safe_float(r.get("門檻天數"), 0)) for r in adjusted_rows if longterm_safe_float(r.get("門檻天數"), 0) > 0})
        if available_days:
            selected_days = available_days[0]
            rows = [r for r in adjusted_rows if int(longterm_safe_float(r.get("門檻天數"), -1)) == selected_days]

    out = {}
    for rec in rows:
        broker = str(rec.get("分點", "")).strip()
        code = str(rec.get("事件代碼", "")).strip() or _longterm_extract_event_code(rec.get("事件類型", ""))
        if not broker or not code:
            continue
        out[(broker, code)] = _longterm_pct_text(rec.get("修正勝率_已知虧損視為敗"))
    return out, selected_days


def _longterm_find_winrate_header_info(values):
    """找出勝率統計內既有的勝率欄位，並判斷旁邊是否已有修正勝率欄。"""
    for row_idx, row in enumerate(values or [], start=1):
        cells = [str(x).strip() for x in row]
        if "分點" not in cells or "事件類型" not in cells or "勝率" not in cells:
            continue
        winrate_idx = cells.index("勝率")
        has_adjusted_next = (winrate_idx + 1 < len(cells) and str(cells[winrate_idx + 1]).strip().startswith("修正勝率"))
        return {
            "header_row": row_idx,
            "winrate_col_index0": winrate_idx,
            "adjusted_col_index0": winrate_idx + 1,
            "has_adjusted_next": has_adjusted_next,
        }
    return None


def _longterm_clear_old_bottom_winrate_block(ws, values):
    """清掉舊版曾寫在勝率統計底部的長期留單修正勝率區塊。"""
    marker = "長期留單修正勝率"
    marker_row = None
    for idx, row in enumerate(values or [], start=1):
        first = str(row[0]).strip() if row else ""
        if first.startswith(marker):
            marker_row = idx
            break
    if marker_row is None:
        return False
    try:
        current_row_count = int(getattr(ws, "row_count", len(values)) or len(values))
        clear_range = f"A{marker_row}:Z{current_row_count}"
        gsheet_api_call("清除舊版勝率統計底部長期修正勝率區塊", ws.batch_clear, [clear_range])
        print(f"  🧹 已清除舊版勝率統計底部長期修正勝率區塊：第 {marker_row} 列起")
        return True
    except Exception as e:
        print(f"  ⚠️ 清除舊版勝率統計底部長期修正勝率區塊失敗：{type(e).__name__}: {e}")
        return False


def upload_longterm_adjusted_winrate_to_gsheet(adjusted_rows, target_date):
    """
    longterm 模式只把修正勝率回填到既有「勝率統計」表格的「勝率」旁邊。

    不新增長期留單明細 / 分點彙總工作表，完整明細只保留在 Excel artifact。
    若勝率旁邊尚未有修正勝率欄，第一次 longterm 會插入一欄；之後重跑只更新該欄。
    """
    if not LONGTERM_UPDATE_WINRATE_SHEET:
        print("  ✅ LONGTERM_UPDATE_WINRATE_SHEET=0，略過更新 Google Sheet 勝率統計。")
        return False
    if not GSHEET_RESULT_ENABLED or not gsheet_enabled():
        print("  ⚠️ 未設定 GCP_SERVICE_KEY，略過 Google Sheet 長期修正勝率同步。")
        return False

    title = "勝率統計"
    ws = get_or_create_worksheet(title, rows=200, cols=30)
    if ws is None:
        print("  ⚠️ 找不到或無法建立 Google Sheet：勝率統計，略過長期修正勝率同步。")
        return False

    try:
        existing_values = ws.get_all_values() or []
    except Exception:
        existing_values = []

    if not existing_values:
        print("  ⚠️ 勝率統計工作表目前沒有資料，無法回填修正勝率。")
        return False

    _longterm_clear_old_bottom_winrate_block(ws, existing_values)
    try:
        existing_values = ws.get_all_values() or []
    except Exception:
        existing_values = existing_values or []

    info = _longterm_find_winrate_header_info(existing_values)
    if not info:
        print("  ⚠️ 找不到勝率統計表頭中的『勝率』欄，無法回填修正勝率。")
        return False

    adjusted_idx0 = int(info["adjusted_col_index0"])
    adjusted_col_1based = adjusted_idx0 + 1

    try:
        sheet_id = int(ws.id)
    except Exception:
        sheet_id = None

    if not info.get("has_adjusted_next"):
        if sheet_id is None:
            print("  ⚠️ 無法取得勝率統計 sheetId，無法插入修正勝率欄。")
            return False
        requests = [{
            "insertDimension": {
                "range": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": adjusted_idx0,
                    "endIndex": adjusted_idx0 + 1,
                },
                "inheritFromBefore": True,
            }
        }]
        _gsheet_batch_update(requests)
        print(f"  ✅ 已在勝率欄旁新增欄位：{LONGTERM_WINRATE_SHEET_HEADER}")
        try:
            existing_values = ws.get_all_values() or []
        except Exception:
            existing_values = existing_values or []
    else:
        print(f"  ✅ 勝率欄旁已存在修正勝率欄，直接更新：{LONGTERM_WINRATE_SHEET_HEADER}")

    winrate_map, selected_days = _longterm_build_inline_winrate_map(adjusted_rows)
    if not winrate_map:
        print("  ⚠️ 沒有可回填的長期修正勝率資料。")
        return False

    max_row = max(len(existing_values), 1)
    column_values = [[""] for _ in range(max_row)]
    updated_count = 0
    header_count = 0
    current_header = None

    for row_idx, row in enumerate(existing_values, start=1):
        cells = [str(x).strip() for x in row]
        if "分點" in cells and "事件類型" in cells and "勝率" in cells:
            # 每個分點區塊都有自己的表頭；同一欄都要顯示修正勝率。
            current_header = {h: i for i, h in enumerate(cells) if h}
            column_values[row_idx - 1] = [LONGTERM_WINRATE_SHEET_HEADER]
            header_count += 1
            continue

        if not current_header:
            continue

        broker_col = current_header.get("分點")
        event_col = current_header.get("事件類型")
        if broker_col is None or event_col is None:
            continue
        if broker_col >= len(cells) or event_col >= len(cells):
            continue

        broker = str(cells[broker_col]).strip()
        event_type = str(cells[event_col]).strip()
        if not broker or broker.startswith("分點：") or not event_type:
            continue
        if broker == "分點" or event_type == "事件類型":
            continue

        event_code = _longterm_extract_event_code(event_type)
        value = winrate_map.get((broker, event_code))
        if value is None:
            # 若某分點 / 類型沒有對應修正資料，維持空白，不改其他欄位。
            continue
        column_values[row_idx - 1] = [value]
        updated_count += 1

    col_letter = get_column_letter(adjusted_col_1based)
    try:
        gsheet_api_call(
            "回填勝率統計修正勝率欄",
            ws.update,
            values=column_values,
            range_name=f"{col_letter}1:{col_letter}{max_row}",
            value_input_option="USER_ENTERED",
        )

        # 只針對新增 / 更新的修正勝率欄做最小格式處理，不重套整張表。
        try:
            requests = [{
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 0,
                        "endRowIndex": max_row,
                        "startColumnIndex": adjusted_idx0,
                        "endColumnIndex": adjusted_idx0 + 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "horizontalAlignment": "CENTER",
                            "verticalAlignment": "MIDDLE",
                        }
                    },
                    "fields": "userEnteredFormat.horizontalAlignment,userEnteredFormat.verticalAlignment",
                }
            }]
            _gsheet_batch_update(requests)
        except Exception:
            pass

        print(
            f"  ☁️ 已更新 Google Sheet：勝率統計｜回填 {LONGTERM_WINRATE_SHEET_HEADER} "
            f"{updated_count:,} 筆｜門檻 {selected_days} 日｜目標日期 {target_date}"
        )
        return True
    except Exception as e:
        print(f"  ⚠️ 回填 Google Sheet 勝率統計修正勝率失敗：{type(e).__name__}: {e}")
        return False


def run_longterm_workflow(warrants, broker_map, output_path, program_start):
    print("\n【Longterm】長期留單補價格與修正勝率分析...")
    history_df = load_history_cache()
    history_df, identity_stats = repair_history_metadata_from_warrants(history_df, warrants)
    if identity_stats.get("rows", 0) > 0:
        history_df = save_history_cache(history_df, fetched_items=None, previous_history_empty=False)
        print(f"  🧭 Longterm 已修正歷史身分配對：{identity_stats['rows']:,} 列")
    if history_df is None or history_df.empty:
        print("  ⚠️ 長期留單分析停止：快取_分點歷史為空。請先跑 daily 或 repair 建立歷史快取。")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    target_date = longterm_target_date_from_history(history_df)
    print(f"  ✅ 長期留單統計日期：{target_date}")
    hist_start, hist_end, hist_trade_days = longterm_history_range_info(history_df)
    print(f"  ✅ 歷史快取範圍：{hist_start or '-'} ～ {hist_end or '-'}｜交易日數 {hist_trade_days:,}｜保留設定 {HISTORY_RETENTION_TRADING_DAYS} 個交易日")

    open_lots = rebuild_longterm_open_lots_from_history(history_df, target_date)
    if not open_lots:
        print("  ⚠️ 長期留單分析停止：沒有重建出未出清部位。")
        elapsed = time.time() - program_start
        print(f"\n⏱️ 總執行時間：{elapsed:.2f} 秒")
        return

    min_days = min(LONGTERM_OPEN_DAYS)
    long_open_lots = [lot for lot in open_lots if int(lot.get("持有天數", 0) or 0) >= min_days]
    print(f"  ✅ 未出清買進批次：{len(open_lots):,} 筆｜{min_days} 日以上：{len(long_open_lots):,} 筆")

    price_cache = load_price_cache()
    price_cache = ensure_longterm_warrant_prices(price_cache, long_open_lots, target_date)
    priced_rows_all = attach_longterm_price_and_pnl(long_open_lots, price_cache, target_date)

    threshold_rows_map = {days: filter_longterm_rows_by_days(priced_rows_all, days) for days in LONGTERM_OPEN_DAYS}
    threshold_summary_map = {days: summarize_longterm_rows(threshold_rows_map.get(days, []), days) for days in LONGTERM_OPEN_DAYS}
    overview_rows = build_longterm_overview(threshold_rows_map, history_df, target_date)

    # 用現有 A/B/C/D/E 事件統計提供「原始已出清勝率」基準，再把長期風險做修正勝率。
    items = items_from_history_cache(history_df)
    item_map = {(item["broker_code"], item["warrant_code"]): item for item in items}
    daily_records = build_daily_records(items)
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [amount_events.get(code, []) for code in AMOUNT_CLASS_CODES]
    stat_records = collect_stat_records(a_events, b_events, c_events, d_events, e_events)
    summary_map, _ = make_summary_map(stat_records)
    adjusted_rows = build_longterm_adjusted_winrate_rows(summary_map, threshold_rows_map)

    longterm_output_path = os.path.join(
        OUTPUT_DIR,
        f"warrant_longterm_open_positions_{datetime.today().strftime('%Y%m%d')}.xlsx",
    )
    build_longterm_workbook(
        longterm_output_path,
        overview_rows,
        threshold_summary_map,
        threshold_rows_map,
        adjusted_rows,
    )

    if LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET:
        print("  ⚠️ LONGTERM_UPLOAD_FULL_WORKBOOK_TO_GSHEET=1：將完整長期留單 Excel 同步到 Google Sheet，可能占用大量格數。")
        upload_excel_to_google_sheet(longterm_output_path)
    else:
        print("  ✅ 長期留單明細 / 分點彙總僅輸出 Excel，不同步到 Google Sheet，避免占用格數。")

    upload_longterm_adjusted_winrate_to_gsheet(adjusted_rows, target_date)

    for row in overview_rows:
        print(
            f"  ✅ {row.get('門檻天數')}日以上：留單 {int(row.get('留單筆數') or 0):,} 筆｜"
            f"有價格 {int(row.get('有價格筆數') or 0):,}｜缺價 {int(row.get('缺價筆數') or 0):,}｜"
            f"賠錢 {int(row.get('賠錢筆數') or 0):,}｜龜苓膏 {int(row.get('疑似龜苓膏筆數') or 0):,}"
        )

    elapsed = time.time() - program_start
    print(f"\n{'=' * 70}")
    print("✅ 長期留單分析完成！")
    print(f"📄 {longterm_output_path}")
    print(f"⏱️ 總執行時間：{elapsed:.2f} 秒")

# ══════════════════════════════════════════════════════════════════════
# 空結果安全同步判斷
# ══════════════════════════════════════════════════════════════════════

def evaluate_empty_result_source_completeness(
    broker_map,
    candidates=None,
    candidates_to_fetch=None,
    history_cache_df=None,
    allow_empty_candidates=False,
):
    reasons = []
    if not LIVE_WARRANT_SNAPSHOT_READY:
        reasons.append("FinMind 權證清單與標的對照未完整取得")

    active_labels = set(TARGET_PATTERNS.keys())
    missing_brokers = sorted(active_labels - set((broker_map or {}).keys()))
    if missing_brokers:
        reasons.append(f"追蹤分點代號不完整：缺少 {len(missing_brokers)} 個")

    if not _FINMIND_TARGET_DATE_OK:
        reasons.append(f"FinMind 目標交易日 {_FINMIND_TARGET_DATE or '-'} 全市場日檔未完整取得")

    return not reasons, reasons



# ══════════════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════════════

def main():
    _GROUP_OUTCOME_SALE_ROWS_CACHE.clear()
    program_start = time.time()
    configure_run_mode()

    if not FINMIND_TOKEN:
        raise RuntimeError("缺少 FINMIND_API_0714，請先在 GitHub Secrets 設定同名 Secret。")

    today_fn = datetime.today().strftime("%Y%m%d")
    output_path = os.path.join(OUTPUT_DIR, f"warrant_backtest_ABCDE_{today_fn}.xlsx")

    print(f"\n認購權證特定分點買超回測 ABCDE 版 | {today_fn}")
    print("資料來源：FinMind TaiwanStockWarrantTradingDailyReport 全市場日 Parquet")
    print("新制分類：同一分點 × 同一標的 × 同一天 = 1 筆事件")
    print(f"進入條件：事件內至少 1 檔權證單日買進金額 >= {AMOUNT_THRESH // 10000}萬")
    print("A：100–159萬｜B：160–249萬｜C：250–499萬｜D：500–999萬｜E：1000萬以上")
    print(f"程式版本：{PROGRAM_BUILD_ID}")
    print(f"工作流模式：WORKFLOW_MODE={WORKFLOW_MODE}")
    print(
        "TOP15估值備援："
        f"LP={'ON' if TOP15_LP_QUOTE_FALLBACK_ENABLED else 'OFF'}｜"
        f"LP無日期接受={'ON' if TOP15_LP_ALLOW_IMPLICIT_LATEST_DATE else 'OFF'}｜"
        f"最近成交價={'ON' if TOP15_LAST_TRADE_FALLBACK_ENABLED else 'OFF'}"
        f"({TOP15_LAST_TRADE_FALLBACK_MAX_DAYS}日)"
    )
    print(f"執行模式：RUN_MODE={RUN_MODE}｜分點數：{len(TARGET_PATTERNS)}")
    print("=" * 70)

    # 權證中繼資料與券商資料可同時抓取；兩者均有本機 Parquet 快取。
    with ThreadPoolExecutor(max_workers=2) as ex:
        warrant_future = ex.submit(get_all_call_warrants)
        broker_future = ex.submit(find_broker_codes, None)
        warrants = warrant_future.result()
        broker_map = filter_broker_map_for_active_targets(broker_future.result())

    if not warrants or not broker_map:
        print("  ⚠️ FinMind 權證或分點中繼資料不完整，本次停止，不修改 Google Sheet。")
        return

    run_automatic_cache_maintenance(warrants)

    if workflow_is_longterm():
        run_longterm_workflow(warrants, broker_map, output_path, program_start)
        return

    requested_target_date = latest_finmind_trading_date_on_or_before(datetime.today())
    print(f"  ✅ 本次要求目標交易日：{requested_target_date}")

    # daily：最新日尚未發布就快速結束，避免排程反覆空抓。
    # repair：不可在此提前 return；若最新日尚未發布，往前找最近已發布交易日，
    # 再以該日為基準補抓完整歷史區間。
    requested_raw, requested_status = download_finmind_warrant_day(
        requested_target_date,
        force_refresh=FINMIND_FORCE_REFRESH_TARGET_DATE,
    )
    target_date, target_raw, target_status, used_target_fallback = resolve_finmind_refresh_target(
        requested_target_date,
        preloaded_raw=requested_raw,
        preloaded_status=requested_status,
    )

    if target_status != "ok":
        if workflow_is_repair():
            print(
                f"  ⚠️ 完整修補模式找不到可用的 FinMind 全市場權證分點日檔："
                f"要求日={requested_target_date}｜status={target_status}。"
                "為避免以不完整來源修補，本次停止且不修改 Google Sheet。"
            )
        else:
            print(
                f"  ⏹️ FinMind 尚未提供 {requested_target_date} 全市場權證分點資料"
                f"（status={target_status}）。本次每日流程快速結束，"
                "不抓空白資料、不補價格、不修改 Google Sheet。"
            )
        return

    if used_target_fallback:
        print(
            f"  🔧 完整修補基準日：{target_date}｜"
            f"原要求日 {requested_target_date} 尚未發布，但歷史補抓會繼續執行。"
        )
    else:
        print(f"  ✅ 本次實際目標交易日：{target_date}")

    print(f"  ✅ 目標日全市場權證分點原始列數：{len(target_raw):,}")
    history_cache_df = load_history_cache()
    history_cache_df, refresh_status = refresh_history_from_finmind(
        warrants,
        broker_map,
        history_cache_df,
        target_date,
        preloaded_target_df=target_raw,
    )

    source_complete, source_reasons = evaluate_empty_result_source_completeness(
        broker_map,
        history_cache_df=history_cache_df,
        allow_empty_candidates=True,
    )
    if not source_complete:
        print("  ⚠️ FinMind 來源完整性未確認，本次停止，不修改 Google Sheet。")
        for reason in source_reasons:
            print(f"    - {reason}")
        return

    items = items_from_history_cache(history_cache_df)
    if not items:
        print("  ✅ 目標日完整日檔已取得，但追蹤分點沒有可用資料；仍會依既有歷史完成修補與報表計算，不會刪除工作表。")

    item_map = {(item["broker_code"], item["warrant_code"]): item for item in items}
    daily_records = build_daily_records(items)

    print("【Step 3b】建立 A/B/C/D/E 金額強度事件...")
    amount_events = build_amount_class_events(daily_records, item_map)
    a_events, b_events, c_events, d_events, e_events = [
        amount_events.get(code, []) for code in AMOUNT_CLASS_CODES
    ]
    print(
        f"  ✅ 金額強度事件：A:{len(a_events):,}｜B:{len(b_events):,}｜"
        f"C:{len(c_events):,}｜D:{len(d_events):,}｜E:{len(e_events):,}"
    )

    persistent_price_cache = load_price_cache()
    all_changed_price_codes = set()
    price_cache, changed_codes = fetch_all_prices(
        a_events, b_events, c_events, d_events, e_events,
        persistent_price_cache=persistent_price_cache,
        defer_save=True,
    )
    all_changed_price_codes.update(changed_codes)

    top15_detail_rows, top15_consensus_rows = build_top15_position_detail_and_consensus_rows(
        a_events, b_events, c_events, d_events, e_events,
        item_map,
        price_cache,
        data_scope="全分點" if RUN_MODE == 2 else "精選五分點",
        allow_price_fetch=True,
        persistent_price_cache=persistent_price_cache,
        defer_price_save=True,
        price_changed_codes=all_changed_price_codes,
    )

    selected_items = []
    selected_item_map = {}
    selected_events = ([], [], [], [], [])
    selected_top15_detail_rows = []
    selected_top15_consensus_rows = []

    if RUN_MODE == 2:
        selected_items = filter_items_for_selected_scope(items)
        selected_item_map, sa, sb, sc, sd, se = build_price_prefetch_context_from_items(selected_items)
        selected_events = (sa, sb, sc, sd, se)
        selected_top15_detail_rows, selected_top15_consensus_rows = build_top15_position_detail_and_consensus_rows(
            sa, sb, sc, sd, se,
            selected_item_map,
            price_cache,
            data_scope="精選五分點",
            allow_price_fetch=False,
        )
        print(f"  ✅ 已從同一份 FinMind 歷史資料切出精選五分點：{len(selected_items):,} 組，不重抓 API。")

    if all_changed_price_codes:
        save_price_cache(persistent_price_cache, changed_codes=all_changed_price_codes)
    else:
        print("  ✅ 價格快取沒有新增或更新，略過儲存。")

    warrant_consensus_7d_rows = None
    broker_10d_detail_rows = None
    broker_10d_winrate_rank_rows = None
    if workflow_is_repair() and RUN_MODE == 2:
        print("  🔧 完整修補模式：重建延伸報表資料，不影響每日流程。")
        warrant_consensus_7d_rows = build_7d_warrant_consensus_top15_rows(items, target_date)
        broker_10d_detail_rows = build_10d_broker_underlying_detail_rows(items, price_cache, target_date)
        broker_10d_winrate_rank_rows = build_10d_broker_winrate_rank_rows(broker_10d_detail_rows)

    build_excel(
        a_events, b_events, c_events, d_events, e_events,
        item_map, price_cache, items, output_path,
        top15_detail_rows, top15_consensus_rows,
        warrant_consensus_7d_rows=warrant_consensus_7d_rows,
        broker_10d_detail_rows=broker_10d_detail_rows,
        broker_10d_winrate_rank_rows=broker_10d_winrate_rank_rows,
    )

    extra_scope_values = None
    if RUN_MODE == 2:
        selected_output_path = os.path.join(OUTPUT_DIR, f"warrant_backtest_ABCDE_selected5_{today_fn}.xlsx")
        build_selected_scope_excel(
            selected_items,
            selected_events,
            selected_item_map,
            price_cache,
            selected_output_path,
            selected_top15_detail_rows,
            selected_top15_consensus_rows,
        )
        extra_scope_values = read_excel_values_by_title(
            selected_output_path,
            allowed_titles=set(DAILY_RESULT_SHEET_TITLES),
        )

    upload_excel_to_google_sheet(
        output_path,
        data_scope="全分點" if RUN_MODE == 2 else "精選五分點",
        allowed_titles=(FULL_REPAIR_RESULT_SHEET_TITLES if workflow_is_repair() else DAILY_RESULT_SHEET_TITLES),
        extra_scope_values=extra_scope_values,
    )

    elapsed = time.time() - program_start
    print(f"\n{'=' * 70}")
    print("✅ 完成！")
    print(f"📄 {output_path}")
    print(f"⏱️ 總執行時間：{elapsed:.2f} 秒")



if __name__ == "__main__":
    main()
